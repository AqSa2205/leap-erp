from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Sum, F
from django.http import HttpResponse
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import ManpowerSheet, ManpowerLineItem
from .forms import SheetForm, LineItemForm, EmployeeCreateForm, FilterForm, ImportForm

# All cost field names used in DB-level aggregation
COST_FIELDS = [
    'gross_salary', 'iqama_cost', 'service_transfer_visa_fee',
    'gosi_cost', 'vacation_pay', 'exe_cost', 'eosb',
    'air_ticket', 'insurance_cost', 'project_allowance',
    'ppe', 'engineering_council_cost', 'other_expenditures',
]


class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    def test_func(self):
        return self.request.user.is_super_admin_user or self.request.user.is_admin_user


# ─── Sheet Views ─────────────────────────────────────────────────────────────


class SheetListView(AdminRequiredMixin, ListView):
    model = ManpowerLineItem
    template_name = 'manpower/sheet_list.html'
    context_object_name = 'employees'
    paginate_by = 25

    def get_queryset(self):
        queryset = ManpowerLineItem.objects.select_related('sheet')

        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(employee_name__icontains=search) |
                Q(department__icontains=search) |
                Q(designation__icontains=search) |
                Q(location_project__icontains=search) |
                Q(sheet__title__icontains=search)
            )

        return queryset.order_by('sheet', 'order', 'pk')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = FilterForm(self.request.GET)
        context['total_sheets'] = ManpowerSheet.objects.count()
        context['total_employees'] = ManpowerLineItem.objects.count()
        return context


class SheetDetailView(AdminRequiredMixin, DetailView):
    model = ManpowerSheet
    template_name = 'manpower/sheet_detail.html'
    context_object_name = 'sheet'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['line_items'] = self.object.line_items.all()
        return context


class SheetCreateView(AdminRequiredMixin, CreateView):
    model = ManpowerLineItem
    form_class = EmployeeCreateForm
    template_name = 'manpower/sheet_form.html'
    success_url = reverse_lazy('manpower:sheet_list')

    def form_valid(self, form):
        messages.success(self.request, 'Employee added successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Employee'
        context['button_text'] = 'Save Employee'
        return context


class NewSheetView(AdminRequiredMixin, CreateView):
    """Create a new ManpowerSheet (parent container) only."""
    model = ManpowerSheet
    form_class = SheetForm
    template_name = 'manpower/new_sheet_form.html'
    success_url = reverse_lazy('manpower:sheet_create')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Manpower sheet created. You can now add employees to it.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Create New Sheet'
        context['button_text'] = 'Create Sheet'
        return context


class SheetUpdateView(AdminRequiredMixin, UpdateView):
    model = ManpowerSheet
    form_class = SheetForm
    template_name = 'manpower/new_sheet_form.html'
    success_url = reverse_lazy('manpower:sheet_list')

    def form_valid(self, form):
        messages.success(self.request, 'Manpower sheet updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Manpower Sheet'
        context['button_text'] = 'Update Sheet'
        return context


class SheetDeleteView(AdminRequiredMixin, DeleteView):
    model = ManpowerSheet
    template_name = 'manpower/sheet_confirm_delete.html'
    success_url = reverse_lazy('manpower:sheet_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Manpower sheet deleted successfully.')
        return super().delete(request, *args, **kwargs)


# ─── Line Item Views ─────────────────────────────────────────────────────────


class LineItemCreateView(AdminRequiredMixin, CreateView):
    model = ManpowerLineItem
    form_class = LineItemForm
    template_name = 'manpower/lineitem_form.html'

    def form_valid(self, form):
        form.instance.sheet_id = self.kwargs['pk']
        messages.success(self.request, 'Employee added successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('manpower:sheet_detail', kwargs={'pk': self.kwargs['pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sheet'] = get_object_or_404(ManpowerSheet, pk=self.kwargs['pk'])
        context['title'] = 'Add Employee'
        context['button_text'] = 'Add Employee'
        return context


class LineItemUpdateView(AdminRequiredMixin, UpdateView):
    model = ManpowerLineItem
    form_class = LineItemForm
    template_name = 'manpower/lineitem_form.html'

    def form_valid(self, form):
        messages.success(self.request, 'Employee updated successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('manpower:sheet_detail', kwargs={'pk': self.object.sheet_id})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sheet'] = self.object.sheet
        context['title'] = 'Edit Employee'
        context['button_text'] = 'Update Employee'
        return context


class LineItemDeleteView(AdminRequiredMixin, DeleteView):
    model = ManpowerLineItem
    template_name = 'manpower/sheet_confirm_delete.html'

    def get_success_url(self):
        return reverse('manpower:sheet_detail', kwargs={'pk': self.object.sheet_id})

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Employee removed successfully.')
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_lineitem'] = True
        return context


# ─── Import / Export ──────────────────────────────────────────────────────────


@login_required
def sheet_import(request):
    if not (request.user.is_super_admin_user or request.user.is_admin_user):
        messages.error(request, 'You do not have permission to import manpower sheets.')
        return redirect('manpower:sheet_list')

    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active

                # Detect header row
                header_row = None
                headers = []
                for row_num in range(1, 6):
                    row_values = [cell.value for cell in ws[row_num]]
                    for val in row_values:
                        if val and str(val).strip().upper() in (
                            'EMPLOYEES NAME', 'EMPLOYEE NAME', 'NAME',
                            'SR. NO.', 'SR.NO.', 'SR. NO', 'S.NO', 'S.NO.',
                            'GROSS SALARY', 'DESIGNATION', 'DEPARTMENT',
                        ):
                            header_row = row_num
                            headers = row_values
                            break
                    if header_row:
                        break

                if not header_row:
                    header_row = 1
                    headers = [cell.value for cell in ws[1]]

                # Build header map
                header_map = {}
                for idx, h in enumerate(headers):
                    if h:
                        header_map[str(h).strip().lower()] = idx

                sheet_title = ws.title or excel_file.name.rsplit('.', 1)[0]

                sheet = ManpowerSheet.objects.create(
                    title=sheet_title[:255],
                    date=datetime.now().date(),
                    created_by=request.user,
                )

                imported_count = 0

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if not any(row):
                        continue

                    def get_value(possible_headers):
                        for h in possible_headers:
                            if h in header_map:
                                idx = header_map[h]
                                if idx < len(row):
                                    return row[idx]
                        return None

                    name = get_value([
                        'employees name', 'employee name', 'name', 'full name',
                    ])
                    if not name or not str(name).strip():
                        continue

                    def parse_date_val(val):
                        if val is None:
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        if isinstance(val, date):
                            return val
                        return None

                    def parse_decimal(val):
                        if val is None:
                            return 0
                        try:
                            return round(float(str(val)), 2)
                        except (ValueError, TypeError):
                            return 0

                    ManpowerLineItem.objects.create(
                        sheet=sheet,
                        order=imported_count,
                        employee_name=str(name).strip()[:255],
                        department=str(get_value([
                            'department', 'dept',
                        ]) or '').strip()[:255],
                        classification=str(get_value([
                            'classification of staff', 'classification', 'class',
                        ]) or '').strip()[:255],
                        location_project=str(get_value([
                            'location/project', 'location', 'project', 'site',
                        ]) or '').strip()[:255],
                        designation=str(get_value([
                            'designation', 'position', 'title',
                        ]) or '').strip()[:255],
                        doj=parse_date_val(get_value([
                            'doj', 'date of joining', 'joining date',
                        ])),
                        demobilization_date=parse_date_val(get_value([
                            'employees demobilization date', 'demobilization date',
                            'demob date',
                        ])),
                        date_of_birth=parse_date_val(get_value([
                            'date of birth', 'dob', 'birth date',
                        ])),
                        budget_cost=parse_decimal(get_value([
                            'budget cost', 'budget',
                        ])),
                        gross_salary=parse_decimal(get_value([
                            'gross salary', 'salary', 'basic salary',
                        ])),
                        iqama_cost=parse_decimal(get_value([
                            'iqama cost', 'iqama',
                        ])),
                        service_transfer_visa_fee=parse_decimal(get_value([
                            'services transfer & visa fee',
                            'service transfer & visa fee',
                            'services transfer',
                            'visa fee',
                        ])),
                        gosi_cost=parse_decimal(get_value([
                            'gosi cost', 'gosi',
                        ])),
                        vacation_pay=parse_decimal(get_value([
                            'vacation pay', 'vacation',
                        ])),
                        exe_cost=parse_decimal(get_value([
                            'exe cost', 'exe',
                        ])),
                        eosb=parse_decimal(get_value([
                            'eosb', 'end of service', 'end of service benefit',
                        ])),
                        air_ticket=parse_decimal(get_value([
                            'air ticket', 'airticket', 'ticket',
                        ])),
                        insurance_cost=parse_decimal(get_value([
                            'insurance cost', 'insurance',
                        ])),
                        project_allowance=parse_decimal(get_value([
                            'project allowance', 'allowance',
                        ])),
                        ppe=parse_decimal(get_value([
                            'ppe',
                        ])),
                        engineering_council_cost=parse_decimal(get_value([
                            'engineering council cost', 'engineering council',
                        ])),
                        other_expenditures=parse_decimal(get_value([
                            'other expenditures', 'other expenses', 'others',
                        ])),
                    )
                    imported_count += 1

                if imported_count:
                    messages.success(
                        request,
                        f'Successfully imported sheet "{sheet.title}" with {imported_count} employees.'
                    )
                else:
                    sheet.delete()
                    messages.warning(
                        request,
                        'No employees were found. Check that your Excel file has '
                        'an "Employees Name" or "Name" column.'
                    )

                return redirect('manpower:sheet_list')

            except Exception as e:
                messages.error(request, f'Error importing file: {str(e)}')
    else:
        form = ImportForm()

    return render(request, 'manpower/sheet_import.html', {'form': form})


@login_required
def sheet_export(request, pk):
    if not (request.user.is_super_admin_user or request.user.is_admin_user):
        messages.error(request, 'You do not have permission to export.')
        return redirect('manpower:sheet_list')

    sheet = get_object_or_404(ManpowerSheet, pk=pk)
    items = sheet.line_items.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet.title[:31]

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    currency_format = '#,##0.00'

    # Sheet info rows
    ws.cell(row=1, column=1, value='Title:').font = bold_font
    ws.cell(row=1, column=2, value=sheet.title)
    ws.cell(row=2, column=1, value='Project Reference:').font = bold_font
    ws.cell(row=2, column=2, value=sheet.project_reference or '-')
    ws.cell(row=3, column=1, value='Date:').font = bold_font
    ws.cell(row=3, column=2, value=sheet.date.strftime('%d %b %Y'))
    if sheet.notes:
        ws.cell(row=4, column=1, value='Notes:').font = bold_font
        ws.cell(row=4, column=2, value=sheet.notes)

    # Data header row
    data_start = 6
    data_headers = [
        'Sr. No.', 'Employee Name', 'Department', 'Classification of Staff',
        'Location/Project', 'Designation', 'DOJ', 'Demobilization Date',
        'Date of Birth', 'Employee Age',
        # Costing
        'Budget Cost', 'Gross Salary', 'Iqama Cost',
        'Services Transfer & Visa Fee', 'GOSI Cost', 'Vacation Pay',
        'EXE Cost', 'EOSB', 'Air Ticket', 'Insurance Cost',
        'Project Allowance', 'PPE', 'Engineering Council Cost',
        'Other Expenditures',
        # Totals
        'Total Yearly Cost', 'Monthly Cost', 'Daily Cost', 'Hourly Cost',
    ]
    for col, header in enumerate(data_headers, 1):
        cell = ws.cell(row=data_start, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_num, item in enumerate(items, 1):
        r = data_start + row_num
        data = [
            row_num,
            item.employee_name,
            item.department,
            item.classification,
            item.location_project,
            item.designation,
            item.doj.strftime('%Y-%m-%d') if item.doj else '',
            item.demobilization_date.strftime('%Y-%m-%d') if item.demobilization_date else '',
            item.date_of_birth.strftime('%Y-%m-%d') if item.date_of_birth else '',
            item.employee_age or '',
            float(item.budget_cost),
            float(item.gross_salary),
            float(item.iqama_cost),
            float(item.service_transfer_visa_fee),
            float(item.gosi_cost),
            float(item.vacation_pay),
            float(item.exe_cost),
            float(item.eosb),
            float(item.air_ticket),
            float(item.insurance_cost),
            float(item.project_allowance),
            float(item.ppe),
            float(item.engineering_council_cost),
            float(item.other_expenditures),
            float(item.total_yearly_cost),
            float(item.monthly_cost),
            float(item.daily_cost),
            float(item.hourly_cost),
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=r, column=col, value=value)
            cell.border = thin_border
            # Apply currency format to cost columns (11-28)
            if col >= 11:
                cell.number_format = currency_format

    # Grand total row
    total_row = data_start + len(items) + 1
    for col in range(1, len(data_headers) + 1):
        ws.cell(row=total_row, column=col, value='').border = thin_border
    total_label = ws.cell(row=total_row, column=2, value='GRAND TOTAL')
    total_label.font = Font(bold=True, size=11)

    grand_yearly = sum(float(i.total_yearly_cost) for i in items)
    grand_monthly = sum(float(i.monthly_cost) for i in items)
    grand_daily = sum(float(i.daily_cost) for i in items)
    grand_hourly = sum(float(i.hourly_cost) for i in items)

    for col_idx, val in [(25, grand_yearly), (26, grand_monthly), (27, grand_daily), (28, grand_hourly)]:
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(bold=True, size=11)
        cell.number_format = currency_format
        cell.border = thin_border

    # Column widths
    column_widths = [
        8, 25, 16, 18, 18, 18, 14, 14, 14, 10,
        14, 14, 12, 18, 12, 12, 12, 12, 12, 14,
        14, 10, 18, 16,
        16, 14, 12, 12,
    ]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    safe_title = sheet.title.replace(' ', '_')[:30]
    filename = f'manpower_{safe_title}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
