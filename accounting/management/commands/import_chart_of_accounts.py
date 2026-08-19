"""Import / refresh the Chart of Accounts from the finance team's workbook.

Re-runnable: accounts are matched on `code`, so a second run updates names and
types rather than duplicating. Nothing is ever deleted — accounts missing from
the file can be deactivated with --deactivate-missing, which keeps historic
references intact.

    python manage.py import_chart_of_accounts "ERP final_Babar.xlsx"
    python manage.py import_chart_of_accounts <file> --dry-run
    python manage.py import_chart_of_accounts <file> --sheet changing
"""
import re

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounting.models import Account

# Workbook layout (1-based columns), as sent by finance:
#   B Reporting code #   C Balance Sheet Description
#   D G/L code #         E G/L Description          F Internal Type
COL_REPORTING_CODE = 2
COL_REPORTING_NAME = 3
COL_GL_CODE = 4
COL_GL_NAME = 5
COL_INTERNAL_TYPE = 6
FIRST_DATA_ROW = 3          # rows 1-2 are the title and the header

VALID_TYPES = {t for t, _ in Account.INTERNAL_TYPE_CHOICES}


def _clean(value):
    """Trim a cell and collapse embedded newlines/runs of whitespace."""
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()


def _ancestor_codes(code):
    """Candidate parent codes, nearest first.

    Zeroes the trailing digits one at a time: 1200001 -> 1200000, then
    1200000 -> ... -> 1000000. Yielding nearest-first lets the caller pick the
    closest ancestor that actually exists, so a chart with a missing
    intermediate heading still links up instead of becoming a root.

    Candidates are only *shapes*, not guaranteed parents: zeroing one digit of
    4100025 gives 4100020, which in this chart is a real leaf account (Rental
    Equipment) rather than a heading. The caller must therefore accept only
    candidates that are headings — see _resolve_parent.
    """
    for k in range(1, len(code)):
        candidate = code[:len(code) - k] + '0' * k
        if candidate != code:
            yield candidate


def _resolve_parent(code, by_code):
    """Nearest ancestor of `code` that is a heading, or None for a root.

    Only `View` accounts may parent others — that is exactly what the type
    means in this chart. Without the check, an account like 4100025 would be
    filed under its sibling 4100020 simply because zeroing a digit matched a
    code that happens to exist.
    """
    for candidate in _ancestor_codes(code):
        account = by_code.get(candidate)
        if account is not None and account.internal_type == Account.TYPE_VIEW:
            return account
    return None


class Command(BaseCommand):
    help = 'Import or refresh the Chart of Accounts from an .xlsx workbook.'

    def add_arguments(self, parser):
        parser.add_argument('path', help='Path to the .xlsx file.')
        parser.add_argument('--sheet', default=None,
                            help='Worksheet name (default: the first sheet).')
        parser.add_argument('--dry-run', action='store_true',
                            help='Report what would change without writing.')
        parser.add_argument('--deactivate-missing', action='store_true',
                            help='Mark accounts absent from the file inactive.')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as exc:                      # pragma: no cover
            raise CommandError('openpyxl is required to read .xlsx files.') from exc

        path, dry_run = options['path'], options['dry_run']
        try:
            workbook = openpyxl.load_workbook(path, data_only=True)
        except FileNotFoundError as exc:
            raise CommandError(f'File not found: {path}') from exc

        sheet_name = options['sheet']
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise CommandError(
                f'Sheet {sheet_name!r} not in workbook. Available: {workbook.sheetnames}')
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        self.stdout.write(f'Reading {path} — sheet {sheet.title!r}')

        parsed, duplicates, bad_types = self._parse(sheet)
        if not parsed:
            raise CommandError('No account rows found — check the sheet layout.')

        for row, code, first_row in duplicates:
            self.stdout.write(self.style.WARNING(
                f'  ! row {row}: duplicate code {code} (first seen row {first_row}) — skipped'))
        for row, code, value in bad_types:
            self.stdout.write(self.style.WARNING(
                f'  ! row {row}: account {code} has unrecognised Internal Type '
                f'{value!r} — imported as {Account.TYPE_REGULAR}'))

        if dry_run:
            self._report_plan(parsed)
            self.stdout.write(self.style.WARNING('\nDry run — nothing written.'))
            return

        with transaction.atomic():
            created, updated = self._upsert(parsed)
            linked, orphans = self._link_parents(parsed)
            deactivated = self._deactivate_missing(parsed) if options['deactivate_missing'] else 0

        self.stdout.write(self.style.SUCCESS(
            f'\nImported {len(parsed)} accounts: {created} created, {updated} updated.'))
        self.stdout.write(f'Parent links set: {linked}.')
        if orphans:
            self.stdout.write(self.style.WARNING(
                f'Top-level (no parent found): {", ".join(orphans)}'))
        if deactivated:
            self.stdout.write(self.style.WARNING(
                f'Deactivated {deactivated} account(s) absent from the file.'))
        self._report_plan(parsed)

    # ── parsing ──────────────────────────────────────────────────────────
    def _parse(self, sheet):
        """Return (rows, duplicates, bad_types).

        `rows` is a list of dicts keyed by code, in sheet order. Heading rows
        carry the reporting code/name in columns B/C and repeat the same value
        in D/E; leaf rows fill only D/E. Either way the G/L pair wins.
        """
        parsed, seen, duplicates, bad_types = [], {}, [], []
        for row_no in range(FIRST_DATA_ROW, sheet.max_row + 1):
            reporting_code = _clean(sheet.cell(row_no, COL_REPORTING_CODE).value)
            reporting_name = _clean(sheet.cell(row_no, COL_REPORTING_NAME).value)
            gl_code = _clean(sheet.cell(row_no, COL_GL_CODE).value)
            gl_name = _clean(sheet.cell(row_no, COL_GL_NAME).value)
            internal_type = _clean(sheet.cell(row_no, COL_INTERNAL_TYPE).value)

            code = gl_code or reporting_code
            name = gl_name or reporting_name
            if not code:
                continue
            # Excel hands back numeric codes as floats ("1110001.0").
            code = code[:-2] if code.endswith('.0') else code
            if not code.isdigit():
                continue

            if code in seen:
                duplicates.append((row_no, code, seen[code]))
                continue
            seen[code] = row_no

            if internal_type not in VALID_TYPES:
                if internal_type:
                    bad_types.append((row_no, code, internal_type))
                internal_type = Account.TYPE_REGULAR

            parsed.append({
                'code': code,
                'name': name or code,
                'internal_type': internal_type,
                'reporting_code': reporting_code if reporting_code != code else '',
                'reporting_name': reporting_name if reporting_name != name else '',
                'source_row': row_no,
            })
        return parsed, duplicates, bad_types

    # ── writing ──────────────────────────────────────────────────────────
    def _upsert(self, parsed):
        created = updated = 0
        for item in parsed:
            _, was_created = Account.objects.update_or_create(
                code=item['code'],
                defaults={
                    'name': item['name'],
                    'internal_type': item['internal_type'],
                    'reporting_code': item['reporting_code'],
                    'reporting_name': item['reporting_name'],
                    'source_row': item['source_row'],
                    'is_active': True,
                })
            created += was_created
            updated += not was_created
        return created, updated

    def _link_parents(self, parsed):
        """Point every account at its nearest existing ancestor.

        Done as a second pass so ordering in the sheet never matters: by now
        every account in the file exists, whether it was created just now or
        on an earlier run.
        """
        by_code = {a.code: a for a in Account.objects.all()}
        linked, orphans, to_update = 0, [], []
        for item in parsed:
            account = by_code[item['code']]
            parent = _resolve_parent(item['code'], by_code)
            if parent is not None:
                linked += 1
            else:
                orphans.append(item['code'])
            if account.parent_id != getattr(parent, 'pk', None):
                account.parent = parent
                to_update.append(account)
        if to_update:
            Account.objects.bulk_update(to_update, ['parent'])
        return linked, orphans

    def _deactivate_missing(self, parsed):
        codes = {item['code'] for item in parsed}
        return Account.objects.exclude(code__in=codes).filter(is_active=True).update(
            is_active=False)

    # ── reporting ────────────────────────────────────────────────────────
    def _report_plan(self, parsed):
        by_type, by_class = {}, {}
        for item in parsed:
            by_type[item['internal_type']] = by_type.get(item['internal_type'], 0) + 1
            label = Account.CLASS_LABELS.get(item['code'][:1], 'Unclassified')
            by_class[label] = by_class.get(label, 0) + 1
        self.stdout.write('\nBy internal type:')
        for key in sorted(by_type):
            self.stdout.write(f'  {key:12s} {by_type[key]:4d}')
        self.stdout.write('By class:')
        for key in sorted(by_class):
            self.stdout.write(f'  {key:26s} {by_class[key]:4d}')
