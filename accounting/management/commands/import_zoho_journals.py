"""Import Zoho Books journal entries as JV vouchers.

Zoho remains the statutory book; this brings its entries across so the ERP's
account ledgers are complete. Invoices and bills raised in Zoho already produced
journal entries there, so importing them here is what makes those documents
appear against their revenue and expense accounts.

Re-runnable: vouchers are matched on their journal number, so a second run
updates rather than duplicating.

    python manage.py import_zoho_journals Journals.csv
    python manage.py import_zoho_journals Journals.xlsx --dry-run

Column names are matched loosely (case- and space-insensitive, several known
aliases each), because Zoho's export headers vary by version and locale. If a
required column can't be found the command lists the headers it actually saw
rather than failing with a KeyError.
"""
import csv
import os
import re
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from accounting.models import Account, Partner, Voucher, VoucherLine

# Each field maps to the header aliases we accept. First match wins.
COLUMN_ALIASES = {
    'number':      ['journal number', 'journal no', 'journal#', 'entry number',
                    'reference number', 'journal id'],
    'date':        ['journal date', 'date', 'entry date', 'transaction date'],
    'account_code': ['account code', 'account number', 'gl code', 'code'],
    'account_name': ['account', 'account name', 'ledger', 'ledger name'],
    'description': ['description', 'notes', 'narration', 'memo', 'line description'],
    'debit':       ['debit', 'debit amount', 'debits'],
    'credit':      ['credit', 'credit amount', 'credits'],
    'partner':     ['contact name', 'contact', 'customer name', 'vendor name',
                    'partner', 'name'],
    'currency':    ['currency', 'currency code'],
    'status':      ['status', 'journal status'],
}
REQUIRED = ('number', 'date', 'debit', 'credit')


def _norm(header):
    return re.sub(r'[^a-z0-9]+', ' ', str(header or '').strip().lower()).strip()


def _to_decimal(value):
    """Parse a money cell. Blank, '-' and unparseable text all mean zero —
    Zoho writes an empty cell for the unused side of every line."""
    if value is None:
        return Decimal('0')
    text = str(value).strip().replace(',', '').replace(' ', '')
    if not text or text == '-':
        return Decimal('0')
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal('0')


def _to_date(value):
    from datetime import date, datetime
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or '').strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%d %b %Y',
                '%d/%m/%y', '%Y/%m/%d'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = 'Import Zoho Books journal entries as JV vouchers.'

    def add_arguments(self, parser):
        parser.add_argument('path', help='Journals export (.csv or .xlsx).')
        parser.add_argument('--sheet', default=None, help='Worksheet name for .xlsx.')
        parser.add_argument('--dry-run', action='store_true',
                            help='Report what would happen without writing.')
        parser.add_argument('--post', action='store_true',
                            help='Post imported vouchers instead of leaving them draft.')

    def handle(self, *args, **options):
        path = options['path']
        if not os.path.exists(path):
            raise CommandError(f'File not found: {path}')

        rows, headers = self._read(path, options['sheet'])

        # Headers are checked before rows: if the columns can't be recognised,
        # that is the actionable problem, and reporting "no data rows" for a
        # file full of data under unexpected headers sends you looking in the
        # wrong place.
        mapping = self._map_columns(headers)
        missing = [f for f in REQUIRED if f not in mapping]
        if missing:
            raise CommandError(
                f'Could not find column(s) for: {", ".join(missing)}.\n'
                f'Headers in the file: {", ".join(str(h) for h in headers)}\n'
                f'Add the right alias to COLUMN_ALIASES in this command.')

        if not rows:
            raise CommandError('Headers look right, but the file has no data rows.')

        grouped, problems = self._group(rows, mapping)
        for problem in problems:
            self.stdout.write(self.style.WARNING(f'  ! {problem}'))

        if options['dry_run']:
            self._report(grouped)
            self.stdout.write(self.style.WARNING('\nDry run — nothing written.'))
            return

        with transaction.atomic():
            created, updated, skipped = self._write(grouped, post=options['post'])

        self.stdout.write(self.style.SUCCESS(
            f'\n{created} voucher(s) created, {updated} updated, {skipped} skipped.'))
        self._report(grouped)

    # ── reading ──────────────────────────────────────────────────────────
    def _read(self, path, sheet_name):
        if path.lower().endswith(('.xlsx', '.xlsm')):
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            if sheet_name and sheet_name not in wb.sheetnames:
                raise CommandError(
                    f'Sheet {sheet_name!r} not found. Available: {wb.sheetnames}')
            ws = wb[sheet_name] if sheet_name else wb.worksheets[0]
            data = list(ws.iter_rows(values_only=True))
            if not data:
                return [], []
            headers = list(data[0])
            return [dict(zip(headers, r)) for r in data[1:]], headers
        with open(path, newline='', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            return list(reader), (reader.fieldnames or [])

    def _map_columns(self, headers):
        """Resolve each field to an actual header in the file."""
        by_norm = {_norm(h): h for h in headers if h is not None}
        mapping = {}
        for field, aliases in COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in by_norm:
                    mapping[field] = by_norm[alias]
                    break
        return mapping

    # ── grouping ─────────────────────────────────────────────────────────
    def _group(self, rows, mapping):
        """Collapse the flat export into {journal number: {header, lines}}.

        Zoho writes one row per journal LINE, repeating the header fields, so
        the number is what stitches a voucher back together.
        """
        def cell(row, field):
            key = mapping.get(field)
            return row.get(key) if key else None

        grouped, problems = {}, []
        accounts_by_code = {a.code: a for a in Account.objects.all()}
        accounts_by_name = {a.name.strip().lower(): a for a in Account.objects.all()}

        for index, row in enumerate(rows, start=2):
            number = str(cell(row, 'number') or '').strip()
            if not number:
                continue
            date = _to_date(cell(row, 'date'))
            if date is None:
                problems.append(f'row {index}: unreadable date for journal {number} — skipped')
                continue

            code = str(cell(row, 'account_code') or '').strip()
            code = code[:-2] if code.endswith('.0') else code
            name = str(cell(row, 'account_name') or '').strip()
            account = accounts_by_code.get(code) or accounts_by_name.get(name.lower())
            if account is None:
                problems.append(
                    f'row {index}: journal {number} references unknown account '
                    f'{code or name!r} — line skipped')
                continue
            if not account.is_postable:
                problems.append(
                    f'row {index}: {account.code} {account.name} is a heading — line skipped')
                continue

            debit, credit = _to_decimal(cell(row, 'debit')), _to_decimal(cell(row, 'credit'))
            if not debit and not credit:
                continue

            entry = grouped.setdefault(number, {
                'number': number, 'date': date,
                'narration': str(cell(row, 'description') or '').strip(),
                'currency': str(cell(row, 'currency') or 'SAR').strip() or 'SAR',
                'partner_name': str(cell(row, 'partner') or '').strip(),
                'lines': [],
            })
            entry['lines'].append({
                'account': account,
                'description': str(cell(row, 'description') or '').strip()[:255],
                'debit': debit, 'credit': credit,
            })
        return grouped, problems

    # ── writing ──────────────────────────────────────────────────────────
    def _write(self, grouped, post=False):
        created = updated = skipped = 0
        for number, entry in grouped.items():
            debit = sum((l['debit'] for l in entry['lines']), Decimal('0'))
            credit = sum((l['credit'] for l in entry['lines']), Decimal('0'))
            if debit != credit:
                self.stdout.write(self.style.WARNING(
                    f'  ! journal {number} does not balance '
                    f'(Dr {debit} vs Cr {credit}) — skipped'))
                skipped += 1
                continue

            partner = None
            if entry['partner_name']:
                partner, _ = Partner.objects.get_or_create(
                    name=entry['partner_name'],
                    defaults={'kind': Partner.KIND_BOTH, 'notes': 'Created by Zoho journal import.'})

            voucher, was_created = Voucher.objects.update_or_create(
                voucher_type=Voucher.TYPE_JV, number=number,
                defaults={
                    'date': entry['date'],
                    'narration': entry['narration'],
                    'partner': partner,
                    'currency': entry['currency'],
                    'source': 'zoho',
                    # An imported voucher lands as draft unless --post is given,
                    # so a bad export can be reviewed before it reaches a ledger.
                    'status': Voucher.STATUS_POSTED if post else Voucher.STATUS_DRAFT,
                })
            # Replace the lines wholesale: a re-import must reflect the file,
            # not merge with whatever was there before.
            voucher.lines.all().delete()
            for order, line in enumerate(entry['lines']):
                VoucherLine.objects.create(
                    voucher=voucher, account=line['account'], partner=partner,
                    description=line['description'], debit=line['debit'],
                    credit=line['credit'], order=order)
            created += was_created
            updated += not was_created
        return created, updated, skipped

    def _report(self, grouped):
        total_lines = sum(len(e['lines']) for e in grouped.values())
        value = sum(
            sum((l['debit'] for l in e['lines']), Decimal('0'))
            for e in grouped.values())
        self.stdout.write(f'\nJournals: {len(grouped)}  ·  lines: {total_lines}')
        self.stdout.write(f'Total debits: {value:,.2f}')
