import datetime
import tempfile
from decimal import Decimal
from io import StringIO
from pathlib import Path

import openpyxl
from django.core.exceptions import ValidationError
from django.core.management import call_command
from django.core.management.base import CommandError
from django.db import IntegrityError
from django.test import TestCase
from django.urls import reverse

from accounts.models import Role, User
from accounting.models import (
    Account, Document, DocumentLine, Partner, Voucher, VoucherLine, build_tree,
)

# (reporting_code, reporting_name, gl_code, gl_name, internal_type) — the five
# populated columns of the finance workbook, starting at column B.
SAMPLE_ROWS = [
    ('1000000', 'Assets', '1000000', 'Assets', 'View'),
    ('1100000', 'Current Assets', '1100000', 'Current Assets', 'View'),
    ('1110000', 'Cash in hand', '1110000', 'Cash in hand', 'View'),
    (None, None, '1110001', 'Cash in hand', 'Regular'),
    (None, None, '1110002', 'Petty cash', 'Regular'),
    ('1120000', 'Cash at Banks', '1120000', 'Cash at Banks', 'View'),
    (None, None, '1120001', 'ALINMA Bank', 'Liquidity'),
    ('4000000', 'Cost of Sale', '4000000', 'Cost of Sale', 'View'),
    ('4100000', 'Cost of Projects', '4100000', 'Cost of Projects', 'View'),
    (None, None, '4100020', 'Rental Equipment', 'Regular'),
    # 4100025 is the regression guard: zeroing its last digit yields 4100020,
    # which is a real *sibling* leaf above. It must still land under 4100000.
    (None, None, '4100025', 'Material Logistic & Freight', 'Regular'),
]


def write_workbook(rows, path, sheet_name='changing'):
    """Build a workbook shaped like the finance file: title row, header row,
    then data from row 3 in columns B–F."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.cell(1, 1, 'Chart of Accounts')
    for col, label in enumerate(['Reporting code #', 'Balance Sheet Description',
                                 'G/L code #', 'G/L Description', 'Internal Type'], start=2):
        ws.cell(2, col, label)
    for i, row in enumerate(rows, start=3):
        for col, value in enumerate(row, start=2):
            if value is not None:
                ws.cell(i, col, value)
    wb.save(path)
    return path


class ChartImportTests(TestCase):
    """The import_chart_of_accounts command."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        self.path = write_workbook(SAMPLE_ROWS, str(Path(self.dir) / 'coa.xlsx'))

    def _import(self, *args):
        out = StringIO()
        call_command('import_chart_of_accounts', self.path, *args, stdout=out)
        return out.getvalue()

    def test_imports_every_account(self):
        self._import()
        self.assertEqual(Account.objects.count(), len(SAMPLE_ROWS))
        bank = Account.objects.get(code='1120001')
        self.assertEqual(bank.name, 'ALINMA Bank')
        self.assertEqual(bank.internal_type, Account.TYPE_LIQUIDITY)

    def test_rerun_updates_rather_than_duplicating(self):
        self._import()
        rows = [list(r) for r in SAMPLE_ROWS]
        rows[4][3] = 'Petty cash (renamed)'
        write_workbook(rows, self.path)
        output = self._import()
        self.assertEqual(Account.objects.count(), len(SAMPLE_ROWS))
        self.assertEqual(Account.objects.get(code='1110002').name, 'Petty cash (renamed)')
        self.assertIn('0 created', output)

    def test_parent_is_nearest_heading_not_a_sibling_leaf(self):
        """Regression: 4100025 must not be filed under the leaf 4100020."""
        self._import()
        self.assertEqual(Account.objects.get(code='4100025').parent.code, '4100000')
        self.assertEqual(Account.objects.get(code='4100020').parent.code, '4100000')

    def test_parent_chain_and_roots(self):
        self._import()
        self.assertEqual(Account.objects.get(code='1110001').parent.code, '1110000')
        self.assertEqual(Account.objects.get(code='1110000').parent.code, '1100000')
        self.assertIsNone(Account.objects.get(code='1000000').parent)
        roots = set(Account.objects.filter(parent__isnull=True).values_list('code', flat=True))
        self.assertEqual(roots, {'1000000', '4000000'})

    def test_no_regular_account_ever_parents_another(self):
        self._import()
        for account in Account.objects.exclude(internal_type=Account.TYPE_VIEW):
            self.assertFalse(account.children.exists(), f'{account.code} has children')

    def test_duplicate_code_keeps_first_and_warns(self):
        rows = SAMPLE_ROWS + [(None, None, '1110001', 'Impostor', 'Regular')]
        write_workbook(rows, self.path)
        output = self._import()
        self.assertIn('duplicate code 1110001', output)
        self.assertEqual(Account.objects.get(code='1110001').name, 'Cash in hand')

    def test_unknown_internal_type_falls_back_to_regular(self):
        rows = SAMPLE_ROWS + [(None, None, '1110003', 'Odd one', 'Nonsense')]
        write_workbook(rows, self.path)
        output = self._import()
        self.assertIn('unrecognised Internal Type', output)
        self.assertEqual(Account.objects.get(code='1110003').internal_type,
                         Account.TYPE_REGULAR)

    def test_dry_run_writes_nothing(self):
        output = self._import('--dry-run')
        self.assertIn('Dry run', output)
        self.assertEqual(Account.objects.count(), 0)

    def test_deactivate_missing_keeps_the_row(self):
        self._import()
        stale = Account.objects.create(code='9999999', name='Gone', internal_type='Regular')
        self._import('--deactivate-missing')
        stale.refresh_from_db()
        self.assertFalse(stale.is_active)
        self.assertTrue(Account.objects.get(code='1110001').is_active)

    def test_missing_file_raises(self):
        with self.assertRaises(CommandError):
            call_command('import_chart_of_accounts', str(Path(self.dir) / 'nope.xlsx'),
                         stdout=StringIO())

    def test_unknown_sheet_raises(self):
        with self.assertRaises(CommandError):
            call_command('import_chart_of_accounts', self.path, '--sheet', 'missing',
                         stdout=StringIO())


class AccountModelTests(TestCase):

    def test_is_postable_excludes_only_view(self):
        view = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        leaf = Account.objects.create(code='1110001', name='Cash', internal_type='Regular',
                                      parent=view)
        self.assertFalse(view.is_postable)
        self.assertTrue(leaf.is_postable)
        self.assertEqual(list(Account.objects.postable()), [leaf])
        self.assertEqual(list(Account.objects.headings()), [view])

    def test_account_class_from_leading_digit(self):
        self.assertEqual(
            Account(code='4100001', name='x').account_class, 'Cost of Sale')
        self.assertEqual(Account(code='9100001', name='x').account_class, '')

    def test_build_tree_depths(self):
        a = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        b = Account.objects.create(code='1100000', name='Current', internal_type='View', parent=a)
        c = Account.objects.create(code='1110001', name='Cash', internal_type='Regular', parent=b)
        depths = {x.code: x.tree_depth for x in build_tree(Account.objects.all())}
        self.assertEqual(depths, {'1000000': 0, '1100000': 1, '1110001': 2})

    def test_build_tree_treats_filtered_out_parents_as_roots(self):
        """A searched/filtered list must not render orphaned indentation."""
        a = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        c = Account.objects.create(code='1110001', name='Cash', internal_type='Regular', parent=a)
        rows = build_tree(Account.objects.filter(pk=c.pk))
        self.assertEqual(rows[0].tree_depth, 0)


class ChartViewTests(TestCase):

    def setUp(self):
        self.view = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        self.cash = Account.objects.create(code='1110001', name='Petty cash',
                                           internal_type='Regular', parent=self.view)
        self.bank = Account.objects.create(code='1120001', name='ALINMA Bank',
                                           internal_type='Liquidity', parent=self.view)

    def _user(self, role_name, username):
        role, _ = Role.objects.get_or_create(name=role_name)
        user = User.objects.create_user(username, password='x')
        user.role = role
        user.save()
        return user

    def test_finance_and_super_admin_can_open(self):
        for role_name, username in [(Role.SUPER_ADMIN, 'sa'), (Role.FINANCE_HEAD, 'fh'),
                                    (Role.FINANCE_MANAGER, 'fm'), (Role.FINANCE_REP, 'fr')]:
            self.client.force_login(self._user(role_name, username))
            self.assertEqual(self.client.get(reverse('accounting:chart')).status_code, 200,
                             f'{role_name} should be allowed')

    def test_other_roles_are_denied(self):
        self.client.force_login(self._user(Role.SALES_REP, 'rep'))
        self.assertEqual(self.client.get(reverse('accounting:chart')).status_code, 403)

    def test_anonymous_is_redirected_to_login(self):
        self.assertEqual(self.client.get(reverse('accounting:chart')).status_code, 302)

    def test_lists_accounts_with_totals(self):
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'))
        self.assertContains(resp, 'ALINMA Bank')
        self.assertEqual(resp.context['total_count'], 3)
        self.assertEqual(resp.context['postable_count'], 2)
        self.assertEqual(resp.context['heading_count'], 1)

    def test_search_filters_by_code_and_name(self):
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'), {'q': 'ALINMA'})
        self.assertEqual([a.code for a in resp.context['rows']], ['1120001'])
        resp = self.client.get(reverse('accounting:chart'), {'q': '1110001'})
        self.assertEqual([a.code for a in resp.context['rows']], ['1110001'])

    def test_type_and_postable_filters(self):
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'), {'type': 'Liquidity'})
        self.assertEqual([a.code for a in resp.context['rows']], ['1120001'])
        resp = self.client.get(reverse('accounting:chart'), {'postable': '0'})
        self.assertEqual([a.code for a in resp.context['rows']], ['1000000'])

    def test_inactive_hidden_unless_requested(self):
        self.bank.is_active = False
        self.bank.save()
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'))
        self.assertNotIn('1120001', [a.code for a in resp.context['rows']])
        resp = self.client.get(reverse('accounting:chart'), {'inactive': '1'})
        self.assertIn('1120001', [a.code for a in resp.context['rows']])

    def test_empty_state_when_nothing_imported(self):
        # Children first: parent is PROTECTed, so a heading cannot be deleted
        # out from under its accounts.
        Account.objects.filter(parent__isnull=False).delete()
        Account.objects.all().delete()
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'))
        self.assertContains(resp, 'import_chart_of_accounts')

    def test_heading_with_children_cannot_be_deleted(self):
        from django.db.models.deletion import ProtectedError
        with self.assertRaises(ProtectedError):
            self.view.delete()

    def test_rows_carry_subtree_counts(self):
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'))
        by_code = {a.code: a.counts for a in resp.context['rows']}
        self.assertEqual(by_code['1000000']['direct'], 2)
        self.assertEqual(by_code['1000000']['postable'], 2)
        self.assertEqual(by_code['1110001']['direct'], 0)

    def test_counts_ignore_the_active_filter(self):
        """A heading reports its true size even while the list is filtered."""
        self.client.force_login(self._user(Role.SUPER_ADMIN, 'sa'))
        resp = self.client.get(reverse('accounting:chart'), {'type': 'View'})
        counts = {a.code: a.counts for a in resp.context['rows']}
        self.assertEqual(counts['1000000']['direct'], 2)


class AccountDetailTests(TestCase):
    """Drill-down: a heading lists what is filed beneath it."""

    def setUp(self):
        self.assets = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        self.current = Account.objects.create(code='1100000', name='Current Assets',
                                              internal_type='View', parent=self.assets)
        self.banks = Account.objects.create(code='1120000', name='Cash at Banks',
                                            internal_type='View', parent=self.current)
        self.alinma = Account.objects.create(code='1120001', name='ALINMA Bank',
                                             internal_type='Liquidity', parent=self.banks)
        self.sab = Account.objects.create(code='1120002', name='SAB SAR BANK',
                                          internal_type='Liquidity', parent=self.banks)
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('sa', password='x')
        self.user.role = role
        self.user.save()
        self.client.force_login(self.user)

    def _get(self, code):
        return self.client.get(reverse('accounting:account_detail', args=[code]))

    def test_heading_lists_its_children(self):
        resp = self._get('1120000')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual([c.code for c in resp.context['children']], ['1120001', '1120002'])
        self.assertContains(resp, 'ALINMA Bank')

    def test_breadcrumb_runs_outermost_first(self):
        resp = self._get('1120001')
        self.assertEqual([a.code for a in resp.context['ancestors']],
                         ['1000000', '1100000', '1120000'])

    def test_leaf_shows_siblings_instead(self):
        resp = self._get('1120001')
        self.assertEqual(resp.context['children'], [])
        self.assertEqual([s.code for s in resp.context['siblings']], ['1120002'])

    def test_counts_cover_the_whole_subtree(self):
        resp = self._get('1000000')
        self.assertEqual(resp.context['counts'],
                         {'direct': 1, 'total': 4, 'postable': 2})

    def test_top_level_has_no_parent_or_siblings(self):
        resp = self._get('1000000')
        self.assertIsNone(resp.context['account'].parent)
        self.assertEqual(resp.context['siblings'], [])

    def test_unknown_code_404s(self):
        self.assertEqual(self._get('9999999').status_code, 404)

    def test_requires_finance_or_super_admin(self):
        role, _ = Role.objects.get_or_create(name=Role.SALES_REP)
        rep = User.objects.create_user('rep', password='x')
        rep.role = role
        rep.save()
        self.client.force_login(rep)
        self.assertEqual(self._get('1000000').status_code, 403)


class VoucherTests(TestCase):
    """JV / PV / RV — the double-entry records."""

    def setUp(self):
        self.assets = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        self.bank = Account.objects.create(code='1120001', name='ALINMA Bank',
                                           internal_type='Liquidity', parent=self.assets)
        self.expense = Account.objects.create(code='4100006', name='Local Procurement',
                                              internal_type='Regular', parent=self.assets)
        self.partner = Partner.objects.create(name='Al Ghad Al Taqni Est',
                                              kind=Partner.KIND_VENDOR)

    def _voucher(self, vtype=Voucher.TYPE_PV, number='PV-0001'):
        return Voucher.objects.create(voucher_type=vtype, number=number,
                                      date=datetime.date(2026, 8, 1))

    def _balance(self, voucher, amount='1000.00'):
        VoucherLine.objects.create(voucher=voucher, account=self.expense,
                                   debit=Decimal(amount), order=0)
        VoucherLine.objects.create(voucher=voucher, account=self.bank,
                                   credit=Decimal(amount), order=1)

    def test_balanced_voucher_posts(self):
        v = self._voucher()
        self._balance(v)
        self.assertTrue(v.is_balanced)
        v.post()
        v.refresh_from_db()
        self.assertEqual(v.status, Voucher.STATUS_POSTED)

    def test_unbalanced_voucher_refuses_to_post(self):
        v = self._voucher()
        VoucherLine.objects.create(voucher=v, account=self.expense, debit=Decimal('1000'))
        VoucherLine.objects.create(voucher=v, account=self.bank, credit=Decimal('900'))
        self.assertFalse(v.is_balanced)
        with self.assertRaises(ValidationError):
            v.post()
        v.refresh_from_db()
        self.assertEqual(v.status, Voucher.STATUS_DRAFT)

    def test_empty_voucher_is_not_balanced_at_zero(self):
        """A voucher with no lines must not slip through as 'balances at 0'."""
        with self.assertRaises(ValidationError):
            self._voucher().post()

    def test_posting_twice_is_refused(self):
        v = self._voucher()
        self._balance(v)
        v.post()
        with self.assertRaises(ValidationError):
            v.post()

    def test_line_cannot_be_both_debit_and_credit(self):
        line = VoucherLine(voucher=self._voucher(), account=self.bank,
                           debit=Decimal('10'), credit=Decimal('10'))
        with self.assertRaises(ValidationError):
            line.clean()

    def test_line_needs_an_amount(self):
        with self.assertRaises(ValidationError):
            VoucherLine(voucher=self._voucher(), account=self.bank).clean()

    def test_line_rejects_negative_amounts(self):
        with self.assertRaises(ValidationError):
            VoucherLine(voucher=self._voucher(), account=self.bank,
                        debit=Decimal('-5')).clean()

    def test_line_cannot_post_to_a_heading(self):
        """View accounts group other accounts and must never carry a posting."""
        line = VoucherLine(voucher=self._voucher(), account=self.assets,
                           debit=Decimal('10'))
        with self.assertRaises(ValidationError):
            line.clean()

    def test_totals_and_headline_amount(self):
        v = self._voucher()
        self._balance(v, '2500.50')
        self.assertEqual(v.total_debit, Decimal('2500.50'))
        self.assertEqual(v.total_credit, Decimal('2500.50'))
        self.assertEqual(v.amount, Decimal('2500.50'))

    def test_number_unique_within_a_type_but_not_across(self):
        self._voucher(Voucher.TYPE_JV, '0001')
        self._voucher(Voucher.TYPE_RV, '0001')      # different type — fine
        with self.assertRaises(IntegrityError):
            self._voucher(Voucher.TYPE_JV, '0001')

    def test_all_three_voucher_types_exist(self):
        self.assertEqual({c[0] for c in Voucher.TYPE_CHOICES}, {'JV', 'PV', 'RV'})


class DocumentTests(TestCase):
    """Customer invoices and vendor bills — one model, two directions."""

    def setUp(self):
        root = Account.objects.create(code='3000000', name='Revenues', internal_type='View')
        self.income = Account.objects.create(code='3000101', name='Sales - Projects',
                                             internal_type='Regular', parent=root)
        self.customer = Partner.objects.create(name='Saudi Aramco', kind=Partner.KIND_CUSTOMER)
        self.vendor = Partner.objects.create(name='PT Neo Energy', kind=Partner.KIND_VENDOR)
        self.inv = Document.objects.create(
            kind=Document.KIND_INVOICE, number='INV-001', partner=self.customer,
            date=datetime.date(2026, 8, 1), due_date=datetime.date(2026, 9, 1),
            subtotal=Decimal('100000'), tax_total=Decimal('15000'),
            total=Decimal('115000'), status=Document.STATUS_OPEN)

    def test_invoices_and_bills_split_by_kind(self):
        Document.objects.create(kind=Document.KIND_BILL, number='BILL-001',
                                partner=self.vendor, date=datetime.date(2026, 8, 2))
        self.assertEqual([d.number for d in Document.objects.invoices()], ['INV-001'])
        self.assertEqual([d.number for d in Document.objects.bills()], ['BILL-001'])

    def test_balance_due_tracks_payment(self):
        self.assertEqual(self.inv.balance_due, Decimal('115000'))
        self.inv.amount_paid = Decimal('40000')
        self.assertEqual(self.inv.balance_due, Decimal('75000'))

    def test_overdue_only_while_money_is_outstanding(self):
        self.inv.due_date = datetime.date(2020, 1, 1)
        self.assertTrue(self.inv.is_overdue)
        self.inv.amount_paid = Decimal('115000')
        self.inv.status = Document.STATUS_PAID
        self.assertFalse(self.inv.is_overdue)

    def test_not_overdue_without_a_due_date(self):
        self.inv.due_date = None
        self.assertFalse(self.inv.is_overdue)

    def test_lines_total_surfaces_an_import_that_does_not_add_up(self):
        DocumentLine.objects.create(document=self.inv, description='Phase 1',
                                    account=self.income, amount=Decimal('60000'))
        DocumentLine.objects.create(document=self.inv, description='Phase 2',
                                    account=self.income, amount=Decimal('30000'))
        # Stored subtotal is 100,000 but the lines only reach 90,000. The gap
        # must stay visible rather than being silently reconciled away.
        self.assertEqual(self.inv.lines_total, Decimal('90000'))
        self.assertNotEqual(self.inv.lines_total, self.inv.subtotal)

    def test_document_line_cannot_post_to_a_heading(self):
        heading = Account.objects.get(code='3000000')
        with self.assertRaises(ValidationError):
            DocumentLine(document=self.inv, account=heading).clean()

    def test_number_unique_within_a_kind_but_not_across(self):
        Document.objects.create(kind=Document.KIND_BILL, number='INV-001',
                                partner=self.vendor, date=datetime.date(2026, 8, 3))
        with self.assertRaises(IntegrityError):
            Document.objects.create(kind=Document.KIND_INVOICE, number='INV-001',
                                    partner=self.customer, date=datetime.date(2026, 8, 4))

    def test_outstanding_excludes_paid_and_void(self):
        Document.objects.create(kind=Document.KIND_BILL, number='B-PAID',
                                partner=self.vendor, date=datetime.date(2026, 8, 5),
                                status=Document.STATUS_PAID)
        Document.objects.create(kind=Document.KIND_BILL, number='B-VOID',
                                partner=self.vendor, date=datetime.date(2026, 8, 6),
                                status=Document.STATUS_VOID)
        self.assertEqual([d.number for d in Document.objects.outstanding()], ['INV-001'])


class AccountLedgerTests(TestCase):
    """Every transaction against an account, with a running balance."""

    def setUp(self):
        self.assets = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        self.banks = Account.objects.create(code='1120000', name='Cash at Banks',
                                            internal_type='View', parent=self.assets)
        self.alinma = Account.objects.create(code='1120001', name='ALINMA Bank',
                                             internal_type='Liquidity', parent=self.banks)
        self.sab = Account.objects.create(code='1120002', name='SAB SAR Bank',
                                          internal_type='Liquidity', parent=self.banks)
        rev = Account.objects.create(code='3000000', name='Revenues', internal_type='View')
        self.income = Account.objects.create(code='3000101', name='Sales - Projects',
                                             internal_type='Regular', parent=rev)
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('sa', password='x')
        self.user.role = role
        self.user.save()
        self.client.force_login(self.user)

    def _voucher(self, number, day, status=Voucher.STATUS_POSTED):
        return Voucher.objects.create(
            voucher_type=Voucher.TYPE_RV, number=number,
            date=datetime.date(2026, 8, day), status=status)

    def _pair(self, voucher, debit_account, credit_account, amount):
        VoucherLine.objects.create(voucher=voucher, account=debit_account,
                                   debit=Decimal(amount), order=0)
        VoucherLine.objects.create(voucher=voucher, account=credit_account,
                                   credit=Decimal(amount), order=1)

    def _ledger(self, code, **params):
        return self.client.get(reverse('accounting:account_ledger', args=[code]), params)

    def test_running_balance_accumulates(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-2', 2), self.alinma, self.income, '250')
        resp = self._ledger('1120001')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual([r['balance'] for r in resp.context['rows']],
                         [Decimal('1000'), Decimal('1250')])
        self.assertEqual(resp.context['closing'], Decimal('1250'))

    def test_credit_account_balance_reads_positive(self):
        """Revenue is credit-natured, so its balance must not show negative."""
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        resp = self._ledger('3000101')
        self.assertEqual(resp.context['closing'], Decimal('1000'))

    def test_debit_account_reduced_by_a_credit(self):
        v = self._voucher('RV-1', 1)
        self._pair(v, self.alinma, self.income, '1000')
        v2 = self._voucher('RV-2', 2)
        self._pair(v2, self.income, self.alinma, '400')   # money back out of the bank
        resp = self._ledger('1120001')
        self.assertEqual(resp.context['closing'], Decimal('600'))

    def test_drafts_excluded_by_default(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-D', 2, Voucher.STATUS_DRAFT), self.alinma,
                   self.income, '9999')
        resp = self._ledger('1120001')
        self.assertEqual(len(resp.context['rows']), 1)
        self.assertEqual(resp.context['closing'], Decimal('1000'))

    def test_drafts_included_on_request(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-D', 2, Voucher.STATUS_DRAFT), self.alinma,
                   self.income, '500')
        resp = self._ledger('1120001', drafts='1')
        self.assertEqual(len(resp.context['rows']), 2)
        self.assertEqual(resp.context['closing'], Decimal('1500'))
        self.assertTrue(resp.context['include_drafts'])

    def test_opening_balance_carries_prior_activity(self):
        """A filtered ledger must not start from zero."""
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-2', 15), self.alinma, self.income, '500')
        resp = self._ledger('1120001', **{'from': '2026-08-10'})
        self.assertEqual(resp.context['opening'], Decimal('1000'))
        self.assertEqual(len(resp.context['rows']), 1)
        self.assertEqual(resp.context['closing'], Decimal('1500'))

    def test_date_to_excludes_later_activity(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-2', 20), self.alinma, self.income, '500')
        resp = self._ledger('1120001', to='2026-08-10')
        self.assertEqual(len(resp.context['rows']), 1)
        self.assertEqual(resp.context['closing'], Decimal('1000'))

    def test_heading_rolls_up_its_children(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-2', 2), self.sab, self.income, '250')
        resp = self._ledger('1120000')          # Cash at Banks — a View heading
        self.assertTrue(resp.context['is_rollup'])
        self.assertEqual(len(resp.context['rows']), 2)
        self.assertEqual(resp.context['closing'], Decimal('1250'))

    def test_leaf_is_not_a_rollup(self):
        resp = self._ledger('1120001')
        self.assertFalse(resp.context['is_rollup'])

    def test_period_totals(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        self._pair(self._voucher('RV-2', 2), self.income, self.alinma, '300')
        resp = self._ledger('1120001')
        self.assertEqual(resp.context['period_debit'], Decimal('1000'))
        self.assertEqual(resp.context['period_credit'], Decimal('300'))

    def test_empty_ledger_renders(self):
        resp = self._ledger('1120001')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['rows'], [])
        self.assertEqual(resp.context['closing'], Decimal('0'))

    def test_bad_date_is_ignored_not_a_500(self):
        self._pair(self._voucher('RV-1', 1), self.alinma, self.income, '1000')
        resp = self._ledger('1120001', **{'from': 'not-a-date'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['closing'], Decimal('1000'))

    def test_unknown_account_404s(self):
        self.assertEqual(self._ledger('9999999').status_code, 404)

    def test_requires_finance_or_super_admin(self):
        role, _ = Role.objects.get_or_create(name=Role.SALES_REP)
        rep = User.objects.create_user('rep2', password='x')
        rep.role = role
        rep.save()
        self.client.force_login(rep)
        self.assertEqual(self._ledger('1120001').status_code, 403)

    def test_natural_side_by_class(self):
        self.assertEqual(self.alinma.natural_side, 'debit')    # 1 assets
        self.assertEqual(self.income.natural_side, 'credit')   # 3 revenue
        self.assertEqual(
            Account(code='4100001', name='x').natural_side, 'debit')   # cost of sale
        self.assertEqual(
            Account(code='2120001', name='x').natural_side, 'credit')  # liability


class DocumentPostingTests(TestCase):
    """Posting a document produces the double-entry that puts it in the ledger."""

    def setUp(self):
        from accounting.models import AccountingSettings
        assets = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        self.ar = Account.objects.create(code='1140000', name='Accounts Receivable',
                                         internal_type='Receivable', parent=assets)
        self.vat_out = Account.objects.create(code='2180001', name='Output VAT',
                                              internal_type='Regular', parent=assets)
        self.vat_in = Account.objects.create(code='1300001', name='VAT Input',
                                             internal_type='Regular', parent=assets)
        self.ap = Account.objects.create(code='2120000', name='Accounts Payable',
                                         internal_type='Payable', parent=assets)
        self.income = Account.objects.create(code='3000101', name='Sales - Projects',
                                             internal_type='Regular', parent=assets)
        self.cogs = Account.objects.create(code='4100006', name='Local Procurement',
                                           internal_type='Regular', parent=assets)
        cfg = AccountingSettings.load()
        cfg.default_receivable_account = self.ar
        cfg.default_payable_account = self.ap
        cfg.output_tax_account = self.vat_out
        cfg.input_tax_account = self.vat_in
        cfg.save()
        self.customer = Partner.objects.create(name='Saudi Aramco', kind=Partner.KIND_CUSTOMER)
        self.vendor = Partner.objects.create(name='Al Ghad', kind=Partner.KIND_VENDOR)

    def _invoice(self, source='manual', subtotal='100000', tax='15000', total='115000'):
        doc = Document.objects.create(
            kind=Document.KIND_INVOICE, number='INV-001', partner=self.customer,
            date=datetime.date(2026, 8, 1), source=source,
            subtotal=Decimal(subtotal), tax_total=Decimal(tax), total=Decimal(total))
        DocumentLine.objects.create(document=doc, description='Phase 1',
                                    account=self.income, amount=Decimal(subtotal))
        return doc

    def _bill(self):
        doc = Document.objects.create(
            kind=Document.KIND_BILL, number='BILL-001', partner=self.vendor,
            date=datetime.date(2026, 8, 2), subtotal=Decimal('20000'),
            tax_total=Decimal('3000'), total=Decimal('23000'))
        DocumentLine.objects.create(document=doc, description='Cable',
                                    account=self.cogs, amount=Decimal('20000'))
        return doc

    def test_invoice_posts_debit_receivable_credit_revenue_and_tax(self):
        v = self._invoice().build_voucher()
        self.assertTrue(v.is_balanced)
        self.assertEqual(v.voucher_type, Voucher.TYPE_JV)
        by_account = {l.account.code: (l.debit, l.credit) for l in v.lines.all()}
        self.assertEqual(by_account['1140000'], (Decimal('115000'), Decimal('0')))
        self.assertEqual(by_account['3000101'], (Decimal('0'), Decimal('100000')))
        self.assertEqual(by_account['2180001'], (Decimal('0'), Decimal('15000')))

    def test_bill_posts_debit_expense_and_input_tax_credit_payable(self):
        v = self._bill().build_voucher()
        self.assertTrue(v.is_balanced)
        by_account = {l.account.code: (l.debit, l.credit) for l in v.lines.all()}
        self.assertEqual(by_account['4100006'], (Decimal('20000'), Decimal('0')))
        self.assertEqual(by_account['1300001'], (Decimal('3000'), Decimal('0')))
        self.assertEqual(by_account['2120000'], (Decimal('0'), Decimal('23000')))

    def test_posted_voucher_is_linked_and_document_opened(self):
        doc = self._invoice()
        v = doc.build_voucher()
        doc.refresh_from_db()
        self.assertEqual(doc.voucher_id, v.pk)
        self.assertEqual(doc.status, Document.STATUS_OPEN)
        self.assertEqual(v.status, Voucher.STATUS_POSTED)

    # ── the double-count guard ───────────────────────────────────────────

    def test_zoho_sourced_document_refuses_to_post(self):
        """Zoho already recorded the entry; its journal export supplies it."""
        doc = self._invoice(source='zoho')
        with self.assertRaises(ValidationError):
            doc.build_voucher()
        self.assertIsNone(doc.voucher_id)

    def test_cannot_post_the_same_document_twice(self):
        doc = self._invoice()
        doc.build_voucher()
        with self.assertRaises(ValidationError):
            doc.build_voucher()
        self.assertEqual(Voucher.objects.count(), 1)

    # ── refusing to guess ────────────────────────────────────────────────

    def test_refuses_when_lines_and_tax_do_not_reach_the_total(self):
        doc = self._invoice(subtotal='90000', tax='15000', total='115000')
        with self.assertRaises(ValidationError):
            doc.build_voucher()
        self.assertEqual(Voucher.objects.count(), 0)

    def test_refuses_without_a_control_account(self):
        from accounting.models import AccountingSettings
        cfg = AccountingSettings.load()
        cfg.default_receivable_account = None
        cfg.save()
        with self.assertRaises(ValidationError):
            self._invoice().build_voucher()

    def test_refuses_with_tax_but_no_tax_account(self):
        from accounting.models import AccountingSettings
        cfg = AccountingSettings.load()
        cfg.output_tax_account = None
        cfg.save()
        with self.assertRaises(ValidationError):
            self._invoice().build_voucher()

    def test_refuses_a_document_with_no_coded_lines(self):
        doc = Document.objects.create(
            kind=Document.KIND_INVOICE, number='INV-EMPTY', partner=self.customer,
            date=datetime.date(2026, 8, 1), total=Decimal('100'))
        with self.assertRaises(ValidationError):
            doc.build_voucher()

    def test_partner_account_wins_over_the_company_default(self):
        own = Account.objects.create(code='1140001', name='Saudi Aramco',
                                     internal_type='Receivable',
                                     parent=Account.objects.get(code='1000000'))
        self.customer.receivable_account = own
        self.customer.save()
        v = self._invoice().build_voucher()
        codes = {l.account.code for l in v.lines.all()}
        self.assertIn('1140001', codes)
        self.assertNotIn('1140000', codes)

    def test_zero_tax_document_posts_without_a_tax_line(self):
        doc = self._invoice(subtotal='100000', tax='0', total='100000')
        v = doc.build_voucher()
        self.assertTrue(v.is_balanced)
        self.assertEqual(v.lines.count(), 2)

    def test_posted_document_appears_in_the_account_ledger(self):
        """The whole point — an invoice must show against its revenue account."""
        self._invoice().build_voucher()
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        user = User.objects.create_user('sa_led', password='x')
        user.role = role
        user.save()
        self.client.force_login(user)
        resp = self.client.get(reverse('accounting:account_ledger', args=['3000101']))
        self.assertEqual(len(resp.context['rows']), 1)
        self.assertEqual(resp.context['closing'], Decimal('100000'))


class ZohoJournalImportTests(TestCase):
    """Importing Zoho's journal export as JV vouchers."""

    def setUp(self):
        self.dir = tempfile.mkdtemp()
        root = Account.objects.create(code='1000000', name='Assets', internal_type='View')
        self.bank = Account.objects.create(code='1120001', name='ALINMA Bank',
                                           internal_type='Liquidity', parent=root)
        self.cogs = Account.objects.create(code='4100006', name='Local Procurement',
                                           internal_type='Regular', parent=root)
        self.heading = root

    def _write_csv(self, rows, headers=None):
        import csv as _csv
        path = str(Path(self.dir) / 'journals.csv')
        headers = headers or ['Journal Date', 'Journal Number', 'Account Code',
                              'Account', 'Description', 'Contact Name', 'Debit', 'Credit']
        with open(path, 'w', newline='', encoding='utf-8') as f:
            w = _csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
        return path

    def _run(self, path, *args):
        out = StringIO()
        call_command('import_zoho_journals', path, *args, stdout=out)
        return out.getvalue()

    def _balanced_rows(self, number='JN-001', amount='5000'):
        return [
            ['2026-08-01', number, '4100006', 'Local Procurement', 'Cable', 'Al Ghad', amount, ''],
            ['2026-08-01', number, '1120001', 'ALINMA Bank', 'Cable', 'Al Ghad', '', amount],
        ]

    def test_imports_a_balanced_journal(self):
        self._run(self._write_csv(self._balanced_rows()))
        v = Voucher.objects.get(number='JN-001')
        self.assertEqual(v.voucher_type, Voucher.TYPE_JV)
        self.assertEqual(v.source, 'zoho')
        self.assertEqual(v.lines.count(), 2)
        self.assertTrue(v.is_balanced)

    def test_lands_as_draft_unless_post_is_given(self):
        self._run(self._write_csv(self._balanced_rows()))
        self.assertEqual(Voucher.objects.get(number='JN-001').status, Voucher.STATUS_DRAFT)
        self._run(self._write_csv(self._balanced_rows()), '--post')
        self.assertEqual(Voucher.objects.get(number='JN-001').status, Voucher.STATUS_POSTED)

    def test_reimport_updates_rather_than_duplicating(self):
        path = self._write_csv(self._balanced_rows())
        self._run(path)
        self._run(self._write_csv(self._balanced_rows(amount='7000')))
        self.assertEqual(Voucher.objects.filter(number='JN-001').count(), 1)
        v = Voucher.objects.get(number='JN-001')
        self.assertEqual(v.lines.count(), 2)          # not 4
        self.assertEqual(v.total_debit, Decimal('7000'))

    def test_unbalanced_journal_is_skipped_with_a_warning(self):
        rows = [
            ['2026-08-01', 'JN-BAD', '4100006', 'Local Procurement', 'x', '', '5000', ''],
            ['2026-08-01', 'JN-BAD', '1120001', 'ALINMA Bank', 'x', '', '', '4000'],
        ]
        out = self._run(self._write_csv(rows))
        self.assertIn('does not balance', out)
        self.assertFalse(Voucher.objects.filter(number='JN-BAD').exists())

    def test_unknown_account_line_is_reported_not_fatal(self):
        rows = self._balanced_rows() + [
            ['2026-08-02', 'JN-002', '9999999', 'Nope', 'x', '', '10', ''],
        ]
        out = self._run(self._write_csv(rows))
        self.assertIn('unknown account', out)
        self.assertTrue(Voucher.objects.filter(number='JN-001').exists())

    def test_heading_account_line_is_refused(self):
        rows = [
            ['2026-08-01', 'JN-H', '1000000', 'Assets', 'x', '', '10', ''],
            ['2026-08-01', 'JN-H', '1120001', 'ALINMA Bank', 'x', '', '', '10'],
        ]
        out = self._run(self._write_csv(rows))
        self.assertIn('is a heading', out)

    def test_partner_is_created_from_the_contact_column(self):
        self._run(self._write_csv(self._balanced_rows()))
        self.assertTrue(Partner.objects.filter(name='Al Ghad').exists())

    def test_headers_match_case_and_spacing_insensitively(self):
        path = self._write_csv(
            self._balanced_rows(),
            headers=['JOURNAL  DATE', 'journal_number', 'GL Code', 'Ledger Name',
                     'Notes', 'Customer Name', 'DEBIT AMOUNT', 'credit amount'])
        self._run(path)
        self.assertTrue(Voucher.objects.filter(number='JN-001').exists())

    def test_missing_required_column_lists_the_headers_it_saw(self):
        path = self._write_csv([], headers=['Something', 'Else'])
        with self.assertRaises(CommandError) as ctx:
            self._run(path)
        self.assertIn('Headers in the file', str(ctx.exception))

    def test_dry_run_writes_nothing(self):
        out = self._run(self._write_csv(self._balanced_rows()), '--dry-run')
        self.assertIn('Dry run', out)
        self.assertEqual(Voucher.objects.count(), 0)

    def test_missing_file_raises(self):
        with self.assertRaises(CommandError):
            self._run(str(Path(self.dir) / 'nope.csv'))

    def test_imported_journal_reaches_the_account_ledger(self):
        self._run(self._write_csv(self._balanced_rows()), '--post')
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        user = User.objects.create_user('sa_zj', password='x')
        user.role = role
        user.save()
        self.client.force_login(user)
        resp = self.client.get(reverse('accounting:account_ledger', args=['4100006']))
        self.assertEqual(resp.context['closing'], Decimal('5000'))


class PartnerTests(TestCase):

    def test_kind_flags(self):
        c = Partner.objects.create(name='C', kind=Partner.KIND_CUSTOMER)
        v = Partner.objects.create(name='V', kind=Partner.KIND_VENDOR)
        b = Partner.objects.create(name='B', kind=Partner.KIND_BOTH)
        self.assertEqual((c.is_customer, c.is_vendor), (True, False))
        self.assertEqual((v.is_customer, v.is_vendor), (False, True))
        self.assertEqual((b.is_customer, b.is_vendor), (True, True))
