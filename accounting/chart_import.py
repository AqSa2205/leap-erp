"""Reading the finance team's chart-of-accounts workbook.

Shared by the management command and the upload screen, so there is exactly one
definition of what the workbook means. A second parser would drift from this
one and the two would disagree about the chart everything else is coded
against.

Format is detected from the file's own magic bytes rather than its extension.
The revision that prompted this was named `.xls` and genuinely was one — an
OLE2 compound document, which openpyxl cannot open — but finance's tooling
relabels these freely in both directions, and an extension is a claim while
the first eight bytes are a fact.

Nothing here deletes an account. Codes are matched and updated; an account
absent from a new revision can be deactivated, which keeps historic postings
and their references intact. A chart of accounts is referenced by everything
downstream, so losing one silently would take its ledger history with it.
"""
import re

XLSX_MAGIC = b'PK\x03\x04'                       # zip container
XLS_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'  # OLE2 compound document

# Workbook layout (1-based columns), as sent by finance:
#   B Reporting code #   C Balance Sheet Description
#   D G/L code #         E G/L Description          F Internal Type
COL_REPORTING_CODE = 2
COL_REPORTING_NAME = 3
COL_GL_CODE = 4
COL_GL_NAME = 5
COL_INTERNAL_TYPE = 6
FIRST_DATA_ROW = 3          # rows 1-2 are the title and the header
LAST_COLUMN = 6


class ChartImportError(Exception):
    """Something about the workbook a person needs to fix."""


def _clean(value):
    """Trim a cell and collapse embedded newlines/runs of whitespace."""
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()


def sniff_format(head):
    """Identify a workbook from its leading bytes."""
    if head.startswith(XLSX_MAGIC):
        return 'xlsx'
    if head.startswith(XLS_MAGIC):
        return 'xls'
    return None


def read_grid(source, sheet_name=None):
    """Return (sheet_name, rows) with rows as lists of raw cell values.

    `source` is a path or a file-like object. Both readers are normalised to
    the same plain grid so nothing downstream needs to know which one ran.
    """
    if hasattr(source, 'read'):
        source.seek(0)
        payload = source.read()
    else:
        with open(source, 'rb') as handle:
            payload = handle.read()

    kind = sniff_format(payload[:8])
    if kind == 'xlsx':
        return _read_xlsx(payload, sheet_name)
    if kind == 'xls':
        return _read_xls(payload, sheet_name)
    raise ChartImportError(
        'That is not an Excel workbook. Expected a .xlsx or .xls file — if it '
        'opens in Excel, try File → Save As and choose Excel Workbook.')


def _read_xlsx(payload, sheet_name):
    import io
    try:
        import openpyxl
    except ImportError as exc:                                # pragma: no cover
        raise ChartImportError('openpyxl is required to read .xlsx files.') from exc

    workbook = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    if sheet_name and sheet_name not in workbook.sheetnames:
        raise ChartImportError(
            f'Sheet {sheet_name!r} is not in the workbook. Available: '
            f'{", ".join(workbook.sheetnames)}')
    sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = [[sheet.cell(r, c).value for c in range(1, LAST_COLUMN + 1)]
            for r in range(1, sheet.max_row + 1)]
    return sheet.title, rows


def _read_xls(payload, sheet_name):
    try:
        import xlrd
    except ImportError as exc:                                # pragma: no cover
        raise ChartImportError(
            'xlrd is required to read legacy .xls files. Re-save the workbook '
            'as .xlsx, or install xlrd.') from exc

    workbook = xlrd.open_workbook(file_contents=payload)
    names = workbook.sheet_names()
    if sheet_name and sheet_name not in names:
        raise ChartImportError(
            f'Sheet {sheet_name!r} is not in the workbook. Available: '
            f'{", ".join(names)}')
    sheet = workbook.sheet_by_name(sheet_name) if sheet_name else workbook.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        rows.append([sheet.cell_value(r, c) if c < sheet.ncols else None
                     for c in range(LAST_COLUMN)])
    return sheet.name, rows


def parse_rows(rows):
    """Turn a raw grid into account dicts.

    Returns (parsed, duplicates, bad_types). Heading rows carry the reporting
    code/name in columns B/C and repeat the same value in D/E; leaf rows fill
    only D/E. Either way the G/L pair wins.
    """
    from .models import Account
    valid_types = {t for t, _ in Account.INTERNAL_TYPE_CHOICES}

    parsed, seen, duplicates, bad_types = [], {}, [], []
    for index, raw in enumerate(rows[FIRST_DATA_ROW - 1:], start=FIRST_DATA_ROW):
        cells = list(raw) + [None] * (LAST_COLUMN - len(raw))
        reporting_code = _clean(cells[COL_REPORTING_CODE - 1])
        reporting_name = _clean(cells[COL_REPORTING_NAME - 1])
        gl_code = _clean(cells[COL_GL_CODE - 1])
        gl_name = _clean(cells[COL_GL_NAME - 1])
        internal_type = _clean(cells[COL_INTERNAL_TYPE - 1])

        code = gl_code or reporting_code
        name = gl_name or reporting_name
        if not code:
            continue
        # Excel hands numeric codes back as floats ("1110001.0").
        code = code[:-2] if code.endswith('.0') else code
        if not code.isdigit():
            continue

        if code in seen:
            duplicates.append((index, code, seen[code]))
            continue
        seen[code] = index

        if internal_type not in valid_types:
            if internal_type:
                bad_types.append((index, code, internal_type))
            internal_type = Account.TYPE_REGULAR

        parsed.append({
            'code': code,
            'name': name or code,
            'internal_type': internal_type,
            'reporting_code': reporting_code if reporting_code != code else '',
            'reporting_name': reporting_name if reporting_name != name else '',
            'source_row': index,
        })
    return parsed, duplicates, bad_types


def ancestor_codes(code):
    """Candidate parent codes, nearest first.

    Zeroes the trailing digits one at a time: 1200001 -> 1200000, then
    1200000 -> ... -> 1000000. Yielding nearest-first lets the caller pick the
    closest ancestor that actually exists, so a chart with a missing
    intermediate heading still links up instead of becoming a root.

    Candidates are only *shapes*, not guaranteed parents: zeroing one digit of
    4100025 gives 4100020, which in this chart is a real leaf account (Rental
    Equipment) rather than a heading. The caller must therefore accept only
    candidates that are headings — see resolve_parent.
    """
    for k in range(1, len(code)):
        candidate = code[:len(code) - k] + '0' * k
        if candidate != code:
            yield candidate


def resolve_parent(code, by_code):
    """Nearest ancestor of `code` that is a heading, or None for a root.

    Only `View` accounts may parent others — that is exactly what the type
    means in this chart. Without the check, an account like 4100025 would be
    filed under its sibling 4100020 simply because zeroing a digit matched a
    code that happens to exist.
    """
    from .models import Account

    for candidate in ancestor_codes(code):
        account = by_code.get(candidate)
        if account is not None and account.internal_type == Account.TYPE_VIEW:
            return account
    return None


def plan(parsed):
    """What applying this revision would do, without doing it.

    The preview is the whole safety story for the upload screen: a chart of
    accounts is the structure every ledger entry is coded against, and the
    difference between 'four new accounts' and 'two hundred renamed' is the
    difference between a routine revision and a mistake caught in time.
    """
    from .models import Account

    existing = {a.code: a for a in Account.objects.all()}
    incoming = {item['code']: item for item in parsed}

    created, updated, unchanged = [], [], []
    for code, item in incoming.items():
        current = existing.get(code)
        if current is None:
            created.append(item)
        elif (current.name != item['name']
                or current.internal_type != item['internal_type']
                or not current.is_active):
            updated.append({
                'code': code,
                'from_name': current.name, 'to_name': item['name'],
                'from_type': current.internal_type, 'to_type': item['internal_type'],
                'reactivated': not current.is_active,
            })
        else:
            unchanged.append(item)

    missing = [a for code, a in existing.items()
               if code not in incoming and a.is_active]

    return {
        'total': len(parsed),
        'created': created,
        'updated': updated,
        'unchanged': unchanged,
        'missing': missing,
    }


def apply(parsed, deactivate_missing=False):
    """Upsert the parsed rows, relink parents, optionally retire the absentees."""
    from .models import Account

    created = updated = 0
    for item in parsed:
        _, was_created = Account.objects.update_or_create(
            code=item['code'],
            defaults={
                'name': item['name'],
                'internal_type': item['internal_type'],
                'reporting_code': item['reporting_code'],
                'reporting_name': item['reporting_name'],
                'source_row': item['source_row'],
                'is_active': True,
            })
        created += was_created
        updated += not was_created

    # Second pass so ordering in the sheet never matters: by now every account
    # in the file exists, whether created just now or on an earlier run.
    by_code = {a.code: a for a in Account.objects.all()}
    linked, orphans, to_update = 0, [], []
    for item in parsed:
        account = by_code[item['code']]
        parent = resolve_parent(item['code'], by_code)
        if parent is not None:
            linked += 1
        else:
            orphans.append(item['code'])
        if account.parent_id != getattr(parent, 'pk', None):
            account.parent = parent
            to_update.append(account)
    if to_update:
        Account.objects.bulk_update(to_update, ['parent'])

    deactivated = 0
    if deactivate_missing:
        codes = {item['code'] for item in parsed}
        deactivated = Account.objects.exclude(code__in=codes).filter(
            is_active=True).update(is_active=False)

    return {
        'created': created, 'updated': updated,
        'linked': linked, 'orphans': orphans, 'deactivated': deactivated,
    }
