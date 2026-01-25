from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q, Count
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import ContactDatabase
from .forms import ContactDatabaseForm, ContactDatabaseFilterForm, ContactImportForm


class ContactListView(LoginRequiredMixin, ListView):
    """List all contacts with filtering by category."""
    model = ContactDatabase
    template_name = 'contacts/contact_list.html'
    context_object_name = 'contacts'
    paginate_by = 25

    def get_queryset(self):
        queryset = ContactDatabase.objects.all()

        # Get filter parameters
        search = self.request.GET.get('search', '')
        category = self.request.GET.get('category', '')
        status = self.request.GET.get('status', '')
        region = self.request.GET.get('region', '')

        if search:
            queryset = queryset.filter(
                Q(organisation_name__icontains=search) |
                Q(contact_name__icontains=search) |
                Q(contact_email__icontains=search) |
                Q(title__icontains=search)
            )
        if category:
            queryset = queryset.filter(category=category)
        if status:
            queryset = queryset.filter(status=status)
        if region:
            queryset = queryset.filter(region__icontains=region)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ContactDatabaseFilterForm(self.request.GET)
        context['total_count'] = ContactDatabase.objects.count()

        # Get counts by category
        category_counts = ContactDatabase.objects.values('category').annotate(
            count=Count('id')
        ).order_by('category')
        context['category_counts'] = {
            item['category']: item['count'] for item in category_counts
        }

        # Get current category for highlighting
        context['current_category'] = self.request.GET.get('category', '')

        return context


class ContactByCategoryView(LoginRequiredMixin, ListView):
    """List contacts filtered by a specific category."""
    model = ContactDatabase
    template_name = 'contacts/contact_by_category.html'
    context_object_name = 'contacts'
    paginate_by = 25

    def get_queryset(self):
        category = self.kwargs.get('category')
        queryset = ContactDatabase.objects.filter(category=category)

        # Apply additional filters
        search = self.request.GET.get('search', '')
        status = self.request.GET.get('status', '')
        region = self.request.GET.get('region', '')

        if search:
            queryset = queryset.filter(
                Q(organisation_name__icontains=search) |
                Q(contact_name__icontains=search) |
                Q(contact_email__icontains=search) |
                Q(title__icontains=search)
            )
        if status:
            queryset = queryset.filter(status=status)
        if region:
            queryset = queryset.filter(region__icontains=region)

        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        category = self.kwargs.get('category')
        context['category'] = category
        context['category_display'] = dict(ContactDatabase.CATEGORY_CHOICES).get(category, category)
        context['filter_form'] = ContactDatabaseFilterForm(self.request.GET)
        context['total_count'] = self.get_queryset().count()
        return context


class ContactDetailView(LoginRequiredMixin, DetailView):
    """View contact details."""
    model = ContactDatabase
    template_name = 'contacts/contact_detail.html'
    context_object_name = 'contact'


class ContactCreateView(LoginRequiredMixin, CreateView):
    """Create a new contact."""
    model = ContactDatabase
    form_class = ContactDatabaseForm
    template_name = 'contacts/contact_form.html'
    success_url = reverse_lazy('contacts:contact_list')

    def get_initial(self):
        initial = super().get_initial()
        # Pre-fill category if provided in URL
        category = self.request.GET.get('category')
        if category:
            initial['category'] = category
        return initial

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Contact created successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Contact'
        context['button_text'] = 'Save Contact'
        return context


class ContactUpdateView(LoginRequiredMixin, UpdateView):
    """Update an existing contact."""
    model = ContactDatabase
    form_class = ContactDatabaseForm
    template_name = 'contacts/contact_form.html'
    success_url = reverse_lazy('contacts:contact_list')

    def form_valid(self, form):
        messages.success(self.request, 'Contact updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Contact'
        context['button_text'] = 'Update Contact'
        return context


class ContactDeleteView(LoginRequiredMixin, DeleteView):
    """Delete a contact."""
    model = ContactDatabase
    template_name = 'contacts/contact_confirm_delete.html'
    success_url = reverse_lazy('contacts:contact_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Contact deleted successfully.')
        return super().delete(request, *args, **kwargs)


@login_required
def contact_import(request):
    """Import contacts from Excel file."""
    if request.method == 'POST':
        form = ContactImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            category = form.cleaned_data['category']

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active

                # Get headers from first row
                headers = [cell.value for cell in ws[1] if cell.value]
                header_map = {h.lower().strip(): idx for idx, h in enumerate(headers)}

                imported_count = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not any(row):
                        continue

                    # Map Excel columns to model fields
                    def get_value(possible_headers):
                        for h in possible_headers:
                            if h in header_map:
                                idx = header_map[h]
                                if idx < len(row):
                                    return row[idx]
                        return ''

                    organisation = get_value(['organisation name', 'organization name', 'company', 'organisation'])
                    if not organisation:
                        continue

                    contact = ContactDatabase(
                        category=category,
                        notice_identifier=str(get_value(['notice identifier', 'notice id']) or ''),
                        notice_type='contract',
                        organisation_name=str(organisation)[:255],
                        title=str(get_value(['title', 'project title']) or '')[:500],
                        description=str(get_value(['description']) or ''),
                        postcode=str(get_value(['postcode', 'post code']) or '')[:20],
                        region=str(get_value(['region']) or '')[:255],
                        contact_name=str(get_value(['contact name', 'name']) or '')[:255],
                        contact_email=str(get_value(['contact email', 'email']) or ''),
                        contact_address=str(get_value(['contact address 1', 'contact address', 'address']) or ''),
                        contact_telephone=str(get_value(['contact telephone', 'telephone', 'phone']) or '')[:100],
                        contact_website=str(get_value(['contact website', 'website']) or '')[:500],
                        cpv_codes=str(get_value(['cpv codes', 'cpv']) or '')[:255],
                        comments=str(get_value(['comments', 'notes']) or ''),
                        created_by=request.user
                    )

                    # Handle status
                    status_value = str(get_value(['status']) or '').lower()
                    if 'award' in status_value:
                        contact.status = 'awarded'
                    elif 'open' in status_value:
                        contact.status = 'open'
                    elif 'close' in status_value:
                        contact.status = 'closed'
                    elif 'pend' in status_value:
                        contact.status = 'pending'
                    else:
                        contact.status = 'unknown'

                    # Handle serial number
                    sr = get_value(['sr', 'sr#', 'serial'])
                    if sr:
                        try:
                            contact.serial_number = int(sr)
                        except (ValueError, TypeError):
                            pass

                    # Handle published date
                    pub_date = get_value(['published date', 'date'])
                    if pub_date:
                        if isinstance(pub_date, datetime):
                            contact.published_date = pub_date.date()

                    contact.save()
                    imported_count += 1

                messages.success(request, f'Successfully imported {imported_count} contacts.')
                return redirect('contacts:contact_list')

            except Exception as e:
                messages.error(request, f'Error importing file: {str(e)}')
    else:
        form = ContactImportForm()

    return render(request, 'contacts/contact_import.html', {'form': form})


@login_required
def contact_export(request):
    """Export contacts to Excel."""
    category = request.GET.get('category', '')
    search = request.GET.get('search', '')
    status = request.GET.get('status', '')

    queryset = ContactDatabase.objects.all()

    if category:
        queryset = queryset.filter(category=category)
    if search:
        queryset = queryset.filter(
            Q(organisation_name__icontains=search) |
            Q(contact_name__icontains=search) |
            Q(contact_email__icontains=search)
        )
    if status:
        queryset = queryset.filter(status=status)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    # Styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        'Category', 'Notice ID', 'Notice Type', 'SR#', 'Status',
        'Organisation Name', 'Title', 'Contact Name', 'Contact Email',
        'Contact Telephone', 'Region', 'Postcode', 'Published Date', 'Comments'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for row_num, contact in enumerate(queryset, 2):
        data = [
            contact.get_category_display(),
            contact.notice_identifier,
            contact.get_notice_type_display(),
            contact.serial_number,
            contact.get_status_display(),
            contact.organisation_name,
            contact.title,
            contact.contact_name,
            contact.contact_email,
            contact.contact_telephone,
            contact.region,
            contact.postcode,
            contact.published_date.strftime('%Y-%m-%d') if contact.published_date else '',
            contact.comments
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

    # Adjust column widths
    column_widths = [15, 20, 12, 8, 12, 30, 40, 20, 30, 15, 20, 12, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'contacts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
