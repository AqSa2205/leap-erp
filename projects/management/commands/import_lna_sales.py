"""
Management command to import LNA Sales Report Excel data into the database.

Usage:
    python manage.py import_lna_sales "path/to/LNA SALES REPORT TILL SEPT 2025.xlsx"
    python manage.py import_lna_sales "path/to/file.xlsx" --dry-run
    python manage.py import_lna_sales "path/to/file.xlsx" --clear-existing
"""

import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from projects.models import Project, Region, ProjectStatus


# ── Status mapping helpers ───────────────────────────────────────────────

STATUS_MAP = {
    'submitted': 'Submitted',
    'under process': 'IP',
    'in-progress': 'IP',
    'in progress': 'IP',
    'not submitted': 'Closed',
    'close': 'Closed',
    'closed': 'Closed',
    'awarded': 'Awarded',
    'ongoing': 'Ongoing',
    'on going': 'Ongoing',
    'on goin': 'Ongoing',
    'open': 'Open',
    'hold': 'Hold',
    'lost': 'Lost',
    'hot lead': 'Hot Lead',
    'hot': 'Hot Lead',
}

# Section header overrides the status column
SECTION_STATUS_OVERRIDE = {
    'non-active bids': 'Hold',
    'non active bids': 'Hold',
    'lost': 'Lost',
    'hot lead': 'Hot Lead',
    'hot leads': 'Hot Lead',
    'budgetrory': 'Open',
    'budgetary': 'Open',
    'working leads': 'Open',
    'ongoing project': 'Ongoing',
    'ongoing projects': 'Ongoing',
}


def parse_decimal(val, default=Decimal('0')):
    """Safely parse a value to Decimal."""
    if val is None:
        return default
    try:
        cleaned = str(val).replace(',', '').replace('£', '').replace('$', '').strip()
        if not cleaned or cleaned == '-':
            return default
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return default


def parse_date(val):
    """Safely parse a value to a date."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    try:
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y', '%d/%m/%Y'):
            try:
                return datetime.datetime.strptime(str(val).strip(), fmt).date()
            except ValueError:
                continue
    except Exception:
        pass
    return None


def clean_str(val, max_len=None):
    """Safely convert to stripped string, optionally truncated."""
    if val is None:
        return ''
    s = str(val).strip()
    if max_len:
        s = s[:max_len]
    return s


def parse_quarter(val):
    """Parse PO Award Quarter to Q1-Q4 format."""
    if not val:
        return ''
    s = str(val).strip().upper()
    for q in ('Q1', 'Q2', 'Q3', 'Q4'):
        if q in s:
            return q
    return ''


class Command(BaseCommand):
    help = 'Import LNA sales report from Excel into the projects database'

    def add_arguments(self, parser):
        parser.add_argument('filepath', type=str, help='Path to the Excel file')
        parser.add_argument(
            '--dry-run', action='store_true',
            help='Preview what would be imported without saving',
        )
        parser.add_argument(
            '--clear-existing', action='store_true',
            help='Delete all existing LNA projects before importing',
        )

    def handle(self, *args, **options):
        filepath = options['filepath']
        dry_run = options['dry_run']
        clear_existing = options['clear_existing']

        self.stdout.write(f'Loading workbook: {filepath}')
        wb = load_workbook(filepath, data_only=True)
        self.stdout.write(f'Sheets found: {wb.sheetnames}')

        # Get LNA region
        try:
            self.region = Region.objects.get(code='LNA')
        except Region.DoesNotExist:
            self.stderr.write(self.style.ERROR('LNA region not found. Run load_initial_data first.'))
            return

        # Cache statuses
        self.status_cache = {}
        for s in ProjectStatus.objects.all():
            self.status_cache[s.name] = s

        if clear_existing and not dry_run:
            count = Project.objects.filter(region=self.region).count()
            Project.objects.filter(region=self.region).delete()
            self.stdout.write(self.style.WARNING(f'Deleted {count} existing LNA projects'))

        self.created = 0
        self.updated = 0
        self.skipped = 0
        self.errors = []
        self.dry_run = dry_run

        # ── 1. Pipeline sheets (richest data) ──
        pipeline_sheets = [
            'Aramco-Sept-25-Smart Safety',
            'Aramco-Sept-25',
            "Ma'aden Sept-25",
            'EPC-May-25',
            'Special Project Unit',
            'Central Region Lead-EPC',
        ]
        for sheet_name in pipeline_sheets:
            if sheet_name in wb.sheetnames:
                self.import_pipeline_sheet(wb[sheet_name], sheet_name)

        # ── 2. Budgetary sheet ──
        if 'BDGTYR-2025' in wb.sheetnames:
            self.import_pipeline_sheet(wb['BDGTYR-2025'], 'BDGTYR-2025')

        # ── 3. Proposal Summary & Smart Safety Proposal ──
        for sheet_name in ['Proposal Summary', 'Smart Safety Proposal']:
            if sheet_name in wb.sheetnames:
                self.import_proposal_summary(wb[sheet_name], sheet_name)

        # ── 4. Ongoing projects ──
        if 'Ongoing projects' in wb.sheetnames:
            self.import_ongoing_awarded(wb['Ongoing projects'], 'Ongoing projects')

        # ── 5. Awarded Projects ──
        if 'Awarded Projects' in wb.sheetnames:
            self.import_awarded(wb['Awarded Projects'], 'Awarded Projects')

        # ── 6. Actual Sales ──
        for sheet_name in ['Actual Sales 2025', 'Actual Sales 2024']:
            if sheet_name in wb.sheetnames:
                year = sheet_name.split()[-1]
                self.import_actual_sales(wb[sheet_name], sheet_name, year)

        # ── Summary ──
        prefix = '[DRY RUN] ' if dry_run else ''
        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'{prefix}Import complete: {self.created} created, '
            f'{self.updated} updated, {self.skipped} skipped'
        ))
        if self.errors:
            self.stdout.write(self.style.WARNING(f'{len(self.errors)} errors:'))
            for err in self.errors[:20]:
                self.stdout.write(f'  - {err}')
            if len(self.errors) > 20:
                self.stdout.write(f'  ... and {len(self.errors) - 20} more')

    def get_status(self, status_text, section_name=''):
        """Resolve a status name to a ProjectStatus object."""
        # Section override takes priority
        section_lower = section_name.lower().strip()
        if section_lower in SECTION_STATUS_OVERRIDE:
            target = SECTION_STATUS_OVERRIDE[section_lower]
            if target in self.status_cache:
                return self.status_cache[target]

        # Map from status text
        if status_text:
            status_lower = status_text.lower().strip()
            target = STATUS_MAP.get(status_lower)
            if target and target in self.status_cache:
                return self.status_cache[target]

        # Default to Submitted
        return self.status_cache.get('Submitted', list(self.status_cache.values())[0])

    def save_project(self, proposal_ref, defaults, sheet_name, row_num):
        """Create or update a project, handling errors."""
        if not proposal_ref or proposal_ref in ('-', 'None', ''):
            self.skipped += 1
            return

        proposal_ref = clean_str(proposal_ref, 50)

        # Skip rows that are clearly not project data
        if not defaults.get('project_name'):
            self.skipped += 1
            return

        try:
            if self.dry_run:
                exists = Project.objects.filter(proposal_reference=proposal_ref).exists()
                action = 'UPDATE' if exists else 'CREATE'
                self.stdout.write(
                    f'  [{action}] {proposal_ref}: {defaults.get("project_name", "")[:60]}'
                )
                if exists:
                    self.updated += 1
                else:
                    self.created += 1
                return

            obj, created = Project.objects.update_or_create(
                proposal_reference=proposal_ref,
                defaults=defaults,
            )
            if created:
                self.created += 1
            else:
                self.updated += 1
        except Exception as e:
            self.errors.append(f'{sheet_name} row {row_num}: {proposal_ref} - {e}')

    # ──────────────────────────────────────────────────────────────────────
    # Pipeline sheets (Aramco, Ma'aden, EPC, Special Projects, Central Region)
    # Header row 4, data rows after section headers
    # Columns: A=S/N, B=Project Name, C=Proposal Ref, D=Client RFQ,
    #          E=Submission Date, F=Owner(client), G=EPC, H=Bid Status,
    #          I=Est Value SAR, J=Est Value USD, K=Per Annum, L=GP,
    #          M=PO Award Q, N=Success Quotient, O=Min Achievement,
    #          P=Est PO date (some sheets), Q/P=Contact With, R/Q=Remarks
    # ──────────────────────────────────────────────────────────────────────

    def import_pipeline_sheet(self, ws, sheet_name):
        self.stdout.write(f'\nProcessing: {sheet_name}')

        # Find header row dynamically
        header_row = None
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
            row_vals = [c.value for c in row if c.value]
            row_text = ' '.join(str(v) for v in row_vals)
            if 'Project Name' in row_text and 'S/N' in row_text:
                header_row = row[0].row
                for cell in row:
                    if cell.value:
                        headers[str(cell.value).strip()] = cell.column - 1  # 0-indexed
                break

        if not header_row:
            self.stdout.write(self.style.WARNING(f'  No header row found, skipping'))
            return

        # Build column index map
        def col(name, *alternates):
            idx = headers.get(name)
            if idx is not None:
                return idx
            for alt in alternates:
                idx = headers.get(alt)
                if idx is not None:
                    return idx
            return None

        ci = {
            'sn': col('S/N'),
            'name': col('Project Name'),
            'ref': col('Leap Proposal Reference', 'Leap Proposal Reference '),
            'rfq': col('Client RFQ Ref Number'),
            'date': col('Submission Date', 'Submission Date/Deadline'),
            'owner_client': col('Owner'),
            'epc': col('EPC'),
            'status': col('Bid Status'),
            'est_sar': col('Est. Value (SAR)'),
            'po_q': col(' PO Award - Q', 'PO Award - Q'),
            'sq': col('Success Quotient'),
            'min_ach': col('Minimum Achievement'),
            'est_po_date': col('Est PO date'),
            'contact': col('Contact With'),
            'remarks': col('Remarks'),
        }

        current_section = 'Active Bids'
        count = 0

        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=False):
            vals = [c.value for c in row]

            # Detect section headers (col B has text, col A is empty)
            a_val = vals[0] if len(vals) > 0 else None
            b_val = vals[1] if len(vals) > 1 else None

            if b_val and not a_val:
                section_text = str(b_val).strip()
                if section_text.lower() not in ('total value', 'procurement'):
                    current_section = section_text
                continue

            # Skip non-data rows
            if not isinstance(a_val, (int, float)):
                continue

            def g(key):
                idx = ci.get(key)
                if idx is not None and idx < len(vals):
                    return vals[idx]
                return None

            proposal_ref = g('ref')
            project_name = g('name')
            if not proposal_ref or not project_name:
                continue

            status = self.get_status(clean_str(g('status')), current_section)
            est_value = parse_decimal(g('est_sar'))
            sq = parse_decimal(g('sq'))
            min_ach = parse_decimal(g('min_ach'))

            # Build notes with client/owner info
            owner_client = clean_str(g('owner_client'))
            notes_parts = []
            if owner_client:
                notes_parts.append(f'Client: {owner_client}')
            notes_parts.append(f'Source: {sheet_name} / {current_section}')

            defaults = {
                'project_name': clean_str(project_name, 500),
                'client_rfq_reference': clean_str(g('rfq'), 255),
                'region': self.region,
                'status': status,
                'epc': clean_str(g('epc'), 200),
                'estimated_value': est_value,
                'submission_deadline': parse_date(g('date')),
                'estimated_po_date': parse_date(g('est_po_date')),
                'po_award_quarter': parse_quarter(g('po_q')),
                'success_quotient': sq,
                'minimum_achievement': min_ach if min_ach else None,
                'contact_with': clean_str(g('contact'), 255),
                'remarks': clean_str(g('remarks')),
                'notes': ' | '.join(notes_parts),
            }

            self.save_project(proposal_ref, defaults, sheet_name, row[0].row)
            count += 1

        self.stdout.write(f'  Processed {count} rows')

    # ──────────────────────────────────────────────────────────────────────
    # Proposal Summary / Smart Safety Proposal
    # Header row 2, different columns
    # A=S/N, B=Project Name, C=Proposal Ref, D=Client RFQ,
    # E=Submission Deadline, F=Owner, G=EPC,
    # H=Submission Date to LNA, I=Submission Date to Client,
    # J=Status, K=Remarks, L=Action By
    # ──────────────────────────────────────────────────────────────────────

    def import_proposal_summary(self, ws, sheet_name):
        self.stdout.write(f'\nProcessing: {sheet_name}')

        # Find header row
        header_row = None
        headers = {}
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
            row_text = ' '.join(str(c.value) for c in row if c.value)
            if 'Project Name' in row_text:
                header_row = row[0].row
                for cell in row:
                    if cell.value:
                        headers[str(cell.value).strip()] = cell.column - 1
                break

        if not header_row:
            self.stdout.write(self.style.WARNING(f'  No header row found, skipping'))
            return

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if not row or not isinstance(row[0], (int, float)):
                continue

            # Column positions: A=0, B=1, C=2, D=3, E=4, F=5, G=6, H=7, I=8, J=9, K=10, L=11
            sn = row[0]
            project_name = row[1] if len(row) > 1 else None
            proposal_ref = row[2] if len(row) > 2 else None
            client_rfq = row[3] if len(row) > 3 else None
            sub_deadline = row[4] if len(row) > 4 else None
            owner_client = row[5] if len(row) > 5 else None
            epc = row[6] if len(row) > 6 else None
            status_text = row[9] if len(row) > 9 else None
            remarks_text = row[10] if len(row) > 10 else None

            if not proposal_ref or not project_name:
                continue

            # Only import if not already in DB (pipeline sheets have richer data)
            if Project.objects.filter(proposal_reference=clean_str(proposal_ref, 50)).exists():
                self.skipped += 1
                continue

            status = self.get_status(clean_str(status_text))

            notes_parts = []
            if owner_client:
                notes_parts.append(f'Client: {clean_str(owner_client)}')
            notes_parts.append(f'Source: {sheet_name}')

            defaults = {
                'project_name': clean_str(project_name, 500),
                'client_rfq_reference': clean_str(client_rfq, 255),
                'region': self.region,
                'status': status,
                'epc': clean_str(epc, 200),
                'estimated_value': Decimal('0'),
                'submission_deadline': parse_date(sub_deadline),
                'remarks': clean_str(remarks_text),
                'notes': ' | '.join(notes_parts),
            }

            self.save_project(proposal_ref, defaults, sheet_name, sn)
            count += 1

        self.stdout.write(f'  Processed {count} rows')

    # ──────────────────────────────────────────────────────────────────────
    # Ongoing projects
    # A=S/N, B=Project Name, C=Proposal Ref, D=Owner, E=EPC,
    # F=Status, G=Est Value SAR, H=Est Value USD
    # ──────────────────────────────────────────────────────────────────────

    def import_ongoing_awarded(self, ws, sheet_name):
        self.stdout.write(f'\nProcessing: {sheet_name}')

        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=False):
            row_text = ' '.join(str(c.value) for c in row if c.value)
            if 'Project Name' in row_text and 'S/N' in row_text:
                header_row = row[0].row
                break

        if not header_row:
            self.stdout.write(self.style.WARNING(f'  No header row found, skipping'))
            return

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if not row or not isinstance(row[0], (int, float)):
                continue

            proposal_ref = row[2] if len(row) > 2 else None
            project_name = row[1] if len(row) > 1 else None
            owner_client = row[3] if len(row) > 3 else None
            epc = row[4] if len(row) > 4 else None
            status_text = row[5] if len(row) > 5 else None
            est_sar = row[6] if len(row) > 6 else None

            if not proposal_ref or not project_name:
                continue

            status = self.get_status(clean_str(status_text), 'Ongoing Project')

            notes_parts = []
            if owner_client:
                notes_parts.append(f'Client: {clean_str(owner_client)}')
            notes_parts.append(f'Source: {sheet_name}')

            defaults = {
                'project_name': clean_str(project_name, 500),
                'region': self.region,
                'status': status,
                'epc': clean_str(epc, 200),
                'estimated_value': parse_decimal(est_sar),
                'notes': ' | '.join(notes_parts),
            }

            self.save_project(proposal_ref, defaults, sheet_name, row[0])
            count += 1

        self.stdout.write(f'  Processed {count} rows')

    # ──────────────────────────────────────────────────────────────────────
    # Awarded Projects
    # A=S/N, B=Project Name, C=Proposal Ref, D=Owner, E=EPC,
    # F=Status, G=Est Value SAR, H=Est Value USD, ..., O=PO Number
    # ──────────────────────────────────────────────────────────────────────

    def import_awarded(self, ws, sheet_name):
        self.stdout.write(f'\nProcessing: {sheet_name}')

        count = 0
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
            if not row or not isinstance(row[0], (int, float)):
                continue

            proposal_ref = row[2] if len(row) > 2 else None
            project_name = row[1] if len(row) > 1 else None
            owner_client = row[3] if len(row) > 3 else None
            epc = row[4] if len(row) > 4 else None
            est_sar = row[6] if len(row) > 6 else None
            po_number = row[14] if len(row) > 14 else None

            if not proposal_ref or not project_name:
                continue

            status = self.get_status('Awarded')

            notes_parts = []
            if owner_client:
                notes_parts.append(f'Client: {clean_str(owner_client)}')
            notes_parts.append(f'Source: {sheet_name}')

            defaults = {
                'project_name': clean_str(project_name, 500),
                'region': self.region,
                'status': status,
                'epc': clean_str(epc, 200),
                'estimated_value': parse_decimal(est_sar),
                'po_number': clean_str(po_number, 100),
                'notes': ' | '.join(notes_parts),
            }

            self.save_project(proposal_ref, defaults, sheet_name, row[0])
            count += 1

        self.stdout.write(f'  Processed {count} rows')

    # ──────────────────────────────────────────────────────────────────────
    # Actual Sales 2024/2025
    # A=S/N, B=Project Name, C=Project Code, D=PO Number,
    # E=Customer, F=Client, G=PO Value, H=Status, I=Completion %
    # ──────────────────────────────────────────────────────────────────────

    def import_actual_sales(self, ws, sheet_name, year):
        self.stdout.write(f'\nProcessing: {sheet_name}')

        # Find header row
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=False):
            row_text = ' '.join(str(c.value) for c in row if c.value)
            if 'Project Name' in row_text and 'PO' in row_text:
                header_row = row[0].row
                break

        if not header_row:
            self.stdout.write(self.style.WARNING(f'  No header row found, skipping'))
            return

        count = 0
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, values_only=True):
            if not row or not isinstance(row[0], (int, float)):
                continue

            project_name = row[1] if len(row) > 1 else None
            project_code = row[2] if len(row) > 2 else None
            po_number = row[3] if len(row) > 3 else None
            customer = row[4] if len(row) > 4 else None
            client = row[5] if len(row) > 5 else None
            po_value = row[6] if len(row) > 6 else None
            status_text = row[7] if len(row) > 7 else None

            if not project_name:
                continue

            # Use project_code as proposal_reference if available,
            # otherwise generate one
            if project_code and clean_str(project_code):
                proposal_ref = clean_str(project_code, 50)
            else:
                # Generate a reference from PO number or sequential
                po_clean = clean_str(po_number, 30)
                if po_clean:
                    proposal_ref = f'LNA-SALES-{year}-{po_clean}'[:50]
                else:
                    proposal_ref = f'LNA-SALES-{year}-{int(row[0])}'

            status = self.get_status(clean_str(status_text))
            # Actual sales projects are typically ongoing or closed
            if status.name == 'Submitted':
                status = self.status_cache.get('Ongoing', status)

            notes_parts = []
            if customer:
                notes_parts.append(f'Customer: {clean_str(customer)}')
            if client:
                notes_parts.append(f'Client: {clean_str(client)}')
            notes_parts.append(f'Source: {sheet_name}')

            defaults = {
                'project_name': clean_str(project_name, 500),
                'region': self.region,
                'status': status,
                'epc': clean_str(customer, 200),
                'estimated_value': parse_decimal(po_value),
                'actual_sales': parse_decimal(po_value),
                'po_number': clean_str(po_number, 100),
                'year': year,
                'notes': ' | '.join(notes_parts),
            }

            self.save_project(proposal_ref, defaults, sheet_name, row[0])
            count += 1

        self.stdout.write(f'  Processed {count} rows')
