import io
import json
import os
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404, HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.urls import reverse, reverse_lazy
from django.utils.http import urlencode
from django.db import transaction
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from django.core.files.base import ContentFile
from django.utils import timezone
import openpyxl

from .models import Project, Region, ProjectStatus, ProjectHistory, Document, ProjectRevision, PipelineEmail
from .forms import ProjectForm, ProjectFilterForm, DocumentForm, DocumentFilterForm
from .email_parsing import parse_eml_file, EmailParseError
from . import graph_mail
from notifications.services import notify_users
from accounts.permissions import CapabilityRequiredMixin


class ProjectPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Base mixin for project permissions"""

    def get_queryset(self):
        """Filter queryset based on user role"""
        queryset = Project.objects.select_related('status', 'region', 'owner').all()
        user = self.request.user

        if user.is_super_admin_user:
            return queryset
        # Procurement only kicks in after a project is Won, so they only
        # ever need to see Won-status projects in their region.
        if getattr(user, 'is_procurement_user', False):
            return queryset.filter(
                region=user.region,
                status__category='won',
            )
        # Admin/manager, finance, and the proposal team get a region-scoped
        # *view* of the pipeline. Sales REPS are deliberately excluded — they
        # see only the projects they own (fall through to owner=user below).
        region_scoped = (
            user.is_admin_user
            or user.is_manager_user
            or getattr(user, 'is_finance_team_user', False)
            or getattr(user, 'is_proposal_team_user', False)
        )

        if region_scoped:
            # A region-scoped user with no region assigned would otherwise fall
            # through to seeing everything, so return nothing and tell them why
            # (once per request, not once per queryset evaluation).
            if user.region_id is None:
                if not getattr(self.request, '_no_region_warning_shown', False):
                    messages.warning(
                        self.request,
                        'Your account has no region assigned. You can only view projects you own.'
                    )
                    self.request._no_region_warning_shown = True
                return queryset.none()
            return queryset.filter(region=user.region)
        return queryset.filter(owner=user)


class ProjectListView(CapabilityRequiredMixin, ProjectPermissionMixin, ListView):
    """List all projects with filtering"""
    capability = 'pipeline.access'
    model = Project
    template_name = 'projects/project_list.html'
    context_object_name = 'projects'
    paginate_by = 25

    def test_func(self):
        return True  # All authenticated users can view

    # Mapping of consolidated regions to database region codes
    REGION_CODE_MAP = {
        'LNUK': ['UK', 'GLB', 'LNUK'],  # UK and Global together
        'LNA': ['LNA'],
        'PA': ['PA'],
        'NEO-Dubai': ['NEO-Dubai'],
        'NEO-KSA': ['NEO-KSA'],
    }

    def get_queryset(self):
        queryset = super().get_queryset()

        # Apply filters
        search = self.request.GET.get('search')
        region = self.request.GET.get('region')
        year = self.request.GET.get('year')
        status = self.request.GET.get('status')
        workflow_stage = self.request.GET.get('workflow_stage')
        quarter = self.request.GET.get('quarter')
        owner = self.request.GET.get('owner')

        if search:
            queryset = queryset.filter(
                Q(project_name__icontains=search) |
                Q(proposal_reference__icontains=search) |
                Q(client_rfq_reference__icontains=search)
            )
        if region:
            # Use consolidated region mapping
            region_codes = self.REGION_CODE_MAP.get(region, [])
            if region_codes:
                queryset = queryset.filter(region__code__in=region_codes)
        if year:
            queryset = queryset.filter(year=year)
        if status:
            queryset = queryset.filter(status_id=status)
        if workflow_stage == 'none':
            # Projects that have no costing sheet yet (workflow not started).
            queryset = queryset.filter(costing_sheets__isnull=True)
        elif workflow_stage:
            # Any costing sheet at this stage. distinct() guards against a
            # project with multiple sheets appearing more than once.
            queryset = queryset.filter(
                costing_sheets__workflow_stage=workflow_stage
            ).distinct()
        if quarter:
            queryset = queryset.filter(po_award_quarter=quarter)
        if owner:
            queryset = queryset.filter(owner_id=owner)

        # Prefetch costing sheets + their SOW items so each project's row
        # can resolve its derived contract-total without an N+1 storm.
        # ``sales_resolved`` is attached per project in get_context_data().
        return queryset.prefetch_related(
            'costing_sheets',
            'costing_sheets__scope_of_work_items',
        ).order_by('-updated_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = ProjectFilterForm(self.request.GET, user=self.request.user)

        # Add summary stats
        queryset = self.get_queryset()
        context['total_count'] = queryset.count()

        # Exchange rates are loaded once per request and injected onto every
        # sheet so contract_total avoids a per-row ExchangeRate query.
        from costing.models import ExchangeRate, pipeline_stage_badge, WORKFLOW_STAGE_SEQUENCE
        rates = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}

        # Headline value split: Won vs Hot Leads, in SAR (converted so
        # cross-region projects sum correctly). Computed over the FULL
        # filtered queryset — not just the current page — so it reflects
        # every matching project. Prefetch sections__line_items here because
        # the SAR value comes from each sheet's contract_total (grand_total).
        summary_qs = queryset.filter(
            status__category__in=['won', 'hot_lead']
        ).prefetch_related('costing_sheets__sections__line_items')
        context['value_summary'] = _pipeline_value_summary(summary_qs, rates)

        # Decorate each visible row with its resolved sales-value dict so
        # the template can render the same "Actual / Costing / Estimate /
        # Not started" logic the detail page uses. The prefetch on
        # get_queryset() means this is O(rows) without extra queries.

        def _stage_rank(s):
            try:
                return WORKFLOW_STAGE_SEQUENCE.index(s.workflow_stage)
            except ValueError:
                return -1

        for project in context.get('projects', []):
            sheets = list(project.costing_sheets.all())
            for s in sheets:
                s.set_rates_cache(rates)
            project.sales_resolved = _resolve_project_sales_value(project, sheets)
            # Commercial-pipeline workflow badge (furthest BOM→Sales→Finance
            # stage across this project's costing sheets; None = not started).
            project.pipeline_stage = pipeline_stage_badge(sheets)
            # The furthest-along sheet drives the cycle-time columns (management
            # view). Tie-break on the most recently updated.
            project.cycle_sheet = (
                max(sheets, key=lambda s: (_stage_rank(s), s.updated_at))
                if sheets else None)

        # Regions for import modal
        context['regions'] = Region.objects.filter(is_active=True)

        return context


class ProjectDetailView(ProjectPermissionMixin, DetailView):
    """View project details"""
    model = Project
    template_name = 'projects/project_detail.html'
    context_object_name = 'project'

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        project = self.object
        # Edit/delete permission for this specific project (the model helpers
        # need the project instance, which templates can't pass to a method).
        context['can_edit'] = self.request.user.can_edit_project(project)
        context['can_delete'] = self.request.user.can_delete_project()
        context['history'] = project.history.select_related(
            'old_status', 'new_status', 'changed_by'
        ).all()[:10]
        context['revisions'] = project.revisions.select_related('created_by').all()

        # ── Workflow data: everything linked to this proposal_reference ──
        # RFQ documents
        rfq_docs = Document.objects.filter(project=project, document_type='rfq')
        # Costing sheets (BOM and Costing are the same record)
        costing_sheets = list(project.costing_sheets.select_related('created_by').all())
        # Technical proposals (write-ups)
        from proposals.models import TechnicalProposal
        tech_proposals = list(
            TechnicalProposal.objects.filter(project=project).order_by('-updated_at')
        )
        # Vendor quotes + Commercial Proposal PDFs reached via any linked costing sheet
        from costing.models import VendorQuote, CostingSheetRevision
        vendor_quotes = VendorQuote.objects.filter(sheet__in=costing_sheets).count()
        commercial_pdfs = CostingSheetRevision.objects.filter(sheet__in=costing_sheets).count()

        context['workflow'] = {
            'rfq_count':            rfq_docs.count(),
            'rfq_first_url':        rfq_docs.first().file.url if rfq_docs.exists() else None,
            'costing_sheets':       costing_sheets,
            'tech_proposals':       tech_proposals,
            'vendor_quote_count':   vendor_quotes,
            'commercial_pdf_count': commercial_pdfs,
        }

        # ── Derived "Actual Sales / Costing" panel value ─────────────────
        # Same resolution rule used by the Commercial Pipeline list — see
        # _resolve_project_sales_value. Display-only, never writes back to
        # the actual_sales column. Pre-load exchange rates and inject onto
        # every sheet so contract_total doesn't hit the DB once per sheet.
        from costing.models import ExchangeRate
        rates = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}
        for s in costing_sheets:
            s.set_rates_cache(rates)
        resolved = _resolve_project_sales_value(project, costing_sheets)
        context['sales_value_source']   = resolved['source']
        context['sales_value_amount']   = resolved['amount']
        context['sales_value_note']     = resolved['note']
        context['sales_value_sheet']    = resolved['sheet']
        context['sales_value_currency'] = resolved['currency']

        # ── Timeline: flat, chronologically sorted event stream ─────────
        # Each event: {when, icon, color, title, subtitle, actor, url}
        events = []

        def _actor_name(u):
            return (u.get_full_name() or u.username) if u else ''

        # Project created
        if project.created_at:
            events.append({
                'when':     project.created_at,
                'icon':     'bi-folder-plus',
                'color':    '#1a1a1a',
                'title':    'Pipeline entry created',
                'subtitle': project.project_name or project.proposal_reference,
                'actor':    _actor_name(project.owner),
                'url':      '',
            })

        # RFQ documents uploaded
        for doc in rfq_docs.select_related('uploaded_by'):
            events.append({
                'when':     doc.uploaded_at,
                'icon':     'bi-file-earmark-arrow-up',
                'color':    '#0dcaf0',
                'title':    'Client RFQ uploaded',
                'subtitle': doc.name,
                'actor':    _actor_name(doc.uploaded_by),
                'url':      reverse_lazy('projects:document_detail', kwargs={'pk': doc.pk}),
            })

        # Costing sheet lifecycle
        for sheet in costing_sheets:
            if sheet.created_at:
                events.append({
                    'when':     sheet.created_at,
                    'icon':     'bi-list-check',
                    'color':    '#ffc107',
                    'title':    'BOM started',
                    'subtitle': sheet.title,
                    'actor':    _actor_name(sheet.created_by),
                    'url':      reverse_lazy('costing:detail', kwargs={'pk': sheet.pk}) + '?view=bom',
                })
            if getattr(sheet, 'handed_over_at', None):
                events.append({
                    'when':     sheet.handed_over_at,
                    'icon':     'bi-arrow-right-square',
                    'color':    '#0d6efd',
                    'title':    'BOM handed over to sales',
                    'subtitle': sheet.title,
                    'actor':    _actor_name(getattr(sheet, 'handed_over_by', None)),
                    'url':      reverse_lazy('costing:detail', kwargs={'pk': sheet.pk}),
                })
            if getattr(sheet, 'costing_started_at', None):
                events.append({
                    'when':     sheet.costing_started_at,
                    'icon':     'bi-cash-coin',
                    'color':    '#198754',
                    'title':    'Costing in progress',
                    'subtitle': sheet.title,
                    'actor':    _actor_name(getattr(sheet, 'costing_started_by', None)),
                    'url':      reverse_lazy('costing:detail', kwargs={'pk': sheet.pk}),
                })
            if getattr(sheet, 'finalized_at', None):
                events.append({
                    'when':     sheet.finalized_at,
                    'icon':     'bi-check2-circle',
                    'color':    '#212529',
                    'title':    'Costing finalized',
                    # The proposal team must not see costing figures — drop the grand total.
                    'subtitle': (sheet.title if self.request.user.is_proposal_team_user
                                 else f'{sheet.title} · Grand total {sheet.grand_total:.2f} {sheet.output_currency}'),
                    'actor':    _actor_name(getattr(sheet, 'finalized_by', None)),
                    'url':      reverse_lazy('costing:detail', kwargs={'pk': sheet.pk}),
                })

        # Commercial Proposal PDF exports + vendor quotes link straight to the
        # priced PDF / supplier cost documents, so they must never appear in the
        # proposal team's activity feed (they work from the BOM only).
        if not self.request.user.is_proposal_team_user:
            from costing.models import CostingSheetRevision as _Rev
            for rev in _Rev.objects.filter(sheet__in=costing_sheets, export_format='pdf').select_related('created_by', 'sheet'):
                events.append({
                    'when':     rev.created_at,
                    'icon':     'bi-file-earmark-pdf',
                    'color':    '#C41E3A',
                    'title':    f'Commercial Proposal {rev.revision_label} exported',
                    'subtitle': rev.change_summary or rev.sheet.title,
                    'actor':    _actor_name(rev.created_by),
                    'url':      rev.file.url if rev.file else reverse_lazy('costing:detail', kwargs={'pk': rev.sheet_id}),
                })

            # Vendor quotes uploaded
            from costing.models import VendorQuote as _VQ
            for vq in _VQ.objects.filter(sheet__in=costing_sheets).select_related('uploaded_by', 'sheet'):
                events.append({
                    'when':     vq.uploaded_at,
                    'icon':     'bi-receipt',
                    'color':    '#6f42c1',
                    'title':    f'Vendor quote from {vq.vendor_name}',
                    'subtitle': f'{vq.quote_reference or "no ref"} · {vq.sheet.title}',
                    'actor':    _actor_name(vq.uploaded_by),
                    'url':      reverse_lazy('costing:detail', kwargs={'pk': vq.sheet_id}),
                })

        # Technical proposals
        for tp in tech_proposals:
            if tp.created_at:
                events.append({
                    'when':     tp.created_at,
                    'icon':     'bi-file-earmark-richtext',
                    'color':    '#0dcaf0',
                    'title':    'Technical Proposal created',
                    'subtitle': tp.title,
                    'actor':    _actor_name(tp.created_by),
                    'url':      reverse_lazy('proposals:detail', kwargs={'pk': tp.pk}),
                })

        # Pipeline-entry revisions
        for rev in project.revisions.select_related('created_by'):
            events.append({
                'when':     rev.created_at,
                'icon':     'bi-bookmark-star',
                'color':    '#fd7e14',
                'title':    f'Revision {rev.revision_label} saved',
                'subtitle': rev.notes or 'Snapshot of pipeline state',
                'actor':    _actor_name(rev.created_by),
                'url':      reverse_lazy('projects:revision_detail', kwargs={'pk': project.pk, 'revision_pk': rev.pk}),
            })

        # Sort newest first
        events.sort(key=lambda e: e['when'] or project.created_at, reverse=True)
        context['timeline'] = events

        # ── Pipeline email linking ("Add Emails" / "Delink") ──────────────
        # Purely additive: doesn't touch any of the context keys above.
        context['active_linked_email'] = project.linked_emails.filter(is_active=True).first()
        context['email_documents'] = project.documents.filter(source_pipeline_email__isnull=False)
        context['other_documents'] = project.documents.filter(source_pipeline_email__isnull=True)

        return context


def _exchange_rates_json():
    """Return a JSON-safe dict of {currency_code: rate_to_usd} for use by the
    project form's USD → local-currency auto-conversion. Stays in sync with
    whatever the costing module's ExchangeRate rows hold."""
    import json
    from costing.models import ExchangeRate
    rates = {
        r.currency_code: float(r.rate_to_usd)
        for r in ExchangeRate.objects.all()
    }
    return json.dumps(rates)


def _lna_form_context(obj):
    """Context for the project form's LNA auto-reference behaviour: the LNA
    region id (so JS knows when to lock the field) and the number to preview."""
    from .models import next_lna_reference_number, parse_lna_reference
    lna = Region.objects.filter(code='LNA').first()
    number = next_lna_reference_number()
    if obj and getattr(obj, 'pk', None) and obj.proposal_reference:
        parsed = parse_lna_reference(obj.proposal_reference)
        if parsed:
            number = parsed[0]
    return {
        'lna_region_id': lna.id if lna else '',
        'lna_preview_number': number,
    }

@login_required
def next_lna_reference_preview(request):
    """Tiny JSON endpoint: returns the current next-free LNA number,
    calculated fresh at the moment it's called. Used by the draft-resume
    button so it never shows a stale number, without needing a page reload."""
    from django.http import JsonResponse
    from .models import next_lna_reference_number
    return JsonResponse({'number': next_lna_reference_number()})


PRICED_WORKFLOW_STAGES = {
    'costing_in_progress', 'finalized', 'finance_review', 'finance_approved',
}
PENDING_WORKFLOW_STAGES = {'bom_in_progress', 'ready_for_costing'}


def _pipeline_value_summary(projects, rates):
    """Split the pipeline's contract value into Won vs Hot-Lead buckets.

    All amounts are converted to SAR so projects from different regions
    (SAR, GBP, …) can be summed into one figure. Each project's value is
    resolved with the SAME rule the rows use (_resolve_project_sales_value):
    the live costing-sheet contract total when a priced sheet exists, else
    actual sales (won) / estimate as a fallback.

    ``rates`` maps currency_code -> rate_to_usd (units of that currency per
    1 USD, e.g. SAR≈3.75, GBP≈0.79). To convert amount X→SAR we go via USD:
    ``X / rate_to_usd[cur] * rate_to_usd['SAR']``.

    Returns a dict of Decimals + counts, including ``hot_no_costing_count``:
    hot-lead projects that don't yet have a priced costing sheet.
    """
    sar_rate = rates.get('SAR') or Decimal('1')

    def _to_sar(amount, currency):
        if not amount:
            return Decimal('0')
        cur = (currency or 'SAR').upper()
        if cur == 'SAR':
            return Decimal(amount)
        cur_rate = rates.get(cur)
        if not cur_rate or not sar_rate:
            return Decimal(amount)          # missing rate → leave as-is
        return (Decimal(amount) / cur_rate) * sar_rate

    summary = {
        'won_value_sar':        Decimal('0'),
        'won_count':            0,
        'hot_value_sar':        Decimal('0'),
        'hot_count':            0,
        'hot_no_costing_count': 0,
    }
    for project in projects:
        category = project.status_category
        if category not in ('won', 'hot_lead'):
            continue
        sheets = list(project.costing_sheets.all())
        for s in sheets:
            s.set_rates_cache(rates)
        resolved = _resolve_project_sales_value(project, sheets)
        value_sar = _to_sar(resolved['amount'], resolved['currency'])
        if category == 'won':
            summary['won_value_sar'] += value_sar
            summary['won_count'] += 1
        else:  # hot_lead
            summary['hot_value_sar'] += value_sar
            summary['hot_count'] += 1
            if resolved['source'] != 'costing':
                summary['hot_no_costing_count'] += 1
    return summary


def _resolve_project_sales_value(project, costing_sheets=None):
    """Derive the "Actual Sales / Costing" display for a single project.

    Returns a dict shaped:
        {
            'source':   'actual' | 'costing' | 'estimate' | 'none',
            'amount':   Decimal | None,
            'note':     str,           # subtitle under the big number
            'sheet':    CostingSheet | None,
            'currency': 'GBP' | 'SAR', # which currency to format `amount` in
        }

    Resolution order (costing wins over actual_sales — the costing sheet
    is the live source of truth; actual_sales can go stale):
      1. Any linked costing sheet has a grand_total > 0 → "Costing Total
         (live)" with the highest grand_total. Stage doesn't matter; we
         look at the number itself.
      2. Actual sales recorded → "Actual Sales".
      3. Otherwise → "Estimated Price" using the project's estimated_value,
         with the subtitle "Costing not started". Keeps the column useful
         for early-stage pipeline entries without adding visual noise.
      4. No estimate either → bare "Costing not started" message.

    Display-only — never writes back to actual_sales.
    """
    region_code = project.region.code if project.region_id else ''
    local_ccy = 'GBP' if region_code in ('UK', 'GLB') else 'SAR'

    if costing_sheets is None:
        costing_sheets = list(project.costing_sheets.all())
    else:
        costing_sheets = list(costing_sheets)

    # Pick the sheet with the largest non-zero contract_total — that
    # matches the "MAIN — TOTAL CONTRACT PRICE" line on the costing PDF
    # (A.1 + A.3* + A.2 Services), in the sheet's own output currency.
    priced_sheets = [s for s in costing_sheets if s.contract_total and s.contract_total > 0]
    if priced_sheets:
        latest_priced = max(priced_sheets, key=lambda s: s.contract_total)
        return {
            'source':   'costing',
            'amount':   latest_priced.contract_total,
            'note':     f'Live contract total from costing sheet "{latest_priced.title}" · {latest_priced.get_workflow_stage_display()}',
            'sheet':    latest_priced,
            'currency': latest_priced.contract_total_currency,
        }

    if project.actual_sales and project.actual_sales > 0:
        return {
            'source':   'actual',
            'amount':   project.actual_sales,
            'note':     None,
            'sheet':    None,
            'currency': local_ccy,
        }

    # No priced costing yet — show the estimated value as a placeholder.
    if project.estimated_value and project.estimated_value > 0:
        return {
            'source':   'estimate',
            'amount':   project.estimated_value,
            'note':     'Costing not started',
            'sheet':    None,
            'currency': local_ccy,
        }

    return {
        'source':   'none',
        'amount':   None,
        'note':     'Costing not started',
        'sheet':    None,
        'currency': local_ccy,
    }


class ProjectCreateView(LoginRequiredMixin, CreateView):
    """Create a new project"""
    model = Project
    form_class = ProjectForm
    template_name = 'projects/project_form.html'
    success_url = reverse_lazy('projects:list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_initial(self):
        # Coming back from the live-inbox "Add Emails" flow (see
        # link_pipeline_email_new) with a freshly picked email — re-fetch
        # its summary so the hidden picked_email_json field (and therefore
        # the "Email to be attached" section) is populated on this fresh
        # GET. Only on GET: this form has no <form action>, so a real
        # Create POST still carries the same ?picked_email=... — without
        # this guard we'd hit Graph pointlessly (and unsafely) on every
        # real submission too.
        initial = super().get_initial()
        if self.request.method == 'GET':
            message_id = self.request.GET.get('picked_email')
            if message_id and _can_use_pipeline_email_feature(self.request.user):
                # Always this user's own mailbox — never taken from the
                # request. See _user_mailbox()'s docstring.
                mailbox = _user_mailbox(self.request.user)
                try:
                    if not mailbox:
                        raise graph_mail.GraphMailError(
                            'No mailbox is linked to your account.')
                    summary = graph_mail.get_message_summary(mailbox, message_id)
                    # Default type, same as the existing-project attach flow —
                    # editable in the "Email to be attached" section, and
                    # whatever's chosen there rides along in this same field.
                    for attachment in summary['attachments']:
                        attachment['document_type'] = 'vendor_quotation'
                    # Without this, each attachment's .url is never set here
                    # (only the existing-project attach flow set it), so the
                    # "Email to be attached" panel renders a dead href="" —
                    # clicking a document just reloads the create page.
                    _add_attachment_urls([summary], None)
                    initial['picked_email_json'] = json.dumps(summary)
                except graph_mail.GraphMailError as exc:
                    messages.warning(self.request, f'Could not load that email: {exc}')
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['exchange_rates_json'] = _exchange_rates_json()
        context.update(_lna_form_context(getattr(self, 'object', None)))
        context['document_type_choices'] = Document.DOCUMENT_TYPE_CHOICES
        from drafts.models import FormDraft
        context['draft'] = FormDraft.objects.filter(
            user=self.request.user, form_key='project_create', object_id=None
        ).first()
        return context

    def form_valid(self, form):
        from drafts.models import FormDraft
        FormDraft.objects.filter(
            user=self.request.user, form_key='project_create', object_id=None
        ).delete()

        form.instance.created_by = self.request.user
        if not form.instance.owner:
            form.instance.owner = self.request.user
        messages.success(self.request, 'Project created successfully.')
        response = super().form_valid(form)

        # Notify the region's sales + proposal team (and managers) so they can
        # see and start working on the new project. Region-scoped visibility is
        # granted by ProjectPermissionMixin.
        from accounts.models import User, Role
        project = self.object

        # A new commercial-pipeline project starts its costing workflow at
        # "BOM not started" — auto-create the sheet so it shows on the Costing
        # list and the proposal team can pick it up ("Start BOM").
        from costing.models import CostingSheet
        CostingSheet.objects.create(
            project=project,
            title=project.project_name or project.proposal_reference or f'Project {project.pk}',
            customer_reference=project.proposal_reference or '',
            customer_name=project.customer or '',
            end_user=project.end_user or '',
            contact_person=project.contact_with or '',
            created_by=self.request.user,
            workflow_stage='bom_not_started',
        )

        # Sales REPS are not notified region-wide: under owner-only scoping
        # projects:detail 404s for a rep who doesn't own the project, so a
        # region-wide notice would send most of them to a dead link. The owning
        # rep is added back explicitly below — they are the one rep who can
        # open it.
        team_roles = [
            Role.MANAGER, Role.PROPOSAL_HEAD, Role.PROPOSAL_REP,
        ]
        recipients = User.objects.filter(
            Q(role__name__in=team_roles, region=project.region)
            # Only the OWNING rep, and only when they are a rep — every other
            # role's reach is untouched by owner-scoping, so nothing is added
            # back for them. notify_users() already drops the actor.
            | Q(pk=project.owner_id, role__name=Role.SALES_REP)
        ).filter(is_active=True).distinct()
        notify_users(
            recipients=recipients,
            verb='created a new pipeline project',
            actor=self.request.user,
            target=project,
            target_url=reverse_lazy('projects:detail', kwargs={'pk': project.pk}),
            description=(
                f'New project "{project.project_name}" in '
                f'{project.region or "—"} is ready for the team.'
            ),
            level='info',
            send_email=True,
        )

        # An email picked from the live inbox while filling out this form
        # (see link_pipeline_email_new / ProjectForm.picked_email_json) is
        # only materialized into a real PipelineEmail + Documents now that
        # the project actually exists. The project/costing sheet/notification
        # above are already committed by this point, so a failure here must
        # never surface as an error on top of an otherwise-successful create.
        picked_email_json = form.cleaned_data.get('picked_email_json')
        if picked_email_json and _can_use_pipeline_email_feature(self.request.user):
            try:
                picked = json.loads(picked_email_json)
                # Keyed by position, not a['name'] — see
                # _attach_email_to_project()'s docstring for why a filename
                # picked here (from Graph's attachment listing) can't be
                # trusted to match the filename parse_eml_file() re-derives.
                attachment_types = {
                    str(index): a.get('document_type')
                    for index, a in enumerate(picked.get('attachments', []))
                }
                # Always this user's own mailbox — never taken from
                # picked_email_json (which rode back from the browser in a
                # hidden field, so it can't be trusted for this) or any
                # other request value. See _user_mailbox()'s docstring.
                # Checked explicitly here (not left to fetch_raw_message_
                # bytes() raising below and the broad except catching it) —
                # every other _user_mailbox() call site in this file guards
                # the empty case up front, and this one should too rather
                # than relying on a downstream function happening to.
                mailbox = _user_mailbox(self.request.user)
                if not mailbox:
                    raise graph_mail.GraphMailError('No mailbox is linked to your account.')
                _attach_email_to_project(
                    project, self.request.user, mailbox, picked['id'],
                    attachment_types=attachment_types)
            except Exception:
                messages.warning(
                    self.request,
                    'The project was created, but the picked email could not be attached '
                    'automatically — you can attach it again from the project page.'
                )

        return response


class ProjectUpdateView(ProjectPermissionMixin, UpdateView):
    """Update a project"""
    model = Project
    form_class = ProjectForm
    template_name = 'projects/project_form.html'

    def test_func(self):
        # Finance gets a region-scoped *view* of the pipeline (via the mixin
        # queryset), but not edit rights — exclude them. Everyone else unchanged.
        return not getattr(self.request.user, 'is_finance_team_user', False)

    def get_queryset(self):
        # Editing is narrower than viewing: super admins edit anything,
        # admins/managers edit their region, and everyone else (incl. sales &
        # proposal, who now *view* the whole region) edits only projects they
        # own. This stops the widened view-queryset from granting region-wide
        # project editing.
        qs = Project.objects.all()
        user = self.request.user
        if user.is_super_admin_user:
            return qs
        if user.is_admin_user or user.is_manager_user:
            return qs.filter(region=user.region)
        return qs.filter(owner=user)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['exchange_rates_json'] = _exchange_rates_json()
        context.update(_lna_form_context(getattr(self, 'object', None)))
        return context

    def form_valid(self, form):
        old_owner = Project.objects.get(pk=self.object.pk).owner

        # Track status change
        if 'status' in form.changed_data:
            old_status = Project.objects.get(pk=self.object.pk).status
            ProjectHistory.objects.create(
                project=self.object,
                old_status=old_status,
                new_status=form.cleaned_data['status'],
                changed_by=self.request.user
            )

        messages.success(self.request, 'Project updated successfully.')
        response = super().form_valid(form)

        # Notify on owner change
        new_owner = self.object.owner
        if 'owner' in form.changed_data and old_owner != new_owner:
            target_url = str(reverse_lazy('projects:detail', kwargs={'pk': self.object.pk}))
            recipients = set()
            if old_owner:
                recipients.add(old_owner)
            if new_owner:
                recipients.add(new_owner)
            notify_users(
                recipients=recipients,
                verb=f'changed project owner for "{self.object.project_name}"',
                actor=self.request.user,
                target=self.object,
                target_url=target_url,
                description=f'Owner changed from {old_owner or "None"} to {new_owner or "None"}',
                level='warning',
                send_email=True,
            )

        return response

    def get_success_url(self):
        return reverse_lazy('projects:detail', kwargs={'pk': self.object.pk})


class ProjectDeleteView(ProjectPermissionMixin, DeleteView):
    """Soft-delete a project (admin only).

    The row is retained and simply hidden, so nothing cascades into finance
    rows or history, and costing sheets / POs keep their project link. Deleted
    projects can be restored from the recycle bin.
    """
    model = Project
    template_name = 'projects/project_confirm_delete.html'
    success_url = reverse_lazy('projects:list')

    def test_func(self):
        return self.request.user.can_delete_project()

    def form_valid(self, form):
        project = self.get_object()
        project.soft_delete(user=self.request.user)
        messages.success(
            self.request,
            f'Project "{project.project_name}" moved to the recycle bin. '
            f'You can restore it from Projects → Recycle Bin.')
        return redirect(self.success_url)


class ProjectRecycleBinView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Soft-deleted projects, most recently deleted first."""
    model = Project
    template_name = 'projects/project_recycle_bin.html'
    context_object_name = 'projects'
    paginate_by = 50

    def test_func(self):
        return self.request.user.can_delete_project()

    def get_queryset(self):
        # all_objects to see past the default manager's is_deleted filter.
        qs = (Project.all_objects
              .filter(is_deleted=True)
              .select_related('status', 'region', 'owner', 'deleted_by')
              .order_by('-deleted_at'))
        user = self.request.user
        if not user.is_super_admin_user:
            qs = qs.filter(region=user.region)
        return qs


@login_required
@require_POST
def project_restore(request, pk):
    """Restore a soft-deleted project from the recycle bin."""
    if not request.user.can_delete_project():
        messages.error(request, 'You do not have permission to restore projects.')
        return redirect('projects:recycle_bin')

    project = get_object_or_404(Project.all_objects, pk=pk, is_deleted=True)
    if (not request.user.is_super_admin_user
            and project.region_id != request.user.region_id):
        raise Http404('Project not found.')

    project.restore()
    messages.success(request, f'Project "{project.project_name}" restored.')
    return redirect('projects:detail', pk=project.pk)


class ProjectImportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Import projects from an Excel file"""

    def test_func(self):
        return self.request.user.is_super_admin_user or self.request.user.is_admin_user or self.request.user.is_manager_user

    def post(self, request):
        excel_file = request.FILES.get('excel_file')
        region_id = request.POST.get('region')

        if not excel_file:
            messages.error(request, 'Please select an Excel file to import.')
            return redirect('projects:list')

        if not region_id:
            messages.error(request, 'Please select a region.')
            return redirect('projects:list')

        try:
            region = Region.objects.get(pk=region_id)
        except Region.DoesNotExist:
            messages.error(request, 'Invalid region selected.')
            return redirect('projects:list')

        try:
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            ws = wb.active

            # Auto-detect header row by scanning for known column names
            header_row = None
            header_map = {}
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
                for cell in row:
                    val = str(cell.value).strip() if cell.value else ''
                    if val.lower() in ('leap proposal reference', 'project name'):
                        header_row = row_idx
                        break
                if header_row:
                    break

            if not header_row:
                messages.error(request, 'Could not find header row. Ensure the Excel file has a "Leap Proposal Reference" or "Project Name" column.')
                return redirect('projects:list')

            # Build column index mapping from header row
            COLUMN_MAP = {
                'project name': 'project_name',
                'leap proposal reference': 'proposal_reference',
                'client rfq ref number': 'client_rfq_reference',
                'submission date': 'submission_deadline',
                'owner': 'owner',
                'epc': 'customer',
                'customer': 'customer',
                'end user': 'end_user',
                'bid status': 'status',
                'est. value (sar)': 'estimated_value',
                'est. value ($usd)': 'estimated_value_usd',
                'est. value (sar) - per annum': 'estimated_value_per_annum',
                'est. gp': 'estimated_gp',
                'po award - q': 'po_award_quarter',
                'success quotient': 'success_quotient',
                'minimum achievement': 'minimum_achievement',
                'contact with': 'contact_with',
                'remarks': 'remarks',
            }

            for cell in list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False))[0]:
                col_name = str(cell.value).strip().lower() if cell.value else ''
                if col_name in COLUMN_MAP:
                    header_map[COLUMN_MAP[col_name]] = cell.column - 1  # 0-based index

            if 'proposal_reference' not in header_map:
                messages.error(request, 'Excel file must contain a "Leap Proposal Reference" column.')
                return redirect('projects:list')

            # Cache lookups
            from accounts.models import User
            users = {u.get_full_name().lower(): u for u in User.objects.filter(is_active=True) if u.get_full_name()}
            users.update({u.username.lower(): u for u in User.objects.filter(is_active=True)})
            statuses = {s.name.lower(): s for s in ProjectStatus.objects.filter(is_active=True)}

            # Get a default status for new projects
            default_status = ProjectStatus.objects.filter(is_active=True).first()

            imported_count = 0
            errors = []
            row_num = header_row  # for error reporting

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                row_num += 1

                # Get values by column index
                def get_val(field_name):
                    idx = header_map.get(field_name)
                    if idx is not None and idx < len(row):
                        return row[idx]
                    return None

                proposal_ref = get_val('proposal_reference')
                if not proposal_ref:
                    continue  # Skip rows without proposal reference
                proposal_ref = str(proposal_ref).strip()
                if not proposal_ref:
                    continue

                try:
                    defaults = {'region': region}

                    # Project name
                    project_name = get_val('project_name')
                    if project_name:
                        defaults['project_name'] = str(project_name).strip()
                    else:
                        defaults['project_name'] = proposal_ref

                    # Client RFQ Reference
                    client_rfq = get_val('client_rfq_reference')
                    if client_rfq:
                        defaults['client_rfq_reference'] = str(client_rfq).strip()

                    # Submission Date
                    sub_date = get_val('submission_deadline')
                    if sub_date:
                        import datetime
                        if isinstance(sub_date, datetime.datetime):
                            defaults['submission_deadline'] = sub_date.date()
                        elif isinstance(sub_date, datetime.date):
                            defaults['submission_deadline'] = sub_date
                        else:
                            try:
                                from django.utils.dateparse import parse_date
                                parsed = parse_date(str(sub_date).strip())
                                if parsed:
                                    defaults['submission_deadline'] = parsed
                            except (ValueError, TypeError):
                                pass

                    # Owner
                    owner_val = get_val('owner')
                    if owner_val:
                        owner_key = str(owner_val).strip().lower()
                        matched_user = users.get(owner_key)
                        if matched_user:
                            defaults['owner'] = matched_user

                    # Customer (was EPC)
                    customer_val = get_val('customer')
                    if customer_val:
                        defaults['customer'] = str(customer_val).strip()

                    # End User
                    end_user_val = get_val('end_user')
                    if end_user_val:
                        defaults['end_user'] = str(end_user_val).strip()

                    # Bid Status
                    status_val = get_val('status')
                    if status_val:
                        status_key = str(status_val).strip().lower()
                        matched_status = statuses.get(status_key)
                        if matched_status:
                            defaults['status'] = matched_status
                        elif default_status:
                            defaults['status'] = default_status
                    elif default_status:
                        defaults['status'] = default_status

                    # Decimal fields
                    decimal_fields = [
                        'estimated_value', 'estimated_value_usd',
                        'estimated_value_per_annum', 'estimated_gp',
                        'success_quotient', 'minimum_achievement',
                    ]
                    for field_name in decimal_fields:
                        val = get_val(field_name)
                        if val is not None:
                            try:
                                defaults[field_name] = Decimal(str(val).strip().replace(',', ''))
                            except (InvalidOperation, ValueError):
                                pass

                    # PO Award Quarter
                    po_q = get_val('po_award_quarter')
                    if po_q:
                        po_q_str = str(po_q).strip().upper()
                        if po_q_str in ('Q1', 'Q2', 'Q3', 'Q4'):
                            defaults['po_award_quarter'] = po_q_str

                    # Contact With
                    contact = get_val('contact_with')
                    if contact:
                        defaults['contact_with'] = str(contact).strip()

                    # Remarks
                    remarks = get_val('remarks')
                    if remarks:
                        defaults['remarks'] = str(remarks).strip()

                    Project.objects.update_or_create(
                        proposal_reference=proposal_ref,
                        defaults=defaults,
                    )
                    imported_count += 1

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

            wb.close()

            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} project(s).')
                # Notify importer on completion
                from notifications.services import create_notification
                create_notification(
                    recipient=request.user,
                    verb=f'Excel import completed: {imported_count} project(s)',
                    target_url=str(reverse_lazy('projects:list')),
                    description=f'Successfully imported {imported_count} project(s) into region {region.name}.',
                    level='success',
                )
            if errors:
                error_summary = '; '.join(errors[:5])
                if len(errors) > 5:
                    error_summary += f' ... and {len(errors) - 5} more errors'
                messages.warning(request, f'Import completed with errors: {error_summary}')
            if imported_count == 0 and not errors:
                messages.warning(request, 'No projects were found in the Excel file.')

        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')

        return redirect('projects:list')


# Document Views
def _documents_visible_to(user):
    """Scoped Document queryset, mirroring ProjectPermissionMixin.

    - Super admin: every document.
    - Admin/manager: documents on projects in their region, plus any
      standalone (project=None) documents they uploaded themselves.
    - Anyone else: documents on projects they own, plus their own uploads
      (standalone or otherwise).
    """
    qs = Document.objects.select_related('project', 'project__region', 'uploaded_by')
    if user.is_super_admin_user:
        return qs
    if (user.is_admin_user or user.is_manager_user
            or getattr(user, 'is_proposal_team_user', False)):
        # Proposal team (head + reps) see documents (Client RFQ etc.) on every
        # project in their region, plus their own standalone uploads.
        return qs.filter(
            Q(project__region=user.region) |
            Q(project__isnull=True, uploaded_by=user)
        ).distinct()
    return qs.filter(
        Q(project__owner=user) |
        Q(uploaded_by=user)
    ).distinct()


class DocumentListView(LoginRequiredMixin, ListView):
    """List all documents with filtering"""
    model = Document
    template_name = 'projects/document_list.html'
    context_object_name = 'documents'
    paginate_by = 25

    def get_queryset(self):
        queryset = _documents_visible_to(self.request.user)

        # Apply filters
        search = self.request.GET.get('search')
        document_type = self.request.GET.get('document_type')
        project = self.request.GET.get('project')

        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(reference_number__icontains=search) |
                Q(vendor_name__icontains=search) |
                Q(description__icontains=search)
            )
        if document_type:
            queryset = queryset.filter(document_type=document_type)
        if project:
            queryset = queryset.filter(project_id=project)

        return queryset.order_by('-uploaded_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = DocumentFilterForm(self.request.GET)
        context['total_count'] = self.get_queryset().count()

        # Group by document type for summary, scoped to what the user can see
        scoped = _documents_visible_to(self.request.user)
        type_counts = scoped.values('document_type').annotate(count=Count('id'))
        context['type_counts'] = {t['document_type']: t['count'] for t in type_counts}

        return context


class DocumentCreateView(LoginRequiredMixin, CreateView):
    """Upload a new document"""
    model = Document
    form_class = DocumentForm
    template_name = 'projects/document_form.html'
    success_url = reverse_lazy('projects:document_list')

    def get_initial(self):
        initial = super().get_initial()
        # Pre-fill project if passed in URL
        project_id = self.request.GET.get('project')
        if project_id:
            initial['project'] = project_id
        return initial

    def form_valid(self, form):
        form.instance.uploaded_by = self.request.user
        messages.success(self.request, 'Document uploaded successfully.')
        response = super().form_valid(form)

        # Redirect back to project detail if project was specified
        project_id = self.request.GET.get('project')
        if project_id:
            return redirect('projects:detail', pk=project_id)
        return response


class DocumentDetailView(LoginRequiredMixin, DetailView):
    """View document details"""
    model = Document
    template_name = 'projects/document_detail.html'
    context_object_name = 'document'

    def get_queryset(self):
        return _documents_visible_to(self.request.user)


class DocumentUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Edit document metadata (and optionally replace the file)."""
    model = Document
    form_class = DocumentForm
    template_name = 'projects/document_form.html'

    def get_queryset(self):
        return _documents_visible_to(self.request.user)

    def test_func(self):
        document = self.get_object()
        user = self.request.user
        return (
            user.is_super_admin_user
            or user.is_admin_user
            or document.uploaded_by == user
        )

    def get_success_url(self):
        return reverse('projects:document_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # If the user replaced the file, delete the previous R2 object so we
        # don't accumulate orphans on every edit.
        old_file = None
        if 'file' in form.changed_data and self.object.pk:
            old_file = Document.objects.get(pk=self.object.pk).file
        response = super().form_valid(form)
        if old_file and old_file.name and old_file.name != self.object.file.name:
            old_file.storage.delete(old_file.name)
        messages.success(self.request, 'Document updated successfully.')
        return response


class DocumentDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Delete a document"""
    model = Document
    template_name = 'projects/document_confirm_delete.html'
    success_url = reverse_lazy('projects:document_list')

    def test_func(self):
        document = self.get_object()
        user = self.request.user
        # Allow deletion if user is admin or uploaded the document
        return user.is_super_admin_user or user.is_admin_user or document.uploaded_by == user

    def form_valid(self, form):
        document = self.get_object()
        # Delete the file from storage
        if document.file:
            document.file.delete(save=False)
        messages.success(self.request, 'Document deleted successfully.')
        return super().form_valid(form)


@login_required
def add_project_document(request, pk):
    """Add document to a specific project"""
    project = get_object_or_404(Project, pk=pk)

    if request.method == 'POST':
        form = DocumentForm(request.POST, request.FILES)
        if form.is_valid():
            document = form.save(commit=False)
            document.project = project
            document.uploaded_by = request.user
            document.save()
            messages.success(request, 'Document uploaded successfully.')
            return redirect('projects:detail', pk=pk)
    else:
        form = DocumentForm(initial={'project': project})

    return render(request, 'projects/document_form.html', {
        'form': form,
        'project': project,
        'title': f'Add Document to {project.project_name}'
    })


@login_required
def add_project_documents_bulk(request, pk):
    """Upload several documents to a project in one go — same Document
    model/fields as add_project_document, just one file input accepting
    multiple files instead of uploading them one at a time. Each file gets
    its own document type, picked in the preview list before upload, same
    as if it had been uploaded individually via add_project_document."""
    project = _scoped_project_or_404(request, pk)

    if request.method == 'POST':
        uploaded_files = request.FILES.getlist('files')
        if not uploaded_files:
            messages.error(request, 'Please choose at least one file to upload.')
            return redirect('projects:add_documents_bulk', pk=pk)

        valid_types = dict(Document.DOCUMENT_TYPE_CHOICES)
        for index, uploaded_file in enumerate(uploaded_files):
            document_type = request.POST.get(f'document_type_{index}') or 'other'
            if document_type not in valid_types:
                document_type = 'other'
            Document.objects.create(
                project=project,
                document_type=document_type,
                name=uploaded_file.name,
                uploaded_by=request.user,
                file=uploaded_file,
            )
        messages.success(
            request, f'{len(uploaded_files)} document(s) uploaded successfully.')
        return redirect('projects:detail', pk=pk)

    return render(request, 'projects/document_bulk_form.html', {
        'project': project,
        'document_type_choices': Document.DOCUMENT_TYPE_CHOICES,
    })


def _can_use_pipeline_email_feature(user):
    """Add Emails (live inbox, attach, delink) is restricted to Sales,
    Admin, and Super Admin — the roles that actually work commercial
    pipeline entries day to day. Everyone else still sees the linked
    email / Email Documents on the project detail page (read-only);
    they just can't attach, pick, or delink one."""
    return bool(
        user.is_super_admin_user or user.is_admin_user
        or getattr(user, 'is_sales_rep_user', False)
    )


def _user_mailbox(user):
    """The ONE mailbox `user` is allowed to browse through "Add Emails" —
    never taken from a request parameter, always derived here from who's
    actually logged in, so there is no value anywhere (URL, POST field,
    hidden JSON) that could be tampered with to reach someone else's
    mailbox. Returns '' if this user has no mailbox to browse — callers
    must treat that as "nothing to show", not fall back to guessing one.

    Resolution order:
    1. This user's own MonitoredMailbox, if an admin has linked one and
       left it active.
    2. The legacy single PIPELINE_EMAIL_MAILBOX setting, but ONLY while
       the MonitoredMailbox table has never had a single row created in
       it — so the feature doesn't go dark for everyone the moment this
       ships, before an admin has had a chance to link anyone.
       The instant even one row has EVER existed, this fallback stops
       applying for everyone else too — checked by whether any row
       exists at all, active or not. That distinction matters: a row
       being deactivated (an admin revoking one employee's access) must
       never cause OTHER unlinked users to start getting the legacy
       mailbox — it would silently hand a deactivated or never-linked
       user access to a mailbox nobody authorised them for. Only "the
       per-employee system has literally never been touched" counts as
       the pre-rollout state this fallback exists for."""
    from .models import MonitoredMailbox
    try:
        return MonitoredMailbox.objects.get(owner=user, is_active=True).email_address
    except MonitoredMailbox.DoesNotExist:
        pass
    if not MonitoredMailbox.objects.exists():
        return settings.PIPELINE_EMAIL_MAILBOX or ''
    return ''


def _scoped_project_or_404(request, pk):
    """Same region/ownership visibility rule as ProjectPermissionMixin
    (used by every class-based project view) — for the plain function-based
    pipeline-email views below, which otherwise did a bare
    get_object_or_404(Project, pk=pk) and let any Sales Rep act on any
    OTHER rep's project just by guessing its id, bypassing the scoping
    every other project view already enforces."""
    mixin = ProjectPermissionMixin()
    mixin.request = request
    return get_object_or_404(mixin.get_queryset(), pk=pk)


def _delink_active_pipeline_email(project, user):
    """Mark the project's currently active linked email (if any) as
    delinked. This NEVER deletes the PipelineEmail row or any Document it
    created — it only flips is_active off, so the email drops out of the
    'currently linked' widget while its documents stay exactly where they
    are under Project Documents."""
    active = project.linked_emails.filter(is_active=True).first()
    if active:
        active.is_active = False
        active.delinked_by = user
        active.delinked_at = timezone.now()
        active.save(update_fields=['is_active', 'delinked_by', 'delinked_at'])
    return active


def _safe_attachment_filename(filename):
    """A crafted attachment filename like '../../x.pdf' would otherwise
    reach Django's storage backend as-is and raise an uncaught
    SuspiciousFileOperation mid-attach (see _attach_email_to_project) —
    os.path.basename() strips any directory-traversal components before
    that ever happens, and the length clamp keeps it comfortably under
    Document.name's max_length=255 (a real attachment filename is never
    anywhere near this long).

    basename() alone still lets '.' or '..' straight through unchanged —
    those aren't traversal sequences by themselves, but Django's storage
    backend treats them as reserved names and raises SuspiciousFileOperation
    for them too, which would otherwise still crash this uncaught."""
    name = os.path.basename((filename or '').strip()) or 'attachment'
    if name in ('.', '..'):
        name = 'attachment'
    return name[:200]


def _attach_email_to_project(project, user, mailbox, message_id, attachment_types=None):
    """Fetch one message from `mailbox` (Microsoft Graph —
    see projects/graph_mail.py) and turn it into a PipelineEmail + Document
    rows on `project`, tagged with source_pipeline_email (see
    projects/models.py) so they show up under Project Documents.

    attachment_types is an optional {index: document_type} map, index being
    the attachment's 0-based position within this email — NOT its filename.
    Filename used to be the key, but the picker UI's filenames come from
    Graph's attachment listing while parse_eml_file() below re-derives
    filenames independently (different RFC 2047 decoding, synthesised
    "attachment_N.ext" names when none is present) — a mismatch between the
    two silently reverted the user's choice to 'vendor_quotation' with no
    indication anything had gone wrong. Position is the one thing both
    parsers agree on, since both just walk the same MIME parts in order.
    Anything missing/invalid falls back to 'vendor_quotation', matching this
    feature's original default before per-document types existed.

    If an email is already linked, adding a new one automatically delinks
    the old one first — this matches "add a new email, removing the one
    before" — but the old email's documents are never touched or removed.
    That delink, and every row created below, happens inside one
    transaction.atomic() block, and the delink happens LAST: if anything
    above it fails (a bad filename, a database error), the whole thing
    rolls back and the previously-linked email is left exactly as it was,
    rather than ending up delinked with a half-built replacement.

    Raises GraphMailError/EmailParseError on failure — callers decide how
    to surface that. Returns (pipeline_email, created_document_count)."""
    attachment_types = attachment_types or {}
    valid_types = dict(Document.DOCUMENT_TYPE_CHOICES)

    raw_bytes = graph_mail.fetch_raw_message_bytes(mailbox, message_id)
    parsed = parse_eml_file(io.BytesIO(raw_bytes))

    raw_filename = f"pipeline_email_{project.pk}_{timezone.now():%Y%m%d%H%M%S}.eml"

    with transaction.atomic():
        # Captured BEFORE the new email is created, and delinked as this
        # specific object at the end — not re-queried afterward. The new
        # PipelineEmail below is also created with is_active=True, so a
        # fresh "whichever email is active" query at the end would match
        # both of them and — ordered newest-first — delink the one just
        # created instead of the actual previous one.
        previous_active = project.linked_emails.filter(is_active=True).first()

        pipeline_email = PipelineEmail.objects.create(
            project=project,
            source_mailbox=mailbox,
            subject=parsed['subject'],
            sender_name=parsed['sender_name'],
            sender_email=parsed['sender_email'],
            recipients=parsed['recipients'],
            sent_at=parsed['sent_at'],
            body_text=parsed['body_text'],
            raw_file=ContentFile(raw_bytes, name=raw_filename),
            raw_filename=raw_filename,
            is_active=True,
            linked_by=user,
        )

        created_count = 0
        for index, attachment in enumerate(parsed['attachments']):
            document_type = attachment_types.get(str(index)) or 'vendor_quotation'
            if document_type not in valid_types:
                document_type = 'vendor_quotation'
            safe_filename = _safe_attachment_filename(attachment['filename'])
            document = Document(
                project=project,
                document_type=document_type,
                name=safe_filename,
                uploaded_by=user,
                source_pipeline_email=pipeline_email,
            )
            document.file.save(
                safe_filename,
                ContentFile(attachment['content']),
                save=False,
            )
            document.save()
            created_count += 1

        if previous_active:
            previous_active.is_active = False
            previous_active.delinked_by = user
            previous_active.delinked_at = timezone.now()
            previous_active.save(update_fields=['is_active', 'delinked_by', 'delinked_at'])

    return pipeline_email, created_count


@login_required
def link_pipeline_email(request, pk):
    """Attach an email from the live monitored inbox to an existing
    commercial pipeline entry — see _attach_email_to_project()."""
    project = _scoped_project_or_404(request, pk)
    if not _can_use_pipeline_email_feature(request.user):
        messages.error(request, 'Add Emails is restricted to Sales, Admin, and Super Admin users.')
        return redirect('projects:detail', pk=pk)

    mailbox = _user_mailbox(request.user)

    if request.method == 'POST':
        message_id = request.POST.get('message_id')
        if not message_id:
            messages.error(request, 'Please choose an email to attach.')
            return redirect('projects:link_pipeline_email', pk=pk)
        if not mailbox:
            messages.error(request, 'No mailbox is linked to your account — ask an admin to link one.')
            return redirect('projects:link_pipeline_email', pk=pk)

        # Keyed by position within the email, not filename — see
        # _attach_email_to_project()'s docstring for why.
        attachment_types = dict(zip(
            request.POST.getlist('doc_index'), request.POST.getlist('doc_type')))

        try:
            pipeline_email, created_count = _attach_email_to_project(
                project, request.user, mailbox, message_id, attachment_types=attachment_types)
        except (graph_mail.GraphMailError, EmailParseError) as exc:
            messages.error(request, f'Could not attach that email: {exc}')
            return redirect('projects:link_pipeline_email', pk=pk)

        if created_count:
            messages.success(
                request,
                f'Email "{pipeline_email.subject or "(no subject)"}" linked — '
                f'{created_count} document(s) added to Project Documents.'
            )
        else:
            messages.success(
                request,
                f'Email "{pipeline_email.subject or "(no subject)"}" linked. '
                f'No attachments were found inside it.'
            )
        return redirect('projects:detail', pk=pk)

    inbox_messages = []
    inbox_error = None
    if not mailbox:
        inbox_error = 'No mailbox is linked to your account yet — ask an admin to link one in Django admin (Projects → Monitored mailboxes).'
    else:
        try:
            inbox_messages = graph_mail.list_inbox_messages(mailbox)
        except graph_mail.GraphMailError as exc:
            inbox_error = str(exc)
    _add_attachment_urls(inbox_messages, project.pk)

    return render(request, 'projects/pipeline_email_form.html', {
        'mailbox': mailbox,
        'project': project,
        'active_email': project.linked_emails.filter(is_active=True).first(),
        'inbox_messages': inbox_messages,
        'inbox_error': inbox_error,
        'document_type_choices': Document.DOCUMENT_TYPE_CHOICES,
    })


def _add_attachment_urls(inbox_messages, project_pk):
    """Precompute each attachment's open/preview URL onto its dict, so the
    template doesn't need to branch on whether a project exists yet.

    message_id/attachment_id ride as query params, not path segments —
    Microsoft documents that Graph ids may contain characters (notably '/')
    that Django's <str:> path converter rejects. With ids in the path,
    reverse() raised NoReverseMatch for any such message, which took this
    whole page down — one oddly-encoded id anywhere in the inbox, and
    nobody could use Add Emails at all. Query params have no such
    restriction. `mailbox` is deliberately NOT one of these params — see
    _user_mailbox()'s docstring: the mailbox is always derived from
    request.user server-side, on both this list view and
    view_pipeline_inbox_attachment, precisely so there's never a client-
    supplied value that could point at someone else's mailbox."""
    if project_pk:
        url_name, kwargs = 'projects:view_pipeline_inbox_attachment', {'pk': project_pk}
    else:
        url_name, kwargs = 'projects:view_pipeline_inbox_attachment_new', {}
    base_url = reverse(url_name, kwargs=kwargs)
    for msg in inbox_messages:
        for att in msg['attachments']:
            att['url'] = base_url + '?' + urlencode({
                'message_id': msg['id'], 'attachment_id': att['id']})


@login_required
def link_pipeline_email_new(request):
    """Same live inbox as link_pipeline_email, but for a pipeline entry
    that doesn't exist yet (the "Add Emails" button on the create form).
    Nothing is saved to the database here — picking an email just carries
    its message_id back to the create form via a query param, where it
    rides along in that form's own draft-autosave state (ProjectForm.
    picked_email_json) until the entry is actually created."""
    if not _can_use_pipeline_email_feature(request.user):
        messages.error(request, 'Add Emails is restricted to Sales, Admin, and Super Admin users.')
        return redirect('projects:create')

    mailbox = _user_mailbox(request.user)

    if request.method == 'POST':
        message_id = request.POST.get('message_id')
        if not message_id:
            messages.error(request, 'Please choose an email.')
            return redirect('projects:link_pipeline_email_new')
        return redirect(reverse('projects:create') + '?' + urlencode({'picked_email': message_id}))

    inbox_messages = []
    inbox_error = None
    if not mailbox:
        inbox_error = 'No mailbox is linked to your account yet — ask an admin to link one in Django admin (Projects → Monitored mailboxes).'
    else:
        try:
            inbox_messages = graph_mail.list_inbox_messages(mailbox)
        except graph_mail.GraphMailError as exc:
            inbox_error = str(exc)
    _add_attachment_urls(inbox_messages, None)

    return render(request, 'projects/pipeline_email_form.html', {
        'mailbox': mailbox,
        'project': None,
        'active_email': None,
        'inbox_messages': inbox_messages,
        'inbox_error': inbox_error,
    })


# Content types safe to render inline in the browser from
# view_pipeline_inbox_attachment below. Both the bytes and the content
# type there come straight from an attachment in the externally-reachable
# monitored mailbox — anyone who can email that inbox controls both. An
# attachment named "quote.html" with contentType "text/html" (or
# "image/svg+xml", which can also carry a <script>) served inline would
# execute the sender's JavaScript on the ERP's own origin, with whichever
# staff member's session happened to click it. Anything not on this list
# is forced to download instead of rendering in the browser.
_INLINE_SAFE_CONTENT_TYPES = {
    'application/pdf',
    'image/png', 'image/jpeg', 'image/gif', 'image/webp', 'image/bmp',
}


@login_required
def view_pipeline_inbox_attachment(request, pk=None):
    """Open/preview one attachment straight from the live inbox, before it's
    attached to anything — read-only, nothing is saved. Works both for an
    existing pipeline entry (pk given) and while creating a new one (pk
    is None — see link_pipeline_email_new).

    message_id/attachment_id come from query params, not path segments —
    see _add_attachment_urls() for why. The mailbox itself is never a
    request param at all — see _user_mailbox()'s docstring."""
    message_id = request.GET.get('message_id')
    attachment_id = request.GET.get('attachment_id')
    if not message_id or not attachment_id:
        raise Http404('Missing message_id/attachment_id.')
    if not _can_use_pipeline_email_feature(request.user):
        messages.error(request, 'Add Emails is restricted to Sales, Admin, and Super Admin users.')
        if pk is not None:
            return redirect('projects:detail', pk=pk)
        return redirect('projects:create')
    # Scoped like every other pipeline-email view, and checked only after
    # the permission gate above: a bare get_object_or_404(Project, ...)
    # ahead of the gate told a user who cannot use this feature at all
    # which project ids exist, by returning 404 for one and a redirect for
    # the other.
    if pk is not None:
        _scoped_project_or_404(request, pk)
    mailbox = _user_mailbox(request.user)
    if not mailbox:
        raise Http404('No mailbox is linked to your account.')
    try:
        filename, content_type, content = graph_mail.fetch_attachment_bytes(
            mailbox, message_id, attachment_id)
    except graph_mail.GraphMailError as exc:
        messages.error(request, f'Could not open that document: {exc}')
        if pk is not None:
            return redirect('projects:link_pipeline_email', pk=pk)
        return redirect('projects:link_pipeline_email_new')

    disposition = 'inline' if content_type in _INLINE_SAFE_CONTENT_TYPES else 'attachment'
    response = HttpResponse(content, content_type=content_type)
    response['Content-Disposition'] = (
        f'{disposition}; filename="{_safe_header_filename(filename)}"')
    # Defense in depth on top of the allowlist above: stops a browser from
    # ever second-guessing the declared content type and rendering
    # something dangerous anyway.
    response['X-Content-Type-Options'] = 'nosniff'
    return response


def _safe_header_filename(filename):
    """Strip characters that would break out of the quoted filename or inject
    a new header line. `filename` comes straight from Graph — untrusted,
    external data — so this must fail closed rather than let a crafted
    filename (e.g. containing '\\r\\n') reach HttpResponse and either crash
    the request or smuggle an extra header."""
    cleaned = (filename or '').replace('\r', '').replace('\n', '').replace('"', "'")
    return cleaned or 'attachment'


@login_required
@require_POST
def delink_pipeline_email(request, pk):
    """Remove the currently linked email from a pipeline entry. The
    Documents it created are NEVER deleted — only the email link itself."""
    project = _scoped_project_or_404(request, pk)
    if not _can_use_pipeline_email_feature(request.user):
        messages.error(request, 'Add Emails is restricted to Sales, Admin, and Super Admin users.')
        return redirect('projects:detail', pk=pk)
    delinked = _delink_active_pipeline_email(project, request.user)
    if delinked:
        messages.success(request, 'Email delinked. Its documents remain under Project Documents.')
    else:
        messages.info(request, 'There is no linked email to remove.')
    return redirect('projects:detail', pk=pk)


# ─── Commercial Pipeline Revisions ───────────────────────────

def _build_project_snapshot(project):
    """Return a dict capturing the Project's state at this moment.

    Foreign keys are expanded to name/code so the snapshot remains
    readable even if related rows are later renamed or deleted.
    """
    def _dec(v):
        return None if v is None else str(v)

    return {
        'snapshot_version': 1,
        'identity': {
            'serial_number': project.serial_number,
            'project_name': project.project_name,
            'proposal_reference': project.proposal_reference,
            'client_rfq_reference': project.client_rfq_reference,
            'po_number': project.po_number,
        },
        'parties': {
            'customer': project.customer,
            'end_user': project.end_user,
            'owner': (project.owner.get_full_name() or project.owner.username) if project.owner else None,
            'owner_username': project.owner.username if project.owner else None,
            'contact_with': project.contact_with,
        },
        'classification': {
            'project_stage': project.project_stage,
            'project_stage_display': project.get_project_stage_display() if project.project_stage else None,
            'region': {
                'code': project.region.code if project.region else None,
                'name': project.region.name if project.region else None,
                'currency': project.region.currency if project.region else None,
            },
            'status': {
                'name': project.status.name if project.status else None,
                'category': project.status.category if project.status else None,
                'color': project.status.color if project.status else None,
            },
        },
        'dates': {
            'submission_deadline': project.submission_deadline.isoformat() if project.submission_deadline else None,
            'estimated_po_date': project.estimated_po_date.isoformat() if project.estimated_po_date else None,
            'year': project.year,
            'po_award_quarter': project.po_award_quarter,
        },
        'financials': {
            'estimated_value': _dec(project.estimated_value),
            'estimated_value_usd': _dec(project.estimated_value_usd),
            'estimated_value_per_annum': _dec(project.estimated_value_per_annum),
            'estimated_gp': _dec(project.estimated_gp),
            'success_quotient': _dec(project.success_quotient),
            'minimum_achievement': _dec(project.minimum_achievement),
            'actual_sales': _dec(project.actual_sales),
        },
        'narrative': {
            'remarks': project.remarks,
            'notes': project.notes,
            'portal_url': project.portal_url,
        },
        'metadata': {
            'created_at': project.created_at.isoformat() if project.created_at else None,
            'updated_at': project.updated_at.isoformat() if project.updated_at else None,
            'created_by': (project.created_by.get_full_name() or project.created_by.username) if project.created_by else None,
        },
    }


class ProjectRevisionCreateView(ProjectPermissionMixin, View):
    """POST: create a new revision of the current project state."""

    def test_func(self):
        # Same as ProjectUpdateView: finance has view access, not write access.
        return not getattr(self.request.user, 'is_finance_team_user', False)

    def post(self, request, pk):
        project = get_object_or_404(Project, pk=pk)
        label = ProjectRevision.next_label_for(project)
        notes = (request.POST.get('notes') or '').strip()
        ProjectRevision.objects.create(
            project=project,
            revision_label=label,
            snapshot=_build_project_snapshot(project),
            notes=notes,
            created_by=request.user,
        )
        messages.success(request, f'Revision {label} saved.')
        return redirect('projects:detail', pk=project.pk)


class ProjectRevisionDetailView(ProjectPermissionMixin, View):
    """GET: render a past revision as a read-only page."""

    def test_func(self):
        return True

    def get(self, request, pk, revision_pk):
        project = get_object_or_404(Project, pk=pk)
        revision = get_object_or_404(ProjectRevision, pk=revision_pk, project=project)
        return render(request, 'projects/project_revision_detail.html', {
            'project': project,
            'revision': revision,
            'snapshot': revision.snapshot,
        })


# ─── Pipeline Print to PDF ───────────────────────────────────
def _apply_pipeline_filters(request, queryset):
    """Apply the same filter parameters ProjectListView uses.

    Kept here so the print view renders exactly what the user sees on
    screen after clicking Filter, without rebuilding the logic.
    """
    region_map = {
        'LNUK':      ['UK', 'GLB', 'LNUK'],
        'LNA':       ['LNA'],
        'PA':        ['PA'],
        'NEO-Dubai': ['NEO-Dubai'],
        'NEO-KSA':   ['NEO-KSA'],
    }
    g = request.GET
    search   = g.get('search')
    region   = g.get('region')
    year     = g.get('year')
    status   = g.get('status')
    workflow_stage = g.get('workflow_stage')
    quarter  = g.get('quarter')
    owner    = g.get('owner')
    if search:
        queryset = queryset.filter(
            Q(project_name__icontains=search) |
            Q(proposal_reference__icontains=search) |
            Q(client_rfq_reference__icontains=search)
        )
    if region:
        codes = region_map.get(region, [])
        if codes:
            queryset = queryset.filter(region__code__in=codes)
    if year:
        queryset = queryset.filter(year=year)
    if status:
        queryset = queryset.filter(status_id=status)
    if workflow_stage == 'none':
        queryset = queryset.filter(costing_sheets__isnull=True)
    elif workflow_stage:
        queryset = queryset.filter(
            costing_sheets__workflow_stage=workflow_stage
        ).distinct()
    if quarter:
        queryset = queryset.filter(po_award_quarter=quarter)
    if owner:
        queryset = queryset.filter(owner_id=owner)
    return queryset.order_by('-updated_at')


@login_required
def pipeline_print_pdf(request):
    """Render the currently-filtered Commercial Pipeline list as PDF.

    Honors the same ?search=/region/status/workflow_stage/owner/year/quarter
    querystring as the on-screen list view so users can preview in the
    browser then "Print PDF" and get the exact same rows.
    """
    from io import BytesIO
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from django.http import HttpResponse
    from django.utils import timezone

    # Permission scoping — mirror ProjectPermissionMixin.get_queryset
    user = request.user
    qs = Project.objects.select_related('status', 'region', 'owner').all()
    if not user.is_super_admin_user:
        if user.is_admin_user or user.is_manager_user:
            qs = qs.filter(region=user.region)
        else:
            qs = qs.filter(owner=user)
    qs = _apply_pipeline_filters(request, qs)

    total_value = sum((p.estimated_value or Decimal('0')) for p in qs)

    # Filter summary line for the PDF header
    summary_parts = []
    for key in ('search', 'region', 'year', 'status', 'workflow_stage', 'quarter', 'owner'):
        val = request.GET.get(key)
        if val:
            summary_parts.append(f'{key}={val}')
    filter_summary = ' · '.join(summary_parts) if summary_parts else 'All entries (no filter)'

    # Build PDF
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=10*mm, rightMargin=10*mm,
        topMargin=12*mm, bottomMargin=12*mm,
        title='Commercial Pipeline',
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Heading1'], fontSize=16, alignment=TA_LEFT, textColor=colors.HexColor('#1a1a1a'))
    meta = ParagraphStyle('meta', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#6c757d'))
    small = ParagraphStyle('small', parent=styles['Normal'], fontSize=8, leading=10)

    elements = []
    elements.append(Paragraph('Commercial Pipeline', h1))
    elements.append(Paragraph(
        f'Generated {timezone.localtime().strftime("%d %b %Y %H:%M")} · {qs.count()} entries · Filter: {filter_summary}',
        meta,
    ))
    elements.append(Spacer(1, 6))

    # Paragraph parses input as mini-XML, so any '<' or '&' in user-typed
    # fields would crash the parser ("unclosed tags"). Escape first.
    from xml.sax.saxutils import escape as _xml_escape
    def _safe(text, limit=None):
        s = '' if text is None else str(text)
        if limit:
            s = s[:limit]
        return _xml_escape(s) or '-'

    header = ['Reference', 'Project', 'Client', 'Region', 'Status', 'Value', 'Owner', 'Updated']
    data = [header]
    for p in qs:
        owner_name = (p.owner.get_full_name() or p.owner.username) if p.owner_id else '-'
        data.append([
            Paragraph(_safe(p.proposal_reference), small),
            Paragraph(_safe(p.project_name, 80), small),
            Paragraph(_safe(p.customer, 40), small),
            _safe(p.region.code) if p.region_id else '-',
            _safe(p.status.name) if p.status_id else '-',
            f'{(p.estimated_value or 0):,.0f}',
            _safe(owner_name),
            p.updated_at.strftime('%d %b %Y') if p.updated_at else '-',
        ])
    data.append([
        Paragraph('<b>TOTAL</b>', small), '', '', '', '',
        f'{total_value:,.0f}', '', '',
    ])

    tbl = Table(
        data,
        colWidths=[28*mm, 70*mm, 45*mm, 18*mm, 28*mm, 24*mm, 32*mm, 22*mm],
        repeatRows=1,
    )
    tbl.setStyle(TableStyle([
        ('BACKGROUND',  (0, 0), (-1, 0), colors.HexColor('#1a1a1a')),
        ('TEXTCOLOR',   (0, 0), (-1, 0), colors.white),
        ('FONTNAME',    (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, 0), 8),
        ('ALIGN',       (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN',       (5, 1), (5, -1), 'RIGHT'),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE',    (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND',  (0, -1), (-1, -1), colors.HexColor('#C41E3A')),
        ('TEXTCOLOR',   (0, -1), (-1, -1), colors.white),
        ('FONTNAME',    (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID',        (0, 0), (-1, -1), 0.3, colors.HexColor('#dee2e6')),
    ]))
    elements.append(tbl)

    doc.build(elements)
    resp = HttpResponse(buf.getvalue(), content_type='application/pdf')
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    resp['Content-Disposition'] = f'inline; filename="commercial_pipeline_{stamp}.pdf"'
    return resp


@login_required
def project_recovery(request):
    """Super-admin browser tool to restore a hard-deleted Project from a Render
    Point-in-Time-Recovery copy (set RECOVERY_DATABASE_URL on the web service).
    Preview first, then apply. See projects/recovery.py."""
    if not getattr(request.user, 'is_super_admin_user', False):
        messages.error(request, 'Project recovery is restricted to super administrators.')
        return redirect('projects:list')

    from .recovery import (recovery_available, run_recovery, run_recovery_all,
                           list_deleted_projects)
    ctx = {
        'available': recovery_available(),
        'reference': (request.POST.get('reference') or '').strip(),
        'pk_raw': (request.POST.get('pk') or '').strip(),
        'with_cascaded': request.POST.get('with_cascaded', '1') == '1',
        'report': None,
        'all_report': None,
        'deleted_list': None,
        'error': None,
        'committed': False,
    }
    # Always show how many projects are missing (whole-pipeline view).
    if ctx['available']:
        try:
            ctx['deleted_list'] = list_deleted_projects()
        except (RuntimeError, Exception):
            ctx['deleted_list'] = None

    if request.method == 'POST' and ctx['available']:
        action = request.POST.get('action')   # preview / apply / preview_all / apply_all
        try:
            if action in ('preview_all', 'apply_all'):
                ctx['all_report'] = run_recovery_all(
                    with_cascaded=ctx['with_cascaded'], commit=(action == 'apply_all'))
                ctx['committed'] = ctx['all_report']['committed']
                if ctx['committed']:
                    messages.success(
                        request, f"Recovered {ctx['all_report']['count']} deleted project(s).")
                    ctx['deleted_list'] = list_deleted_projects()   # refresh
            else:
                pk = int(ctx['pk_raw']) if ctx['pk_raw'].isdigit() else None
                report = run_recovery(
                    reference=ctx['reference'] or None, pk=pk,
                    with_cascaded=ctx['with_cascaded'], commit=(action == 'apply'))
                ctx['report'] = report
                ctx['committed'] = report['committed']
                if report['committed']:
                    messages.success(
                        request, f"Recovered project {report['reference']} (id {report['pk']}).")
        except (ValueError, RuntimeError) as e:
            ctx['error'] = str(e)
    return render(request, 'projects/project_recovery.html', ctx)
