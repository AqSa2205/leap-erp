import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, FormView, TemplateView
from django.urls import reverse_lazy, reverse
from django.db.models import Q, Count, Sum
from django.db.models import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import PermissionDenied
from datetime import datetime, date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Employee, Asset, AssetAssignment, Vehicle, VehicleDocument, EmployeeDocument, LeaveType, Holiday, LeaveEntitlement, LeaveRecord, AttendanceRecord, AttendanceSettings, WorkingDay, WFHRecord, LeaveRequest, AttendanceException, LeaveRevokeRequest, AttendanceExceptionRevokeRequest
from .forms import (
    EmployeeForm, EmployeeFilterForm, EmployeeImportForm,
    AssetForm, AssetFilterForm, AssetImportForm,
    VehicleForm, VehicleFilterForm, EmployeeDocumentForm, VehicleDocumentForm,
    LeaveTypeForm, HolidayForm, WorkingDayForm, WFHRecordForm,
    AttendanceSettingsForm, LeaveRequestForm, AttendanceExceptionForm,
    EmployeeHierarchyForm, ExceptionGrantForm,
)
from .leave_services import generate_year_entitlements
from .attendance_services import derive_status
from .attendance_matrix import period_range, build_matrix, display_status_no_record, cell_time_lines
from .scoping import scoped_employee_ids, scope_employee_queryset, scope_asset_queryset, can_manage_hr_scoped


class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Gate for the Administration section.

    The `admin` role is deliberately NOT here. Administration is owned by
    super_admin and erp_admin; `admin` keeps its access everywhere else in the
    app (costing, procurement, pipeline) but is locked out of this section
    entirely, so hiding it from the sidebar is matched by the views refusing a
    direct URL too.
    """
    def test_func(self):
        u = self.request.user
        return u.is_super_admin_user or u.is_erp_admin_user


class SuperAdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Historically super-admin-only; erp_admin now has the whole
    Administration section, so it passes here too. Kept as a separate mixin
    because the name marks these as the more sensitive screens."""
    def test_func(self):
        u = self.request.user
        return u.is_super_admin_user or u.is_erp_admin_user


class HRScopedAccessMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Access gate for the shared HR feature set (attendance, leave, assets,
    team exceptions, org chart). Passes for the company-wide admin tiers
    (super_admin / admin / erp_admin) AND the team-scoped roles (project_manager
    / site_manager / document_controller).

    Views using this mixin MUST scope their own employee querysets through
    ``scope_employee_queryset(self.request.user, qs)`` — this only opens the
    door; it does not restrict WHICH rows are shown. For a scoped role with no
    reports, the queryset comes back empty and they simply see nothing."""
    def test_func(self):
        return can_manage_hr_scoped(self.request.user)


def is_designated_approver(user):
    """True if `user` currently holds active approval authority for leave
    requests — an active LeaveDashboardAccess grant (managed on the Leave
    Access page, sidebar Leave -> Leave Access). This — not a username check
    — is what makes Aamna Khan and Ali Sultan (or whoever holds these rows)
    able to actually approve/reject. Note this is NOT the same as being a
    Super Admin: Super Admin status alone grants viewing, never approval
    authority by itself (see can_view_leave_dashboard) — and, since override
    access became configurable, not necessarily override authority either
    (see has_override_access)."""
    from .models import LeaveDashboardAccess
    return LeaveDashboardAccess.objects.filter(user=user, is_active=True).exists()


def has_override_access(user):
    """Access gate for the Super Admin 'override' escape hatch on leave
    requests (force-approve/disapprove when the real approver is
    unavailable) — configured via OverrideAccessSettings, managed from the
    Leave Access page (sidebar Leave -> Leave Access). Exactly one mode is
    authoritative: by default (MODE_ALL_SUPER_ADMINS) every Super Admin has
    it, matching the original hardcoded behavior; narrowed to
    MODE_SPECIFIC_ROLES or MODE_SPECIFIC_EMPLOYEES, only that hand-picked
    set does — Super Admin status alone no longer suffices once narrowed."""
    from .models import OverrideAccessSettings, OverrideAccessRole, OverrideAccessEmployee
    config = OverrideAccessSettings.get_solo()
    if config.mode == OverrideAccessSettings.MODE_SPECIFIC_ROLES:
        return bool(user.role_id and OverrideAccessRole.objects.filter(role_id=user.role_id).exists())
    if config.mode == OverrideAccessSettings.MODE_SPECIFIC_EMPLOYEES:
        return OverrideAccessEmployee.objects.filter(user=user).exists()
    return user.is_super_admin_user


def can_view_leave_dashboard(user):
    """Access gate for the master Leave Request dashboard (the FIFO queue +
    detail pages). Super Admins always see it; anyone with an active
    LeaveDashboardAccess grant also sees it, AND that same grant is what
    makes them a real approver on new leave requests (see
    submit_leave_request); anyone with override access (see
    has_override_access) must be able to see it too, or granting them
    override authority would be pointless — they couldn't reach the page to
    use it."""
    # erp_admin is included because Leave Requests sits in the Administration
    # section, which that role now owns in full — without it the sidebar would
    # offer a link the view refuses.
    return bool(user.is_super_admin_user or user.is_erp_admin_user
                or is_designated_approver(user) or has_override_access(user))


def _is_head_manager_role(user):
    """Any of the 'head of department' roles that get organization-wide
    visibility into the attendance-exception queue even if they currently
    have zero direct reports of their own (e.g. a newly-appointed head)."""
    return bool(
        user.is_admin_user or user.is_super_admin_user or user.is_erp_admin_user or user.is_ai_head_user
        or user.is_finance_head_user or user.is_procurement_manager_user or user.is_proposal_head_user
    )


def can_view_team_exceptions(user):
    """Access gate for the Team Exceptions page: actively assigned as a main
    manager OR a secondary manager to at least one active employee, OR holds
    one of the head-manager roles (checked via the confirmed is_*_user
    boolean properties)."""
    if _is_head_manager_role(user):
        return True
    emp = getattr(user, 'employee_profile', None)
    if not emp:
        return False
    return bool(
        emp.main_reports.filter(is_active=True).exists()
        or emp.secondary_reports.filter(is_active=True).exists()
    )


def is_direct_manager_of_anyone(user):
    """True if `user` is the live main_manager of at least one active
    employee — the org-chart-derived counterpart to the Role-based
    scoped_employee_ids(). Deliberately checks direct reports only (never
    walks get_downstream_employee_ids) and deliberately does NOT touch
    scoped_employee_ids/scope_asset_queryset — a manager's leave-logging
    power must not leak into Asset/Attendance visibility as a side effect."""
    emp = getattr(user, 'employee_profile', None)
    return bool(emp and emp.main_reports.filter(is_active=True).exists())


def can_decide_attendance_exception(user, exc):
    """True if user is the assigned main manager, OR user is HR (global
    visibility/override), OR user has an employee_profile that is upstream of
    exc.employee in the main-manager hierarchy (override path), OR user has an
    employee_profile that is one of exc.employee's secondary_managers (checked
    directly/flatly, not via the recursive is_manager_of walk — secondary-
    manager grants are non-hierarchical).

    NOTE: a secondary manager's True here must ONLY feed the OVERRIDE
    eligibility gate, never the plain 'decide' action — that action's
    authorization stays strictly exc.employee.main_manager.user_id ==
    request.user.id (see TeamExceptionsView.post's 'decide' branch), because
    secondary managers are, by definition, not main managers.

    Deliberately checks the LIVE exc.employee.main_manager relationship, not
    the exc.main_manager snapshot taken at submission time — if the employee's
    main_manager is assigned/corrected via the Org Chart page after a request
    was already submitted, the newly (or correctly) assigned manager must be
    recognized immediately, not permanently stuck seeing only the Override
    control for that pre-existing request. exc.main_manager itself is left
    untouched as historical/audit metadata (who was originally supposed to
    decide) — it's just no longer what authorization reads."""
    if exc.employee.user_id and exc.employee.user_id == user.id:
        # Self-approval guardrail: nobody — not even a Super Admin/HR with
        # otherwise-blanket override rights — may decide or override their
        # own attendance exception request.
        return False
    if user.is_document_controller_user:
        # Document Controllers get the full (subtree-scoped) HR feature set but
        # NO approval authority anywhere — so they can never decide or override
        # an exception, even for an employee they'd otherwise manage.
        return False
    if exc.employee.main_manager and exc.employee.main_manager.user_id == user.id:
        return True
    if user.is_super_admin_user or user.is_admin_user or user.is_erp_admin_user:
        return True
    emp = getattr(user, 'employee_profile', None)
    if not emp:
        return False
    if emp.is_manager_of(exc.employee):
        return True
    return exc.employee.secondary_managers.filter(pk=emp.pk).exists()


@login_required
def hr_dashboard(request):
    """Comprehensive HR Admin Dashboard with employees, assets, vehicles, and assignments."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('dashboard:index')

    from hr.models import AssetHandover
    employees = Employee.objects.all()
    assets = Asset.objects.all()
    vehicles = Vehicle.objects.all()

    # Employee stats
    total_employees = employees.count()
    active_employees = employees.filter(is_active=True).count()
    contract_breakdown = {}
    for ct in ['permanent', 'yearly', 'ajeer']:
        contract_breakdown[ct] = employees.filter(contract_type=ct).count()

    # Top nationalities
    nationalities = employees.values('nationality').annotate(
        count=Count('id')
    ).order_by('-count')[:6]

    # Top deployments
    deployments = employees.values('deployment').annotate(
        count=Count('id')
    ).exclude(deployment='').order_by('-count')[:6]

    # Asset stats. "Assigned" means the asset has an active AssetHandover -
    # AssetHandover, not the legacy in_stock field, is the source of truth
    # for current custody. Decommissioned assets are excluded from both
    # buckets (retired, not available and not held by anyone).
    total_assets = assets.count()
    assigned_asset_ids = set(AssetHandover.objects.filter(
        status='active', asset__isnull=False).values_list('asset_id', flat=True))
    assets_assigned = len(assigned_asset_ids)
    assets_in_stock = assets.exclude(pk__in=assigned_asset_ids).exclude(is_decommissioned=True).count()
    total_asset_value = assets.aggregate(val=Sum('price'))['val'] or 0

    # Vehicle stats
    total_vehicles = vehicles.count()
    valid_vehicles = vehicles.filter(vehicle_status='valid').count()
    compliance_issues = sum(1 for v in vehicles if v.has_compliance_issue)

    # Vehicle makers breakdown
    makers = vehicles.values('vehicle_maker').annotate(
        count=Count('id')
    ).order_by('-count')[:6]

    # Employee-Asset-Vehicle assignments, built from active AssetHandover
    # rows (the source of truth for custody) rather than fuzzy name
    # matching against the legacy employee_name/driver_name fields.
    assigned_asset_handovers = AssetHandover.objects.filter(
        status='active', asset__isnull=False).select_related('asset', 'employee')
    assigned_vehicle_handovers = AssetHandover.objects.filter(
        status='active', vehicle__isnull=False).select_related('vehicle', 'employee')
    assets_by_employee = {}
    for h in assigned_asset_handovers:
        assets_by_employee.setdefault(h.employee_id, []).append(h.asset)
    vehicles_by_employee = {}
    assigned_vehicle_ids = set()
    for h in assigned_vehicle_handovers:
        vehicles_by_employee.setdefault(h.employee_id, []).append(h.vehicle)
        assigned_vehicle_ids.add(h.vehicle_id)

    assignments = []
    unassigned_employees = []
    for emp in employees.filter(is_active=True).order_by('full_name'):
        emp_assets = assets_by_employee.get(emp.pk, [])
        emp_vehicles = vehicles_by_employee.get(emp.pk, [])
        if emp_assets or emp_vehicles:
            assignments.append({
                'employee': emp,
                'assets': emp_assets,
                'vehicles': emp_vehicles,
            })
        else:
            unassigned_employees.append(emp)

    # Unassigned assets & vehicles
    unassigned_assets = assets.exclude(pk__in=assigned_asset_ids).exclude(is_decommissioned=True)
    unassigned_vehicles = vehicles.exclude(pk__in=assigned_vehicle_ids)

    # Recent employees
    recent_employees = employees.order_by('-created_at')[:5]

    # Asset types breakdown
    asset_types = assets.values('asset_type').annotate(count=Count('id')).exclude(asset_type='').order_by('-count')[:8]

    # JSON data for Chart.js
    nationality_labels = [n['nationality'] or 'Unknown' for n in nationalities]
    nationality_data = [n['count'] for n in nationalities]
    maker_labels = [m['vehicle_maker'] for m in makers]
    maker_data = [m['count'] for m in makers]
    asset_type_labels = [a['asset_type'] for a in asset_types]
    asset_type_data = [a['count'] for a in asset_types]

    context = {
        'total_employees': total_employees,
        'active_employees': active_employees,
        'inactive_employees': total_employees - active_employees,
        'contract_breakdown': contract_breakdown,
        'nationalities': nationalities,
        'deployments': deployments,
        'total_assets': total_assets,
        'assets_in_stock': assets_in_stock,
        'assets_assigned': assets_assigned,
        'total_asset_value': total_asset_value,
        'total_vehicles': total_vehicles,
        'valid_vehicles': valid_vehicles,
        'compliance_issues': compliance_issues,
        'makers': makers,
        'assignments': assignments,
        'unassigned_employees': unassigned_employees[:20],
        'unassigned_assets': unassigned_assets,
        'unassigned_vehicles': unassigned_vehicles,
        'recent_employees': recent_employees,
        'asset_types': asset_types,
        # Chart.js JSON
        'nationality_labels': json.dumps(nationality_labels),
        'nationality_data': json.dumps(nationality_data),
        'maker_labels': json.dumps(maker_labels),
        'maker_data': json.dumps(maker_data),
        'asset_type_labels': json.dumps(asset_type_labels),
        'asset_type_data': json.dumps(asset_type_data),
        'contract_data': json.dumps([contract_breakdown.get('permanent', 0), contract_breakdown.get('yearly', 0), contract_breakdown.get('ajeer', 0)]),
    }
    return render(request, 'hr/dashboard.html', context)


def _edit_leave_retry_redirect(request, base_url_name, target, edit_form, extra_params=None, error_text=None):
    """A failed leave-request edit (POST-redirect-GET) used to redirect
    with only a generic 'check the form' flash message, discarding both
    the real per-field reason AND whatever the employee had just typed —
    the reopened panel silently reverted to the pre-edit values, which
    read as the edit having been thrown away. This carries the actual
    field errors and the attempted values back as query params so the
    calling view's context-building can restore them onto the matching
    row (see my_profile/LeaveRequestListView) instead of losing them.

    Pass ``error_text`` to surface a service-level failure (a ValueError from
    edit_leave_request that the form itself didn't catch — balance/overlap/
    concurrent decision) with the same value-preserving behaviour. The message
    is shown inline on the row only, not also as a top-of-page flash, to avoid
    the user reading the same error twice."""
    from urllib.parse import urlencode
    if error_text is None:
        error_text = ' '.join(str(e) for errs in edit_form.errors.values() for e in errs)
    params = {
        'opened_leave': target.pk,
        'edit_error': error_text,
        'retry_leave_type': request.POST.get('leave_type', ''),
        'retry_start_date': request.POST.get('start_date', ''),
        'retry_end_date': request.POST.get('end_date', ''),
        'retry_reason': request.POST.get('employee_reason', ''),
    }
    if extra_params:
        params.update(extra_params)
    return redirect(f"{reverse(base_url_name)}?{urlencode(params)}")


def _apply_leave_edit_retry(request, r):
    """Read side of _edit_leave_retry_redirect: restore a failed edit's attempted
    values + inline error onto row `r`. Defaults to the row's stored values, and
    overrides them from the retry_* query params only for the row whose edit
    actually failed (matched by opened_leave). Shared by my_profile and
    LeaveRequestListView so the two copies can't drift."""
    from django.utils.dateparse import parse_date
    r.retry_leave_type_id = r.leave_type_id
    r.retry_start_date = r.start_date
    r.retry_end_date = r.end_date
    r.retry_reason = r.employee_reason
    r.edit_error = None
    if request.GET.get('edit_error') and request.GET.get('opened_leave') == str(r.pk):
        r.edit_error = request.GET.get('edit_error')
        retry_lt = request.GET.get('retry_leave_type')
        if retry_lt and retry_lt.isdigit():
            r.retry_leave_type_id = int(retry_lt)
        r.retry_start_date = parse_date(request.GET.get('retry_start_date') or '') or r.start_date
        r.retry_end_date = parse_date(request.GET.get('retry_end_date') or '') or r.end_date
        r.retry_reason = request.GET.get('retry_reason', r.employee_reason)


@login_required
def my_profile(request):
    """Self-service portal: the logged-in user's own HR record — attendance,
    leave balance, assets held, documents (iqama/passport), and any vehicle.
    Shows a friendly prompt if the account isn't linked to an employee yet."""
    from decimal import Decimal
    from django.db.models import Q
    from django.utils import timezone
    from .models import AttendanceRecord, Asset, Vehicle, AttendanceException
    from .forms import LeaveRequestForm

    emp = getattr(request.user, 'employee_profile', None)
    context = {'employee': emp}

    is_leave_request_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'request_leave')
    if is_leave_request_post:
        form = LeaveRequestForm(request.POST, request.FILES, fixed_employee=emp)
        if form.is_valid():
            from hr.leave_approval_services import submit_leave_request
            try:
                submit_leave_request(
                    employee=emp, leave_type=form.cleaned_data['leave_type'],
                    start_date=form.cleaned_data['start_date'], end_date=form.cleaned_data['end_date'],
                    employee_reason=form.cleaned_data['employee_reason'], document=form.cleaned_data['document'],
                    created_by=request.user,
                )
                messages.success(request, 'Leave request submitted and sent for approval.')
                return redirect('hr:my_profile')
            except ValueError as exc:
                # The form already passed its own (unlocked) balance/overlap
                # check — reaching here means submit_leave_request's locked,
                # authoritative re-check caught something that changed in the
                # meantime (most likely a genuinely concurrent submission).
                messages.error(request, str(exc))
                context['leave_request_submit_failed'] = True
    else:
        form = LeaveRequestForm(fixed_employee=emp) if emp else LeaveRequestForm()
    context['leave_request_form'] = form

    is_edit_leave_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'edit_leave_request')
    if is_edit_leave_post:
        from hr.leave_approval_services import edit_leave_request
        target = LeaveRequest.objects.filter(pk=request.POST.get('request_id')).first()
        if target is None or target.created_by_id != request.user.id:
            return HttpResponse('You did not submit this request.', status=403)
        edit_form = LeaveRequestForm(
            request.POST, request.FILES, fixed_employee=target.employee, exclude_request_id=target.pk,
            allow_leave_type=target.leave_type, initial={'document': target.document})
        if edit_form.is_valid():
            try:
                edit_leave_request(
                    target, request.user, leave_type=edit_form.cleaned_data['leave_type'],
                    start_date=edit_form.cleaned_data['start_date'], end_date=edit_form.cleaned_data['end_date'],
                    employee_reason=edit_form.cleaned_data['employee_reason'],
                    document=edit_form.cleaned_data['document'])
                messages.success(request, 'Leave request updated.')
            except ValueError as exc:
                # Service-level failure (balance/overlap/concurrent decision) —
                # preserve the attempted values and surface the real reason
                # inline, same as a form-invalid edit, instead of reverting.
                return _edit_leave_retry_redirect(
                    request, 'hr:my_profile', target, edit_form, error_text=str(exc))
        else:
            return _edit_leave_retry_redirect(request, 'hr:my_profile', target, edit_form)
        return redirect(f"{reverse('hr:my_profile')}?opened_leave={target.pk}")

    is_cancel_leave_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'cancel_leave_request')
    if is_cancel_leave_post:
        from hr.leave_approval_services import cancel_leave_request
        target = LeaveRequest.objects.filter(pk=request.POST.get('request_id')).first()
        if target is None or target.created_by_id != request.user.id:
            return HttpResponse('You did not submit this request.', status=403)
        try:
            cancel_leave_request(target, request.user, request.POST.get('reason', ''))
            messages.success(request, 'Leave request cancelled.')
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect('hr:my_profile')

    is_request_revoke_leave_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'request_leave_revoke')
    if is_request_revoke_leave_post:
        from hr.leave_approval_services import request_leave_revoke
        target = LeaveRequest.objects.filter(pk=request.POST.get('request_id')).first()
        if target is None or not (target.employee.user_id and target.employee.user_id == request.user.id):
            return HttpResponse('This is not your leave request.', status=403)
        try:
            request_leave_revoke(target, request.user, request.POST.get('reason', ''))
            messages.success(request, 'Revoke requested — it will not take effect until reviewed.')
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect('hr:my_profile')

    is_attendance_exception_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'submit_attendance_exception')
    if is_attendance_exception_post:
        exception_form = AttendanceExceptionForm(request.POST)
        if exception_form.is_valid():
            from hr.attendance_exception_services import submit_attendance_exception
            try:
                submit_attendance_exception(
                    employee=emp, event_date=exception_form.cleaned_data['event_date'],
                    event_start_time=exception_form.cleaned_data['event_start_time'],
                    reason_category=exception_form.cleaned_data['reason_category'],
                    custom_reason=exception_form.cleaned_data['custom_reason'],
                    employee_comment=exception_form.cleaned_data['employee_comment'],
                    created_by=request.user,
                )
                messages.success(request, 'Attendance exception submitted and sent to your manager.')
            except ValueError as exc:
                messages.error(request, str(exc))
            return redirect('hr:my_profile')
    else:
        exception_form = AttendanceExceptionForm()
    context['attendance_exception_form'] = exception_form

    is_edit_exception_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'edit_attendance_exception')
    if is_edit_exception_post:
        from hr.attendance_exception_services import edit_attendance_exception
        target = AttendanceException.objects.filter(pk=request.POST.get('request_id')).first()
        if target is None or target.created_by_id != request.user.id:
            return HttpResponse('You did not submit this request.', status=403)
        edit_exc_form = AttendanceExceptionForm(request.POST)
        if edit_exc_form.is_valid():
            try:
                edit_attendance_exception(
                    target, request.user, event_date=edit_exc_form.cleaned_data['event_date'],
                    event_start_time=edit_exc_form.cleaned_data['event_start_time'],
                    reason_category=edit_exc_form.cleaned_data['reason_category'],
                    custom_reason=edit_exc_form.cleaned_data['custom_reason'],
                    employee_comment=edit_exc_form.cleaned_data['employee_comment'])
                messages.success(request, 'Attendance exception updated.')
            except ValueError as exc:
                messages.error(request, str(exc))
        else:
            messages.error(request, 'Could not update the exception — please check the form.')
        return redirect('hr:my_profile')

    is_delete_exception_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'delete_attendance_exception')
    if is_delete_exception_post:
        from hr.attendance_exception_services import delete_attendance_exception
        target = AttendanceException.objects.filter(pk=request.POST.get('request_id')).first()
        if target is None or target.created_by_id != request.user.id:
            return HttpResponse('You did not submit this request.', status=403)
        try:
            delete_attendance_exception(target, request.user)
            messages.success(request, 'Attendance exception deleted.')
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect('hr:my_profile')

    is_request_revoke_exception_post = (
        emp and request.method == 'POST' and request.POST.get('action') == 'request_attendance_exception_revoke')
    if is_request_revoke_exception_post:
        from hr.attendance_exception_services import request_attendance_exception_revoke
        target = AttendanceException.objects.filter(pk=request.POST.get('request_id')).first()
        if target is None or not (target.employee.user_id and target.employee.user_id == request.user.id):
            return HttpResponse('This is not your attendance exception.', status=403)
        try:
            request_attendance_exception_revoke(target, request.user, request.POST.get('reason', ''))
            messages.success(request, 'Revoke requested — it will not take effect until reviewed.')
        except ValueError as exc:
            messages.error(request, str(exc))
        return redirect('hr:my_profile')

    if emp:
        today = timezone.localtime(timezone.now()).date()
        # Assets in custody - AssetHandover is the source of truth for
        # current custody (see also hr_dashboard and asset_detail.html).
        context['assets'] = Asset.objects.filter(
            handovers__employee=emp, handovers__status='active').distinct().order_by('asset_name')
        # Documents (iqama/passport copies, contracts, etc.).
        context['documents'] = emp.documents.all()
        # Leave balance for the current year. The summary total only counts
        # standard accrued allowances (Annual) — conditional/incidental
        # leave (Sick, Marriage, Umrah, etc.) still appears in the per-type rows but
        # is excluded from the aggregate so it doesn't inflate it.
        entitlements = (
            emp.leave_entitlements.filter(year=today.year)
            .select_related('leave_type'))
        context['entitlements'] = entitlements
        accumulative = [e for e in entitlements if e.leave_type.is_accumulative]
        context['leave_total_entitled'] = sum((e.entitled_days for e in accumulative), Decimal('0'))
        context['leave_total_remaining'] = sum((e.remaining_days for e in accumulative), Decimal('0'))
        context['leave_total_exception'] = sum((e.exception_days for e in accumulative), Decimal('0'))
        # This employee's own leave requests (pending/approved/disapproved), newest first.
        context['leave_requests'] = list(
            emp.leave_requests.select_related('leave_type', 'overridden_by')
            .prefetch_related('approvals__approver', 'notes__author')
            .order_by('-created_at')[:10])
        # Include any (now-inactive) leave type still attached to a shown request
        # so the edit dropdown never silently switches it to another type.
        _used_lt_ids = {r.leave_type_id for r in context['leave_requests']}
        context['active_leave_types'] = LeaveType.objects.filter(
            Q(is_active=True) | Q(pk__in=_used_lt_ids)).order_by('name')
        for r in context['leave_requests']:
            # Only employee-facing notes (is_internal=False) — internal HR
            # notes must never leak here. Filtered in Python against the
            # prefetch cache rather than a second queryset, so this doesn't
            # cost an extra query per row.
            r.visible_notes = [n for n in r.notes.all() if not n.is_internal]
            # Edit spans the whole pending window, same as Cancel below —
            # editing after a decision resets that decision (see
            # edit_leave_request), so decided_approvers drives the "this
            # will reset X's decision" warning in the edit form.
            r.can_edit = bool(r.created_by_id == request.user.id and r.status == 'pending')
            r.decided_approvers = [
                a.approver.get_full_name() or a.approver.username
                for a in r.approvals.all() if a.decision != 'pending']
            # Restore a failed edit's attempted values + inline error (shared
            # with LeaveRequestListView so the logic can't drift).
            _apply_leave_edit_retry(request, r)
            # Cancel spans the whole pending window — before any decision
            # this is a same-day change of mind (no reason required); once
            # an approver has decided, it still works but requires a
            # reason and notifies them. Either way it's never a hard
            # delete: the row stays visible as 'Cancelled' in History.
            r.can_cancel = bool(r.created_by_id == request.user.id and r.status == 'pending')
            r.can_request_revoke = bool(
                r.employee.user_id == request.user.id and r.status == 'approved'
                and not r.revoke_requests.filter(status='pending').exists())
            r.pending_revoke_request = r.revoke_requests.filter(status='pending').first()
        # This employee's own attendance exceptions, newest event first.
        context['my_attendance_exceptions'] = list(emp.attendance_exceptions.order_by('-event_date')[:10])
        for e in context['my_attendance_exceptions']:
            e.can_edit_delete = bool(e.created_by_id == request.user.id and e.status == 'pending')
            e.can_request_revoke = bool(
                e.employee.user_id == request.user.id and e.status == 'approved'
                and not e.revoke_requests.filter(status='pending').exists())
            e.pending_revoke_request = e.revoke_requests.filter(status='pending').first()
        # Attendance summary for the current month.
        # Attendance for the selected month (defaults to current month; user can
        # navigate Prev/Next via ?date=YYYY-MM-DD in the query string).
        from hr.attendance_matrix import period_range
        attendance_anchor = _parse_date(request.GET.get('date'))
        month_start, month_end = period_range('month', attendance_anchor)
        month_records = AttendanceRecord.objects.filter(
            employee=emp, date__gte=month_start, date__lte=month_end)
        records_by_date = {r.date: r for r in month_records}
        summary = {}
        for rec in month_records:
            summary[rec.status] = summary.get(rec.status, 0) + 1
        context['attendance_summary'] = summary
        context['attendance_month'] = month_start
        # The same totals the Attendance Register ranks people on, for this
        # employee's own month. Built by the same function, so what someone
        # sees about themselves here and what HR sees about them there can
        # never be two different numbers.
        from hr.attendance_summary import build_attendance_summary
        context['my_attendance_totals'] = build_attendance_summary(
            [emp], month_start, month_end)[emp.pk]
        context['late_warning_threshold'] = _late_warning_threshold()
        # Day-by-day list for the full month view + PDF export.
        from hr.models import AttendanceSettings, WorkingDay
        weekend_days = AttendanceSettings.load().weekend_day_set()
        working_days = set(WorkingDay.objects.filter(
            is_active=True, date__range=(month_start, month_end)).values_list('date', flat=True))
        exceptions_by_date = {
            e.event_date: e for e in AttendanceException.objects.filter(
                employee=emp, status='approved',
                event_date__gte=month_start, event_date__lte=month_end)
        }
        daily_attendance = []
        from hr.models import LateQuery
        queries_by_record = {
            q.attendance_record_id: q for q in LateQuery.objects.filter(
                employee=emp, attendance_record__in=month_records).order_by('created_at')
        }
        d = month_start
        while d <= month_end:
            rec = records_by_date.get(d)
            is_weekend = d.weekday() in weekend_days and d not in working_days
            existing_query = queries_by_record.get(rec.pk) if rec else None
            can_raise_query = bool(
                rec and rec.status == 'late'
                and (existing_query is None or existing_query.status == 'rejected'))
            exc = exceptions_by_date.get(d)
            exception_reason = None
            if exc:
                exception_reason = exc.custom_reason.strip() or exc.get_reason_category_display()
            daily_attendance.append({
                'date': d,
                'status': rec.status if rec else '',
                'record': rec,
                'is_weekend': is_weekend,
                'check_in': rec.check_in if rec else None,
                'check_out': rec.check_out if rec else None,
                'late_query': existing_query,
                'can_raise_query': can_raise_query,
                'exception_reason': exception_reason,
            })
            d += timedelta(days=1)
        context['daily_attendance'] = daily_attendance
        context['attendance_prev_month'] = (month_start - timedelta(days=1)).replace(day=1)
        if emp.user_id:
            from notifications.models import Notification
            context['my_late_emails'] = Notification.objects.filter(
                recipient=emp.user, verb='You were late 3 times this month').order_by('-created_at')[:20]
        context['my_late_queries'] = LateQuery.objects.filter(
            employee=emp).select_related('attendance_record').order_by('-created_at')[:20]
        next_month_first = (month_end + timedelta(days=1))
        context['attendance_next_month'] = next_month_first
        context['recent_leaves'] = emp.leave_records.select_related(
            'leave_type').order_by('-start_date')[:5]
        # Direct Reports list — each report gets EVERY in-behalf (manager-
        # logged) request attached as a plain in-memory list, newest first,
        # batched in one query rather than N+1. Scoped to created_by=this
        # viewer AND logged_by_manager=True: a manager sees the status of
        # what THEY logged, never a report's own self-submitted requests —
        # that stays private to the employee and HR's approval queue.
        # Deliberately a list, not just the most recent one — a manager who
        # logs two separate leave periods for the same report must see the
        # status of both, not have the older one silently disappear.
        direct_reports = list(emp.main_reports.all())
        if direct_reports:
            from collections import defaultdict
            in_behalf_by_employee = defaultdict(list)
            for req in (LeaveRequest.objects.filter(
                            employee_id__in=[r.pk for r in direct_reports],
                            created_by=request.user, logged_by_manager=True)
                        .select_related('leave_type').order_by('employee_id', '-created_at')):
                # created_by == request.user is already guaranteed by the
                # queryset filter above, so only the status check is needed
                # here (unlike the top-level leave_requests list, which
                # mixes self-submitted and manager-logged rows and needs
                # the full creator check). Cancel spans the whole pending
                # window (see the top-level loop above for why).
                req.can_cancel = bool(req.status == 'pending')
                in_behalf_by_employee[req.employee_id].append(req)
            for r in direct_reports:
                r.in_behalf_requests = in_behalf_by_employee.get(r.pk, [])
        context['direct_reports'] = direct_reports

        # Reporting Structure "expand" control: the further-down reporting
        # chain (grandreports and beyond) that direct reports alone don't
        # show. Excludes direct reports themselves so expanding never
        # duplicates what's already listed above it.
        direct_report_ids = [r.pk for r in direct_reports]
        context['downstream_reports'] = Employee.objects.filter(
            pk__in=emp.get_downstream_employee_ids()).exclude(pk__in=direct_report_ids)
        # Vehicles - AssetHandover is the source of truth for current
        # custody (see also hr_dashboard and vehicle_detail.html).
        context['vehicles'] = Vehicle.objects.filter(
            handovers__employee=emp, handovers__status='active').distinct()
        from hr.models import AssetHandover
        context['asset_handovers'] = AssetHandover.objects.filter(
            employee=emp).select_related('asset', 'vehicle').order_by('-created_at')[:20]
    return render(request, 'hr/my_profile.html', context)


@login_required
def build_attendance_pdf(emp, month_start, month_end):
    # Builds the same monthly attendance PDF used by my_attendance_export_pdf
    # (and reused for the automatic late-threshold email), returned as a
    # BytesIO buffer rather than an HttpResponse, so callers decide what to
    # do with it (download, or attach to an email).
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from hr.models import AttendanceRecord, AttendanceSettings, WorkingDay
    import io

    month_records = AttendanceRecord.objects.filter(
        employee=emp, date__gte=month_start, date__lte=month_end)
    records_by_date = {r.date: r for r in month_records}
    weekend_days = AttendanceSettings.load().weekend_day_set()
    working_days = set(WorkingDay.objects.filter(
        is_active=True, date__range=(month_start, month_end)).values_list('date', flat=True))
    from hr.models import AttendanceException
    exceptions_by_date = {
        e.event_date: e for e in AttendanceException.objects.filter(
            employee=emp, status='approved',
            event_date__gte=month_start, event_date__lte=month_end)
    }
    COLOR_PRESENT = colors.HexColor('#C6E0B4')
    COLOR_ABSENT = colors.HexColor('#F4B6C2')
    COLOR_WFH = colors.HexColor('#BDD7EE')
    COLOR_WEEKEND = colors.HexColor('#D9D2E9')

    def color_for(status, is_weekend):
        if status == 'present':
            return COLOR_PRESENT
        if status == 'absent':
            return COLOR_ABSENT
        if status == 'wfh':
            return COLOR_WFH
        if is_weekend:
            return COLOR_WEEKEND
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=18)
    styles = getSampleStyleSheet()
    elements = [Paragraph(f'{emp.full_name} - Attendance ({month_start.strftime("%B %Y")})', styles['Title'])]
    reason_style = ParagraphStyle('reason', parent=styles['Normal'], fontSize=8, leading=10)

    # The same totals My Profile and the Attendance Register show, from the
    # same function. This PDF is also what the automatic late-threshold email
    # attaches, so it is often the first place someone reads their own figures
    # - it must not be the one place they are computed differently.
    from hr.attendance_summary import build_attendance_summary
    from hr.models import LATE_WARNING_THRESHOLD
    totals = build_attendance_summary([emp], month_start, month_end)[emp.pk]
    rate = f"{totals['rate']}%" if totals['rate'] is not None else '-'
    summary_style = ParagraphStyle('sum', parent=styles['Normal'], fontSize=8,
                                   leading=10, textColor=colors.HexColor('#555555'))
    summary = Table(
        [['On time', 'Late', 'Absent', 'On-time rate', 'Warnings'],
         [str(totals['on_time']), str(totals['late']), str(totals['absent']),
          rate, str(totals['warnings'])]],
        colWidths=[89, 89, 89, 89, 89])
    summary.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(summary)
    elements.append(Paragraph(
        'The on-time rate counts only days an arrival was assessed - leave, '
        'holidays, weekends and working from home carry no check-in time. '
        f'A warning is {LATE_WARNING_THRESHOLD} lates in one calendar month. '
        'If a day below is marked Late in error, raise a query from My Profile.',
        summary_style))
    elements.append(Paragraph('&nbsp;', summary_style))

    data = [['Date', 'Day', 'Status', 'Check In', 'Check Out', 'Reason']]
    bg_commands = []
    d = month_start
    row_idx = 1
    while d <= month_end:
        rec = records_by_date.get(d)
        status = rec.status if rec else ''
        is_weekend = d.weekday() in weekend_days and d not in working_days
        label = status.title() if status else ('Weekend' if is_weekend else '-')
        check_in = rec.check_in.strftime('%H:%M') if rec and rec.check_in else ''
        check_out = rec.check_out.strftime('%H:%M') if rec and rec.check_out else ''
        exc = exceptions_by_date.get(d)
        reason = ''
        if exc:
            from xml.sax.saxutils import escape as _xml_escape
            raw_reason = exc.custom_reason.strip() or exc.get_reason_category_display()
            reason = Paragraph(_xml_escape(raw_reason), reason_style) if raw_reason else ''
        data.append([d.strftime('%d %b %Y'), d.strftime('%A'), label, check_in, check_out, reason])
        color = color_for(status, is_weekend)
        if color is not None:
            bg_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), color))
        row_idx += 1
        d += timedelta(days=1)

    table = Table(data, colWidths=[85, 75, 65, 55, 55, 110], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ] + bg_commands))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer


@login_required
@require_POST
def raise_late_query(request):
    from hr.models import AttendanceRecord, LateQuery
    emp = getattr(request.user, 'employee_profile', None)
    if emp is None:
        messages.error(request, 'Your account is not linked to an employee record.')
        return redirect('hr:my_profile')
    record_id = request.POST.get('attendance_record_id')
    message_text = (request.POST.get('message') or '').strip()
    if not record_id or not str(record_id).isdigit():
        messages.error(request, 'That attendance record could not be found.')
        return redirect('hr:my_profile')
    rec = AttendanceRecord.objects.filter(pk=record_id, employee=emp, status='late').first()
    if rec is None:
        messages.error(request, 'That attendance record could not be found.')
        return redirect('hr:my_profile')
    redirect_url = reverse('hr:my_profile') + '?date=' + rec.date.replace(day=1).strftime('%Y-%m-%d')
    if not message_text:
        messages.error(request, 'Please explain why you believe this Late mark is incorrect.')
        return redirect(redirect_url)
    if LateQuery.objects.filter(attendance_record=rec).exclude(status='rejected').exists():
        messages.error(request, 'A query has already been raised for this day.')
        return redirect(redirect_url)
    LateQuery.objects.create(employee=emp, attendance_record=rec, message=message_text)
    from notifications.services import notify_users
    from accounts.models import User
    admins = User.objects.filter(is_active=True)
    hr_admins = [u for u in admins if (u.is_super_admin_user or u.is_admin_user or u.is_erp_admin_user) and u.pk != emp.user_id]
    notify_users(
        recipients=hr_admins,
        verb=emp.full_name + ' raised a query about a Late mark on ' + rec.date.strftime('%d %b %Y'),
        description=message_text, level='info',
        target_url=reverse('hr:team_exceptions') + '?tab=late_queries',
    )
    messages.success(request, 'Your query has been submitted to HR.')
    return redirect(redirect_url)


@login_required
def my_attendance_export_pdf(request):

    emp = getattr(request.user, 'employee_profile', None)
    if emp is None:
        messages.error(request, 'Your account is not linked to an employee record.')
        return redirect('hr:my_profile')

    anchor = _parse_date(request.GET.get('date'))
    month_start, month_end = period_range('month', anchor)
    buffer = build_attendance_pdf(emp, month_start, month_end)
    filename = f'my_attendance_{month_start.strftime("%B_%Y").lower()}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

class EmployeeListView(AdminRequiredMixin, ListView):
    model = Employee
    template_name = 'hr/employee_list.html'
    context_object_name = 'employees'
    paginate_by = 25

    def get_queryset(self):
        queryset = Employee.objects.all()

        search = self.request.GET.get('search', '')
        contract_type = self.request.GET.get('contract_type', '')
        nationality = self.request.GET.get('nationality', '')
        deployment = self.request.GET.get('deployment', '')
        work_location = self.request.GET.get('work_location', '')
        status = self.request.GET.get('status', '')

        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search) |
                Q(iqama_number__icontains=search) |
                Q(work_email__icontains=search) |
                Q(mobile_number__icontains=search)
            )
        if contract_type:
            queryset = queryset.filter(contract_type=contract_type)
        if nationality:
            queryset = queryset.filter(nationality__icontains=nationality)
        if deployment:
            queryset = queryset.filter(deployment__icontains=deployment)
        if work_location:
            queryset = queryset.filter(work_location=work_location)
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)

        return queryset.order_by('full_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = EmployeeFilterForm(self.request.GET)
        context['total_count'] = Employee.objects.count()
        context['active_count'] = Employee.objects.filter(is_active=True).count()

        # Counts by contract type
        contract_counts = Employee.objects.values('contract_type').annotate(
            count=Count('id')
        ).order_by('contract_type')
        context['contract_counts'] = {
            item['contract_type']: item['count'] for item in contract_counts
        }

        # Counts by nationality (top nationalities)
        nationality_counts = Employee.objects.exclude(nationality='').values('nationality').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        context['nationality_counts'] = nationality_counts

        return context


class EmployeeDetailView(AdminRequiredMixin, DetailView):
    model = Employee
    template_name = 'hr/employee_detail.html'
    context_object_name = 'employee'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['documents'] = self.object.documents.all()
        context['doc_form'] = EmployeeDocumentForm()
        return context


class EmployeeCreateView(AdminRequiredMixin, CreateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'hr/employee_form.html'
    success_url = reverse_lazy('hr:employee_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        response = super().form_valid(form)
        # Auto-generate this year's leave balances for every newly added
        # employee — the same rows an admin would otherwise have to create
        # by hand via the Entitlements page's "Generate" action.
        if self.object.is_active:
            from hr.leave_services import generate_entitlements_for_employee
            generate_entitlements_for_employee(self.object, timezone.now().year, actor=self.request.user)
        messages.success(self.request, 'Employee created successfully.')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Employee'
        context['button_text'] = 'Save Employee'
        return context


class EmployeeUpdateView(AdminRequiredMixin, UpdateView):
    model = Employee
    form_class = EmployeeForm
    template_name = 'hr/employee_form.html'
    success_url = reverse_lazy('hr:employee_list')

    def form_valid(self, form):
        old_location = Employee.objects.filter(pk=self.object.pk).values_list('work_location', flat=True).first()
        response = super().form_valid(form)
        new_location = self.object.work_location
        if old_location != new_location:
            from hr.leave_services import apply_work_location_transfer
            apply_work_location_transfer(
                self.object, old_location=old_location, new_location=new_location,
                year=timezone.now().year, actor=self.request.user)
        messages.success(self.request, 'Employee updated successfully.')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Employee'
        context['button_text'] = 'Update Employee'
        return context


class EmployeeGrantExceptionDaysView(SuperAdminRequiredMixin, FormView):
    """HR action: adjust an employee's balance for one leave type/year —
    creates one audited LeaveExceptionGrant row (positive or negative,
    derived from the new-total-vs-current diff; see ExceptionGrantForm).
    Gated by the same override-access permission as the balance-override
    escape hatch on Leave Requests. Reachable from an employee's own Leave
    Summary page, or from the Entitlements page's per-leave-type breakdown
    (which pre-fills leave_type/year via query params, since that page
    already shows the exact row being adjusted)."""
    form_class = ExceptionGrantForm
    template_name = 'hr/exception_grant_form.html'

    def test_func(self):
        return has_override_access(self.request.user)

    def dispatch(self, request, *args, **kwargs):
        self.employee = get_object_or_404(Employee, pk=kwargs['pk'])
        return super().dispatch(request, *args, **kwargs)

    def _prefilled_leave_type_and_year(self):
        """Only trusted when BOTH arrive together from a per-row Adjust
        link — a lone leave_type or lone year isn't enough to safely lock
        the fields (the other one would default silently)."""
        leave_type_id = self.request.GET.get('leave_type')
        year = self.request.GET.get('year')
        if leave_type_id and leave_type_id.isdigit() and year and year.isdigit():
            leave_type = LeaveType.objects.filter(pk=leave_type_id).first()
            if leave_type:
                return leave_type, int(year)
        return None, None

    def _current_total(self, leave_type, year):
        entitlement = LeaveEntitlement.objects.filter(
            employee=self.employee, leave_type=leave_type, year=year).first()
        if entitlement:
            return entitlement.effective_entitled_days
        return leave_type.default_days_for(self.employee.work_location)

    def get_initial(self):
        initial = {'year': timezone.now().year}
        leave_type, year = self._prefilled_leave_type_and_year()
        if leave_type:
            initial['leave_type'] = leave_type.pk
            initial['year'] = year
            initial['new_total_days'] = self._current_total(leave_type, year)
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['employee'] = self.employee
        leave_type, year = self._prefilled_leave_type_and_year()
        ctx['locked_leave_type'] = leave_type
        ctx['locked_year'] = year
        if leave_type:
            ctx['current_total_days'] = self._current_total(leave_type, year)
        return ctx

    def form_valid(self, form):
        from hr.leave_approval_services import grant_exception_days
        leave_type = form.cleaned_data['leave_type']
        year = form.cleaned_data['year']
        new_total = form.cleaned_data['new_total_days']
        current_total = self._current_total(leave_type, year)
        delta = new_total - current_total
        if delta == 0:
            messages.info(
                self.request,
                f'{self.employee.full_name} already has {new_total} day(s) for {leave_type.name} in {year} '
                '— no change made.')
            return redirect('hr:leave_summary', pk=self.employee.pk)
        grant_exception_days(
            employee=self.employee, leave_type=leave_type, year=year,
            days=delta, granted_by=self.request.user, reason=form.cleaned_data['reason'])
        verb = 'Granted' if delta > 0 else 'Deducted'
        messages.success(
            self.request,
            f'{verb} {abs(delta)} day(s) — {self.employee.full_name} now has {new_total} total day(s) '
            f'for {leave_type.name} in {year}.')
        return redirect('hr:leave_summary', pk=self.employee.pk)


class EmployeeDeleteView(AdminRequiredMixin, DeleteView):
    model = Employee
    template_name = 'hr/employee_confirm_delete.html'
    success_url = reverse_lazy('hr:employee_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Employee deleted successfully.')
        return super().delete(request, *args, **kwargs)


@login_required
def employee_import(request):
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have permission to import employees.')
        return redirect('hr:employee_list')

    if request.method == 'POST':
        form = EmployeeImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active

                # Detect header row — look for "NAME" or "S/No." in first 5 rows
                header_row = None
                headers = []
                for row_num in range(1, 6):
                    row_values = [cell.value for cell in ws[row_num]]
                    for val in row_values:
                        if val and str(val).strip().upper() in ('NAME', 'S/NO.', 'S/NO'):
                            header_row = row_num
                            headers = row_values
                            break
                    if header_row:
                        break

                if not header_row:
                    # Fallback: use first row
                    header_row = 1
                    headers = [cell.value for cell in ws[1]]

                # Build header map (lowercase stripped header → column index)
                header_map = {}
                for idx, h in enumerate(headers):
                    if h:
                        header_map[str(h).strip().lower()] = idx

                imported_count = 0
                updated_count = 0

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if not any(row):
                        continue

                    def get_value(possible_headers):
                        for h in possible_headers:
                            if h in header_map:
                                idx = header_map[h]
                                if idx < len(row):
                                    return row[idx]
                        return None

                    name = get_value(['name', 'full name', 'employee name'])
                    if not name or not str(name).strip():
                        continue

                    iqama = get_value(['iqamah/passport', 'iqama/passport', 'iqamah', 'iqama', 'passport', 'iqama no', 'iqama no.'])
                    if not iqama or not str(iqama).strip():
                        continue

                    iqama_str = str(iqama).strip()

                    # Parse dates
                    def parse_date(val):
                        if val is None:
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        try:
                            from datetime import date
                            if isinstance(val, date):
                                return val
                        except Exception:
                            pass
                        return None

                    dob = parse_date(get_value(['date of birth', 'dob', 'birth date']))
                    joining = parse_date(get_value(['joining date', 'date of joining', 'join date']))

                    # Parse contract type
                    contract_raw = str(get_value(['contract type', 'contract']) or '').strip().lower()
                    contract_type = ''
                    if 'permanent' in contract_raw:
                        contract_type = 'permanent'
                    elif 'yearly' in contract_raw:
                        contract_type = 'yearly'
                    elif 'ajeer' in contract_raw:
                        contract_type = 'ajeer'

                    # Parse marital status
                    marital_raw = str(get_value(['marital status', 'marital']) or '').strip().lower()
                    marital_status = ''
                    if 'married' in marital_raw:
                        marital_status = 'married'
                    elif 'single' in marital_raw:
                        marital_status = 'single'

                    # Blood group — exact match
                    blood_raw = str(get_value(['blood group', 'blood']) or '').strip()
                    valid_blood = [c[0] for c in Employee.BLOOD_GROUP_CHOICES]
                    blood_group = blood_raw if blood_raw in valid_blood else ''

                    # Handle two email columns — personal (col ~11) and work (col ~15)
                    personal_email = str(get_value(['personal email', 'personal e-mail']) or '').strip()
                    work_email = str(get_value(['work email', 'work e-mail', 'official email']) or '').strip()

                    # If we have generic "email" headers, openpyxl may see them positionally
                    # The Excel has two email columns — try to pick them up
                    if not personal_email and not work_email:
                        # Look for any 'email' column
                        email_val = str(get_value(['email', 'e-mail']) or '').strip()
                        if email_val:
                            work_email = email_val

                    defaults = {
                        'full_name': str(name).strip()[:255],
                        'designation': str(get_value(['designation', 'position', 'title']) or '').strip()[:255],
                        'qualification': str(get_value(['qualification', 'qualifications']) or '').strip()[:255],
                        'date_of_birth': dob,
                        'joining_date': joining,
                        'nationality': str(get_value(['nationality']) or '').strip()[:100],
                        'marital_status': marital_status,
                        'blood_group': blood_group,
                        'personal_email': personal_email,
                        'documents_link': str(get_value(['documents', 'documents link', 'document link', 'doc link']) or '').strip()[:500],
                        'deployment': str(get_value(['deployment', 'location', 'site']) or '').strip()[:100],
                        'contract_type': contract_type,
                        'work_email': work_email,
                        'mobile_number': str(get_value(['mobile no', 'mobile number', 'mobile', 'phone', 'contact no']) or '').strip()[:20],
                        'created_by': request.user,
                    }

                    imported_emp, created = Employee.objects.update_or_create(
                        iqama_number=iqama_str,
                        defaults=defaults,
                    )

                    if created:
                        imported_count += 1
                        if imported_emp.is_active:
                            from hr.leave_services import generate_entitlements_for_employee
                            generate_entitlements_for_employee(
                                imported_emp, timezone.now().year, actor=request.user)
                    else:
                        updated_count += 1

                msg_parts = []
                if imported_count:
                    msg_parts.append(f'{imported_count} new employees imported')
                if updated_count:
                    msg_parts.append(f'{updated_count} employees updated')
                if msg_parts:
                    messages.success(request, 'Successfully ' + ' and '.join(msg_parts) + '.')
                else:
                    messages.warning(request, 'No employees were imported. Check file format.')

                return redirect('hr:employee_list')

            except Exception as e:
                messages.error(request, f'Error importing file: {str(e)}')
    else:
        form = EmployeeImportForm()

    return render(request, 'hr/employee_import.html', {'form': form})


@login_required
def employee_export(request):
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have permission to export employees.')
        return redirect('hr:employee_list')

    queryset = Employee.objects.all()

    # Apply filters
    search = request.GET.get('search', '')
    contract_type = request.GET.get('contract_type', '')
    nationality = request.GET.get('nationality', '')
    deployment = request.GET.get('deployment', '')

    if search:
        queryset = queryset.filter(
            Q(full_name__icontains=search) |
            Q(iqama_number__icontains=search) |
            Q(work_email__icontains=search)
        )
    if contract_type:
        queryset = queryset.filter(contract_type=contract_type)
    if nationality:
        queryset = queryset.filter(nationality__icontains=nationality)
    if deployment:
        queryset = queryset.filter(deployment__icontains=deployment)

    queryset = queryset.order_by('full_name')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employees'

    # Styles (matching Leap red header)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = [
        'S/No.', 'Iqama/Passport', 'Name', 'Designation', 'Qualification',
        'Date of Birth', 'Joining Date', 'Nationality', 'Marital Status',
        'Blood Group', 'Personal Email', 'Documents', 'Deployment',
        'Contract Type', 'Work Email', 'Mobile No', 'Status',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, emp in enumerate(queryset, 2):
        data = [
            row_num - 1,
            emp.iqama_number,
            emp.full_name,
            emp.designation,
            emp.qualification,
            emp.date_of_birth.strftime('%Y-%m-%d') if emp.date_of_birth else '',
            emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else '',
            emp.nationality,
            emp.get_marital_status_display(),
            emp.blood_group,
            emp.personal_email,
            emp.documents_link,
            emp.deployment,
            emp.get_contract_type_display(),
            emp.work_email,
            emp.mobile_number,
            'Active' if emp.is_active else 'Inactive',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

    # Column widths
    column_widths = [8, 18, 25, 20, 20, 14, 14, 15, 14, 12, 28, 30, 15, 14, 28, 16, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'employees_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── Asset Views ─────────────────────────────────────────────────────────────


class AssetListView(HRScopedAccessMixin, ListView):
    model = Asset
    template_name = 'hr/asset_list.html'
    context_object_name = 'assets'
    paginate_by = 25

    def _scoped_base(self):
        # Team-scoped roles see only assets currently held by their reports
        # (read-only); admin tiers see the whole inventory.
        return scope_asset_queryset(self.request.user, Asset.objects.all())

    def get_queryset(self):
        queryset = self._scoped_base()
        from hr.models import AssetHandover

        search = self.request.GET.get('search', '')
        asset_type = self.request.GET.get('asset_type', '')
        condition = self.request.GET.get('condition', '')
        in_stock = self.request.GET.get('in_stock', '')
        status = self.request.GET.get('status', '')

        if search:
            queryset = queryset.filter(
                Q(asset_name__icontains=search) |
                Q(serial_number__icontains=search) |
                Q(employee_name__icontains=search) |
                Q(invoice_number__icontains=search)
            )
        if asset_type:
            queryset = queryset.filter(asset_type__icontains=asset_type)
        if condition:
            queryset = queryset.filter(condition=condition)
        # "In stock" / "assigned" now reflect an active AssetHandover, not
        # the legacy in_stock field (see also hr_dashboard, my_profile,
        # asset_detail.html - AssetHandover is the sole source of truth
        # for current custody).
        if in_stock in ('true', 'false'):
            assigned_ids = AssetHandover.objects.filter(
                status='active', asset__isnull=False).values_list('asset_id', flat=True)
            if in_stock == 'true':
                queryset = queryset.exclude(pk__in=assigned_ids)
            else:
                queryset = queryset.filter(pk__in=assigned_ids)
        if status == 'decommissioned':
            queryset = queryset.filter(is_decommissioned=True)
        elif status == 'in_service':
            queryset = queryset.filter(is_decommissioned=False)

        return queryset.order_by('asset_name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = AssetFilterForm(self.request.GET)
        # Stat cards reflect the viewer's scope (all inventory for admin tiers;
        # only the team's held assets for scoped roles).
        base = self._scoped_base()
        context['total_count'] = base.count()
        from hr.models import AssetHandover
        assigned_ids = AssetHandover.objects.filter(
            status='active', asset__isnull=False).values_list('asset_id', flat=True)
        context['assigned_count'] = base.filter(pk__in=assigned_ids).count()
        context['in_stock_count'] = base.exclude(pk__in=assigned_ids).count()
        context['total_value'] = base.aggregate(total=Sum('price'))['total'] or 0

        # Counts by asset type (top types)
        type_counts = base.exclude(asset_type='').values('asset_type').annotate(
            count=Count('id')
        ).order_by('-count')[:5]
        context['type_counts'] = type_counts

        # Set of serial numbers (lower-cased) that appear on more than one
        # currently-assigned asset. The list template uses this to flag rows
        # at a glance instead of running an .exists() query per row.
        conflict_serials = (
            Asset.objects
            .exclude(serial_number='')
            .filter(assignments__returned_at__isnull=True)
            .values('serial_number')
            .annotate(c=Count('id', distinct=True))
            .filter(c__gt=1)
            .values_list('serial_number', flat=True)
        )
        context['conflict_serials'] = {s.lower() for s in conflict_serials}
        context['conflict_count'] = len(context['conflict_serials'])

        # Inventory operations (add/edit/issue/import/export/decommission) are
        # admin-only; team-scoped roles get a read-only view of their team's
        # assets. Drives which action buttons render.
        u = self.request.user
        context['can_manage_assets'] = bool(
            u.is_super_admin_user or u.is_admin_user or u.is_erp_admin_user)

        return context


class AssetDetailView(HRScopedAccessMixin, DetailView):
    model = Asset
    template_name = 'hr/asset_detail.html'
    context_object_name = 'asset'

    def get_queryset(self):
        # Team-scoped roles can only open an asset their team currently holds
        # (out-of-scope 404s); admin tiers see everything.
        return scope_asset_queryset(self.request.user, super().get_queryset())


def _asset_employees_context():
    """Active employees for the asset form's 'pick employee' dropdown, plus a
    map (pk -> name/designation) so selecting one auto-fills the text fields.
    Returned as a dict for json_script (which serialises it in the template)."""
    emps = list(Employee.objects.filter(is_active=True).order_by('full_name'))
    data = {
        str(e.pk): {'name': e.full_name, 'designation': e.designation or ''}
        for e in emps
    }
    return emps, data


class AssetCreateView(AdminRequiredMixin, CreateView):
    model = Asset
    form_class = AssetForm
    template_name = 'hr/asset_form.html'
    success_url = reverse_lazy('hr:asset_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Asset created successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add New Asset'
        context['button_text'] = 'Save Asset'
        context['employees'], context['employees_json'] = _asset_employees_context()
        return context


class AssetUpdateView(AdminRequiredMixin, UpdateView):
    model = Asset
    form_class = AssetForm
    template_name = 'hr/asset_form.html'
    success_url = reverse_lazy('hr:asset_list')

    def form_valid(self, form):
        messages.success(self.request, 'Asset updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Asset'
        context['button_text'] = 'Update Asset'
        context['employees'], context['employees_json'] = _asset_employees_context()
        return context


class AssetDeleteView(AdminRequiredMixin, DeleteView):
    model = Asset
    template_name = 'hr/asset_confirm_delete.html'
    success_url = reverse_lazy('hr:asset_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Asset deleted successfully.')
        return super().delete(request, *args, **kwargs)


@login_required
@require_POST
def asset_decommission(request, pk):
    """Mark an asset as out of service (dead / no longer usable). It can never
    be in stock once decommissioned."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:asset_list')
    from django.utils import timezone
    asset = get_object_or_404(Asset, pk=pk)
    asset.is_decommissioned = True
    asset.in_stock = False  # out of service can't be in stock
    if not asset.decommissioned_on:
        asset.decommissioned_on = timezone.localdate()
    reason = (request.POST.get('reason') or '').strip()
    if reason:
        asset.decommission_reason = reason
    asset.save(update_fields=['is_decommissioned', 'in_stock',
                              'decommissioned_on', 'decommission_reason', 'updated_at'])
    messages.success(request, f'"{asset.asset_name}" marked out of service.')
    return redirect(request.POST.get('next') or 'hr:asset_list')


@login_required
@require_POST
def asset_restore(request, pk):
    """Restore a decommissioned asset back into service."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:asset_list')
    asset = get_object_or_404(Asset, pk=pk)
    asset.is_decommissioned = False
    asset.decommissioned_on = None
    asset.decommission_reason = ''
    asset.save(update_fields=['is_decommissioned', 'decommissioned_on',
                              'decommission_reason', 'updated_at'])
    messages.success(request, f'"{asset.asset_name}" restored to service.')
    return redirect(request.POST.get('next') or 'hr:asset_list')


@login_required
def asset_import(request):
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have permission to import assets.')
        return redirect('hr:asset_list')

    if request.method == 'POST':
        form = AssetImportForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']

            try:
                wb = openpyxl.load_workbook(excel_file, data_only=True)
                ws = wb.active

                # Detect header row — look for "Asset Name" or "S. No" in first 5 rows
                header_row = None
                headers = []
                for row_num in range(1, 6):
                    row_values = [cell.value for cell in ws[row_num]]
                    for val in row_values:
                        if val and str(val).strip().upper() in (
                            'ASSET NAME', 'S. NO', 'S.NO', 'S. NO.', 'SERIAL NO.',
                            'SERIAL NO', 'ASSET TYPE',
                        ):
                            header_row = row_num
                            headers = row_values
                            break
                    if header_row:
                        break

                if not header_row:
                    header_row = 2
                    headers = [cell.value for cell in ws[2]]

                # Build header map
                header_map = {}
                for idx, h in enumerate(headers):
                    if h:
                        header_map[str(h).strip().lower()] = idx

                imported_count = 0
                updated_count = 0

                for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    if not any(row):
                        continue

                    def get_value(possible_headers):
                        for h in possible_headers:
                            if h in header_map:
                                idx = header_map[h]
                                if idx < len(row):
                                    return row[idx]
                        return None

                    asset_name = get_value(['asset name', 'asset'])
                    if not asset_name or not str(asset_name).strip():
                        continue

                    def parse_date(val):
                        if val is None:
                            return None
                        if isinstance(val, datetime):
                            return val.date()
                        if isinstance(val, date):
                            return val
                        return None

                    # Parse in_stock
                    in_stock_raw = get_value(['in stock?', 'in stock', 'stock'])
                    in_stock = False
                    if in_stock_raw is not None:
                        in_stock_str = str(in_stock_raw).strip().lower()
                        in_stock = in_stock_str in ('yes', 'true', '1', 'y', 'in stock')

                    # Parse condition
                    condition_raw = str(get_value(['condition']) or '').strip().lower()
                    condition = ''
                    if 'new' in condition_raw:
                        condition = 'new'
                    elif 'used' in condition_raw:
                        condition = 'used'

                    # Parse quantity
                    qty_raw = get_value(['qty', 'quantity'])
                    quantity = 1
                    if qty_raw is not None:
                        try:
                            quantity = int(float(str(qty_raw)))
                            if quantity < 1:
                                quantity = 1
                        except (ValueError, TypeError):
                            quantity = 1

                    # Parse price
                    price_raw = get_value(['price (sar)', 'price', 'price(sar)'])
                    price = None
                    if price_raw is not None:
                        try:
                            price = round(float(str(price_raw)), 2)
                        except (ValueError, TypeError):
                            price = None

                    serial_number = str(get_value(['serial no.', 'serial no', 'serial number', 'serial']) or '').strip()

                    # Return date is free text
                    return_date_raw = get_value(['return date', 'return'])
                    return_date = str(return_date_raw).strip() if return_date_raw else ''
                    # If it's a datetime object, format it
                    if isinstance(return_date_raw, (datetime, date)):
                        return_date = return_date_raw.strftime('%Y-%m-%d') if hasattr(return_date_raw, 'strftime') else str(return_date_raw)

                    defaults = {
                        'asset_type': str(get_value(['asset type', 'type']) or '').strip()[:50],
                        'specifications': str(get_value(['specifications', 'specs', 'specification']) or '').strip()[:255],
                        'invoice_number': str(get_value(['invoice number', 'invoice no', 'invoice no.', 'invoice']) or '').strip()[:100],
                        'employee_name': str(get_value(['employee name', 'employee', 'name']) or '').strip()[:255],
                        'department': str(get_value(['department', 'dept']) or '').strip()[:100],
                        'designation': str(get_value(['designation', 'position']) or '').strip()[:255],
                        'handover_date': parse_date(get_value(['handover date', 'handover'])),
                        'handover_by': str(get_value(['handover by', 'handed by']) or '').strip()[:255],
                        'condition': condition,
                        'return_date': return_date[:100],
                        'return_to': str(get_value(['return to', 'returned to']) or '').strip()[:255],
                        'quantity': quantity,
                        'purchase_date': parse_date(get_value(['purchase date', 'purchased date', 'purchase'])),
                        'price': price,
                        'planned_life': str(get_value(['planned asset life', 'planned life', 'asset life']) or '').strip()[:50],
                        'in_stock': in_stock,
                        'created_by': request.user,
                    }

                    _, created = Asset.objects.update_or_create(
                        asset_name=str(asset_name).strip()[:255],
                        serial_number=serial_number[:255],
                        defaults=defaults,
                    )

                    if created:
                        imported_count += 1
                    else:
                        updated_count += 1

                msg_parts = []
                if imported_count:
                    msg_parts.append(f'{imported_count} new assets imported')
                if updated_count:
                    msg_parts.append(f'{updated_count} assets updated')
                if msg_parts:
                    messages.success(request, 'Successfully ' + ' and '.join(msg_parts) + '.')
                else:
                    messages.warning(request, 'No assets were imported. Check file format.')

                return redirect('hr:asset_list')

            except Exception as e:
                messages.error(request, f'Error importing file: {str(e)}')
    else:
        form = AssetImportForm()

    return render(request, 'hr/asset_import.html', {'form': form})


@login_required
def asset_export(request):
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have permission to export assets.')
        return redirect('hr:asset_list')

    queryset = Asset.objects.all()

    # Apply filters
    search = request.GET.get('search', '')
    asset_type = request.GET.get('asset_type', '')
    condition = request.GET.get('condition', '')
    in_stock = request.GET.get('in_stock', '')

    if search:
        queryset = queryset.filter(
            Q(asset_name__icontains=search) |
            Q(serial_number__icontains=search) |
            Q(employee_name__icontains=search)
        )
    if asset_type:
        queryset = queryset.filter(asset_type__icontains=asset_type)
    if condition:
        queryset = queryset.filter(condition=condition)
    if in_stock == 'true':
        queryset = queryset.filter(in_stock=True)
    elif in_stock == 'false':
        queryset = queryset.filter(in_stock=False)

    queryset = queryset.order_by('asset_name')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Assets'

    # Styles (matching Leap red header)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = [
        'S. No', 'Asset Name', 'Asset Type', 'Serial No.', 'Specifications',
        'Invoice Number', 'Employee Name', 'Department', 'Designation',
        'Handover Date', 'Handover By', 'Condition', 'Return Date',
        'Return To', 'QTY', 'Purchase Date', 'Price (SAR)',
        'Planned Asset Life', 'Asset Current Age', 'In Stock?',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_num, asset in enumerate(queryset, 2):
        data = [
            row_num - 1,
            asset.asset_name,
            asset.asset_type,
            asset.serial_number,
            asset.specifications,
            asset.invoice_number,
            asset.employee_name,
            asset.department,
            asset.designation,
            asset.handover_date.strftime('%Y-%m-%d') if asset.handover_date else '',
            asset.handover_by,
            asset.get_condition_display() if asset.condition else '',
            asset.return_date,
            asset.return_to,
            asset.quantity,
            asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
            float(asset.price) if asset.price else '',
            asset.planned_life,
            asset.current_age,
            'Yes' if asset.in_stock else 'No',
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border

    # Column widths
    column_widths = [8, 22, 14, 22, 30, 16, 22, 16, 18, 14, 16, 12, 18, 16, 8, 14, 14, 18, 18, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'assets_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════
# EMPLOYEE DOCUMENTS
# ═══════════════════════════════════════════════════════════════

@login_required
def leave_request_document_download(request, pk):
    """Stream a LeaveRequest's uploaded document only to the employee it
    belongs to, a designated approver, or a super admin — never via a public
    media URL. This deliberately does NOT follow the EmployeeDocument/
    medical_certificate pattern (those link straight to MEDIA_URL); this field
    holds sensitive personal documents and needs a real permission check."""
    from django.http import FileResponse, Http404
    from .models import LeaveRequest
    # Look up without get_object_or_404 and check authorization before ever
    # confirming the object exists — otherwise a nonexistent pk (404) and an
    # existing-but-forbidden pk (403) are distinguishable, letting any
    # authenticated user enumerate valid LeaveRequest ids. Both "doesn't
    # exist" and "exists but you can't see it" collapse to the same 404.
    leave_request = LeaveRequest.objects.filter(pk=pk).first()
    user = request.user
    is_owner = bool(leave_request) and leave_request.employee.user_id == user.id
    authorized = leave_request is not None and (
        is_owner or is_designated_approver(user) or user.is_super_admin_user)
    if not authorized:
        raise Http404('No such leave request.')
    if not leave_request.document:
        raise Http404('No document attached to this request.')
    return FileResponse(leave_request.document.open('rb'), as_attachment=True,
                        filename=leave_request.document.name.rsplit('/', 1)[-1])


@login_required
def employee_document_upload(request, pk):
    """Upload a document for an employee."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:employee_list')

    employee = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST':
        form = EmployeeDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.employee = employee
            doc.uploaded_by = request.user
            doc.save()
            messages.success(request, f'Document "{doc.title}" uploaded.')
        else:
            messages.error(request, 'Please fix the errors below.')

    return redirect('hr:employee_detail', pk=pk)


@login_required
def employee_document_delete(request, pk):
    """Delete an employee document."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:employee_list')

    doc = get_object_or_404(EmployeeDocument, pk=pk)
    employee_pk = doc.employee_id
    if doc.file:
        doc.file.delete(save=False)
    doc.delete()
    messages.success(request, 'Document deleted.')
    return redirect('hr:employee_detail', pk=employee_pk)


# ═══════════════════════════════════════════════════════════════
# VEHICLES
# ═══════════════════════════════════════════════════════════════

class VehicleListView(AdminRequiredMixin, ListView):
    model = Vehicle
    template_name = 'hr/vehicle_list.html'
    context_object_name = 'vehicles'
    paginate_by = 25

    def get_queryset(self):
        queryset = Vehicle.objects.all()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        if search:
            queryset = queryset.filter(
                Q(plate_number__icontains=search) |
                Q(vehicle_maker__icontains=search) |
                Q(vehicle_model__icontains=search) |
                Q(driver_name__icontains=search) |
                Q(chassis_number__icontains=search)
            )
        if status:
            queryset = queryset.filter(vehicle_status=status)
        compliance = self.request.GET.get('compliance')
        if compliance == 'issues':
            pks = [v.pk for v in queryset if v.has_compliance_issue]
            queryset = queryset.filter(pk__in=pks)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = VehicleFilterForm(self.request.GET)
        context['total_count'] = self.get_queryset().count()
        all_vehicles = Vehicle.objects.all()
        context['valid_count'] = all_vehicles.filter(vehicle_status='valid').count()
        context['compliance_issues'] = sum(1 for v in all_vehicles if v.has_compliance_issue)
        return context


class VehicleCreateView(AdminRequiredMixin, CreateView):
    model = Vehicle
    form_class = VehicleForm
    template_name = 'hr/vehicle_form.html'
    success_url = reverse_lazy('hr:vehicle_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Vehicle'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Vehicle added.')
        return super().form_valid(form)


class VehicleDetailView(AdminRequiredMixin, DetailView):
    model = Vehicle
    template_name = 'hr/vehicle_detail.html'
    context_object_name = 'vehicle'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['documents'] = self.object.documents.all()
        context['doc_form'] = VehicleDocumentForm()
        return context


@login_required
def vehicle_document_upload(request, pk):
    """Upload a document for a vehicle."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:vehicle_list')

    vehicle = get_object_or_404(Vehicle, pk=pk)
    if request.method == 'POST':
        form = VehicleDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=False)
            doc.vehicle = vehicle
            doc.uploaded_by = request.user
            doc.save()
            messages.success(request, f'Document "{doc.title}" uploaded.')
        else:
            errs = '; '.join(f'{f}: {", ".join(e)}' for f, e in form.errors.items())
            messages.error(request, f'Could not upload document. {errs}')
    return redirect('hr:vehicle_detail', pk=pk)


@login_required
def vehicle_document_edit(request, pk):
    """Edit a vehicle document's details, optionally replacing the file.
    Replacing the file reclaims the old one via the central pre_save signal."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:vehicle_list')

    doc = get_object_or_404(VehicleDocument, pk=pk)
    if request.method == 'POST':
        form = VehicleDocumentForm(request.POST, request.FILES, instance=doc)
        if form.is_valid():
            form.save()
            messages.success(request, f'Document "{doc.title}" updated.')
            return redirect('hr:vehicle_detail', pk=doc.vehicle_id)
        messages.error(request, 'Please fix the errors below.')
    else:
        form = VehicleDocumentForm(instance=doc)
    return render(request, 'hr/vehicle_document_edit.html', {
        'form': form, 'doc': doc, 'vehicle': doc.vehicle})


@login_required
def vehicle_document_delete(request, pk):
    """Delete a vehicle document (its file is reclaimed by the cleanup signal)."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:vehicle_list')

    doc = get_object_or_404(VehicleDocument, pk=pk)
    vehicle_pk = doc.vehicle_id
    doc.delete()
    messages.success(request, 'Document deleted.')
    return redirect('hr:vehicle_detail', pk=vehicle_pk)


class VehicleUpdateView(AdminRequiredMixin, UpdateView):
    model = Vehicle
    form_class = VehicleForm
    template_name = 'hr/vehicle_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit: {self.object.plate_number}'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Vehicle updated.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('hr:vehicle_detail', kwargs={'pk': self.object.pk})


class VehicleDeleteView(AdminRequiredMixin, DeleteView):
    model = Vehicle
    template_name = 'hr/vehicle_confirm_delete.html'
    success_url = reverse_lazy('hr:vehicle_list')

    def form_valid(self, form):
        messages.success(self.request, 'Vehicle deleted.')
        return super().form_valid(form)


@login_required
def vehicle_import(request):
    """Import vehicles from MOI Excel export."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:vehicle_list')

    if request.method != 'POST':
        return redirect('hr:vehicle_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select a file.')
        return redirect('hr:vehicle_list')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        header_row = None
        for r in range(1, min(20, ws.max_row + 1)):
            val = str(ws.cell(row=r, column=1).value or '').strip()
            if 'plate' in val.lower():
                header_row = r
                break
        if not header_row:
            messages.error(request, 'Could not find header row.')
            return redirect('hr:vehicle_list')

        imported = 0
        for r in range(header_row + 1, ws.max_row + 1):
            plate = ws.cell(row=r, column=1).value
            if not plate:
                continue
            plate_str = str(plate).strip()
            if not plate_str:
                continue

            def cv(col):
                v = ws.cell(row=r, column=col).value
                return str(v).strip() if v is not None else ''

            mvpi_raw = cv(16).lower()
            mvpi = 'valid' if mvpi_raw == 'valid' else ('expired' if 'expir' in mvpi_raw else 'not_exist')
            ins_raw = cv(17).lower()
            ins = 'valid' if ins_raw == 'valid' else ('expired' if 'expir' in ins_raw else 'not_exist')
            rest_raw = cv(18).lower()
            rest = 'unrestricted' if 'unrestrict' in rest_raw else 'restricted'
            vstatus_raw = cv(10).lower()
            vstatus = 'valid' if vstatus_raw == 'valid' else 'expired'

            Vehicle.objects.update_or_create(
                plate_number=plate_str,
                defaults={
                    'plate_type': cv(2),
                    'branch_name': cv(3),
                    'vehicle_maker': cv(4),
                    'vehicle_model': cv(5),
                    'model_year': cv(6),
                    'sequence_number': cv(7),
                    'chassis_number': cv(8),
                    'major_color': cv(9),
                    'vehicle_status': vstatus,
                    'ownership_date': cv(11),
                    'license_expiry': cv(12),
                    'inspection_expiry': cv(13),
                    'driver_id': cv(14),
                    'driver_name': cv(15),
                    'mvpi_status': mvpi,
                    'insurance_status': ins,
                    'restriction_status': rest,
                    'license_issue_date': cv(19),
                    'body_type': cv(21),
                    'created_by': request.user,
                },
            )
            imported += 1

        messages.success(request, f'Imported {imported} vehicles.')
        return redirect('hr:vehicle_list')

    except Exception as e:
        messages.error(request, f'Import error: {str(e)}')
        return redirect('hr:vehicle_list')


@login_required
def vehicle_export(request):
    """Export vehicles to Excel."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:vehicle_list')

    vehicles = Vehicle.objects.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vehicles'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        'Plate Number', 'Plate Type', 'Branch', 'Maker', 'Model', 'Year',
        'Sequence No.', 'Chassis No.', 'Color', 'Status',
        'Ownership Date', 'License Expiry', 'Inspection Expiry',
        'Driver ID', 'Driver Name', 'MVPI', 'Insurance', 'Restriction',
        'License Issue Date', 'Body Type',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_num, v in enumerate(vehicles, 2):
        data = [
            v.plate_number, v.plate_type, v.branch_name, v.vehicle_maker,
            v.vehicle_model, v.model_year, v.sequence_number, v.chassis_number,
            v.major_color, v.get_vehicle_status_display(),
            v.ownership_date, v.license_expiry, v.inspection_expiry,
            v.driver_id, v.driver_name,
            v.get_mvpi_status_display(), v.get_insurance_status_display(),
            v.get_restriction_status_display(), v.license_issue_date, v.body_type,
        ]
        for col, val in enumerate(data, 1):
            ws.cell(row=row_num, column=col, value=val).border = thin_border

    widths = [16, 14, 10, 12, 16, 8, 14, 22, 10, 10, 14, 14, 14, 14, 30, 12, 12, 14, 14, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vehicles_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ─── Leave Type CRUD ──────────────────────────────────────────


class LeaveTypeListView(AdminRequiredMixin, ListView):
    model = LeaveType
    template_name = 'hr/leavetype_list.html'
    context_object_name = 'leave_types'
    paginate_by = 25


class LeaveTypeCreateView(AdminRequiredMixin, CreateView):
    model = LeaveType
    form_class = LeaveTypeForm
    template_name = 'hr/leavetype_form.html'
    success_url = reverse_lazy('hr:leavetype_list')

    def form_valid(self, form):
        # is_accumulative is a fixed business rule (only Annual is ever True),
        # not an admin-editable toggle — it's no longer on this form, so set
        # the sensible default explicitly here rather than relying on the
        # model field's default=True (which would wrongly mark every new
        # custom leave type as accumulative).
        form.instance.is_accumulative = (form.instance.code == 'annual')
        messages.success(self.request, 'Leave type created successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Leave Type'
        context['button_text'] = 'Save Leave Type'
        return context


class LeaveTypeUpdateView(AdminRequiredMixin, UpdateView):
    model = LeaveType
    form_class = LeaveTypeForm
    template_name = 'hr/leavetype_form.html'
    success_url = reverse_lazy('hr:leavetype_list')

    def form_valid(self, form):
        # Snapshot the pre-save day count so we know, after saving, whether it
        # actually changed (LeaveType.save() already propagates a changed
        # default_annual_days to every existing entitlement of this type, for
        # every year — this is deliberate, tested behavior, see
        # LeaveTypeDaySyncTests). Here we additionally keep the *current* year
        # explicitly, visibly in sync via the same service the manual "Re-apply"
        # button on the Entitlements page uses, so an admin editing a leave
        # type gets the same confirmation without a separate manual click —
        # historical years are intentionally left to the model-level sync
        # above, not retroactively rewritten by this extra step.
        old_days = None
        if self.object and self.object.pk:
            old_days = (LeaveType.objects.filter(pk=self.object.pk)
                        .values_list('default_annual_days', flat=True).first())
        response = super().form_valid(form)
        messages.success(self.request, 'Leave type updated successfully.')
        if old_days is not None and old_days != self.object.default_annual_days:
            from hr.leave_services import reapply_leave_type_defaults
            current_year = timezone.now().year
            updated = reapply_leave_type_defaults(year=current_year, leave_type=self.object)
            messages.success(
                self.request,
                f'Re-applied leave-type day counts to {updated} entitlement(s) for {current_year}.')
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Leave Type'
        context['button_text'] = 'Save Changes'
        return context


class LeaveTypeDeleteView(AdminRequiredMixin, DeleteView):
    model = LeaveType
    template_name = 'hr/leavetype_confirm_delete.html'
    success_url = reverse_lazy('hr:leavetype_list')

    def form_valid(self, form):
        # LeaveEntitlement/LeaveRecord both use on_delete=PROTECT against
        # LeaveType, so a type with any employee history can't be deleted —
        # catch that instead of letting it 500.
        try:
            response = super().form_valid(form)
        except ProtectedError:
            messages.error(
                self.request,
                f'"{self.object.name}" can\'t be deleted — it\'s already used in an '
                'employee\'s entitlement or leave history. Set it to Inactive instead.')
            return redirect('hr:leavetype_list')
        messages.success(self.request, 'Leave type deleted successfully.')
        return response


# ─── Holiday CRUD ─────────────────────────────────────────────

class HolidayListView(AdminRequiredMixin, ListView):
    model = Holiday
    template_name = 'hr/holiday_list.html'
    context_object_name = 'holidays'
    ordering = ['date']
    paginate_by = 25


class HolidayCreateView(AdminRequiredMixin, CreateView):
    model = Holiday
    form_class = HolidayForm
    template_name = 'hr/holiday_form.html'
    success_url = reverse_lazy('hr:holiday_list')

    def form_valid(self, form):
        messages.success(self.request, 'Holiday created successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Holiday'
        context['button_text'] = 'Save Holiday'
        return context


class HolidayUpdateView(AdminRequiredMixin, UpdateView):
    model = Holiday
    form_class = HolidayForm
    template_name = 'hr/holiday_form.html'
    success_url = reverse_lazy('hr:holiday_list')

    def form_valid(self, form):
        messages.success(self.request, 'Holiday updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Holiday'
        context['button_text'] = 'Save Changes'
        return context


class HolidayDeleteView(AdminRequiredMixin, DeleteView):
    model = Holiday
    template_name = 'hr/holiday_confirm_delete.html'
    success_url = reverse_lazy('hr:holiday_list')

    def form_valid(self, form):
        messages.success(self.request, 'Holiday deleted.')
        return super().form_valid(form)


# ─── WorkingDay CRUD ──────────────────────────────────────────

class WorkingDayListView(AdminRequiredMixin, ListView):
    model = WorkingDay
    template_name = 'hr/workingday_list.html'
    context_object_name = 'working_days'
    paginate_by = 25

class WorkingDayCreateView(AdminRequiredMixin, CreateView):
    model = WorkingDay
    form_class = WorkingDayForm
    template_name = 'hr/workingday_form.html'
    success_url = reverse_lazy('hr:workingday_list')
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs); ctx['title'] = 'Add Working Day'; ctx['button_text'] = 'Save'; return ctx

class WorkingDayUpdateView(AdminRequiredMixin, UpdateView):
    model = WorkingDay
    form_class = WorkingDayForm
    template_name = 'hr/workingday_form.html'
    success_url = reverse_lazy('hr:workingday_list')
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs); ctx['title'] = 'Edit Working Day'; ctx['button_text'] = 'Save Changes'; return ctx

class WorkingDayDeleteView(AdminRequiredMixin, DeleteView):
    model = WorkingDay
    template_name = 'hr/workingday_confirm_delete.html'
    success_url = reverse_lazy('hr:workingday_list')


# ─── Leave Records, Summary & Entitlement Generation ─────────────────────────


class LeaveRecordDeleteView(AdminRequiredMixin, DeleteView):
    model = LeaveRecord
    template_name = 'hr/leaverecord_confirm_delete.html'

    def get_success_url(self):
        return reverse_lazy('hr:leave_summary', kwargs={'pk': self.object.employee_id})


class LeaveRequestListView(SuperAdminRequiredMixin, ListView):
    """FIFO queue: oldest pending first, then a paginated history of decided requests."""
    model = LeaveRequest
    template_name = 'hr/leave_request_list.html'
    context_object_name = 'pending_requests'
    paginate_by = None  # pending list is expected to stay small; history is paginated separately below

    def test_func(self):
        # Broader than SuperAdminRequiredMixin's plain super-admin check: a designated
        # approver needs to see the queue of requests awaiting their decision, not just
        # reach individual detail pages via a notification link. Mirrors the same
        # widening already applied to LeaveRequestDetailView. can_view_leave_dashboard
        # also grants access to anyone with an active LeaveDashboardAccess record,
        # managed by Super Admins from the Leave Access page.
        return can_view_leave_dashboard(self.request.user)

    def get_queryset(self):
        qs = LeaveRequest.objects.filter(status='pending').select_related('employee', 'leave_type')
        sort = self.request.GET.get('sort')
        if sort == 'vacation_date':
            return qs.order_by('start_date')
        if sort == 'submitted_newest':
            return qs.order_by('-created_at')
        return qs.order_by('created_at')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        history_sort = self.request.GET.get('history_sort')
        history_order = 'decided_at' if history_sort == 'oldest' else '-decided_at'
        ctx['decided_requests'] = (
            LeaveRequest.objects.exclude(status='pending')
            .select_related('employee', 'leave_type', 'overridden_by')
            .prefetch_related('approvals__approver')
            .order_by(history_order)[:50])
        ctx['current_history_sort'] = 'oldest' if history_sort == 'oldest' else 'newest'
        ctx['current_sort'] = self.request.GET.get('sort') or 'submitted'
        ctx['can_revoke'] = has_override_access(self.request.user)
        ctx['pending_leave_revoke_requests'] = (
            LeaveRevokeRequest.objects.filter(status='pending')
            .select_related('leave_request__employee', 'leave_request__leave_type', 'requested_by')
            .order_by('created_at'))
        # Everything the VIEWER has personally logged for someone else —
        # not scoped to direct reports, so this is where an HR/Super Admin
        # user (who can log a request for any employee company-wide via
        # "Log Request" above, and may have no employee_profile of their
        # own to check My Profile with) finds and edits/cancels what they
        # logged. Scoped purely to created_by=self.request.user, matching
        # exactly what edit_leave_request/cancel_leave_request check.
        my_logged_requests = list(
            LeaveRequest.objects.filter(created_by=self.request.user)
            .select_related('employee', 'leave_type').prefetch_related('approvals__approver')
            .order_by('-created_at')[:20])
        for req in my_logged_requests:
            # Edit spans the whole pending window, same as Cancel — editing
            # after a decision resets that decision (see edit_leave_request),
            # so decided_approvers drives the "this will reset X's decision"
            # warning in the edit form.
            req.can_edit = bool(req.status == 'pending')
            req.decided_approvers = [
                a.approver.get_full_name() or a.approver.username
                for a in req.approvals.all() if a.decision != 'pending']
            req.can_cancel = bool(req.status == 'pending')
            # Restore a failed edit's attempted values + inline error (shared
            # with my_profile via _apply_leave_edit_retry).
            _apply_leave_edit_retry(self.request, req)
        ctx['my_logged_requests'] = my_logged_requests
        # Include any (now-inactive) leave type still on a shown request so the
        # edit dropdown never silently switches it to another type.
        _used_lt_ids = {req.leave_type_id for req in my_logged_requests}
        ctx['active_leave_types'] = LeaveType.objects.filter(
            Q(is_active=True) | Q(pk__in=_used_lt_ids)).order_by('name')
        return ctx

    def post(self, request, *args, **kwargs):
        from django.http import Http404
        from hr.leave_approval_services import (
            revoke_leave_request, decide_leave_revoke_request, edit_leave_request, cancel_leave_request,
        )
        action = request.POST.get('action')
        if action == 'edit_leave_request':
            from .forms import LeaveRequestForm
            target = LeaveRequest.objects.filter(pk=request.POST.get('request_id')).first()
            if target is None or target.created_by_id != request.user.id:
                return HttpResponse('You did not submit this request.', status=403)
            edit_form = LeaveRequestForm(
                request.POST, request.FILES, fixed_employee=target.employee, exclude_request_id=target.pk,
                allow_leave_type=target.leave_type, initial={'document': target.document})
            if edit_form.is_valid():
                try:
                    edit_leave_request(
                        target, request.user, leave_type=edit_form.cleaned_data['leave_type'],
                        start_date=edit_form.cleaned_data['start_date'], end_date=edit_form.cleaned_data['end_date'],
                        employee_reason=edit_form.cleaned_data['employee_reason'],
                        document=edit_form.cleaned_data['document'])
                    messages.success(request, 'Leave request updated.')
                except ValueError as exc:
                    # Service-level failure — preserve values + show the reason
                    # inline on the row, same as a form-invalid edit.
                    return _edit_leave_retry_redirect(
                        request, 'hr:leave_request_list', target, edit_form,
                        extra_params={'tab': 'logged'}, error_text=str(exc))
            else:
                return _edit_leave_retry_redirect(
                    request, 'hr:leave_request_list', target, edit_form, extra_params={'tab': 'logged'})
            # opened_leave + tab=logged tell the page which row's edit panel
            # to re-expand, and which tab to switch to, after the reload —
            # so the change is actually visible instead of collapsing back
            # to hidden on a tab the user isn't looking at.
            return redirect(f"{reverse('hr:leave_request_list')}?opened_leave={target.pk}&tab=logged")
        elif action == 'cancel_leave_request':
            target = LeaveRequest.objects.filter(pk=request.POST.get('request_id')).first()
            if target is None or target.created_by_id != request.user.id:
                return HttpResponse('You did not submit this request.', status=403)
            try:
                cancel_leave_request(target, request.user, request.POST.get('reason', ''))
                messages.success(request, 'Leave request cancelled.')
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == 'revoke':
            if not has_override_access(request.user):
                return HttpResponse('You do not have override access to revoke this request.', status=403)
            target = LeaveRequest.objects.filter(pk=request.POST.get('request_id')).first()
            if target is None:
                raise Http404('No such leave request.')
            try:
                revoke_leave_request(target, request.user, request.POST.get('reason', ''))
                messages.success(request, 'Leave request revoked.')
            except ValueError as exc:
                messages.error(request, str(exc))
        elif action == 'decide_revoke_request':
            revoke_req = LeaveRevokeRequest.objects.filter(pk=request.POST.get('revoke_request_id')).first()
            if revoke_req is None:
                raise Http404('No such revoke request.')
            try:
                decide_leave_revoke_request(
                    revoke_req, request.user, request.POST.get('decision'),
                    decision_note=request.POST.get('decision_note', ''))
                messages.success(request, 'Revoke request decided.')
            except ValueError as exc:
                messages.error(request, str(exc))
        return redirect('hr:leave_request_list')


class LeaveRequestCreateView(HRScopedAccessMixin, FormView):
    # HRScopedAccessMixin covers the Role-based tiers (admin/super_admin/
    # erp_admin, plus site/project_manager, document_controller); test_func
    # is overridden below to also admit anyone who is the live main_manager
    # of at least one active employee, independent of Role — direct reports
    # only, never the downstream subtree (see is_direct_manager_of_anyone).
    # This view serves the "Add Leave" buttons (Entitlements -> Leave
    # Summary / Attendance Matrix) and My Profile's "My Team" section. A
    # submitter can only *log* a request here — approving it in the queue
    # stays with designated approvers, unchanged for manager-logged
    # requests too.
    form_class = LeaveRequestForm
    template_name = 'hr/leave_request_form.html'

    def test_func(self):
        return can_manage_hr_scoped(self.request.user) or is_direct_manager_of_anyone(self.request.user)

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        if 'employee' not in form.fields:
            return form
        # Restrict the employee dropdown to the union of the submitter's
        # Role-based team scope and their own live direct reports. A scoped
        # role (or plain manager) thus cannot log leave for anyone outside
        # that set (an out-of-scope pk would also fail ModelChoiceField
        # validation).
        ids = scoped_employee_ids(self.request.user)
        if ids is None:
            return form  # unrestricted (admin tiers) — direct reports add nothing new
        emp = getattr(self.request.user, 'employee_profile', None)
        direct_report_ids = set(emp.main_reports.filter(is_active=True).values_list('pk', flat=True)) if emp else set()
        form.fields['employee'].queryset = form.fields['employee'].queryset.filter(pk__in=(ids | direct_report_ids))
        return form

    def get_initial(self):
        initial = super().get_initial()
        employee_id = self.request.GET.get('employee')
        if employee_id:
            initial['employee'] = employee_id
        return initial

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # A plain direct manager (no Role, no LeaveDashboardAccess/override
        # access) can submit here but cannot view the Leave Requests queue —
        # that's a separate, more restrictive page. The template's Back/
        # Cancel links must not point them at a page they'll immediately
        # get a 403 on.
        ctx['can_view_queue'] = can_view_leave_dashboard(self.request.user)
        # This view has multiple entry points (Leave Summary's "Add Leave",
        # the sidebar, Attendance Matrix, My Profile's "My Team"). Only
        # Leave Summary passes back=leave_summary&employee=<id>, so the
        # Back/Cancel links can return there instead of defaulting to the
        # queue/profile fallback below.
        employee_id = self.request.GET.get('employee')
        if self.request.GET.get('back') == 'leave_summary' and employee_id and employee_id.isdigit():
            ctx['back_url'] = reverse('hr:leave_summary', kwargs={'pk': employee_id})
            ctx['back_label'] = 'Back to Leave Summary'
        elif ctx['can_view_queue']:
            ctx['back_url'] = reverse('hr:leave_request_list')
            ctx['back_label'] = 'Back to Queue'
        else:
            ctx['back_url'] = reverse('hr:my_profile')
            ctx['back_label'] = 'Back to My Profile'
        return ctx

    def _success_redirect(self):
        if can_view_leave_dashboard(self.request.user):
            return redirect('hr:leave_request_list')
        return redirect('hr:my_profile')

    def form_valid(self, form):
        from hr.leave_approval_services import submit_leave_request, grant_exception_days
        from hr.leave_services import preview_leave_shortfall

        employee = form.cleaned_data['employee']
        leave_type = form.cleaned_data['leave_type']
        start_date = form.cleaned_data['start_date']
        end_date = form.cleaned_data['end_date']
        confirmed = bool(self.request.POST.get('confirm_grant_and_log'))
        can_grant = has_override_access(self.request.user)

        # Super Admins only: before creating anything, warn inline in this
        # same form if the entered request exceeds the employee's balance,
        # and offer a one-click "Log Anyway" that grants the exact shortfall
        # as an exception (so the employee's limit itself reflects it,
        # e.g. 45 -> 47) and then logs the request normally. Anyone without
        # override access skips straight to the existing behavior below —
        # their over-cap Site request still gets held for normal approval,
        # per validate_leave_submission's location-based rule.
        if can_grant and not confirmed:
            shortfall = preview_leave_shortfall(employee, leave_type, start_date, end_date)
            if shortfall:
                return self.render_to_response(self.get_context_data(
                    form=form, balance_shortfall=shortfall, balance_warning_employee=employee,
                    balance_warning_leave_type=leave_type))

        try:
            if confirmed:
                if not can_grant:
                    raise PermissionDenied
                shortfall = preview_leave_shortfall(employee, leave_type, start_date, end_date)
                if shortfall:
                    grant_exception_days(
                        employee=employee, leave_type=leave_type, year=start_date.year, days=shortfall,
                        granted_by=self.request.user,
                        reason=f'Auto-granted via Log Request by '
                               f'{self.request.user.get_full_name() or self.request.user.username} to allow a '
                               f'{start_date}–{end_date} request that exceeded the balance by {shortfall} day(s).')
            submit_leave_request(
                employee=employee, leave_type=leave_type, start_date=start_date, end_date=end_date,
                employee_reason=form.cleaned_data['employee_reason'], document=form.cleaned_data['document'],
                created_by=self.request.user,
            )
        except ValueError as exc:
            # The form already passed its own (unlocked) balance/overlap
            # check — reaching here means submit_leave_request's locked,
            # authoritative re-check caught something that changed in the
            # meantime (most likely a genuinely concurrent submission).
            form.add_error(None, str(exc))
            return self.form_invalid(form)
        messages.success(self.request, 'Leave request logged and sent for approval.')
        return self._success_redirect()


class EmployeeLeaveSummaryView(HRScopedAccessMixin, DetailView):
    model = Employee
    template_name = 'hr/leave_summary.html'
    context_object_name = 'employee'

    def get_queryset(self):
        # Team-scoped roles can only open a leave summary for their own reports.
        return scope_employee_queryset(self.request.user, super().get_queryset())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        year = _int_or(self.request.GET.get('year'), timezone.now().year)
        ctx['year'] = year
        ctx['entitlements'] = LeaveEntitlement.objects.filter(
            employee=self.object, year=year).select_related('leave_type')
        ctx['records'] = self.object.leave_records.filter(
            start_date__year=year).select_related('leave_type')
        ctx['can_grant_exception'] = has_override_access(self.request.user)
        return ctx


@login_required
def entitlement_year(request):
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')
    year = _int_or(request.GET.get('year'), timezone.now().year)
    if request.method == 'POST':
        post_year = _int_or(request.POST.get('year'), year)
        if request.POST.get('action') == 'reapply':
            from hr.leave_services import reapply_leave_type_defaults
            updated = reapply_leave_type_defaults(post_year)
            messages.success(
                request,
                f'Re-applied leave-type day counts to {updated} entitlement(s) for {post_year}.')
        else:
            created = generate_year_entitlements(post_year, actor=request.user)
            messages.success(request, f'Generated {created} entitlement row(s) for {post_year}.')
        return redirect(f"{reverse('hr:entitlement_year')}?year={post_year}")
    entitlements = (LeaveEntitlement.objects.filter(year=year)
                    .select_related('employee', 'leave_type')
                    .order_by('employee__full_name', 'leave_type__name'))

    # Batch the taken-days and exception-grant totals per (employee, leave_type)
    # in one query each (avoid N+1) — mirrors LeaveEntitlement.taken_days/
    # exception_days, but batched for a whole-company table instead of one
    # row at a time.
    from decimal import Decimal
    from collections import OrderedDict
    from .models import LeaveExceptionGrant
    taken_map = {}
    for r in (LeaveRecord.objects.filter(start_date__year=year)
              .values('employee_id', 'leave_type_id')
              .annotate(t=Sum('days'))):
        taken_map[(r['employee_id'], r['leave_type_id'])] = r['t'] or Decimal('0')
    exception_map = {}
    for r in (LeaveExceptionGrant.objects.filter(year=year)
              .values('employee_id', 'leave_type_id')
              .annotate(t=Sum('days'))):
        exception_map[(r['employee_id'], r['leave_type_id'])] = r['t'] or Decimal('0')

    # Group entitlement rows under each employee with combined totals.
    groups = OrderedDict()
    for e in entitlements:
        g = groups.get(e.employee_id)
        if g is None:
            g = groups[e.employee_id] = {
                'employee': e.employee, 'rows': [],
                'total_entitled': Decimal('0'), 'total_taken': Decimal('0'),
                'total_exception': Decimal('0'), 'total_remaining': Decimal('0'),
            }
        taken = taken_map.get((e.employee_id, e.leave_type_id), Decimal('0'))
        exception = exception_map.get((e.employee_id, e.leave_type_id), Decimal('0'))
        remaining = e.entitled_days + exception - taken
        g['rows'].append({'leave_type': e.leave_type, 'entitled': e.entitled_days,
                          'taken': taken, 'exception': exception, 'remaining': remaining})
        # Only standard accrued leave (Annual) counts toward the top-level totals.
        # Conditional/incidental leave (Sick, Marriage, Death of Family Member, Umrah,
        # New Born) is shown in the per-type breakdown above but excluded from the summary.
        if e.leave_type.is_accumulative:
            g['total_entitled'] += e.entitled_days
            g['total_taken'] += taken
            g['total_exception'] += exception
            g['total_remaining'] += remaining

    return render(request, 'hr/entitlement_year.html',
                  {'year': year, 'groups': list(groups.values()),
                   'entitlement_count': entitlements.count(),
                   'can_adjust_balance': has_override_access(request.user)})


class AttendanceHistoryView(HRScopedAccessMixin, DetailView):
    model = Employee
    template_name = 'hr/attendance_history.html'
    context_object_name = 'employee'

    def get_queryset(self):
        # Team-scoped roles can only open history for their own reports (an
        # out-of-team employee 404s); admin tiers see everyone.
        return scope_employee_queryset(self.request.user, super().get_queryset())

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        now = timezone.now()
        year = _int_or(self.request.GET.get('year'), now.year)
        month = _int_or(self.request.GET.get('month'), now.month, lo=1, hi=12)
        qs = self.object.attendance.filter(date__year=year, date__month=month).order_by('date')
        counts = {row['status']: row['n'] for row in qs.values('status').annotate(n=Count('id'))}
        total_hours = qs.aggregate(s=Sum('hours_worked'))['s'] or 0
        ctx.update({
            'year': year, 'month': month, 'records': qs,
            'summary': {
                'present': counts.get('present', 0), 'absent': counts.get('absent', 0),
                'leave': counts.get('leave', 0), 'holiday': counts.get('holiday', 0),
                'weekend': counts.get('weekend', 0), 'total_hours': total_hours,
                'late': counts.get('late', 0), 'wfh': counts.get('wfh', 0),
            },
        })
        return ctx


def _parse_date(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except (TypeError, ValueError):
        return timezone.now().date()


def _int_or(value, default, lo=None, hi=None):
    """Parse an int query param, falling back to `default` on bad/out-of-range input."""
    try:
        n = int(value)
    except (TypeError, ValueError):
        return default
    if (lo is not None and n < lo) or (hi is not None and n > hi):
        return default
    return n


@login_required
def attendance_grid(request):
    if not can_manage_hr_scoped(request.user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')

    day = _parse_date(request.GET.get('date') or (request.POST.get('date') if request.method == 'POST' else None))
    # Office / Site segregation. 'office' tab = office + unassigned (so nobody is
    # hidden until back-filled); 'site' tab = site only.
    location = (request.GET.get('location') or request.POST.get('location') or 'office')
    # Team-scoped roles (PM / Site Manager / Doc Controller) only see their own
    # reports; admin tiers see everyone. POST below iterates `employees`, so this
    # also blocks a scoped role from writing attendance outside their team.
    emp_qs = scope_employee_queryset(request.user, Employee.objects.filter(is_active=True))
    if location == 'site':
        emp_qs = emp_qs.filter(work_location='site')
    else:
        location = 'office'
        emp_qs = emp_qs.exclude(work_location='site')
    employees = list(emp_qs.order_by('full_name'))
    from hr.models import LateQuery
    approved_late_query_employee_ids = set(LateQuery.objects.filter(
        attendance_record__date=day, status='approved').values_list('employee_id', flat=True))

    if request.method == 'POST':
        for emp in employees:
            ci = request.POST.get(f'check_in_{emp.pk}') or None
            co = request.POST.get(f'check_out_{emp.pk}') or None
            ci_t = datetime.strptime(ci, '%H:%M').time() if ci else None
            co_t = datetime.strptime(co, '%H:%M').time() if co else None
            # Reconcile per-day WFH flag before deriving status.
            wants_wfh = request.POST.get(f'wfh_{emp.pk}') == '1'
            if wants_wfh:
                WFHRecord.objects.get_or_create(
                    employee=emp, start_date=day, end_date=day,
                    defaults={'created_by': request.user})
            else:
                # Only delete single-day records; never touch multi-day WFH.
                WFHRecord.objects.filter(employee=emp, start_date=day, end_date=day).delete()
            status, hours = derive_status(emp, day, ci_t, co_t, approved_late_query_employee_ids)
            AttendanceRecord.objects.update_or_create(
                employee=emp, date=day,
                defaults={'check_in': ci_t, 'check_out': co_t, 'status': status,
                          'hours_worked': hours, 'created_by': request.user})
        messages.success(request, f'Attendance saved for {day:%Y-%m-%d}.')
        return redirect(f"{reverse('hr:attendance_grid')}?date={day:%Y-%m-%d}&location={location}")

    existing = {r.employee_id: r for r in AttendanceRecord.objects.filter(date=day)}
    rows = []
    for emp in employees:
        rec = existing.get(emp.pk)
        preview_status, _ph = derive_status(emp, day, rec.check_in if rec else None,
                                            rec.check_out if rec else None, approved_late_query_employee_ids)
        locked = preview_status in ('leave', 'holiday', 'weekend')
        is_wfh = WFHRecord.objects.filter(employee=emp, start_date=day, end_date=day).exists()
        rows.append({'employee': emp, 'record': rec,
                     'status': rec.status if rec else preview_status, 'locked': locked,
                     'is_wfh': is_wfh})
    return render(request, 'hr/attendance_grid.html', {
        'day': day, 'rows': rows, 'location': location})


def _late_warning_threshold():
    """Lates in a month that earn a warning - read from the model so the
    register's explanation cannot drift from the rule that sends them."""
    from hr.models import LATE_WARNING_THRESHOLD
    return LATE_WARNING_THRESHOLD


@login_required
def attendance_settings(request):
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')
    obj = AttendanceSettings.load()
    if request.method == 'POST':
        form = AttendanceSettingsForm(request.POST)
        if form.is_valid():
            obj.weekend_days = ','.join(form.cleaned_data['weekend_days'])
            obj.expected_in_by = form.cleaned_data.get('expected_in_by') or obj.expected_in_by
            obj.save()
            messages.success(request, 'Attendance settings saved.')
            return redirect('hr:attendance_settings')
    else:
        form = AttendanceSettingsForm()
        form.initial_from(obj)
    return render(request, 'hr/attendance_settings.html', {'form': form})


@login_required
@require_POST
def attendance_regenerate(request):
    """Re-derive stored status for all records on a given date (after leave/holiday edits)."""
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')
    day = _parse_date(request.POST.get('date'))
    from hr.models import LateQuery
    approved_late_query_employee_ids = set(LateQuery.objects.filter(
        attendance_record__date=day, status='approved').values_list('employee_id', flat=True))
    n = 0
    for rec in AttendanceRecord.objects.filter(date=day).select_related('employee'):
        status, hours = derive_status(rec.employee, day, rec.check_in, rec.check_out, approved_late_query_employee_ids)
        AttendanceRecord.objects.filter(pk=rec.pk).update(status=status, hours_worked=hours)
        n += 1
    messages.success(request, f'Regenerated {n} record(s) for {day:%Y-%m-%d}.')
    return redirect(f"{reverse('hr:attendance_grid')}?date={day:%Y-%m-%d}")


@login_required
def attendance_matrix(request):
    if not can_manage_hr_scoped(request.user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')
    period = request.GET.get('period') if request.GET.get('period') in ('week', 'month') else 'month'
    anchor = _parse_date(request.GET.get('date'))
    start, end = period_range(period, anchor)
    # Office / Site segregation (office tab = office + unassigned).
    location = request.GET.get('location') or 'office'
    # Team-scoped roles see only their own reports; admin tiers see everyone.
    emp_qs = scope_employee_queryset(request.user, Employee.objects.filter(is_active=True))
    if location == 'site':
        emp_qs = emp_qs.filter(work_location='site')
    else:
        location = 'office'
        emp_qs = emp_qs.exclude(work_location='site')
    employees = list(emp_qs.order_by('full_name'))
    # Filter bar: which people to show, and how to rank them. Applied before
    # build_matrix so the grid draws only the rows that survive, in order.
    from hr.attendance_summary import (
        apply_register_filters, resolve_register_filters,
        SHOW_CHOICES, ORDER_CHOICES,
    )
    show, order = resolve_register_filters(request.GET)
    employees, totals, filter_counts = apply_register_filters(
        employees, start, end, show, order)
    days, rows, weekend_dates = build_matrix(employees, start, end, with_weekend_dates=True)
    for row in rows:
        row['totals'] = totals[row['employee'].pk]
    prev_anchor = start - timedelta(days=1)
    next_anchor = end + timedelta(days=1)
    return render(request, 'hr/attendance_matrix.html', {
        'period': period, 'anchor': anchor, 'start': start, 'end': end,
        'days': days, 'rows': rows, 'location': location,
        'prev_anchor': prev_anchor, 'next_anchor': next_anchor,
        'today': timezone.now().date(),
        'leave_types': LeaveType.objects.filter(is_active=True).order_by('name'),
        'weekend_dates': weekend_dates,
        'show': show, 'order': order,
        'show_choices': SHOW_CHOICES, 'order_choices': ORDER_CHOICES,
        'filter_counts': filter_counts,
        'late_warning_threshold': _late_warning_threshold(),
    })


@login_required
@login_required
def attendance_matrix_export_excel(request):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if not can_manage_hr_scoped(request.user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')

    period = request.GET.get('period') if request.GET.get('period') in ('week', 'month') else 'month'
    anchor = _parse_date(request.GET.get('date'))
    start, end = period_range(period, anchor)
    location = request.GET.get('location') or 'office'
    # Team-scoped roles export only their own reports; admin tiers export all.
    emp_qs = scope_employee_queryset(request.user, Employee.objects.filter(is_active=True))
    if location == 'site':
        emp_qs = emp_qs.filter(work_location='site')
    else:
        location = 'office'
        emp_qs = emp_qs.exclude(work_location='site')
    employees = list(emp_qs.order_by('full_name'))
    # Same filter bar as the register: an export that quietly ignored it would
    # hand back a different population from the one on screen.
    from hr.attendance_summary import apply_register_filters, resolve_register_filters
    show, order = resolve_register_filters(request.GET)
    employees, _totals, _counts = apply_register_filters(
        employees, start, end, show, order)
    days, rows = build_matrix(employees, start, end)

    status_labels = {'': '-', 'leave': 'L', 'holiday': 'H', 'weekend': 'WE', 'wfh': 'WFH', 'present': 'P', 'absent': 'A'}
    from hr.models import LeaveRecord, Holiday
    leave_type_map = {}
    for lr in LeaveRecord.objects.filter(employee_id__in=[e.pk for e in employees], start_date__lte=end, end_date__gte=start).select_related('leave_type'):
        dd = max(lr.start_date, start)
        last = min(lr.end_date, end)
        while dd <= last:
            leave_type_map[(lr.employee_id, dd)] = lr.leave_type.code
            dd += timedelta(days=1)
    holiday_category_map = {h.date: h.category for h in Holiday.objects.filter(date__range=(start, end))}

    COLOR_PRESENT = 'C6E0B4'
    COLOR_ABSENT = 'C00000'
    COLOR_LATE = 'F4B183'
    COLOR_WEEKEND = 'D9D2E9'
    COLOR_ANNUAL_LEAVE = 'FCE4D6'
    COLOR_NATIONAL = '00B050'
    COLOR_EID = 'FFFF00'
    COLOR_OTHER = 'BDD7EE'  # light blue - shared fallback for WFH, non-Annual leave types, and ungrouped holidays

    def fill_for(cell_data, emp_id):
        status = cell_data['status']
        if status == 'present':
            return COLOR_PRESENT
        if status == '':
            return None
        if status == 'absent':
            return COLOR_ABSENT
        if status == 'late':
            return COLOR_LATE
        if status == 'weekend':
            return COLOR_WEEKEND
        if status == 'leave':
            code = leave_type_map.get((emp_id, cell_data['date']))
            if code == 'annual':
                return COLOR_ANNUAL_LEAVE
            return COLOR_OTHER
        if status == 'holiday':
            cat = holiday_category_map.get(cell_data['date'], 'other')
            if cat == 'saudi_national_day':
                return COLOR_NATIONAL
            if cat == 'eid':
                return COLOR_EID
            return COLOR_OTHER
        # AUTHOR'S NOTE: 'wfh' status currently falls through to COLOR_OTHER below,
        # same as Sick/Marriage/other leave types and ungrouped holidays.
        # To give WFH its own distinct color in the future: add
        # if status == 'wfh': return COLOR_WFH here (above this comment),
        # define a new COLOR_WFH constant near the other COLOR_ constants above,
        # and add a matching test to AttendanceExportColorTests in hr/tests.py.
        return COLOR_OTHER

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ATTENDANCE MATRIX'

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(row=1, column=1, value=f'Attendance Matrix ({location.title()}) - {start.strftime("%d-%b-%Y")} to {end.strftime("%d-%b-%Y")}').font = Font(bold=True, size=13)

    header_row = 3
    ws.cell(row=header_row, column=1, value='Employee').font = header_font
    ws.cell(row=header_row, column=1).fill = header_fill
    ws.cell(row=header_row, column=1).border = thin_border
    for col, day in enumerate(days, 2):
        cell = ws.cell(row=header_row, column=col, value=day.strftime('%d-%b'))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    row = header_row
    for r in rows:
        row += 1
        ws.cell(row=row, column=1, value=r['employee'].full_name).border = thin_border
        for col, cell_data in enumerate(r['cells'], 2):
            label = status_labels.get(cell_data['status'], cell_data['status'])
            # Stack the check-in (and check-out) time beneath the status letter,
            # e.g. "P" over "09:12-17:30". Non-present days keep just the letter.
            times = cell_time_lines(cell_data)
            display_lines = ['-'.join(times)] if times else []
            if cell_data.get('exception_reason'):
                display_lines.append(cell_data['exception_reason'])
            value = label if not display_lines else f'{label}\n' + '\n'.join(display_lines)
            c = ws.cell(row=row, column=col, value=value)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            color_value = fill_for(cell_data, r['employee'].pk)
            if color_value is not None:
                c.fill = PatternFill(start_color=color_value, end_color=color_value, fill_type='solid')

    ws.column_dimensions['A'].width = 25
    for col in range(2, len(days) + 2):
        # Wide enough for "09:12-17:30" so the time line doesn't clip.
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f'attendance_register_{start.strftime("%B_%Y").lower()}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def attendance_matrix_export_pdf(request):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    import io

    if not can_manage_hr_scoped(request.user):
        messages.error(request, 'Admin access required.')
        return redirect('hr:hr_dashboard')

    period = request.GET.get('period') if request.GET.get('period') in ('week', 'month') else 'month'
    anchor = _parse_date(request.GET.get('date'))
    start, end = period_range(period, anchor)
    location = request.GET.get('location') or 'office'
    # Team-scoped roles export only their own reports; admin tiers export all.
    emp_qs = scope_employee_queryset(request.user, Employee.objects.filter(is_active=True))
    if location == 'site':
        emp_qs = emp_qs.filter(work_location='site')
    else:
        location = 'office'
        emp_qs = emp_qs.exclude(work_location='site')
    employees = list(emp_qs.order_by('full_name'))
    # Same filter bar as the register: an export that quietly ignored it would
    # hand back a different population from the one on screen.
    from hr.attendance_summary import apply_register_filters, resolve_register_filters
    show, order = resolve_register_filters(request.GET)
    employees, _totals, _counts = apply_register_filters(
        employees, start, end, show, order)
    days, rows = build_matrix(employees, start, end)

    status_labels = {'': '-', 'leave': 'L', 'holiday': 'H', 'weekend': 'WE', 'wfh': 'WFH', 'present': 'P', 'absent': 'A'}
    from hr.models import LeaveRecord, Holiday
    leave_type_map = {}
    for lr in LeaveRecord.objects.filter(employee_id__in=[e.pk for e in employees], start_date__lte=end, end_date__gte=start).select_related('leave_type'):
        dd = max(lr.start_date, start)
        last = min(lr.end_date, end)
        while dd <= last:
            leave_type_map[(lr.employee_id, dd)] = lr.leave_type.code
            dd += timedelta(days=1)
    holiday_category_map = {h.date: h.category for h in Holiday.objects.filter(date__range=(start, end))}

    COLOR_PRESENT = colors.HexColor('#C6E0B4')
    COLOR_ABSENT = colors.HexColor('#C00000')
    COLOR_LATE = colors.HexColor('#F4B183')
    COLOR_WEEKEND = colors.HexColor('#D9D2E9')
    COLOR_ANNUAL_LEAVE = colors.HexColor('#FCE4D6')
    COLOR_NATIONAL = colors.HexColor('#00B050')
    COLOR_EID = colors.HexColor('#FFFF00')
    COLOR_OTHER = colors.HexColor('#BDD7EE')  # light blue - shared fallback for WFH, non-Annual leave types, and ungrouped holidays

    def pdf_fill_for(cell_data, emp_id):
        status = cell_data['status']
        if status == 'present':
            return COLOR_PRESENT
        if status == '':
            return None
        if status == 'absent':
            return COLOR_ABSENT
        if status == 'late':
            return COLOR_LATE
        if status == 'weekend':
            return COLOR_WEEKEND
        if status == 'leave':
            code = leave_type_map.get((emp_id, cell_data['date']))
            if code == 'annual':
                return COLOR_ANNUAL_LEAVE
            return COLOR_OTHER
        if status == 'holiday':
            cat = holiday_category_map.get(cell_data['date'], 'other')
            if cat == 'saudi_national_day':
                return COLOR_NATIONAL
            if cat == 'eid':
                return COLOR_EID
            return COLOR_OTHER
        # AUTHOR'S NOTE: 'wfh' status currently falls through to COLOR_OTHER below,
        # same as Sick/Marriage/other leave types and ungrouped holidays.
        # To give WFH its own distinct color in the future: add
        # if status == 'wfh': return COLOR_WFH here (above this comment),
        # define a new COLOR_WFH constant near the other COLOR_ constants above,
        # and add a matching test to AttendanceExportColorTests in hr/tests.py.
        return COLOR_OTHER
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            leftMargin=18, rightMargin=18, topMargin=24, bottomMargin=18)
    styles = getSampleStyleSheet()
    # Small styles so the status letter + stacked times fit a narrow day column.
    name_style = ParagraphStyle('emp', fontSize=6, leading=7)
    cell_style = ParagraphStyle('cell', fontSize=5, leading=6, alignment=1)  # 1 = centre
    elements = [Paragraph(f'Attendance Matrix ({location.title()}) - {start.strftime("%d-%b-%Y")} to {end.strftime("%d-%b-%Y")}', styles['Title'])]

    header = ['Employee'] + [d.strftime('%d-%b') for d in days]
    data = [header]
    bg_commands = []
    for row_idx, r in enumerate(rows, 1):
        row_data = [Paragraph(r['employee'].full_name, name_style)]
        for col_idx, c in enumerate(r['cells'], 1):
            label = status_labels.get(c['status'], c['status'])
            times = cell_time_lines(c)
            # Stack the status letter over the check-in/out times, each on its
            display_lines = list(times)
            if c.get('exception_reason'):
                from xml.sax.saxutils import escape as _xml_escape
                display_lines.append(_xml_escape(c['exception_reason']))
            row_data.append(Paragraph('<br/>'.join([label] + display_lines), cell_style) if display_lines else label)
            # Only emit a BACKGROUND for cells that actually have a fill colour;
            # blank/unrecorded days return None (no fill). Emitting a None colour
            # here would crash reportlab at render (setFillColor(None)).
            color_value = pdf_fill_for(c, r['employee'].pk)
            if color_value is not None:
                bg_commands.append(('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), color_value))
        data.append(row_data)

    # Fixed column widths: a fair share of the usable width per day column.
    usable = landscape(A4)[0] - 36  # page width minus left+right margins
    emp_w = 90
    day_w = (usable - emp_w) / max(len(days), 1)
    col_widths = [emp_w] + [day_w] * len(days)

    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1F4E79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTSIZE', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ALIGN', (1, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ] + bg_commands))
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    filename = f'attendance_register_{start.strftime("%B_%Y").lower()}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@require_POST
def attendance_mark_leave(request):
    if not can_manage_hr_scoped(request.user):
        raise PermissionDenied
    try:
        payload = json.loads(request.body or '{}')
        emp_id = int(payload['employee'])
        day = datetime.strptime(payload['date'], '%Y-%m-%d').date()
        lt_id = int(payload['leave_type'])
    except (ValueError, KeyError, TypeError):
        return JsonResponse({'error': 'Bad payload'}, status=400)

    employee = Employee.objects.filter(pk=emp_id, is_active=True).first()
    leave_type = LeaveType.objects.filter(pk=lt_id, is_active=True).first()
    if employee is None or leave_type is None:
        return JsonResponse({'error': 'Unknown employee or leave type'}, status=400)

    # Team-scoped roles may only mark leave for their own reports.
    scope = scoped_employee_ids(request.user)
    if scope is not None and employee.pk not in scope:
        return JsonResponse({'error': 'Employee is not in your team.'}, status=403)

    # Certificate-required leave (e.g. Sick) can't be granted from the one-click
    # grid action — there's nowhere to attach the file. Send them to Add Leave.
    if leave_type.requires_medical_certificate:
        return JsonResponse(
            {'error': f'{leave_type.name} needs a medical certificate — use the Add Leave form.'},
            status=400)

    # Guard against double-booking: a day already inside any leave (incl. a
    # multi-day record showing stale 'present' in the matrix) must not get a
    # second overlapping LeaveRecord that would double-count the balance.
    if LeaveRecord.objects.filter(employee=employee, start_date__lte=day, end_date__gte=day).exists():
        return JsonResponse({'error': 'Already on leave that day.'}, status=400)

    with transaction.atomic():
        lr = LeaveRecord.objects.create(
            employee=employee, leave_type=leave_type,
            start_date=day, end_date=day, created_by=request.user)
        # Don't blank check_in/check_out — if the day already had clock times,
        # preserving them keeps the mark->unmark round trip lossless (unmark
        # re-derives back to present). hours_worked is nulled (no hours on leave).
        AttendanceRecord.objects.update_or_create(
            employee=employee, date=day,
            defaults={'status': 'leave', 'hours_worked': None, 'created_by': request.user})
    return JsonResponse({'ok': True, 'status': 'leave', 'leave_record_id': lr.pk})


@login_required
@require_POST
def attendance_unmark_leave(request):
    if not can_manage_hr_scoped(request.user):
        raise PermissionDenied
    try:
        lr_id = int(json.loads(request.body or '{}')['leave_record_id'])
    except (ValueError, KeyError, TypeError):
        return JsonResponse({'error': 'Bad payload'}, status=400)

    lr = LeaveRecord.objects.filter(pk=lr_id).first()
    if lr is None:
        return JsonResponse({'error': 'Not found'}, status=404)
    # Team-scoped roles may only unmark leave for their own reports.
    scope = scoped_employee_ids(request.user)
    if scope is not None and lr.employee_id not in scope:
        return JsonResponse({'error': 'Employee is not in your team.'}, status=403)
    if lr.start_date != lr.end_date:
        return JsonResponse({'error': 'Part of a multi-day leave — edit from the leave summary.'}, status=400)

    emp, emp_id, day = lr.employee, lr.employee_id, lr.start_date
    with transaction.atomic():
        lr.delete()
        ar = AttendanceRecord.objects.filter(employee_id=emp_id, date=day).first()
        if ar and (ar.check_in or ar.check_out):
            status, hours = derive_status(ar.employee, day, ar.check_in, ar.check_out)
            AttendanceRecord.objects.filter(pk=ar.pk).update(status=status, hours_worked=hours)
            new_status = status
        else:
            if ar:
                ar.delete()  # leave-only row -> restore blank/derived cell
            new_status = display_status_no_record(day, emp)
    return JsonResponse({'ok': True, 'status': new_status})


# ─── WFH Records ──────────────────────────────────────────────────────────────


class WFHRecordListView(AdminRequiredMixin, ListView):
    model = WFHRecord
    template_name = 'hr/wfhrecord_list.html'
    context_object_name = 'wfh_records'
    paginate_by = 25

    def get_queryset(self):
        return WFHRecord.objects.select_related('employee').all()


class WFHRecordCreateView(AdminRequiredMixin, CreateView):
    model = WFHRecord
    form_class = WFHRecordForm
    template_name = 'hr/wfhrecord_form.html'
    success_url = reverse_lazy('hr:wfh_list')

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'WFH recorded.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['title'] = 'Record WFH'
        ctx['button_text'] = 'Save WFH'
        return ctx


class WFHRecordDeleteView(AdminRequiredMixin, DeleteView):
    model = WFHRecord
    template_name = 'hr/wfhrecord_confirm_delete.html'
    success_url = reverse_lazy('hr:wfh_list')


class LeaveRequestDetailView(SuperAdminRequiredMixin, DetailView):
    model = LeaveRequest
    template_name = 'hr/leave_request_detail.html'
    context_object_name = 'leave_request'

    def test_func(self):
        # Broader than SuperAdminRequiredMixin's plain super-admin check: a designated
        # approver must also be able to open the detail page and record their decision.
        # Whether they're the *specific* pending approver for *this* request is enforced
        # per-action in post() (see the 'decide' branch) — this is just page access.
        # can_view_leave_dashboard also grants access to anyone with an active
        # LeaveDashboardAccess record, managed from the Leave Access page.
        return can_view_leave_dashboard(self.request.user)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # A user must never be able to act on their own leave request —
        # decide, edit a decision, add a note, or override — regardless of
        # what other authority (approver, override access) they hold.
        ctx['is_own_request'] = bool(
            self.object.employee.user_id and self.object.employee.user_id == self.request.user.id)
        ctx['is_approver'] = is_designated_approver(self.request.user)
        ctx['can_override'] = has_override_access(self.request.user) and not ctx['is_own_request']
        my_approval = self.object.approvals.filter(approver=self.request.user).first()
        ctx['my_approval'] = my_approval if not ctx['is_own_request'] else None
        ctx['visible_notes'] = self.object.notes.filter(is_internal=False) if not self.request.user.is_super_admin_user \
            else self.object.notes.all()
        # No fixed time limit — a decision stays editable for as long as the
        # overall request is still pending (see
        # hr.leave_approval_services._reconcile: finalizing requires every
        # approver to agree, so 'pending' here also covers a split decision
        # awaiting resolution, not just an outstanding vote).
        ctx['can_edit_decision'] = bool(
            my_approval and my_approval.decision in ('approved', 'disapproved')
            and self.object.status == 'pending')
        # A genuine conflict: every approver has decided, but not
        # unanimously — the request stays pending until one of them changes
        # their vote. Surfaced so the page can explain why it hasn't
        # finalized instead of leaving that a mystery.
        ctx['has_conflicting_decisions'] = bool(
            self.object.status == 'pending' and self.object.approvals.exists()
            and not self.object.approvals.filter(decision='pending').exists())
        if self.object.exceeds_balance:
            ctx['entitlement'] = LeaveEntitlement.objects.filter(
                employee=self.object.employee, leave_type=self.object.leave_type,
                year=self.object.start_date.year).first()
        return ctx

    def post(self, request, *args, **kwargs):
        from hr.leave_approval_services import record_approver_decision, edit_approver_decision, override_finalize
        self.object = self.get_object()
        action = request.POST.get('action')

        if action == 'decide':
            approval = self.object.approvals.filter(approver=request.user).first()
            if not approval:
                return HttpResponse('You are not a designated approver for this request.', status=403)
            try:
                if approval.decision == 'pending':
                    record_approver_decision(self.object, request.user, request.POST.get('decision'),
                                             comment=request.POST.get('comment', ''))
                    messages.success(request, 'Your decision has been recorded.')
                else:
                    edit_approver_decision(self.object, request.user, request.POST.get('decision'),
                                           edit_note=request.POST.get('edit_note', ''))
                    messages.success(request, 'Your decision has been updated.')
            except ValueError as exc:
                messages.error(request, str(exc))

        elif action == 'override':
            if not has_override_access(request.user):
                return HttpResponse('You do not have override access for this request.', status=403)
            try:
                override_finalize(self.object, request.user, request.POST.get('decision'),
                                  request.POST.get('reason', ''))
                messages.success(request, 'Request finalized via override.')
            except ValueError as exc:
                messages.error(request, str(exc))

        elif action == 'revoke':
            if not has_override_access(request.user):
                return HttpResponse('You do not have override access to revoke this request.', status=403)
            from hr.leave_approval_services import revoke_leave_request
            try:
                revoke_leave_request(self.object, request.user, request.POST.get('reason', ''))
                messages.success(request, 'Leave request revoked.')
            except ValueError as exc:
                messages.error(request, str(exc))

        elif action == 'add_note':
            # Restricted to the assigned approver(s) for THIS request, not any
            # Super Admin — a Super Admin with no approval row here is
            # view-only aside from the override escape hatch above.
            if not self.object.approvals.filter(approver=request.user).exists():
                return HttpResponse('You are not a designated approver for this request.', status=403)
            from .models import LeaveRequestNote
            note_text = request.POST.get('note', '').strip()
            if note_text:
                LeaveRequestNote.objects.create(
                    leave_request=self.object, author=request.user, note=note_text,
                    is_internal=bool(request.POST.get('is_internal')),
                )
                messages.success(request, 'Note added.')

        return redirect('hr:leave_request_detail', pk=self.object.pk)


class TeamExceptionsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Manager/HR queue for the attendance-exception workflow. Deliberately a
    single self-contained page (list + decide/override actions together) —
    the single-approver-plus-override model doesn't need a separate detail
    page per request the way the multi-approver leave workflow does; managers
    should be able to act on a pending row without leaving the queue."""
    model = AttendanceException
    template_name = 'hr/team_exceptions.html'
    context_object_name = 'pending_exceptions'

    def test_func(self):
        return can_view_team_exceptions(self.request.user)

    VALID_TABS = ('direct', 'secondary', 'all', 'late_queries')

    def _resolve_tab(self):
        """?tab=direct|secondary|all, defaulting to 'direct' if absent or
        invalid. The 'all' (All Organization Requests) tab is HR/Super-Admin
        only — a non-HR user requesting it is silently downgraded to 'direct'
        rather than erroring or leaking org-wide data."""
        user = self.request.user
        is_hr = bool(user.is_super_admin_user or user.is_admin_user or user.is_erp_admin_user)
        requested = self.request.GET.get('tab')
        if requested not in self.VALID_TABS:
            return 'direct'
        if requested in ('all', 'late_queries') and not is_hr:
            return 'direct'
        return requested

    def _tab_queryset(self, tab, user, emp):
        """Base (not yet status-filtered) queryset of AttendanceExceptions
        belonging to the given tab's employee set.

        - 'all' (HR-only, enforced by _resolve_tab): every request in the
          system, no exclusions — including the viewer's own (visibility
          only; the self-approval guardrail still blocks the buttons on that
          row via can_decide_attendance_exception / is_assigned_manager).
        - 'direct'/'secondary': scoped to the viewer's own live
          main_reports/secondary_reports, with the viewer's own request
          entirely excluded (not just button-blocked) per the self-approval
          visibility rule.
        """
        if tab == 'all':
            return AttendanceException.objects.all()
        if not emp:
            # A head-manager-role user with no linked Employee record has
            # nothing of their own to scope by — empty queue, not an error.
            return AttendanceException.objects.none()
        if tab == 'secondary':
            return AttendanceException.objects.filter(
                employee__secondary_managers=emp).exclude(employee__user=user)
        # Default: direct reports only.
        return AttendanceException.objects.filter(
            employee__main_manager=emp).exclude(employee__user=user)

    def get_queryset(self):
        # Soonest-deadline-first: event_date/event_start_time order doubles as
        # decision-deadline order (deadline = event_start + fixed 24h), and is
        # simpler to reason about for managers than sorting by a derived value.
        user = self.request.user
        emp = getattr(user, 'employee_profile', None)
        tab = self._resolve_tab()
        if tab == 'late_queries':
            from hr.models import AttendanceException
            return AttendanceException.objects.none()
        # 'pending' AND 'expired' both belong in the active/actionable queue —
        # an auto-expired-by-cron request was never actually decided by a
        # human, so it must stay visible and actionable here, not silently
        # vanish into history.
        return (self._tab_queryset(tab, user, emp).filter(status__in=('pending', 'expired'))
                .select_related('employee', 'employee__main_manager', 'main_manager')
                .order_by('event_date', 'event_start_time'))

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        user = self.request.user
        emp = getattr(user, 'employee_profile', None)
        tab = self._resolve_tab()
        is_hr = bool(user.is_super_admin_user or user.is_admin_user or user.is_erp_admin_user)

        pending = list(ctx['pending_exceptions'])
        # History is now the natural complement of the widened active queue:
        # only genuinely human-decided outcomes (approved/rejected) — expired
        # requests stay in the active queue above until a human actually
        # decides them.
        history_sort = self.request.GET.get('history_sort')
        history_order = 'decided_at' if history_sort == 'oldest' else '-decided_at'
        ctx['current_history_sort'] = 'oldest' if history_sort == 'oldest' else 'newest'
        decided = list(
            self._tab_queryset(tab, user, emp).filter(status__in=('approved', 'rejected', 'revoked'))
            .select_related('employee', 'employee__main_manager', 'main_manager', 'decided_by', 'overridden_by')
            .order_by(history_order)[:50])

        # Precompute per-row action visibility here (rather than re-deriving
        # the hierarchy/HR logic in the template): 'can_decide' is the normal
        # assigned-manager path; 'can_override' is offered to HR/upstream
        # viewers, but suppressed when the normal decide control already
        # covers the same person, to avoid showing two redundant controls on
        # the same row.
        now = timezone.now()
        for exc in pending + decided:
            # Self-approval guardrail: even if exc.employee.main_manager
            # somehow points back at the requester themselves, the decide
            # action must never be offered on one's own request.
            #
            # Checks the LIVE exc.employee.main_manager relationship, not the
            # exc.main_manager snapshot taken at submission — otherwise a
            # manager assigned/corrected via the Org Chart page after a
            # request was already submitted would be stuck seeing only the
            # Override control for that pre-existing request forever, even
            # though can_decide_attendance_exception's hierarchy-override
            # check already recognizes them live.
            is_assigned_manager = bool(
                exc.employee.main_manager and exc.employee.main_manager.user_id == user.id
                and exc.employee.user_id != user.id)
            # Widened alongside decide_attendance_exception's own allowed
            # starting statuses: an auto-expired request must still be
            # decidable normally by the real assigned manager, not forced
            # through Override.
            exc.can_decide = is_assigned_manager and exc.status in ('pending', 'expired')
            exc.can_override = (
                exc.status in ('pending', 'expired')
                and can_decide_attendance_exception(user, exc)
                and not exc.can_decide)
            # Countdown display state (Task 3 fix): a request can legitimately
            # be submitted up to 7 days before its event, so "time left until
            # decision_deadline" can read as up to ~8 days while the event
            # hasn't started yet — misleading, since the real 24h decision
            # window only opens at event_start. Before the event starts, the
            # template shows a static "Starts <date>" instead of a ticking
            # countdown; only once the event has started does it show the
            # real live countdown toward decision_deadline (at most 24h away).
            exc.event_has_started = bool(exc.event_start and now >= exc.event_start)

        ctx['pending_exceptions'] = pending
        ctx['decided_exceptions'] = decided
        ctx['tab'] = tab
        ctx['is_hr_viewer'] = is_hr
        # The "All Organization Requests" tab is only rendered at all for
        # HR/Super-Admin — "Direct Reports"/"Secondary Reports" are always
        # shown (even empty) for everyone who can reach this page.
        ctx['show_all_tab'] = is_hr
        ctx['can_revoke'] = has_override_access(user)
        ctx['pending_exception_revoke_requests'] = (
            AttendanceExceptionRevokeRequest.objects.filter(
                status='pending', attendance_exception__in=self._tab_queryset(tab, user, emp))
            .select_related('attendance_exception__employee', 'requested_by')
            .order_by('created_at'))
        # Per-tab pending counts for the tab labels — always all three
        # (regardless of which tab is currently selected), so switching tabs
        # doesn't need a second request to know what the other tabs hold.
        # Each count folds in pending revoke requests too, not just plain
        # pending/expired exceptions — a revoke request needs a decision
        # exactly as much as a new request does, so a tab reading "0" while
        # a revoke sits there awaiting review would be misleading.
        def _tab_count(tab_name):
            base = self._tab_queryset(tab_name, user, emp)
            return (
                base.filter(status__in=('pending', 'expired')).count()
                + AttendanceExceptionRevokeRequest.objects.filter(
                    status='pending', attendance_exception__in=base).count())
        ctx['direct_tab_count'] = _tab_count('direct')
        ctx['secondary_tab_count'] = _tab_count('secondary')
        if ctx['show_all_tab']:
            ctx['all_tab_count'] = _tab_count('all')
        from hr.models import LateQuery
        from notifications.models import Notification
        if tab == 'late_queries':
            ctx['pending_late_queries'] = LateQuery.objects.filter(
                status='pending').select_related('employee', 'attendance_record').order_by('-created_at')
            ctx['decided_late_queries'] = LateQuery.objects.filter(
                status__in=('approved', 'rejected')).select_related(
                    'employee', 'attendance_record', 'decided_by').order_by('-decided_at')[:50]
            ctx['late_email_notifications'] = Notification.objects.filter(
                verb='You were late 3 times this month').select_related('recipient').order_by('-created_at')[:50]
        ctx['late_queries_tab_count'] = LateQuery.objects.filter(status='pending').count() if is_hr else 0
        return ctx

    def post(self, request, *args, **kwargs):
        from django.http import Http404
        from hr.attendance_exception_services import decide_attendance_exception, override_attendance_exception
        exc = AttendanceException.objects.filter(pk=request.POST.get('exc_id')).first()
        action = request.POST.get('action')

        if action == 'decide':
            # Checks the LIVE exc.employee.main_manager relationship, not the
            # exc.main_manager submission-time snapshot — see
            # can_decide_attendance_exception's docstring for why.
            authorized = (
                exc is not None and exc.employee.main_manager and exc.employee.main_manager.user_id == request.user.id
                and exc.employee.user_id != request.user.id)
            if not authorized:
                # Collapse "doesn't exist" and "exists but not yours" into one
                # 404 — same pk-existence-oracle fix already applied to the
                # leave workflow's document download view.
                raise Http404('No such attendance exception.')
            try:
                decide_attendance_exception(
                    exc, request.user, request.POST.get('decision'),
                    # Field named 'comment', matching the leave-request
                    # workflow's optional decision-comment convention
                    # (templates/hr/leave_request_detail.html).
                    note=request.POST.get('comment', ''))
                messages.success(request, 'Decision recorded.')
            except ValueError as e:
                messages.error(request, str(e))

        elif action == 'override':
            authorized = exc is not None and can_decide_attendance_exception(request.user, exc)
            if not authorized:
                raise Http404('No such attendance exception.')
            try:
                override_attendance_exception(
                    exc, request.user, request.POST.get('decision'),
                    request.POST.get('reason', ''))
                messages.success(request, 'Request finalized via override.')
            except ValueError as e:
                messages.error(request, str(e))

        elif action == 'revoke_direct':
            if not has_override_access(request.user):
                return HttpResponse('You do not have override access to revoke this request.', status=403)
            if exc is None:
                raise Http404('No such attendance exception.')
            from hr.attendance_exception_services import revoke_attendance_exception
            try:
                revoke_attendance_exception(exc, request.user, request.POST.get('reason', ''))
                messages.success(request, 'Attendance exception revoked.')
            except ValueError as e:
                messages.error(request, str(e))

        elif action == 'decide_revoke_request':
            from hr.attendance_exception_services import decide_attendance_exception_revoke_request
            revoke_req = AttendanceExceptionRevokeRequest.objects.filter(
                pk=request.POST.get('revoke_request_id')).first()
            if revoke_req is None:
                raise Http404('No such revoke request.')
            try:
                decide_attendance_exception_revoke_request(
                    revoke_req, request.user, request.POST.get('decision'),
                    decision_note=request.POST.get('decision_note', ''))
                messages.success(request, 'Revoke request decided.')
            except ValueError as e:
                messages.error(request, str(e))

        elif action == 'decide_late_query':
            from hr.models import LateQuery
            if not (request.user.is_super_admin_user or request.user.is_admin_user or request.user.is_erp_admin_user):
                raise Http404('No such query.')
            query_id = request.POST.get('query_id')
            if not query_id or not str(query_id).isdigit():
                raise Http404('No such query.')
            query = LateQuery.objects.filter(pk=query_id, status='pending').first()
            if query is None:
                raise Http404('No such query.')
            decision = request.POST.get('decision')
            if decision not in ('approved', 'rejected'):
                messages.error(request, 'Invalid decision.')
            else:
                query.status = decision
                query.decision_note = request.POST.get('comment', '')
                query.decided_by = request.user
                query.decided_at = timezone.now()
                query.save()
                if decision == 'approved':
                    from hr.attendance_services import regenerate_attendance_record
                    regenerate_attendance_record(query.employee, query.attendance_record.date)
                if query.employee.user_id:
                    from notifications.services import create_notification
                    create_notification(
                        recipient=query.employee.user,
                        verb=f'Your late-mark query for {query.attendance_record.date:%d %b %Y} was {decision}',
                        actor=request.user, description=query.decision_note, level='info',
                        target_url=reverse('hr:my_profile'),
                    )
                messages.success(request, 'Query decision recorded.')
        # Preserve whatever tab the manager was viewing when they acted.
        tab = request.POST.get('tab', '')
        url = reverse('hr:team_exceptions')
        if tab:
            url = f'{url}?tab={tab}'
        return redirect(url)


class LeaveAccessView(SuperAdminRequiredMixin, TemplateView):
    """Super-Admin-only page (linked from the sidebar's Leave section) for
    the Leave Request workflow's permission surface: who can view the Leave
    Request dashboard and approve requests (LeaveDashboardAccess), who can
    force-approve/disapprove when the real approver is unavailable
    (OverrideAccessSettings/Role/Employee), and the balance-hold settings
    for over-balance submissions. Previously the Org Chart page's second
    tab — Org Chart itself is purely about the reporting hierarchy now."""
    template_name = 'hr/leave_access.html'

    def get_context_data(self, **kwargs):
        from django.db.models import Q
        from accounts.models import User, Role
        from .models import LeaveDashboardAccess, OverrideAccessSettings, OverrideAccessRole, OverrideAccessEmployee
        ctx = super().get_context_data(**kwargs)

        ctx['access_grants'] = (
            LeaveDashboardAccess.objects.filter(is_active=True)
            .select_related('user', 'granted_by').order_by('user__username'))

        access_search = (self.request.GET.get('access_search') or '').strip()
        ctx['access_search'] = access_search
        access_results = None
        if access_search:
            granted_user_ids = set(
                LeaveDashboardAccess.objects.filter(is_active=True).values_list('user_id', flat=True))
            access_results = (
                Employee.objects.filter(is_active=True, user__isnull=False)
                .filter(
                    Q(full_name__icontains=access_search)
                    | Q(user__username__icontains=access_search)
                    | Q(user__email__icontains=access_search))
                .exclude(user_id__in=granted_user_ids)
                .select_related('user').order_by('full_name')[:25])
        ctx['access_results'] = access_results

        # ── Override Access (who can use the leave-request override escape
        # hatch) — lives on this same page since it's specifically about
        # Leave Request access, not a general-purpose setting. ──
        override_settings = OverrideAccessSettings.get_solo()
        ctx['override_settings'] = override_settings
        ctx['override_mode_choices'] = OverrideAccessSettings.MODE_CHOICES
        ctx['current_super_admins'] = User.objects.filter(
            role__name='super_admin', is_active=True).order_by('username')
        ctx['override_roles'] = Role.objects.all().order_by('name')
        ctx['override_role_ids'] = set(
            OverrideAccessRole.objects.values_list('role_id', flat=True))
        ctx['override_employees'] = (
            OverrideAccessEmployee.objects.select_related('user', 'added_by').order_by('user__username'))

        override_emp_search = (self.request.GET.get('override_emp_search') or '').strip()
        ctx['override_emp_search'] = override_emp_search
        override_emp_results = None
        if override_emp_search:
            granted_override_user_ids = set(
                OverrideAccessEmployee.objects.values_list('user_id', flat=True))
            override_emp_results = (
                Employee.objects.filter(is_active=True, user__isnull=False)
                .filter(
                    Q(full_name__icontains=override_emp_search)
                    | Q(user__username__icontains=override_emp_search)
                    | Q(user__email__icontains=override_emp_search))
                .exclude(user_id__in=granted_override_user_ids)
                .select_related('user').order_by('full_name')[:25])
        ctx['override_emp_results'] = override_emp_results
        return ctx

    def post(self, request, *args, **kwargs):
        from accounts.models import Role
        from .models import LeaveDashboardAccess, OverrideAccessSettings, OverrideAccessRole, OverrideAccessEmployee
        access_url = reverse('hr:leave_access')

        grant_employee_id = request.POST.get('grant_dashboard_access')
        if grant_employee_id:
            grant_employee = get_object_or_404(Employee, pk=grant_employee_id)
            if not grant_employee.user_id:
                messages.error(request, f'{grant_employee.full_name} has no linked user account.')
            else:
                LeaveDashboardAccess.objects.update_or_create(
                    user=grant_employee.user,
                    defaults={'is_active': True, 'granted_by': request.user})
                messages.success(
                    request,
                    f'Granted {grant_employee.full_name} access to the Leave Request dashboard '
                    'and approval authority on new leave requests.')
            return redirect(access_url)

        revoke_user_id = request.POST.get('revoke_dashboard_access')
        if revoke_user_id:
            grant = get_object_or_404(LeaveDashboardAccess, user_id=revoke_user_id, is_active=True)
            display_name = grant.user.get_full_name() or grant.user.username
            grant.delete()
            messages.success(
                request,
                f'Revoked {display_name}\'s Leave Request dashboard access and approval authority.')
            return redirect(access_url)

        override_mode = request.POST.get('set_override_mode')
        if override_mode:
            if override_mode not in dict(OverrideAccessSettings.MODE_CHOICES):
                messages.error(request, 'Not a valid override access mode.')
                return redirect(access_url)
            config = OverrideAccessSettings.get_solo()
            config.mode = override_mode
            config.updated_by = request.user
            config.save(update_fields=['mode', 'updated_by', 'updated_at'])
            messages.success(
                request, f'Override access mode set to "{config.get_mode_display()}".')
            return redirect(access_url)

        if 'set_balance_hold' in request.POST:
            config = OverrideAccessSettings.get_solo()
            config.allow_site_balance_hold = bool(request.POST.get('allow_site_balance_hold'))
            config.allow_office_balance_hold = bool(request.POST.get('allow_office_balance_hold'))
            config.updated_by = request.user
            config.save(update_fields=['allow_site_balance_hold', 'allow_office_balance_hold', 'updated_by', 'updated_at'])
            messages.success(request, 'Balance-hold settings updated.')
            return redirect(access_url)

        toggle_role_id = request.POST.get('toggle_override_role')
        if toggle_role_id:
            role = get_object_or_404(Role, pk=toggle_role_id)
            existing = OverrideAccessRole.objects.filter(role=role).first()
            if existing:
                existing.delete()
                messages.success(request, f'Removed override access for the {role.get_name_display()} role.')
            else:
                OverrideAccessRole.objects.create(role=role)
                messages.success(request, f'Granted override access to the {role.get_name_display()} role.')
            return redirect(access_url)

        grant_override_employee_id = request.POST.get('grant_override_employee')
        if grant_override_employee_id:
            grant_employee = get_object_or_404(Employee, pk=grant_override_employee_id)
            if not grant_employee.user_id:
                messages.error(request, f'{grant_employee.full_name} has no linked user account.')
            else:
                OverrideAccessEmployee.objects.update_or_create(
                    user=grant_employee.user, defaults={'added_by': request.user})
                messages.success(request, f'Granted {grant_employee.full_name} override access.')
            return redirect(access_url)

        revoke_override_employee_id = request.POST.get('revoke_override_employee')
        if revoke_override_employee_id:
            grant = get_object_or_404(OverrideAccessEmployee, user_id=revoke_override_employee_id)
            display_name = grant.user.get_full_name() or grant.user.username
            grant.delete()
            messages.success(request, f'Revoked {display_name}\'s override access.')
            return redirect(access_url)

        return redirect(access_url)


class OrgChartView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """Super-Admin-only page: assign each employee's main_manager/secondary_managers,
    and view the resulting hierarchy tree. Uses a per-row EmployeeHierarchyForm
    (see hr/forms.py) rather than one giant multi-employee submit, so a mistake
    on one row can never affect any other row's data."""
    model = Employee
    template_name = 'hr/org_chart.html'
    context_object_name = 'employees'

    def test_func(self):
        u = self.request.user
        # Super Admin + ERP Admin get the management page; team-scoped roles get
        # a read-only tree of their own subtree (enforced by _can_manage below +
        # scoping in get_queryset/get_context_data and blocked in post()).
        return bool(u.is_super_admin_user or u.is_erp_admin_user or u.is_team_scoped_hr_user)

    def _can_manage(self):
        """Who may edit the hierarchy (assignment table + save). Company-wide
        admin tiers only — team-scoped roles get a read-only view."""
        u = self.request.user
        return bool(u.is_super_admin_user or u.is_erp_admin_user)

    def get_queryset(self):
        qs = Employee.objects.filter(is_active=True).select_related('main_manager').order_by('full_name')
        # Team-scoped roles only ever see their own reports in the table.
        qs = scope_employee_queryset(self.request.user, qs)
        search = (self.request.GET.get('search') or '').strip()
        if search:
            qs = qs.filter(full_name__icontains=search)
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['search'] = self.request.GET.get('search', '')

        forms_by_employee = {}
        for emp in ctx['employees']:
            row_form = EmployeeHierarchyForm(
                employee=emp,
                initial={'main_manager': emp.main_manager_id,
                        'secondary_managers': list(emp.secondary_managers.values_list('pk', flat=True))})
            # Each row's widgets live outside their own <form> element (a
            # <form> can't validly wrap <td>s that belong to a shared <tr>/
            # <table> structure), so they're associated with their row's
            # <form id="emp-form-{pk}"> purely via the HTML `form` attribute
            # — standards-compliant, no JS required.
            form_id = f'emp-form-{emp.pk}'
            row_form.fields['main_manager'].widget.attrs['form'] = form_id
            row_form.fields['secondary_managers'].widget.attrs['form'] = form_id
            forms_by_employee[emp.pk] = row_form
        ctx['forms_by_employee'] = forms_by_employee
        # (employee, form) pairs in display order — simpler for the template
        # to iterate than doing a dict lookup by pk per row.
        ctx['employee_rows'] = [(emp, forms_by_employee[emp.pk]) for emp in ctx['employees']]
        # Root nodes (no main_manager) for the hierarchy tree — deliberately
        # independent of the search filter above, so the tree at the bottom
        # of the page always shows the whole org, even while searching the
        # assignment table above it. The template recursively walks
        # main_reports.all() from each root via _org_chart_node.html.
        if self._can_manage():
            ctx['root_employees'] = Employee.objects.filter(
                is_active=True, main_manager__isnull=True).order_by('full_name')
        else:
            # Team-scoped role: root the tree at themselves, so it shows only
            # their own subtree instead of every company-wide root node.
            emp = getattr(self.request.user, 'employee_profile', None)
            ctx['root_employees'] = (
                Employee.objects.filter(pk=emp.pk, is_active=True) if emp
                else Employee.objects.none())
        ctx['can_manage_org'] = self._can_manage()
        return ctx

    def post(self, request, *args, **kwargs):
        # Team-scoped roles have a read-only org chart — no mutations at all.
        if not self._can_manage():
            raise PermissionDenied

        employee = get_object_or_404(Employee, pk=request.POST.get('employee_id'))
        # Every successful action redirects with a `#row-<pk>` fragment so
        # page-load JS can flash/highlight the specific row that changed —
        # there's no clean way to target a single row from a generic
        # messages-framework flash after a full-page POST-redirect-GET, so
        # this fragment + a transient CSS class is the row-level acknowledgment,
        # on top of the always-visible persistent per-row indicators.
        row_redirect_url = f"{reverse('hr:org_chart')}#row-{employee.pk}"

        form = EmployeeHierarchyForm(request.POST, employee=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f'Updated hierarchy assignment for {employee.full_name}.')
            return redirect(row_redirect_url)
        else:
            for field, errors in form.errors.items():
                for e in errors:
                    messages.error(request, f'{employee.full_name}: {e}')
        return redirect('hr:org_chart')


@login_required
def asset_handover_create(request):
    """HR creates and signs a new handover in one step."""
    from hr.models import Asset, Vehicle, Employee
    from hr.asset_handover_services import create_asset_handover, get_default_authorizer
    from procurement.views import _read_and_validate_signature
    from accounts.models import User, Role
    from django.core.files.base import ContentFile

    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have access to create asset handovers.')
        return redirect('hr:hr_dashboard')

    if request.method == 'POST':
        item_kind = request.POST.get('item_kind')
        item_id = request.POST.get('item_id')
        employee = Employee.objects.filter(pk=request.POST.get('employee_id')).first()
        authorizer = User.objects.filter(pk=request.POST.get('authorizer_id'), role__name=Role.SUPER_ADMIN).first()
        item = None
        if item_kind == 'asset':
            item = Asset.objects.filter(pk=item_id).first()
        elif item_kind == 'vehicle':
            item = Vehicle.objects.filter(pk=item_id).first()

        if not (item and employee and authorizer):
            messages.error(request, 'Please select a valid item, employee, and authorizer.')
            return redirect('hr:asset_handover_create')

        png_bytes, error_response = _read_and_validate_signature(request)
        if error_response is not None:
            messages.error(request, 'Signature could not be saved. Please try again.')
            return redirect('hr:asset_handover_create')

        try:
            handover = create_asset_handover(
                employee=employee, item=item, issued_by=request.user,
                issued_signature=ContentFile(png_bytes, name='issued.png'),
                authorizer=authorizer,
                accessories=request.POST.get('accessories', ''),
                software_installed=request.POST.get('software_installed', ''),
                remember_authorizer=request.POST.get('remember_authorizer') == 'on',
            )
        except ValueError as e:
            messages.error(request, str(e))
            return redirect('hr:asset_handover_create')
        messages.success(request, 'Handover created and sent for authorization.')
        return redirect('hr:asset_handover_detail', pk=handover.pk)
    preselected_item_kind = request.GET.get('item_kind', 'asset')
    preselected_item_id = request.GET.get('item_id', '')
    default_accessories = ''
    default_software_installed = ''
    if preselected_item_kind == 'asset' and preselected_item_id:
        preselected_asset = Asset.objects.filter(pk=preselected_item_id).first()
        if preselected_asset:
            default_accessories = preselected_asset.accessories
            default_software_installed = preselected_asset.software_installed


    context = {
        'assets': Asset.objects.filter(is_decommissioned=False).order_by('asset_name'),
        'vehicles': Vehicle.objects.all().order_by('plate_number'),
        'employees': Employee.objects.all().order_by('full_name'),
        'super_admins': User.objects.filter(is_active=True, role__name=Role.SUPER_ADMIN).order_by('username'),
        'default_authorizer': get_default_authorizer(request.user),
        'preselected_item_kind': preselected_item_kind,
        'preselected_item_id': preselected_item_id,
        'default_accessories': default_accessories,
        'default_software_installed': default_software_installed,
    }
    return render(request, 'hr/asset_handover_create.html', context)


@login_required
def asset_handover_detail(request, pk):
    from hr.models import AssetHandover
    handover = get_object_or_404(AssetHandover, pk=pk)
    is_hr = bool(request.user.is_super_admin_user or request.user.is_erp_admin_user)
    is_employee = bool(handover.employee.user_id and handover.employee.user_id == request.user.id)
    is_authorizer = bool(handover.authorizer_id and handover.authorizer_id == request.user.id)
    if not (is_hr or is_employee or is_authorizer):
        messages.error(request, 'You do not have access to this handover.')
        return redirect('hr:hr_dashboard')
    context = {
        'handover': handover,
        'can_authorize': is_authorizer and handover.status == 'pending_authorization',
        'can_receive': is_employee and handover.status == 'pending_receipt',
        'can_initiate_return': is_hr and handover.status == 'active',
        'can_acknowledge_return': is_employee and handover.status == 'pending_return_confirmation',
    }
    return render(request, 'hr/asset_handover_detail.html', context)

@login_required
@require_POST
def asset_handover_authorize(request, pk):
    from hr.models import AssetHandover
    from hr.asset_handover_services import authorize_handover
    from procurement.views import _read_and_validate_signature
    from django.core.files.base import ContentFile

    handover = get_object_or_404(AssetHandover, pk=pk)
    if handover.authorizer_id != request.user.id or not request.user.is_super_admin_user:
        messages.error(request, 'You are not the assigned authorizer for this handover.')
        return redirect('hr:asset_handover_detail', pk=pk)

    png_bytes, error_response = _read_and_validate_signature(request)
    if error_response is not None:
        messages.error(request, 'Signature could not be saved. Please try again.')
        return redirect('hr:asset_handover_detail', pk=pk)

    try:
        authorize_handover(handover, request.user, ContentFile(png_bytes, name='authorized.png'))
        messages.success(request, 'Handover authorized and signed.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('hr:asset_handover_detail', pk=pk)


@login_required
@require_POST
def asset_handover_receive(request, pk):
    from hr.models import AssetHandover
    from hr.asset_handover_services import receive_handover
    from procurement.views import _read_and_validate_signature
    from django.core.files.base import ContentFile

    handover = get_object_or_404(AssetHandover, pk=pk)
    if not (handover.employee.user_id and handover.employee.user_id == request.user.id):
        messages.error(request, 'You are not the recipient of this handover.')
        return redirect('hr:asset_handover_detail', pk=pk)

    png_bytes, error_response = _read_and_validate_signature(request)
    if error_response is not None:
        messages.error(request, 'Signature could not be saved. Please try again.')
        return redirect('hr:asset_handover_detail', pk=pk)

    try:
        receive_handover(handover, ContentFile(png_bytes, name='received.png'))
        messages.success(request, 'Handover received and signed. The asset is now registered to you.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('hr:asset_handover_detail', pk=pk)


@login_required
@require_POST
def asset_handover_initiate_return(request, pk):
    from hr.models import AssetHandover
    from hr.asset_handover_services import initiate_return
    from procurement.views import _read_and_validate_signature
    from django.core.files.base import ContentFile

    handover = get_object_or_404(AssetHandover, pk=pk)
    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have access to start a return.')
        return redirect('hr:asset_handover_detail', pk=pk)

    png_bytes, error_response = _read_and_validate_signature(request)
    if error_response is not None:
        messages.error(request, 'Signature could not be saved. Please try again.')
        return redirect('hr:asset_handover_detail', pk=pk)

    try:
        initiate_return(handover, request.user, ContentFile(png_bytes, name='return_received.png'),
                         remarks=request.POST.get('return_remarks', ''))
        messages.success(request, 'Return request sent. Waiting on the employee to confirm.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('hr:asset_handover_detail', pk=pk)


@login_required
@require_POST
def asset_handover_acknowledge_return(request, pk):
    from hr.models import AssetHandover
    from hr.asset_handover_services import acknowledge_return
    from procurement.views import _read_and_validate_signature
    from django.core.files.base import ContentFile

    handover = get_object_or_404(AssetHandover, pk=pk)
    if not (handover.employee.user_id and handover.employee.user_id == request.user.id):
        messages.error(request, 'You are not the current holder of this item.')
        return redirect('hr:asset_handover_detail', pk=pk)

    png_bytes, error_response = _read_and_validate_signature(request)
    if error_response is not None:
        messages.error(request, 'Signature could not be saved. Please try again.')
        return redirect('hr:asset_handover_detail', pk=pk)

    try:
        acknowledge_return(handover, ContentFile(png_bytes, name='returned.png'))
        messages.success(request, 'Return confirmed. Custody released.')
    except ValueError as e:
        messages.error(request, str(e))
    return redirect('hr:asset_handover_detail', pk=pk)


@login_required
def asset_handover_list(request):
    """HR dashboard: every asset/vehicle's current owner and, if a new
    handover is already in progress for it, the next owner too."""
    from hr.models import AssetHandover

    if not (request.user.is_super_admin_user or request.user.is_erp_admin_user):
        messages.error(request, 'You do not have access to this page.')
        return redirect('hr:hr_dashboard')

    active = AssetHandover.objects.filter(status='active').select_related(
        'asset', 'vehicle', 'employee')
    in_progress = AssetHandover.objects.filter(
        status__in=('pending_authorization', 'pending_receipt', 'pending_return_confirmation')
    ).select_related('asset', 'vehicle', 'employee')

    rows = {}
    for h in active:
        key = ('asset', h.asset_id) if h.asset_id else ('vehicle', h.vehicle_id)
        rows[key] = {'item': h.item, 'current_owner': h.employee, 'next_owner': None, 'status': 'Active', 'handover': h}
    for h in in_progress:
        key = ('asset', h.asset_id) if h.asset_id else ('vehicle', h.vehicle_id)
        if key in rows:
            rows[key]['next_owner'] = h.employee
            rows[key]['status'] = 'Transfer pending' if h.status != 'pending_return_confirmation' else 'Return pending'
        else:
            rows[key] = {'item': h.item, 'current_owner': None, 'next_owner': h.employee,
                          'status': 'Pending', 'handover': h}

    context = {'rows': list(rows.values())}
    return render(request, 'hr/asset_handover_list.html', context)


@login_required
def asset_active_handover_redirect(request, item_id):
    from hr.models import AssetHandover, Asset, Vehicle
    item_kind = request.GET.get('item_kind', 'asset')
    if item_kind == 'vehicle':
        item = Vehicle.objects.filter(pk=item_id).first()
        filter_kwargs = {'vehicle_id': item_id}
        fallback_url, fallback_list_url = 'hr:vehicle_detail', 'hr:vehicle_list'
    else:
        item = Asset.objects.filter(pk=item_id).first()
        filter_kwargs = {'asset_id': item_id}
        fallback_url, fallback_list_url = 'hr:asset_detail', 'hr:asset_list'

    # item_kind is caller-supplied (query string) and item_id is a raw pk -
    # if they disagree with reality (e.g. item_kind=vehicle for what's
    # actually an Asset's pk), redirecting straight to fallback_url with
    # that pk could 404 on an unrelated/nonexistent record instead of
    # showing a sensible error. Confirm the item actually exists as the
    # claimed kind first.
    if item is None:
        messages.error(request, 'That item could not be found.')
        return redirect(fallback_list_url)

    handover = AssetHandover.objects.filter(
        status='active', **filter_kwargs).order_by('-created_at').first()
    if handover is None:
        messages.error(request, 'No active handover found.')
        return redirect(fallback_url, pk=item_id)
    return redirect('hr:asset_handover_detail', pk=handover.pk)


@login_required
def asset_handover_export_pdf(request, pk):
    from hr.models import AssetHandover
    handover = get_object_or_404(AssetHandover, pk=pk)
    is_hr = bool(request.user.is_super_admin_user or request.user.is_erp_admin_user)
    is_employee = bool(handover.employee.user_id and handover.employee.user_id == request.user.id)
    is_authorizer = bool(handover.authorizer_id and handover.authorizer_id == request.user.id)
    if not (is_hr or is_employee or is_authorizer):
        messages.error(request, 'You do not have access to this handover.')
        return redirect('hr:hr_dashboard')

    import io
    from io import BytesIO as _BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
                                     Image as RLImage)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from xml.sax.saxutils import escape as _xml_escape
    from django.contrib.staticfiles.finders import find as find_static

    BRAND_RED = colors.HexColor('#C41E3A')

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=18)
    styles = getSampleStyleSheet()
    label_style = ParagraphStyle('label', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#555555'))
    value_style = ParagraphStyle('value', parent=styles['Normal'], fontSize=10)
    section_style = ParagraphStyle('section', parent=styles['Normal'], fontSize=11, spaceBefore=6, spaceAfter=4, alignment=1)
    note_style = ParagraphStyle('note', parent=styles['Normal'], fontSize=7.5, textColor=colors.HexColor('#555555'))

    elements = []

    # ── Header: title/form-number on the left, company logo on the right ──
    logo_path = find_static('images/leap_logo.jpg')
    if logo_path:
        elements.append(RLImage(logo_path, width=55*mm, height=16.6*mm, hAlign='CENTER'))
        elements.append(Spacer(1, 3*mm))
    title_style_center = ParagraphStyle('titleCenter', parent=styles['Title'], alignment=1)
    label_style_center = ParagraphStyle('labelCenter', parent=label_style, alignment=1)
    elements.append(Paragraph('Asset Handover Form', title_style_center))
    elements.append(Paragraph('LNA_ADM/007', label_style_center))
    elements.append(Spacer(1, 4*mm))

    item = handover.item
    if handover.asset_id:
        item_rows = [
            ['Item Type', item.asset_type or '-', 'Model', item.model or '-'],
            ['Serial No.', item.serial_number or '-', 'Part Number', item.part_number or '-'],
            ['Tag Number', item.tag_number or '-', 'Quantity', str(item.quantity)],
            ['Item Description', item.item_description or '-', '', ''],
        ]
    else:
        item_rows = [
            ['Plate Number', item.plate_number, 'Vehicle Model', item.vehicle_model or '-'],
            ['Maker', item.vehicle_maker or '-', '', ''],
        ]
    item_rows.append(['Assigned To', handover.employee.full_name, 'Designation', handover.employee.designation or '-'])

    item_table = Table(item_rows, colWidths=[35*mm, 55*mm, 35*mm, 55*mm])
    item_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(item_table)
    elements.append(Spacer(1, 4*mm))

    if handover.accessories:
        elements.append(Paragraph(f'<b>Accessories:</b> {_xml_escape(handover.accessories)}', value_style))
    if handover.software_installed:
        elements.append(Paragraph(f'<b>Software Installed:</b> {_xml_escape(handover.software_installed)}', value_style))
    elements.append(Spacer(1, 4*mm))

    def _sig_image(field, max_w=45*mm, max_h=16*mm):
        if not field:
            return ''
        try:
            field.open('rb')
            data = field.read()
            field.close()
            return RLImage(_BytesIO(data), width=max_w, height=max_h, kind='proportional')
        except Exception:
            return ''

    elements.append(Paragraph('<u><b>HANDOVER</b></u>', section_style))
    handover_rows = [
        ['Issued By', 'Authorized By', 'Received By'],
        [
            handover.issued_by.get_full_name() or handover.issued_by.username if handover.issued_by_id else '-',
            handover.authorizer.get_full_name() or handover.authorizer.username if handover.authorizer_id else '-',
            handover.employee.full_name,
        ],
        [
            handover.issued_at.strftime('%d %b %Y') if handover.issued_at else '-',
            handover.authorized_at.strftime('%d %b %Y') if handover.authorized_at else '-',
            handover.received_at.strftime('%d %b %Y') if handover.received_at else '-',
        ],
        [
            _sig_image(handover.issued_signature),
            _sig_image(handover.authorized_signature),
            _sig_image(handover.received_signature),
        ],
    ]
    handover_table = Table(handover_rows, colWidths=[60*mm, 60*mm, 60*mm], rowHeights=[None, None, None, 20*mm])
    handover_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), BRAND_RED),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(handover_table)

    if handover.status in ('pending_return_confirmation', 'returned'):
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph('<u><b>ASSET RETURN</b></u>', section_style))
        if handover.return_remarks:
            elements.append(Paragraph(
                f'<b>Remarks upon Return:</b> {_xml_escape(handover.return_remarks)}', value_style))
            elements.append(Spacer(1, 2*mm))
        return_rows = [
            ['Returned By', 'Received By'],
            [
                handover.employee.full_name,
                (handover.return_received_by.get_full_name() or handover.return_received_by.username)
                if handover.return_received_by_id else '-',
            ],
            [
                handover.returned_at.strftime('%d %b %Y') if handover.returned_at else '-',
                handover.return_received_at.strftime('%d %b %Y') if handover.return_received_at else '-',
            ],
            [
                _sig_image(handover.returned_signature),
                _sig_image(handover.return_received_signature),
            ],
        ]
        return_table = Table(return_rows, colWidths=[90*mm, 90*mm], rowHeights=[None, None, None, 20*mm])
        return_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), BRAND_RED),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(return_table)

    # ── Footer note ──
    elements.append(Spacer(1, 8*mm))
    note_text = (
        "NOTE: Please be aware that it is your responsibility to take good care of the company's "
        "property whilst you are using it, whether on your premises or outstationed. This particularly "
        "applies to engineers taking Notebook and Laptops on service assignments.<br/>"
        "If any of the company's property is damaged due to negligence, misuse, mishandling, etc and "
        "requires repair or replacement, the company reserves the right to charge all or part of the "
        "cost to the concerned individual. You are not allowed to install or change or delete any of "
        "the software in the computer. If however the need arises, please refer to the concerned person."
    )
    elements.append(Paragraph(note_text, note_style))

    doc.build(elements)
    buffer.seek(0)
    filename = f'asset_handover_{handover.pk}.pdf'
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
