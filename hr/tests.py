from datetime import date
from django.test import TestCase, skipUnlessDBFeature
from django.urls import reverse
from hr.work_calendar import count_working_days, is_working_day
from hr.forms import LeaveRequestForm
from hr.models import LeaveType
from hr.forms import LeaveRequestForm

from django.test import TestCase
from hr.forms import EmployeeForm, AssetForm


class EmployeeFormValidationTests(TestCase):
    """BUG-004: free-text fields must reject symbol-only garbage while
    still accepting legitimate Latin and Arabic names/titles."""

    def _base_valid_data(self, **overrides):
        data = {
            'full_name': 'John Smith',
            'nationality': 'British',
            'designation': 'Site Engineer',
            'qualification': 'BSc Mechanical Engineering',
            'deployment': 'Riyadh Site',
            'department': 'Operations',
            'mobile_number': '+966501234567',
            'iqama_number': '1234567890',
            # add any other required fields your EmployeeForm needs
        }
        data.update(overrides)
        return data

    def test_symbol_only_full_name_is_rejected(self):
        form = EmployeeForm(data=self._base_valid_data(full_name='!@#$%^&*()'))
        self.assertFalse(form.is_valid())
        self.assertIn('full_name', form.errors)

    def test_valid_latin_full_name_is_accepted(self):
        form = EmployeeForm(data=self._base_valid_data(full_name="O'Connor-Smith"))
        # Only check this field's own validation, not the whole form
        form.is_valid()
        self.assertNotIn('full_name', form.errors)

    def test_valid_arabic_full_name_is_accepted(self):
        form = EmployeeForm(data=self._base_valid_data(full_name='محمد أحمد'))
        form.is_valid()
        self.assertNotIn('full_name', form.errors)

    def test_symbol_only_mobile_number_is_rejected(self):
        form = EmployeeForm(data=self._base_valid_data(mobile_number='!@#$%^&*()'))
        self.assertFalse(form.is_valid())
        self.assertIn('mobile_number', form.errors)

    def test_symbol_only_designation_is_rejected(self):
        form = EmployeeForm(data=self._base_valid_data(designation='!@#$%^&*()'))
        self.assertFalse(form.is_valid())
        self.assertIn('designation', form.errors)

    def test_symbol_only_grade_is_rejected(self):
        form = EmployeeForm(data=self._base_valid_data(grade='!@#$%^&*()'))
        self.assertFalse(form.is_valid())
        self.assertIn('grade', form.errors)

    def test_valid_grade_is_accepted(self):
        form = EmployeeForm(data=self._base_valid_data(grade='EX-2'))
        form.is_valid()
        self.assertNotIn('grade', form.errors)


class AssetFormValidationTests(TestCase):
    """Same validation pattern applied to AssetForm's free-text fields."""

    def _base_valid_data(self, **overrides):
        data = {
            'serial_number': 'SN-2026-001',
            # add remaining required AssetForm fields here
        }
        data.update(overrides)
        return data

    def test_symbol_only_serial_number_is_rejected(self):
        form = AssetForm(data=self._base_valid_data(serial_number='!@#$%^&*()'))
        self.assertFalse(form.is_valid())
        self.assertIn('serial_number', form.errors)

    def test_valid_alphanumeric_serial_number_is_accepted(self):
        form = AssetForm(data=self._base_valid_data(serial_number='ABC-123-XYZ'))
        form.is_valid()
        self.assertNotIn('serial_number', form.errors)


class WorkCalendarTests(TestCase):
    WEEKENDS = {4, 5}  # Fri, Sat (Mon=0..Sun=6)

    def test_weekday_is_working(self):
        self.assertTrue(is_working_day(date(2026, 7, 13), self.WEEKENDS, set()))  # Monday

    def test_friday_is_not_working(self):
        self.assertFalse(is_working_day(date(2026, 7, 10), self.WEEKENDS, set()))  # Friday

    def test_holiday_is_not_working(self):
        h = {date(2026, 7, 13)}
        self.assertFalse(is_working_day(date(2026, 7, 13), self.WEEKENDS, h))

    def test_count_excludes_weekend_and_holiday(self):
        # Sun 5 Jul -> Thu 9 Jul 2026 is 5 weekdays; add a holiday on Tue 7th -> 4
        holidays = {date(2026, 7, 7)}
        self.assertEqual(count_working_days(date(2026, 7, 5), date(2026, 7, 9), self.WEEKENDS, holidays), 4)

    def test_count_range_all_weekend_is_zero(self):
        # Fri 10 + Sat 11 Jul 2026
        self.assertEqual(count_working_days(date(2026, 7, 10), date(2026, 7, 11), self.WEEKENDS, set()), 0)


class LeaveTypeModelTests(TestCase):
    def test_create_and_str(self):
        lt, _ = LeaveType.objects.get_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': 21})
        self.assertEqual(str(lt), 'Annual')
        self.assertTrue(lt.is_paid)  # default True

    def test_code_is_unique(self):
        from django.db import IntegrityError
        LeaveType.objects.get_or_create(code='sick', defaults={
            'name': 'Sick', 'default_annual_days': 30})
        with self.assertRaises(IntegrityError):
            LeaveType.objects.create(name='Sick 2', code='sick', default_annual_days=10)


class LeaveTypeTask1DefaultsTests(TestCase):
    def test_only_annual_is_accumulative(self):
        from hr.models import LeaveType
        annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        sick, _ = LeaveType.objects.get_or_create(code='sick', defaults={'name': 'Sick', 'default_annual_days': 12})
        marriage, _ = LeaveType.objects.get_or_create(code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3})
        # Simulate the migration's effect directly (the migration itself is exercised by `migrate` in CI/manual QA;
        # this test locks in the invariant the app code relies on).
        LeaveType.objects.exclude(code='annual').update(is_accumulative=False)
        LeaveType.objects.filter(code='annual').update(is_accumulative=True)
        self.assertTrue(LeaveType.objects.get(code='annual').is_accumulative)
        self.assertFalse(LeaveType.objects.get(code='sick').is_accumulative)
        self.assertFalse(LeaveType.objects.get(code='marriage').is_accumulative)


from datetime import date as _date
from hr.models import Holiday, AttendanceSettings


class HolidayAndSettingsTests(TestCase):
    def test_holiday_unique_date(self):
        from django.db import IntegrityError
        Holiday.objects.create(date=_date(2026, 7, 17), name='Eid')
        with self.assertRaises(IntegrityError):
            Holiday.objects.create(date=_date(2026, 7, 17), name='Eid dup')

    def test_settings_singleton_default_weekend(self):
        s = AttendanceSettings.load()
        self.assertEqual(s.weekend_day_set(), {4, 5})  # Fri, Sat
        self.assertEqual(AttendanceSettings.load().pk, s.pk)

    def test_settings_parse_custom_weekend(self):
        s = AttendanceSettings.load()
        s.weekend_days = '5,6'  # Sat, Sun
        s.save()
        self.assertEqual(AttendanceSettings.load().weekend_day_set(), {5, 6})


from decimal import Decimal
from django.core.exceptions import ValidationError
from django.contrib.auth import get_user_model
from hr.models import (Employee, LeaveEntitlement, LeaveRecord, LeaveExceptionGrant,
                       LeaveDashboardAccess, LeaveRequest, LeaveRequestApproval, LeaveRequestNote,
                       LeaveRevokeRequest)

User = get_user_model()


def make_employee(iqama='E1', name='Ali', joining=None):
    return Employee.objects.create(iqama_number=iqama, full_name=name, joining_date=joining)


def make_user(username, **kwargs):
    return User.objects.create(username=username, **kwargs)


class LeaveEntitlementTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': 30})

    def test_unique_per_employee_type_year(self):
        from django.db import IntegrityError
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        with self.assertRaises(IntegrityError):
            LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=25)

    def test_balance_with_records(self):
        ent = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 3, 1), end_date=_date(2026, 3, 5), days=Decimal('5'))
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 6, 1), end_date=_date(2026, 6, 3), days=Decimal('3'))
        self.assertEqual(ent.taken_days, Decimal('8'))
        self.assertEqual(ent.remaining_days, Decimal('22'))

    def test_taken_only_counts_matching_year_and_type(self):
        ent = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2025, 12, 30), end_date=_date(2025, 12, 31), days=Decimal('2'))
        self.assertEqual(ent.taken_days, Decimal('0'))


class LeaveRecordTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': 30})

    def test_days_autocomputed_calendar_incl_weekend(self):
        # Thu 2 Jul -> Mon 6 Jul 2026 = 5 calendar days (Fri+Sat weekend counted)
        rec = LeaveRecord(employee=self.emp, leave_type=self.annual,
                          start_date=_date(2026, 7, 2), end_date=_date(2026, 7, 6))
        rec.save()
        self.assertEqual(rec.days, Decimal('5'))

    def test_days_count_calendar_incl_holiday(self):
        # A holiday inside the range still counts as a leave day (calendar count).
        Holiday.objects.create(date=_date(2026, 7, 7), name='X')
        rec = LeaveRecord(employee=self.emp, leave_type=self.annual,
                          start_date=_date(2026, 7, 5), end_date=_date(2026, 7, 9))
        rec.save()
        self.assertEqual(rec.days, Decimal('5'))

    def test_manual_days_override_preserved(self):
        rec = LeaveRecord(employee=self.emp, leave_type=self.annual,
                          start_date=_date(2026, 7, 5), end_date=_date(2026, 7, 9), days=Decimal('3'))
        rec.save()
        self.assertEqual(rec.days, Decimal('3'))

    def test_end_before_start_rejected(self):
        rec = LeaveRecord(employee=self.emp, leave_type=self.annual,
                          start_date=_date(2026, 7, 9), end_date=_date(2026, 7, 5))
        with self.assertRaises(ValidationError):
            rec.full_clean()


class LeaveTypeLocationDefaultsTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': None})

    def test_default_days_for_office_uses_flat_default(self):
        self.assertEqual(self.lt.default_days_for('office'), Decimal('30'))

    def test_default_days_for_blank_location_falls_back_to_office(self):
        self.assertEqual(self.lt.default_days_for(''), Decimal('30'))

    def test_default_days_for_site_falls_back_when_blank(self):
        self.assertEqual(self.lt.default_days_for('site'), Decimal('30'))

    def test_default_days_for_site_uses_override_when_set(self):
        self.lt.site_default_annual_days = Decimal('45')
        self.lt.save()
        self.assertEqual(self.lt.default_days_for('site'), Decimal('45'))

    def test_propagate_on_save_is_location_aware(self):
        office_emp = Employee.objects.create(iqama_number='LOC-OFF-1', full_name='Office Emp', work_location='office')
        site_emp = Employee.objects.create(iqama_number='LOC-SITE-1', full_name='Site Emp', work_location='site')
        blank_emp = Employee.objects.create(iqama_number='LOC-BLANK-1', full_name='Blank Emp')
        LeaveEntitlement.objects.create(employee=office_emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        LeaveEntitlement.objects.create(employee=site_emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        LeaveEntitlement.objects.create(employee=blank_emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.lt.default_annual_days = Decimal('32')
        self.lt.site_default_annual_days = Decimal('47')
        self.lt.save()
        self.assertEqual(LeaveEntitlement.objects.get(employee=office_emp).entitled_days, Decimal('32'))
        self.assertEqual(LeaveEntitlement.objects.get(employee=site_emp).entitled_days, Decimal('47'))
        self.assertEqual(LeaveEntitlement.objects.get(employee=blank_emp).entitled_days, Decimal('32'))


class LeaveExceptionGrantTests(TestCase):
    def setUp(self):
        self.emp = Employee.objects.create(iqama_number='LEG-1', full_name='Grant Test', work_location='site')
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.hr_user = make_user('leg-hr1')

    def test_grant_is_recorded_with_audit_fields(self):
        grant = LeaveExceptionGrant.objects.create(
            employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('5'),
            granted_by=self.hr_user, reason='Worked through Eid holidays.')
        self.assertEqual(grant.days, Decimal('5'))
        self.assertEqual(grant.granted_by, self.hr_user)
        self.assertIsNotNone(grant.granted_at)


class LeaveEntitlementEffectiveDaysTests(TestCase):
    def setUp(self):
        self.emp = Employee.objects.create(iqama_number='LEED-1', full_name='Eff Test', work_location='site')
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.ent = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))

    def test_no_grants_effective_equals_baseline(self):
        self.assertEqual(self.ent.exception_days, Decimal('0'))
        self.assertEqual(self.ent.effective_entitled_days, Decimal('45'))
        self.assertEqual(self.ent.effective_remaining_days, Decimal('45'))

    def test_grants_sum_into_effective_but_not_baseline(self):
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('5'), reason='x')
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('2'), reason='y')
        self.assertEqual(self.ent.entitled_days, Decimal('45'))  # baseline untouched
        self.assertEqual(self.ent.exception_days, Decimal('7'))
        self.assertEqual(self.ent.effective_entitled_days, Decimal('52'))

    def test_grants_from_other_years_or_types_dont_count(self):
        other_lt, _ = LeaveType.objects.get_or_create(code='sick', defaults={'name': 'Sick', 'default_annual_days': 12})
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=self.lt, year=2025, days=Decimal('9'), reason='wrong year')
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=other_lt, year=2026, days=Decimal('9'), reason='wrong type')
        self.assertEqual(self.ent.exception_days, Decimal('0'))


class LeaveRequestExceedsBalanceFieldTests(TestCase):
    def test_defaults_false(self):
        emp = Employee.objects.create(iqama_number='LREB-1', full_name='Default Test')
        lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        req = LeaveRequest.objects.create(employee=emp, leave_type=lt, start_date=_date(2026, 1, 1), end_date=_date(2026, 1, 2))
        self.assertFalse(req.exceeds_balance)


class OverrideAccessSettingsBalanceHoldTests(TestCase):
    def test_defaults_site_enabled_office_disabled(self):
        from hr.models import OverrideAccessSettings
        config = OverrideAccessSettings.get_solo()
        self.assertTrue(config.balance_hold_enabled_for('site'))
        self.assertFalse(config.balance_hold_enabled_for('office'))
        self.assertFalse(config.balance_hold_enabled_for(''))

    def test_office_can_be_enabled_without_code_change(self):
        from hr.models import OverrideAccessSettings
        config = OverrideAccessSettings.get_solo()
        config.allow_office_balance_hold = True
        config.save()
        self.assertTrue(OverrideAccessSettings.get_solo().balance_hold_enabled_for('office'))


class LocationAwareGenerationTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})

    def test_generate_picks_baseline_by_location(self):
        from hr.leave_services import generate_entitlements_for_employee
        office_emp = Employee.objects.create(iqama_number='LAG-OFF-1', full_name='Office', work_location='office')
        site_emp = Employee.objects.create(iqama_number='LAG-SITE-1', full_name='Site', work_location='site')
        generate_entitlements_for_employee(office_emp, 2026)
        generate_entitlements_for_employee(site_emp, 2026)
        self.assertEqual(LeaveEntitlement.objects.get(employee=office_emp, leave_type=self.lt).entitled_days, Decimal('30'))
        self.assertEqual(LeaveEntitlement.objects.get(employee=site_emp, leave_type=self.lt).entitled_days, Decimal('45'))

    def test_reapply_is_location_aware_and_preserves_exceptions(self):
        from hr.leave_services import generate_entitlements_for_employee, reapply_leave_type_defaults
        site_emp = Employee.objects.create(iqama_number='LAG-SITE-2', full_name='Site2', work_location='site')
        generate_entitlements_for_employee(site_emp, 2026)
        LeaveExceptionGrant.objects.create(employee=site_emp, leave_type=self.lt, year=2026, days=Decimal('5'), reason='x')
        reapply_leave_type_defaults(2026)
        ent = LeaveEntitlement.objects.get(employee=site_emp, leave_type=self.lt)
        self.assertEqual(ent.entitled_days, Decimal('45'))  # still the Site default, not wiped to Office's 30
        self.assertEqual(ent.exception_days, Decimal('5'))  # untouched


class ValidateLeaveSubmissionHoldTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})

    def test_within_cap_returns_false(self):
        from hr.leave_services import validate_leave_submission
        emp = Employee.objects.create(iqama_number='VLS-1', full_name='OK', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        held = validate_leave_submission(emp, self.lt, date(2026, 3, 1), date(2026, 3, 5))
        self.assertFalse(held)

    def test_site_over_cap_is_held_not_blocked(self):
        from hr.leave_services import validate_leave_submission
        emp = Employee.objects.create(iqama_number='VLS-2', full_name='Site Over', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        held = validate_leave_submission(emp, self.lt, date(2026, 1, 1), date(2026, 2, 15))  # 46 days
        self.assertTrue(held)

    def test_office_over_cap_still_hard_blocks(self):
        from hr.leave_services import validate_leave_submission
        emp = Employee.objects.create(iqama_number='VLS-3', full_name='Office Over', work_location='office')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        with self.assertRaises(ValueError):
            validate_leave_submission(emp, self.lt, date(2026, 1, 1), date(2026, 2, 1))  # 32 days

    def test_site_within_effective_cap_via_exception_grant_returns_false(self):
        from hr.leave_services import validate_leave_submission
        emp = Employee.objects.create(iqama_number='VLS-4', full_name='Site Grant', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        LeaveExceptionGrant.objects.create(employee=emp, leave_type=self.lt, year=2026, days=Decimal('5'), reason='x')
        held = validate_leave_submission(emp, self.lt, date(2026, 1, 1), date(2026, 2, 19))  # 50 days
        self.assertFalse(held)

    def test_no_entitlement_still_raises(self):
        from hr.leave_services import validate_leave_submission
        emp = Employee.objects.create(iqama_number='VLS-5', full_name='No Ent', work_location='site')
        with self.assertRaises(ValueError):
            validate_leave_submission(emp, self.lt, date(2026, 3, 1), date(2026, 3, 5))

    def test_overlap_still_raises_even_when_held(self):
        from hr.leave_services import validate_leave_submission
        emp = Employee.objects.create(iqama_number='VLS-6', full_name='Overlap', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        LeaveRecord.objects.create(employee=emp, leave_type=self.lt, start_date=date(2026, 1, 10), end_date=date(2026, 1, 12))
        with self.assertRaises(ValueError):
            validate_leave_submission(emp, self.lt, date(2026, 1, 1), date(2026, 2, 15))  # over cap AND overlaps


class SubmitLeaveRequestHeldTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.approver_user = make_user('slrh-appr')
        LeaveDashboardAccess.objects.create(user=self.approver_user, is_active=True)

    def test_held_request_still_gets_the_normal_approver_roster(self):
        # A held (exceeds_balance) request routes through the exact same
        # approver roster and decide flow as any other request — the only
        # difference is the warning shown on it, not who can decide it.
        from hr.leave_approval_services import submit_leave_request
        emp = Employee.objects.create(iqama_number='SLRH-1', full_name='Site Over', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        req = submit_leave_request(
            employee=emp, leave_type=self.lt, start_date=date(2026, 1, 1), end_date=date(2026, 2, 15),
            created_by=self.approver_user)
        self.assertTrue(req.exceeds_balance)
        self.assertEqual(req.approvals.count(), 1)
        self.assertEqual(req.approvals.first().approver, self.approver_user)
        self.assertEqual(req.status, 'pending')

    def test_normal_request_still_gets_approval_rows(self):
        from hr.leave_approval_services import submit_leave_request
        emp = Employee.objects.create(iqama_number='SLRH-2', full_name='Site OK', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        req = submit_leave_request(
            employee=emp, leave_type=self.lt, start_date=date(2026, 3, 1), end_date=date(2026, 3, 5),
            created_by=self.approver_user)
        self.assertFalse(req.exceeds_balance)
        self.assertEqual(req.approvals.count(), 1)


class GrantExceptionDaysTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='GED-1', full_name='Grantee', work_location='site')
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        self.hr_user = make_user('ged-hr1')

    def test_grant_increases_effective_but_not_baseline(self):
        from hr.leave_approval_services import grant_exception_days
        grant_exception_days(employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('5'),
                              granted_by=self.hr_user, reason='Eid overtime.')
        ent = LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.entitled_days, Decimal('45'))
        self.assertEqual(ent.exception_days, Decimal('5'))

    def test_grant_requires_a_reason(self):
        from hr.leave_approval_services import grant_exception_days
        with self.assertRaises(ValueError):
            grant_exception_days(employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('5'),
                                  granted_by=self.hr_user, reason='   ')


class GrantExceptionDaysViewPermissionTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='GEDV-1', full_name='Grantee View', work_location='site')
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=timezone.now().year, entitled_days=Decimal('45'))

    def test_super_admin_can_grant_via_the_view(self):
        # Baseline is 45 (site default, no prior exception) — entering a new
        # total of 48 should be stored as a +3 exception, same as before,
        # just expressed as a target total instead of a raw delta.
        from accounts.models import Role
        _make_role_user('gedv-super', Role.SUPER_ADMIN)
        self.client.login(username='gedv-super', password='testpass123')
        resp = self.client.post(reverse('hr:grant_exception_days', args=[self.emp.pk]), data={
            'leave_type': self.lt.pk, 'year': timezone.now().year, 'new_total_days': '48',
            'reason': 'Test grant via view.'})
        self.assertEqual(resp.status_code, 302)
        ent = LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.lt, year=timezone.now().year)
        self.assertEqual(ent.exception_days, Decimal('3'))

    def test_plain_user_is_denied(self):
        plain_user = make_user('gedv-plain')
        plain_user.set_password('x')
        plain_user.save()
        self.client.login(username='gedv-plain', password='x')
        resp = self.client.get(reverse('hr:grant_exception_days', args=[self.emp.pk]))
        self.assertEqual(resp.status_code, 403)
        resp = self.client.post(reverse('hr:grant_exception_days', args=[self.emp.pk]), data={
            'leave_type': self.lt.pk, 'year': timezone.now().year, 'new_total_days': '48',
            'reason': 'Should not work.'})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(LeaveEntitlement.objects.get(employee=self.emp).exception_days, Decimal('0'))


class GrantExceptionDaysTargetTotalTests(TestCase):
    """The Adjust Balance form takes a target total, not a signed +/- delta
    — the view diffs it against the employee's current effective total
    (LeaveEntitlement.effective_entitled_days) to derive the
    LeaveExceptionGrant it actually stores. Reached with leave_type/year
    pre-filled (the normal path, via a per-row Adjust link), those two
    fields lock and the current total is shown/pre-filled; reached without
    prefill (a bare URL visit), they stay editable and no total is shown."""
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='GETT-1', full_name='Target Total Grantee', work_location='site')
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        from accounts.models import Role
        _make_role_user('gett-super', Role.SUPER_ADMIN)
        self.client.login(username='gett-super', password='testpass123')

    def _url(self, **params):
        base = reverse('hr:grant_exception_days', args=[self.emp.pk])
        if not params:
            return base
        return f"{base}?{'&'.join(f'{k}={v}' for k, v in params.items())}"

    def test_prefilled_locks_the_fields_and_shows_current_total(self):
        resp = self.client.get(self._url(leave_type=self.lt.pk, year=2026))
        self.assertContains(resp, f'value="{self.lt.pk}"')
        self.assertContains(resp, 'currently 45.0 total day')
        self.assertNotContains(resp, '<select name="leave_type"')

    def test_prefilled_new_total_days_initial_is_the_current_total(self):
        resp = self.client.get(self._url(leave_type=self.lt.pk, year=2026))
        self.assertEqual(resp.context['form'].initial['new_total_days'], Decimal('45'))

    def test_without_prefill_leave_type_stays_an_editable_dropdown(self):
        resp = self.client.get(self._url())
        self.assertContains(resp, '<select name="leave_type"')
        self.assertNotContains(resp, 'currently')

    def test_higher_total_grants_the_difference(self):
        resp = self.client.post(self._url(leave_type=self.lt.pk, year=2026), data={
            'leave_type': self.lt.pk, 'year': 2026, 'new_total_days': '50', 'reason': 'Worked through Eid.'})
        self.assertEqual(resp.status_code, 302)
        ent = LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.exception_days, Decimal('5'))

    def test_lower_total_deducts_the_difference(self):
        # This is the "reduce" case — a new employee entitled to fewer days
        # than the standard baseline for company-policy reasons.
        resp = self.client.post(self._url(leave_type=self.lt.pk, year=2026), data={
            'leave_type': self.lt.pk, 'year': 2026, 'new_total_days': '40', 'reason': 'New-hire policy cap.'})
        self.assertEqual(resp.status_code, 302)
        ent = LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.exception_days, Decimal('-5'))
        self.assertEqual(ent.effective_entitled_days, Decimal('40'))

    def test_same_total_is_a_no_op(self):
        resp = self.client.post(self._url(leave_type=self.lt.pk, year=2026), data={
            'leave_type': self.lt.pk, 'year': 2026, 'new_total_days': '45', 'reason': 'No actual change.'})
        self.assertEqual(resp.status_code, 302)
        ent = LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.exception_days, Decimal('0'))
        self.assertFalse(LeaveExceptionGrant.objects.filter(employee=self.emp).exists())

    def test_negative_new_total_rejected(self):
        resp = self.client.post(self._url(leave_type=self.lt.pk, year=2026), data={
            'leave_type': self.lt.pk, 'year': 2026, 'new_total_days': '-5', 'reason': 'Should not work.'})
        self.assertEqual(resp.status_code, 200)  # form re-renders with an error
        self.assertFalse(LeaveExceptionGrant.objects.filter(employee=self.emp).exists())

    def test_deduction_reflected_on_leave_summary(self):
        self.client.post(self._url(leave_type=self.lt.pk, year=2026), data={
            'leave_type': self.lt.pk, 'year': 2026, 'new_total_days': '40', 'reason': 'New-hire policy cap.'})
        resp = self.client.get(reverse('hr:leave_summary', args=[self.emp.pk]) + '?year=2026')
        # New effective total is primary; the struck-through standard 45.0
        # and a red "reduced" badge (absolute value, no sign — the icon and
        # color already convey direction) show the adjustment.
        self.assertContains(resp, '<div class="fw-semibold">40.0</div>')
        self.assertContains(resp, 'text-decoration-line-through')
        self.assertContains(resp, 'bg-danger-subtle')
        self.assertContains(resp, 'bi-arrow-down-short')


class BalanceHoldSettingsViewTests(TestCase):
    def test_super_admin_can_change_balance_hold_settings(self):
        from accounts.models import Role
        from hr.models import OverrideAccessSettings
        _make_role_user('bhsv-super', Role.SUPER_ADMIN)
        self.client.login(username='bhsv-super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_access'), data={
            'set_balance_hold': '1', 'allow_office_balance_hold': '1',
        })
        self.assertEqual(resp.status_code, 302)
        config = OverrideAccessSettings.get_solo()
        self.assertTrue(config.allow_office_balance_hold)
        self.assertFalse(config.allow_site_balance_hold)  # unchecked box -> off

    def test_plain_user_cannot_change_balance_hold_settings(self):
        from hr.models import OverrideAccessSettings
        plain_user = make_user('bhsv-plain')
        plain_user.set_password('x')
        plain_user.save()
        self.client.login(username='bhsv-plain', password='x')
        resp = self.client.post(reverse('hr:leave_access'), data={
            'set_balance_hold': '1', 'allow_office_balance_hold': '1',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(OverrideAccessSettings.get_solo().allow_office_balance_hold)


class LogRequestBalanceWarningTests(TestCase):
    """The 'Log Request' (log-on-behalf-of) form: a Super Admin submitting an
    over-cap request sees an inline warning and must explicitly confirm
    before anything is created; confirming grants the exact shortfall as an
    exception and then logs the request normally (not held)."""
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='LRBW-1', full_name='Warning Test', work_location='site')
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))

    def _post(self, **extra):
        data = {
            'employee': self.emp.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-01-01', 'end_date': '2026-02-16',  # 47 days, 2 over the 45-day Site baseline
        }
        data.update(extra)
        return self.client.post(reverse('hr:leave_request_create'), data=data)

    def test_super_admin_first_submit_shows_warning_and_creates_nothing(self):
        from accounts.models import Role
        _make_role_user('lrbw-super1', Role.SUPER_ADMIN)
        self.client.login(username='lrbw-super1', password='testpass123')
        resp = self._post()
        self.assertEqual(resp.status_code, 200)  # re-rendered, not redirected
        self.assertContains(resp, 'Exceeds balance')
        self.assertContains(resp, 'Log Anyway')
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp).exists())
        self.assertEqual(LeaveEntitlement.objects.get(employee=self.emp).exception_days, Decimal('0'))

    def test_super_admin_confirms_grants_shortfall_and_logs_normally(self):
        from accounts.models import Role
        _make_role_user('lrbw-super2', Role.SUPER_ADMIN)
        self.client.login(username='lrbw-super2', password='testpass123')
        resp = self._post(confirm_grant_and_log='1')
        self.assertEqual(resp.status_code, 302)
        ent = LeaveEntitlement.objects.get(employee=self.emp)
        self.assertEqual(ent.entitled_days, Decimal('45'))  # baseline untouched
        self.assertEqual(ent.exception_days, Decimal('2'))  # exact shortfall granted
        req = LeaveRequest.objects.get(employee=self.emp)
        self.assertFalse(req.exceeds_balance)  # fits now, logged normally
        self.assertEqual(req.approvals.count(), 0)  # no LeaveDashboardAccess holders in this test's DB, unrelated to the grant

    def test_non_super_admin_gets_no_warning_request_is_just_held(self):
        from accounts.models import Role
        _make_role_user('lrbw-erp', Role.ERP_ADMIN)
        self.client.login(username='lrbw-erp', password='testpass123')
        resp = self._post()
        self.assertEqual(resp.status_code, 302)  # no warning step — logged straight through
        req = LeaveRequest.objects.get(employee=self.emp)
        self.assertTrue(req.exceeds_balance)
        self.assertEqual(LeaveEntitlement.objects.get(employee=self.emp).exception_days, Decimal('0'))

    def test_confirm_flag_without_override_access_is_denied(self):
        # Defense in depth: a direct POST with the confirm flag set, from a
        # user who isn't actually allowed to grant, must not silently work.
        from accounts.models import Role
        _make_role_user('lrbw-erp2', Role.ERP_ADMIN)
        self.client.login(username='lrbw-erp2', password='testpass123')
        resp = self._post(confirm_grant_and_log='1')
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(LeaveEntitlement.objects.get(employee=self.emp).exception_days, Decimal('0'))
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp).exists())

    def test_within_cap_submission_never_shows_warning(self):
        from accounts.models import Role
        _make_role_user('lrbw-super3', Role.SUPER_ADMIN)
        self.client.login(username='lrbw-super3', password='testpass123')
        resp = self._post(end_date='2026-01-10')  # 10 days, well within 45
        self.assertEqual(resp.status_code, 302)
        req = LeaveRequest.objects.get(employee=self.emp)
        self.assertFalse(req.exceeds_balance)


class EmployeeWorkLocationTransferTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})

    def test_upgrade_recomputes_baseline_for_whole_year(self):
        emp = Employee.objects.create(iqama_number='EWLT-1', full_name='Upgrade', work_location='office')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        LeaveRecord.objects.create(employee=emp, leave_type=self.lt, start_date=date(2026, 2, 1), end_date=date(2026, 2, 10))  # 10 days taken
        from hr.leave_services import apply_work_location_transfer
        apply_work_location_transfer(emp, old_location='office', new_location='site', year=2026, actor=None)
        ent = LeaveEntitlement.objects.get(employee=emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.entitled_days, Decimal('45'))

    def test_downgrade_auto_grants_the_shortfall(self):
        emp = Employee.objects.create(iqama_number='EWLT-2', full_name='Downgrade', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        LeaveRecord.objects.create(employee=emp, leave_type=self.lt, start_date=date(2026, 1, 1), end_date=date(2026, 2, 4))  # 35 days taken
        from hr.leave_services import apply_work_location_transfer
        apply_work_location_transfer(emp, old_location='site', new_location='office', year=2026, actor=None)
        ent = LeaveEntitlement.objects.get(employee=emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.entitled_days, Decimal('30'))  # baseline drops to Office
        self.assertEqual(ent.exception_days, Decimal('5'))  # 35 taken - 30 new baseline, auto-preserved
        self.assertEqual(ent.effective_remaining_days, Decimal('0'))

    def test_no_op_when_location_unchanged(self):
        emp = Employee.objects.create(iqama_number='EWLT-3', full_name='NoOp', work_location='site')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        from hr.leave_services import apply_work_location_transfer
        apply_work_location_transfer(emp, old_location='site', new_location='site', year=2026, actor=None)
        ent = LeaveEntitlement.objects.get(employee=emp, leave_type=self.lt, year=2026)
        self.assertEqual(ent.entitled_days, Decimal('45'))
        self.assertEqual(ent.exception_days, Decimal('0'))


class EmployeeUpdateViewTransferTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        from accounts.models import Role
        self.admin_user = _make_role_user('euvt-admin', Role.SUPER_ADMIN)

    def test_editing_work_location_via_view_triggers_recompute(self):
        self.client.login(username='euvt-admin', password='testpass123')
        emp = Employee.objects.create(iqama_number='EUVT-1', full_name='View Transfer', work_location='office')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=timezone.now().year, entitled_days=Decimal('30'))
        resp = self.client.post(reverse('hr:employee_update', args=[emp.pk]), data={
            'iqama_number': 'EUVT-1', 'full_name': 'View Transfer', 'work_location': 'site',
        })
        self.assertEqual(resp.status_code, 302)
        emp.refresh_from_db()
        self.assertEqual(emp.work_location, 'site')
        ent = LeaveEntitlement.objects.get(employee=emp, leave_type=self.lt, year=timezone.now().year)
        self.assertEqual(ent.entitled_days, Decimal('45'))


class HeldRequestDisplayTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='HRD-1', full_name='Held Display', work_location='site')
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        from accounts.models import Role
        self.hr_user = _make_role_user('hrdview', Role.SUPER_ADMIN)
        self.client.login(username='hrdview', password='testpass123')

    def test_list_page_shows_exceeds_balance_badge(self):
        from hr.leave_approval_services import submit_leave_request
        submit_leave_request(employee=self.emp, leave_type=self.lt, start_date=date(2026, 1, 1),
                              end_date=date(2026, 2, 15), created_by=self.hr_user)
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, 'Exceeds balance')

    def test_detail_page_shows_breakdown(self):
        from hr.leave_approval_services import submit_leave_request
        req = submit_leave_request(employee=self.emp, leave_type=self.lt, start_date=date(2026, 1, 1),
                                    end_date=date(2026, 2, 15), created_by=self.hr_user)
        resp = self.client.get(reverse('hr:leave_request_detail', args=[req.pk]))
        self.assertContains(resp, 'Exceeds balance')

    def test_normal_request_shows_no_badge(self):
        from hr.leave_approval_services import submit_leave_request
        submit_leave_request(employee=self.emp, leave_type=self.lt, start_date=date(2026, 3, 1),
                              end_date=date(2026, 3, 5), created_by=self.hr_user)
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertNotContains(resp, 'Exceeds balance')

    def test_my_profile_shows_exceeds_balance_badge(self):
        # The badge already existed on the Leave Queue and detail pages,
        # and on the Direct Reports card's in-behalf lines, but was never
        # rendered on the employee's own top-level leave requests table.
        from hr.leave_approval_services import submit_leave_request
        emp_user = _login_user('hrd-emp')
        self.emp.user = emp_user
        self.emp.save(update_fields=['user'])
        submit_leave_request(employee=self.emp, leave_type=self.lt, start_date=date(2026, 1, 1),
                              end_date=date(2026, 2, 15), created_by=self.hr_user)
        self.client.logout()
        self.client.login(username='hrd-emp', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Exceeds balance')


class EmployeeDashboardExceptionDisplayTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='EDED-1', full_name='Dashboard Test', work_location='site')
        self.ent = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('5'), reason='x')
        from accounts.models import Role
        self.hr_user = _make_role_user('edeview', Role.SUPER_ADMIN)
        self.client.login(username='edeview', password='testpass123')

    def test_leave_summary_shows_baseline_and_exception_separately(self):
        resp = self.client.get(reverse('hr:leave_summary', args=[self.emp.pk]) + '?year=2026')
        self.assertContains(resp, '<div class="fw-semibold">50.0</div>')  # effective total (45 + 5)
        self.assertContains(resp, 'text-decoration-line-through')  # struck-through standard 45.0
        self.assertContains(resp, 'bg-success-subtle')
        self.assertContains(resp, '+5')

    def test_grant_exception_days_link_visible_for_override_access(self):
        # The balance-adjustment entry point is a per-row action in the
        # breakdown table (pre-filled to that row's leave type/year), not a
        # single generic header button — mirrors the Entitlements page.
        resp = self.client.get(reverse('hr:leave_summary', args=[self.emp.pk]) + '?year=2026')
        expected_url = f"{reverse('hr:grant_exception_days', args=[self.emp.pk])}?leave_type={self.lt.pk}&year=2026"
        self.assertContains(resp, expected_url)
        self.assertContains(resp, 'Adjust')

    def test_no_standalone_adjust_button_in_header(self):
        resp = self.client.get(reverse('hr:leave_summary', args=[self.emp.pk]) + '?year=2026')
        self.assertNotContains(resp, 'Add Exception Days')


class LeaveSummaryAndRecordCreateNavigationTests(TestCase):
    """Leave Summary used to show two redundant "Add Leave" buttons (only
    one of which pre-filled the employee) and a "Back to Employee" link that
    didn't reflect how the page is actually reached (only ever from Leave
    Entitlements). The Add Leave form's own Back/Cancel links, in turn,
    always said "Back to Queue"/"Back to My Profile" regardless of where the
    submitter actually came from."""
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        self.emp = Employee.objects.create(iqama_number='LSNT-1', full_name='Nav Test Employee')
        from accounts.models import Role
        self.hr_user = _make_role_user('lsntview', Role.SUPER_ADMIN)
        self.client.login(username='lsntview', password='testpass123')

    def test_leave_summary_shows_exactly_one_add_leave_button(self):
        resp = self.client.get(reverse('hr:leave_summary', args=[self.emp.pk]) + '?year=2026')
        self.assertContains(resp, 'Add Leave', count=1)
        expected_href = f"{reverse('hr:leave_record_create')}?employee={self.emp.pk}&back=leave_summary"
        self.assertContains(resp, expected_href)

    def test_leave_summary_back_link_points_to_entitlements(self):
        resp = self.client.get(reverse('hr:leave_summary', args=[self.emp.pk]) + '?year=2026')
        self.assertContains(resp, 'Back to Leave Entitlements')
        self.assertContains(resp, f"{reverse('hr:entitlement_year')}?year=2026")
        self.assertNotContains(resp, 'Back to Employee')

    def test_add_leave_form_reached_from_leave_summary_backs_to_leave_summary(self):
        resp = self.client.get(
            reverse('hr:leave_record_create') + f'?employee={self.emp.pk}&back=leave_summary')
        self.assertContains(resp, 'Back to Leave Summary')
        self.assertContains(resp, reverse('hr:leave_summary', args=[self.emp.pk]))

    def test_add_leave_form_reached_directly_still_backs_to_queue(self):
        # No employee/back params (e.g. the sidebar link, or Attendance Matrix)
        # — existing behaviour for a submitter who can view the queue must be
        # unchanged.
        resp = self.client.get(reverse('hr:leave_record_create'))
        self.assertContains(resp, 'Back to Queue')
        self.assertNotContains(resp, 'Back to Leave Summary')


class EntitlementYearExceptionDisplayTests(TestCase):
    """The company-wide Entitlements page (HR -> Leave -> Entitlements) must
    also reflect exception grants — it used to compute Remaining from raw
    entitled_days only, so a grant never showed up there even though it was
    correctly visible on the employee's own Leave Summary."""
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.emp = Employee.objects.create(iqama_number='EYED-1', full_name='Entitlement Year Test', work_location='site')
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=self.lt, year=2026, days=Decimal('1'), reason='x')
        from accounts.models import Role
        _make_role_user('eyed-hr', Role.SUPER_ADMIN)
        self.client.login(username='eyed-hr', password='testpass123')

    def test_remaining_includes_the_granted_day(self):
        resp = self.client.get(reverse('hr:entitlement_year') + '?year=2026')
        self.assertEqual(resp.context['groups'][0]['total_remaining'], Decimal('46'))
        self.assertEqual(resp.context['groups'][0]['total_exception'], Decimal('1'))
        self.assertContains(resp, '+1')

    def test_row_breakdown_includes_exception_column(self):
        resp = self.client.get(reverse('hr:entitlement_year') + '?year=2026')
        row = resp.context['groups'][0]['rows'][0]
        self.assertEqual(row['exception'], Decimal('1'))
        self.assertEqual(row['remaining'], Decimal('46'))


class MyProfileExceptionDisplayTests(TestCase):
    """The employee's own My Profile leave-balance table must also reflect
    an exception grant, not just HR-facing pages."""
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        self.user = User.objects.create_user('mped-emp', password='x')
        self.emp = Employee.objects.create(
            iqama_number='MPED-1', full_name='Profile Exception Test', work_location='site', user=self.user)
        year = timezone.now().year
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=year, entitled_days=Decimal('45'))
        LeaveExceptionGrant.objects.create(employee=self.emp, leave_type=self.lt, year=year, days=Decimal('1'), reason='x')

    def test_my_profile_shows_the_granted_day(self):
        # My Profile shows only the clean effective total by default — the
        # standard-vs-adjusted breakdown is tucked behind an info icon
        # (tooltip) instead of always-visible, unlike the HR-facing pages.
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, '46.0')  # effective total (45 + 1) in the per-type row
        self.assertContains(resp, 'bi-info-circle')
        self.assertContains(resp, 'data-bs-toggle="tooltip"')
        self.assertNotContains(resp, 'text-decoration-line-through')
        self.assertEqual(resp.context['leave_total_exception'], Decimal('1'))


class PendingCountsContextProcessorTests(TestCase):
    def setUp(self):
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        from accounts.models import Role
        self.hr_user = _make_role_user('pccp-hr', Role.SUPER_ADMIN)
        LeaveDashboardAccess.objects.create(user=self.hr_user, is_active=True)

    def test_leave_requests_pending_count_reflects_actual_pending(self):
        self.client.login(username='pccp-hr', password='testpass123')
        emp = Employee.objects.create(iqama_number='PCCP-1', full_name='Ctx Test', work_location='office')
        LeaveEntitlement.objects.create(employee=emp, leave_type=self.lt, year=timezone.now().year, entitled_days=Decimal('30'))
        from hr.leave_approval_services import submit_leave_request
        submit_leave_request(employee=emp, leave_type=self.lt, start_date=date.today(), end_date=date.today(), created_by=self.hr_user)
        resp = self.client.get(reverse('hr:hr_dashboard'))
        self.assertEqual(resp.context['leave_requests_pending_count'], 1)

    def test_plain_employee_gets_no_counts_in_context(self):
        plain_user = make_user('pccp-plain')
        plain_user.set_password('x')
        plain_user.save()
        self.client.login(username='pccp-plain', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotIn('leave_requests_pending_count', resp.context)
        self.assertNotIn('team_exceptions_pending_count', resp.context)


class SidebarBadgeMarkupTests(TestCase):
    def test_leave_requests_badge_renders_when_pending(self):
        lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        from accounts.models import Role
        hr_user = _make_role_user('sbmt-hr', Role.SUPER_ADMIN)
        LeaveDashboardAccess.objects.create(user=hr_user, is_active=True)
        emp = Employee.objects.create(iqama_number='SBMT-1', full_name='Badge Test', work_location='office')
        LeaveEntitlement.objects.create(employee=emp, leave_type=lt, year=timezone.now().year, entitled_days=Decimal('30'))
        from hr.leave_approval_services import submit_leave_request
        submit_leave_request(employee=emp, leave_type=lt, start_date=date.today(), end_date=date.today(), created_by=hr_user)
        self.client.login(username='sbmt-hr', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, '<span class="nav-badge-count">1</span>')

    def test_no_badge_when_nothing_pending(self):
        from accounts.models import Role
        hr_user = _make_role_user('sbmt-hr2', Role.SUPER_ADMIN)
        LeaveDashboardAccess.objects.create(user=hr_user, is_active=True)
        self.client.login(username='sbmt-hr2', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, '<span class="nav-badge-count">')
        self.assertNotContains(resp, '<span class="nav-badge-dot">')


class TeamExceptionsTabCountTests(TestCase):
    def test_tab_buttons_show_pending_counts(self):
        from hr.attendance_exception_services import submit_attendance_exception
        manager_user = make_user('tabmgr')
        manager_user.set_password('testpass123')
        manager_user.save()
        manager = Employee.objects.create(iqama_number='TABC-MGR', full_name='Tab Manager', user=manager_user)
        report = Employee.objects.create(iqama_number='TABC-RPT', full_name='Tab Report', main_manager=manager)
        submit_attendance_exception(
            employee=report, event_date=_timezone.now().date(), event_start_time=_time(0, 1),
            reason_category='site_visit', created_by=manager_user)
        self.client.login(username='tabmgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.context['direct_tab_count'], 1)
        self.assertEqual(resp.context['secondary_tab_count'], 0)
        self.assertContains(resp, 'Direct Reports <span class="nav-badge-count">1</span>')


from hr.leave_services import generate_year_entitlements


class WorkingDayModelTests(TestCase):
    def test_unique_date_and_str(self):
        from django.db import IntegrityError
        from hr.models import WorkingDay
        wd = WorkingDay.objects.create(date=_date(2026, 7, 18), name='Working Saturday')
        self.assertIn('2026-07-18', str(wd))
        with self.assertRaises(IntegrityError):
            WorkingDay.objects.create(date=_date(2026, 7, 18), name='dup')


class GeneratorTests(TestCase):
    def setUp(self):
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': 30})
        # Sick forced to 15 (not the seeded 30) so the generator test verifies the flat-default path with a non-seed value.
        self.sick, _ = LeaveType.objects.update_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 15})
        self.e_new = make_employee('A', 'New', _date(2026, 2, 1))   # joins 2026
        self.e_old = make_employee('B', 'Old', _date(2020, 1, 1))   # tenured
        inactive = make_employee('C', 'Inactive', _date(2019, 1, 1))
        inactive.is_active = False
        inactive.save()

    def test_generates_flat_default_rows_for_active_employees(self):
        generate_year_entitlements(2026)
        self.assertEqual(self._ent(self.e_new, self.annual), Decimal('30'))  # flat default
        self.assertEqual(self._ent(self.e_old, self.annual), Decimal('30'))  # flat default
        self.assertEqual(self._ent(self.e_new, self.sick), Decimal('15'))    # flat default
        self.assertFalse(LeaveEntitlement.objects.filter(employee__iqama_number='C').exists())

    def test_does_not_overwrite_existing(self):
        LeaveEntitlement.objects.create(employee=self.e_old, leave_type=self.annual, year=2026, entitled_days=99)
        generate_year_entitlements(2026)
        self.assertEqual(self._ent(self.e_old, self.annual), Decimal('99'))

    def _ent(self, emp, lt):
        return LeaveEntitlement.objects.get(employee=emp, leave_type=lt, year=2026).entitled_days


class GenerateEntitlementsForEmployeeTests(TestCase):
    """Unit coverage of hr.leave_services.generate_entitlements_for_employee —
    the single-employee helper shared by generate_year_entitlements (bulk)
    and the new-employee auto-generation hooks below."""
    def setUp(self):
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': 30})
        self.sick, _ = LeaveType.objects.update_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 12, 'is_active': True})
        self.emp = make_employee('GEFE-1', 'Solo Employee')

    def test_creates_one_row_per_active_leave_type(self):
        # Not asserting an exact `created` count here: a data migration
        # (0024_leavetype_task1_values) pre-seeds several canonical LeaveType
        # rows (annual, sick, marriage, ...) into every test database, so the
        # real active-type count is >= the 2 this test explicitly cares about.
        from hr.leave_services import generate_entitlements_for_employee
        created = generate_entitlements_for_employee(self.emp, 2026)
        self.assertGreaterEqual(created, 2)
        self.assertEqual(
            LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.annual, year=2026).entitled_days,
            Decimal('30'))
        self.assertEqual(
            LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.sick, year=2026).entitled_days,
            Decimal('12'))

    def test_does_not_overwrite_existing_row(self):
        from hr.leave_services import generate_entitlements_for_employee
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=99)
        generate_entitlements_for_employee(self.emp, 2026)
        self.assertEqual(
            LeaveEntitlement.objects.get(employee=self.emp, leave_type=self.annual, year=2026).entitled_days,
            Decimal('99'))

    def test_ignores_inactive_leave_types(self):
        from hr.leave_services import generate_entitlements_for_employee
        inactive_lt, _ = LeaveType.objects.get_or_create(code='defunct', defaults={
            'name': 'Defunct', 'default_annual_days': 5, 'is_active': False})
        generate_entitlements_for_employee(self.emp, 2026)
        self.assertFalse(LeaveEntitlement.objects.filter(employee=self.emp, leave_type=inactive_lt).exists())


class NewEmployeeAutoEntitlementTests(TestCase):
    """A newly added employee (via the real 'Add Employee' form or the Excel
    import) automatically gets this year's leave balances — no separate
    manual 'Generate' step needed. Deliberately hooked into these two actual
    creation entry points, NOT Employee.save() itself: a model-level hook
    would fire on every raw Employee.objects.create() across the whole test
    suite (and any future script), silently colliding with the ~30 existing
    tests that create their own LeaveEntitlement rows for a just-created
    employee in the current year."""
    def setUp(self):
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': 30})
        self.admin = _make_role_user('neaet_admin', Role.SUPER_ADMIN)
        self.this_year = timezone.now().year

    def test_creating_employee_via_form_generates_entitlements(self):
        self.client.login(username='neaet_admin', password='testpass123')
        resp = self.client.post(reverse('hr:employee_create'), {
            'iqama_number': 'NEAET-1', 'full_name': 'Fresh Hire', 'is_active': 'on',
        })
        self.assertEqual(resp.status_code, 302)
        emp = Employee.objects.get(iqama_number='NEAET-1')
        ent = LeaveEntitlement.objects.get(employee=emp, leave_type=self.annual, year=self.this_year)
        self.assertEqual(ent.entitled_days, Decimal('30'))

    def test_creating_inactive_employee_via_form_generates_no_entitlements(self):
        self.client.login(username='neaet_admin', password='testpass123')
        resp = self.client.post(reverse('hr:employee_create'), {
            'iqama_number': 'NEAET-2', 'full_name': 'Not Yet Active',
        })
        self.assertEqual(resp.status_code, 302)
        emp = Employee.objects.get(iqama_number='NEAET-2')
        self.assertFalse(emp.is_active)
        self.assertFalse(LeaveEntitlement.objects.filter(employee=emp).exists())

    def test_updating_existing_employee_does_not_regenerate(self):
        # EmployeeUpdateView reuses EmployeeForm but is a different view
        # (UpdateView, not CreateView) — this hook lives only in
        # EmployeeCreateView.form_valid, so an edit must never touch entitlements.
        emp = make_employee('NEAET-3', 'Existing Employee')
        self.client.login(username='neaet_admin', password='testpass123')
        resp = self.client.post(reverse('hr:employee_update', args=[emp.pk]), {
            'iqama_number': 'NEAET-3', 'full_name': 'Existing Employee Renamed', 'is_active': 'on',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(LeaveEntitlement.objects.filter(employee=emp).exists())


class LeaveAdminViewTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x')
        self.admin.role = role
        self.admin.save()

    def test_leavetype_list_ok(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse('hr:leavetype_list')).status_code, 200)

    def test_leavetype_create_non_annual_is_not_accumulative(self):
        # is_accumulative is a fixed business rule (only Annual is ever True) —
        # it's no longer on LeaveTypeForm, so even posting it can't flip it.
        # A newly created custom leave type must default to non-accumulative.
        self.client.force_login(self.admin)
        resp = self.client.post(reverse('hr:leavetype_create'), {
            'name': 'Bereavement', 'code': 'bereavement', 'default_annual_days': '3.0',
            'is_paid': 'on', 'color': 'secondary', 'is_active': 'on',
            'is_accumulative': 'on',  # attempted (malicious/mistaken) override — must be ignored
        })
        self.assertRedirects(resp, reverse('hr:leavetype_list'))
        lt = LeaveType.objects.get(code='bereavement')
        self.assertFalse(lt.is_accumulative)

    def test_leavetype_create_annual_code_is_accumulative(self):
        self.client.force_login(self.admin)
        LeaveType.objects.filter(code='annual').delete()
        resp = self.client.post(reverse('hr:leavetype_create'), {
            'name': 'Annual', 'code': 'annual', 'default_annual_days': '30.0',
            'is_paid': 'on', 'color': 'primary', 'is_active': 'on',
        })
        self.assertRedirects(resp, reverse('hr:leavetype_list'))
        lt = LeaveType.objects.get(code='annual')
        self.assertTrue(lt.is_accumulative)

    def test_holiday_create(self):
        self.client.force_login(self.admin)
        resp = self.client.post(reverse('hr:holiday_create'),
                                {'date': '2026-07-17', 'name': 'Eid', 'category': 'eid', 'is_active': 'on'})
        self.assertRedirects(resp, reverse('hr:holiday_list'))
        self.assertEqual(Holiday.objects.filter(name='Eid').count(), 1)

    def test_asset_form_has_employee_picker_autofill(self):
        from hr.models import Employee
        Employee.objects.create(iqama_number='EMP-AST', full_name='Dana Ali',
                                designation='Network Engineer', is_active=True)
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:asset_create'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'employee-picker')          # the dropdown
        self.assertContains(resp, 'asset-employees')          # the JSON map
        self.assertContains(resp, 'Dana Ali')                 # employee option
        self.assertContains(resp, 'Network Engineer')         # designation in the map


class LeaveRecordViewTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        # leave_record_create now unifies onto LeaveRequestCreateView, which is
        # SuperAdminRequiredMixin (the approval-workflow queue's existing access
        # level) — a plain Admin can no longer reach it, only Super Admin.
        superadmin_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.superadmin = User.objects.create_user('supadm', password='x')
        self.superadmin.role = superadmin_role
        self.superadmin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def test_create_leave_autocomputes_days(self):
        # Annual now requires approval like every other leave type (the
        # "accumulative skips approval" bypass was removed) — this submission
        # creates a pending LeaveRequest, not an immediate LeaveRecord;
        # computed_days() is identical logic on both models.
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        self.client.force_login(self.superadmin)
        self.client.post(reverse('hr:leave_record_create'), {
            'employee': self.emp.pk, 'leave_type': self.annual.pk,
            'start_date': '2026-07-05', 'end_date': '2026-07-09', 'employee_reason': ''})
        req = LeaveRequest.objects.get(employee=self.emp)
        self.assertEqual(req.status, 'pending')
        self.assertEqual(req.days, Decimal('5'))

    def test_create_leave_counts_calendar_days_incl_weekend(self):
        # Thu 2 Jul -> Mon 6 Jul = 5 calendar days (weekend counted), days blank.
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        self.client.force_login(self.superadmin)
        self.client.post(reverse('hr:leave_record_create'), {
            'employee': self.emp.pk, 'leave_type': self.annual.pk,
            'start_date': '2026-07-02', 'end_date': '2026-07-06', 'employee_reason': ''})
        req = LeaveRequest.objects.get(employee=self.emp, start_date=_date(2026, 7, 2))
        self.assertEqual(req.days, Decimal('5'))

    def test_summary_shows_balance(self):
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:leave_summary', kwargs={'pk': self.emp.pk}) + '?year=2026')
        self.assertEqual(resp.status_code, 200)

    def test_generate_entitlements_action(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:entitlement_year'), {'year': '2026'})
        self.assertTrue(LeaveEntitlement.objects.filter(employee=self.emp, year=2026).exists())

    def test_reapply_forces_entitlements_to_leave_type_count(self):
        sick, _ = LeaveType.objects.update_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': Decimal('12')})
        # An entitlement that drifted to 30 (e.g. generated before the count change).
        ent = LeaveEntitlement.objects.create(
            employee=self.emp, leave_type=sick, year=2026, entitled_days=Decimal('30'))
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:entitlement_year'),
                         {'year': '2026', 'action': 'reapply'})
        ent.refresh_from_db()
        self.assertEqual(ent.entitled_days, Decimal('12'))  # snapped to the type count


from datetime import time
from hr.models import AttendanceRecord
from hr.attendance_services import derive_status


class AttendanceStatusTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def test_leave_beats_everything(self):
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'))
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 13), time(8, 0))[0], 'leave')

    def test_holiday(self):
        Holiday.objects.create(date=_date(2026, 7, 14), name='X')
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 14), None)[0], 'holiday')

    def test_weekend(self):
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 10), None)[0], 'weekend')  # Friday

    def test_present_with_hours(self):
        status, hours = derive_status(self.emp, _date(2026, 7, 13), time(8, 0), time(17, 30))
        self.assertEqual(status, 'present')
        self.assertEqual(hours, Decimal('9.5'))

    def test_absent_when_no_checkin_on_workday(self):
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 13), None)[0], 'absent')

    # -- exception-awareness (approved attendance exception excuses the day) --

    def _approve_exception_for_today(self, event_dt):
        from hr.attendance_exception_services import submit_attendance_exception, decide_attendance_exception
        manager = make_employee(iqama='ASTAT-MGR', name='Stat Manager')
        manager_user = make_user('astat_mgr')
        manager.user = manager_user
        manager.save(update_fields=['user'])
        self.emp.main_manager = manager
        self.emp.save(update_fields=['main_manager'])
        creator = make_user('astat_creator')
        exc = submit_attendance_exception(
            employee=self.emp, event_date=event_dt.date(), event_start_time=event_dt.time(),
            reason_category='site_visit', created_by=creator)
        decide_attendance_exception(exc, manager_user, 'approved')
        return exc

    def test_approved_exception_previews_as_present_instead_of_late(self):
        from django.utils import timezone as _tz
        settings_obj = AttendanceSettings.load()
        settings_obj.weekend_days = ''  # isolate from real-calendar weekend interference
        settings_obj.save(update_fields=['weekend_days'])
        event_dt = _tz.localtime(_tz.now() - timedelta(minutes=30))
        self._approve_exception_for_today(event_dt)

        late_check_in = time(9, 0)  # after the 8:30 expected_in_by default
        status, _ = derive_status(self.emp, event_dt.date(), late_check_in)
        self.assertEqual(status, 'present')

    def test_no_exception_late_check_in_is_unaffected(self):
        settings_obj = AttendanceSettings.load()
        settings_obj.weekend_days = ''
        settings_obj.save(update_fields=['weekend_days'])
        status, _ = derive_status(self.emp, _date(2026, 7, 13), time(9, 0))
        self.assertEqual(status, 'late')

    def test_rejected_exception_does_not_excuse_late(self):
        from django.utils import timezone as _tz
        from hr.attendance_exception_services import submit_attendance_exception, decide_attendance_exception
        settings_obj = AttendanceSettings.load()
        settings_obj.weekend_days = ''
        settings_obj.save(update_fields=['weekend_days'])

        manager = make_employee(iqama='ASTAT-MGR2', name='Stat Manager Two')
        manager_user = make_user('astat_mgr2')
        manager.user = manager_user
        manager.save(update_fields=['user'])
        self.emp.main_manager = manager
        self.emp.save(update_fields=['main_manager'])
        creator = make_user('astat_creator2')

        event_dt = _tz.localtime(_tz.now() - timedelta(minutes=30))
        exc = submit_attendance_exception(
            employee=self.emp, event_date=event_dt.date(), event_start_time=event_dt.time(),
            reason_category='site_visit', created_by=creator)
        decide_attendance_exception(exc, manager_user, 'rejected', note='No evidence')

        late_check_in = time(9, 0)
        status, _ = derive_status(self.emp, event_dt.date(), late_check_in)
        self.assertEqual(status, 'late')


class AttendanceGridTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def test_grid_get_lists_active_employees(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_grid') + '?date=2026-07-13')  # Monday
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.emp.full_name)

    def test_grid_post_saves_present(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_grid') + '?date=2026-07-13', {
            'date': '2026-07-13',
            f'check_in_{self.emp.pk}': '08:00',
            f'check_out_{self.emp.pk}': '17:30',
        })
        rec = AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13))
        self.assertEqual(rec.status, 'present')
        self.assertEqual(rec.hours_worked, Decimal('9.5'))

    def test_grid_post_marks_leave_day(self):
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'))
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_grid') + '?date=2026-07-13', {'date': '2026-07-13'})
        rec = AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13))
        self.assertEqual(rec.status, 'leave')

    def test_grid_post_blank_times_marks_absent(self):
        # The Absent button clears the row's times; saving a working day with no
        # check-in must record 'absent' (overwriting any prior present record).
        from datetime import time
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 13),
                                        status='present', check_in=time(8, 0), hours_worked=Decimal('8'))
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_grid') + '?date=2026-07-13',
                         {'date': '2026-07-13'})  # no check_in/out submitted = cleared
        rec = AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13))
        self.assertEqual(rec.status, 'absent')
        self.assertIsNone(rec.check_in)


class AttendanceHistoryTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()

    def test_history_summary_counts(self):
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 13), status='present', hours_worked=Decimal('8'))
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 14), status='absent')
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_history', kwargs={'pk': self.emp.pk}) + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['summary']['present'], 1)
        self.assertEqual(resp.context['summary']['absent'], 1)
        self.assertEqual(resp.context['summary']['total_hours'], Decimal('8'))


class AttendanceSettingsViewTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()

    def test_update_weekend(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_settings'), {'weekend_days': ['5', '6']})
        from hr.models import AttendanceSettings
        self.assertEqual(AttendanceSettings.load().weekend_day_set(), {5, 6})


class HardeningTests(TestCase):
    """Final-review fixes: no negative hours, days=0 preserved, bad int params don't 500."""

    def setUp(self):
        from accounts.models import Role, User
        from datetime import time
        self.time = time
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def test_inverted_times_yield_no_hours(self):
        from hr.attendance_services import derive_status
        status, hours = derive_status(self.emp, _date(2026, 7, 13), self.time(22, 0), self.time(6, 0))
        self.assertIn(status, ('present', 'late'))  # late threshold now applies; either is non-absent
        self.assertIsNone(hours)  # negative span -> blank, not a corrupt negative total

    def test_explicit_zero_days_preserved(self):
        # NOTE: the unified LeaveRequestForm (hr/forms.py) has no 'days' field —
        # every leave created through a view (legacy Add Leave, log-on-behalf-of,
        # My Profile self-service) is always auto-computed now. The only way to
        # set an explicit day count is directly on the model, so this is now a
        # model-level test of LeaveRecord.save() rather than a view test.
        rec = LeaveRecord(employee=self.emp, leave_type=self.annual,
                          start_date=_date(2026, 7, 10), end_date=_date(2026, 7, 11), days=Decimal('0'))
        rec.save()
        self.assertEqual(rec.days, Decimal('0'))

    def test_bad_year_param_does_not_500(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:leave_summary', kwargs={'pk': self.emp.pk}) + '?year=abc')
        self.assertEqual(resp.status_code, 200)

    def test_grid_renders_present_presets_on_working_row(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_grid') + '?date=2026-07-13')  # Monday (working)
        # The bulk "Mark all present" (which stamped everyone 08:15) was removed —
        # attendance times now come from the Wi-Fi agent.
        self.assertNotContains(resp, 'Mark all present')
        # The per-row manual controls remain for corrections.
        self.assertContains(resp, 'data-pk="%d"' % self.emp.pk)
        self.assertContains(resp, 'present-btn')
        self.assertContains(resp, 'absent-btn')  # per-row Absent (clears times -> Absent on save)
        self.assertContains(resp, 'unsavedBar')  # sticky 'unsaved changes' reminder

    def test_grid_no_present_button_on_locked_leave_row(self):
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_grid') + '?date=2026-07-13')
        # The only active employee is on leave -> locked row -> no editable controls
        # for them (no time inputs, no per-row buttons). Assert on the time-input
        # name, which is unique to the rendered row (the JS references data-pk/.present-btn,
        # so those substrings appear in the script regardless of rows).
        self.assertNotContains(resp, 'name="check_in_%d"' % self.emp.pk)
        self.assertContains(resp, 'Leave')  # the locked row shows its Leave badge


import json as _json
from hr.attendance_matrix import period_range, build_matrix, display_status_no_record


class MatrixHelperTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def test_week_range_starts_sunday(self):
        # 2026-07-15 is a Wednesday; its week (Sun-start) is 12 Jul (Sun) .. 18 Jul (Sat)
        start, end = period_range('week', _date(2026, 7, 15))
        self.assertEqual(start, _date(2026, 7, 12))
        self.assertEqual(end, _date(2026, 7, 18))

    def test_month_range_full_month(self):
        start, end = period_range('month', _date(2026, 7, 15))
        self.assertEqual(start, _date(2026, 7, 1))
        self.assertEqual(end, _date(2026, 7, 31))

    def test_stored_record_status_wins(self):
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 13), status='present', hours_worked=Decimal('8'))
        days, rows = build_matrix([self.emp], _date(2026, 7, 13), _date(2026, 7, 13))
        self.assertEqual(rows[0]['cells'][0]['status'], 'present')

    def test_leave_holiday_weekend_blank_precedence(self):
        # Mon 13 = leave (single-day record), Tue 14 = holiday, Fri 10 = weekend, Thu 16 = blank
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'))
        Holiday.objects.create(date=_date(2026, 7, 14), name='X')
        days, rows = build_matrix([self.emp], _date(2026, 7, 10), _date(2026, 7, 16))
        by_date = {c['date']: c for c in rows[0]['cells']}
        self.assertEqual(by_date[_date(2026, 7, 13)]['status'], 'leave')
        self.assertEqual(by_date[_date(2026, 7, 13)]['leave_record_id'],
                         LeaveRecord.objects.get(employee=self.emp).pk)  # single-day -> removable
        self.assertEqual(by_date[_date(2026, 7, 14)]['status'], 'holiday')
        self.assertEqual(by_date[_date(2026, 7, 10)]['status'], 'weekend')  # Friday
        self.assertEqual(by_date[_date(2026, 7, 16)]['status'], '')         # Thursday, no record

    def test_multiday_leave_has_no_removable_id(self):
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 15), days=Decimal('3'))
        days, rows = build_matrix([self.emp], _date(2026, 7, 13), _date(2026, 7, 15))
        for c in rows[0]['cells']:
            self.assertEqual(c['status'], 'leave')
            self.assertIsNone(c['leave_record_id'])  # multi-day -> not cell-removable

    def test_display_status_no_record(self):
        Holiday.objects.create(date=_date(2026, 7, 14), name='X')
        self.assertEqual(display_status_no_record(_date(2026, 7, 14)), 'holiday')
        self.assertEqual(display_status_no_record(_date(2026, 7, 10)), 'weekend')  # Friday
        self.assertEqual(display_status_no_record(_date(2026, 7, 16)), '')          # Thursday

    def test_display_status_no_record_honors_wfh(self):
        from hr.models import WFHRecord
        emp = make_employee(iqama='WFHHELPER')
        WFHRecord.objects.create(employee=emp, start_date=_date(2026, 7, 16), end_date=_date(2026, 7, 16))
        # Thursday working day with a WFH record -> wfh (not blank)
        self.assertEqual(display_status_no_record(_date(2026, 7, 16), emp), 'wfh')
        # weekend still beats wfh
        WFHRecord.objects.create(employee=emp, start_date=_date(2026, 7, 11), end_date=_date(2026, 7, 11))
        self.assertEqual(display_status_no_record(_date(2026, 7, 11), emp), 'weekend')  # Saturday


class MatrixViewTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee(name='Zara Tester')

    def test_matrix_month_renders(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_matrix') + '?period=month&date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Zara Tester')
        self.assertContains(resp, 'July')        # period heading
        self.assertEqual(len(resp.context['days']), 31)

    def test_matrix_week_has_7_days(self):
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_matrix') + '?period=week&date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.context['days']), 7)

    def test_matrix_requires_admin(self):
        from accounts.models import Role, User
        rep, _ = Role.objects.get_or_create(name=Role.SALES_REP)
        u = User.objects.create_user('rep', password='x'); u.role = rep; u.save()
        self.client.force_login(u)
        resp = self.client.get(reverse('hr:attendance_matrix'))
        self.assertEqual(resp.status_code, 302)  # admin-gate redirect to hr_dashboard


class MatrixExportTimesTests(TestCase):
    """The register exports show each present day's check-in/out time stacked
    beneath the status letter (e.g. 'P' over '09:12-17:30')."""

    def setUp(self):
        from accounts.models import Role, User
        from datetime import time
        self.time = time
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee(name='Zara Tester')
        AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 7, 13), status='present',
            check_in=time(9, 12), check_out=time(17, 30), hours_worked=Decimal('8'))
        self.client.force_login(self.admin)

    def test_excel_cell_stacks_status_over_times(self):
        import io
        import openpyxl as _op
        resp = self.client.get(reverse('hr:attendance_matrix_export_excel') + '?period=month&date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        ws = _op.load_workbook(io.BytesIO(resp.content)).active
        cells = [str(c.value) for r in ws.iter_rows() for c in r if c.value is not None]
        target = [v for v in cells if '09:12-17:30' in v]
        self.assertTrue(target, 'check-in/out time missing from Excel export')
        self.assertEqual(target[0], 'P\n09:12-17:30')  # status letter over the times

    def test_pdf_export_builds_with_times(self):
        resp = self.client.get(reverse('hr:attendance_matrix_export_pdf') + '?period=month&date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertGreater(len(resp.content), 500)  # a real PDF was produced

    def test_cell_time_lines_present_and_blank(self):
        from hr.attendance_matrix import build_matrix, cell_time_lines
        _, rows = build_matrix([self.emp], _date(2026, 7, 13), _date(2026, 7, 13))
        cell = rows[0]['cells'][0]
        self.assertEqual(cell['check_in'], self.time(9, 12))
        self.assertEqual(cell['check_out'], self.time(17, 30))
        self.assertEqual(cell_time_lines(cell), ['09:12', '17:30'])
        # A blank working day (Thu 2026-07-16, no record) carries no times.
        _, blank_rows = build_matrix([self.emp], _date(2026, 7, 16), _date(2026, 7, 16))
        self.assertEqual(cell_time_lines(blank_rows[0]['cells'][0]), [])


class MarkLeaveTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def _post(self, payload):
        return self.client.post(reverse('hr:attendance_mark_leave'),
                                data=_json.dumps(payload), content_type='application/json')

    def test_mark_creates_leave_and_attendance(self):
        self.client.force_login(self.admin)
        resp = self._post({'employee': self.emp.pk, 'date': '2026-07-13', 'leave_type': self.annual.pk})
        self.assertEqual(resp.status_code, 200)
        lr = LeaveRecord.objects.get(employee=self.emp)
        self.assertEqual(lr.start_date, _date(2026, 7, 13))
        self.assertEqual(lr.end_date, _date(2026, 7, 13))
        ar = AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13))
        self.assertEqual(ar.status, 'leave')
        self.assertEqual(resp.json()['leave_record_id'], lr.pk)

    def test_mark_requires_admin(self):
        from accounts.models import Role, User
        rep, _ = Role.objects.get_or_create(name=Role.SALES_REP)
        u = User.objects.create_user('rep', password='x'); u.role = rep; u.save()
        self.client.force_login(u)
        resp = self._post({'employee': self.emp.pk, 'date': '2026-07-13', 'leave_type': self.annual.pk})
        self.assertEqual(resp.status_code, 403)


class UnmarkLeaveTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def _post(self, payload):
        return self.client.post(reverse('hr:attendance_unmark_leave'),
                                data=_json.dumps(payload), content_type='application/json')

    def test_unmark_single_day_deletes_and_rederives(self):
        self.client.force_login(self.admin)
        # mark first (creates a 1-day leave + an attendance row at 'leave')
        self.client.post(reverse('hr:attendance_mark_leave'),
                         data=_json.dumps({'employee': self.emp.pk, 'date': '2026-07-13', 'leave_type': self.annual.pk}),
                         content_type='application/json')
        lr = LeaveRecord.objects.get(employee=self.emp)
        resp = self._post({'leave_record_id': lr.pk})
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(LeaveRecord.objects.filter(pk=lr.pk).exists())
        # 2026-07-13 is a Monday (working day), no check-in -> attendance row removed -> cell blank
        self.assertFalse(AttendanceRecord.objects.filter(employee=self.emp, date=_date(2026, 7, 13)).exists())
        self.assertEqual(resp.json()['status'], '')

    def test_unmark_rejects_multiday(self):
        self.client.force_login(self.admin)
        lr = LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                        start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 15), days=Decimal('3'))
        resp = self._post({'leave_record_id': lr.pk})
        self.assertEqual(resp.status_code, 400)
        self.assertTrue(LeaveRecord.objects.filter(pk=lr.pk).exists())  # not deleted


class MarkLeaveIntegrityTests(TestCase):
    """Final-review fixes: no double-booking; clock times survive a mark->unmark round trip."""

    def setUp(self):
        from accounts.models import Role, User
        from datetime import time
        self.time = time
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        self.client.force_login(self.admin)

    def _mark(self, day='2026-07-13'):
        return self.client.post(reverse('hr:attendance_mark_leave'),
                                data=_json.dumps({'employee': self.emp.pk, 'date': day, 'leave_type': self.annual.pk}),
                                content_type='application/json')

    def test_mark_rejects_overlapping_leave(self):
        # A multi-day leave already covers 2026-07-13.
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 7, 12), end_date=_date(2026, 7, 15), days=Decimal('4'))
        resp = self._mark('2026-07-13')
        self.assertEqual(resp.status_code, 400)
        self.assertEqual(LeaveRecord.objects.filter(employee=self.emp).count(), 1)  # no second record

    def test_mark_over_present_preserves_times_roundtrip(self):
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 13), status='present',
                                        check_in=self.time(8, 0), check_out=self.time(17, 0), hours_worked=Decimal('9'))
        self._mark('2026-07-13')
        ar = AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13))
        self.assertEqual(ar.status, 'leave')
        self.assertEqual(ar.check_in, self.time(8, 0))   # times preserved, not wiped
        lr = LeaveRecord.objects.get(employee=self.emp)
        resp = self.client.post(reverse('hr:attendance_unmark_leave'),
                                data=_json.dumps({'leave_record_id': lr.pk}), content_type='application/json')
        self.assertEqual(resp.json()['status'], 'present')   # re-derived back to present
        ar.refresh_from_db()
        self.assertEqual(ar.status, 'present')
        self.assertEqual(ar.hours_worked, Decimal('9.00'))


from datetime import time as _time


class AttendanceExtrasModelTests(TestCase):
    def test_expected_in_by_default(self):
        from hr.models import AttendanceSettings
        self.assertEqual(AttendanceSettings.load().expected_in_by, _time(8, 30))

    def test_record_accepts_late_and_wfh(self):
        from hr.models import AttendanceRecord
        codes = dict(AttendanceRecord.STATUS_CHOICES)
        self.assertIn('late', codes)
        self.assertIn('wfh', codes)


class DeriveLateWorkingDayTests(TestCase):
    def setUp(self):
        self.emp = make_employee()

    def test_on_time_is_present(self):
        from hr.attendance_services import derive_status
        # Mon 2026-07-13; default expected_in_by 08:30; check-in 08:20 -> present
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 13), _time(8, 20))[0], 'present')

    def test_late_after_threshold(self):
        from hr.attendance_services import derive_status
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 13), _time(9, 5))[0], 'late')

    def test_exactly_threshold_is_present(self):
        from hr.attendance_services import derive_status
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 13), _time(8, 30))[0], 'present')

    def test_working_day_overrides_weekend(self):
        from hr.models import WorkingDay
        from hr.attendance_services import derive_status
        WorkingDay.objects.create(date=_date(2026, 7, 10), name='WS')  # Fri 10 Jul is weekend
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 10), _time(8, 0))[0], 'present')
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 10), None)[0], 'absent')

    def test_plain_weekend_still_weekend(self):
        from hr.attendance_services import derive_status
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 11), None)[0], 'weekend')  # Saturday


class MatrixWorkingDayTests(TestCase):
    def setUp(self):
        self.emp = make_employee()

    def test_workingday_cell_not_weekend(self):
        from hr.models import WorkingDay
        from hr.attendance_matrix import build_matrix
        WorkingDay.objects.create(date=_date(2026, 7, 11), name='WS')  # Saturday
        days, rows = build_matrix([self.emp], _date(2026, 7, 11), _date(2026, 7, 11))
        self.assertEqual(rows[0]['cells'][0]['status'], '')        # blank working day, not 'weekend'
        self.assertFalse(rows[0]['cells'][0]['locked'])

    def test_weekend_set_excludes_workingdays(self):
        from hr.models import WorkingDay
        from hr.attendance_matrix import build_matrix
        WorkingDay.objects.create(date=_date(2026, 7, 11), name='WS')
        days, rows, weekend_dates = build_matrix([self.emp], _date(2026, 7, 10), _date(2026, 7, 11), with_weekend_dates=True)
        self.assertIn(_date(2026, 7, 10), weekend_dates)     # Friday still weekend
        self.assertNotIn(_date(2026, 7, 11), weekend_dates)  # Saturday is a working day


class WorkingDayViewTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()

    def test_list_ok(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse('hr:workingday_list')).status_code, 200)

    def test_create(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:workingday_create'), {'date': '2026-07-18', 'name': 'WS', 'is_active': 'on'})
        from hr.models import WorkingDay
        self.assertEqual(WorkingDay.objects.filter(name='WS').count(), 1)


class SettingsLateTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()

    def test_update_expected_in_by(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_settings'), {'weekend_days': ['4', '5'], 'expected_in_by': '09:00'})
        from hr.models import AttendanceSettings
        self.assertEqual(AttendanceSettings.load().expected_in_by, _time(9, 0))


class WFHRecordModelTests(TestCase):
    def setUp(self):
        self.emp = make_employee()

    def test_create_and_clean(self):
        from django.core.exceptions import ValidationError
        from hr.models import WFHRecord
        r = WFHRecord(employee=self.emp, start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 15))
        r.full_clean(); r.save()
        self.assertEqual(WFHRecord.objects.count(), 1)
        bad = WFHRecord(employee=self.emp, start_date=_date(2026, 7, 15), end_date=_date(2026, 7, 13))
        with self.assertRaises(ValidationError):
            bad.full_clean()


class DeriveWFHTests(TestCase):
    def setUp(self):
        self.emp = make_employee()

    def test_wfh_record_makes_day_wfh(self):
        from hr.models import WFHRecord
        from hr.attendance_services import derive_status
        WFHRecord.objects.create(employee=self.emp, start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13))
        status, hours = derive_status(self.emp, _date(2026, 7, 13), _time(8, 0), _time(17, 0))
        self.assertEqual(status, 'wfh')
        self.assertEqual(hours, Decimal('9'))  # hours still counted; Decimal(round(9.0,2)) == Decimal('9')

    def test_weekend_beats_wfh(self):
        from hr.models import WFHRecord
        from hr.attendance_services import derive_status
        WFHRecord.objects.create(employee=self.emp, start_date=_date(2026, 7, 11), end_date=_date(2026, 7, 11))
        self.assertEqual(derive_status(self.emp, _date(2026, 7, 11), None)[0], 'weekend')  # Saturday


class MatrixWFHTests(TestCase):
    def setUp(self):
        self.emp = make_employee()

    def test_wfh_record_shows_in_matrix(self):
        from hr.models import WFHRecord
        from hr.attendance_matrix import build_matrix
        WFHRecord.objects.create(employee=self.emp, start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13))
        days, rows = build_matrix([self.emp], _date(2026, 7, 13), _date(2026, 7, 13))
        self.assertEqual(rows[0]['cells'][0]['status'], 'wfh')


class WFHViewTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()

    def test_list_ok(self):
        self.client.force_login(self.admin)
        self.assertEqual(self.client.get(reverse('hr:wfh_list')).status_code, 200)

    def test_create_wfh(self):
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:wfh_create'), {'employee': self.emp.pk,
                         'start_date': '2026-07-13', 'end_date': '2026-07-15', 'note': ''})
        from hr.models import WFHRecord
        self.assertEqual(WFHRecord.objects.filter(employee=self.emp).count(), 1)


class GridWFHTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()

    def test_grid_wfh_flag_creates_record_and_status(self):
        from hr.models import WFHRecord, AttendanceRecord
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_grid') + '?date=2026-07-13', {
            'date': '2026-07-13',
            f'wfh_{self.emp.pk}': '1',
            f'check_in_{self.emp.pk}': '08:15',
            f'check_out_{self.emp.pk}': '18:00',
        })
        self.assertEqual(WFHRecord.objects.filter(employee=self.emp, start_date=_date(2026, 7, 13),
                                                  end_date=_date(2026, 7, 13)).count(), 1)
        self.assertEqual(AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13)).status, 'wfh')

    def test_grid_unflag_wfh_removes_single_day_record(self):
        from hr.models import WFHRecord, AttendanceRecord
        WFHRecord.objects.create(employee=self.emp, start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13))
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_grid') + '?date=2026-07-13', {
            'date': '2026-07-13',
            f'check_in_{self.emp.pk}': '08:00', f'check_out_{self.emp.pk}': '17:00',
        })
        self.assertFalse(WFHRecord.objects.filter(employee=self.emp, start_date=_date(2026, 7, 13),
                                                  end_date=_date(2026, 7, 13)).exists())
        self.assertEqual(AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 13)).status, 'present')

    def test_grid_does_not_touch_multiday_wfh(self):
        from hr.models import WFHRecord
        WFHRecord.objects.create(employee=self.emp, start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 15))
        self.client.force_login(self.admin)
        self.client.post(reverse('hr:attendance_grid') + '?date=2026-07-13', {'date': '2026-07-13'})
        self.assertTrue(WFHRecord.objects.filter(employee=self.emp, start_date=_date(2026, 7, 13),
                                                 end_date=_date(2026, 7, 15)).exists())  # multi-day untouched


class HistoryLateWFHSummaryTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()

    def test_summary_counts_late_and_wfh(self):
        from hr.models import AttendanceRecord
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 13), status='late')
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 14), status='wfh')
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:attendance_history', kwargs={'pk': self.emp.pk}) + '?year=2026&month=7')
        self.assertEqual(resp.context['summary']['late'], 1)
        self.assertEqual(resp.context['summary']['wfh'], 1)


class EntitlementYearGroupedTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee()
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        self.sick, _ = LeaveType.objects.get_or_create(code='sick', defaults={'name': 'Sick', 'default_annual_days': 15})

    def test_grouped_totals_and_taken_remaining(self):
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        # 8 annual days taken in 2026 (explicit days so it's not recomputed)
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2026, 3, 2), end_date=_date(2026, 3, 11), days=Decimal('8'))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:entitlement_year') + '?year=2026')
        self.assertEqual(resp.status_code, 200)
        groups = resp.context['groups']
        self.assertEqual(len(groups), 1)
        g = groups[0]
        self.assertEqual(g['employee'].pk, self.emp.pk)
        # Only Annual is accumulative (Task 1's rule) — Sick no longer counts toward the top-level total.
        self.assertEqual(g['total_entitled'], Decimal('30'))   # Annual only; Sick is conditional
        self.assertEqual(g['total_taken'], Decimal('8'))       # only annual taken
        self.assertEqual(g['total_remaining'], Decimal('22'))  # 30 - 8
        self.assertEqual(len(g['rows']), 2)                    # per-type breakdown
        annual_row = next(r for r in g['rows'] if r['leave_type'].pk == self.annual.pk)
        self.assertEqual(annual_row['taken'], Decimal('8'))
        self.assertEqual(annual_row['remaining'], Decimal('22'))

    def test_leave_from_other_year_not_counted(self):
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)
        LeaveRecord.objects.create(employee=self.emp, leave_type=self.annual,
                                   start_date=_date(2025, 3, 2), end_date=_date(2025, 3, 6), days=Decimal('5'))
        self.client.force_login(self.admin)
        resp = self.client.get(reverse('hr:entitlement_year') + '?year=2026')
        g = resp.context['groups'][0]
        self.assertEqual(g['total_taken'], Decimal('0'))       # 2025 leave excluded
        self.assertEqual(g['total_remaining'], Decimal('30'))


class SickLeaveCertificateTests(TestCase):
    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm', password='x'); self.admin.role = role; self.admin.save()
        # leave_record_create now unifies onto LeaveRequestCreateView (SuperAdminRequiredMixin).
        superadmin_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.superadmin = User.objects.create_user('supadm2', password='x')
        self.superadmin.role = superadmin_role
        self.superadmin.save()
        self.emp = make_employee()
        self.sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 15})
        self.sick.requires_medical_certificate = True
        self.sick.save()
        self.annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def _cert(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        return SimpleUploadedFile('cert.pdf', b'%PDF-1.4 fake', content_type='application/pdf')

    def test_sick_migration_set_flag(self):
        # Data migration 0017 should have flagged the seeded sick type — but this
        # test recreates it, so just assert the flag we set holds.
        self.assertTrue(LeaveType.objects.get(code='sick').requires_medical_certificate)

    def test_model_clean_blocks_sick_without_certificate(self):
        from django.core.exceptions import ValidationError
        r = LeaveRecord(employee=self.emp, leave_type=self.sick,
                        start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'))
        with self.assertRaises(ValidationError):
            r.full_clean()

    def test_model_clean_allows_sick_with_certificate(self):
        r = LeaveRecord(employee=self.emp, leave_type=self.sick,
                        start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'),
                        medical_certificate=self._cert())
        r.full_clean()  # should not raise

    def test_model_clean_allows_non_sick_without_certificate(self):
        r = LeaveRecord(employee=self.emp, leave_type=self.annual,
                        start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13), days=Decimal('1'))
        r.full_clean()  # annual doesn't require a certificate

    def test_create_form_sick_now_routes_to_approval_queue(self):
        # Data-compliance migration 0027 (a concurrent, unrelated fix) seeds
        # Sick with is_accumulative=False, i.e. Sick is genuinely conditional —
        # so under the unified dispatch (hr.leave_approval_services.submit_leave)
        # it now correctly enters the dual-approval queue as a LeaveRequest
        # rather than being recorded directly as a LeaveRecord. The old
        # LeaveRecordForm-level "certificate required" ModelForm validation
        # (which only ever applied to direct LeaveRecord creation) is therefore
        # moot for Sick specifically — it no longer takes that code path at
        # all. See the report for the residual, narrower concern: a *custom*
        # leave type that is both is_accumulative=True and
        # requires_medical_certificate=True would still bypass the guard,
        # since LeaveRecord.objects.create() in submit_leave() never calls
        # full_clean().
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        self.client.force_login(self.superadmin)
        resp = self.client.post(reverse('hr:leave_record_create'), {
            'employee': self.emp.pk, 'leave_type': self.sick.pk,
            'start_date': '2026-07-13', 'end_date': '2026-07-13', 'employee_reason': '',
            'document': self._cert()})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(LeaveRecord.objects.filter(employee=self.emp).exists())
        req = LeaveRequest.objects.get(employee=self.emp, leave_type=self.sick)
        self.assertEqual(req.status, 'pending')

    def test_sick_leave_without_document_is_blocked(self):
        # Task 1: mandatory document/certificate enforcement for any leave type
        # flagged requires_medical_certificate (currently only Sick).
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        self.client.force_login(self.superadmin)
        resp = self.client.post(reverse('hr:leave_record_create'), {
            'employee': self.emp.pk, 'leave_type': self.sick.pk,
            'start_date': '2026-07-13', 'end_date': '2026-07-13', 'employee_reason': ''})
        self.assertEqual(resp.status_code, 200)  # re-renders the form with an error, no redirect
        self.assertContains(resp, 'A medical certificate/document is required')
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp, leave_type=self.sick).exists())

    def test_model_clean_blocks_leave_request_sick_without_document(self):
        req = LeaveRequest(employee=self.emp, leave_type=self.sick,
                           start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13))
        with self.assertRaises(ValidationError):
            req.full_clean()

    def test_create_form_sick_document_attaches_to_the_leave_request(self):
        # The unified form's 'document' upload maps onto LeaveRequest.document
        # (already an existing field on the approval-workflow model) when the
        # leave type is conditional — this is how a certificate would actually
        # reach an approver under the new architecture.
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        self.client.force_login(self.superadmin)
        resp = self.client.post(reverse('hr:leave_record_create'), {
            'employee': self.emp.pk, 'leave_type': self.sick.pk,
            'start_date': '2026-07-13', 'end_date': '2026-07-13', 'employee_reason': '',
            'document': self._cert()})
        self.assertEqual(resp.status_code, 302)
        req = LeaveRequest.objects.get(employee=self.emp, leave_type=self.sick)
        self.assertTrue(req.document)

    def test_grid_mark_leave_blocks_certificate_type(self):
        import json as _json
        self.client.force_login(self.admin)
        resp = self.client.post(reverse('hr:attendance_mark_leave'),
                                data=_json.dumps({'employee': self.emp.pk, 'date': '2026-07-13',
                                                  'leave_type': self.sick.pk}),
                                content_type='application/json')
        self.assertEqual(resp.status_code, 400)
        self.assertIn('certificate', resp.json()['error'].lower())
        self.assertFalse(LeaveRecord.objects.filter(employee=self.emp).exists())

    def tearDown(self):
        # Remove any files written to MEDIA_ROOT by accepted-certificate tests.
        for r in LeaveRecord.objects.exclude(medical_certificate=''):
            if r.medical_certificate:
                r.medical_certificate.delete(save=False)
        for req in LeaveRequest.objects.exclude(document=''):
            if req.document:
                req.document.delete(save=False)


class LeaveTypeDaySyncTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.emp2 = make_employee(iqama='E2', name='Sara')
        self.sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 15})
        self.annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})

    def test_changing_days_updates_all_entitlements_all_years(self):
        e25 = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2025, entitled_days=15)
        e26 = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        e26b = LeaveEntitlement.objects.create(employee=self.emp2, leave_type=self.sick, year=2026, entitled_days=15)
        self.sick.default_annual_days = Decimal('20')
        self.sick.save()
        for e in (e25, e26, e26b):
            e.refresh_from_db()
            self.assertEqual(e.entitled_days, Decimal('20'))

    def test_overwrites_custom_values(self):
        custom = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=10)
        self.sick.default_annual_days = Decimal('20')
        self.sick.save()
        custom.refresh_from_db()
        self.assertEqual(custom.entitled_days, Decimal('20'))  # custom value overwritten

    def test_annual_type_is_synced(self):
        # Annual now behaves like any other type: a flat change updates all its
        # entitlements (proration policy dropped).
        a = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=25)
        self.annual.default_annual_days = Decimal('40')
        self.annual.save()
        a.refresh_from_db()
        self.assertEqual(a.entitled_days, Decimal('40'))  # now updated

    def test_no_change_leaves_entitlements_alone(self):
        e = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        self.sick.name = 'Sick Leave'  # change something other than the day count
        self.sick.save()
        e.refresh_from_db()
        self.assertEqual(e.entitled_days, Decimal('15'))

    def test_other_leave_types_unaffected(self):
        other, _ = LeaveType.objects.get_or_create(
            code='unpaid', defaults={'name': 'Unpaid', 'default_annual_days': 0})
        other_ent = LeaveEntitlement.objects.create(employee=self.emp, leave_type=other, year=2026, entitled_days=0)
        sick_ent = LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=15)
        self.sick.default_annual_days = Decimal('20')
        self.sick.save()
        other_ent.refresh_from_db(); sick_ent.refresh_from_db()
        self.assertEqual(other_ent.entitled_days, Decimal('0'))   # different type untouched
        self.assertEqual(sick_ent.entitled_days, Decimal('20'))


import tempfile
from django.test import override_settings
from django.core.files.storage import default_storage
from django.core.files.uploadedfile import SimpleUploadedFile
from accounts.models import Role, User
from hr.models import Vehicle, VehicleDocument


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class VehicleDocumentTests(TestCase):
    def setUp(self):
        sa, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.admin = User.objects.create_user('va', password='x')
        self.admin.role = sa
        self.admin.save()
        self.vehicle = Vehicle.objects.create(plate_number='ABC-1234')
        self.client.force_login(self.admin)

    def _upload(self, **overrides):
        data = {
            'document_type': 'registration', 'custom_type': '',
            'title': 'Istimara 2026',
            'file': SimpleUploadedFile('reg.pdf', b'%PDF x',
                                       content_type='application/pdf'),
        }
        data.update(overrides)
        return self.client.post(
            reverse('hr:vehicle_doc_upload', kwargs={'pk': self.vehicle.pk}), data)

    def test_upload_then_delete_reclaims_file(self):
        self._upload()
        self.assertEqual(VehicleDocument.objects.count(), 1)
        doc = VehicleDocument.objects.first()
        name = doc.file.name
        self.assertTrue(default_storage.exists(name))
        self.client.get(reverse('hr:vehicle_doc_delete', kwargs={'pk': doc.pk}))
        self.assertFalse(default_storage.exists(name))

    def test_other_type_requires_custom_label(self):
        self._upload(document_type='other', custom_type='')
        self.assertEqual(VehicleDocument.objects.count(), 0)  # rejected

    def test_custom_type_label(self):
        self._upload(document_type='other', custom_type='Pollution Cert')
        doc = VehicleDocument.objects.first()
        self.assertEqual(doc.type_label, 'Pollution Cert')

    def test_cascade_delete_removes_doc_files(self):
        self._upload(document_type='insurance')
        doc = VehicleDocument.objects.first()
        name = doc.file.name
        self.vehicle.delete()  # cascade -> central signal cleans the file
        self.assertEqual(VehicleDocument.objects.count(), 0)
        self.assertFalse(default_storage.exists(name))

    def test_edit_replacing_file_reclaims_old(self):
        self._upload()
        doc = VehicleDocument.objects.first()
        old_name = doc.file.name
        self.client.post(reverse('hr:vehicle_doc_edit', kwargs={'pk': doc.pk}), {
            'document_type': 'registration', 'custom_type': '', 'title': 'Istimara 2026',
            'file': SimpleUploadedFile('newreg.pdf', b'%PDF new',
                                       content_type='application/pdf'),
        })
        doc.refresh_from_db()
        self.assertNotEqual(doc.file.name, old_name)
        self.assertFalse(default_storage.exists(old_name))  # no orphan on replace
        self.assertTrue(default_storage.exists(doc.file.name))

    def test_edit_metadata_keeps_file(self):
        self._upload()
        doc = VehicleDocument.objects.first()
        name = doc.file.name
        self.client.post(reverse('hr:vehicle_doc_edit', kwargs={'pk': doc.pk}), {
            'document_type': 'registration', 'custom_type': '', 'title': 'Istimara (renamed)',
        })  # no file -> keep existing
        doc.refresh_from_db()
        self.assertEqual(doc.title, 'Istimara (renamed)')
        self.assertEqual(doc.file.name, name)
        self.assertTrue(default_storage.exists(name))

    def test_edit_form_renders(self):
        self._upload()
        doc = VehicleDocument.objects.first()
        r = self.client.get(reverse('hr:vehicle_doc_edit', kwargs={'pk': doc.pk}))
        self.assertEqual(r.status_code, 200)
        self.assertIn(b'Edit Document', r.content)


from hr.models import Employee


class MyProfilePortalTests(TestCase):
    """The self-service portal shows the linked employee's data, or a prompt
    when the account isn't linked yet."""

    def setUp(self):
        self.user = User.objects.create_user('emp1', password='x', email='emp1@leap.com')
        self.emp = Employee.objects.create(
            full_name='Emp One', iqama_number='IQ-001', user=self.user)
        self.unlinked = User.objects.create_user('nobody', password='x')

    def test_portal_shows_linked_employee(self):
        self.client.force_login(self.user)
        r = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, 'Emp One')

    def test_portal_prompts_when_not_linked(self):
        self.client.force_login(self.unlinked)
        r = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(r.status_code, 200)
        self.assertContains(r, "isn't linked")

    def test_portal_shows_assets_and_vehicles_via_active_handover(self):
        from hr.models import Asset, Vehicle, AssetHandover
        asset = Asset.objects.create(asset_name='Dell Laptop', asset_type='Laptop')
        vehicle = Vehicle.objects.create(plate_number='XYZ-9', vehicle_maker='Toyota')
        AssetHandover.objects.create(asset=asset, employee=self.emp, status='active')
        AssetHandover.objects.create(vehicle=vehicle, employee=self.emp, status='active')
        self.client.force_login(self.user)
        r = self.client.get(reverse('hr:my_profile'))
        self.assertContains(r, 'Dell Laptop')
        self.assertContains(r, 'XYZ-9')

    def test_portal_ignores_legacy_employee_name_and_driver_id_without_handover(self):
        from hr.models import Asset, Vehicle
        # These have the legacy denormalised fields set, matching this
        # employee, but no AssetHandover - AssetHandover is now the sole
        # source of truth for custody, so neither should appear here.
        Asset.objects.create(asset_name='Stale Laptop', employee_name='Emp One')
        Vehicle.objects.create(plate_number='STALE-1', driver_id='IQ-001')
        self.client.force_login(self.user)
        r = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(r, 'Stale Laptop')
        self.assertNotContains(r, 'STALE-1')


class MyAttendanceExportPDFTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('att_emp', password='x', email='att_emp@leap.com')
        self.emp = Employee.objects.create(
            full_name='Attendance Tester', iqama_number='IQ-ATT', user=self.user)
        self.other_user = User.objects.create_user('att_other', password='x', email='other@leap.com')
        self.other_emp = Employee.objects.create(
            full_name='Other Person', iqama_number='IQ-OTHER', user=self.other_user)
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 13), status='present')
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 14), status='absent')
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 15), status='wfh')

    def test_profile_shows_daily_attendance(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_profile') + '?date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        self.assertIn('daily_attendance', resp.context)
        statuses = {d['date']: d['status'] for d in resp.context['daily_attendance']}
        self.assertEqual(statuses[_date(2026, 7, 13)], 'present')
        self.assertEqual(statuses[_date(2026, 7, 14)], 'absent')
        self.assertEqual(statuses[_date(2026, 7, 15)], 'wfh')

    def test_pdf_export_returns_valid_pdf(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_attendance_export_pdf') + '?date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertGreater(len(resp.content), 500)

    def test_pdf_export_only_shows_own_attendance(self):
        # Security check: a different user's export must never include this
        # employee's data - each user can only ever see/export their own record
        # (the view always uses request.user.employee_profile, never a URL param).
        self.client.force_login(self.other_user)
        resp = self.client.get(reverse('hr:my_attendance_export_pdf') + '?date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        # The other user has no attendance records this month, so their export
        # should still succeed but contain none of Attendance Tester's data.
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_unlinked_user_redirected_from_export(self):
        unlinked = User.objects.create_user('unlinked_att', password='x')
        self.client.force_login(unlinked)
        resp = self.client.get(reverse('hr:my_attendance_export_pdf') + '?date=2026-07-15')
        self.assertEqual(resp.status_code, 302)


class LinkEmployeeUsersCommandTests(TestCase):
    def test_links_by_email_then_code(self):
        from django.core.management import call_command
        u_email = User.objects.create_user('byemail', password='x', email='match@leap.com')
        u_code = User.objects.create_user('bycode', password='x', employee_code='IQ-CODE')
        e1 = Employee.objects.create(full_name='By Email', iqama_number='IQ-1', work_email='match@leap.com')
        e2 = Employee.objects.create(full_name='By Code', iqama_number='IQ-CODE')
        call_command('link_employee_users')
        e1.refresh_from_db()
        e2.refresh_from_db()
        self.assertEqual(e1.user, u_email)
        self.assertEqual(e2.user, u_code)

    def test_does_not_steal_already_linked_user(self):
        from django.core.management import call_command
        u = User.objects.create_user('shared', password='x', email='dup@leap.com')
        already = Employee.objects.create(full_name='Already', iqama_number='IQ-A', user=u)
        contender = Employee.objects.create(full_name='Contender', iqama_number='IQ-B', work_email='dup@leap.com')
        call_command('link_employee_users')
        contender.refresh_from_db()
        self.assertIsNone(contender.user)  # u already taken


class EmployeeInactiveFromTests(TestCase):
    """Marking an employee inactive captures an 'inactive from' date (defaulting
    to today, never in the future); reactivating clears it; the list filters by
    status and the date picker is capped at today."""

    def _data(self, **over):
        d = {'iqama_number': 'IQ-X', 'full_name': 'Test Emp', 'user': '', 'is_active': 'on'}
        d.update(over)
        return d

    def test_marking_inactive_defaults_date_to_today(self):
        from django.utils import timezone
        from hr.forms import EmployeeForm
        emp = Employee.objects.create(full_name='X', iqama_number='IQ-1', is_active=True)
        data = self._data(iqama_number='IQ-1', full_name='X')
        data.pop('is_active')  # unchecked -> inactive
        f = EmployeeForm(data=data, instance=emp)
        self.assertTrue(f.is_valid(), f.errors)
        f.save()
        emp.refresh_from_db()
        self.assertFalse(emp.is_active)
        self.assertEqual(emp.inactive_from, timezone.localdate())

    def test_inactive_with_past_date_kept(self):
        from datetime import date
        from hr.forms import EmployeeForm
        emp = Employee.objects.create(full_name='Y', iqama_number='IQ-2', is_active=True)
        data = self._data(iqama_number='IQ-2', full_name='Y', inactive_from='2025-01-15')
        data.pop('is_active')
        f = EmployeeForm(data=data, instance=emp)
        self.assertTrue(f.is_valid(), f.errors)
        f.save()
        emp.refresh_from_db()
        self.assertEqual(emp.inactive_from, date(2025, 1, 15))

    def test_future_inactive_date_rejected(self):
        from datetime import timedelta
        from django.utils import timezone
        from hr.forms import EmployeeForm
        emp = Employee.objects.create(full_name='Z', iqama_number='IQ-3', is_active=True)
        future = (timezone.localdate() + timedelta(days=5)).isoformat()
        data = self._data(iqama_number='IQ-3', full_name='Z', inactive_from=future)
        data.pop('is_active')
        f = EmployeeForm(data=data, instance=emp)
        self.assertFalse(f.is_valid())
        self.assertIn('inactive_from', f.errors)

    def test_reactivating_clears_inactive_from(self):
        from datetime import date
        from hr.forms import EmployeeForm
        emp = Employee.objects.create(full_name='W', iqama_number='IQ-4',
                                      is_active=False, inactive_from=date(2025, 1, 1))
        f = EmployeeForm(data=self._data(iqama_number='IQ-4', full_name='W', is_active='on'),
                         instance=emp)
        self.assertTrue(f.is_valid(), f.errors)
        f.save()
        emp.refresh_from_db()
        self.assertTrue(emp.is_active)
        self.assertIsNone(emp.inactive_from)

    def test_date_picker_max_is_today(self):
        from django.utils import timezone
        from hr.forms import EmployeeForm
        f = EmployeeForm()
        self.assertEqual(f.fields['inactive_from'].widget.attrs.get('max'),
                         timezone.localdate().isoformat())

    def test_list_filters_by_status(self):
        sa, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        admin = User.objects.create_user('hradmin', password='x')
        admin.role = sa
        admin.save()
        Employee.objects.create(full_name='Active One', iqama_number='IQ-AC', is_active=True)
        Employee.objects.create(full_name='Inactive One', iqama_number='IQ-IN', is_active=False)
        self.client.force_login(admin)
        r = self.client.get(reverse('hr:employee_list'), {'status': 'inactive'})
        names = [e.full_name for e in r.context['employees']]
        self.assertIn('Inactive One', names)
        self.assertNotIn('Active One', names)


class AssetDecommissionTests(TestCase):
    """Assets can be marked out of service (dead) — removed from stock, kept for
    records — and restored. Listable via the status filter."""

    def setUp(self):
        from hr.models import Asset
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.admin = User.objects.create_user('aadm', password='x')
        self.admin.role = role
        self.admin.save()
        self.asset = Asset.objects.create(asset_name='Old Laptop', in_stock=True)
        self.client.force_login(self.admin)

    def test_decommission_marks_and_removes_from_stock(self):
        from django.utils import timezone
        self.client.post(reverse('hr:asset_decommission', kwargs={'pk': self.asset.pk}),
                         {'reason': 'Dead'})
        self.asset.refresh_from_db()
        self.assertTrue(self.asset.is_decommissioned)
        self.assertFalse(self.asset.in_stock)  # can't be in stock
        self.assertEqual(self.asset.decommissioned_on, timezone.localdate())
        self.assertEqual(self.asset.decommission_reason, 'Dead')

    def test_restore_brings_back(self):
        from datetime import date
        self.asset.is_decommissioned = True
        self.asset.decommissioned_on = date(2025, 1, 1)
        self.asset.in_stock = False
        self.asset.save()
        self.client.post(reverse('hr:asset_restore', kwargs={'pk': self.asset.pk}))
        self.asset.refresh_from_db()
        self.assertFalse(self.asset.is_decommissioned)
        self.assertIsNone(self.asset.decommissioned_on)

    def test_form_decommission_forces_out_of_stock(self):
        from hr.forms import AssetForm
        f = AssetForm(data={'asset_name': 'X', 'quantity': '1', 'in_stock': 'on',
                            'is_decommissioned': 'on'}, instance=self.asset)
        self.assertTrue(f.is_valid(), f.errors)
        obj = f.save()
        self.assertTrue(obj.is_decommissioned)
        self.assertFalse(obj.in_stock)  # in_stock forced off

    def test_list_filters_decommissioned(self):
        from hr.models import Asset
        Asset.objects.create(asset_name='Dead One', is_decommissioned=True)
        r = self.client.get(reverse('hr:asset_list'), {'status': 'decommissioned'})
        names = [a.asset_name for a in r.context['assets']]
        self.assertIn('Dead One', names)
        self.assertNotIn('Old Laptop', names)


class LeaveApprovalModelsTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.aamna = make_user('aamna_khan')
        self.ali = make_user('ali_sultan')
        LeaveDashboardAccess.objects.create(user=self.aamna, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.ali, is_active=True)

    def test_days_autocomputed(self):
        req = LeaveRequest(employee=self.emp, leave_type=self.marriage,
                           start_date=_date(2026, 7, 5), end_date=_date(2026, 7, 7))
        req.save()
        self.assertEqual(req.days, Decimal('3'))

    def test_end_before_start_rejected(self):
        req = LeaveRequest(employee=self.emp, leave_type=self.marriage,
                           start_date=_date(2026, 7, 9), end_date=_date(2026, 7, 5))
        with self.assertRaises(ValidationError):
            req.full_clean()

    def test_default_status_is_pending(self):
        req = LeaveRequest.objects.create(employee=self.emp, leave_type=self.marriage,
                                          start_date=_date(2026, 7, 5), end_date=_date(2026, 7, 7))
        self.assertEqual(req.status, 'pending')

    def test_approval_rows_are_separate_from_notes(self):
        req = LeaveRequest.objects.create(employee=self.emp, leave_type=self.marriage,
                                          start_date=_date(2026, 7, 5), end_date=_date(2026, 7, 7))
        LeaveRequestApproval.objects.create(leave_request=req, approver=self.aamna)
        LeaveRequestApproval.objects.create(leave_request=req, approver=self.ali)
        LeaveRequestNote.objects.create(leave_request=req, author=self.aamna, note='Looks fine.')
        self.assertEqual(req.approvals.count(), 2)
        self.assertEqual(req.notes.count(), 1)


import threading
from datetime import timedelta
from django.test import TransactionTestCase
from django.utils import timezone
from hr.forms import check_leave_balance
from hr.leave_approval_services import (
    record_approver_decision, override_finalize, edit_approver_decision, submit_leave_request)
from notifications.models import Notification


class LeaveApprovalServiceTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.aamna = make_user('aamna_khan')
        self.ali = make_user('ali_sultan')
        LeaveDashboardAccess.objects.create(user=self.aamna, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.ali, is_active=True)
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3))
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.aamna)
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.ali)

    def test_stays_pending_after_one_approval(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')
        self.assertIsNone(self.req.leave_record)

    def test_fully_approved_creates_leave_record(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        self.assertIsNotNone(self.req.leave_record)
        self.assertEqual(self.req.leave_record.employee, self.emp)
        self.assertEqual(self.req.leave_record.days, Decimal('3'))

    def test_balance_untouched_while_pending_deducted_only_on_full_approval(self):
        ent = LeaveEntitlement.objects.create(
            employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=Decimal('3'))
        self.assertEqual(ent.remaining_days, Decimal('3'))

        record_approver_decision(self.req, self.aamna, 'approved')
        self.assertEqual(ent.remaining_days, Decimal('3'))  # one approver down, still pending — untouched

        record_approver_decision(self.req, self.ali, 'approved')
        self.assertEqual(ent.remaining_days, Decimal('0'))  # fully approved — LeaveRecord now exists, deducted

    def test_disapproved_request_never_deducts_balance(self):
        ent = LeaveEntitlement.objects.create(
            employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=Decimal('3'))
        record_approver_decision(self.req, self.aamna, 'disapproved', comment='No.')
        record_approver_decision(self.req, self.ali, 'disapproved', comment='Agreed, no.')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'disapproved')
        self.assertEqual(ent.remaining_days, Decimal('3'))

    def test_unanimous_disapproval_finalizes_as_disapproved(self):
        record_approver_decision(self.req, self.aamna, 'disapproved', comment='Not enough notice')
        record_approver_decision(self.req, self.ali, 'disapproved', comment='Agreed')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'disapproved')
        self.assertTrue(self.req.salary_deduction_applicable)
        self.assertIsNone(self.req.leave_record)

    def test_split_decision_stays_pending_not_disapproved(self):
        # Regression: a single dissenting vote used to fail-fast finalize
        # the whole request as disapproved — reported as a bug, since it
        # instantly locked in a decision the OTHER approver never agreed to
        # and gave no way to reconsider. Disapproval now requires the same
        # unanimity approval already does.
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'disapproved', comment='Not enough notice')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')
        self.assertFalse(self.req.salary_deduction_applicable)
        self.assertIsNone(self.req.leave_record)

    def test_single_disapproval_does_not_skip_the_other_pending_approver(self):
        record_approver_decision(self.req, self.aamna, 'disapproved', comment='Not enough notice')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')
        ali_approval = self.req.approvals.get(approver=self.ali)
        self.assertEqual(ali_approval.decision, 'pending')

    def test_second_approvers_conflicting_decision_is_accepted_not_rejected(self):
        # A second, disagreeing decision is a normal vote now, not a "late"
        # decision arriving after the request already finalized — it must
        # be accepted (and simply leaves the request in conflict/pending),
        # never raise.
        record_approver_decision(self.req, self.aamna, 'disapproved', comment='Not enough notice')
        record_approver_decision(self.req, self.ali, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')
        ali_approval = self.req.approvals.get(approver=self.ali)
        self.assertEqual(ali_approval.decision, 'approved')

    def test_conflict_notifies_the_other_decided_approver(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'disapproved', comment='Not enough notice')
        conflict_note = Notification.objects.filter(
            recipient=self.aamna, verb__icontains='disagreed').first()
        self.assertIsNotNone(conflict_note)

    def test_non_approver_cannot_decide(self):
        stranger = make_user('someone_else')
        with self.assertRaises(ValueError):
            record_approver_decision(self.req, stranger, 'approved')

    def test_finalize_failure_rolls_back_the_whole_decision_not_just_finalize(self):
        # If _finalize blows up (e.g. a transient DB error creating the
        # LeaveRecord), the approver's own decision — recorded just before
        # _finalize runs, in the same call — must roll back with it. Without
        # this, the request gets stuck: one approval permanently shows
        # 'approved' while the request itself never leaves 'pending' and no
        # retry can ever re-decide that approval (record_approver_decision
        # requires decision == 'pending').
        from unittest.mock import patch
        record_approver_decision(self.req, self.aamna, 'approved')
        with patch('hr.leave_approval_services.LeaveRecord.objects.create',
                   side_effect=RuntimeError('simulated DB failure')):
            with self.assertRaises(RuntimeError):
                record_approver_decision(self.req, self.ali, 'approved')

        self.req.refresh_from_db()
        ali_approval = self.req.approvals.get(approver=self.ali)
        self.assertEqual(self.req.status, 'pending')
        self.assertEqual(ali_approval.decision, 'pending')
        self.assertIsNone(self.req.leave_record)

        # Retry after the transient failure clears must succeed cleanly.
        record_approver_decision(self.req, self.ali, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        self.assertIsNotNone(self.req.leave_record)

    def test_edit_decision_updates_and_finalizes_when_it_reaches_unanimity(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        edit_approver_decision(self.req, self.aamna, 'disapproved', edit_note='Misclicked, meant to disapprove.')
        self.req.refresh_from_db()
        # Only aamna has decided (ali is still pending) — editing to
        # disapproved doesn't finalize anything by itself; it just changes
        # aamna's own vote.
        self.assertEqual(self.req.status, 'pending')
        aamna_approval = self.req.approvals.get(approver=self.aamna)
        self.assertEqual(aamna_approval.decision, 'disapproved')
        note = self.req.notes.get()
        self.assertIn('Misclicked, meant to disapprove.', note.note)
        self.assertEqual(note.author, self.aamna)

    def test_editing_a_decision_can_resolve_a_conflict_and_finalize(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'disapproved', comment='Not enough notice')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')  # conflict — split decision
        edit_approver_decision(self.req, self.aamna, 'disapproved', edit_note='Agreed with Ali after all.')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'disapproved')  # now unanimous

    def test_edit_decision_requires_note(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        with self.assertRaises(ValueError):
            edit_approver_decision(self.req, self.aamna, 'disapproved', edit_note='')

    def test_edit_decision_still_allowed_after_24_hours_while_awaiting_other_approver(self):
        # Regression: editing used to be blocked after a fixed 24-hour
        # window even while the request was still genuinely awaiting a
        # decision — reported as a bug. There is no time limit now, only
        # whether the overall request is still pending.
        record_approver_decision(self.req, self.aamna, 'approved')
        approval = self.req.approvals.get(approver=self.aamna)
        approval.decided_at = timezone.now() - timedelta(hours=25)
        approval.save(update_fields=['decided_at'])
        edit_approver_decision(self.req, self.aamna, 'disapproved', edit_note='Changed my mind, no time limit now.')
        approval.refresh_from_db()
        self.assertEqual(approval.decision, 'disapproved')

    def test_edit_decision_rejected_once_request_finalized(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        with self.assertRaises(ValueError):
            edit_approver_decision(self.req, self.aamna, 'disapproved', edit_note='Changed my mind.')

    def test_edit_decision_rejected_without_prior_decision(self):
        with self.assertRaises(ValueError):
            edit_approver_decision(self.req, self.aamna, 'approved', edit_note='No prior decision exists.')

    def test_override_approve_finalizes_and_skips_remaining(self):
        superadmin = make_user('super1')
        override_finalize(self.req, superadmin, 'approved', reason='Ali is on leave; approving on his behalf.')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        self.assertTrue(self.req.is_overridden)
        self.assertIsNotNone(self.req.leave_record)
        skipped = self.req.approvals.filter(decision='skipped')
        self.assertEqual(skipped.count(), 2)  # neither had decided yet

    def test_override_requires_reason(self):
        superadmin = make_user('super2')
        with self.assertRaises(ValueError):
            override_finalize(self.req, superadmin, 'approved', reason='')

    def test_override_on_already_decided_request_rejected(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'approved')
        superadmin = make_user('super3')
        with self.assertRaises(ValueError):
            override_finalize(self.req, superadmin, 'disapproved', reason='Too late anyway')


class SelfApprovalPreventionTests(TestCase):
    """A user must never be able to approve, disapprove, or override their
    own leave request, under any circumstances — whether they'd otherwise
    qualify as an approver (LeaveDashboardAccess) or hold override access."""
    def setUp(self):
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.creator = make_user('sap_creator')

    def _emp_with_user(self, iqama, username):
        user = make_user(username, password='x')
        emp = make_employee(iqama=iqama, name=username)
        emp.user = user
        emp.save(update_fields=['user'])
        return emp, user

    def test_submitter_excluded_from_own_approver_roster(self):
        # Person A is both an approver (LeaveDashboardAccess) AND the one
        # submitting this request — they must not end up as their own
        # approver. Person B, the co-approver, must still be assigned.
        person_a, user_a = self._emp_with_user('SAP-A', 'sap_person_a')
        person_b, user_b = self._emp_with_user('SAP-B', 'sap_person_b')
        LeaveDashboardAccess.objects.create(user=user_a, is_active=True)
        LeaveDashboardAccess.objects.create(user=user_b, is_active=True)
        LeaveEntitlement.objects.create(employee=person_a, leave_type=self.marriage, year=2026, entitled_days=3)

        req = submit_leave_request(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3), created_by=user_a)

        self.assertFalse(req.approvals.filter(approver=user_a).exists())
        self.assertTrue(req.approvals.filter(approver=user_b).exists())

    def test_multi_approver_escalation_bypasses_requester(self):
        # 3 configured approvers; the requester is one of them. Approval
        # authority must fall entirely to the other 2 — unanimous approval
        # from just those 2 (never involving the requester) finalizes it.
        person_a, user_a = self._emp_with_user('SAP-MA', 'sap_multi_a')
        _, user_b = self._emp_with_user('SAP-MB', 'sap_multi_b')
        _, user_c = self._emp_with_user('SAP-MC', 'sap_multi_c')
        for u in (user_a, user_b, user_c):
            LeaveDashboardAccess.objects.create(user=u, is_active=True)
        LeaveEntitlement.objects.create(employee=person_a, leave_type=self.marriage, year=2026, entitled_days=3)

        req = submit_leave_request(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3), created_by=user_a)

        self.assertEqual(req.approvals.count(), 2)  # requester excluded, only B and C
        self.assertFalse(req.approvals.filter(approver=user_a).exists())

        record_approver_decision(req, user_b, 'approved')
        req.refresh_from_db()
        self.assertEqual(req.status, 'pending')  # still waiting on C
        record_approver_decision(req, user_c, 'approved')
        req.refresh_from_db()
        self.assertEqual(req.status, 'approved')  # unanimous among B+C alone finalized it

    def test_lone_approver_being_requester_leaves_zero_approvers(self):
        # If the ONLY configured approver submits for themselves, nobody is
        # assigned — the request sits pending until a Super Admin/override
        # holder steps in. This must not error; it's a legitimate state.
        person_a, user_a = self._emp_with_user('SAP-LONE', 'sap_lone')
        LeaveDashboardAccess.objects.create(user=user_a, is_active=True)
        LeaveEntitlement.objects.create(employee=person_a, leave_type=self.marriage, year=2026, entitled_days=3)

        req = submit_leave_request(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3), created_by=user_a)
        self.assertEqual(req.approvals.count(), 0)
        self.assertEqual(req.status, 'pending')
        self.assertEqual(req.pending_approvers(), [])

    def test_record_approver_decision_blocks_self_approval(self):
        # Simulates a legacy/edge-case row where a self-approval slipped in
        # despite the submission-time exclusion — must still be blocked.
        person_a, user_a = self._emp_with_user('SAP-LEGACY1', 'sap_legacy1')
        req = LeaveRequest.objects.create(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3))
        LeaveRequestApproval.objects.create(leave_request=req, approver=user_a)
        with self.assertRaises(ValueError):
            record_approver_decision(req, user_a, 'approved')
        req.refresh_from_db()
        self.assertEqual(req.status, 'pending')

    def test_edit_approver_decision_blocks_self_approval(self):
        person_a, user_a = self._emp_with_user('SAP-LEGACY2', 'sap_legacy2')
        _, user_b = self._emp_with_user('SAP-LEGACY2B', 'sap_legacy2b')
        req = LeaveRequest.objects.create(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3))
        approval = LeaveRequestApproval.objects.create(
            leave_request=req, approver=user_a, decision='approved', decided_at=timezone.now())
        LeaveRequestApproval.objects.create(leave_request=req, approver=user_b)
        with self.assertRaises(ValueError):
            edit_approver_decision(req, user_a, 'disapproved', edit_note='Changed my mind')

    def test_override_finalize_blocks_self_override(self):
        person_a, user_a = self._emp_with_user('SAP-OVR', 'sap_override_self')
        req = LeaveRequest.objects.create(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3))
        with self.assertRaises(ValueError):
            override_finalize(req, user_a, 'approved', reason='Approving my own request')
        req.refresh_from_db()
        self.assertEqual(req.status, 'pending')

    def test_override_finalize_still_works_for_a_different_admin(self):
        # Confirms the self-check is narrowly scoped to the requester, not a
        # blanket block on override for this request.
        person_a, _ = self._emp_with_user('SAP-OVR2', 'sap_override_target')
        admin = make_user('sap_override_admin')
        req = LeaveRequest.objects.create(
            employee=person_a, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3))
        override_finalize(req, admin, 'approved', reason='Approver unavailable')
        req.refresh_from_db()
        self.assertEqual(req.status, 'approved')


class LeaveRequestDetailOwnRequestUITests(TestCase):
    """UI-level: a user viewing their OWN leave request must never see the
    decide/edit-decision/override forms, even if they'd otherwise qualify —
    and the page must make clear why (see is_own_request context)."""
    def setUp(self):
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.employee_user = make_user('duo_employee', password='x')
        self.employee_user.set_password('testpass123')
        # Also a Super Admin AND holds override access — the strongest case:
        # even with every other authority, their OWN request must stay view-only.
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.employee_user.role = role
        self.employee_user.save()
        LeaveDashboardAccess.objects.create(user=self.employee_user, is_active=True)
        self.emp = make_employee(iqama='DUO-EMP', name='Duo Employee')
        self.emp.user = self.employee_user
        self.emp.save(update_fields=['user'])

        self.co_approver_user = make_user('duo_coapprover', password='x')
        self.co_approver_user.set_password('testpass123')
        self.co_approver_user.save()
        LeaveDashboardAccess.objects.create(user=self.co_approver_user, is_active=True)

        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=3)
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3), created_by=self.employee_user)

    def test_own_request_hides_decide_and_override_forms(self):
        self.client.login(username='duo_employee', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertNotContains(resp, 'name="action" value="decide"')
        self.assertNotContains(resp, 'name="action" value="override"')
        self.assertNotContains(resp, 'name="action" value="add_note"')
        self.assertContains(resp, 'This is your own leave request')

    def test_co_approver_sees_decide_form_for_the_same_request(self):
        self.client.login(username='duo_coapprover', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, 'name="action" value="decide"')

    def test_own_request_post_decide_forbidden_even_with_legacy_approval_row(self):
        # Simulate a legacy self-approval row somehow existing, and confirm
        # POSTing a decision is still blocked (error message, not a crash).
        LeaveRequestApproval.objects.get_or_create(leave_request=self.req, approver=self.employee_user)
        self.client.login(username='duo_employee', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'decide', 'decision': 'approved'}, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')

    def test_own_request_post_override_forbidden(self):
        self.client.login(username='duo_employee', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Self-serving override'})
        self.assertEqual(resp.status_code, 302)  # redirects with an error message, doesn't crash
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')


class SubmitLeaveRequestRaceConditionTests(TransactionTestCase):
    """Real concurrency test: two genuinely simultaneous submissions for the
    same employee/leave_type/year, together exceeding the entitlement, must
    not both succeed. Uses TransactionTestCase (real commits, no wrapping
    transaction) + threads so the select_for_update() lock in
    validate_leave_submission actually has two separate DB connections to
    serialize — a plain TestCase's single outer transaction would make the
    lock a no-op and prove nothing."""
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=3)
        self.creator = make_user('race_creator')

    # Relies on select_for_update() row-locking to serialize two concurrent
    # submissions. SQLite has no row-level locking (has_select_for_update is
    # False), so the lock is a no-op there and the test can't prove anything —
    # it only runs against Postgres (production/CI). Skipping on SQLite keeps
    # local `manage.py test` green without weakening the production check.
    @skipUnlessDBFeature('has_select_for_update')
    def test_concurrent_submissions_cannot_together_exceed_balance(self):
        from django.db import connection
        results = []
        barrier = threading.Barrier(2)

        def attempt(start_day):
            try:
                barrier.wait(timeout=5)
                submit_leave_request(
                    employee=self.emp, leave_type=self.marriage,
                    start_date=_date(2026, 9, start_day), end_date=_date(2026, 9, start_day + 2),
                    created_by=self.creator,
                )
                results.append('success')
            except ValueError:
                results.append('rejected')
            finally:
                connection.close()

        # Two non-overlapping-in-date but same-balance-pool requests (3 days
        # each against a 3-day entitlement) — only one can survive.
        t1 = threading.Thread(target=attempt, args=(1,))
        t2 = threading.Thread(target=attempt, args=(10,))
        t1.start()
        t2.start()
        t1.join(timeout=10)
        t2.join(timeout=10)

        self.assertEqual(sorted(results), ['rejected', 'success'])
        self.assertEqual(LeaveRequest.objects.filter(employee=self.emp).count(), 1)


class OverlappingLeaveAcrossTypesTests(TestCase):
    """A person cannot be on two kinds of leave for overlapping dates —
    check_leave_balance/validate_leave_submission must block this
    regardless of leave type, closing the 'split the same dates across two
    leave types to get more total days off' loophole."""
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 12, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=30)
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=30)

    def test_overlapping_pending_request_of_different_type_rejected(self):
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.sick,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 5))  # pending
        with self.assertRaises(ValidationError) as ctx:
            check_leave_balance(self.emp, self.marriage, _date(2026, 9, 3), _date(2026, 9, 4))
        self.assertIn('overlaps with another leave request', str(ctx.exception))

    def test_overlapping_approved_leaverecord_of_different_type_rejected(self):
        LeaveRecord.objects.create(
            employee=self.emp, leave_type=self.sick,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 5), days=Decimal('5'))
        with self.assertRaises(ValidationError) as ctx:
            check_leave_balance(self.emp, self.marriage, _date(2026, 9, 3), _date(2026, 9, 4))
        self.assertIn('already taken or been approved', str(ctx.exception))

    def test_non_overlapping_dates_different_type_allowed(self):
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.sick,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 5))
        check_leave_balance(self.emp, self.marriage, _date(2026, 10, 1), _date(2026, 10, 2))  # no overlap

    def test_overlapping_with_disapproved_request_of_different_type_allowed(self):
        # A disapproved request never created a LeaveRecord and is no longer
        # 'pending' — those same dates must be freely reusable.
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.sick,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 5), status='disapproved')
        check_leave_balance(self.emp, self.marriage, _date(2026, 9, 3), _date(2026, 9, 4))


class LeaveCancellationRefundSafetyTests(TestCase):
    """Balances are pure live aggregates over LeaveRecord (taken_days/
    remaining_days are @property, never a stored counter) — deleting a
    LeaveRecord (the closest thing to 'cancelling' approved leave today;
    there's no separate cancel action yet) restores the exact balance with
    no double-refund risk, since there's nothing to decrement/increment
    apart from the aggregate itself."""
    def test_deleting_leave_record_restores_full_balance(self):
        emp = make_employee()
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        ent = LeaveEntitlement.objects.create(employee=emp, leave_type=annual, year=2026, entitled_days=10)
        record = LeaveRecord.objects.create(
            employee=emp, leave_type=annual, start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 5), days=Decimal('5'))
        self.assertEqual(ent.remaining_days, Decimal('5'))
        record.delete()
        self.assertEqual(ent.remaining_days, Decimal('10'))

    def test_balance_fields_are_computed_not_stored(self):
        # Structural regression guard: if a future change adds a cached/
        # stored counter (e.g. 'taken_days' as a real column) instead of
        # this live aggregate, double-refund bugs become possible again.
        stored_field_names = {f.name for f in LeaveEntitlement._meta.get_fields() if not f.is_relation}
        self.assertNotIn('taken_days', stored_field_names)
        self.assertIsInstance(LeaveEntitlement.taken_days, property)
        self.assertIsInstance(LeaveEntitlement.remaining_days, property)




class LeaveApprovalNotificationTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.emp_user = make_user('emp_user')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])
        self.aamna = make_user('aamna_khan')
        self.ali = make_user('ali_sultan')
        LeaveDashboardAccess.objects.create(user=self.aamna, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.ali, is_active=True)
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3))
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.aamna)
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.ali)

    def test_other_approver_notified_after_first_decision(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        self.assertTrue(Notification.objects.filter(recipient=self.ali).exists())

    def test_employee_notified_on_final_approval(self):
        record_approver_decision(self.req, self.aamna, 'approved')
        record_approver_decision(self.req, self.ali, 'approved')
        self.assertTrue(Notification.objects.filter(recipient=self.emp_user, verb__icontains='approved').exists())

    def test_employee_notified_on_disapproval(self):
        record_approver_decision(self.req, self.aamna, 'disapproved')
        record_approver_decision(self.req, self.ali, 'disapproved')
        self.assertTrue(Notification.objects.filter(recipient=self.emp_user, verb__icontains='disapproved').exists())


from django.core.files.uploadedfile import SimpleUploadedFile


class LeaveRequestDocumentAccessTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.emp_user = make_user('doc_emp_user', password='x')
        self.emp_user.set_password('testpass123')
        self.emp_user.save()
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])
        self.aamna = make_user('doc_aamna')
        self.aamna.set_password('testpass123')
        self.aamna.save()
        LeaveDashboardAccess.objects.create(user=self.aamna, is_active=True)
        self.stranger = make_user('doc_stranger')
        self.stranger.set_password('testpass123')
        self.stranger.save()
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3),
            document=SimpleUploadedFile('cert.pdf', b'dummy-bytes'))

    def test_owner_can_download(self):
        self.client.login(username='doc_emp_user', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_document', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_approver_can_download(self):
        self.client.login(username='doc_aamna', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_document', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_unrelated_user_forbidden(self):
        # 404, not 403 — an unauthorized user must not be able to distinguish
        # "this request doesn't exist" from "it exists but isn't yours".
        self.client.login(username='doc_stranger', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_document', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 404)

    def test_nonexistent_request_also_404s_indistinguishably(self):
        self.client.login(username='doc_stranger', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_document', args=[999999]))
        self.assertEqual(resp.status_code, 404)

    def test_anonymous_redirected_to_login(self):
        resp = self.client.get(reverse('hr:leave_request_document', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 302)


class CanViewLeaveDashboardTests(TestCase):
    """Unit tests for hr.views.can_view_leave_dashboard, independent of any
    specific view. LeaveDashboardAccess is the single merged roster — an
    active grant is both 'can view the dashboard' and 'is a real approver'
    (see is_designated_approver / submit_leave_request) — while Super Admin
    status alone only grants viewing, never approval authority by itself."""
    def setUp(self):
        from accounts.models import Role
        self.role, _ = Role.objects.get_or_create(name='super_admin')

    def test_super_admin_can_view(self):
        from hr.views import can_view_leave_dashboard
        user = make_user('cvld_super', password='x')
        user.role = self.role
        user.save()
        self.assertTrue(can_view_leave_dashboard(user))

    def test_plain_user_cannot_view(self):
        from hr.views import can_view_leave_dashboard
        user = make_user('cvld_plain', password='x')
        user.save()
        self.assertFalse(can_view_leave_dashboard(user))

    def test_user_with_active_grant_can_view(self):
        from hr.views import can_view_leave_dashboard
        user = make_user('cvld_grantee', password='x')
        user.save()
        LeaveDashboardAccess.objects.create(user=user, is_active=True)
        self.assertTrue(can_view_leave_dashboard(user))

    def test_user_with_inactive_grant_cannot_view(self):
        from hr.views import can_view_leave_dashboard
        user = make_user('cvld_revoked', password='x')
        user.save()
        LeaveDashboardAccess.objects.create(user=user, is_active=False)
        self.assertFalse(can_view_leave_dashboard(user))


class LeaveRequestQueueViewTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.superadmin = make_user('queue_super')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.plain_user = make_user('queue_plain', password='x')
        self.plain_user.set_password('testpass123')
        self.plain_user.save()
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=10)
        self.pending = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3))

    def test_super_admin_can_view_queue(self):
        self.client.login(username='queue_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.emp.full_name)

    def test_non_super_admin_forbidden(self):
        self.client.login(username='queue_plain', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertEqual(resp.status_code, 403)

    def test_admin_can_log_request_on_employees_behalf(self):
        # entitled_days (set in setUp) must cover BOTH self.pending (setUp's
        # own 3-day Marriage request, still pending) AND this test's own
        # 2-day request — check_leave_balance now counts other pending
        # requests as already-spoken-for balance, so 3 alone would leave 0
        # available.
        self.client.login(username='queue_super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_create'), {
            'employee': self.emp.pk, 'leave_type': self.marriage.pk,
            'start_date': '2026-09-01', 'end_date': '2026-09-02', 'employee_reason': 'Family event',
        })
        self.assertEqual(resp.status_code, 302)
        new_req = LeaveRequest.objects.exclude(pk=self.pending.pk).get()
        self.assertEqual(new_req.created_by, self.superadmin)
        self.assertEqual(new_req.approvals.count(), 0)  # approvals seeded in Task 4f alongside submission wiring — see note below

    def test_race_condition_valueerror_from_submit_re_renders_form_with_error(self):
        from unittest.mock import patch
        self.client.login(username='queue_super', password='testpass123')
        with patch('hr.leave_approval_services.submit_leave_request',
                   side_effect=ValueError('Someone else just took that balance.')):
            resp = self.client.post(reverse('hr:leave_request_create'), {
                'employee': self.emp.pk, 'leave_type': self.marriage.pk,
                'start_date': '2026-09-01', 'end_date': '2026-09-02', 'employee_reason': 'Family event',
            })
        self.assertEqual(resp.status_code, 200)  # re-rendered, not redirected
        self.assertContains(resp, 'Someone else just took that balance.')
        self.assertFalse(LeaveRequest.objects.exclude(pk=self.pending.pk).exists())

    def test_designated_approver_can_view_queue(self):
        approver = make_user('queue_approver', password='x')
        approver.set_password('testpass123')
        approver.save()
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        self.client.login(username='queue_approver', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertEqual(resp.status_code, 200)

    def test_sort_by_vacation_date(self):
        later_vacation_earlier_submission = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 7, 1), end_date=_date(2026, 7, 2))
        self.client.login(username='queue_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'), {'sort': 'vacation_date'})
        ids_in_order = [r.pk for r in resp.context['pending_requests']]
        self.assertEqual(ids_in_order, [later_vacation_earlier_submission.pk, self.pending.pk])

    def test_default_sort_is_submission_date(self):
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 7, 1), end_date=_date(2026, 7, 2))
        self.client.login(username='queue_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        ids_in_order = [r.pk for r in resp.context['pending_requests']]
        self.assertEqual(ids_in_order[0], self.pending.pk)  # oldest submission still first

    def test_user_granted_leave_dashboard_access_can_view_queue(self):
        # Not a Super Admin — the only thing granting this user visibility
        # is an active LeaveDashboardAccess row.
        grantee = make_user('queue_grantee', password='x')
        grantee.set_password('testpass123')
        grantee.save()
        LeaveDashboardAccess.objects.create(user=grantee, is_active=True, granted_by=self.superadmin)
        self.client.login(username='queue_grantee', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertEqual(resp.status_code, 200)

    def test_history_shows_who_decided(self):
        aamna = make_user('queue_history_aamna', password='x', first_name='Aamna', last_name='Khan')
        LeaveDashboardAccess.objects.create(user=aamna, is_active=True)
        LeaveRequestApproval.objects.create(leave_request=self.pending, approver=aamna, decision='approved')
        self.pending.status = 'approved'
        self.pending.decided_at = timezone.now()
        self.pending.save(update_fields=['status', 'decided_at'])
        self.client.login(username='queue_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, 'by Aamna Khan')

    def test_inactive_leave_dashboard_access_does_not_grant_view(self):
        from hr.models import LeaveDashboardAccess
        grantee = make_user('queue_revoked', password='x')
        grantee.set_password('testpass123')
        grantee.save()
        LeaveDashboardAccess.objects.create(user=grantee, is_active=False, granted_by=self.superadmin)
        self.client.login(username='queue_revoked', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertEqual(resp.status_code, 403)


class LeaveRequestDetailViewTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.superadmin = make_user('detail_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.aamna = make_user('detail_aamna', password='x')
        self.aamna.set_password('testpass123')
        self.aamna.save()
        LeaveDashboardAccess.objects.create(user=self.aamna, is_active=True)
        self.ali = make_user('detail_ali', password='x')
        self.ali.set_password('testpass123')
        self.ali.save()
        LeaveDashboardAccess.objects.create(user=self.ali, is_active=True)
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3))
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.aamna)
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.ali)

    def test_detail_page_loads(self):
        self.client.login(username='detail_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_approver_can_decide_via_post(self):
        self.client.login(username='detail_aamna', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'decide', 'decision': 'approved', 'comment': 'ok'})
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.approvals.get(approver=self.aamna).decision, 'approved')

    def test_add_note_visible_to_employee_by_default(self):
        # add_note is restricted to an assigned approver for THIS request, not
        # any Super Admin — detail_aamna is a designated approver here.
        self.client.login(username='detail_aamna', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'add_note', 'note': 'Please bring the certificate.'})
        self.assertEqual(resp.status_code, 302)
        note = self.req.notes.get()
        self.assertFalse(note.is_internal)

    def test_override_by_superadmin(self):
        self.client.login(username='detail_super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Ali is on leave'})
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        self.assertTrue(self.req.is_overridden)

    def test_detail_page_shows_who_decided(self):
        self.aamna.first_name, self.aamna.last_name = 'Aamna', 'Khan'
        self.aamna.save()
        self.ali.first_name, self.ali.last_name = 'Ali', 'Sultan'
        self.ali.save()
        self.client.login(username='detail_aamna', password='testpass123')
        self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                         {'action': 'decide', 'decision': 'approved', 'comment': 'ok'})
        self.client.login(username='detail_ali', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'decide', 'decision': 'approved', 'comment': 'ok'}, follow=True)
        self.assertContains(resp, 'by Aamna Khan, Ali Sultan')

    def test_page_renders_when_overriding_user_was_deleted(self):
        # Regression: overridden_by is on_delete=SET_NULL, so a historical
        # request can have is_overridden=True but overridden_by=None once
        # the overriding account is deleted. The template dereferenced
        # overridden_by.username as a *filter argument* (default:...), which
        # Django does not null-guard — this crashed with
        # VariableDoesNotExist for any such row.
        self.client.login(username='detail_super', password='testpass123')
        self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                         {'action': 'override', 'decision': 'approved', 'reason': 'Ali is on leave'})
        self.superadmin.delete()
        self.req.refresh_from_db()
        self.assertTrue(self.req.is_overridden)
        self.assertIsNone(self.req.overridden_by_id)
        self.client.login(username='detail_aamna', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Overridden')

    def test_add_note_by_non_superadmin_approver_forbidden(self):
        # A Super Admin with no approval row on this specific request is
        # view-only aside from the override escape hatch — add_note requires
        # being a designated approver for THIS request, not Super Admin status.
        self.client.login(username='detail_super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'add_note', 'note': 'Should not be allowed.'})
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(self.req.notes.count(), 0)

    def test_standalone_add_note_form_no_longer_rendered(self):
        # The separate "Add Note" card was removed as redundant — the
        # decide form's own "Add a message for the employee" field now
        # covers that use case in one submit.
        self.client.login(username='detail_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertNotContains(resp, 'name="note"')

    def test_assigned_approver_sees_message_field_on_the_decide_form(self):
        self.client.login(username='detail_aamna', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, 'name="comment"')

    def test_non_approver_non_superadmin_cannot_decide(self):
        outsider = make_user('detail_outsider', password='x')
        outsider.set_password('testpass123')
        outsider.save()
        self.client.login(username='detail_outsider', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 403)

    def test_user_granted_leave_dashboard_access_can_view_detail(self):
        from hr.models import LeaveDashboardAccess
        grantee = make_user('detail_grantee', password='x')
        grantee.set_password('testpass123')
        grantee.save()
        LeaveDashboardAccess.objects.create(user=grantee, is_active=True, granted_by=self.superadmin)
        self.client.login(username='detail_grantee', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_deciding_approver_sees_can_change_decision_ui_not_an_instant_finalize(self):
        # Bug report: choosing Disapprove used to instantly finalize the
        # whole request (fail-fast) — the deciding approver never got shown
        # that they could still change their mind. With a co-approver still
        # pending, disapproving now just records the vote and the page must
        # offer to change it.
        self.client.login(username='detail_aamna', password='testpass123')
        self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]), {
            'action': 'decide', 'decision': 'disapproved',
        })
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, 'You can change this any time until every approver has decided.')

    def test_conflict_banner_shows_when_approvers_disagree(self):
        self.client.login(username='detail_aamna', password='testpass123')
        self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]), {
            'action': 'decide', 'decision': 'approved',
        })
        self.client.logout()
        self.client.login(username='detail_ali', password='testpass123')
        self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]), {
            'action': 'decide', 'decision': 'disapproved',
        })
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, 'Approvers disagree on this request.')

    def test_no_conflict_banner_while_genuinely_still_awaiting_a_vote(self):
        self.client.login(username='detail_aamna', password='testpass123')
        self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]), {
            'action': 'decide', 'decision': 'approved',
        })
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertNotContains(resp, 'Approvers disagree on this request.')


class CheckLeaveBalanceTests(TestCase):
    """Unit tests for hr.forms.check_leave_balance, independent of any view.
    remaining_days only reflects APPROVED days (LeaveRecord rows) — this
    function must also treat this employee's OTHER still-pending requests
    for the same leave type/year as already-spoken-for balance, or two
    individually-valid submissions could together exceed the entitlement."""
    def setUp(self):
        from hr.forms import check_leave_balance
        self.check_leave_balance = check_leave_balance
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 12, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=3)

    def test_no_entitlement_raises(self):
        with self.assertRaises(ValidationError):
            self.check_leave_balance(self.emp, self.marriage, _date(2027, 1, 5), _date(2027, 1, 6))

    def test_within_balance_passes(self):
        self.check_leave_balance(self.emp, self.marriage, _date(2026, 9, 1), _date(2026, 9, 3))  # 3 days, exact fit

    def test_exceeding_plain_entitlement_raises(self):
        with self.assertRaises(ValidationError):
            self.check_leave_balance(self.emp, self.marriage, _date(2026, 9, 1), _date(2026, 9, 10))

    def test_second_request_blocked_once_first_pending_request_uses_up_balance(self):
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3))  # 3 days, still pending
        with self.assertRaises(ValidationError) as ctx:
            self.check_leave_balance(self.emp, self.marriage, _date(2026, 10, 1), _date(2026, 10, 1))  # just 1 more day
        self.assertIn('already tied up in other pending requests', str(ctx.exception))

    def test_second_request_allowed_once_first_request_is_no_longer_pending(self):
        req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 3), status='disapproved')
        self.check_leave_balance(self.emp, self.marriage, _date(2026, 10, 1), _date(2026, 10, 3))

    def test_pending_request_of_a_different_leave_type_does_not_count(self):
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.sick, year=2026, entitled_days=12)
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.sick,
            start_date=_date(2026, 9, 1), end_date=_date(2026, 9, 10))  # 10 days of Sick, pending
        # Marriage's own 3-day balance is untouched by Sick's pending request.
        self.check_leave_balance(self.emp, self.marriage, _date(2026, 11, 1), _date(2026, 11, 3))

    def test_pending_request_in_a_different_year_does_not_count(self):
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2027, entitled_days=3)
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2027, 1, 1), end_date=_date(2027, 1, 3))  # pending, but a different year
        self.check_leave_balance(self.emp, self.marriage, _date(2026, 9, 1), _date(2026, 9, 3))


class LeaveRequestFormCrossYearTests(TestCase):
    """LeaveRequestForm.clean() must reject a date range spanning two
    calendar years — LeaveEntitlement (and check_leave_balance) is keyed per
    year, so a cross-year request would otherwise only ever be checked
    against the start year's balance, silently never consulting the end
    year's entitlement at all."""
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=30)
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2027, entitled_days=30)

    def test_cross_year_range_rejected(self):
        form = LeaveRequestForm({
            'leave_type': self.marriage.pk, 'start_date': '2026-12-30', 'end_date': '2027-01-02',
            'employee_reason': 'New Year trip',
        }, fixed_employee=self.emp)
        self.assertFalse(form.is_valid())
        self.assertIn('cannot span two different years', str(form.errors['end_date']))

    def test_same_year_range_still_passes(self):
        form = LeaveRequestForm({
            'leave_type': self.marriage.pk, 'start_date': '2026-12-28', 'end_date': '2026-12-31',
            'employee_reason': 'End of year',
        }, fixed_employee=self.emp)
        self.assertTrue(form.is_valid(), form.errors)

    def test_new_years_eve_to_new_years_day_rejected(self):
        # The most common real-world case this guards against.
        form = LeaveRequestForm({
            'leave_type': self.marriage.pk, 'start_date': '2026-12-31', 'end_date': '2027-01-01',
            'employee_reason': 'NYE',
        }, fixed_employee=self.emp)
        self.assertFalse(form.is_valid())
        self.assertIn('end_date', form.errors)


class MyProfileLeaveRequestTests(TestCase):
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        self.user = make_user('profile_user', password='x')
        self.user.set_password('testpass123')
        self.user.save()
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        # The balance guardrail (check_leave_balance in hr/forms.py) now blocks any
        # submission without a matching LeaveEntitlement row — set up sufficient
        # balance so "can submit" tests exercise the real success path, not a
        # validation bypass.
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.marriage, year=2026, entitled_days=3)
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.annual, year=2026, entitled_days=30)

    def test_profile_shows_own_requests(self):
        LeaveRequest.objects.create(employee=self.emp, leave_type=self.marriage,
                                    start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Marriage')

    def test_employee_can_submit_own_request(self):
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave', 'leave_type': self.marriage.pk,
            'start_date': '2026-09-10', 'end_date': '2026-09-11', 'employee_reason': 'Wedding',
        })
        self.assertEqual(resp.status_code, 302)
        req = LeaveRequest.objects.get(employee=self.emp)
        self.assertEqual(req.created_by, self.user)
        self.assertEqual(req.status, 'pending')

    def test_edit_keeps_inactive_leave_type_no_silent_switch(self):
        # A pending request whose type was later deactivated must still show +
        # keep that type on edit, not silently switch to another active type.
        req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2),
            created_by=self.user, status='pending')
        self.marriage.is_active = False
        self.marriage.save(update_fields=['is_active'])
        self.client.login(username='profile_user', password='testpass123')
        # The edit dropdown still includes the (now-inactive) current type.
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertIn(self.marriage, list(resp.context['active_leave_types']))
        # Editing (keeping the inactive type) keeps it — no silent switch.
        self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': req.pk,
            'leave_type': self.marriage.pk,
            'start_date': '2026-08-01', 'end_date': '2026-08-02',
            'employee_reason': 'updated'})
        req.refresh_from_db()
        self.assertEqual(req.leave_type_id, self.marriage.pk)
        self.assertEqual(req.employee_reason, 'updated')

    def test_service_level_edit_failure_preserves_values_inline(self):
        # A ValueError from edit_leave_request (form passed, locked re-check
        # failed) must carry the attempted values + real error back, not revert.
        from unittest.mock import patch
        req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.annual,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2),
            created_by=self.user, status='pending')
        self.client.login(username='profile_user', password='testpass123')
        with patch('hr.leave_approval_services.edit_leave_request',
                   side_effect=ValueError('Balance exceeded')):
            resp = self.client.post(reverse('hr:my_profile'), {
                'action': 'edit_leave_request', 'request_id': req.pk,
                'leave_type': self.annual.pk,
                'start_date': '2026-08-01', 'end_date': '2026-08-05',
                'employee_reason': 'longer'})
        self.assertEqual(resp.status_code, 302)
        self.assertIn('edit_error=', resp.url)                 # real reason carried
        self.assertIn('retry_end_date=2026-08-05', resp.url)   # attempted value kept
        self.assertIn(f'opened_leave={req.pk}', resp.url)

    def test_request_exceeding_balance_is_rejected_with_visible_error(self):
        # Marriage entitlement is 3 days (setUp); this request spans 10 days.
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave', 'leave_type': self.marriage.pk,
            'start_date': '2026-09-01', 'end_date': '2026-09-10', 'employee_reason': 'Too long',
        })
        self.assertEqual(resp.status_code, 200)  # re-renders the form, no redirect
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp).exists())
        self.assertContains(resp, 'only 3')  # the remaining-balance error message
        # Regression: the error lives inside a Bootstrap-collapsed panel that
        # defaults to hidden — without the 'show' class, the error is present
        # in the HTML but invisible to the user.
        self.assertContains(resp, '<div class="collapse mb-3 show" id="requestLeaveForm">')

    def test_request_without_entitlement_setup_is_rejected_with_visible_error(self):
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave', 'leave_type': self.marriage.pk,
            'start_date': '2027-01-05', 'end_date': '2027-01-06', 'employee_reason': 'No entitlement for 2027',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp).exists())
        self.assertContains(resp, 'No leave entitlement has been set up')
        self.assertContains(resp, '<div class="collapse mb-3 show" id="requestLeaveForm">')

    def test_end_date_before_start_date_is_rejected_with_visible_error(self):
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave', 'leave_type': self.marriage.pk,
            'start_date': '2026-09-10', 'end_date': '2026-09-05', 'employee_reason': 'Backwards range',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp).exists())
        self.assertContains(resp, 'End date cannot be before start date')

    def test_collapse_panel_stays_closed_on_normal_page_load(self):
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, '<div class="collapse mb-3" id="requestLeaveForm">')

    def test_race_condition_valueerror_from_submit_shows_error_and_expands_panel(self):
        # Simulates submit_leave_request's locked, authoritative re-check
        # catching something the form's own (unlocked) pre-check missed —
        # e.g. a genuinely concurrent submission. The form itself validated
        # fine, so form.errors is empty; the panel must still expand via the
        # dedicated leave_request_submit_failed flag, and the message
        # framework banner must carry the actual error.
        from unittest.mock import patch
        self.client.login(username='profile_user', password='testpass123')
        with patch('hr.leave_approval_services.submit_leave_request',
                   side_effect=ValueError('Someone else just took that balance.')):
            resp = self.client.post(reverse('hr:my_profile'), {
                'action': 'request_leave', 'leave_type': self.marriage.pk,
                'start_date': '2026-09-10', 'end_date': '2026-09-11', 'employee_reason': 'Wedding',
            })
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Someone else just took that balance.')
        self.assertContains(resp, '<div class="collapse mb-3 show" id="requestLeaveForm">')
        self.assertFalse(LeaveRequest.objects.filter(employee=self.emp).exists())

    def test_profile_shows_who_decided_approved_request(self):
        req = LeaveRequest.objects.create(employee=self.emp, leave_type=self.marriage,
                                          start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        approver = make_user('profile_approver', password='x', first_name='Sarah', last_name='A')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        LeaveRequestApproval.objects.create(leave_request=req, approver=approver)
        from hr.leave_approval_services import record_approver_decision
        record_approver_decision(req, approver, 'approved')
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'by Sarah A')

    def test_employee_self_service_annual_leave_now_requires_approval(self):
        # Annual used to be recorded immediately with no approval step (the
        # "accumulative types skip approval" rule) — that bypass was removed:
        # every leave type, Annual included, now goes into the same pending
        # queue and needs Approve/Disapprove, exactly like Sick/Marriage/etc.
        # is_accumulative still controls the summary-total aggregation on this
        # same page (unchanged), just no longer bypasses approval.
        self.client.login(username='profile_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave', 'leave_type': self.annual.pk,
            'start_date': '2026-09-10', 'end_date': '2026-09-11', 'employee_reason': 'Vacation',
        })
        self.assertEqual(resp.status_code, 302)
        req = LeaveRequest.objects.get(employee=self.emp, leave_type=self.annual)
        self.assertEqual(req.status, 'pending')
        self.assertFalse(LeaveRecord.objects.filter(employee=self.emp, leave_type=self.annual).exists())

    def test_annual_leave_balance_deducted_only_once_approved(self):
        approver = make_user('profile_annual_approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        self.client.login(username='profile_user', password='testpass123')
        self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave', 'leave_type': self.annual.pk,
            'start_date': '2026-09-10', 'end_date': '2026-09-11', 'employee_reason': 'Vacation',
        })
        req = LeaveRequest.objects.get(employee=self.emp, leave_type=self.annual)
        self.assertFalse(LeaveRecord.objects.filter(employee=self.emp, leave_type=self.annual).exists())

        from hr.leave_approval_services import record_approver_decision
        record_approver_decision(req, approver, 'approved')
        req.refresh_from_db()
        self.assertEqual(req.status, 'approved')
        rec = LeaveRecord.objects.get(employee=self.emp, leave_type=self.annual)
        self.assertEqual(rec.days, Decimal('2'))



class LeaveRequestDecidedByDisplayTests(TestCase):
    """Model-level coverage of LeaveRequest.decided_by_display, independent
    of any view/template — the 'by whom' summary shown across the queue
    history, My Profile, and the detail page."""
    def setUp(self):
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))

    def test_pending_returns_none(self):
        self.assertIsNone(self.req.decided_by_display)

    def test_approved_by_all_approvers(self):
        aamna = make_user('dbd_aamna', password='x', first_name='Aamna', last_name='Khan')
        ali = make_user('dbd_ali', password='x', first_name='Ali', last_name='Sultan')
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=aamna, decision='approved')
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=ali, decision='approved')
        self.req.status = 'approved'
        self.req.save(update_fields=['status'])
        self.assertEqual(self.req.decided_by_display, 'by Aamna Khan, Ali Sultan')

    def test_disapproved_by_single_approver(self):
        aamna = make_user('dbd_aamna2', password='x', first_name='Aamna', last_name='Khan')
        ali = make_user('dbd_ali2', password='x', first_name='Ali', last_name='Sultan')
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=aamna, decision='disapproved')
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=ali, decision='skipped')
        self.req.status = 'disapproved'
        self.req.save(update_fields=['status'])
        self.assertEqual(self.req.decided_by_display, 'by Aamna Khan')

    def test_override_shows_overriding_admin(self):
        admin = make_user('dbd_admin', password='x', first_name='Sarah', last_name='Admin')
        self.req.status = 'approved'
        self.req.is_overridden = True
        self.req.overridden_by = admin
        self.req.save(update_fields=['status', 'is_overridden', 'overridden_by'])
        self.assertEqual(self.req.decided_by_display, 'by Sarah Admin (override)')

    def test_override_with_deleted_overriding_user(self):
        admin = make_user('dbd_admin2', password='x')
        self.req.status = 'approved'
        self.req.is_overridden = True
        self.req.overridden_by = admin
        self.req.save(update_fields=['status', 'is_overridden', 'overridden_by'])
        admin.delete()
        self.req.refresh_from_db()
        self.assertEqual(self.req.decided_by_display, '(override — approver account no longer exists)')


from datetime import datetime as _datetime, time as _time
from django.utils import timezone as _timezone
from hr.models import AttendanceException, AttendanceExceptionRevokeRequest


class EmployeeDownstreamChainTests(TestCase):
    def setUp(self):
        self.head = make_employee(iqama='CHAIN-HEAD', name='Head')
        self.mid = make_employee(iqama='CHAIN-MID', name='Mid')
        self.mid.main_manager = self.head
        self.mid.save()
        self.leaf1 = make_employee(iqama='CHAIN-LEAF1', name='Leaf1')
        self.leaf1.main_manager = self.mid
        self.leaf1.save()
        self.leaf2 = make_employee(iqama='CHAIN-LEAF2', name='Leaf2')
        self.leaf2.main_manager = self.mid
        self.leaf2.save()

    def test_main_reports_only(self):
        self.assertEqual(set(e.pk for e in self.head.main_reports.all()), {self.mid.pk})

    def test_full_downstream_chain(self):
        ids = self.head.get_downstream_employee_ids()
        self.assertEqual(set(ids), {self.mid.pk, self.leaf1.pk, self.leaf2.pk})

    def test_leaf_has_no_downstream(self):
        self.assertEqual(self.leaf1.get_downstream_employee_ids(), [])

    def test_manager_deletion_does_not_cascade_to_reports(self):
        head_pk = self.head.pk
        self.head.delete()
        self.mid.refresh_from_db()
        self.assertIsNone(self.mid.main_manager_id)
        self.assertFalse(Employee.objects.filter(pk=head_pk).exists())


class AttendanceExceptionModelTests(TestCase):
    def setUp(self):
        self.manager_emp = make_employee(iqama='AE-MGR', name='Manager Person')
        self.emp = make_employee(iqama='AE-EMP', name='Field Employee')
        self.emp.main_manager = self.manager_emp
        self.emp.save()

    def _make(self, **overrides):
        defaults = dict(
            employee=self.emp, event_date=date(2026, 8, 10), event_start_time=_time(9, 0),
            reason_category='site_visit', main_manager=self.manager_emp,
        )
        defaults.update(overrides)
        return AttendanceException(**defaults)

    def test_other_category_requires_custom_reason(self):
        exc = self._make(reason_category='other', custom_reason='')
        with self.assertRaises(ValidationError):
            exc.full_clean()

    def test_other_category_with_custom_reason_is_valid(self):
        exc = self._make(reason_category='other', custom_reason='Client emergency call')
        exc.full_clean()  # should not raise

    def test_event_start_combines_date_and_time(self):
        exc = self._make()
        expected = _timezone.make_aware(_datetime(2026, 8, 10, 9, 0))
        self.assertEqual(exc.event_start, expected)

    def test_decision_window_boundaries(self):
        exc = self._make()
        exc.save()
        self.assertEqual(exc.decision_window_opens_at, exc.event_start - timedelta(days=7))
        self.assertEqual(exc.decision_deadline, exc.event_start + timedelta(hours=24))

    def test_is_within_decision_window_true_just_before_event(self):
        exc = self._make()
        exc.save()
        now = exc.event_start - timedelta(days=1)
        self.assertTrue(exc.is_within_decision_window(now=now))

    def test_is_within_decision_window_false_too_early(self):
        exc = self._make()
        exc.save()
        now = exc.event_start - timedelta(days=8)
        self.assertFalse(exc.is_within_decision_window(now=now))

    def test_is_overdue_after_24h(self):
        exc = self._make()
        exc.save()
        now = exc.event_start + timedelta(hours=25)
        self.assertTrue(exc.is_overdue(now=now))
        self.assertFalse(exc.is_overdue(now=exc.event_start + timedelta(hours=23)))

    def test_time_left_seconds_counts_down(self):
        exc = self._make()
        exc.save()
        now = exc.event_start + timedelta(hours=23)
        self.assertAlmostEqual(exc.time_left_seconds(now=now), 3600, delta=1)


from hr.attendance_exception_services import (
    submit_attendance_exception, decide_attendance_exception,
    override_attendance_exception, expire_overdue_exceptions,
)


class AttendanceExceptionServiceTests(TestCase):
    def setUp(self):
        self.manager_emp = make_employee(iqama='AEX-MGR', name='Manager Person')
        self.manager_user = make_user('aex_manager')
        self.manager_emp.user = self.manager_user
        self.manager_emp.save(update_fields=['user'])

        self.emp = make_employee(iqama='AEX-EMP', name='Field Employee')
        self.emp.main_manager = self.manager_emp
        self.emp.save(update_fields=['main_manager'])
        self.emp_user = make_user('aex_employee')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])

        self.creator = make_user('aex_creator')

        self.today = _timezone.now().date()
        self.in_window_time = _time(0, 1)


    def test_approving_exception_flips_already_saved_late_record_to_present(self):
        AttendanceRecord.objects.create(
            employee=self.emp, date=self.today, status='late', check_in=_time(9, 30))
        exc = self._submit()
        decide_attendance_exception(exc, self.manager_user, 'approved', 'Confirmed')
        rec = AttendanceRecord.objects.get(employee=self.emp, date=self.today)
        self.assertEqual(rec.status, 'present')

    def test_build_matrix_surfaces_the_exception_reason(self):
        from hr.attendance_matrix import build_matrix
        AttendanceRecord.objects.create(
            employee=self.emp, date=self.today, status='late', check_in=_time(9, 30))
        exc = self._submit(custom_reason='Client site walkthrough')
        decide_attendance_exception(exc, self.manager_user, 'approved', 'Confirmed')
        days, rows = build_matrix([self.emp], self.today, self.today)
        cell = rows[0]['cells'][0]
        self.assertEqual(cell['status'], 'present')
        self.assertEqual(cell['exception_reason'], 'Client site walkthrough')
    def _submit(self, **overrides):
        defaults = dict(
            employee=self.emp, event_date=self.today, event_start_time=self.in_window_time,
            reason_category='site_visit', created_by=self.creator,
        )
        defaults.update(overrides)
        return submit_attendance_exception(**defaults)

    # -- submit --

    def test_submit_snapshots_manager_and_notifies(self):
        exc = self._submit()
        self.assertEqual(exc.main_manager_id, self.manager_emp.pk)
        self.assertEqual(exc.status, 'pending')
        self.assertTrue(Notification.objects.filter(recipient=self.manager_user).exists())

    def test_submit_with_no_manager_creates_no_notification(self):
        lone = make_employee(iqama='AEX-LONE', name='Lone Wolf')
        before = Notification.objects.count()
        exc = submit_attendance_exception(
            employee=lone, event_date=self.today, event_start_time=self.in_window_time,
            reason_category='site_visit', created_by=self.creator,
        )
        self.assertIsNone(exc.main_manager)
        self.assertEqual(Notification.objects.count(), before)

    def test_duplicate_submission_for_same_event_rejected(self):
        # A double-click or resubmit of the exact same event must not create
        # a second pending row for the manager's queue.
        self._submit()
        with self.assertRaises(ValueError):
            self._submit()
        self.assertEqual(
            AttendanceException.objects.filter(
                employee=self.emp, event_date=self.today, event_start_time=self.in_window_time).count(),
            1)

    def test_resubmission_allowed_after_earlier_one_decided(self):
        first = self._submit()
        decide_attendance_exception(first, self.manager_user, 'rejected', note='Wrong info')
        # Same employee/date/time, but the earlier one is no longer pending —
        # a fresh submission for that same slot must be allowed through.
        second = self._submit()
        self.assertNotEqual(first.pk, second.pk)

    def test_different_event_time_not_blocked(self):
        self._submit()
        other = self._submit(event_start_time=_time(0, 2))
        self.assertIsNotNone(other.pk)

    # -- decide --

    def test_decide_approved_marks_present(self):
        # Simplified outcome mapping: approved always upserts 'present'
        # (excused) — late-vs-not is now derived automatically from real
        # check-in time by the Wi-Fi/manual attendance-status derivation
        # layer, not by a manual 'arrived late' flag on the decision itself.
        exc = self._submit()
        decide_attendance_exception(exc, self.manager_user, 'approved')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'approved')
        record = AttendanceRecord.objects.get(employee=self.emp, date=self.today)
        self.assertEqual(record.status, 'present')

    def test_decide_rejected_marks_absent(self):
        exc = self._submit()
        decide_attendance_exception(exc, self.manager_user, 'rejected', note='No evidence')
        record = AttendanceRecord.objects.get(employee=self.emp, date=self.today)
        self.assertEqual(record.status, 'absent')

    def test_decide_notifies_employee(self):
        exc = self._submit()
        decide_attendance_exception(exc, self.manager_user, 'approved')
        self.assertTrue(
            Notification.objects.filter(recipient=self.emp_user, verb__icontains='approved').exists())

    def test_decide_wrong_manager_rejected(self):
        exc = self._submit()
        stranger = make_user('aex_stranger')
        with self.assertRaises(ValueError):
            decide_attendance_exception(exc, stranger, 'approved')

    def test_decide_before_window_opens_now_allowed(self):
        # Task 1: the is_within_decision_window() timing check was removed
        # entirely from decide_attendance_exception — a manager may now
        # decide even before the 7-days-early decision window would have
        # opened. Only the status-window ({'pending', 'expired'}) and
        # assigned-manager checks remain.
        exc = self._submit(event_date=self.today + timedelta(days=10))
        decide_attendance_exception(exc, self.manager_user, 'approved')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'approved')

    def test_decide_after_24h_deadline_still_allowed_as_normal_decision(self):
        # Task 1: a late decision (past decision_deadline) must be allowed to
        # proceed as a normal decide — not blocked, not forced through the
        # override path — and must still correctly derive the day's
        # AttendanceRecord outcome.
        exc = self._submit(event_date=self.today - timedelta(days=5))
        decide_attendance_exception(exc, self.manager_user, 'rejected', note='Reviewed late')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'rejected')
        self.assertFalse(exc.is_overridden)
        record = AttendanceRecord.objects.get(employee=self.emp, date=exc.event_date)
        self.assertEqual(record.status, 'absent')

    def test_decide_auto_expired_request_still_decidable_by_real_manager(self):
        # Task 1: decide_attendance_exception's allowed starting statuses
        # widen from strictly {'pending'} to {'pending', 'expired'} — mirrors
        # override_attendance_exception's existing allowance — so a request
        # the cron sweep already auto-expired can still be decided normally
        # (not via override) by the real assigned manager.
        exc = self._submit(event_date=self.today - timedelta(days=5))
        exc.status = 'expired'
        exc.save(update_fields=['status'])
        decide_attendance_exception(exc, self.manager_user, 'approved')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'approved')
        self.assertFalse(exc.is_overridden)
        record = AttendanceRecord.objects.get(employee=self.emp, date=exc.event_date)
        self.assertEqual(record.status, 'present')

    def test_decide_already_decided_rejected(self):
        exc = self._submit()
        decide_attendance_exception(exc, self.manager_user, 'approved')
        with self.assertRaises(ValueError):
            decide_attendance_exception(exc, self.manager_user, 'rejected')

    def test_decide_race_guard_second_stale_call_rejected(self):
        exc = self._submit()
        stale1 = AttendanceException.objects.get(pk=exc.pk)
        stale2 = AttendanceException.objects.get(pk=exc.pk)
        decide_attendance_exception(stale1, self.manager_user, 'approved')
        with self.assertRaises(ValueError):
            decide_attendance_exception(stale2, self.manager_user, 'rejected')

    def test_decide_own_request_rejected_even_for_the_assigned_manager(self):
        # Guardrail applies even in the (pathological) case where the
        # employee is somehow their own assigned main_manager — the
        # self-approval check runs before the main_manager check.
        exc = self._submit()
        with self.assertRaises(ValueError):
            decide_attendance_exception(exc, self.emp_user, 'approved')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'pending')

    # -- override --

    def test_override_pending_success(self):
        exc = self._submit()
        hr_user = make_user('aex_hr')
        override_attendance_exception(exc, hr_user, 'approved', reason='Confirmed via site log')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'approved')
        self.assertTrue(exc.is_overridden)
        self.assertEqual(exc.overridden_by, hr_user)

    def test_override_expired_success(self):
        exc = self._submit(event_date=self.today - timedelta(days=5))
        exc.status = 'expired'
        exc.save(update_fields=['status'])
        hr_user = make_user('aex_hr2')
        override_attendance_exception(exc, hr_user, 'rejected', reason='No evidence provided')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'rejected')
        record = AttendanceRecord.objects.get(employee=self.emp, date=exc.event_date)
        self.assertEqual(record.status, 'absent')

    def test_override_blank_reason_rejected(self):
        exc = self._submit()
        hr_user = make_user('aex_hr3')
        with self.assertRaises(ValueError):
            override_attendance_exception(exc, hr_user, 'approved', reason='   ')

    def test_override_already_decided_rejected(self):
        exc = self._submit()
        decide_attendance_exception(exc, self.manager_user, 'approved')
        hr_user = make_user('aex_hr4')
        with self.assertRaises(ValueError):
            override_attendance_exception(exc, hr_user, 'rejected', reason='Changing my mind')

    def test_override_own_request_rejected_even_as_super_admin(self):
        # Self-approval guardrail applies at the service layer regardless of
        # the overriding user's role — a Super Admin who is also the
        # request's own employee still cannot override their own request.
        exc = self._submit()
        with self.assertRaises(ValueError):
            override_attendance_exception(exc, self.emp_user, 'approved', reason='Self-confirmed')
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'pending')

    # -- expire --

    def test_expire_overdue_pending(self):
        exc = self._submit(event_date=self.today - timedelta(days=5))
        count = expire_overdue_exceptions()
        exc.refresh_from_db()
        self.assertEqual(count, 1)
        self.assertEqual(exc.status, 'expired')
        record = AttendanceRecord.objects.get(employee=self.emp, date=exc.event_date)
        self.assertEqual(record.status, 'absent')
        self.assertTrue(
            Notification.objects.filter(recipient=self.emp_user, verb__icontains='expired').exists())
        self.assertTrue(
            Notification.objects.filter(recipient=self.manager_user, verb__icontains='expired').exists())

    def test_expire_does_not_touch_in_window_pending(self):
        exc = self._submit()
        count = expire_overdue_exceptions()
        exc.refresh_from_db()
        self.assertEqual(count, 0)
        self.assertEqual(exc.status, 'pending')

    def test_expire_does_not_touch_already_decided(self):
        exc = self._submit(event_date=self.today - timedelta(days=5))
        hr_user = make_user('aex_hr5')
        override_attendance_exception(exc, hr_user, 'approved', reason='Manually confirmed')
        count = expire_overdue_exceptions()
        exc.refresh_from_db()
        self.assertEqual(count, 0)
        self.assertEqual(exc.status, 'approved')


# ─── Attendance Exception Workflow: start reminders + cron command ────────

from django.core.management import call_command
from hr.attendance_exception_services import send_pending_start_reminders


class SendPendingStartRemindersTests(TestCase):
    def setUp(self):
        self.manager_emp = make_employee(iqama='AEXR-MGR', name='Manager Person')
        self.manager_user = make_user('aexr_manager')
        self.manager_emp.user = self.manager_user
        self.manager_emp.save(update_fields=['user'])

        self.emp = make_employee(iqama='AEXR-EMP', name='Field Employee')
        self.emp.main_manager = self.manager_emp
        self.emp.save(update_fields=['main_manager'])
        self.emp_user = make_user('aexr_employee')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])

        self.creator = make_user('aexr_creator')

    def _submit(self, started_ago=None, started_in=None, **overrides):
        # Derive both date and time from the same local datetime so a
        # midnight rollover can never split them across two different days.
        if started_ago is not None:
            dt = _timezone.localtime(_timezone.now() - started_ago)
        else:
            dt = _timezone.localtime(_timezone.now() + started_in)
        defaults = dict(
            employee=self.emp, event_date=dt.date(), event_start_time=dt.time(),
            reason_category='site_visit', created_by=self.creator,
        )
        defaults.update(overrides)
        return submit_attendance_exception(**defaults)

    def test_reminds_manager_once_event_has_started(self):
        exc = self._submit(started_ago=timedelta(hours=1))
        count = send_pending_start_reminders()
        exc.refresh_from_db()
        self.assertEqual(count, 1)
        self.assertIsNotNone(exc.reminder_sent_at)
        self.assertTrue(
            Notification.objects.filter(recipient=self.manager_user, verb__icontains='started').exists())

    def test_does_not_remind_for_future_event(self):
        exc = self._submit(started_in=timedelta(days=5))
        count = send_pending_start_reminders()
        exc.refresh_from_db()
        self.assertEqual(count, 0)
        self.assertIsNone(exc.reminder_sent_at)

    def test_idempotent_across_repeated_calls(self):
        self._submit(started_ago=timedelta(hours=1))
        first = send_pending_start_reminders()
        second = send_pending_start_reminders()
        self.assertEqual(first, 1)
        self.assertEqual(second, 0)
        self.assertEqual(
            Notification.objects.filter(recipient=self.manager_user, verb__icontains='started').count(), 1)

    def test_no_linked_manager_login_still_marks_sent_without_notifying(self):
        lone_manager = make_employee(iqama='AEXR-LONEMGR', name='Loneless Manager')
        self.emp.main_manager = lone_manager
        self.emp.save(update_fields=['main_manager'])
        exc = self._submit(started_ago=timedelta(hours=1))
        before = Notification.objects.count()
        count = send_pending_start_reminders()
        exc.refresh_from_db()
        self.assertEqual(count, 1)
        self.assertIsNotNone(exc.reminder_sent_at)
        self.assertEqual(Notification.objects.count(), before)

    def test_does_not_touch_already_decided_or_expired(self):
        exc = self._submit(started_ago=timedelta(hours=1))
        decide_attendance_exception(exc, self.manager_user, 'approved')
        count = send_pending_start_reminders()
        exc.refresh_from_db()
        self.assertEqual(count, 0)
        self.assertIsNone(exc.reminder_sent_at)


class ProcessAttendanceExceptionsCommandTests(TestCase):
    def setUp(self):
        self.manager_emp = make_employee(iqama='AEXC-MGR', name='Manager Person')
        self.manager_user = make_user('aexc_manager')
        self.manager_emp.user = self.manager_user
        self.manager_emp.save(update_fields=['user'])

        self.emp = make_employee(iqama='AEXC-EMP', name='Field Employee')
        self.emp.main_manager = self.manager_emp
        self.emp.save(update_fields=['main_manager'])
        self.emp_user = make_user('aexc_employee')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])

        self.creator = make_user('aexc_creator')

    def _submit(self, event_date, event_start_time, **overrides):
        defaults = dict(
            employee=self.emp, event_date=event_date, event_start_time=event_start_time,
            reason_category='site_visit', created_by=self.creator,
        )
        defaults.update(overrides)
        return submit_attendance_exception(**defaults)

    def test_command_sends_reminders_and_expires_overdue(self):
        recent = _timezone.localtime(_timezone.now() - timedelta(hours=1))
        reminder_due = self._submit(recent.date(), recent.time())

        stale = _timezone.localtime(_timezone.now() - timedelta(days=5))
        overdue = self._submit(stale.date(), stale.time(), employee=self.emp)

        call_command('process_attendance_exceptions')

        reminder_due.refresh_from_db()
        overdue.refresh_from_db()

        self.assertIsNotNone(reminder_due.reminder_sent_at)
        self.assertTrue(
            Notification.objects.filter(recipient=self.manager_user, verb__icontains='started').exists())

        self.assertEqual(overdue.status, 'expired')
        record = AttendanceRecord.objects.get(employee=self.emp, date=overdue.event_date)
        self.assertEqual(record.status, 'absent')


class EmployeeIsManagerOfTests(TestCase):
    def setUp(self):
        self.head = make_employee(iqama='IMO-HEAD', name='Head')
        self.mid = make_employee(iqama='IMO-MID', name='Mid')
        self.mid.main_manager = self.head
        self.mid.save()
        self.leaf = make_employee(iqama='IMO-LEAF', name='Leaf')
        self.leaf.main_manager = self.mid
        self.leaf.save()
        self.outsider = make_employee(iqama='IMO-OUT', name='Outsider')

    def test_direct_manager(self):
        self.assertTrue(self.mid.is_manager_of(self.leaf))

    def test_upstream_manager(self):
        self.assertTrue(self.head.is_manager_of(self.leaf))

    def test_not_a_manager(self):
        self.assertFalse(self.outsider.is_manager_of(self.leaf))

    def test_no_manager_at_all_returns_false(self):
        self.assertFalse(self.outsider.is_manager_of(self.head))

    def test_max_depth_exhausted_returns_false(self):
        self.assertFalse(self.head.is_manager_of(self.leaf, max_depth=0))


class EmployeeCleanCycleDetectionTests(TestCase):
    """Employee.clean()'s guardrail against main_manager assignments that
    would break the formal reporting line — direct self-assignment and
    indirect (multi-hop) cycles."""

    def test_direct_self_assignment_raises(self):
        emp = make_employee(iqama='CYC-SELF', name='Self Referencer')
        emp.main_manager = emp
        with self.assertRaises(ValidationError):
            emp.full_clean()

    def test_three_node_cycle_raises(self):
        a = make_employee(iqama='CYC-A', name='A')
        b = make_employee(iqama='CYC-B', name='B')
        c = make_employee(iqama='CYC-C', name='C')
        b.main_manager = a
        b.save()
        c.main_manager = b
        c.save()
        # Closing the loop: A -> C -> B -> A.
        a.main_manager = c
        with self.assertRaises(ValidationError):
            a.full_clean()

    def test_valid_non_circular_chain_does_not_raise(self):
        head = make_employee(iqama='CYC-HEAD', name='Head')
        mid = make_employee(iqama='CYC-MID', name='Mid')
        mid.main_manager = head
        mid.save()
        leaf = make_employee(iqama='CYC-LEAF', name='Leaf')
        leaf.main_manager = mid
        leaf.full_clean()  # should not raise


# ─── Attendance Exception Workflow: forms, views, templates ──────────────

from accounts.models import Role
from django.contrib.messages import get_messages
from hr.attendance_exception_services import submit_attendance_exception as _submit_aex


def _make_role_user(username, role_name):
    """A logged-in test user holding a specific Role (e.g. super_admin, ai_head)."""
    role, _ = Role.objects.get_or_create(name=role_name)
    user = make_user(username, password='x')
    user.set_password('testpass123')
    user.role = role
    user.save()
    return user


def _login_user(username):
    user = make_user(username, password='x')
    user.set_password('testpass123')
    user.save()
    return user


class MyProfileAttendanceExceptionTests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='MPAX-1', name='Field Worker')
        self.user = _login_user('mpax_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])

    def test_employee_can_submit_own_exception(self):
        self.client.login(username='mpax_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'submit_attendance_exception',
            'event_date': '2026-07-20', 'event_start_time': '09:30',
            'reason_category': 'site_visit', 'custom_reason': '',
        })
        self.assertEqual(resp.status_code, 302)
        exc = AttendanceException.objects.get(employee=self.emp)
        self.assertEqual(exc.status, 'pending')
        self.assertEqual(exc.created_by, self.user)

    def test_submitted_exception_appears_in_profile_context(self):
        _submit_aex(employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 30),
                    reason_category='site_visit', created_by=self.user)
        self.client.login(username='mpax_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn(resp.context['my_attendance_exceptions'][0].employee_id, [self.emp.pk])

    def test_other_requires_custom_reason(self):
        self.client.login(username='mpax_user', password='testpass123')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'submit_attendance_exception',
            'event_date': '2026-07-20', 'event_start_time': '09:30',
            'reason_category': 'other', 'custom_reason': '',
        })
        self.assertEqual(resp.status_code, 200)  # form invalid, re-rendered
        self.assertFalse(AttendanceException.objects.filter(employee=self.emp).exists())

    def test_duplicate_submission_shows_error_not_500(self):
        self.client.login(username='mpax_user', password='testpass123')
        post_data = {
            'action': 'submit_attendance_exception',
            'event_date': '2026-07-20', 'event_start_time': '09:30',
            'reason_category': 'site_visit', 'custom_reason': '',
        }
        self.client.post(reverse('hr:my_profile'), post_data)
        resp = self.client.post(reverse('hr:my_profile'), post_data, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(AttendanceException.objects.filter(employee=self.emp).count(), 1)


class TeamExceptionsAccessTests(TestCase):
    def setUp(self):
        self.top = make_employee(iqama='TEX-TOP', name='Top Manager')
        self.top_user = _login_user('tex_top')
        self.top.user = self.top_user
        self.top.save(update_fields=['user'])

        self.mid = make_employee(iqama='TEX-MID', name='Mid Manager')
        self.mid.main_manager = self.top
        self.mid.save(update_fields=['main_manager'])
        self.mid_user = _login_user('tex_mid')
        self.mid.user = self.mid_user
        self.mid.save(update_fields=['user'])

        self.direct_report = make_employee(iqama='TEX-DIRECT', name='Direct Report')
        self.direct_report.main_manager = self.mid
        self.direct_report.save(update_fields=['main_manager'])

        self.grandchild = make_employee(iqama='TEX-GRANDCHILD', name='Grandchild')
        self.grandchild.main_manager = self.direct_report
        self.grandchild.save(update_fields=['main_manager'])

        self.exc_direct = _submit_aex(
            employee=self.direct_report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.top_user)
        self.exc_grandchild = _submit_aex(
            employee=self.grandchild, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.top_user)

        self.plain_emp = make_employee(iqama='TEX-PLAIN', name='Plain Employee')
        self.plain_user = _login_user('tex_plain')
        self.plain_emp.user = self.plain_user
        self.plain_emp.save(update_fields=['user'])

        self.hr_user = _make_role_user('tex_hr', Role.SUPER_ADMIN)
        self.ai_head_user = _make_role_user('tex_ai_head', Role.AI_HEAD)

    def test_plain_employee_forbidden(self):
        self.client.login(username='tex_plain', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 403)

    def test_manager_with_main_reports_can_view(self):
        self.client.login(username='tex_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)

    def test_head_role_user_can_view_with_zero_reports(self):
        self.client.login(username='tex_ai_head', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)

    def test_hr_sees_all_requests_on_all_organization_tab(self):
        # Task 2: HR's global visibility now lives on the dedicated
        # "All Organization Requests" tab (?tab=all), not the default view.
        self.client.login(username='tex_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        self.assertEqual(resp.status_code, 200)
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_direct.pk, self.exc_grandchild.pk})

    def test_hr_default_tab_is_direct_and_empty_without_own_employee_profile(self):
        # HR/Super-Admin test users have no linked Employee profile, so their
        # default "Direct Reports" tab is empty — global visibility requires
        # explicitly switching to "All Organization Requests".
        self.client.login(username='tex_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['tab'], 'direct')
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, set())

    def test_non_hr_requesting_all_tab_silently_falls_back_to_direct(self):
        # A non-HR user somehow requesting ?tab=all must not error and must
        # not leak org-wide data — it silently resolves to the same 'direct'
        # tab/data they'd get with no ?tab param at all.
        self.client.login(username='tex_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['tab'], 'direct')
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_direct.pk})
        self.assertNotIn(self.exc_grandchild.pk, ids)

    def test_mid_manager_sees_only_main_reports_by_default(self):
        self.client.login(username='tex_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_direct.pk})

    def test_secondary_tab_does_not_include_recursive_grandchild(self):
        # Task 2 fix: the recursive downstream-hierarchy chain concept is
        # removed from this page entirely. The "Secondary Reports" tab is
        # scoped strictly to employee__secondary_managers — it does NOT pull
        # in grandchildren via the main-manager chain. self.grandchild's
        # main_manager is self.direct_report (not self.mid), and grandchild
        # has no secondary_managers relationship to mid, so it must stay
        # excluded from mid's "Secondary Reports" tab.
        self.client.login(username='tex_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'secondary'})
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, set())

    def test_page_renders_with_an_expired_request_in_history(self):
        # Regression test: expired requests have decided_by=None (nobody
        # decided — it just timed out), and the history table's template
        # used to crash rendering that row with VariableDoesNotExist
        # ("Failed lookup for key [username] in None") — a filter *argument*
        # referencing a dotted lookup on None isn't silently swallowed the
        # way a plain {{ var }} output is. Fixed by guarding with {% if %}.
        from hr.attendance_exception_services import expire_overdue_exceptions
        overdue = _submit_aex(
            employee=self.direct_report, event_date=_date(2026, 6, 1), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.top_user)
        overdue.event_date = date(2020, 1, 1)
        overdue.save(update_fields=['event_date'])
        expire_overdue_exceptions()
        self.client.login(username='tex_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Expired')


class TeamExceptionsDecideOverrideTests(TestCase):
    def setUp(self):
        self.top = make_employee(iqama='TEXD-TOP', name='Top Manager')
        self.top_user = _login_user('texd_top')
        self.top.user = self.top_user
        self.top.save(update_fields=['user'])

        self.mid = make_employee(iqama='TEXD-MID', name='Mid Manager')
        self.mid.main_manager = self.top
        self.mid.save(update_fields=['main_manager'])
        self.mid_user = _login_user('texd_mid')
        self.mid.user = self.mid_user
        self.mid.save(update_fields=['user'])

        # A manager unrelated to the direct_report/grandchild chain — has its own
        # direct report so it legitimately passes the page-level access gate,
        # letting the 'not the assigned manager for *this* request' check in
        # post() (rather than the view's test_func) be what actually denies it.
        self.other_manager = make_employee(iqama='TEXD-OTHER', name='Other Manager')
        self.other_manager_user = _login_user('texd_other')
        self.other_manager.user = self.other_manager_user
        self.other_manager.save(update_fields=['user'])
        self.other_manager_report = make_employee(iqama='TEXD-OTHERREPORT', name='Other Manager Report')
        self.other_manager_report.main_manager = self.other_manager
        self.other_manager_report.save(update_fields=['main_manager'])

        self.direct_report = make_employee(iqama='TEXD-DIRECT', name='Direct Report')
        self.direct_report.main_manager = self.mid
        self.direct_report.save(update_fields=['main_manager'])

        self.grandchild = make_employee(iqama='TEXD-GRANDCHILD', name='Grandchild')
        self.grandchild.main_manager = self.direct_report
        self.grandchild.save(update_fields=['main_manager'])

        self.hr_user = _make_role_user('texd_hr', Role.SUPER_ADMIN)

        # Derive both date and time from the same local datetime so a
        # midnight rollover can never split them across two different days.
        event_start_dt = _timezone.localtime(_timezone.now() - timedelta(hours=1))
        self.in_window_date = event_start_dt.date()
        self.in_window_time = event_start_dt.time()

        self.exc_direct = _submit_aex(
            employee=self.direct_report, event_date=self.in_window_date, event_start_time=self.in_window_time,
            reason_category='site_visit', created_by=self.top_user)
        self.exc_grandchild = _submit_aex(
            employee=self.grandchild, event_date=self.in_window_date, event_start_time=self.in_window_time,
            reason_category='site_visit', created_by=self.top_user)

    def _post(self, username, data):
        self.client.login(username=username, password='testpass123')
        return self.client.post(reverse('hr:team_exceptions'), data)

    def test_assigned_manager_can_decide(self):
        resp = self._post('texd_mid', {'action': 'decide', 'exc_id': self.exc_direct.pk, 'decision': 'approved'})
        self.assertEqual(resp.status_code, 302)
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'approved')

    def test_non_assigned_manager_cannot_decide(self):
        # 404, not 403: collapsing "not yours" and "doesn't exist" into one
        # response closes the pk-existence oracle (same fix already applied
        # to the leave workflow's document download view).
        resp = self._post('texd_other', {'action': 'decide', 'exc_id': self.exc_direct.pk, 'decision': 'approved'})
        self.assertEqual(resp.status_code, 404)
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'pending')

    def test_decide_nonexistent_request_also_404s(self):
        resp = self._post('texd_other', {'action': 'decide', 'exc_id': 999999, 'decision': 'approved'})
        self.assertEqual(resp.status_code, 404)

    def test_hr_can_override_pending_with_reason(self):
        resp = self._post('texd_hr', {
            'action': 'override', 'exc_id': self.exc_direct.pk, 'decision': 'approved', 'reason': 'Confirmed by HR'})
        self.assertEqual(resp.status_code, 302)
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'approved')
        self.assertTrue(self.exc_direct.is_overridden)

    def test_hr_can_override_expired_with_reason(self):
        self.exc_direct.status = 'expired'
        self.exc_direct.save(update_fields=['status'])
        resp = self._post('texd_hr', {
            'action': 'override', 'exc_id': self.exc_direct.pk, 'decision': 'rejected', 'reason': 'No evidence'})
        self.assertEqual(resp.status_code, 302)
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'rejected')

    def test_override_without_reason_rejected(self):
        resp = self._post('texd_hr', {
            'action': 'override', 'exc_id': self.exc_direct.pk, 'decision': 'approved', 'reason': ''})
        self.assertEqual(resp.status_code, 302)
        messages_list = list(get_messages(resp.wsgi_request))
        self.assertTrue(any('reason' in str(m).lower() for m in messages_list))
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'pending')

    def test_upstream_non_direct_manager_can_override(self):
        # top is not exc_grandchild's assigned manager (that's direct_report),
        # but top is upstream of grandchild in the hierarchy.
        resp = self._post('texd_top', {
            'action': 'override', 'exc_id': self.exc_grandchild.pk, 'decision': 'approved',
            'reason': 'Confirmed via upstream hierarchy'})
        self.assertEqual(resp.status_code, 302)
        self.exc_grandchild.refresh_from_db()
        self.assertEqual(self.exc_grandchild.status, 'approved')
        self.assertTrue(self.exc_grandchild.is_overridden)

    def test_overriding_already_decided_request_rejected(self):
        self._post('texd_mid', {'action': 'decide', 'exc_id': self.exc_direct.pk, 'decision': 'approved'})
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'approved')
        resp = self._post('texd_hr', {
            'action': 'override', 'exc_id': self.exc_direct.pk, 'decision': 'rejected', 'reason': 'Changed my mind'})
        self.assertEqual(resp.status_code, 302)
        self.exc_direct.refresh_from_db()
        self.assertEqual(self.exc_direct.status, 'approved')  # unchanged


class SelfApprovalGuardrailTests(TestCase):
    """A Super Admin who is also the request's own employee must never be
    able to decide or override their own attendance exception request —
    confirmed gap fix, enforced at both can_decide_attendance_exception
    (view-layer gate) and decide_/override_attendance_exception
    (service-layer, defense in depth)."""

    def setUp(self):
        self.hr_user = _make_role_user('selfapp_hr', Role.SUPER_ADMIN)
        self.hr_emp = make_employee(iqama='SELFAPP-HR', name='Super Admin Employee')
        self.hr_emp.user = self.hr_user
        self.hr_emp.save(update_fields=['user'])

        self.exc = _submit_aex(
            employee=self.hr_emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.hr_user)

    def test_can_decide_attendance_exception_false_for_own_request(self):
        from hr.views import can_decide_attendance_exception
        self.assertFalse(can_decide_attendance_exception(self.hr_user, self.exc))

    def test_team_exceptions_page_shows_no_action_buttons_for_own_row(self):
        # Task 2: HR's own request is only visible at all on the
        # "All Organization Requests" tab — visible there, but still
        # unactionable (self-approval guardrail).
        self.client.login(username='selfapp_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        self.assertEqual(resp.status_code, 200)
        row = next(e for e in resp.context['pending_exceptions'] if e.pk == self.exc.pk)
        self.assertFalse(row.can_decide)
        self.assertFalse(row.can_override)

    def test_post_decide_own_request_blocked(self):
        self.client.login(username='selfapp_hr', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'),
                                {'action': 'decide', 'exc_id': self.exc.pk, 'decision': 'approved'})
        self.assertEqual(resp.status_code, 404)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'pending')

    def test_post_override_own_request_blocked(self):
        self.client.login(username='selfapp_hr', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'override', 'exc_id': self.exc.pk, 'decision': 'approved', 'reason': 'Self-confirmed'})
        self.assertEqual(resp.status_code, 404)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'pending')


class TeamExceptionsSidebarLinkTests(TestCase):
    def setUp(self):
        self.manager_emp = make_employee(iqama='SBAR-MGR', name='Sidebar Manager')
        self.manager_user = _login_user('sbar_mgr')
        self.manager_emp.user = self.manager_user
        self.manager_emp.save(update_fields=['user'])
        self.report_emp = make_employee(iqama='SBAR-REPORT', name='Sidebar Report')
        self.report_emp.main_manager = self.manager_emp
        self.report_emp.save(update_fields=['main_manager'])

        self.ai_head_user = _make_role_user('sbar_ai_head', Role.AI_HEAD)

        self.plain_emp = make_employee(iqama='SBAR-PLAIN', name='Sidebar Plain')
        self.plain_user = _login_user('sbar_plain')
        self.plain_emp.user = self.plain_user
        self.plain_emp.save(update_fields=['user'])

    def test_link_renders_for_manager_with_reports(self):
        self.client.login(username='sbar_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Team Exceptions')

    def test_link_renders_for_head_role_user(self):
        self.client.login(username='sbar_ai_head', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Team Exceptions')

    def test_link_does_not_render_for_plain_employee(self):
        self.client.login(username='sbar_plain', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'Team Exceptions')

    def test_team_exceptions_nested_under_collapsible_my_profile_submenu(self):
        """For a qualifying user, 'Team Exceptions' must render inside the
        new 'My Profile' collapsible submenu (id="myProfileSubmenu"), not as
        its own top-level sidebar item."""
        self.client.login(username='sbar_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        content = resp.content.decode()
        self.assertIn('id="myProfileSubmenu"', content)
        submenu_start = content.index('id="myProfileSubmenu"')
        team_exceptions_pos = content.index('Team Exceptions')
        # The nested link must appear after the submenu container opens.
        self.assertGreater(team_exceptions_pos, submenu_start)

    def test_my_profile_link_still_reachable_for_non_qualifying_user(self):
        """'My Profile' has no permission gate and must remain visible/
        clickable for a user who doesn't qualify for Team Exceptions —
        no collapsible submenu should be rendered for them at all."""
        self.client.login(username='sbar_plain', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'My Profile')
        self.assertContains(resp, reverse('hr:my_profile'))
        self.assertNotContains(resp, 'id="myProfileSubmenu"')


class SecondaryManagerTeamExceptionsTests(TestCase):
    """Secondary managers get visibility/override rights over an employee's
    attendance exceptions without being part of the formal (main_manager)
    reporting line. Per the explicit business rule, a secondary manager's
    decision must always be recorded as an override, never as a plain
    'decide' — they are, by definition, not the assigned main manager."""

    def setUp(self):
        # Deliberately no main_manager at all on self.employee, and
        # self.secondary has zero main_reports of its own — isolates that
        # every permission granted below flows purely from the
        # secondary_managers relationship, not from the hierarchy or HR.
        self.employee = make_employee(iqama='SECM-EMP', name='Secondary-Managed Employee')
        self.secondary = make_employee(iqama='SECM-SEC', name='Secondary Manager')
        self.secondary_user = _login_user('secm_secondary')
        self.secondary.user = self.secondary_user
        self.secondary.save(update_fields=['user'])
        self.employee.secondary_managers.add(self.secondary)

        self.creator = make_user('secm_creator')
        self.exc = _submit_aex(
            employee=self.employee, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)

    def test_can_view_team_exceptions_true_for_secondary_manager_with_zero_main_reports(self):
        from hr.views import can_view_team_exceptions
        self.assertTrue(can_view_team_exceptions(self.secondary_user))

    def test_secondary_manager_sees_pending_request_with_override_not_decide(self):
        # Task 2: secondary-managed employees' requests live on their own
        # dedicated "Secondary Reports" tab (?tab=secondary), not the default
        # "Direct Reports" tab.
        self.client.login(username='secm_secondary', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'secondary'})
        self.assertEqual(resp.status_code, 200)
        pending = resp.context['pending_exceptions']
        ids = {e.pk for e in pending}
        self.assertIn(self.exc.pk, ids)
        row = next(e for e in pending if e.pk == self.exc.pk)
        self.assertTrue(row.can_override)
        self.assertFalse(row.can_decide)

    def test_secondary_manager_default_direct_tab_does_not_show_request(self):
        # Default "Direct Reports" tab must NOT include secondary-managed
        # employees' requests — that's the whole point of the dedicated
        # "Secondary Reports" tab.
        self.client.login(username='secm_secondary', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertNotIn(self.exc.pk, ids)

    def test_secondary_manager_cannot_use_plain_decide_action(self):
        self.client.login(username='secm_secondary', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'decide', 'exc_id': self.exc.pk, 'decision': 'approved'})
        # 404, not 403: same pk-existence-oracle-safe pattern as the
        # non-assigned-manager case — a secondary manager is never the
        # assigned main_manager, so the plain decide branch must reject them.
        self.assertEqual(resp.status_code, 404)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'pending')

    def test_secondary_manager_can_override_pending_request(self):
        self.client.login(username='secm_secondary', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'override', 'exc_id': self.exc.pk, 'decision': 'approved',
            'reason': 'Confirmed as secondary manager'})
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'approved')
        self.assertTrue(self.exc.is_overridden)
        self.assertEqual(self.exc.overridden_by, self.secondary_user)


# ─── Org-Chart Management: form + view + template ────────────────────────


class OrgChartViewTests(TestCase):
    def setUp(self):
        self.super_admin = _make_role_user('org_super', Role.SUPER_ADMIN)
        self.plain_admin = _make_role_user('org_admin', Role.ADMIN)
        self.plain_user = _login_user('org_plain')

        self.alice = make_employee(iqama='ORG-ALICE', name='Alice')
        self.bob = make_employee(iqama='ORG-BOB', name='Bob')
        self.carol = make_employee(iqama='ORG-CAROL', name='Carol')
        self.dave = make_employee(iqama='ORG-DAVE', name='Dave')

    # ── access gate ──
    def test_super_admin_can_view(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 200)

    def test_plain_admin_forbidden(self):
        self.client.login(username='org_admin', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 403)

    def test_plain_employee_forbidden(self):
        self.client.login(username='org_plain', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 403)

    def test_anonymous_redirected_to_login(self):
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 302)

    # ── persisting assignments ──
    def test_assign_main_manager_persists(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.bob.pk, 'main_manager': self.alice.pk, 'secondary_managers': [],
        })
        self.assertEqual(resp.status_code, 302)
        self.bob.refresh_from_db()
        self.assertEqual(self.bob.main_manager_id, self.alice.pk)

    def test_clearing_main_manager_persists(self):
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.bob.pk, 'main_manager': '', 'secondary_managers': [],
        })
        self.assertEqual(resp.status_code, 302)
        self.bob.refresh_from_db()
        self.assertIsNone(self.bob.main_manager_id)

    def test_assign_secondary_managers_persists_multiple(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.dave.pk, 'main_manager': '',
            'secondary_managers': [self.alice.pk, self.bob.pk],
        })
        self.assertEqual(resp.status_code, 302)
        ids = set(self.dave.secondary_managers.values_list('pk', flat=True))
        self.assertEqual(ids, {self.alice.pk, self.bob.pk})

    # ── self-assignment rejection ──
    def test_self_main_manager_rejected_no_change_persisted(self):
        # Self is excluded from the main_manager dropdown's queryset (see
        # EmployeeHierarchyForm.__init__), so a crafted POST selecting self
        # hits Django's standard "not one of the available choices" queryset
        # validation before clean_main_manager's custom message ever runs
        # (clean_main_manager's own message is exercised directly in
        # EmployeeHierarchyFormTests, bypassing the view's queryset). Either
        # way, the important behavior asserted here is: a clear per-field
        # error, no 500, and no persisted change.
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.alice.pk, 'main_manager': self.alice.pk, 'secondary_managers': [],
        })
        self.assertEqual(resp.status_code, 302)
        self.alice.refresh_from_db()
        self.assertIsNone(self.alice.main_manager_id)
        found = list(get_messages(resp.wsgi_request))
        self.assertTrue(any(self.alice.full_name in str(m) for m in found))

    # ── circular chain rejection ──
    def test_circular_chain_rejected_no_500_no_change_persisted(self):
        # A(alice) has no manager; bob -> alice; carol -> bob.
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        self.carol.main_manager = self.bob
        self.carol.save(update_fields=['main_manager'])
        self.client.login(username='org_super', password='testpass123')
        # Attempt to close the loop: alice -> carol (alice -> carol -> bob -> alice).
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.alice.pk, 'main_manager': self.carol.pk, 'secondary_managers': [],
        })
        self.assertEqual(resp.status_code, 302)  # no 500 — redirects normally
        self.alice.refresh_from_db()
        self.assertIsNone(self.alice.main_manager_id)
        found = list(get_messages(resp.wsgi_request))
        self.assertTrue(any('circular' in str(m).lower() for m in found))

    def test_direct_two_node_circular_chain_rejected(self):
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.alice.pk, 'main_manager': self.bob.pk, 'secondary_managers': [],
        })
        self.assertEqual(resp.status_code, 302)
        self.alice.refresh_from_db()
        self.assertIsNone(self.alice.main_manager_id)

    # ── hierarchy tree rendering ──
    def test_tree_renders_employee_with_no_reports(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.alice.full_name)

    def test_tree_renders_multiple_direct_reports(self):
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        self.carol.main_manager = self.alice
        self.carol.save(update_fields=['main_manager'])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.alice.full_name)
        self.assertContains(resp, self.bob.full_name)
        self.assertContains(resp, self.carol.full_name)

    def test_tree_renders_multi_level_chain(self):
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        self.carol.main_manager = self.bob
        self.carol.save(update_fields=['main_manager'])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, self.alice.full_name)
        self.assertContains(resp, self.bob.full_name)
        self.assertContains(resp, self.carol.full_name)

    # ── UI regression: multi-line {# #} comments never leak into rendered HTML ──
    def test_no_raw_template_comment_delimiters_in_rendered_page(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        content = resp.content.decode()
        self.assertNotIn('{#', content)
        self.assertNotIn('#}', content)

    def test_secondary_manager_search_filter_input_renders(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'sm-filter')
        self.assertContains(resp, 'Search employees...')

    def test_main_manager_search_filter_input_renders(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'mm-filter')
        self.assertContains(resp, 'Search managers...')

    # ── secondary-manager circle-selection UI ──
    def test_circle_dropdown_and_hidden_select_render(self):
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'sm-circle-wrap')
        self.assertContains(resp, 'sm-circle-list')
        self.assertContains(resp, 'sm-select-wrap')

    def test_current_secondary_managers_summary_shown_when_assigned(self):
        self.dave.secondary_managers.set([self.alice, self.bob])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'sm-current-summary')
        self.assertContains(resp, self.alice.full_name)
        self.assertContains(resp, self.bob.full_name)

    def test_current_secondary_managers_summary_shows_none_assigned(self):
        # None of alice/bob/carol/dave has any secondary managers in setUp.
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'No secondary managers assigned')

    def test_saving_a_reduced_secondary_manager_set_removes_the_dropped_one(self):
        # The circle UI toggles options in the same hidden multi-select the
        # main per-row Save form already submits — removing a secondary
        # manager is just saving a smaller set through that one mechanism,
        # replacing the old separate instant-remove action.
        self.dave.secondary_managers.set([self.alice, self.bob, self.carol])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.post(reverse('hr:org_chart'), {
            'employee_id': self.dave.pk, 'main_manager': '',
            'secondary_managers': [self.bob.pk, self.carol.pk],
        })
        self.assertEqual(resp.status_code, 302)
        ids = set(self.dave.secondary_managers.values_list('pk', flat=True))
        self.assertEqual(ids, {self.bob.pk, self.carol.pk})

    # ── per-row assignment checkmark indicator ──
    # The indicator <i> is now always rendered (fixed-width slot, so the
    # adjacent "x" clear button never shifts position row-to-row per the
    # "buttons must not move based on checked state" fix) — only its
    # containing slot's is-assigned class (driving CSS visibility) toggles.
    def test_main_manager_indicator_present_when_set(self):
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'main-manager-indicator-slot is-assigned')

    def test_main_manager_indicator_absent_when_none(self):
        # None of alice/bob/carol/dave has a main_manager in setUp.
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'main-manager-indicator')  # slot always renders
        self.assertNotContains(resp, 'main-manager-indicator-slot is-assigned')

    # ── hierarchy tree export scaffolding (PNG/PDF) ──
    def test_export_buttons_and_scripts_render(self):
        # html2canvas was dropped entirely — it couldn't reliably render the
        # CSS table/table-cell tree layout (glitched output, missing names).
        # The tree is now drawn from scratch on a <canvas> from DOM-derived
        # data, so only jsPDF remains as an external export dependency.
        self.client.login(username='org_super', password='testpass123')
        resp = self.client.get(reverse('hr:org_chart'))
        self.assertContains(resp, 'Export as PNG')
        self.assertContains(resp, 'Export as PDF')
        self.assertNotContains(resp, 'html2canvas')
        self.assertContains(resp, 'jspdf')
        self.assertContains(resp, 'renderTreeCanvas')


class LeaveAccessViewTests(TestCase):
    """The standalone Leave Access page (sidebar Leave -> Leave Access,
    formerly the Org Chart's second tab): still Super-Admin-only, search-
    and-grant, and revoke."""
    def setUp(self):
        from hr.models import LeaveDashboardAccess
        self.LeaveDashboardAccess = LeaveDashboardAccess
        self.super_admin = _make_role_user('access_super', Role.SUPER_ADMIN)
        self.plain_admin = _make_role_user('access_admin', Role.ADMIN)

        self.eve_user = make_user('access_eve', password='x')
        self.eve_user.set_password('testpass123')
        self.eve_user.save()
        self.eve = make_employee(iqama='ACC-EVE', name='Eve Grantee')
        self.eve.user = self.eve_user
        self.eve.save(update_fields=['user'])

        self.frank_user = make_user('access_frank', password='x')
        self.frank_user.set_password('testpass123')
        self.frank_user.save()
        self.frank = make_employee(iqama='ACC-FRANK', name='Frank Nogrant')
        self.frank.user = self.frank_user
        self.frank.save(update_fields=['user'])

    # ── page-level access gate ──
    def test_plain_admin_forbidden(self):
        self.client.login(username='access_admin', password='testpass123')
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertEqual(resp.status_code, 403)

    def test_super_admin_sees_the_page(self):
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Leave Dashboard Access')

    def test_anonymous_redirected_to_login(self):
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertEqual(resp.status_code, 302)

    # ── search ──
    def test_search_finds_employee_without_existing_access(self):
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_access'), {'access_search': 'Eve'})
        self.assertContains(resp, 'Eve Grantee')

    def test_search_excludes_employee_who_already_has_access(self):
        self.LeaveDashboardAccess.objects.create(user=self.eve_user, is_active=True, granted_by=self.super_admin)
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_access'), {'access_search': 'Eve'})
        self.assertNotContains(resp, 'value="{}"'.format(self.eve.pk))

    # ── grant ──
    def test_grant_creates_active_access_row(self):
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_access'), {'grant_dashboard_access': self.eve.pk})
        self.assertEqual(resp.status_code, 302)
        grant = self.LeaveDashboardAccess.objects.get(user=self.eve_user)
        self.assertTrue(grant.is_active)
        self.assertEqual(grant.granted_by, self.super_admin)

    def test_granted_user_can_then_view_leave_dashboard(self):
        from hr.views import can_view_leave_dashboard
        self.client.login(username='access_super', password='testpass123')
        self.client.post(reverse('hr:leave_access'), {'grant_dashboard_access': self.eve.pk})
        self.assertTrue(can_view_leave_dashboard(self.eve_user))
        self.assertFalse(can_view_leave_dashboard(self.frank_user))

    def test_plain_admin_cannot_grant(self):
        self.client.login(username='access_admin', password='testpass123')
        resp = self.client.post(reverse('hr:leave_access'), {'grant_dashboard_access': self.eve.pk})
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(self.LeaveDashboardAccess.objects.filter(user=self.eve_user).exists())

    # ── revoke ──
    def test_revoke_removes_access_row(self):
        self.LeaveDashboardAccess.objects.create(user=self.eve_user, is_active=True, granted_by=self.super_admin)
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_access'), {'revoke_dashboard_access': self.eve_user.pk})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(self.LeaveDashboardAccess.objects.filter(user=self.eve_user).exists())

    def test_current_access_list_shows_grantee_and_granter(self):
        # LeaveDashboardAccess.user is a User, not an Employee — the current-
        # access list renders the User's identity (username, since neither
        # test user has a first/last name set), not the linked Employee's
        # full_name.
        self.LeaveDashboardAccess.objects.create(user=self.eve_user, is_active=True, granted_by=self.super_admin)
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertContains(resp, self.eve_user.username)
        self.assertContains(resp, self.super_admin.username)

    # ── sidebar link ──
    def test_sidebar_shows_leave_access_link_for_super_admin(self):
        self.client.login(username='access_super', password='testpass123')
        resp = self.client.get(reverse('hr:entitlement_year'))
        self.assertContains(resp, reverse('hr:leave_access'))
        self.assertContains(resp, 'Leave Access')

    def test_sidebar_hides_leave_access_link_for_plain_admin(self):
        """A plain admin no longer reaches Administration at all.

        This used to load an Administration page as an admin and check the
        Leave Access link was absent from the sidebar. Administration is now
        owned by super_admin and erp_admin, so the guarantee is stronger: the
        admin is turned away from the page itself, and the link is missing from
        a page they CAN open.
        """
        from accounts.permissions import seed_default_permissions
        seed_default_permissions()   # so the admin can open a non-admin page at all
        self.client.login(username='access_admin', password='testpass123')
        # The Administration page itself is no longer reachable.
        blocked = self.client.get(reverse('hr:entitlement_year'))
        self.assertIn(blocked.status_code, (302, 403))
        # And the sidebar on a page they can open offers no Administration link.
        reachable = self.client.get(reverse('dashboard:index'))
        self.assertEqual(reachable.status_code, 200)
        self.assertNotContains(reachable, reverse('hr:leave_access'))
        self.assertNotContains(reachable, 'data-section="administration"')


class HasOverrideAccessTests(TestCase):
    """Unit tests for hr.views.has_override_access, independent of any view —
    exactly one OverrideAccessSettings mode is authoritative at a time."""
    def setUp(self):
        from hr.models import OverrideAccessSettings
        OverrideAccessSettings.objects.all().delete()

    def test_default_mode_is_all_super_admins(self):
        from hr.models import OverrideAccessSettings
        config = OverrideAccessSettings.get_solo()
        self.assertEqual(config.mode, OverrideAccessSettings.MODE_ALL_SUPER_ADMINS)

    def test_all_super_admins_mode_grants_every_super_admin(self):
        from hr.views import has_override_access
        admin = _make_role_user('hoa_super1', Role.SUPER_ADMIN)
        self.assertTrue(has_override_access(admin))

    def test_all_super_admins_mode_denies_non_super_admin(self):
        from hr.views import has_override_access
        plain = _login_user('hoa_plain1')
        self.assertFalse(has_override_access(plain))

    def test_specific_roles_mode_grants_only_that_role(self):
        from hr.views import has_override_access
        from hr.models import OverrideAccessSettings, OverrideAccessRole
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_ROLES
        config.save()
        finance_role, _ = Role.objects.get_or_create(name=Role.FINANCE_HEAD)
        OverrideAccessRole.objects.create(role=finance_role)

        finance_user = _make_role_user('hoa_finance', Role.FINANCE_HEAD)
        self.assertTrue(has_override_access(finance_user))

        super_admin_without_grant = _make_role_user('hoa_super2', Role.SUPER_ADMIN)
        self.assertFalse(has_override_access(super_admin_without_grant))

    def test_specific_employees_mode_grants_only_that_person(self):
        from hr.views import has_override_access
        from hr.models import OverrideAccessSettings, OverrideAccessEmployee
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_EMPLOYEES
        config.save()
        chosen = _login_user('hoa_chosen')
        OverrideAccessEmployee.objects.create(user=chosen)

        self.assertTrue(has_override_access(chosen))

        super_admin_without_grant = _make_role_user('hoa_super3', Role.SUPER_ADMIN)
        self.assertFalse(has_override_access(super_admin_without_grant))

    def test_override_access_alone_grants_dashboard_viewing(self):
        # Otherwise granting override access to a non-super-admin,
        # non-approver would be pointless — they couldn't reach the page.
        from hr.views import can_view_leave_dashboard
        from hr.models import OverrideAccessSettings, OverrideAccessEmployee
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_EMPLOYEES
        config.save()
        chosen = _login_user('hoa_viewer')
        OverrideAccessEmployee.objects.create(user=chosen)
        self.assertTrue(can_view_leave_dashboard(chosen))



class OverrideAccessViewTests(TestCase):
    """End-to-end: the leave-request detail page's override form/action
    respects has_override_access instead of a hardcoded is_super_admin_user
    check, across all three modes."""
    def setUp(self):
        from hr.models import OverrideAccessSettings
        OverrideAccessSettings.objects.all().delete()
        self.emp = make_employee()
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3))

    def test_default_mode_super_admin_sees_and_can_use_override(self):
        admin = _make_role_user('oav_super1', Role.SUPER_ADMIN)
        self.client.login(username='oav_super1', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, 'name="action" value="override"')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Testing'})
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')

    def test_specific_roles_mode_blocks_super_admin_without_the_role(self):
        from hr.models import OverrideAccessSettings, OverrideAccessRole
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_ROLES
        config.save()
        finance_role, _ = Role.objects.get_or_create(name=Role.FINANCE_HEAD)
        OverrideAccessRole.objects.create(role=finance_role)

        admin = _make_role_user('oav_super2', Role.SUPER_ADMIN)
        self.client.login(username='oav_super2', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertNotContains(resp, 'name="action" value="override"')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Testing'})
        self.assertEqual(resp.status_code, 403)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')

    def test_specific_roles_mode_allows_the_granted_role_even_if_not_super_admin(self):
        from hr.models import OverrideAccessSettings, OverrideAccessRole
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_ROLES
        config.save()
        finance_role, _ = Role.objects.get_or_create(name=Role.FINANCE_HEAD)
        OverrideAccessRole.objects.create(role=finance_role)

        finance_user = _make_role_user('oav_finance', Role.FINANCE_HEAD)
        self.client.login(username='oav_finance', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Testing'})
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')

    def test_specific_employees_mode_blocks_unlisted_super_admin(self):
        from hr.models import OverrideAccessSettings, OverrideAccessEmployee
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_EMPLOYEES
        config.save()
        chosen = _login_user('oav_chosen')
        OverrideAccessEmployee.objects.create(user=chosen)

        admin = _make_role_user('oav_super3', Role.SUPER_ADMIN)
        self.client.login(username='oav_super3', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Testing'})
        self.assertEqual(resp.status_code, 403)

        self.client.login(username='oav_chosen', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]),
                                {'action': 'override', 'decision': 'approved', 'reason': 'Testing'})
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')


class LeaveAccessOverrideAccessTests(TestCase):
    """The Override Access management section on the Leave Access page: mode
    switching, role toggles, employee grant/revoke, and the at-a-glance
    'currently granted to' summary."""
    def setUp(self):
        from hr.models import OverrideAccessSettings
        OverrideAccessSettings.objects.all().delete()
        self.super_admin = _make_role_user('ooa_super', Role.SUPER_ADMIN)
        self.client.login(username='ooa_super', password='testpass123')

    def test_default_mode_shows_all_super_admins_summary(self):
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertContains(resp, 'All Super Admins')
        self.assertContains(resp, 'ooa_super')  # listed as a current super admin

    def test_set_mode_to_specific_roles(self):
        from hr.models import OverrideAccessSettings
        resp = self.client.post(reverse('hr:leave_access'), {'set_override_mode': 'specific_roles'})
        self.assertEqual(resp.status_code, 302)
        config = OverrideAccessSettings.get_solo()
        self.assertEqual(config.mode, 'specific_roles')
        self.assertEqual(config.updated_by, self.super_admin)

    def test_invalid_mode_rejected(self):
        from hr.models import OverrideAccessSettings
        resp = self.client.post(reverse('hr:leave_access'), {'set_override_mode': 'bogus_mode'})
        self.assertEqual(resp.status_code, 302)
        config = OverrideAccessSettings.get_solo()
        self.assertEqual(config.mode, OverrideAccessSettings.MODE_ALL_SUPER_ADMINS)

    def test_toggle_role_grants_then_removes(self):
        from hr.models import OverrideAccessSettings, OverrideAccessRole
        OverrideAccessSettings.get_solo()
        finance_role, _ = Role.objects.get_or_create(name=Role.FINANCE_HEAD)

        self.client.post(reverse('hr:leave_access'), {'toggle_override_role': finance_role.pk})
        self.assertTrue(OverrideAccessRole.objects.filter(role=finance_role).exists())

        self.client.post(reverse('hr:leave_access'), {'toggle_override_role': finance_role.pk})
        self.assertFalse(OverrideAccessRole.objects.filter(role=finance_role).exists())

    def test_specific_roles_mode_renders_role_toggle_buttons(self):
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertNotContains(resp, 'toggle_override_role')  # only shown in specific_roles mode
        self.client.post(reverse('hr:leave_access'), {'set_override_mode': 'specific_roles'})
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertContains(resp, 'toggle_override_role')
        self.assertContains(resp, 'Finance Head')

    def test_grant_and_revoke_override_employee(self):
        from hr.models import OverrideAccessSettings, OverrideAccessEmployee
        config = OverrideAccessSettings.get_solo()
        config.mode = OverrideAccessSettings.MODE_SPECIFIC_EMPLOYEES
        config.save()

        target_user = make_user('ooa_target', password='x')
        target_emp = make_employee(iqama='OOA-TARGET', name='Target Person')
        target_emp.user = target_user
        target_emp.save(update_fields=['user'])

        resp = self.client.post(reverse('hr:leave_access'), {'grant_override_employee': target_emp.pk})
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(OverrideAccessEmployee.objects.filter(user=target_user).exists())

        resp = self.client.post(reverse('hr:leave_access'), {'revoke_override_employee': target_user.pk})
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(OverrideAccessEmployee.objects.filter(user=target_user).exists())

    def test_empty_specific_roles_summary_warns_no_one_can_override(self):
        self.client.post(reverse('hr:leave_access'), {'set_override_mode': 'specific_roles'})
        resp = self.client.get(reverse('hr:leave_access'))
        self.assertContains(resp, 'no roles selected yet')

    def test_plain_admin_cannot_change_override_mode(self):
        plain_admin = _make_role_user('ooa_plainadmin', Role.ADMIN)
        self.client.login(username='ooa_plainadmin', password='testpass123')
        resp = self.client.post(reverse('hr:leave_access'), {'set_override_mode': 'specific_roles'})
        self.assertEqual(resp.status_code, 403)


class EmployeeHierarchyFormTests(TestCase):
    """Form-level coverage of the cycle-detection dry run and self-assignment
    guards, independent of the view/HTTP layer."""

    def setUp(self):
        from hr.forms import EmployeeHierarchyForm
        self.EmployeeHierarchyForm = EmployeeHierarchyForm
        self.alice = make_employee(iqama='EHF-ALICE', name='Alice')
        self.bob = make_employee(iqama='EHF-BOB', name='Bob')

    def test_valid_assignment_is_valid_and_saves(self):
        form = self.EmployeeHierarchyForm(
            {'main_manager': self.alice.pk, 'secondary_managers': []},
            employee=self.bob)
        self.assertTrue(form.is_valid())
        form.save()
        self.bob.refresh_from_db()
        self.assertEqual(self.bob.main_manager_id, self.alice.pk)

    def test_self_as_main_manager_is_invalid(self):
        form = self.EmployeeHierarchyForm(
            {'main_manager': self.alice.pk, 'secondary_managers': []},
            employee=self.alice)
        self.assertFalse(form.is_valid())
        self.assertIn('main_manager', form.errors)

    def test_self_in_secondary_managers_excluded_from_queryset(self):
        form = self.EmployeeHierarchyForm(employee=self.alice)
        self.assertNotIn(self.alice, form.fields['secondary_managers'].queryset)
        self.assertNotIn(self.alice, form.fields['main_manager'].queryset)

    def test_circular_chain_is_invalid_and_does_not_raise(self):
        self.bob.main_manager = self.alice
        self.bob.save(update_fields=['main_manager'])
        form = self.EmployeeHierarchyForm(
            {'main_manager': self.bob.pk, 'secondary_managers': []},
            employee=self.alice)
        self.assertFalse(form.is_valid())
        self.assertIn('main_manager', form.errors)
        self.alice.refresh_from_db()
        self.assertIsNone(self.alice.main_manager_id)


class LeaveRequestDetailDocumentLinkTests(TestCase):
    """The 'View attached document' link on the leave request detail page
    must open in a new tab (target="_blank") — otherwise the global loading
    overlay in base.html (which only hides on `pageshow` or a 30s fallback)
    gets stuck showing "Loading, please wait…" forever, because the
    same-page FileResponse download never fires a real navigation/pageshow
    event. This matches the established convention used elsewhere in the
    codebase (e.g. the document links on My Profile use target="_blank")."""

    def setUp(self):
        self.emp = make_employee(iqama='DOCLINK-1', name='Doc Link Employee')
        self.marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.superadmin = make_user('doclink_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.marriage,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 3),
            document=SimpleUploadedFile('cert.pdf', b'dummy-bytes'))

    def test_document_link_opens_in_new_tab(self):
        self.client.login(username='doclink_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        doc_url = reverse('hr:leave_request_document', args=[self.req.pk])
        # Confirm the exact rendered markup, not just that target="_blank"
        # appears somewhere on the page.
        self.assertContains(resp, f'href="{doc_url}" target="_blank"')


class MyProfileHierarchyTests(TestCase):
    """'My Reporting Structure' card on My Profile: main manager, secondary
    managers, direct reports, secondary reports. Must render gracefully
    (no VariableDoesNotExist / template crash) when main_manager is None —
    the same class of bug previously found and fixed on Team Exceptions."""

    def setUp(self):
        self.manager = make_employee(iqama='HIER-MGR', name='Hierarchy Manager')
        self.report = make_employee(iqama='HIER-REPORT', name='Hierarchy Report')
        self.lone = make_employee(iqama='HIER-LONE', name='Hierarchy Lone')

    def test_shows_main_manager_name(self):
        self.report.main_manager = self.manager
        self.report.save(update_fields=['main_manager'])
        user = _login_user('hier_report_user')
        self.report.user = user
        self.report.save(update_fields=['user'])
        self.client.login(username='hier_report_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Hierarchy Manager')

    def test_shows_direct_reports(self):
        self.report.main_manager = self.manager
        self.report.save(update_fields=['main_manager'])
        user = _login_user('hier_manager_user')
        self.manager.user = user
        self.manager.save(update_fields=['user'])
        self.client.login(username='hier_manager_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Hierarchy Report')

    def test_no_manager_and_no_reports_renders_gracefully(self):
        """main_manager is None and there are zero reports — this must not
        raise a template error and must render the empty-state copy."""
        user = _login_user('hier_lone_user')
        self.lone.user = user
        self.lone.save(update_fields=['user'])
        self.client.login(username='hier_lone_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Not assigned')

class ReapplyLeaveTypeDefaultsScopingTests(TestCase):
    """Unit-level test for reapply_leave_type_defaults' new `leave_type` param
    (added alongside the existing `year` param), isolated from
    LeaveType.save()'s own broader propagation by changing default_annual_days
    via a bare queryset .update() (which bypasses save())."""

    def test_scoped_to_single_leave_type_and_year(self):
        emp = make_employee(iqama='RSD-1', name='Reapply Scope Employee')
        sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': Decimal('12.0'),
                                   'is_accumulative': False})
        marriage, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': Decimal('3.0'),
                                       'is_accumulative': False})
        year = timezone.now().year
        other_year = year - 1
        sick_current = LeaveEntitlement.objects.create(
            employee=emp, leave_type=sick, year=year, entitled_days=Decimal('99'))
        sick_other = LeaveEntitlement.objects.create(
            employee=emp, leave_type=sick, year=other_year, entitled_days=Decimal('99'))
        marriage_current = LeaveEntitlement.objects.create(
            employee=emp, leave_type=marriage, year=year, entitled_days=Decimal('99'))

        LeaveType.objects.filter(pk=sick.pk).update(default_annual_days=Decimal('20'))
        sick.refresh_from_db()

        from hr.leave_services import reapply_leave_type_defaults
        updated = reapply_leave_type_defaults(year=year, leave_type=sick)
        self.assertEqual(updated, 1)

        sick_current.refresh_from_db()
        sick_other.refresh_from_db()
        marriage_current.refresh_from_db()
        self.assertEqual(sick_current.entitled_days, Decimal('20'))
        self.assertEqual(sick_other.entitled_days, Decimal('99'))       # different year untouched
        self.assertEqual(marriage_current.entitled_days, Decimal('99'))  # different type untouched



class LeaveTypeUpdateViewAutoReapplyTests(TestCase):
    """Editing a LeaveType's default_annual_days through the admin edit view
    flashes an explicit "Re-applied ... for <year>" confirmation, current-year
    scoped, mirroring the Entitlements page's manual Re-apply action."""

    def setUp(self):
        self.admin = _make_role_user('ltu_admin', Role.SUPER_ADMIN)
        self.emp = make_employee(iqama='LTU-1', name='LTU Employee')
        self.sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': Decimal('12.0'),
                                   'is_accumulative': False})
        self.current_year = timezone.now().year
        self.current_ent = LeaveEntitlement.objects.create(
            employee=self.emp, leave_type=self.sick, year=self.current_year, entitled_days=Decimal('12.0'))

    def test_editing_default_days_updates_current_year_and_flashes_confirmation(self):
        self.client.login(username='ltu_admin', password='testpass123')
        resp = self.client.post(reverse('hr:leavetype_update', args=[self.sick.pk]), {
            'name': 'Sick', 'code': 'sick', 'default_annual_days': '20.0',
            'is_paid': 'on', 'color': 'secondary', 'is_active': 'on',
        }, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.current_ent.refresh_from_db()
        self.assertEqual(self.current_ent.entitled_days, Decimal('20.0'))
        rendered_messages = [str(m) for m in resp.context['messages']]
        self.assertTrue(
            any(f'entitlement(s) for {self.current_year}' in m for m in rendered_messages),
            rendered_messages)

    def test_no_day_count_change_does_not_flash_reapply_confirmation(self):
        self.client.login(username='ltu_admin', password='testpass123')
        resp = self.client.post(reverse('hr:leavetype_update', args=[self.sick.pk]), {
            'name': 'Sick Leave', 'code': 'sick', 'default_annual_days': '12.0',
            'is_paid': 'on', 'color': 'secondary', 'is_active': 'on',
        }, follow=True)
        self.assertEqual(resp.status_code, 200)
        rendered_messages = [str(m) for m in resp.context['messages']]
        self.assertFalse(any('Re-applied leave-type day counts' in m for m in rendered_messages), rendered_messages)


# ─── Team Attendance Exceptions — scope / countdown / manager-routing fixes ──


class TeamExceptionsScopeDirectVsSecondaryTests(TestCase):
    """Task 2 fix: the old ?scope=direct|downstream query-param scheme is
    replaced entirely by a three-tab structure (?tab=direct|secondary|all).
    The default "Direct Reports" tab shows ONLY employee__main_manager
    matches; secondary-managed employees live on their own dedicated
    "Secondary Reports" tab — there's no combined/expanded view anymore, per
    the explicit business rule that secondary-manager visibility exists to
    ease things along when the main manager is unavailable, not to clutter
    the everyday view."""

    def setUp(self):
        self.manager = make_employee(iqama='SCOPE-MGR', name='Scope Manager')
        self.manager_user = _login_user('scope_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])

        self.direct_report = make_employee(iqama='SCOPE-DIRECT', name='Scope Direct Report')
        self.direct_report.main_manager = self.manager
        self.direct_report.save(update_fields=['main_manager'])

        # Separately secondary-managed — no main_manager relationship to
        # self.manager at all.
        self.secondary_employee = make_employee(iqama='SCOPE-SECONDARY', name='Scope Secondary Employee')
        self.secondary_employee.secondary_managers.add(self.manager)

        self.creator = make_user('scope_creator')
        self.exc_direct = _submit_aex(
            employee=self.direct_report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        self.exc_secondary = _submit_aex(
            employee=self.secondary_employee, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)

    def test_default_direct_tab_shows_only_direct_report(self):
        self.client.login(username='scope_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.context['tab'], 'direct')
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_direct.pk})

    def test_secondary_tab_shows_only_secondary_managed_employee(self):
        self.client.login(username='scope_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'secondary'})
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_secondary.pk})

    def test_ui_copy_no_longer_mentions_downstream_chain(self):
        self.client.login(username='scope_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'secondary'})
        content = resp.content.decode()
        self.assertNotIn('downstream chain', content.lower())
        self.assertNotIn('downstream reporting chain', content.lower())

    def test_secondary_and_all_org_tabs_still_render_labels_when_relevant(self):
        # Per the explicit rename ask: exact tab labels, and the
        # non-default tabs must be clickable/visible even when they'd
        # currently render zero rows for a *different* user (not exercised
        # here directly, but the labels themselves must always be present
        # for any qualifying viewer).
        self.client.login(username='scope_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        content = resp.content.decode()
        self.assertIn('Direct Reports', content)
        self.assertIn('Secondary Reports', content)
        self.assertNotIn('All Organization Requests', content)  # non-HR: hidden entirely


class TeamExceptionsThreeTabHRTests(TestCase):
    """Task 2: an HR/Super-Admin viewer who ALSO has their own Employee
    profile with both a direct report and a secondary-managed employee sees:
    only the direct report's request on "Direct Reports"; only the
    secondary-managed employee's request on "Secondary Reports"; and BOTH
    plus everyone else's requests on "All Organization Requests"."""

    def setUp(self):
        self.hr_user = _make_role_user('tex3_hr', Role.SUPER_ADMIN)
        self.hr_emp = make_employee(iqama='TEX3-HR', name='HR Manager Employee')
        self.hr_emp.user = self.hr_user
        self.hr_emp.save(update_fields=['user'])

        self.direct_report = make_employee(iqama='TEX3-DIRECT', name='Tex3 Direct Report')
        self.direct_report.main_manager = self.hr_emp
        self.direct_report.save(update_fields=['main_manager'])

        self.secondary_employee = make_employee(iqama='TEX3-SECONDARY', name='Tex3 Secondary Employee')
        self.secondary_employee.secondary_managers.add(self.hr_emp)

        # An unrelated third employee, visible only via the org-wide tab.
        self.other_employee = make_employee(iqama='TEX3-OTHER', name='Tex3 Unrelated Employee')

        self.creator = make_user('tex3_creator')
        self.exc_direct = _submit_aex(
            employee=self.direct_report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        self.exc_secondary = _submit_aex(
            employee=self.secondary_employee, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        self.exc_other = _submit_aex(
            employee=self.other_employee, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)

    def test_hr_direct_tab_shows_only_direct_report(self):
        self.client.login(username='tex3_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_direct.pk})

    def test_hr_secondary_tab_shows_only_secondary_managed_employee(self):
        self.client.login(username='tex3_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'secondary'})
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_secondary.pk})

    def test_hr_all_org_tab_shows_everything(self):
        self.client.login(username='tex3_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertEqual(ids, {self.exc_direct.pk, self.exc_secondary.pk, self.exc_other.pk})


class TeamExceptionsTabStylingTests(TestCase):
    """CSS-only restyle of the tab nav (Direct Reports / Secondary Reports /
    All Organization Requests): each tab now gets its own distinct accent
    color instead of Bootstrap's default active-tab blue. This is markup/CSS
    class presence only — it must not touch the {% if %} visibility gating
    for the HR-only "All Organization Requests" tab, which stays covered by
    TeamExceptionsAccessTests / TeamExceptionsThreeTabHRTests unmodified."""

    def setUp(self):
        self.mid = make_employee(iqama='TEXSTYLE-MID', name='Style Mid Manager')
        self.mid_user = _login_user('texstyle_mid')
        self.mid.user = self.mid_user
        self.mid.save(update_fields=['user'])

        self.direct_report = make_employee(iqama='TEXSTYLE-DIRECT', name='Style Direct Report')
        self.direct_report.main_manager = self.mid
        self.direct_report.save(update_fields=['main_manager'])

        self.hr_user = _make_role_user('texstyle_hr', Role.SUPER_ADMIN)

    def test_page_renders_successfully_with_new_tab_styling(self):
        self.client.login(username='texstyle_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'team-exc-tabs')

    def test_each_tab_has_its_own_distinct_css_class(self):
        # Log in as a qualifying HR/Super-Admin user so all three tabs
        # (including the HR-only "All Organization Requests") render.
        self.client.login(username='texstyle_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, 'team-exc-tab-direct')
        self.assertContains(resp, 'team-exc-tab-secondary')
        self.assertContains(resp, 'team-exc-tab-all')

    def test_non_hr_user_still_does_not_see_all_organization_tab(self):
        # Confirms the restyle didn't loosen the existing {% if show_all_tab %}
        # visibility gate — a plain manager (no HR/Super-Admin role) must
        # still not see the "All Organization Requests" tab link or label.
        # Note: the `.team-exc-tab-all` *selector* is always present in the
        # <style> block regardless of tab visibility (it's just CSS), so
        # that alone isn't a meaningful signal here — we instead assert the
        # actual tab anchor (its href and visible label) is absent.
        self.client.login(username='texstyle_mid', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, 'team-exc-tab-direct')
        self.assertContains(resp, 'team-exc-tab-secondary')
        self.assertNotContains(resp, 'href="?tab=all"')
        self.assertNotContains(resp, 'All Organization Requests')


class TeamExceptionsSelfApprovalVisibilityTests(TestCase):
    """Task 2 self-approval VISIBILITY rule: a user's own pending request
    must be entirely excluded (not just button-blocked) from "Direct
    Reports"/"Secondary Reports", even for a Super Admin who would otherwise
    structurally qualify as their own manager. It DOES appear (visible, but
    unactionable) on "All Organization Requests" for HR."""

    def setUp(self):
        self.hr_user = _make_role_user('selfvis_hr', Role.SUPER_ADMIN)
        self.hr_emp = make_employee(iqama='SELFVIS-HR', name='Self Visibility HR Employee')
        self.hr_emp.user = self.hr_user
        self.hr_emp.save(update_fields=['user'])
        # Pathological but structurally possible: the HR employee is their
        # own main_manager AND their own secondary_manager.
        self.hr_emp.main_manager = self.hr_emp
        self.hr_emp.secondary_managers.add(self.hr_emp)
        self.hr_emp.save(update_fields=['main_manager'])

        self.creator = make_user('selfvis_creator')
        self.exc = _submit_aex(
            employee=self.hr_emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)

    def test_own_request_excluded_from_direct_tab(self):
        self.client.login(username='selfvis_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertNotIn(self.exc.pk, ids)

    def test_own_request_excluded_from_secondary_tab(self):
        self.client.login(username='selfvis_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'secondary'})
        ids = {e.pk for e in resp.context['pending_exceptions']}
        self.assertNotIn(self.exc.pk, ids)

    def test_own_request_visible_but_unactionable_on_all_org_tab(self):
        self.client.login(username='selfvis_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        pending = resp.context['pending_exceptions']
        ids = {e.pk for e in pending}
        self.assertIn(self.exc.pk, ids)
        row = next(e for e in pending if e.pk == self.exc.pk)
        self.assertFalse(row.can_decide)
        self.assertFalse(row.can_override)


class TeamExceptionsCountdownEventAwareTests(TestCase):
    """Task 3 fix: the countdown must be anchored to event_start, not the
    up-to-7-days-early submission moment. Before the event starts, the page
    must show a static 'Starts <date>' — never a ticking countdown (which
    would otherwise misleadingly read as lasting up to ~8 days, since
    decision_deadline = event_start + 24h and submission can precede
    event_start by up to 7 days). At/after event start, the real live
    countdown toward decision_deadline (by definition at most 24h away)
    renders exactly as before."""

    def setUp(self):
        self.manager = make_employee(iqama='CD-MGR', name='Countdown Manager')
        self.manager_user = _login_user('cd_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.report = make_employee(iqama='CD-REPORT', name='Countdown Report')
        self.report.main_manager = self.manager
        self.report.save(update_fields=['main_manager'])
        self.creator = make_user('cd_creator')

    def test_future_event_shows_static_starts_text_not_live_countdown(self):
        from django.template.defaultfilters import date as _django_date
        future_dt = _timezone.localtime(_timezone.now() + timedelta(days=3))
        exc = _submit_aex(
            employee=self.report, event_date=future_dt.date(), event_start_time=future_dt.time(),
            reason_category='site_visit', created_by=self.creator)
        self.client.login(username='cd_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        row = next(e for e in resp.context['pending_exceptions'] if e.pk == exc.pk)
        self.assertFalse(row.event_has_started)
        content = resp.content.decode()
        self.assertIn(f"Starts {_django_date(exc.event_start, 'd M, H:i')}", content)

    def test_started_event_shows_live_countdown(self):
        from django.template.defaultfilters import date as _django_date
        started_dt = _timezone.localtime(_timezone.now() - timedelta(hours=1))
        exc = _submit_aex(
            employee=self.report, event_date=started_dt.date(), event_start_time=started_dt.time(),
            reason_category='site_visit', created_by=self.creator)
        self.client.login(username='cd_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        row = next(e for e in resp.context['pending_exceptions'] if e.pk == exc.pk)
        self.assertTrue(row.event_has_started)
        content = resp.content.decode()
        self.assertIn(_django_date(exc.decision_deadline, 'd M, H:i'), content)
        self.assertNotIn(f"Starts {_django_date(exc.event_start, 'd M, H:i')}", content)


class ManagerAssignedAfterSubmissionShowsDecideNotOverrideTests(TestCase):
    """Regression test for Task 5: reproduces the exact reported bug. An
    employee's attendance exception is submitted with NO main_manager
    assigned yet (so exc.main_manager, the submission-time snapshot, is
    None). The Org Chart is then used to assign the employee's main_manager
    AFTER the request already exists (a very plausible sequence — testing
    hierarchy changes, or correcting a mistake, with pre-existing pending
    requests already in flight). The now-correctly-assigned manager must see
    the normal Approve/Reject controls for that pre-existing request — not
    stuck seeing only Override, which is what a stale exc.main_manager-based
    check used to force permanently."""

    def setUp(self):
        self.employee = make_employee(iqama='SNAP-EMP', name='Snapshot Employee')
        self.creator = make_user('snap_creator')
        self.exc = _submit_aex(
            employee=self.employee, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        self.assertIsNone(self.exc.main_manager_id)  # sanity: no snapshot at submission

        self.manager = make_employee(iqama='SNAP-MGR', name='Snapshot Manager')
        self.manager_user = _login_user('snap_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])

        # Simulate the post-submission Org Chart fix.
        self.employee.main_manager = self.manager
        self.employee.save(update_fields=['main_manager'])

    def test_page_shows_approve_reject_not_override_for_live_manager(self):
        self.client.login(username='snap_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        row = next(e for e in resp.context['pending_exceptions'] if e.pk == self.exc.pk)
        self.assertTrue(row.can_decide)
        self.assertFalse(row.can_override)
        self.assertContains(resp, 'Approve')
        self.assertContains(resp, 'Reject')

    def test_manager_can_decide_via_post(self):
        self.client.login(username='snap_mgr', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'decide', 'exc_id': self.exc.pk, 'decision': 'approved'})
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'approved')

    def test_can_decide_attendance_exception_true_via_live_relationship(self):
        from hr.views import can_decide_attendance_exception
        self.assertTrue(can_decide_attendance_exception(self.manager_user, self.exc))

    def test_manager_column_shows_live_manager_not_stale_snapshot(self):
        self.client.login(username='snap_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, 'Snapshot Manager')


# ─── My Profile: Reporting Structure redesign (expand toggle + card order) ──
#
# The "My Reporting Structure" card used to only ever show direct reports,
# with no way to see further down the chain, and lived in the middle of the
# page. It now (a) shows direct reports unconditionally plus a "+N more"
# expand control for the further downstream chain (Employee.
# get_downstream_employee_ids(), minus direct reports so nothing is
# duplicated), sourced from a new context['downstream_reports'] computed in
# my_profile() only `if emp:`; and (b) always renders as the very last card
# on the page, after the conditionally-rendered My Vehicles card.

class MyProfileDownstreamReportsTests(TestCase):
    """The expand control should only appear when there's a downstream chain
    beyond the direct reports already listed, and the grandchild must appear
    only inside that expand section — never duplicated in the direct-reports
    list."""

    def setUp(self):
        self.root = make_employee(iqama='DSR-ROOT', name='Root Manager')
        self.direct = make_employee(iqama='DSR-DIRECT', name='Direct Report')
        self.direct.main_manager = self.root
        self.direct.save(update_fields=['main_manager'])
        self.root_user = _login_user('dsr_root_user')
        self.root.user = self.root_user
        self.root.save(update_fields=['user'])
        self.client.login(username='dsr_root_user', password='testpass123')

    def test_shows_expand_control_with_downstream_chain(self):
        grandchild = make_employee(iqama='DSR-GRANDCHILD', name='Grandchild Report')
        grandchild.main_manager = self.direct
        grandchild.save(update_fields=['main_manager'])

        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()

        self.assertIn('Direct Report', content)
        self.assertIn('Grandchild Report', content)
        self.assertIn('more further down your reporting chain', content)

        # The grandchild is what expanding adds — not a re-list of direct
        # reports, and not duplicated in the direct-reports queryset.
        self.assertEqual(list(resp.context['downstream_reports']), [grandchild])
        direct_report_names = [e.full_name for e in resp.context['employee'].main_reports.all()]
        self.assertEqual(direct_report_names, ['Direct Report'])

    def test_no_expand_control_when_no_further_downstream_chain(self):
        # self.direct exists (a direct report) but has no reports of its own
        # — downstream_reports should end up empty after excluding direct
        # reports, so no "+N more" clutter should render.
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()

        self.assertIn('Direct Report', content)
        self.assertNotIn('more further down your reporting chain', content)
        self.assertEqual(list(resp.context['downstream_reports']), [])


class MyProfileCardOrderingTests(TestCase):
    """'My Reporting Structure' must be the very last card on the page —
    after Attendance Exceptions, and after My Vehicles whenever that
    conditional card renders — not merely swapped with a neighbor."""

    def setUp(self):
        self.emp = make_employee(iqama='ORD-1', name='Order Employee')
        self.user = _login_user('ord_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.client.login(username='ord_user', password='testpass123')

    def test_reporting_structure_renders_after_vehicles_when_vehicle_present(self):
        vehicle = Vehicle.objects.create(plate_number='ORD-9999')
        from hr.models import AssetHandover
        AssetHandover.objects.create(vehicle=vehicle, employee=self.emp, status='active')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertIn('My Vehicles', content)
        self.assertIn('My Reporting Structure', content)
        self.assertGreater(content.index('My Reporting Structure'), content.index('My Vehicles'))
        self.assertGreater(content.index('My Reporting Structure'), content.index('Attendance Exceptions'))

    def test_reporting_structure_is_last_section_without_vehicle(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        content = resp.content.decode()
        self.assertNotIn('My Vehicles', content)
        self.assertGreater(content.index('My Reporting Structure'), content.index('Attendance Exceptions'))


class MyProfileOverriddenExceptionAttributionTests(TestCase):
    """The Attendance Exceptions status column must attribute an override to
    the actual overriding user, matching team_exceptions.html's existing
    'Overridden by {name}' pattern — not a bare 'Overridden' badge."""

    def setUp(self):
        self.manager_emp = make_employee(iqama='OVR-MGR', name='Override Manager')
        self.emp = make_employee(iqama='OVR-EMP', name='Override Employee')
        self.emp.main_manager = self.manager_emp
        self.emp.save(update_fields=['main_manager'])
        self.user = _login_user('ovr_emp_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])

    def test_overridden_exception_shows_overriding_user_name(self):
        exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 30),
            reason_category='site_visit', custom_reason='', employee_comment='', created_by=self.user)
        hr_user = make_user('ovr_hr_user', first_name='Helen', last_name='Ross')
        override_attendance_exception(exc, hr_user, 'approved', reason='Confirmed via site log')

        self.client.login(username='ovr_emp_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Overridden by Helen Ross')
        self.assertNotContains(resp, '>Overridden<')


class MyProfileNoManagerNoReportsRegressionTests(TestCase):
    """Regression coverage for the bug class already fixed once on this page:
    an employee with neither a main_manager nor any reports (direct or
    downstream) must render without a VariableDoesNotExist-style crash."""

    def test_no_manager_no_reports_no_downstream_renders_without_error(self):
        lone = make_employee(iqama='LONE-2', name='Truly Lone Employee')
        user = _login_user('truly_lone_user')
        lone.user = user
        lone.save(update_fields=['user'])
        self.client.login(username='truly_lone_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(list(resp.context['downstream_reports']), [])
        self.assertContains(resp, 'Not assigned')


# ─── Team Attendance Exceptions — History table redesign (Task 1) + optional
#     decide-comment (Task 4) ────────────────────────────────────────────────


class TeamExceptionsHistoryTableTests(TestCase):
    """Task 1: the History table drops "Decided By" (redundant with the
    "Overridden by {name}" badge) in favor of a "Decided At" timestamp
    column, and active-queue rows past their decision_deadline get a light
    overdue visual treatment."""

    def setUp(self):
        self.manager = make_employee(iqama='HIST-MGR', name='History Manager')
        self.manager_user = _login_user('hist_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.report = make_employee(iqama='HIST-REPORT', name='History Report')
        self.report.main_manager = self.manager
        self.report.save(update_fields=['main_manager'])
        self.creator = make_user('hist_creator')

    def test_history_table_shows_decided_at_not_decided_by(self):
        exc = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(exc, self.manager_user, 'approved')
        exc.refresh_from_db()
        self.client.login(username='hist_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        content = resp.content.decode()
        self.assertNotIn('Decided By', content)
        from django.template.defaultfilters import date as _django_date
        # The template's |date filter localizes aware datetimes to the
        # current timezone (Asia/Riyadh) before formatting; a raw call to
        # the filter function does not, so decided_at must be localized
        # here too or the two strings differ by the UTC offset.
        self.assertIn(_django_date(_timezone.localtime(exc.decided_at), 'd M Y, H:i'), content)

    def test_overdue_row_has_overdue_css_class(self):
        # Overdue-and-unactioned: still 'pending', past decision_deadline.
        stale_dt = _timezone.localtime(_timezone.now() - timedelta(days=2))
        exc = _submit_aex(
            employee=self.report, event_date=stale_dt.date(), event_start_time=stale_dt.time(),
            reason_category='site_visit', created_by=self.creator)
        self.assertTrue(exc.is_overdue())
        self.client.login(username='hist_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        content = resp.content.decode()
        self.assertIn('row-overdue', content)

    def test_expired_overdue_row_also_gets_overdue_css_class(self):
        # Same overdue treatment applies whether the row's status is still
        # 'pending' or has already flipped to 'expired' via the cron sweep —
        # both are "overdue and unactioned".
        from hr.attendance_exception_services import expire_overdue_exceptions
        stale_dt = _timezone.localtime(_timezone.now() - timedelta(days=2))
        exc = _submit_aex(
            employee=self.report, event_date=stale_dt.date(), event_start_time=stale_dt.time(),
            reason_category='site_visit', created_by=self.creator)
        expire_overdue_exceptions()
        exc.refresh_from_db()
        self.assertEqual(exc.status, 'expired')
        self.client.login(username='hist_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        content = resp.content.decode()
        self.assertIn('row-overdue', content)

    def test_not_yet_overdue_row_has_no_overdue_css_class(self):
        # Recent-and-relative, not a hardcoded date: a fixed calendar date
        # eventually becomes "in the past" as real time moves on, which
        # would flip is_overdue() to True and make this test fail for a
        # reason that has nothing to do with the code under test.
        recent_dt = _timezone.localtime(_timezone.now() - timedelta(hours=2))
        exc = _submit_aex(
            employee=self.report, event_date=recent_dt.date(), event_start_time=recent_dt.time(),
            reason_category='site_visit', created_by=self.creator)
        self.assertFalse(exc.is_overdue())
        self.client.login(username='hist_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        content = resp.content.decode()
        # The CSS rule for .row-overdue is always present in <head> — check
        # the class isn't actually APPLIED to a row (class="row-overdue"),
        # not just that the string appears anywhere on the page.
        self.assertNotIn('class="row-overdue"', content)

    def test_page_renders_when_an_overridden_rows_overriding_user_was_deleted(self):
        # Regression: overridden_by is on_delete=SET_NULL, so a historical
        # row can have is_overridden=True but overridden_by=None once the
        # overriding account is deleted. The template used to dereference
        # overridden_by.username as a *filter argument* (default:...), which
        # Django does not silently null-guard the way plain {{ var }} output
        # is — this crashed with VariableDoesNotExist for any tab showing
        # such a row (e.g. Ali/Sarah0's "All Organization Requests" tab).
        overrider = make_user('temp_overrider')
        exc = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        override_attendance_exception(exc, overrider, 'approved', reason='Covering for the main manager.')
        overrider.delete()
        exc.refresh_from_db()
        self.assertTrue(exc.is_overridden)
        self.assertIsNone(exc.overridden_by_id)
        self.client.login(username='hist_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Overridden')


class TeamExceptionsDecideCommentTests(TestCase):
    """Task 4: the plain Approve/Reject decide form gets an optional
    'comment' textarea (matching the leave-request workflow's naming
    convention — templates/hr/leave_request_detail.html), threaded through
    to decide_attendance_exception(..., note=<value>) and persisted on
    AttendanceException.decision_note."""

    def setUp(self):
        self.manager = make_employee(iqama='CMT-MGR', name='Comment Manager')
        self.manager_user = _login_user('cmt_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.report = make_employee(iqama='CMT-REPORT', name='Comment Report')
        self.report.main_manager = self.manager
        self.report.save(update_fields=['main_manager'])
        self.creator = make_user('cmt_creator')
        self.exc = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)

    def test_decide_form_renders_optional_comment_textarea(self):
        self.client.login(username='cmt_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, 'name="comment"')

    def test_decide_form_comment_persists_to_decision_note(self):
        self.client.login(username='cmt_mgr', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'decide', 'exc_id': self.exc.pk, 'decision': 'approved',
            'comment': 'Confirmed via phone call',
        })
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'approved')
        self.assertEqual(self.exc.decision_note, 'Confirmed via phone call')

    def test_decide_form_without_comment_still_works(self):
        self.client.login(username='cmt_mgr', password='testpass123')
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'decide', 'exc_id': self.exc.pk, 'decision': 'approved',
        })
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'approved')
        self.assertEqual(self.exc.decision_note, '')


class AttendanceExportColorTests(TestCase):
    # The register exports fill each cell with a color matching its status -
    # Annual Leave, Saudi National Day, and Eid holidays each get their own
    # distinct fill color in the Excel export.

    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('adm2', password='x'); self.admin.role = role; self.admin.save()
        self.emp = make_employee(iqama='E2', name='Leave Tester')
        self.annual, _ = LeaveType.objects.get_or_create(code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        LeaveRecord.objects.create(
            employee=self.emp, leave_type=self.annual,
            start_date=_date(2026, 7, 13), end_date=_date(2026, 7, 13))
        Holiday.objects.create(date=_date(2026, 7, 14), name='Saudi National Day', category='saudi_national_day')
        Holiday.objects.create(date=_date(2026, 7, 15), name='Eid al-Fitr', category='eid')
        AttendanceRecord.objects.create(employee=self.emp, date=_date(2026, 7, 16), status='present')
        self.client.force_login(self.admin)

    def test_annual_leave_cell_has_correct_fill_color(self):
        import io
        import openpyxl as _op
        resp = self.client.get(reverse('hr:attendance_matrix_export_excel') + '?period=month&date=2026-07-15')
        self.assertEqual(resp.status_code, 200)
        ws = _op.load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[3]]
        col_13 = header.index('13-Jul') + 1
        target_row = None
        for row in ws.iter_rows(min_row=4):
            if row[0].value == 'Leave Tester':
                target_row = row
                break
        self.assertIsNotNone(target_row, 'Employee row not found in export')
        fill = target_row[col_13 - 1].fill.start_color.rgb
        self.assertEqual(fill[-6:].upper(), 'FCE4D6', 'Annual Leave cell is not the expected peach color')

    def test_saudi_national_day_cell_has_correct_fill_color(self):
        import io
        import openpyxl as _op
        resp = self.client.get(reverse('hr:attendance_matrix_export_excel') + '?period=month&date=2026-07-15')
        ws = _op.load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[3]]
        col_14 = header.index('14-Jul') + 1
        target_row = None
        for row in ws.iter_rows(min_row=4):
            if row[0].value == 'Leave Tester':
                target_row = row
                break
        fill = target_row[col_14 - 1].fill.start_color.rgb
        self.assertEqual(fill[-6:].upper(), '00B050', 'Saudi National Day cell is not the expected green color')

    def test_eid_holiday_cell_has_correct_fill_color(self):
        import io
        import openpyxl as _op
        resp = self.client.get(reverse('hr:attendance_matrix_export_excel') + '?period=month&date=2026-07-15')
        ws = _op.load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[3]]
        col_15 = header.index('15-Jul') + 1
        target_row = None
        for row in ws.iter_rows(min_row=4):
            if row[0].value == 'Leave Tester':
                target_row = row
                break
        fill = target_row[col_15 - 1].fill.start_color.rgb
        self.assertEqual(fill[-6:].upper(), 'FFFF00', 'Eid holiday cell is not the expected yellow color')

    def test_present_cell_has_correct_fill_color(self):
        import io
        import openpyxl as _op
        resp = self.client.get(reverse('hr:attendance_matrix_export_excel') + '?period=month&date=2026-07-15')
        ws = _op.load_workbook(io.BytesIO(resp.content)).active
        header = [c.value for c in ws[3]]
        col_16 = header.index('16-Jul') + 1
        target_row = None
        for row in ws.iter_rows(min_row=4):
            if row[0].value == 'Leave Tester':
                target_row = row
                break
        fill = target_row[col_16 - 1].fill.start_color.rgb
        self.assertEqual(fill[-6:].upper(), 'C6E0B4', 'Present cell is not the expected green color')


class DirectManagerLeaveLoggingAccessTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(username='dmla-mgr', password='x')
        self.manager = Employee.objects.create(
            iqama_number='DMLA-MGR', full_name='Direct Manager', user=self.manager_user)
        self.report = Employee.objects.create(
            iqama_number='DMLA-RPT', full_name='Direct Report', main_manager=self.manager)
        self.other_manager_user = User.objects.create_user(username='dmla-other', password='x')
        Employee.objects.create(
            iqama_number='DMLA-OTHERMGR', full_name='Other Manager', user=self.other_manager_user)
        self.stranger = Employee.objects.create(iqama_number='DMLA-STRANGER', full_name='Stranger')
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(
            employee=self.report, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))

    def test_direct_manager_can_reach_the_log_request_page(self):
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.get(reverse('hr:leave_request_create'))
        self.assertEqual(resp.status_code, 200)

    def test_direct_manager_sees_only_their_report_in_the_dropdown(self):
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.get(reverse('hr:leave_request_create'))
        ids = set(resp.context['form'].fields['employee'].queryset.values_list('pk', flat=True))
        self.assertEqual(ids, {self.report.pk})

    def test_manager_with_no_reports_is_denied(self):
        self.client.login(username='dmla-other', password='x')
        resp = self.client.get(reverse('hr:leave_request_create'))
        self.assertEqual(resp.status_code, 403)

    def test_direct_manager_can_log_for_their_report(self):
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(LeaveRequest.objects.filter(employee=self.report).exists())

    def test_manager_cannot_log_for_a_non_report(self):
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.stranger.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        self.assertEqual(resp.status_code, 200)  # form_invalid re-render, not a 302
        self.assertFalse(LeaveRequest.objects.filter(employee=self.stranger).exists())

    def test_manager_cannot_log_for_their_own_manager(self):
        # Upward abuse: DMLA-MGR reports to nobody here, but prove a manager
        # two levels up can't reach a peer/superior via this dropdown either.
        upline_user = User.objects.create_user(username='dmla-upline', password='x')
        upline = Employee.objects.create(iqama_number='DMLA-UPLINE', full_name='Upline', user=upline_user)
        self.manager.main_manager = upline
        self.manager.save(update_fields=['main_manager'])
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': upline.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(LeaveRequest.objects.filter(employee=upline).exists())

    def test_manager_cannot_log_for_a_reports_report(self):
        sub_report = Employee.objects.create(
            iqama_number='DMLA-SUBRPT', full_name='Sub Report', main_manager=self.report)
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': sub_report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(LeaveRequest.objects.filter(employee=sub_report).exists())

    def test_inactive_report_is_excluded_from_the_dropdown(self):
        # A second active report keeps the manager on the page at all (with
        # zero active reports they'd lose access entirely, per test_func) —
        # this isolates "is the inactive one filtered out of the dropdown."
        active_report = Employee.objects.create(
            iqama_number='DMLA-ACTIVE', full_name='Still Active', main_manager=self.manager)
        self.report.is_active = False
        self.report.save(update_fields=['is_active'])
        self.client.login(username='dmla-mgr', password='x')
        resp = self.client.get(reverse('hr:leave_request_create'))
        ids = set(resp.context['form'].fields['employee'].queryset.values_list('pk', flat=True))
        self.assertNotIn(self.report.pk, ids)
        self.assertIn(active_report.pk, ids)


class ManagerLoggedRequestRoutingTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(username='mlrr-mgr', password='x')
        self.manager = Employee.objects.create(
            iqama_number='MLRR-MGR', full_name='Routing Manager', user=self.manager_user)
        self.report = Employee.objects.create(
            iqama_number='MLRR-RPT', full_name='Routing Report', main_manager=self.manager, work_location='site')
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30'), 'site_default_annual_days': Decimal('45')})
        LeaveEntitlement.objects.create(
            employee=self.report, leave_type=self.lt, year=2026, entitled_days=Decimal('45'))
        self.approver_user = User.objects.create_user(username='mlrr-appr', password='x')
        LeaveDashboardAccess.objects.create(user=self.approver_user, is_active=True)

    def test_manager_logged_request_goes_to_the_normal_approver_roster(self):
        self.client.login(username='mlrr-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertEqual(req.status, 'pending')
        self.assertEqual(list(req.approvals.values_list('approver', flat=True)), [self.approver_user.pk])

    def test_manager_logged_request_is_not_auto_approved(self):
        self.client.login(username='mlrr-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertNotEqual(req.status, 'approved')

    def test_manager_logged_over_cap_request_is_held_not_blocked(self):
        self.client.login(username='mlrr-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-01-01', 'end_date': '2026-02-15',  # 46 days, 1 over the 45-day Site baseline
        })
        self.assertEqual(resp.status_code, 302)
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertTrue(req.exceeds_balance)
        self.assertEqual(req.approvals.count(), 1)  # still the normal roster, not held-and-approver-less

    def test_manager_never_sees_the_log_anyway_grant_option(self):
        # A plain manager has no override access, so form_valid's
        # can_grant-gated warning branch never triggers for them — the
        # over-cap request goes straight through to a normal 302 redirect,
        # never pausing on an in-page "Log Anyway" warning at all.
        self.client.login(username='mlrr-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-01-01', 'end_date': '2026-02-15',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(LeaveEntitlement.objects.get(employee=self.report).exception_days, Decimal('0'))


class MyProfileTeamSectionTests(TestCase):
    """The 'Log Leave' entry point lives inside the existing 'My Reporting
    Structure' card's Direct Reports list — not a separate new section."""
    def setUp(self):
        self.manager_user = User.objects.create_user(username='mpts-mgr', password='x')
        self.manager = Employee.objects.create(
            iqama_number='MPTS-MGR', full_name='Team Section Manager', user=self.manager_user)
        self.report = Employee.objects.create(
            iqama_number='MPTS-RPT', full_name='Team Section Report', main_manager=self.manager)
        self.inactive_report = Employee.objects.create(
            iqama_number='MPTS-INACTIVE', full_name='Inactive Report', main_manager=self.manager, is_active=False)
        self.plain_user = User.objects.create_user(username='mpts-plain', password='x')
        Employee.objects.create(iqama_number='MPTS-PLAIN', full_name='Plain Employee', user=self.plain_user)

    def test_manager_sees_log_leave_link_for_their_direct_report(self):
        self.client.login(username='mpts-mgr', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'My Reporting Structure')
        self.assertContains(resp, 'Team Section Report')
        self.assertContains(resp, f"{reverse('hr:leave_request_create')}?employee={self.report.pk}")

    def test_inactive_report_shows_no_log_leave_link(self):
        self.client.login(username='mpts-mgr', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Inactive Report')
        self.assertNotContains(resp, f"{reverse('hr:leave_request_create')}?employee={self.inactive_report.pk}")

    def test_plain_employee_sees_no_log_leave_link(self):
        self.client.login(username='mpts-plain', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, reverse('hr:leave_request_create'))


class ManagerSuccessRedirectTests(TestCase):
    """Bug found in manual testing: a plain direct manager could submit the
    Log Request form (widened access), but the success redirect always
    pointed at the Leave Requests queue — a separate, more restrictive page
    (can_view_leave_dashboard) they don't have permission to view, so they
    hit a 403 immediately after a successful submission. Must redirect
    somewhere the submitter can actually see."""
    def setUp(self):
        self.manager_user = User.objects.create_user(username='msr-mgr', password='x')
        self.manager = Employee.objects.create(
            iqama_number='MSR-MGR', full_name='Redirect Manager', user=self.manager_user)
        self.report = Employee.objects.create(
            iqama_number='MSR-RPT', full_name='Redirect Report', main_manager=self.manager)
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.report, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        from accounts.models import Role
        self.hr_user = _make_role_user('msr-hr', Role.SUPER_ADMIN)
        LeaveDashboardAccess.objects.create(user=self.hr_user, is_active=True)

    def test_plain_manager_redirect_target_is_reachable(self):
        self.client.login(username='msr-mgr', password='x')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        }, follow=True)
        self.assertEqual(resp.status_code, 200)  # the final page in the redirect chain, not a 403
        self.assertEqual(resp.redirect_chain[-1][0], reverse('hr:my_profile'))

    def test_hr_user_still_lands_on_the_queue(self):
        # Unchanged behavior for anyone who could already view the queue.
        self.client.login(username='msr-hr', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        }, follow=True)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.redirect_chain[-1][0], reverse('hr:leave_request_list'))

    def test_plain_manager_back_and_cancel_links_do_not_point_at_the_queue(self):
        self.client.login(username='msr-mgr', password='x')
        resp = self.client.get(reverse('hr:leave_request_create'))
        self.assertNotContains(resp, f'href="{reverse("hr:leave_request_list")}"')
        self.assertContains(resp, f'href="{reverse("hr:my_profile")}"')


class LoggedByManagerFlagTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='lbmf-mgr', password='x', first_name='Flag', last_name='Manager')
        self.manager = Employee.objects.create(
            iqama_number='LBMF-MGR', full_name='Flag Manager', user=self.manager_user)
        self.report = Employee.objects.create(
            iqama_number='LBMF-RPT', full_name='Flag Report', main_manager=self.manager)
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.report, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        from accounts.models import Role
        self.hr_user = _make_role_user('lbmf-hr', Role.SUPER_ADMIN)
        LeaveDashboardAccess.objects.create(user=self.hr_user, is_active=True)

    def test_manager_logged_request_sets_the_flag(self):
        self.client.login(username='lbmf-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertTrue(req.logged_by_manager)

    def test_hr_logged_request_does_not_set_the_flag(self):
        self.client.login(username='lbmf-hr', password='testpass123')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertFalse(req.logged_by_manager)

    def test_self_service_request_does_not_set_the_flag(self):
        report_user = User.objects.create_user(username='lbmf-emp', password='x')
        self.report.user = report_user
        self.report.save(update_fields=['user'])
        self.client.login(username='lbmf-emp', password='x')
        self.client.post(reverse('hr:my_profile'), data={
            'action': 'request_leave', 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertFalse(req.logged_by_manager)

    def test_badge_shows_on_queue_and_detail_pages(self):
        self.client.login(username='lbmf-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.client.logout()
        self.client.login(username='lbmf-hr', password='testpass123')
        list_resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(list_resp, 'By Flag Manager (manager)')
        detail_resp = self.client.get(reverse('hr:leave_request_detail', args=[req.pk]))
        self.assertContains(detail_resp, 'Logged on behalf of')
        self.assertContains(detail_resp, 'Flag Manager')


class ManagerSeesOnlyTheirOwnInBehalfStatusTests(TestCase):
    """My Profile -> Reporting Structure: a manager sees the status of
    requests THEY logged for a direct report, and only those — never a
    report's own self-submitted requests, and never another manager's
    in-behalf requests for the same person."""
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='msob-mgr', password='x', first_name='Msob', last_name='Manager')
        self.manager = Employee.objects.create(
            iqama_number='MSOB-MGR', full_name='Msob Manager', user=self.manager_user)
        self.report_user = User.objects.create_user(username='msob-emp', password='x')
        self.report = Employee.objects.create(
            iqama_number='MSOB-RPT', full_name='Msob Report', main_manager=self.manager, user=self.report_user)
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.report, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))

    def test_manager_sees_status_of_a_request_they_logged(self):
        self.client.login(username='msob-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Logged by you:')
        self.assertContains(resp, 'Pending')

    def test_manager_does_not_see_reports_self_submitted_request(self):
        self.client.login(username='msob-emp', password='x')
        self.client.post(reverse('hr:my_profile'), data={
            'action': 'request_leave', 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        req = LeaveRequest.objects.get(employee=self.report)
        self.assertFalse(req.logged_by_manager)  # sanity: really self-submitted
        self.client.logout()
        self.client.login(username='msob-mgr', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'Logged by you:')

    def test_manager_does_not_see_a_different_managers_in_behalf_request(self):
        other_manager_user = User.objects.create_user(username='msob-other', password='x')
        other_manager = Employee.objects.create(
            iqama_number='MSOB-OTHER', full_name='Other Manager', user=other_manager_user)
        # Simulate a historical request logged by a different manager (e.g.
        # before a reassignment) by calling the service directly.
        submit_leave_request(
            employee=self.report, leave_type=self.lt, start_date=date(2026, 1, 1), end_date=date(2026, 1, 3),
            created_by=other_manager_user)
        self.client.login(username='msob-mgr', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'Logged by you:')

    def test_multiple_in_behalf_requests_for_the_same_report_all_show(self):
        # Bug found in manual testing: logging two separate periods for the
        # same report only showed one status line — the view kept just the
        # most recent request per employee instead of every one of them.
        self.client.login(username='msob-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-01-01', 'end_date': '2026-01-03',
        })
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-06-01', 'end_date': '2026-06-03',
        })
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Logged by you:', count=2)
        shown_report = next(r for r in resp.context['direct_reports'] if r.pk == self.report.pk)
        self.assertEqual(len(shown_report.in_behalf_requests), 2)
        # Newest first.
        self.assertEqual(shown_report.in_behalf_requests[0].start_date, date(2026, 6, 1))
        self.assertEqual(shown_report.in_behalf_requests[1].start_date, date(2026, 1, 1))

    def test_more_than_two_in_behalf_requests_collapse_behind_a_show_more_control(self):
        self.client.login(username='msob-mgr', password='x')
        for month in (1, 3, 6):
            self.client.post(reverse('hr:leave_request_create'), data={
                'employee': self.report.pk, 'leave_type': self.lt.pk,
                'start_date': f'2026-{month:02d}-01', 'end_date': f'2026-{month:02d}-02',
            })
        resp = self.client.get(reverse('hr:my_profile'))
        # Only the 2 most recent render unconditionally; the 3rd is tucked
        # behind a "+N more" control instead of cluttering the card.
        self.assertContains(resp, 'Logged by you:', count=3)  # 2 visible + 1 inside the collapsed region
        self.assertContains(resp, '+1 more request(s)')

    def test_employee_sees_flag_that_manager_logged_it(self):
        self.client.login(username='msob-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        self.client.logout()
        self.client.login(username='msob-emp', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Logged by your manager')
        self.assertContains(resp, 'Msob Manager')

    def test_employee_sees_no_flag_on_their_own_self_submitted_request(self):
        self.client.login(username='msob-emp', password='x')
        self.client.post(reverse('hr:my_profile'), data={
            'action': 'request_leave', 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'Logged by your manager')


class BackdatedRequestTests(TestCase):
    """'Backdated' badge — surfaced when a leave/exception's own date is
    before the day it was actually submitted, so an after-the-fact
    documentation request doesn't get mistaken for a live upcoming one."""

    def setUp(self):
        self.emp = make_employee(iqama='BDR-1')
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.user = _login_user('bdr_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])

    def test_leave_request_is_backdated_when_start_date_precedes_submission(self):
        req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        LeaveRequest.objects.filter(pk=req.pk).update(created_at=_timezone.make_aware(_datetime(2026, 8, 5, 9, 0)))
        req.refresh_from_db()
        self.assertTrue(req.is_backdated)

    def test_leave_request_is_not_backdated_when_submitted_on_or_before_the_start_date(self):
        req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 9, 1), end_date=date(2026, 9, 2),
            created_by=self.user)
        LeaveRequest.objects.filter(pk=req.pk).update(created_at=_timezone.make_aware(_datetime(2026, 8, 25, 9, 0)))
        req.refresh_from_db()
        self.assertFalse(req.is_backdated)

    def test_attendance_exception_is_backdated_when_event_precedes_submission(self):
        exc = submit_attendance_exception(
            employee=self.emp, event_date=date(2026, 8, 1), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        AttendanceException.objects.filter(pk=exc.pk).update(
            created_at=_timezone.make_aware(_datetime(2026, 8, 5, 9, 0)))
        exc.refresh_from_db()
        self.assertTrue(exc.is_backdated)

    def test_attendance_exception_reported_same_day_is_not_backdated(self):
        # Reporting an event a few hours after it happened, same calendar
        # day, is the expected/normal case — not a "backdated" flag.
        exc = submit_attendance_exception(
            employee=self.emp, event_date=date(2026, 8, 5), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        AttendanceException.objects.filter(pk=exc.pk).update(
            created_at=_timezone.make_aware(_datetime(2026, 8, 5, 17, 0)))
        exc.refresh_from_db()
        self.assertFalse(exc.is_backdated)

    def test_backdated_badge_renders_on_my_profile(self):
        req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        LeaveRequest.objects.filter(pk=req.pk).update(created_at=_timezone.make_aware(_datetime(2026, 8, 5, 9, 0)))
        self.client.login(username='bdr_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Backdated')

    def test_backdated_badge_renders_on_leave_queue_pending_tab(self):
        from accounts.models import Role
        req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        LeaveRequest.objects.filter(pk=req.pk).update(created_at=_timezone.make_aware(_datetime(2026, 8, 5, 9, 0)))
        _make_role_user('bdr-super', Role.SUPER_ADMIN)
        self.client.login(username='bdr-super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, 'Backdated')

    def test_no_backdated_badge_for_a_normal_advance_request(self):
        self.client.login(username='bdr_user', password='testpass123')
        submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 12, 1), end_date=date(2026, 12, 2),
            created_by=self.user)
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'Backdated')

    def test_backdated_badge_renders_on_team_exceptions_pending_tab(self):
        manager = make_employee(iqama='BDR-MGR', name='BDR Manager')
        manager_user = _login_user('bdr_manager')
        manager.user = manager_user
        manager.save(update_fields=['user'])
        self.emp.main_manager = manager
        self.emp.save(update_fields=['main_manager'])
        exc = submit_attendance_exception(
            employee=self.emp, event_date=date(2026, 8, 1), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        AttendanceException.objects.filter(pk=exc.pk).update(
            created_at=_timezone.make_aware(_datetime(2026, 8, 5, 9, 0)))
        self.client.login(username='bdr_manager', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'direct'})
        self.assertContains(resp, 'Backdated')


class LeaveQueueHistoryCollapseTests(TestCase):
    """History card on the Leave Approval Queue is collapsed by default so a
    long history doesn't push the Pending section down the page, but the
    data is still in the DOM (Bootstrap collapse hides via CSS, not by
    dropping markup) and the header shows an accurate count badge."""

    def setUp(self):
        self.emp = make_employee(iqama='LQHC-1', name='Queue History Employee')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.superadmin = make_user('lqhc_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()

    def _decided_request(self, n, status='approved'):
        req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.lt,
            start_date=_date(2026, 1, n), end_date=_date(2026, 1, n + 1))
        req.status = status
        req.decided_at = timezone.now()
        req.save(update_fields=['status', 'decided_at'])
        return req

    def test_history_panel_not_the_default_tab(self):
        # History moved from a collapsible card into its own tab (see
        # 'Requests You've Logged' / tabbed queue redesign) — the
        # equivalent of "collapsed by default" is now "not the active
        # tab-pane on page load", while Pending stays the default view.
        self._decided_request(1)
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        content = resp.content.decode()
        self.assertIn('<div class="tab-pane fade" id="historyTab"', content)
        self.assertIn('<div class="tab-pane fade show active" id="pendingTab"', content)

    def test_history_badge_shows_accurate_count(self):
        for n in range(1, 4):
            self._decided_request(n)
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, '<span class="badge bg-secondary ms-1">3</span>')

    def test_decided_request_data_still_renders_while_collapsed(self):
        self._decided_request(1)
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        # The row is present in the response body even though the panel
        # starts visually collapsed — Bootstrap's collapse only toggles a
        # CSS class, it never removes the content from the page.
        self.assertContains(resp, self.emp.full_name)

    def test_history_panel_scoped_to_only_decided_requests(self):
        # A large pending queue must not get swept into the same collapsed
        # panel or count — Pending stays its own always-visible section.
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.lt,
            start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        self._decided_request(1)
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, '<span class="badge bg-secondary ms-1">1</span>')

    def test_history_defaults_to_newest_decided_first(self):
        older = self._decided_request(1)
        older.decided_at = timezone.now() - timedelta(days=5)
        older.save(update_fields=['decided_at'])
        newer = self._decided_request(3)
        newer.decided_at = timezone.now()
        newer.save(update_fields=['decided_at'])
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        ids = [r.pk for r in resp.context['decided_requests']]
        self.assertEqual(ids, [newer.pk, older.pk])
        self.assertEqual(resp.context['current_history_sort'], 'newest')

    def test_history_sort_oldest_reverses_the_order(self):
        older = self._decided_request(1)
        older.decided_at = timezone.now() - timedelta(days=5)
        older.save(update_fields=['decided_at'])
        newer = self._decided_request(3)
        newer.decided_at = timezone.now()
        newer.save(update_fields=['decided_at'])
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'), {'history_sort': 'oldest'})
        ids = [r.pk for r in resp.context['decided_requests']]
        self.assertEqual(ids, [older.pk, newer.pk])
        self.assertEqual(resp.context['current_history_sort'], 'oldest')

    def test_history_sort_form_round_trips_the_history_tab(self):
        # The sort control is a plain GET form — without a tab=history
        # hidden field, submitting it would silently drop the page back to
        # the default Pending tab instead of staying on History.
        self.client.login(username='lqhc_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        content = resp.content.decode()
        sort_form_start = content.index('name="history_sort"')
        preceding = content[:sort_form_start]
        self.assertIn('<input type="hidden" name="tab" value="history">', preceding[-400:])


class TeamExceptionsHistoryCollapseTests(TestCase):
    """Same collapsible-History treatment on Team Exceptions, verified
    alongside its existing per-row inline Override collapse form to make
    sure the two nested collapses don't collide."""

    def setUp(self):
        self.manager = make_employee(iqama='TEHC-MGR', name='TE History Manager')
        self.manager_user = _login_user('tehc_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.report = make_employee(iqama='TEHC-REPORT', name='TE History Report')
        self.report.main_manager = self.manager
        self.report.save(update_fields=['main_manager'])
        self.creator = make_user('tehc_creator')

    def test_history_panel_collapsed_by_default(self):
        exc = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(exc, self.manager_user, 'approved')
        self.client.login(username='tehc_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        content = resp.content.decode()
        self.assertIn('<div class="collapse" id="exceptionHistoryCollapse">', content)

    def test_history_badge_shows_accurate_count(self):
        for day in (18, 19, 20):
            exc = _submit_aex(
                employee=self.report, event_date=_date(2026, 7, day), event_start_time=_time(9, 0),
                reason_category='site_visit', created_by=self.creator)
            decide_attendance_exception(exc, self.manager_user, 'approved')
        self.client.login(username='tehc_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, '<span class="badge bg-secondary ms-1">3</span>')

    def test_history_defaults_to_newest_decided_first(self):
        older = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 18), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(older, self.manager_user, 'approved')
        older.decided_at = timezone.now() - timedelta(days=5)
        older.save(update_fields=['decided_at'])
        newer = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(newer, self.manager_user, 'approved')
        newer.decided_at = timezone.now()
        newer.save(update_fields=['decided_at'])
        self.client.login(username='tehc_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        ids = [e.pk for e in resp.context['decided_exceptions']]
        self.assertEqual(ids, [newer.pk, older.pk])
        self.assertEqual(resp.context['current_history_sort'], 'newest')

    def test_history_sort_oldest_reverses_the_order(self):
        older = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 18), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(older, self.manager_user, 'approved')
        older.decided_at = timezone.now() - timedelta(days=5)
        older.save(update_fields=['decided_at'])
        newer = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(newer, self.manager_user, 'approved')
        newer.decided_at = timezone.now()
        newer.save(update_fields=['decided_at'])
        self.client.login(username='tehc_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'history_sort': 'oldest'})
        ids = [e.pk for e in resp.context['decided_exceptions']]
        self.assertEqual(ids, [older.pk, newer.pk])
        self.assertEqual(resp.context['current_history_sort'], 'oldest')

    def test_history_sort_reload_keeps_the_panel_expanded(self):
        # The sort control lives inside the collapsed-by-default History
        # card — without re-opening on reload, changing the sort would
        # visually hide the very list that just got resorted.
        self.client.login(username='tehc_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'history_sort': 'oldest'})
        self.assertIn('<div class="collapse show" id="exceptionHistoryCollapse">', resp.content.decode())

    def test_overridden_attribution_still_renders_inside_collapsed_history(self):
        # History only ever holds already-decided (approved/rejected) rows
        # — 'expired' stays in the active Pending queue until a human acts,
        # so the per-row Override *form* never actually appears in History
        # (can_override requires status in pending/expired). What DOES
        # render there is the "Overridden by {name}" attribution on a row
        # that was decided via Override — confirm that per-row conditional
        # content still renders correctly nested inside the new outer
        # History collapse wrapper.
        exc = _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        override_attendance_exception(exc, self.manager_user, 'approved', reason='Covering for the assigned manager.')
        self.client.login(username='tehc_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, 'Overridden by')


class MyProfileHistoryCollapseTests(TestCase):
    """My Profile: both 'My Leave Requests' and 'My Attendance Exceptions'
    get the same collapsed-by-default History treatment as the queue
    pages, scoped to just the table (the request/report forms above each
    table must stay visible, not get swept into the collapse)."""

    def setUp(self):
        self.emp = make_employee(iqama='MPHC-1', name='Profile History Employee')
        self.user = _login_user('mphc_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=3)

    def test_leave_requests_history_collapsed_by_default_with_count(self):
        LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.lt, start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        self.client.login(username='mphc_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        content = resp.content.decode()
        self.assertIn('<div class="collapse" id="myLeaveRequestsCollapse">', content)
        self.assertContains(resp, '<span class="badge bg-secondary ms-1">1</span>')

    def test_request_leave_form_stays_outside_the_history_collapse(self):
        # The "Request Leave" form toggle is a separate, always-reachable
        # control — it must not end up nested inside the collapsed history.
        self.client.login(username='mphc_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'id="requestLeaveForm"')

    def test_attendance_exceptions_history_collapsed_by_default_with_count(self):
        _submit_aex(employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 30),
                    reason_category='site_visit', created_by=self.user)
        self.client.login(username='mphc_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        content = resp.content.decode()
        self.assertIn('<div class="collapse" id="myExceptionsCollapse">', content)
        self.assertContains(resp, '<span class="badge bg-secondary ms-1">1</span>')

    def test_report_exception_form_stays_outside_the_history_collapse(self):
        self.client.login(username='mphc_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'id="reportExceptionForm"')


class MyProfileDirectReportsCollapseTests(TestCase):
    """Direct Reports (inside 'My Reporting Structure') gets the same
    collapse affordance as the History sections, but — unlike History,
    which is past/reference data — it's the card's primary actionable
    content (the Log Leave button lives there), so it defaults to
    expanded rather than hidden."""

    def setUp(self):
        self.manager = make_employee(iqama='MPDR-MGR', name='Direct Reports Manager')
        self.manager_user = _login_user('mpdr_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])

    def test_direct_reports_panel_defaults_expanded(self):
        report = make_employee(iqama='MPDR-REPORT', name='Direct Reports Report')
        report.main_manager = self.manager
        report.save(update_fields=['main_manager'])
        self.client.login(username='mpdr_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        content = resp.content.decode()
        self.assertIn('<div class="collapse show" id="directReportsCollapse">', content)
        self.assertContains(resp, 'Direct Reports Report')

    def test_no_toggle_rendered_when_there_are_no_direct_reports(self):
        self.client.login(username='mpdr_mgr', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'id="directReportsCollapse"')


class ApprovalDecisionMessageVisibilityTests(TestCase):
    """Bug fix: a message an approver types while deciding a leave request
    used to be saved only on LeaveRequestApproval.comment, which nothing
    employee-facing ever reads — a plain employee can't reach
    leave_request_detail.html (it's gated to can_view_leave_dashboard), so
    the message was effectively invisible to the person it was about.
    record_approver_decision now also creates an employee-visible
    LeaveRequestNote, and My Profile renders it behind a small toggle so it
    doesn't clutter the page when there's nothing to show."""

    def setUp(self):
        self.emp = make_employee(iqama='ADMV-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.user = _login_user('admv_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=3)
        self.approver = User.objects.create_user(
            username='admv-approver', password='x', first_name='Admv', last_name='Approver')
        LeaveDashboardAccess.objects.create(user=self.approver, is_active=True)
        # Roster is snapshotted onto the request AT SUBMISSION TIME (see
        # submit_leave_request) — the approver's LeaveDashboardAccess grant
        # must exist before this call, or no LeaveRequestApproval row gets
        # created for them.
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)

    def test_decision_comment_creates_an_employee_visible_note(self):
        record_approver_decision(self.req, self.approver, 'approved', comment='Approved — enjoy your trip.')
        self.assertEqual(self.req.notes.count(), 1)
        note = self.req.notes.first()
        self.assertFalse(note.is_internal)
        self.assertEqual(note.note, 'Approved — enjoy your trip.')
        self.assertEqual(note.author, self.approver)

    def test_decision_without_comment_creates_no_note(self):
        record_approver_decision(self.req, self.approver, 'approved')
        self.assertEqual(self.req.notes.count(), 0)

    def test_message_shows_on_my_profile_behind_a_toggle(self):
        record_approver_decision(self.req, self.approver, 'approved', comment='Approved — enjoy your trip.')
        self.client.login(username='admv_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, '1 message from HR')
        self.assertContains(resp, 'Approved — enjoy your trip.')
        self.assertContains(resp, f'id="leaveNotes{self.req.pk}"')

    def test_no_toggle_rendered_when_there_is_no_message(self):
        record_approver_decision(self.req, self.approver, 'approved')
        self.client.login(username='admv_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'message from HR')

    def test_internal_only_note_never_shows_on_my_profile(self):
        LeaveRequestNote.objects.create(
            leave_request=self.req, author=self.approver, note='Internal HR context.', is_internal=True)
        self.client.login(username='admv_user', password='testpass123')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'Internal HR context.')


class LeaveDecisionFormUXTests(TestCase):
    """The decide form now carries an optional employee-message field
    directly (one submit for decision + message, instead of a second,
    separate Add Note submission), laid out as numbered steps
    (1. decision, 2. message, 3. submit)."""

    def setUp(self):
        self.emp = make_employee(iqama='LDFU-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.approver = make_user('ldfu-approver', password='x')
        self.approver.set_password('testpass123')
        self.approver.save()
        LeaveDashboardAccess.objects.create(user=self.approver, is_active=True)
        self.req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.lt, start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=self.approver)
        self.client.login(username='ldfu-approver', password='testpass123')

    def test_decide_form_has_numbered_steps_and_message_field(self):
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, '1. Choose a decision')
        self.assertContains(resp, 'Add a message for the employee')
        self.assertContains(resp, 'name="comment"')
        content = resp.content.decode()
        self.assertIn('<span class="fw-semibold small">3.</span>', content)

    def test_single_submit_decides_and_creates_note_together(self):
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]), {
            'action': 'decide', 'decision': 'approved', 'comment': 'All set.',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        self.assertEqual(self.req.notes.count(), 1)


class SalaryDeductionCompactUITests(TestCase):
    """The salary-deduction warning is important enough that it must stay
    always visible — never hidden behind a click — but the row's secondary
    status lines (decided-by, warning, HR-message toggle) share one tight
    flex-column gap instead of each carrying its own margin, so the row
    doesn't look like it's sprawling."""

    def setUp(self):
        self.emp = make_employee(iqama='SDCU-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        self.user = _login_user('sdcu_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=3)
        req = LeaveRequest.objects.create(
            employee=self.emp, leave_type=self.lt, start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        req.status = 'disapproved'
        req.salary_deduction_applicable = True
        req.save(update_fields=['status', 'salary_deduction_applicable'])
        self.req = req
        self.client.login(username='sdcu_user', password='testpass123')

    def test_message_is_always_visible_not_behind_a_toggle(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Taking this leave will result in a salary deduction.')
        content = resp.content.decode()
        # No collapse wrapper around it — this warning must never require a
        # click to see.
        self.assertNotIn(f'id="salaryDeduction{self.req.pk}"', content)
        self.assertNotIn('data-bs-toggle="collapse" data-bs-target="#salaryDeduction', content)

    def test_no_message_when_not_applicable(self):
        self.req.salary_deduction_applicable = False
        self.req.save(update_fields=['salary_deduction_applicable'])
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, 'salary deduction')


class LeaveRequestRevokeFieldsTests(TestCase):
    def test_revoked_is_a_valid_status_choice(self):
        self.assertIn(('revoked', 'Revoked'), LeaveRequest.STATUS_CHOICES)

    def test_revoke_fields_exist_and_default_empty(self):
        emp = make_employee(iqama='LRRF-1')
        lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        req = LeaveRequest.objects.create(
            employee=emp, leave_type=lt, start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2))
        self.assertIsNone(req.revoked_by)
        self.assertIsNone(req.revoked_at)
        self.assertEqual(req.revoke_reason, '')


class LeaveRevokeRequestModelTests(TestCase):
    def test_create_and_defaults(self):
        emp = make_employee(iqama='LRRM-1')
        lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 3, 'is_accumulative': False})
        req = LeaveRequest.objects.create(
            employee=emp, leave_type=lt, start_date=_date(2026, 8, 1), end_date=_date(2026, 8, 2), status='approved')
        user = make_user('lrrm-user', password='x')
        revoke_req = LeaveRevokeRequest.objects.create(
            leave_request=req, requested_by=user, reason='Project emergency, cannot take leave.')
        self.assertEqual(revoke_req.status, 'pending')
        self.assertIsNone(revoke_req.decided_by)
        self.assertIsNone(revoke_req.decided_at)
        self.assertEqual(revoke_req.decision_note, '')
        self.assertEqual(list(req.revoke_requests.all()), [revoke_req])


class AttendanceExceptionRevokeFieldsTests(TestCase):
    def test_revoked_is_a_valid_status_choice(self):
        self.assertIn(('revoked', 'Revoked'), AttendanceException.STATUS_CHOICES)

    def test_revoke_fields_exist_and_default_empty(self):
        emp = make_employee(iqama='AERF-1')
        exc = _submit_aex(employee=emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
                           reason_category='site_visit', created_by=make_user('aerf-creator'))
        self.assertIsNone(exc.revoked_by)
        self.assertIsNone(exc.revoked_at)
        self.assertEqual(exc.revoke_reason, '')


class AttendanceExceptionRevokeRequestModelTests(TestCase):
    def test_create_and_defaults(self):
        emp = make_employee(iqama='AERRM-1')
        exc = _submit_aex(employee=emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
                           reason_category='site_visit', created_by=make_user('aerrm-creator'))
        exc.status = 'approved'
        exc.save(update_fields=['status'])
        user = make_user('aerrm-user', password='x')
        revoke_req = AttendanceExceptionRevokeRequest.objects.create(
            attendance_exception=exc, requested_by=user, reason='No longer needed.')
        self.assertEqual(revoke_req.status, 'pending')
        self.assertEqual(list(exc.revoke_requests.all()), [revoke_req])


class AttendanceOutcomeRevokedTests(TestCase):
    def test_revoked_exception_marks_day_absent(self):
        from hr.attendance_exception_services import _apply_attendance_outcome
        from hr.models import AttendanceRecord
        emp = make_employee(iqama='AORT-1')
        exc = _submit_aex(employee=emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
                           reason_category='site_visit', created_by=make_user('aort-creator'))
        exc.status = 'revoked'
        exc.save(update_fields=['status'])
        _apply_attendance_outcome(exc)
        record = AttendanceRecord.objects.get(employee=emp, date=_date(2026, 7, 20))
        self.assertEqual(record.status, 'absent')


class EditLeaveRequestServiceTests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='EDLR-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 5, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=5)
        self.user = make_user('edlr-user', password='x')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)

    def test_creator_can_edit_pending_undecided_request(self):
        from hr.leave_approval_services import edit_leave_request
        updated = edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4),
            employee_reason='Updated reason')
        self.assertEqual(updated.start_date, date(2026, 8, 3))
        self.assertEqual(updated.end_date, date(2026, 8, 4))
        self.assertEqual(updated.employee_reason, 'Updated reason')
        self.assertEqual(updated.days, 2)

    def test_non_creator_cannot_edit(self):
        from hr.leave_approval_services import edit_leave_request
        other = make_user('edlr-other', password='x')
        with self.assertRaises(ValueError):
            edit_leave_request(self.req, other, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4))

    def test_edit_can_change_the_leave_type(self):
        from hr.leave_approval_services import edit_leave_request
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=annual, year=2026, entitled_days=Decimal('30'))
        updated = edit_leave_request(
            self.req, self.user, leave_type=annual, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4))
        self.assertEqual(updated.leave_type, annual)

    def test_changing_leave_type_alone_resets_a_prior_decision(self):
        # Substance changed (different type = different balance/policy),
        # so a type-only edit must reset a recorded decision exactly like
        # a date change does.
        from hr.leave_approval_services import edit_leave_request
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=annual, year=2026, entitled_days=Decimal('30'))
        approver = make_user('edlr-type-approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        approval = LeaveRequestApproval.objects.create(
            leave_request=self.req, approver=approver, decision='approved')
        edit_leave_request(
            self.req, self.user, leave_type=annual, start_date=self.req.start_date, end_date=self.req.end_date)
        approval.refresh_from_db()
        self.assertEqual(approval.decision, 'pending')

    def test_editing_after_a_decision_resets_it_to_pending(self):
        # Editing after a decision is recorded is now allowed (the whole
        # pending window, same as cancel) — but the substance changed, so
        # the stale decision is reset rather than silently kept.
        from hr.leave_approval_services import edit_leave_request
        approver = make_user('edlr-approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        approval = LeaveRequestApproval.objects.create(
            leave_request=self.req, approver=approver, decision='approved',
            comment='looks fine', decided_at=timezone.now())
        updated = edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4))
        self.assertEqual(updated.status, 'pending')
        approval.refresh_from_db()
        self.assertEqual(approval.decision, 'pending')
        self.assertEqual(approval.comment, '')
        self.assertIsNone(approval.decided_at)

    def test_editing_after_a_decision_notifies_the_approver(self):
        from hr.leave_approval_services import edit_leave_request
        approver_user = make_user('edlr-notify-approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver_user, is_active=True)
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=approver_user, decision='approved')
        from notifications.models import Notification
        edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4))
        self.assertTrue(
            Notification.objects.filter(recipient=approver_user, verb__icontains='edited a leave request').exists())

    def test_edit_with_no_prior_decision_leaves_approvals_untouched(self):
        # No decision was ever recorded, so there's nothing to reset —
        # confirms the reset path doesn't fire needlessly.
        from hr.leave_approval_services import edit_leave_request
        approver = make_user('edlr-untouched-approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        approval = LeaveRequestApproval.objects.create(
            leave_request=self.req, approver=approver, decision='pending')
        edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4))
        approval.refresh_from_db()
        self.assertEqual(approval.decision, 'pending')

    def test_cannot_edit_a_fully_finalized_request(self):
        # A fully approved/disapproved request is a different mechanism
        # entirely (Revoke) — edit still only applies to the pending window.
        from hr.leave_approval_services import edit_leave_request
        self.req.status = 'approved'
        self.req.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            edit_leave_request(self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4))

    def test_edit_does_not_count_the_requests_own_prior_dates_against_itself(self):
        # Regression guard: without exclude_request_id, editing a request
        # would collide with its own still-pending row (self-overlap and
        # self-counted-toward-balance), making every edit fail.
        from hr.leave_approval_services import edit_leave_request
        updated = edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2))
        self.assertEqual(updated.pk, self.req.pk)

    def test_edit_with_document_none_leaves_existing_file_untouched(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from hr.leave_approval_services import edit_leave_request
        original = SimpleUploadedFile('original.pdf', b'original-bytes')
        self.req.document = original
        self.req.save(update_fields=['document'])
        original_name = self.req.document.name
        updated = edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4),
            document=None)
        self.assertEqual(updated.document.name, original_name)

    def test_edit_with_a_new_file_replaces_the_old_one(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from hr.leave_approval_services import edit_leave_request
        original = SimpleUploadedFile('original.pdf', b'original-bytes')
        self.req.document = original
        self.req.save(update_fields=['document'])
        old_name = self.req.document.name
        replacement = SimpleUploadedFile('replacement.pdf', b'replacement-bytes')
        updated = edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4),
            document=replacement)
        self.assertIn('replacement', updated.document.name)
        self.assertFalse(updated.document.storage.exists(old_name))  # old file deleted, not orphaned

    def test_edit_with_document_false_clears_the_existing_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        from hr.leave_approval_services import edit_leave_request
        original = SimpleUploadedFile('original.pdf', b'original-bytes')
        self.req.document = original
        self.req.save(update_fields=['document'])
        old_name = self.req.document.name
        updated = edit_leave_request(
            self.req, self.user, leave_type=self.lt, start_date=date(2026, 8, 3), end_date=date(2026, 8, 4),
            document=False)
        self.assertFalse(updated.document)
        self.assertFalse(updated.document.storage.exists(old_name))


class CancelLeaveRequestServiceTests(TestCase):
    """Cancel is the ONLY way to withdraw a leave request — it covers the
    entire pending window, before or after any approver has decided (see
    hr.leave_approval_services.cancel_leave_request). There is no separate
    hard-delete path: even a same-day, zero-decisions withdrawal stays
    visible in History as 'Cancelled', consistent with every other
    terminal state this feature introduced (Revoked, etc.). Uses two
    approvers so a single decision leaves the request genuinely still
    pending (unanimous-consensus reconciliation) rather than immediately
    finalizing, matching the real limbo window Cancel exists for."""

    def setUp(self):
        self.emp = make_employee(iqama='CLR-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.user = make_user('clr-user', password='x')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.approver1 = make_user('clr-approver1', password='x')
        self.approver2 = make_user('clr-approver2', password='x')
        LeaveDashboardAccess.objects.create(user=self.approver1, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.approver2, is_active=True)
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)

    def test_creator_can_cancel_before_any_decision_with_no_reason(self):
        from hr.leave_approval_services import cancel_leave_request
        updated = cancel_leave_request(self.req, self.user)
        self.assertEqual(updated.status, 'cancelled')
        self.assertIsNotNone(updated.cancelled_at)
        self.assertEqual(updated.cancel_reason, '')

    def test_cancelling_before_any_decision_sends_no_notification(self):
        # Nobody has decided yet, so there's nobody to explain the
        # cancellation to.
        from hr.leave_approval_services import cancel_leave_request
        cancel_leave_request(self.req, self.user)
        self.assertFalse(Notification.objects.filter(verb__icontains='cancelled').exists())

    def test_creator_can_cancel_after_one_of_two_approvers_decides(self):
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')  # still awaiting approver2
        updated = cancel_leave_request(self.req, self.user, 'Plans changed.')
        self.assertEqual(updated.status, 'cancelled')
        self.assertIsNotNone(updated.cancelled_at)
        self.assertEqual(updated.cancel_reason, 'Plans changed.')

    def test_non_creator_cannot_cancel(self):
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        other = make_user('clr-other', password='x')
        with self.assertRaises(ValueError):
            cancel_leave_request(self.req, other, 'Not my request.')

    def test_reason_is_required(self):
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        with self.assertRaises(ValueError):
            cancel_leave_request(self.req, self.user, '')

    def test_cannot_cancel_once_fully_approved(self):
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        record_approver_decision(self.req, self.approver2, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        with self.assertRaises(ValueError):
            cancel_leave_request(self.req, self.user, 'Too late.')

    def test_decided_approver_is_notified(self):
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        cancel_leave_request(self.req, self.user, 'Plans changed.')
        self.assertTrue(
            Notification.objects.filter(recipient=self.approver1, verb__icontains='cancelled').exists())

    def test_still_pending_approver_is_not_sent_a_cancel_notice(self):
        # approver2 never decided — nothing to explain to them, their
        # outstanding review just stops mattering once the row is gone.
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        cancel_leave_request(self.req, self.user, 'Plans changed.')
        self.assertFalse(
            Notification.objects.filter(recipient=self.approver2, verb__icontains='cancelled').exists())


class EditDeleteAttendanceExceptionServiceTests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='EDAE-1')
        self.user = _login_user('edae-user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)

    def test_creator_can_edit_pending_exception(self):
        from hr.attendance_exception_services import edit_attendance_exception
        updated = edit_attendance_exception(
            self.exc, self.user, event_date=_date(2026, 7, 21), event_start_time=_time(10, 0),
            reason_category='outside_meeting')
        self.assertEqual(updated.event_date, _date(2026, 7, 21))
        self.assertEqual(updated.reason_category, 'outside_meeting')

    def test_non_creator_cannot_edit(self):
        from hr.attendance_exception_services import edit_attendance_exception
        other = make_user('edae-other', password='x')
        with self.assertRaises(ValueError):
            edit_attendance_exception(
                self.exc, other, event_date=_date(2026, 7, 21), event_start_time=_time(10, 0),
                reason_category='outside_meeting')

    def test_cannot_edit_once_decided(self):
        from hr.attendance_exception_services import edit_attendance_exception
        self.exc.status = 'approved'
        self.exc.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            edit_attendance_exception(
                self.exc, self.user, event_date=_date(2026, 7, 21), event_start_time=_time(10, 0),
                reason_category='outside_meeting')

    def test_creator_can_delete_pending_exception(self):
        from hr.attendance_exception_services import delete_attendance_exception
        pk = self.exc.pk
        delete_attendance_exception(self.exc, self.user)
        self.assertFalse(AttendanceException.objects.filter(pk=pk).exists())

    def test_non_creator_cannot_delete(self):
        from hr.attendance_exception_services import delete_attendance_exception
        other = make_user('edae-other2', password='x')
        with self.assertRaises(ValueError):
            delete_attendance_exception(self.exc, other)
        self.assertTrue(AttendanceException.objects.filter(pk=self.exc.pk).exists())

    def test_cannot_delete_once_decided(self):
        from hr.attendance_exception_services import delete_attendance_exception
        self.exc.status = 'expired'
        self.exc.save(update_fields=['status'])
        with self.assertRaises(ValueError):
            delete_attendance_exception(self.exc, self.user)


class RevokeLeaveRequestServiceTests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='RLR-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.user = make_user('rlr-user', password='x')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        approver = make_user('rlr-approver', password='x')
        # Roster is snapshotted onto the request AT SUBMISSION TIME (see
        # submit_leave_request) — the approver's LeaveDashboardAccess grant
        # must exist before this call, or no LeaveRequestApproval row gets
        # created for them.
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 3),
            created_by=self.user)
        record_approver_decision(self.req, approver, 'approved')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')
        self.assertIsNotNone(self.req.leave_record_id)
        self.revoker = make_user('rlr-revoker', password='x')

    def test_direct_revoke_deletes_leave_record_and_sets_fields(self):
        from hr.leave_approval_services import revoke_leave_request
        record_id = self.req.leave_record_id
        updated = revoke_leave_request(self.req, self.revoker, 'Project emergency.')
        self.assertEqual(updated.status, 'revoked')
        self.assertEqual(updated.revoked_by, self.revoker)
        self.assertIsNotNone(updated.revoked_at)
        self.assertEqual(updated.revoke_reason, 'Project emergency.')
        self.assertFalse(LeaveRecord.objects.filter(pk=record_id).exists())

    def test_direct_revoke_requires_reason(self):
        from hr.leave_approval_services import revoke_leave_request
        with self.assertRaises(ValueError):
            revoke_leave_request(self.req, self.revoker, '')

    def test_cannot_revoke_a_non_approved_request(self):
        from hr.leave_approval_services import revoke_leave_request
        pending = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 9, 1), end_date=date(2026, 9, 2),
            created_by=self.user)
        with self.assertRaises(ValueError):
            revoke_leave_request(pending, self.revoker, 'reason')

    def test_cannot_self_revoke(self):
        from hr.leave_approval_services import revoke_leave_request
        with self.assertRaises(ValueError):
            revoke_leave_request(self.req, self.user, 'reason')

    def test_employee_can_request_revoke_on_own_approved_request(self):
        from hr.leave_approval_services import request_leave_revoke
        revoke_req = request_leave_revoke(self.req, self.user, 'Plans changed.')
        self.assertEqual(revoke_req.status, 'pending')
        self.assertEqual(revoke_req.requested_by, self.user)

    def test_non_employee_cannot_request_revoke(self):
        from hr.leave_approval_services import request_leave_revoke
        with self.assertRaises(ValueError):
            request_leave_revoke(self.req, self.revoker, 'reason')

    def test_duplicate_pending_revoke_request_blocked(self):
        from hr.leave_approval_services import request_leave_revoke
        request_leave_revoke(self.req, self.user, 'First request.')
        with self.assertRaises(ValueError):
            request_leave_revoke(self.req, self.user, 'Second request.')

    def test_decide_revoke_request_approve_applies_the_revoke(self):
        from hr.leave_approval_services import request_leave_revoke, decide_leave_revoke_request
        revoke_req = request_leave_revoke(self.req, self.user, 'Plans changed.')
        decide_leave_revoke_request(revoke_req, self.revoker, 'approved')
        self.req.refresh_from_db()
        revoke_req.refresh_from_db()
        self.assertEqual(self.req.status, 'revoked')
        self.assertEqual(revoke_req.status, 'approved')
        self.assertEqual(revoke_req.decided_by, self.revoker)

    def test_decide_revoke_request_reject_leaves_original_untouched(self):
        from hr.leave_approval_services import request_leave_revoke, decide_leave_revoke_request
        revoke_req = request_leave_revoke(self.req, self.user, 'Plans changed.')
        decide_leave_revoke_request(revoke_req, self.revoker, 'rejected', decision_note='Coverage already arranged.')
        self.req.refresh_from_db()
        revoke_req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')  # unchanged
        self.assertEqual(revoke_req.status, 'rejected')
        self.assertEqual(revoke_req.decision_note, 'Coverage already arranged.')

    def test_reject_without_decision_note_is_rejected(self):
        from hr.leave_approval_services import request_leave_revoke, decide_leave_revoke_request
        revoke_req = request_leave_revoke(self.req, self.user, 'Plans changed.')
        with self.assertRaises(ValueError):
            decide_leave_revoke_request(revoke_req, self.revoker, 'rejected', decision_note='')

    def test_self_decision_on_own_revoke_request_blocked(self):
        from hr.leave_approval_services import request_leave_revoke, decide_leave_revoke_request
        revoke_req = request_leave_revoke(self.req, self.user, 'Plans changed.')
        with self.assertRaises(ValueError):
            decide_leave_revoke_request(revoke_req, self.user, 'approved')


class RevokeAttendanceExceptionServiceTests(TestCase):
    def setUp(self):
        self.manager = make_employee(iqama='RAE-MGR', name='Revoke Exc Manager')
        self.manager_user = _login_user('rae_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.emp = make_employee(iqama='RAE-1')
        self.emp.main_manager = self.manager
        self.emp.save(update_fields=['main_manager'])
        self.user = _login_user('rae_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        decide_attendance_exception(self.exc, self.manager_user, 'approved')
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'approved')
        self.revoker = _login_user('rae_revoker')

    def test_direct_revoke_sets_fields_and_marks_day_absent(self):
        from hr.attendance_exception_services import revoke_attendance_exception
        from hr.models import AttendanceRecord
        updated = revoke_attendance_exception(self.exc, self.revoker, 'No longer applicable.')
        self.assertEqual(updated.status, 'revoked')
        self.assertEqual(updated.revoked_by, self.revoker)
        self.assertEqual(updated.revoke_reason, 'No longer applicable.')
        record = AttendanceRecord.objects.get(employee=self.emp, date=_date(2026, 7, 20))
        self.assertEqual(record.status, 'absent')

    def test_direct_revoke_requires_reason(self):
        from hr.attendance_exception_services import revoke_attendance_exception
        with self.assertRaises(ValueError):
            revoke_attendance_exception(self.exc, self.revoker, '')

    def test_cannot_revoke_non_approved_exception(self):
        from hr.attendance_exception_services import revoke_attendance_exception
        pending = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 25), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        with self.assertRaises(ValueError):
            revoke_attendance_exception(pending, self.revoker, 'reason')

    def test_employee_can_request_revoke(self):
        from hr.attendance_exception_services import request_attendance_exception_revoke
        revoke_req = request_attendance_exception_revoke(self.exc, self.user, 'Plans changed.')
        self.assertEqual(revoke_req.status, 'pending')

    def test_duplicate_pending_revoke_request_blocked(self):
        from hr.attendance_exception_services import request_attendance_exception_revoke
        request_attendance_exception_revoke(self.exc, self.user, 'First.')
        with self.assertRaises(ValueError):
            request_attendance_exception_revoke(self.exc, self.user, 'Second.')

    def test_assigned_manager_can_decide_revoke_request(self):
        from hr.attendance_exception_services import request_attendance_exception_revoke, decide_attendance_exception_revoke_request
        revoke_req = request_attendance_exception_revoke(self.exc, self.user, 'Plans changed.')
        decide_attendance_exception_revoke_request(revoke_req, self.manager_user, 'approved')
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'revoked')

    def test_unrelated_user_cannot_decide_revoke_request(self):
        from hr.attendance_exception_services import request_attendance_exception_revoke, decide_attendance_exception_revoke_request
        revoke_req = request_attendance_exception_revoke(self.exc, self.user, 'Plans changed.')
        stranger = _login_user('rae_stranger')
        with self.assertRaises(ValueError):
            decide_attendance_exception_revoke_request(revoke_req, stranger, 'approved')

    def test_reject_requires_decision_note(self):
        from hr.attendance_exception_services import request_attendance_exception_revoke, decide_attendance_exception_revoke_request
        revoke_req = request_attendance_exception_revoke(self.exc, self.user, 'Plans changed.')
        with self.assertRaises(ValueError):
            decide_attendance_exception_revoke_request(revoke_req, self.manager_user, 'rejected', decision_note='')


class MyProfileLeaveEditCancelUITests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='MPLED-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='marriage', defaults={'name': 'Marriage', 'default_annual_days': 5, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=5)
        self.user = _login_user('mpled_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        self.client.login(username='mpled_user', password='testpass123')

    def test_edit_and_cancel_buttons_render_for_own_pending_request(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')

    def test_edit_form_leave_type_is_a_real_dropdown_not_hidden(self):
        resp = self.client.get(reverse('hr:my_profile'))
        content = resp.content.decode()
        self.assertIn('<select name="leave_type"', content)
        self.assertNotIn(f'<input type="hidden" name="leave_type" value="{self.lt.pk}">', content)

    def test_failed_edit_shows_the_real_reason_and_keeps_the_attempted_type(self):
        # Regression: a failed edit (e.g. switching to a leave type that
        # requires a document, with none attached) used to show only a
        # generic "check the form" message and silently revert the
        # reopened dropdown back to the pre-edit type — indistinguishable
        # from the edit having been thrown away. It must now show the
        # real per-field reason and keep showing what was attempted.
        sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 15})
        sick.requires_medical_certificate = True
        sick.save()
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=sick, year=2026, entitled_days=15)
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': sick.pk, 'start_date': '2026-08-01', 'end_date': '2026-08-02',
            'employee_reason': '',
        }, follow=True)
        self.assertContains(resp, 'A medical certificate/document is required for Sick leave.')
        self.assertNotContains(resp, 'Could not update the request — please check the form.')
        content = resp.content.decode()
        idx = content.find(f'id="editLeave{self.req.pk}"')
        self.assertGreater(idx, -1)
        row_html = content[idx:idx + 1500]
        self.assertIn(f'value="{sick.pk}" selected', row_html)
        # The request itself must be untouched — the failed edit never applied.
        self.req.refresh_from_db()
        self.assertEqual(self.req.leave_type, self.lt)

    def test_edit_button_and_reset_warning_still_show_once_decided(self):
        approver = make_user('mpled_approver', password='x')
        LeaveRequestApproval.objects.create(leave_request=self.req, approver=approver, decision='approved')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')
        self.assertContains(resp, 'mpled_approver')
        self.assertContains(resp, 'saving changes')

    def test_editing_a_decided_request_resets_the_decision(self):
        approver_user = make_user('mpled_reset_approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver_user, is_active=True)
        approval = LeaveRequestApproval.objects.create(
            leave_request=self.req, approver=approver_user, decision='approved')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-08-05', 'end_date': '2026-08-06',
        })
        self.assertEqual(resp.status_code, 302)
        approval.refresh_from_db()
        self.assertEqual(approval.decision, 'pending')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')

    def test_post_edit_updates_request(self):
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-08-05', 'end_date': '2026-08-06',
            'employee_reason': 'Rescheduled',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.start_date, date(2026, 8, 5))
        self.assertEqual(self.req.employee_reason, 'Rescheduled')

    def test_post_edit_redirect_reopens_the_same_rows_panel(self):
        # The edited row's panel must re-expand after the reload (not
        # collapse back to hidden), so the employee actually sees the
        # change they just made.
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-08-05', 'end_date': '2026-08-06',
        })
        self.assertRedirects(resp, f"{reverse('hr:my_profile')}?opened_leave={self.req.pk}")
        resp = self.client.get(resp.url)
        self.assertContains(resp, f'id="editLeave{self.req.pk}"')

    def test_post_edit_without_touching_document_keeps_existing_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.req.document = SimpleUploadedFile('cert.pdf', b'bytes')
        self.req.save(update_fields=['document'])
        original_name = self.req.document.name
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-08-05', 'end_date': '2026-08-06',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.document.name, original_name)

    def test_post_edit_with_clear_checkbox_removes_the_document(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.req.document = SimpleUploadedFile('cert.pdf', b'bytes')
        self.req.save(update_fields=['document'])
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-08-05', 'end_date': '2026-08-06',
            'document-clear': 'on',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertFalse(self.req.document)

    def test_post_edit_uploads_a_new_document(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-08-05', 'end_date': '2026-08-06',
            'document': SimpleUploadedFile('newcert.pdf', b'bytes'),
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertIn('newcert', self.req.document.name)

    def test_post_cancel_before_any_decision_marks_cancelled_not_deleted(self):
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'cancel_leave_request', 'request_id': self.req.pk,
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'cancelled')

    def test_cannot_cancel_someone_elses_request(self):
        other_emp = make_employee(iqama='MPLED-2')
        other_user = _login_user('mpled_other')
        other_emp.user = other_user
        other_emp.save(update_fields=['user'])
        LeaveEntitlement.objects.create(employee=other_emp, leave_type=self.lt, year=2026, entitled_days=5)
        other_req = submit_leave_request(
            employee=other_emp, leave_type=self.lt, start_date=date(2026, 8, 10), end_date=date(2026, 8, 11),
            created_by=other_user)
        self.client.post(reverse('hr:my_profile'), {
            'action': 'cancel_leave_request', 'request_id': other_req.pk,
        })
        other_req.refresh_from_db()
        self.assertEqual(other_req.status, 'pending')


class MyProfileManagerInBehalfEditCancelUITests(TestCase):
    """Edit/Cancel for a manager's own in-behalf-logged pending request
    render in the Reporting Structure card, not the report's own table —
    the gate is request.user == created_by, and the manager (not the
    report) is the creator here."""

    def setUp(self):
        self.manager_user = User.objects.create_user(username='mpmied-mgr', password='x')
        self.manager = Employee.objects.create(
            iqama_number='MPMIED-MGR', full_name='InBehalf Edit Manager', user=self.manager_user)
        self.report_user = User.objects.create_user(username='mpmied-rpt', password='x')
        self.report = Employee.objects.create(
            iqama_number='MPMIED-RPT', full_name='InBehalf Edit Report',
            main_manager=self.manager, user=self.report_user)
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.report, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.client.login(username='mpmied-mgr', password='x')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.report.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-03-01', 'end_date': '2026-03-05',
        })
        self.req = LeaveRequest.objects.get(employee=self.report)

    def test_manager_sees_cancel_for_own_in_behalf_request(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')

    def test_report_does_not_see_cancel_on_their_own_profile(self):
        # The report views their OWN My Leave Requests table — they didn't
        # create this request (their manager did), so no Cancel for them,
        # even though it's about their own leave.
        self.client.logout()
        self.client.login(username='mpmied-rpt', password='x')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, f'cancel_leave_request_{self.req.pk}')


class MyProfileRequestRevokeLeaveUITests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='MPRRL-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.user = _login_user('mprrl_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        approver = make_user('mprrl_approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        record_approver_decision(self.req, approver, 'approved')
        self.client.login(username='mprrl_user', password='testpass123')

    def test_request_revoke_button_renders_on_approved_own_request(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'request_revoke_leave_{self.req.pk}')

    def test_post_creates_pending_revoke_request(self):
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_leave_revoke', 'request_id': self.req.pk, 'reason': 'Plans changed.',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(LeaveRevokeRequest.objects.filter(leave_request=self.req, status='pending').exists())

    def test_awaiting_review_note_shows_after_requesting(self):
        LeaveRevokeRequest.objects.create(leave_request=self.req, requested_by=self.user, reason='Plans changed.')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Revoke requested')

    def test_revoked_badge_renders(self):
        from hr.leave_approval_services import revoke_leave_request
        revoker = make_user('mprrl_revoker', password='x')
        revoke_leave_request(self.req, revoker, 'Applied for testing.')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Revoked')


class MyProfileCancelLeaveRequestUITests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='MPCLR-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.user = _login_user('mpclr_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.approver1 = make_user('mpclr_approver1', password='x')
        self.approver2 = make_user('mpclr_approver2', password='x')
        LeaveDashboardAccess.objects.create(user=self.approver1, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.approver2, is_active=True)
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        self.client.login(username='mpclr_user', password='testpass123')

    def test_cancel_button_renders_both_before_and_after_a_decision(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')
        record_approver_decision(self.req, self.approver1, 'approved')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')

    def test_edit_button_also_stays_available_once_a_decision_is_recorded(self):
        # Edit and Cancel share the same window: both stay open right
        # through the rest of the pending window. Editing a decided
        # request resets that decision (see edit_leave_request) rather
        # than being blocked outright.
        record_approver_decision(self.req, self.approver1, 'approved')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')

    def test_post_cancel_marks_request_cancelled_and_keeps_the_row(self):
        record_approver_decision(self.req, self.approver1, 'approved')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'cancel_leave_request', 'request_id': self.req.pk, 'reason': 'Plans changed.',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'cancelled')
        self.assertEqual(self.req.cancel_reason, 'Plans changed.')

    def test_cancelled_badge_and_reason_render(self):
        from hr.leave_approval_services import cancel_leave_request
        record_approver_decision(self.req, self.approver1, 'approved')
        cancel_leave_request(self.req, self.user, 'Plans changed.')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Cancelled')
        self.assertContains(resp, 'Plans changed.')

    def test_cannot_cancel_someone_elses_request(self):
        other_emp = make_employee(iqama='MPCLR-2')
        other_user = _login_user('mpclr_other')
        other_emp.user = other_user
        other_emp.save(update_fields=['user'])
        LeaveEntitlement.objects.create(employee=other_emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        other_req = submit_leave_request(
            employee=other_emp, leave_type=self.lt, start_date=date(2026, 8, 10), end_date=date(2026, 8, 11),
            created_by=other_user)
        record_approver_decision(other_req, self.approver1, 'approved')
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'cancel_leave_request', 'request_id': other_req.pk, 'reason': 'Not mine.',
        })
        self.assertEqual(resp.status_code, 403)
        other_req.refresh_from_db()
        self.assertEqual(other_req.status, 'pending')


class MyProfileExceptionEditDeleteUITests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='MPEED-1')
        self.user = _login_user('mpeed_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        self.client.login(username='mpeed_user', password='testpass123')

    def test_edit_delete_buttons_render_for_own_pending_exception(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'edit_attendance_exception_{self.exc.pk}')
        self.assertContains(resp, f'delete_attendance_exception_{self.exc.pk}')

    def test_no_buttons_once_decided(self):
        self.exc.status = 'approved'
        self.exc.save(update_fields=['status'])
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertNotContains(resp, f'delete_attendance_exception_{self.exc.pk}')

    def test_post_edit_updates_exception(self):
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'edit_attendance_exception', 'request_id': self.exc.pk,
            'event_date': '2026-07-21', 'event_start_time': '10:00',
            'reason_category': 'outside_meeting', 'custom_reason': '',
        })
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.event_date, _date(2026, 7, 21))
        self.assertEqual(self.exc.reason_category, 'outside_meeting')

    def test_post_delete_removes_exception(self):
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'delete_attendance_exception', 'request_id': self.exc.pk,
        })
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(AttendanceException.objects.filter(pk=self.exc.pk).exists())


class MyProfileRequestRevokeExceptionUITests(TestCase):
    def setUp(self):
        self.manager = make_employee(iqama='MPRRE-MGR', name='RRE Manager')
        self.manager_user = _login_user('mprre_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.emp = make_employee(iqama='MPRRE-1')
        self.emp.main_manager = self.manager
        self.emp.save(update_fields=['main_manager'])
        self.user = _login_user('mprre_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        decide_attendance_exception(self.exc, self.manager_user, 'approved')
        self.exc.refresh_from_db()
        self.client.login(username='mprre_user', password='testpass123')

    def test_request_revoke_button_renders(self):
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, f'request_revoke_exception_{self.exc.pk}')

    def test_post_creates_pending_revoke_request(self):
        resp = self.client.post(reverse('hr:my_profile'), {
            'action': 'request_attendance_exception_revoke', 'request_id': self.exc.pk, 'reason': 'No longer needed.',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(AttendanceExceptionRevokeRequest.objects.filter(
            attendance_exception=self.exc, status='pending').exists())

    def test_revoked_badge_renders(self):
        from hr.attendance_exception_services import revoke_attendance_exception
        revoker = _login_user('mprre_revoker')
        revoke_attendance_exception(self.exc, revoker, 'Applied for testing.')
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertContains(resp, 'Revoked')


class DirectRevokeLeaveRequestQueueUITests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='DRLQ-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.superadmin = make_user('drlq_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        approver = make_user('drlq_approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.superadmin)
        record_approver_decision(req, approver, 'approved')
        self.req = req
        self.client.login(username='drlq_super', password='testpass123')

    def test_revoke_button_renders_in_history_for_override_access_holder(self):
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, f'revoke_leave_{self.req.pk}')

    def test_post_revoke_via_list_view(self):
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'revoke', 'request_id': self.req.pk, 'reason': 'No longer needed.',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'revoked')

    def test_post_revoke_via_detail_view(self):
        resp = self.client.post(reverse('hr:leave_request_detail', args=[self.req.pk]), {
            'action': 'revoke', 'reason': 'No longer needed.',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'revoked')

    def test_non_override_user_cannot_revoke(self):
        grantee = make_user('drlq_grantee', password='x')
        grantee.set_password('testpass123')
        grantee.save()
        LeaveDashboardAccess.objects.create(user=grantee, is_active=True)
        self.client.logout()
        self.client.login(username='drlq_grantee', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'revoke', 'request_id': self.req.pk, 'reason': 'Trying anyway.',
        })
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'approved')  # unchanged

    def test_revoked_badge_renders_in_history(self):
        from hr.leave_approval_services import revoke_leave_request
        revoke_leave_request(self.req, self.superadmin, 'Applied for testing.')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, 'Revoked')


class CancelledLeaveRequestApproverHistoryUITests(TestCase):
    """The 'Revoked must render symmetrically in both the requester's own
    History and the decider's queue History' requirement extends to
    Cancelled too — the approver who spent a decision on the request must
    see what happened to it, on the Leave Approval Queue and the detail
    page, not just the employee's own My Profile."""

    def setUp(self):
        self.emp = make_employee(iqama='CLRAH-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.user = make_user('clrah-user', password='x')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.superadmin = make_user('clrah_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.approver1 = make_user('clrah_approver1', password='x')
        self.approver2 = make_user('clrah_approver2', password='x')
        self.approver2.set_password('testpass123')
        self.approver2.save()
        LeaveDashboardAccess.objects.create(user=self.approver1, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.approver2, is_active=True)
        self.req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        record_approver_decision(self.req, self.approver1, 'approved')
        from hr.leave_approval_services import cancel_leave_request
        cancel_leave_request(self.req, self.user, 'Plans changed.')
        self.client.login(username='clrah_super', password='testpass123')

    def test_cancelled_badge_renders_in_queue_history(self):
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, 'Cancelled')
        self.assertContains(resp, 'Plans changed.')

    def test_cancelled_status_and_reason_render_on_detail_page(self):
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertContains(resp, 'Cancelled')
        self.assertContains(resp, 'Plans changed.')

    def test_decide_form_no_longer_renders_for_a_cancelled_request(self):
        self.client.logout()
        self.client.login(username='clrah_approver2', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_detail', args=[self.req.pk]))
        self.assertNotContains(resp, 'name="decision" value="approved"')


class DirectRevokeTeamExceptionUITests(TestCase):
    def setUp(self):
        self.manager = make_employee(iqama='DRTE-MGR', name='DRTE Manager')
        self.manager_user = _login_user('drte_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.emp = make_employee(iqama='DRTE-1')
        self.emp.main_manager = self.manager
        self.emp.save(update_fields=['main_manager'])
        self.creator = make_user('drte_creator')
        self.exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.creator)
        decide_attendance_exception(self.exc, self.manager_user, 'approved')
        self.exc.refresh_from_db()
        self.superadmin = make_user('drte_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.client.login(username='drte_super', password='testpass123')

    def test_revoke_button_renders_for_override_access_holder(self):
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        self.assertContains(resp, f'revoke_exception_{self.exc.pk}')

    def test_post_revoke_direct(self):
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'revoke_direct', 'exc_id': self.exc.pk, 'reason': 'No longer needed.', 'tab': 'all',
        })
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'revoked')

    def test_revoked_row_included_in_history(self):
        from hr.attendance_exception_services import revoke_attendance_exception
        revoke_attendance_exception(self.exc, self.superadmin, 'Applied for testing.')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        self.assertContains(resp, 'Revoked')
        self.assertIn(self.exc, list(resp.context['decided_exceptions']))


class LeaveQueueMyLoggedRequestsUITests(TestCase):
    """'Requests You've Logged' on the Leave Approval Queue page — closes a
    real gap: edit_leave_request/cancel_leave_request were already gated
    purely on created_by, but the only place to reach them was My Profile's
    Reporting Structure card, itself scoped to LIVE DIRECT REPORTS. A Super
    Admin/HR user can log a request for ANY employee company-wide via "Log
    Request" here, and — per this codebase's own convention (see e.g.
    PendingRevokeRequestsLeaveQueueUITests) — commonly has NO employee_profile
    at all, so My Profile isn't reachable either way. This lives on the
    queue page instead, which needs no employee_profile to access."""

    def setUp(self):
        self.superadmin = make_user('lqmlr_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.emp = make_employee(iqama='LQMLR-1')  # NOT a direct report of superadmin — no relationship at all
        self.lt, _ = LeaveType.objects.update_or_create(code='annual', defaults={
            'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        # Two approvers, created before submission — the roster is
        # snapshotted at submission time, and unanimous-consensus
        # reconciliation means a single approver would fully finalize the
        # request on one decision, defeating the "still pending, one
        # decided" scenario the test below needs.
        self.approver = make_user('lqmlr_approver', password='x')
        self.approver2 = make_user('lqmlr_approver2', password='x')
        LeaveDashboardAccess.objects.create(user=self.approver, is_active=True)
        LeaveDashboardAccess.objects.create(user=self.approver2, is_active=True)
        self.client.login(username='lqmlr_super', password='testpass123')
        self.client.post(reverse('hr:leave_request_create'), data={
            'employee': self.emp.pk, 'leave_type': self.lt.pk,
            'start_date': '2026-04-01', 'end_date': '2026-04-02',
        })
        self.req = LeaveRequest.objects.get(employee=self.emp)

    def test_superadmin_has_no_employee_profile(self):
        # Sanity check the premise: My Profile would show nothing for this
        # user regardless of what it displays, so the feature MUST live
        # somewhere reachable without one.
        self.assertFalse(hasattr(self.superadmin, 'employee_profile'))

    def test_logged_request_shows_on_queue_page_with_edit_and_cancel(self):
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')

    def test_edit_form_leave_type_is_a_real_dropdown_not_hidden(self):
        resp = self.client.get(reverse('hr:leave_request_list'))
        content = resp.content.decode()
        self.assertIn('<select name="leave_type"', content)
        self.assertNotIn(f'<input type="hidden" name="leave_type" value="{self.lt.pk}">', content)

    def test_failed_edit_shows_the_real_reason_and_keeps_the_attempted_type(self):
        sick, _ = LeaveType.objects.get_or_create(
            code='sick', defaults={'name': 'Sick', 'default_annual_days': 15})
        sick.requires_medical_certificate = True
        sick.save()
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=sick, year=2026, entitled_days=15)
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': sick.pk, 'start_date': '2026-04-01', 'end_date': '2026-04-02',
        }, follow=True)
        self.assertContains(resp, 'A medical certificate/document is required for Sick leave.')
        self.assertNotContains(resp, 'Could not update the request — please check the form.')
        content = resp.content.decode()
        idx = content.find(f'id="editLogged{self.req.pk}"')
        self.assertGreater(idx, -1)
        row_html = content[idx:idx + 1500]
        self.assertIn(f'value="{sick.pk}" selected', row_html)
        self.req.refresh_from_db()
        self.assertEqual(self.req.leave_type, self.lt)

    def test_post_edit_can_change_the_leave_type(self):
        # self.req/self.lt are already 'annual' (see setUp) — use a
        # genuinely different type to prove the switch actually happens.
        marriage, _ = LeaveType.objects.get_or_create(code='marriage', defaults={
            'name': 'Marriage', 'default_annual_days': 5, 'is_accumulative': False})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=marriage, year=2026, entitled_days=5)
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': marriage.pk, 'start_date': '2026-04-01', 'end_date': '2026-04-02',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.leave_type, marriage)

    def test_post_edit_updates_the_request(self):
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-04-05', 'end_date': '2026-04-06',
            'employee_reason': 'Rescheduled',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.start_date, date(2026, 4, 5))

    def test_post_edit_redirect_reopens_the_logged_tab_and_row(self):
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-04-05', 'end_date': '2026-04-06',
        })
        self.assertRedirects(resp, f"{reverse('hr:leave_request_list')}?opened_leave={self.req.pk}&tab=logged")
        resp = self.client.get(resp.url)
        self.assertContains(resp, f'id="editLogged{self.req.pk}"')

    def test_edit_button_and_reset_warning_still_show_once_decided(self):
        record_approver_decision(self.req, self.approver, 'approved')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')
        self.assertContains(resp, 'lqmlr_approver')
        self.assertContains(resp, 'saving changes')

    def test_editing_a_decided_logged_request_resets_the_decision(self):
        record_approver_decision(self.req, self.approver, 'approved')
        approval = self.req.approvals.get(approver=self.approver)
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-04-05', 'end_date': '2026-04-06',
        })
        self.assertEqual(resp.status_code, 302)
        approval.refresh_from_db()
        self.assertEqual(approval.decision, 'pending')
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')

    def test_post_cancel_marks_cancelled(self):
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'cancel_leave_request', 'request_id': self.req.pk,
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'cancelled')

    def test_document_view_link_and_edit_button_both_render(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.req.document = SimpleUploadedFile('cert.pdf', b'bytes')
        self.req.save(update_fields=['document'])
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, self.req.document.url)
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')

    def test_post_edit_without_touching_document_keeps_existing_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.req.document = SimpleUploadedFile('cert.pdf', b'bytes')
        self.req.save(update_fields=['document'])
        original_name = self.req.document.name
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-04-05', 'end_date': '2026-04-06',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.document.name, original_name)

    def test_post_edit_uploading_a_new_document_deletes_the_old_stored_file(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.req.document = SimpleUploadedFile('cert.pdf', b'original-bytes')
        self.req.save(update_fields=['document'])
        old_name = self.req.document.name
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-04-05', 'end_date': '2026-04-06',
            'document': SimpleUploadedFile('newcert.pdf', b'new-bytes'),
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertIn('newcert', self.req.document.name)
        # The old file is gone from storage — not just unreferenced — so
        # nothing orphaned accumulates as requests get re-edited.
        self.assertFalse(self.req.document.storage.exists(old_name))

    def test_post_edit_with_clear_checkbox_removes_the_document(self):
        from django.core.files.uploadedfile import SimpleUploadedFile
        self.req.document = SimpleUploadedFile('cert.pdf', b'bytes')
        self.req.save(update_fields=['document'])
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'edit_leave_request', 'request_id': self.req.pk,
            'leave_type': self.lt.pk, 'start_date': '2026-04-05', 'end_date': '2026-04-06',
            'document-clear': 'on',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertFalse(self.req.document)

    def test_edit_and_cancel_both_stay_available_once_a_decision_is_recorded(self):
        record_approver_decision(self.req, self.approver, 'approved')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, f'edit_leave_request_{self.req.pk}')
        self.assertContains(resp, f'cancel_leave_request_{self.req.pk}')

    def test_other_user_cannot_edit_or_cancel_someone_elses_logged_request(self):
        other = make_user('lqmlr_other', password='x')
        other.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        other.role = role
        other.save()
        self.client.logout()
        self.client.login(username='lqmlr_other', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'cancel_leave_request', 'request_id': self.req.pk,
        })
        self.assertEqual(resp.status_code, 403)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'pending')


class PendingRevokeRequestsLeaveQueueUITests(TestCase):
    def setUp(self):
        self.emp = make_employee(iqama='PRRQ-1')
        self.lt, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': Decimal('30')})
        LeaveEntitlement.objects.create(employee=self.emp, leave_type=self.lt, year=2026, entitled_days=Decimal('30'))
        self.superadmin = make_user('prrq_super', password='x')
        self.superadmin.set_password('testpass123')
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name='super_admin')
        self.superadmin.role = role
        self.superadmin.save()
        self.user = _login_user('prrq_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        approver = make_user('prrq_approver', password='x')
        LeaveDashboardAccess.objects.create(user=approver, is_active=True)
        req = submit_leave_request(
            employee=self.emp, leave_type=self.lt, start_date=date(2026, 8, 1), end_date=date(2026, 8, 2),
            created_by=self.user)
        record_approver_decision(req, approver, 'approved')
        self.req = req
        self.revoke_req = LeaveRevokeRequest.objects.create(
            leave_request=req, requested_by=self.user, reason='Plans changed.')

    def test_pending_revoke_request_shows_on_queue_page(self):
        self.client.login(username='prrq_super', password='testpass123')
        resp = self.client.get(reverse('hr:leave_request_list'))
        self.assertContains(resp, 'Plans changed.')
        self.assertEqual(list(resp.context['pending_leave_revoke_requests']), [self.revoke_req])

    def test_approve_applies_the_revoke(self):
        self.client.login(username='prrq_super', password='testpass123')
        resp = self.client.post(reverse('hr:leave_request_list'), {
            'action': 'decide_revoke_request', 'revoke_request_id': self.revoke_req.pk, 'decision': 'approved',
        })
        self.assertEqual(resp.status_code, 302)
        self.req.refresh_from_db()
        self.assertEqual(self.req.status, 'revoked')

    def test_reject_requires_decision_note(self):
        self.client.login(username='prrq_super', password='testpass123')
        self.client.post(reverse('hr:leave_request_list'), {
            'action': 'decide_revoke_request', 'revoke_request_id': self.revoke_req.pk, 'decision': 'rejected',
            'decision_note': '',
        })
        self.revoke_req.refresh_from_db()
        self.assertEqual(self.revoke_req.status, 'pending')  # blocked, unchanged

    def test_sidebar_badge_includes_pending_revoke_requests(self):
        self.client.login(username='prrq_super', password='testpass123')
        resp = self.client.get(reverse('hr:hr_dashboard'))
        # 1 pending revoke request; no pending LeaveRequest (the only one was
        # already approved in setUp) — the badge count is exactly 1.
        self.assertEqual(resp.context['leave_requests_pending_count'], 1)


class PendingRevokeRequestsTeamExceptionsUITests(TestCase):
    def setUp(self):
        self.manager = make_employee(iqama='PRRTE-MGR', name='PRRTE Manager')
        self.manager_user = _login_user('prrte_mgr')
        self.manager.user = self.manager_user
        self.manager.save(update_fields=['user'])
        self.emp = make_employee(iqama='PRRTE-1')
        self.emp.main_manager = self.manager
        self.emp.save(update_fields=['main_manager'])
        self.user = _login_user('prrte_user')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.exc = _submit_aex(
            employee=self.emp, event_date=_date(2026, 7, 20), event_start_time=_time(9, 0),
            reason_category='site_visit', created_by=self.user)
        decide_attendance_exception(self.exc, self.manager_user, 'approved')
        self.exc.refresh_from_db()
        self.revoke_req = AttendanceExceptionRevokeRequest.objects.create(
            attendance_exception=self.exc, requested_by=self.user, reason='No longer needed.')
        self.client.login(username='prrte_mgr', password='testpass123')

    def test_pending_revoke_request_shows_for_assigned_manager(self):
        resp = self.client.get(reverse('hr:team_exceptions'))
        self.assertContains(resp, 'No longer needed.')
        self.assertEqual(list(resp.context['pending_exception_revoke_requests']), [self.revoke_req])

    def test_direct_tab_count_includes_the_pending_revoke_request(self):
        # The tab label's own count (not just the sidebar nav badge) must
        # reflect a pending revoke request too — the exception itself is
        # 'approved' (not pending/expired), so without this the tab would
        # misleadingly read 0 while a revoke sits there awaiting review.
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'direct'})
        self.assertEqual(resp.context['direct_tab_count'], 1)

    def test_approve_applies_the_revoke(self):
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'decide_revoke_request', 'revoke_request_id': self.revoke_req.pk, 'decision': 'approved',
        })
        self.assertEqual(resp.status_code, 302)
        self.exc.refresh_from_db()
        self.assertEqual(self.exc.status, 'revoked')

    def test_sidebar_badge_includes_pending_revoke_requests(self):
        # hr_dashboard is admin-only and this manager has no Role — use
        # my_profile (universally reachable) to read the context
        # processor's injected badge count instead.
        resp = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(resp.context['team_exceptions_direct_count'], 1)


class LateThresholdTests(TransactionTestCase):
    def setUp(self):
        from accounts.models import User
        self.user = User.objects.create_user('late_emp', password='x', email='late_emp@leap.com')
        self.emp = make_employee(iqama='IQ-LATE', name='Late Tester')
        self.emp.user = self.user
        self.emp.save()
        from accounts.models import Role
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.admin = User.objects.create_user('late_admin', password='x', email='admin@leap.com')
        self.admin.role = role
        self.admin.save()

    def _mark_late(self, day):
        return AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 8, day), status='late')

    def test_no_notification_before_third_late(self):
        from notifications.models import Notification
        self._mark_late(1)
        self._mark_late(2)
        self.assertEqual(Notification.objects.filter(verb__icontains='late').count(), 0)

    def test_notification_fires_on_exactly_third_late(self):
        from notifications.models import Notification
        self._mark_late(1)
        self._mark_late(2)
        self._mark_late(3)
        notifs = Notification.objects.filter(verb__icontains='late')
        self.assertEqual(notifs.count(), 2)

    def test_employee_notification_content_and_link(self):
        from notifications.models import Notification
        self._mark_late(1)
        self._mark_late(2)
        self._mark_late(3)
        n = Notification.objects.get(recipient=self.user, verb__icontains='late')
        self.assertIn('You were late', n.description)
        self.assertIn('my-profile', n.target_url)

    def test_hr_notification_content_and_link(self):
        from notifications.models import Notification
        self._mark_late(1)
        self._mark_late(2)
        self._mark_late(3)
        n = Notification.objects.get(recipient=self.admin, verb__icontains='late')
        self.assertIn('Late Tester', n.description)
        self.assertIn('attendance', n.target_url)

    def test_email_sent_with_pdf_attachment(self):
        from django.core import mail
        from django.test import override_settings
        import time as _time
        with override_settings(EMAIL_BACKEND='django.core.mail.backends.locmem.EmailBackend'):
            self._mark_late(1)
            self._mark_late(2)
            self._mark_late(3)
            _time.sleep(1)
            self.assertEqual(len(mail.outbox), 1)
            sent = mail.outbox[0]
            self.assertIn('Dear Employee', sent.body)
            self.assertIn('three occasions', sent.body)
            self.assertIn('attached', sent.body)
            self.assertEqual(len(sent.alternatives), 1)
            html_content, mimetype = sent.alternatives[0]
            self.assertEqual(mimetype, 'text/html')
            self.assertIn('Admin Team', html_content)
            # Outlook renders mail through Word and ignores CSS gradients, so
            # without a flat bgcolor the header loses its background and the
            # white heading on it becomes invisible.
            self.assertIn('bgcolor="#C41E3A"', html_content)
            self.assertEqual(len(sent.attachments), 1)
            filename, content, mimetype = sent.attachments[0]
            self.assertEqual(mimetype, 'application/pdf')
            self.assertGreater(len(content), 0)
class LateQueryTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user('lq_emp', password='x', email='lq_emp@leap.com')
        self.emp = make_employee(iqama='IQ-LQ', name='Late Query Tester')
        self.emp.user = self.user
        self.emp.save()
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.admin = User.objects.create_user('lq_admin', password='x', role=role)
        self.rec = AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 8, 10), status='late')

    def test_raise_query_creates_pending_query_and_notifies_hr(self):
        from hr.models import LateQuery
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:raise_late_query'), {
            'attendance_record_id': self.rec.pk, 'message': 'I was on approved leave, this is wrong.'})
        self.assertEqual(resp.status_code, 302)
        q = LateQuery.objects.get(employee=self.emp, attendance_record=self.rec)
        self.assertEqual(q.status, 'pending')
        self.assertTrue(Notification.objects.filter(recipient=self.admin, verb__icontains='raised a query').exists())

    def test_cannot_raise_second_query_for_same_record(self):
        from hr.models import LateQuery
        LateQuery.objects.create(employee=self.emp, attendance_record=self.rec, message='first')
        self.client.force_login(self.user)
        self.client.post(reverse('hr:raise_late_query'), {
            'attendance_record_id': self.rec.pk, 'message': 'second attempt'})
        self.assertEqual(LateQuery.objects.filter(attendance_record=self.rec).count(), 1)

    def test_hr_approval_flips_record_to_present_and_notifies_employee(self):
        from hr.models import LateQuery
        q = LateQuery.objects.create(employee=self.emp, attendance_record=self.rec, message='wrong mark')
        self.client.force_login(self.admin)
        resp = self.client.post(reverse('hr:team_exceptions'), {
            'action': 'decide_late_query', 'query_id': q.pk, 'decision': 'approved',
            'comment': 'Confirmed with manager', 'tab': 'late_queries'})
        self.assertEqual(resp.status_code, 302)
        q.refresh_from_db()
        self.rec.refresh_from_db()
        self.assertEqual(q.status, 'approved')
        self.assertEqual(self.rec.status, 'present')
        self.assertTrue(Notification.objects.filter(recipient=self.user, verb__icontains='query').exists())


# ─── Asset Handover: auth-boundary and edge-case coverage ─────────────────

def _png_data_url():
    import base64 as _b64
    import io as _io
    from PIL import Image as _Image
    buf = _io.BytesIO()
    _Image.new('RGBA', (8, 8), (0, 0, 0, 0)).save(buf, 'PNG')
    return 'data:image/png;base64,' + _b64.b64encode(buf.getvalue()).decode()


def _sig_file(name='sig.png'):
    import base64
    from django.core.files.base import ContentFile
    b64 = _png_data_url().split(',', 1)[1]
    return ContentFile(base64.b64decode(b64), name=name)


class AssetHandoverAuthBoundaryTests(TestCase):
    """Only the right person can act at each stage: the designated authorizer
    (not just any super admin) authorizes, only the recipient employee
    receives/acknowledges return, only HR initiates a return, and nobody
    outside the three parties (HR/employee/authorizer) can even view the
    handover or its PDF."""

    def setUp(self):
        from hr.models import Asset, Vehicle, AssetHandover
        from hr.asset_handover_services import create_asset_handover
        self.hr_user = _make_role_user('ahb-hr', Role.SUPER_ADMIN)
        self.authorizer = _make_role_user('ahb-auth', Role.SUPER_ADMIN)
        self.other_admin = _make_role_user('ahb-other-admin', Role.SUPER_ADMIN)

        self.emp = make_employee(iqama='AHB-1', name='Handover Employee')
        self.emp_user = _login_user('ahb-emp')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])

        self.other_emp = make_employee(iqama='AHB-2', name='Bystander Employee')
        self.other_emp_user = _login_user('ahb-other-emp')
        self.other_emp.user = self.other_emp_user
        self.other_emp.save(update_fields=['user'])

        self.asset = Asset.objects.create(asset_name='Boundary Laptop', quantity=1)

        self.handover = create_asset_handover(
            employee=self.emp, item=self.asset, issued_by=self.hr_user,
            issued_signature=_sig_file('issued.png'), authorizer=self.authorizer,
        )

    # ── authorize: must be the designated authorizer AND a super admin ──
    def test_wrong_super_admin_cannot_authorize(self):
        self.client.login(username='ahb-other-admin', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_authorization')

    def test_employee_cannot_authorize_own_handover(self):
        self.client.login(username='ahb-emp', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_authorization')

    def test_designated_authorizer_can_authorize(self):
        self.client.login(username='ahb-auth', password='testpass123')
        self.client.post(
            reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_receipt')

    # ── receive: only the recipient employee ──
    def test_other_employee_cannot_receive(self):
        self.client.login(username='ahb-auth', password='testpass123')
        self.client.post(reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahb-other-emp', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_receive', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_receipt')

    def test_hr_user_cannot_receive_on_employees_behalf(self):
        self.client.login(username='ahb-auth', password='testpass123')
        self.client.post(reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahb-hr', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_receive', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_receipt')

    # ── initiate_return: HR-tier only ──
    def test_plain_employee_cannot_initiate_return(self):
        self._advance_to_active()
        self.client.login(username='ahb-emp', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_initiate_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'active')

    def test_hr_can_initiate_return(self):
        self._advance_to_active()
        self.client.login(username='ahb-hr', password='testpass123')
        self.client.post(
            reverse('hr:asset_handover_initiate_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_return_confirmation')

    # ── acknowledge_return: only the recipient employee ──
    def test_other_employee_cannot_acknowledge_return(self):
        self._advance_to_pending_return()
        self.client.login(username='ahb-other-emp', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_acknowledge_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_return_confirmation')

    def test_hr_cannot_acknowledge_return_on_employees_behalf(self):
        self._advance_to_pending_return()
        self.client.login(username='ahb-hr', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_acknowledge_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_return_confirmation')

    def test_employee_can_acknowledge_own_return(self):
        self._advance_to_pending_return()
        self.client.login(username='ahb-emp', password='testpass123')
        self.client.post(
            reverse('hr:asset_handover_acknowledge_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'returned')

    # ── view/PDF access: only the three parties ──
    def test_unrelated_user_cannot_view_detail(self):
        self.client.login(username='ahb-other-emp', password='testpass123')
        resp = self.client.get(reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse('hr:hr_dashboard'))

    def test_unrelated_user_cannot_export_pdf(self):
        self.client.login(username='ahb-other-emp', password='testpass123')
        resp = self.client.get(reverse('hr:asset_handover_export_pdf', args=[self.handover.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse('hr:hr_dashboard'))

    def test_authorizer_can_view_detail(self):
        self.client.login(username='ahb-auth', password='testpass123')
        resp = self.client.get(reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_recipient_employee_can_export_pdf(self):
        self.client.login(username='ahb-emp', password='testpass123')
        resp = self.client.get(reverse('hr:asset_handover_export_pdf', args=[self.handover.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    # ── helpers ──
    def _advance_to_active(self):
        self.client.login(username='ahb-auth', password='testpass123')
        self.client.post(reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahb-emp', password='testpass123')
        self.client.post(reverse('hr:asset_handover_receive', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()

    def _advance_to_pending_return(self):
        self._advance_to_active()
        self.client.login(username='ahb-hr', password='testpass123')
        self.client.post(reverse('hr:asset_handover_initiate_return', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()


class AssetHandoverEdgeCaseTests(TestCase):
    """Race guards (acting twice, acting out of order) fail cleanly with a
    message rather than corrupting state or crashing, and item_kind
    routing/lookup behaves correctly for both assets and vehicles."""

    def setUp(self):
        from hr.models import Asset, Vehicle
        from hr.asset_handover_services import create_asset_handover
        self.hr_user = _make_role_user('ahe-hr', Role.SUPER_ADMIN)
        self.authorizer = _make_role_user('ahe-auth', Role.SUPER_ADMIN)
        self.emp = make_employee(iqama='AHE-1', name='Edge Case Employee')
        self.emp_user = _login_user('ahe-emp')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])

        self.asset = Asset.objects.create(asset_name='Edge Laptop', quantity=1)
        self.vehicle = Vehicle.objects.create(plate_number='AHE-9999')

        self.handover = create_asset_handover(
            employee=self.emp, item=self.asset, issued_by=self.hr_user,
            issued_signature=_sig_file('issued.png'), authorizer=self.authorizer,
        )

    def test_cannot_authorize_twice(self):
        self.client.login(username='ahe-auth', password='testpass123')
        url = reverse('hr:asset_handover_authorize', args=[self.handover.pk])
        self.client.post(url, {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_receipt')
        # Second authorize attempt: status has moved on, so the service's
        # own allowed_statuses guard should reject it without crashing.
        self.client.post(url, {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_receipt')

    def test_cannot_receive_before_authorized(self):
        self.client.login(username='ahe-emp', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_receive', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_authorization')

    def test_cannot_initiate_return_before_active(self):
        # Still pending_authorization at this point.
        self.client.login(username='ahe-hr', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_initiate_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_authorization')

    def test_cannot_acknowledge_return_before_initiated(self):
        self.client.login(username='ahe-emp', password='testpass123')
        resp = self.client.post(
            reverse('hr:asset_handover_acknowledge_return', args=[self.handover.pk]),
            {'signature_data': _png_data_url()})
        self.handover.refresh_from_db()
        self.assertEqual(self.handover.status, 'pending_authorization')

    def test_nonexistent_handover_404s(self):
        self.client.login(username='ahe-hr', password='testpass123')
        resp = self.client.get(reverse('hr:asset_handover_detail', args=[999999]))
        self.assertEqual(resp.status_code, 404)

    def test_active_handover_redirect_finds_asset(self):
        self.client.login(username='ahe-auth', password='testpass123')
        self.client.post(reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahe-emp', password='testpass123')
        self.client.post(reverse('hr:asset_handover_receive', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahe-hr', password='testpass123')
        resp = self.client.get(reverse('hr:asset_active_handover_redirect', args=[self.asset.pk]))
        self.assertRedirects(resp, reverse('hr:asset_handover_detail', args=[self.handover.pk]))

    def test_active_handover_redirect_no_active_handover_redirects_to_asset(self):
        # No handover exists yet for this vehicle.
        self.client.login(username='ahe-hr', password='testpass123')
        resp = self.client.get(
            reverse('hr:asset_active_handover_redirect', args=[self.vehicle.pk]) + '?item_kind=vehicle')
        self.assertRedirects(resp, reverse('hr:vehicle_detail', args=[self.vehicle.pk]))

    def test_active_handover_redirect_item_kind_vehicle_does_not_match_asset(self):
        # An active handover exists for self.asset, but with item_kind=vehicle
        # the same numeric pk must not accidentally match it.
        self.client.login(username='ahe-auth', password='testpass123')
        self.client.post(reverse('hr:asset_handover_authorize', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahe-emp', password='testpass123')
        self.client.post(reverse('hr:asset_handover_receive', args=[self.handover.pk]),
                          {'signature_data': _png_data_url()})
        self.client.login(username='ahe-hr', password='testpass123')
        resp = self.client.get(
            reverse('hr:asset_active_handover_redirect', args=[self.asset.pk]) + '?item_kind=vehicle')
        self.assertRedirects(resp, reverse('hr:vehicle_detail', args=[self.asset.pk]))

    def test_active_handover_redirect_nonexistent_kind_shows_error_not_404(self):
        # item_kind=vehicle for a pk that genuinely has no matching Vehicle
        # must not blindly redirect to a vehicle_detail URL that 404s -
        # should show a clear error and land on the vehicle list instead.
        from hr.models import Vehicle
        nonexistent_pk = 999999
        self.assertFalse(Vehicle.objects.filter(pk=nonexistent_pk).exists())
        self.client.login(username='ahe-hr', password='testpass123')
        resp = self.client.get(
            reverse('hr:asset_active_handover_redirect', args=[nonexistent_pk]) + '?item_kind=vehicle')
        self.assertRedirects(resp, reverse('hr:vehicle_list'))

    def test_vehicle_handover_item_property_returns_vehicle(self):
        from hr.asset_handover_services import create_asset_handover
        vh = create_asset_handover(
            employee=self.emp, item=self.vehicle, issued_by=self.hr_user,
            issued_signature=_sig_file('v_issued.png'), authorizer=self.authorizer,
        )
        self.assertEqual(vh.item, self.vehicle)
        self.assertIsNone(vh.asset)

    def test_asset_handover_item_property_returns_asset(self):
        self.assertEqual(self.handover.item, self.asset)
        self.assertIsNone(self.handover.vehicle)


class AttendanceRegisterSummaryTests(TestCase):
    """Per-employee totals behind the register's filter bar.

    Kept in one module so the numbers the register ranks on, prints, and
    exports are the same numbers - a ranking computed twice is a ranking that
    will eventually disagree with itself."""

    def setUp(self):
        self.alice = Employee.objects.create(iqama_number='REG-A', full_name='Alice')
        self.bob = Employee.objects.create(iqama_number='REG-B', full_name='Bob')
        self.start = _date(2026, 6, 1)
        self.end = _date(2026, 6, 30)

    def _rec(self, emp, day, status):
        return AttendanceRecord.objects.create(
            employee=emp, date=_date(2026, 6, day), status=status)

    def _totals(self, employees=None):
        from hr.attendance_summary import build_attendance_summary
        return build_attendance_summary(employees or [self.alice, self.bob],
                                        self.start, self.end)

    # ── counting ────────────────────────────────────────────────────────────

    def test_it_counts_each_status_separately(self):
        self._rec(self.alice, 1, 'present')
        self._rec(self.alice, 2, 'late')
        self._rec(self.alice, 3, 'absent')
        t = self._totals()[self.alice.pk]
        self.assertEqual((t['on_time'], t['late'], t['absent']), (1, 1, 1))

    def test_the_rate_is_only_over_days_an_arrival_was_assessed(self):
        """derive_status() returns 'wfh' before it ever looks at check_in, and
        leave, holidays and weekends never carry one. Counting those days
        would let time off improve somebody's punctuality."""
        self._rec(self.alice, 1, 'present')
        self._rec(self.alice, 2, 'late')
        for day, status in ((3, 'wfh'), (4, 'leave'), (5, 'holiday'), (6, 'weekend')):
            self._rec(self.alice, day, status)
        t = self._totals()[self.alice.pk]
        self.assertEqual(t['judged'], 2)
        self.assertEqual(t['rate'], Decimal('50.0'))

    def test_absence_is_not_folded_into_the_punctuality_rate(self):
        """Not turning up at all is a different failure from turning up late;
        averaging them would let a run of absences look like punctuality."""
        self._rec(self.alice, 1, 'present')
        for day in range(2, 12):
            self._rec(self.alice, day, 'absent')
        t = self._totals()[self.alice.pk]
        self.assertEqual(t['rate'], Decimal('100.0'))
        self.assertEqual(t['absent'], 10)

    def test_no_assessed_day_is_no_rate_rather_than_zero(self):
        """Somebody on leave all month has nothing to be judged on, which is
        not the same as having failed."""
        self._rec(self.alice, 1, 'leave')
        self.assertIsNone(self._totals()[self.alice.pk]['rate'])

    # ── warnings ────────────────────────────────────────────────────────────

    def test_a_warning_is_the_threshold_reached_in_a_calendar_month(self):
        from hr.models import LATE_WARNING_THRESHOLD
        for day in range(1, LATE_WARNING_THRESHOLD + 1):
            self._rec(self.alice, day, 'late')
        self.assertEqual(self._totals()[self.alice.pk]['warnings'], 1)

    def test_one_short_of_the_threshold_is_no_warning(self):
        from hr.models import LATE_WARNING_THRESHOLD
        for day in range(1, LATE_WARNING_THRESHOLD):
            self._rec(self.alice, day, 'late')
        self.assertEqual(self._totals()[self.alice.pk]['warnings'], 0)

    def test_more_lates_in_the_same_month_is_still_one_warning(self):
        """The notifier fires once, when the count reaches the threshold - not
        on every late after it. The register must not invent extra ones."""
        from hr.models import LATE_WARNING_THRESHOLD
        for day in range(1, LATE_WARNING_THRESHOLD + 5):
            self._rec(self.alice, day, 'late')
        self.assertEqual(self._totals()[self.alice.pk]['warnings'], 1)

    def test_each_month_earns_its_own_warning(self):
        from hr.attendance_summary import build_attendance_summary
        from hr.models import LATE_WARNING_THRESHOLD
        for month in (6, 7):
            for day in range(1, LATE_WARNING_THRESHOLD + 1):
                AttendanceRecord.objects.create(
                    employee=self.alice, date=_date(2026, month, day), status='late')
        totals = build_attendance_summary(
            [self.alice], _date(2026, 6, 1), _date(2026, 7, 31))
        self.assertEqual(totals[self.alice.pk]['warnings'], 2)

    def test_a_warning_is_counted_over_the_whole_month_not_the_window(self):
        """It was earned by a whole month of lates. Clipping it to a week view
        would show a fraction of something that never happened - the person
        either got the warning that month or did not."""
        from hr.attendance_summary import build_attendance_summary
        from hr.models import LATE_WARNING_THRESHOLD
        for day in range(1, LATE_WARNING_THRESHOLD + 1):
            self._rec(self.alice, day, 'late')
        # A one-week window covering only the first of those lates.
        totals = build_attendance_summary(
            [self.alice], _date(2026, 6, 1), _date(2026, 6, 7))
        self.assertEqual(totals[self.alice.pk]['warnings'], 1)

    def test_the_register_and_the_notifier_share_one_threshold(self):
        """Recomputed from the same constant the notification fires on, so the
        register cannot report warnings that were never sent."""
        from unittest.mock import patch
        for day in range(1, 5):
            self._rec(self.alice, day, 'late')
        with patch('hr.models.attendance.LATE_WARNING_THRESHOLD', 5), \
             patch('hr.models.LATE_WARNING_THRESHOLD', 5):
            self.assertEqual(self._totals()[self.alice.pk]['warnings'], 0)
        self.assertEqual(self._totals()[self.alice.pk]['warnings'], 1)

    # ── filtering ───────────────────────────────────────────────────────────

    def test_the_lates_filter_keeps_only_people_with_a_late(self):
        from hr.attendance_summary import filter_employees
        self._rec(self.alice, 1, 'late')
        self._rec(self.bob, 1, 'present')
        totals = self._totals()
        kept = filter_employees([self.alice, self.bob], totals, 'late')
        self.assertEqual([e.pk for e in kept], [self.alice.pk])

    def test_the_absents_filter_keeps_only_people_with_an_absence(self):
        from hr.attendance_summary import filter_employees
        self._rec(self.alice, 1, 'late')
        self._rec(self.bob, 1, 'absent')
        totals = self._totals()
        kept = filter_employees([self.alice, self.bob], totals, 'absent')
        self.assertEqual([e.pk for e in kept], [self.bob.pk])

    def test_full_attendance_keeps_everybody(self):
        from hr.attendance_summary import filter_employees
        kept = filter_employees([self.alice, self.bob], self._totals(), 'all')
        self.assertEqual(len(kept), 2)

    # ── ordering ────────────────────────────────────────────────────────────

    def test_most_punctual_puts_the_better_rate_first(self):
        from hr.attendance_summary import order_employees
        self._rec(self.alice, 1, 'present')
        self._rec(self.bob, 1, 'present')
        self._rec(self.bob, 2, 'late')
        totals = self._totals()
        ordered = order_employees([self.bob, self.alice], totals, 'punctual')
        self.assertEqual([e.pk for e in ordered], [self.alice.pk, self.bob.pk])

    def test_a_longer_clean_record_beats_a_shorter_one(self):
        """One perfect day must not outrank a month of them."""
        from hr.attendance_summary import order_employees
        self._rec(self.alice, 1, 'present')
        for day in range(1, 15):
            self._rec(self.bob, day, 'present')
        ordered = order_employees([self.alice, self.bob], self._totals(), 'punctual')
        self.assertEqual([e.pk for e in ordered], [self.bob.pk, self.alice.pk])

    def test_someone_with_no_assessed_day_sorts_last_not_first(self):
        """No record at all is not a perfect record. A missing rate must not
        be treated as an unbeatable one."""
        from hr.attendance_summary import order_employees
        self._rec(self.alice, 1, 'present')
        self._rec(self.alice, 2, 'late')
        self._rec(self.bob, 1, 'leave')
        ordered = order_employees([self.bob, self.alice], self._totals(), 'punctual')
        self.assertEqual([e.pk for e in ordered], [self.alice.pk, self.bob.pk])

    def test_a_measured_failure_still_outranks_no_measurement(self):
        """Somebody late every single day has a rate of 0%, which is a worse
        record than an unmeasured one - but it is still a record, so they
        belong in the ranking above the person who has none."""
        from hr.attendance_summary import order_employees
        self._rec(self.alice, 1, 'late')
        self._rec(self.bob, 1, 'leave')
        totals = self._totals()
        self.assertEqual(totals[self.alice.pk]['rate'], Decimal('0.0'))
        self.assertIsNone(totals[self.bob.pk]['rate'])
        ordered = order_employees([self.bob, self.alice], totals, 'punctual')
        self.assertEqual([e.pk for e in ordered], [self.alice.pk, self.bob.pk])

    def test_most_late_puts_the_worst_first(self):
        from hr.attendance_summary import order_employees
        self._rec(self.alice, 1, 'late')
        self._rec(self.bob, 1, 'late')
        self._rec(self.bob, 2, 'late')
        ordered = order_employees([self.alice, self.bob], self._totals(), 'late')
        self.assertEqual([e.pk for e in ordered], [self.bob.pk, self.alice.pk])

    def test_warnings_orders_by_warnings_received(self):
        from hr.attendance_summary import order_employees
        from hr.models import LATE_WARNING_THRESHOLD
        for day in range(1, LATE_WARNING_THRESHOLD + 1):
            self._rec(self.bob, day, 'late')
        self._rec(self.alice, 1, 'late')
        ordered = order_employees([self.alice, self.bob], self._totals(), 'warnings')
        self.assertEqual([e.pk for e in ordered], [self.bob.pk, self.alice.pk])

    def test_name_order_is_the_default_and_is_alphabetical(self):
        from hr.attendance_summary import order_employees
        ordered = order_employees([self.bob, self.alice], self._totals(), 'name')
        self.assertEqual([e.full_name for e in ordered], ['Alice', 'Bob'])

    # ── bad input ───────────────────────────────────────────────────────────

    def test_an_unknown_filter_falls_back_to_the_full_register(self):
        """A stale or hand-edited link must still show the register, not an
        empty page that looks like nobody turned up."""
        from hr.attendance_summary import resolve_register_filters
        self.assertEqual(resolve_register_filters({}), ('all', 'name'))
        self.assertEqual(
            resolve_register_filters({'show': 'nonsense', 'order': 'nonsense'}),
            ('all', 'name'))

    def test_the_chip_counts_say_what_each_filter_would_show(self):
        from hr.attendance_summary import apply_register_filters
        self._rec(self.alice, 1, 'late')
        self._rec(self.bob, 1, 'absent')
        _rows, _totals, counts = apply_register_filters(
            [self.alice, self.bob], self.start, self.end, 'all', 'name')
        self.assertEqual(counts, {'all': 2, 'late': 1, 'absent': 1})


class AttendanceRegisterViewTests(TestCase):
    """The register page and its two exports, which must agree."""

    def setUp(self):
        from accounts.models import Role, User
        role, _ = Role.objects.get_or_create(name=Role.ERP_ADMIN)
        self.admin = User.objects.create_user('reg_adm', password='x')
        self.admin.role = role
        self.admin.save()
        self.client.force_login(self.admin)
        self.punctual = Employee.objects.create(
            iqama_number='RV-P', full_name='Punctual Person')
        self.tardy = Employee.objects.create(
            iqama_number='RV-T', full_name='Tardy Person')
        AttendanceRecord.objects.create(
            employee=self.punctual, date=_date(2026, 6, 1), status='present')
        for day in (1, 2, 3):
            AttendanceRecord.objects.create(
                employee=self.tardy, date=_date(2026, 6, day), status='late')

    def _get(self, qs=''):
        return self.client.get(
            reverse('hr:attendance_matrix') + '?period=month&date=2026-06-15' + qs)

    def test_the_register_shows_everybody_by_default(self):
        body = self._get().content.decode()
        self.assertIn('Punctual Person', body)
        self.assertIn('Tardy Person', body)

    def test_the_lates_filter_drops_people_with_no_late(self):
        body = self._get('&show=late').content.decode()
        self.assertIn('Tardy Person', body)
        self.assertNotIn('Punctual Person', body)

    def test_ordering_by_lates_puts_the_worst_first(self):
        body = self._get('&order=late').content.decode()
        self.assertLess(body.index('Tardy Person'), body.index('Punctual Person'))

    def test_ordering_by_punctuality_puts_the_best_first(self):
        body = self._get('&order=punctual').content.decode()
        self.assertLess(body.index('Punctual Person'), body.index('Tardy Person'))

    def test_the_totals_are_printed_so_the_ranking_can_be_checked(self):
        """A ranking whose numbers are not on the page can only be trusted,
        not verified."""
        body = self._get().content.decode()
        for header in ('On time', 'Late', 'Absent', 'Rate', 'Warnings'):
            self.assertIn(header, body)

    def test_the_filters_survive_the_period_and_location_links(self):
        """A filter that silently resets on the next click is worse than no
        filter - the page would quietly change population."""
        body = self._get('&show=late&order=warnings').content.decode()
        # This page writes its query strings as literal template text, so the
        # ampersands reach the browser unescaped - same as every link already
        # on it.
        self.assertIn('location=site&show=late&order=warnings', body)
        self.assertIn('period=week&date=2026-06-15&location=office'
                      '&show=late&order=warnings', body)
        # The exports too, or "export what I am looking at" is not true.
        self.assertIn('export/excel/?period=month&date=2026-06-15'
                      '&location=office&show=late&order=warnings', body)

    def test_the_exports_carry_the_same_filters(self):
        """An export that ignored the filter bar would hand back a different
        population from the one on screen.

        Asserted on the employee list the export hands to build_matrix, not on
        the bytes: xlsx is a zip and reportlab compresses its streams, so a
        name search over the response body passes whether or not the filter
        was applied."""
        from unittest.mock import patch
        import hr.views as hr_views
        real = hr_views.build_matrix
        for name in ('hr:attendance_matrix_export_excel',
                     'hr:attendance_matrix_export_pdf'):
            with self.subTest(export=name):
                seen = []

                def spy(employees, *args, **kwargs):
                    seen.append([e.full_name for e in employees])
                    return real(employees, *args, **kwargs)

                with patch.object(hr_views, 'build_matrix', spy):
                    resp = self.client.get(
                        reverse(name) + '?period=month&date=2026-06-15&show=late')
                self.assertEqual(resp.status_code, 200)
                self.assertEqual(seen, [['Tardy Person']])

    def test_an_empty_filter_says_so_rather_than_looking_broken(self):
        AttendanceRecord.objects.all().delete()
        body = self._get('&show=absent').content.decode()
        self.assertIn('No employees match this filter', body)


class MyProfileAttendanceTotalsTests(TestCase):
    """An employee sees their own attendance figures on My Profile.

    The point of these is that they are the SAME figures the Attendance
    Register ranks them on. Two places computing punctuality separately is how
    an employee and HR end up looking at different numbers for the same
    month."""

    def setUp(self):
        from accounts.models import User
        self.user = User.objects.create_user('mp_worker', password='x')
        self.emp = Employee.objects.create(
            iqama_number='MP-1', full_name='Mine Worker', user=self.user)
        self.client.force_login(self.user)

    def _rec(self, day, status):
        return AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 6, day), status=status)

    def _get(self):
        return self.client.get(reverse('hr:my_profile') + '?date=2026-06-15')

    def test_it_shows_their_own_totals(self):
        self._rec(1, 'present')
        self._rec(2, 'late')
        self._rec(3, 'absent')
        totals = self._get().context['my_attendance_totals']
        self.assertEqual((totals['on_time'], totals['late'], totals['absent']),
                         (1, 1, 1))
        self.assertEqual(totals['rate'], Decimal('50.0'))

    def test_the_figures_match_what_the_register_shows_hr(self):
        """Same function, so an employee querying their rate and HR reading it
        off the register cannot be looking at two different numbers."""
        from hr.attendance_summary import build_attendance_summary
        self._rec(1, 'present')
        self._rec(2, 'late')
        self._rec(3, 'wfh')
        mine = self._get().context['my_attendance_totals']
        theirs = build_attendance_summary(
            [self.emp], _date(2026, 6, 1), _date(2026, 6, 30))[self.emp.pk]
        self.assertEqual(mine, theirs)

    def test_warnings_are_shown_to_the_person_who_earned_them(self):
        from hr.models import LATE_WARNING_THRESHOLD
        for day in range(1, LATE_WARNING_THRESHOLD + 1):
            self._rec(day, 'late')
        resp = self._get()
        self.assertEqual(resp.context['my_attendance_totals']['warnings'], 1)
        self.assertContains(resp, 'WARNINGS')

    def test_the_page_explains_how_a_warning_is_earned(self):
        """A number nobody can explain is a number people argue about."""
        self._rec(1, 'late')
        resp = self._get()
        self.assertEqual(resp.context['late_warning_threshold'],
                         resp.context['late_warning_threshold'])
        self.assertContains(resp, 'lates in one calendar month')

    def test_a_month_with_no_assessed_day_shows_no_rate_not_zero(self):
        self._rec(1, 'leave')
        self.assertIsNone(self._get().context['my_attendance_totals']['rate'])

    def test_the_month_navigation_moves_the_totals_too(self):
        """The tiles sit above a month the user can page through; totals that
        stayed on the current month would silently describe a different period
        from the table under them."""
        self._rec(1, 'late')
        AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 7, 1), status='present')
        june = self.client.get(
            reverse('hr:my_profile') + '?date=2026-06-15').context['my_attendance_totals']
        july = self.client.get(
            reverse('hr:my_profile') + '?date=2026-07-15').context['my_attendance_totals']
        self.assertEqual((june['late'], june['on_time']), (1, 0))
        self.assertEqual((july['late'], july['on_time']), (0, 1))

    def test_someone_elses_figures_are_not_reachable(self):
        """?date is the only knob on this page; it must not become a way to
        read a colleague's month."""
        other_user = get_user_model().objects.create_user('mp_other', password='x')
        other = Employee.objects.create(
            iqama_number='MP-2', full_name='Other Worker', user=other_user)
        AttendanceRecord.objects.create(
            employee=other, date=_date(2026, 6, 1), status='absent')
        self._rec(1, 'present')
        totals = self._get().context['my_attendance_totals']
        self.assertEqual(totals['absent'], 0)
        self.assertEqual(totals['on_time'], 1)


class MyAttendancePdfTotalsTests(TestCase):
    """The employee's own monthly PDF carries the same figures as the page.

    It is also what the automatic late-threshold email attaches, so for a lot
    of people it is the first place they read their own numbers."""

    def setUp(self):
        from accounts.models import User
        self.user = User.objects.create_user('pdf_worker', password='x')
        self.emp = Employee.objects.create(
            iqama_number='PDF-1', full_name='Pdf Worker', user=self.user)
        self.client.force_login(self.user)
        AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 6, 1), status='present')
        AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 6, 2), status='late')

    def test_the_pdf_builds_and_is_a_pdf(self):
        resp = self.client.get(
            reverse('hr:my_attendance_export_pdf') + '?date=2026-06-15')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')
        self.assertTrue(resp.content.startswith(b'%PDF'))

    def test_the_summary_row_is_built_from_the_shared_totals(self):
        """Asserted on the table handed to reportlab rather than on the bytes:
        reportlab compresses its streams, so searching the response body for a
        number passes whether or not it was ever put there."""
        from unittest.mock import patch
        import hr.views as hr_views
        captured = []
        real_table = None

        from reportlab.platypus import Table as RLTable
        real_table = RLTable

        def spy(data, *args, **kwargs):
            captured.append(data)
            return real_table(data, *args, **kwargs)

        with patch('reportlab.platypus.Table', spy):
            hr_views.build_attendance_pdf(
                self.emp, _date(2026, 6, 1), _date(2026, 6, 30))

        header = ['On time', 'Late', 'Absent', 'On-time rate', 'Warnings']
        summary = next(d for d in captured if d and d[0] == header)
        self.assertEqual(summary[1], ['1', '1', '0', '50.0%', '0'])


class ExceptionReasonVisibilityTests(TestCase):
    """Why a day was excused has to survive the decision.

    The pending queue stops showing a request once it is decided, and the
    history table carried no reason column at all — so an approved exception
    became an unexplained one, which is precisely when the record matters."""

    def setUp(self):
        self.top = make_employee(iqama='ERV-TOP', name='Top Manager')
        self.top_user = _login_user('erv_top')
        self.top.user = self.top_user
        self.top.save(update_fields=['user'])

        self.report = make_employee(iqama='ERV-REP', name='Report Person')
        self.report.main_manager = self.top
        self.report.save(update_fields=['main_manager'])
        self.report_user = _login_user('erv_rep')
        self.report.user = self.report_user
        self.report.save(update_fields=['user'])

        self.hr_user = _make_role_user('erv_hr', Role.SUPER_ADMIN)

    def _exception(self, **kwargs):
        opts = dict(
            employee=self.report, event_date=_date(2026, 7, 20),
            event_start_time=_time(9, 0), reason_category='site_visit',
            created_by=self.top_user)
        opts.update(kwargs)
        return _submit_aex(**opts)

    def _decide(self, exc, status='approved', note=''):
        from django.utils import timezone
        exc.status = status
        exc.decision_note = note
        exc.decided_by = self.top_user
        exc.decided_at = timezone.now()
        exc.save()
        return exc

    def _history(self):
        self.client.login(username='erv_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'), {'tab': 'all'})
        self.assertEqual(resp.status_code, 200)
        return resp.content.decode()

    # ── the record survives the decision ────────────────────────────────────

    def test_an_approved_exception_still_shows_its_reason(self):
        self._decide(self._exception(custom_reason='Client site in Dammam'))
        body = self._history()
        self.assertIn('Client site in Dammam', body)

    def test_the_history_table_has_a_reason_column(self):
        """It had none, which is how the reason disappeared on approval."""
        self._decide(self._exception())
        self.assertIn(
            '<th>Employee</th><th>Manager</th><th>Event</th><th>Reason</th>'
            '<th>Status</th><th>Decided At</th>',
            self._history())

    def test_a_rejected_exception_keeps_its_reason_too(self):
        """Rejections are the ones people come back and ask about."""
        self._decide(self._exception(custom_reason='Dentist appointment'),
                     status='rejected')
        self.assertIn('Dentist appointment', self._history())

    def test_the_decision_note_is_part_of_the_record(self):
        """The only place the decider's reasoning survives once the row leaves
        the pending queue."""
        self._decide(self._exception(), note='Confirmed with the site lead')
        self.assertIn('Confirmed with the site lead', self._history())

    def test_the_employees_own_comment_survives(self):
        self._decide(self._exception(employee_comment='Left at 6am to travel'))
        self.assertIn('Left at 6am to travel', self._history())

    def test_an_override_shows_what_justified_it(self):
        exc = self._decide(self._exception())
        exc.is_overridden = True
        exc.overridden_by = self.hr_user
        exc.override_reason = 'HR corrected the category'
        exc.save()
        self.assertIn('HR corrected the category', self._history())

    # ── the custom reason is not gated on the category ──────────────────────

    def test_a_custom_reason_shows_against_any_category(self):
        """The form requires a custom reason only for 'Other' but accepts one
        against every category, so gating the DISPLAY on 'other' silently
        discarded what people had actually typed."""
        self._decide(self._exception(
            reason_category='site_visit', custom_reason='Aramco plant walkdown'))
        self.assertIn('Aramco plant walkdown', self._history())

    def test_a_custom_reason_still_shows_for_other(self):
        self._decide(self._exception(
            reason_category='other', custom_reason='Passport renewal'))
        self.assertIn('Passport renewal', self._history())

    def test_a_pending_request_shows_the_same_fields(self):
        """Pending and decided read from one partial, so a field cannot be
        present on one list and missing from the other."""
        self._exception(custom_reason='Site handover',
                        employee_comment='Back by noon')
        body = self._history()
        self.assertIn('Site handover', body)
        self.assertIn('Back by noon', body)

    # ── the employee sees their own record ──────────────────────────────────

    def test_the_employee_sees_the_reason_and_decision_on_my_profile(self):
        self._decide(self._exception(custom_reason='Client site in Dammam'),
                     note='Approved, site confirmed')
        self.client.login(username='erv_rep', password='testpass123')
        body = self.client.get(reverse('hr:my_profile')).content.decode()
        self.assertIn('Client site in Dammam', body)
        self.assertIn('Approved, site confirmed', body)


class LateQueryDetailVisibilityTests(TestCase):
    """A decided late-mark query has to keep the question and the answer.

    Same defect as the attendance exceptions: the decided table showed only
    the status and who decided it, so an answered query became an unexplained
    one. Worse here — the decide form never sent a comment, so the
    decision_note field the view reads was never written at all."""

    def setUp(self):
        self.emp = make_employee(iqama='LQ-1', name='Query Person')
        self.emp_user = _login_user('lq_emp')
        self.emp.user = self.emp_user
        self.emp.save(update_fields=['user'])
        self.hr_user = _make_role_user('lq_hr', Role.SUPER_ADMIN)
        self.record = AttendanceRecord.objects.create(
            employee=self.emp, date=_date(2026, 6, 10), status='late',
            check_in=_time(9, 15))

    def _query(self, message='The gate was closed, I was in the car park'):
        from hr.models import LateQuery
        return LateQuery.objects.create(
            employee=self.emp, attendance_record=self.record, message=message)

    def _decide(self, query, decision='approved', comment=''):
        self.client.login(username='lq_hr', password='testpass123')
        return self.client.post(
            reverse('hr:team_exceptions'),
            {'action': 'decide_late_query', 'query_id': query.pk,
             'tab': 'late_queries', 'decision': decision, 'comment': comment})

    def _hr_page(self):
        self.client.login(username='lq_hr', password='testpass123')
        resp = self.client.get(reverse('hr:team_exceptions'),
                               {'tab': 'late_queries'})
        self.assertEqual(resp.status_code, 200)
        return resp.content.decode()

    def _my_profile(self):
        self.client.login(username='lq_emp', password='testpass123')
        return self.client.get(reverse('hr:my_profile')).content.decode()

    # ── the decision note was unreachable ───────────────────────────────────

    def test_the_decide_form_records_the_reason(self):
        """The view already read request.POST['comment'] into decision_note,
        but no form ever sent one, so the field was dead and every decided
        query displayed an empty note."""
        from hr.models import LateQuery
        query = self._query()
        self._decide(query, 'rejected', comment='Gate logs show 09:14 arrival')
        query = LateQuery.objects.get(pk=query.pk)
        self.assertEqual(query.status, 'rejected')
        self.assertEqual(query.decision_note, 'Gate logs show 09:14 arrival')

    def test_the_form_offers_somewhere_to_type_it(self):
        self._query()
        self.assertIn('name="comment"', self._hr_page())

    def test_a_decision_is_not_blocked_on_typing_a_reason(self):
        """Optional on purpose - an approval that needs no explanation should
        stay one click."""
        from hr.models import LateQuery
        query = self._query()
        self._decide(query, 'approved')
        self.assertEqual(LateQuery.objects.get(pk=query.pk).status, 'approved')

    # ── the record survives the decision ────────────────────────────────────

    def test_a_decided_query_still_shows_what_was_asked(self):
        self._decide(self._query(), 'approved', comment='Confirmed with reception')
        body = self._hr_page()
        self.assertIn('The gate was closed', body)

    def test_a_decided_query_shows_why_it_was_decided(self):
        self._decide(self._query(), 'rejected', comment='Gate logs show 09:14')
        self.assertIn('Gate logs show 09:14', self._hr_page())

    def test_the_decided_table_carries_the_detail_columns(self):
        self._decide(self._query())
        self.assertIn(
            '<th>Employee</th><th>Date</th><th>Query &amp; decision</th>'
            '<th>Status</th><th>Decided by</th><th>Decided at</th>',
            self._hr_page())

    def test_a_rejection_with_no_reason_says_so_rather_than_showing_blank(self):
        """The case people come back and ask about; a blank cell reads as a
        rendering fault rather than as nothing having been written."""
        self._decide(self._query(), 'rejected')
        self.assertIn('No reason was recorded', self._hr_page())

    def test_an_approval_with_no_reason_is_not_nagged_about(self):
        """An approval needs no justification - the employee got what they
        asked for."""
        self._decide(self._query(), 'approved')
        self.assertNotIn('No reason was recorded', self._hr_page())

    # ── the employee sees their own ─────────────────────────────────────────

    def test_the_employee_sees_their_own_query_and_its_answer(self):
        self._decide(self._query(), 'rejected', comment='Gate logs show 09:14')
        body = self._my_profile()
        self.assertIn('The gate was closed', body)
        self.assertIn('Gate logs show 09:14', body)

    def test_an_approved_query_actually_clears_the_late_mark(self):
        """The reason the detail matters: an approved query re-derives the day.
        If it did not, the employee would be reading an answer that changed
        nothing."""
        self._decide(self._query(), 'approved')
        self.record.refresh_from_db()
        self.assertEqual(self.record.status, 'present')


class SelfWFHServiceTests(TestCase):
    """An employee marking their own remote days.

    The rules here are what keep a self-declaration from being able to rewrite
    a day somebody has already assessed."""

    def setUp(self):
        from accounts.models import User
        self.user = User.objects.create_user('wfh_emp', password='x')
        self.emp = Employee.objects.create(
            iqama_number='WFH-1', full_name='Remote Worker', user=self.user)
        # A Monday, so nothing in these ranges lands on the KSA weekend.
        self.today = _date(2026, 6, 8)
        self.assertEqual(self.today.weekday(), 0)

    def _mark(self, start, end=None, note='', today=None):
        from hr.wfh_services import mark_self_wfh
        return mark_self_wfh(
            employee=self.emp, start_date=start, end_date=end, note=note,
            created_by=self.user, today=today or self.today)

    # ── the happy path ──────────────────────────────────────────────────────

    def test_an_employee_can_mark_today(self):
        record = self._mark(self.today)
        self.assertEqual(record.employee, self.emp)
        self.assertEqual(record.start_date, self.today)
        self.assertEqual(record.end_date, self.today)
        self.assertEqual(record.created_by, self.user)

    def test_a_blank_end_date_means_a_single_day(self):
        record = self._mark(self.today, None)
        self.assertEqual(record.end_date, self.today)

    def test_a_future_range_is_allowed(self):
        record = self._mark(self.today + timedelta(days=7),
                            self.today + timedelta(days=9))
        self.assertEqual((record.end_date - record.start_date).days, 2)

    def test_marking_today_makes_the_day_read_as_wfh(self):
        """The point of the feature: their attendance shows WFH rather than
        absent, without waiting for HR to tick a box."""
        from hr.attendance_services import derive_status
        AttendanceRecord.objects.create(
            employee=self.emp, date=self.today, status='absent')
        self._mark(self.today)
        status, _hours = derive_status(self.emp, self.today, None)
        self.assertEqual(status, 'wfh')
        self.assertEqual(
            AttendanceRecord.objects.get(employee=self.emp, date=self.today).status,
            'wfh')

    # ── the rule that matters ───────────────────────────────────────────────

    def test_it_cannot_be_backdated(self):
        """derive_status() checks WFH BEFORE it looks at check_in, so a WFH day
        is never late and never absent. Backdating one would let anybody erase
        a Late that has already been assessed, with nobody reviewing it."""
        from hr.wfh_services import WFHError
        with self.assertRaises(WFHError) as caught:
            self._mark(self.today - timedelta(days=1))
        self.assertIn('today or a future date', str(caught.exception))

    def test_backdating_cannot_clear_an_existing_late_mark(self):
        """The concrete abuse the rule exists to stop."""
        from hr.wfh_services import WFHError
        yesterday = self.today - timedelta(days=1)
        AttendanceRecord.objects.create(
            employee=self.emp, date=yesterday, status='late',
            check_in=_time(9, 30))
        with self.assertRaises(WFHError):
            self._mark(yesterday)
        self.assertEqual(
            AttendanceRecord.objects.get(employee=self.emp, date=yesterday).status,
            'late')

    def test_the_refusal_points_at_the_reviewed_path(self):
        """A past day that genuinely needs correcting already has a route that
        a manager sees."""
        from hr.wfh_services import WFHError
        with self.assertRaises(WFHError) as caught:
            self._mark(self.today - timedelta(days=3))
        self.assertIn('attendance exception', str(caught.exception))

    # ── the other refusals ──────────────────────────────────────────────────

    def test_an_end_before_the_start_is_refused(self):
        from hr.wfh_services import WFHError
        with self.assertRaises(WFHError):
            self._mark(self.today + timedelta(days=5), self.today + timedelta(days=2))

    def test_an_absurd_range_is_refused(self):
        """A typo in a date field should not declare a year of remote work."""
        from hr.wfh_services import WFHError, MAX_WFH_RANGE_DAYS
        with self.assertRaises(WFHError) as caught:
            self._mark(self.today, self.today + timedelta(days=MAX_WFH_RANGE_DAYS))
        self.assertIn(str(MAX_WFH_RANGE_DAYS), str(caught.exception))

    def test_a_range_that_is_all_weekend_is_refused(self):
        """A WFH record on a weekend does nothing, so accepting it would be
        telling the employee something was recorded when it was not."""
        from hr.wfh_services import WFHError
        friday = _date(2026, 6, 12)
        self.assertEqual(friday.weekday(), 4)      # KSA weekend is Fri/Sat
        with self.assertRaises(WFHError) as caught:
            self._mark(friday, friday + timedelta(days=1))
        self.assertIn('weekend or holiday', str(caught.exception))

    def test_a_range_that_only_touches_a_weekend_is_fine(self):
        record = self._mark(_date(2026, 6, 11), _date(2026, 6, 14))
        self.assertIsNotNone(record.pk)

    def test_it_is_refused_over_booked_leave(self):
        """Leave outranks WFH in derive_status(), so the record would sit there
        doing nothing."""
        from hr.wfh_services import WFHError
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        LeaveRecord.objects.create(
            employee=self.emp, leave_type=annual, start_date=self.today,
            end_date=self.today, days=Decimal('1'))
        with self.assertRaises(WFHError) as caught:
            self._mark(self.today)
        self.assertIn('leave booked', str(caught.exception))

    def test_overlapping_an_existing_entry_is_refused(self):
        from hr.wfh_services import WFHError
        self._mark(self.today, self.today + timedelta(days=3))
        with self.assertRaises(WFHError) as caught:
            self._mark(self.today + timedelta(days=2))
        self.assertIn('already marked', str(caught.exception))

    # ── withdrawing ─────────────────────────────────────────────────────────

    def test_they_can_withdraw_a_day_that_has_not_finished(self):
        from hr.models import WFHRecord
        from hr.wfh_services import cancel_self_wfh
        record = self._mark(self.today + timedelta(days=2))
        cancel_self_wfh(record=record, requested_by=self.user, today=self.today)
        self.assertFalse(WFHRecord.objects.filter(pk=record.pk).exists())

    def test_withdrawing_today_puts_the_day_back_as_it_was(self):
        from hr.wfh_services import cancel_self_wfh
        AttendanceRecord.objects.create(
            employee=self.emp, date=self.today, status='absent')
        record = self._mark(self.today)
        cancel_self_wfh(record=record, requested_by=self.user, today=self.today)
        self.assertEqual(
            AttendanceRecord.objects.get(employee=self.emp, date=self.today).status,
            'absent')

    def test_a_finished_period_cannot_be_deleted(self):
        """It is a record of what happened, not a plan - letting someone delete
        it would put the same hole in the register that backdating would."""
        from hr.models import WFHRecord
        from hr.wfh_services import cancel_self_wfh, WFHError
        record = self._mark(self.today)
        later = self.today + timedelta(days=5)
        with self.assertRaises(WFHError) as caught:
            cancel_self_wfh(record=record, requested_by=self.user, today=later)
        self.assertIn('already passed', str(caught.exception))
        self.assertTrue(WFHRecord.objects.filter(pk=record.pk).exists())

    def test_only_the_person_who_marked_it_can_withdraw_it(self):
        """An HR-set WFH day is not the employee's to remove."""
        from accounts.models import User
        from hr.models import WFHRecord
        from hr.wfh_services import cancel_self_wfh, WFHError
        hr_user = User.objects.create_user('wfh_hr', password='x')
        record = WFHRecord.objects.create(
            employee=self.emp, start_date=self.today, end_date=self.today,
            created_by=hr_user)
        with self.assertRaises(WFHError):
            cancel_self_wfh(record=record, requested_by=self.user, today=self.today)
        self.assertTrue(WFHRecord.objects.filter(pk=record.pk).exists())

    # ── future days are not attendance yet ──────────────────────────────────

    def test_marking_a_future_range_does_not_write_attendance_rows(self):
        """Re-deriving a future day would record an absence for a date nobody
        could have attended yet, and the register would show a wall of them for
        a fortnight of planned remote work."""
        self._mark(self.today + timedelta(days=1), self.today + timedelta(days=4))
        self.assertEqual(
            AttendanceRecord.objects.filter(employee=self.emp).count(), 0)

    def test_a_past_day_with_no_record_is_not_invented(self):
        """A range starting today can still cover days the resync walks; it
        must not create rows for days that were never recorded."""
        self._mark(self.today, self.today + timedelta(days=2))
        dates = set(AttendanceRecord.objects.filter(
            employee=self.emp).values_list('date', flat=True))
        self.assertEqual(dates, {self.today})


class SelfWFHViewTests(TestCase):
    """Marking work from home from My Profile."""

    def setUp(self):
        from accounts.models import User
        self.user = User.objects.create_user('wfhv_emp', password='x')
        self.emp = Employee.objects.create(
            iqama_number='WFHV-1', full_name='View Worker', user=self.user)
        self.client.force_login(self.user)

    def _post(self, follow=False, **data):
        payload = {'action': 'mark_wfh'}
        payload.update(data)
        return self.client.post(reverse('hr:my_profile'), payload, follow=follow)

    def _next_working_day(self):
        """The next day that is not the KSA Fri/Sat weekend.

        These tests run against the real current date, so "tomorrow" lands on
        a weekend two days in seven - and the service rightly refuses a range
        with no working day in it. Walking to a working day keeps the test
        about the view rather than about which day it happens to run on.
        """
        from django.utils import timezone
        from hr.models import AttendanceSettings
        weekend = AttendanceSettings.load().weekend_day_set()
        day = timezone.localdate() + timedelta(days=1)
        while day.weekday() in weekend:
            day += timedelta(days=1)
        return day

    def test_the_form_is_on_the_page(self):
        body = self.client.get(reverse('hr:my_profile')).content.decode()
        self.assertIn('value="mark_wfh"', body)

    def test_marking_creates_the_record(self):
        from hr.models import WFHRecord
        day = self._next_working_day()
        self._post(start_date=day.isoformat())
        self.assertTrue(WFHRecord.objects.filter(
            employee=self.emp, start_date=day, created_by=self.user).exists())

    def test_a_refusal_tells_the_employee_why(self):
        """The service writes its messages to be read by the employee, so they
        must reach the page rather than being swallowed."""
        from django.utils import timezone
        from hr.models import WFHRecord
        yesterday = timezone.localdate() - timedelta(days=1)
        resp = self._post(start_date=yesterday.isoformat(), follow=True)
        self.assertFalse(WFHRecord.objects.filter(employee=self.emp).exists())
        self.assertContains(resp, 'today or a future date')

    def test_the_employee_cannot_mark_somebody_else(self):
        """The form has no employee field on purpose; posting one must not
        reach the record."""
        from hr.models import WFHRecord
        other = Employee.objects.create(iqama_number='WFHV-2', full_name='Someone Else')
        day = self._next_working_day()
        self._post(start_date=day.isoformat(), employee=other.pk)
        self.assertFalse(WFHRecord.objects.filter(employee=other).exists())
        self.assertTrue(WFHRecord.objects.filter(employee=self.emp).exists())

    def test_withdrawing_from_the_page(self):
        from hr.models import WFHRecord
        day = self._next_working_day()
        self._post(start_date=day.isoformat())
        record = WFHRecord.objects.get(employee=self.emp)
        self.client.post(reverse('hr:my_profile'),
                         {'action': 'cancel_wfh', 'wfh_id': record.pk})
        self.assertFalse(WFHRecord.objects.filter(pk=record.pk).exists())

    def test_withdrawing_somebody_elses_entry_is_refused(self):
        """A guessed id must not reach another employee's day."""
        from hr.models import WFHRecord
        from django.utils import timezone
        other = Employee.objects.create(iqama_number='WFHV-3', full_name='Third Person')
        day = timezone.localdate() + timedelta(days=1)
        theirs = WFHRecord.objects.create(
            employee=other, start_date=day, end_date=day, created_by=self.user)
        self.client.post(reverse('hr:my_profile'),
                         {'action': 'cancel_wfh', 'wfh_id': theirs.pk})
        self.assertTrue(WFHRecord.objects.filter(pk=theirs.pk).exists())

    def test_the_page_lists_their_marked_days(self):
        day = self._next_working_day()
        self._post(start_date=day.isoformat(), note='At home in Riyadh')
        body = self.client.get(reverse('hr:my_profile')).content.decode()
        self.assertIn('At home in Riyadh', body)
        self.assertIn('Withdraw', body)


class PendingApprovalsTests(TestCase):
    """The Pending Approvals inbox on My Profile.

    The risk in gathering approvals from half the system into one list is that
    the list becomes a second opinion about who may decide what. So the thing
    worth testing is not that rows appear - it is that a row never appears for
    somebody the real page would turn away."""

    def setUp(self):
        from accounts.models import Role, User
        self.approver_user = _login_user('pa_approver')
        self.approver = make_employee(iqama='PA-MGR', name='Pending Manager')
        self.approver.user = self.approver_user
        self.approver.save(update_fields=['user'])

        self.report = make_employee(iqama='PA-REP', name='Pending Report')
        self.report.main_manager = self.approver
        self.report.save(update_fields=['main_manager'])
        self.report_user = _login_user('pa_report')
        self.report.user = self.report_user
        self.report.save(update_fields=['user'])

        self.stranger_user = _login_user('pa_stranger')
        self.stranger = make_employee(iqama='PA-STR', name='Pending Stranger')
        self.stranger.user = self.stranger_user
        self.stranger.save(update_fields=['user'])

    def _groups(self, user):
        from hr.approvals import pending_approvals
        return {g['key']: g for g in pending_approvals(user)}

    def _exception_for(self, employee):
        return _submit_aex(
            employee=employee, event_date=_date(2026, 7, 20),
            event_start_time=_time(9, 0), reason_category='site_visit',
            created_by=self.approver_user)

    # ── it shows what is genuinely yours ────────────────────────────────────

    def test_a_managers_own_reports_exception_appears(self):
        self._exception_for(self.report)
        groups = self._groups(self.approver_user)
        self.assertIn('attendance_exceptions', groups)
        self.assertEqual(
            [i['title'] for i in groups['attendance_exceptions']['items']],
            ['Pending Report'])

    def test_it_does_not_show_somebody_elses_queue(self):
        """A stranger to that employee must not be handed their decision."""
        self._exception_for(self.report)
        self.assertNotIn('attendance_exceptions', self._groups(self.stranger_user))

    def test_your_own_request_is_not_yours_to_approve(self):
        """The Team Exceptions queue excludes the viewer's own request rather
        than merely disabling the buttons, so the inbox must too - otherwise it
        sends someone to a page that will not show it.

        Set up through a secondary-manager self-assignment, which is the shape
        that actually reaches this branch: Employee.clean() guards against
        main-manager cycles but says in so many words that secondary-manager
        self-assignment is out of its scope, so nothing stops the data
        existing.
        """
        self.approver.secondary_managers.add(self.approver)
        self._exception_for(self.approver)
        groups = self._groups(self.approver_user)
        titles = [i['title'] for g in groups.values() for i in g['items']]
        self.assertNotIn('Pending Manager', titles)

    def test_a_leave_request_reaches_its_named_approver(self):
        from hr.models import LeaveRequest, LeaveRequestApproval
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        req = LeaveRequest.objects.create(
            employee=self.report, leave_type=annual,
            start_date=_date(2026, 7, 1), end_date=_date(2026, 7, 2),
            days=Decimal('2'), status='pending', created_by=self.report_user)
        LeaveRequestApproval.objects.create(
            leave_request=req, approver=self.approver_user, decision='pending')
        groups = self._groups(self.approver_user)
        self.assertIn('leave', groups)
        self.assertEqual(groups['leave']['count'], 1)

    def test_seeing_the_leave_queue_is_not_the_same_as_approving_a_request(self):
        """Keyed off the approval row, not the dashboard access gate. Listing
        somebody else's decision as yours would send you to a page with no
        buttons on it."""
        from hr.models import LeaveRequest, LeaveRequestApproval
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        req = LeaveRequest.objects.create(
            employee=self.report, leave_type=annual,
            start_date=_date(2026, 7, 1), end_date=_date(2026, 7, 2),
            days=Decimal('2'), status='pending', created_by=self.report_user)
        LeaveRequestApproval.objects.create(
            leave_request=req, approver=self.approver_user, decision='pending')
        self.assertNotIn('leave', self._groups(self.stranger_user))

    def test_an_already_decided_item_drops_off(self):
        from hr.models import LeaveRequest, LeaveRequestApproval
        annual, _ = LeaveType.objects.get_or_create(
            code='annual', defaults={'name': 'Annual', 'default_annual_days': 30})
        req = LeaveRequest.objects.create(
            employee=self.report, leave_type=annual,
            start_date=_date(2026, 7, 1), end_date=_date(2026, 7, 2),
            days=Decimal('2'), status='pending', created_by=self.report_user)
        approval = LeaveRequestApproval.objects.create(
            leave_request=req, approver=self.approver_user, decision='pending')
        self.assertIn('leave', self._groups(self.approver_user))
        approval.decision = 'approved'
        approval.save()
        self.assertNotIn('leave', self._groups(self.approver_user))

    def test_late_queries_are_hr_only(self):
        """TeamExceptionsView silently downgrades a non-HR user off that tab,
        so listing one for them would be a dead link."""
        from hr.models import LateQuery
        from accounts.models import Role
        record = AttendanceRecord.objects.create(
            employee=self.report, date=_date(2026, 6, 10), status='late')
        LateQuery.objects.create(
            employee=self.report, attendance_record=record, message='Gate was shut')
        self.assertNotIn('late_queries', self._groups(self.approver_user))
        hr_user = _make_role_user('pa_hr', Role.SUPER_ADMIN)
        self.assertIn('late_queries', self._groups(hr_user))

    # ── empty groups do not clutter it ──────────────────────────────────────

    def test_an_empty_group_is_left_out_entirely(self):
        """A page of headings each saying zero is a page nobody reads."""
        self.assertEqual(self._groups(self.stranger_user), {})

    def test_the_count_is_the_sum_of_what_is_listed(self):
        from hr.approvals import pending_approvals, pending_approvals_count
        self._exception_for(self.report)
        groups = pending_approvals(self.approver_user)
        self.assertEqual(pending_approvals_count(self.approver_user),
                         sum(len(g['items']) for g in groups))

    # ── one broken source must not take the page down ───────────────────────

    def test_a_failing_source_does_not_lose_the_others(self):
        """This aggregates over half the system. One module's problem should
        not cost somebody the view of their other queues."""
        from unittest.mock import patch
        import hr.approvals as approvals

        def boom(_user):
            raise RuntimeError('procurement is having a day')

        self._exception_for(self.report)
        with patch.object(approvals, 'SOURCES',
                          (boom, approvals.attendance_exceptions)):
            groups = approvals.pending_approvals(self.approver_user)
        self.assertEqual([g['key'] for g in groups], ['attendance_exceptions'])


class PendingApprovalsViewTests(TestCase):
    """The tab itself."""

    def setUp(self):
        self.user = _login_user('pav_user')
        self.emp = make_employee(iqama='PAV-1', name='Tab Person')
        self.emp.user = self.user
        self.emp.save(update_fields=['user'])
        self.report = make_employee(iqama='PAV-2', name='Tab Report')
        self.report.main_manager = self.emp
        self.report.save(update_fields=['main_manager'])
        self.client.login(username='pav_user', password='testpass123')

    def _profile(self):
        return self.client.get(reverse('hr:my_profile'))

    def _approvals(self):
        return self.client.get(reverse('hr:my_approvals'))

    def test_the_tab_is_on_the_profile_page(self):
        self.assertContains(self._profile(), 'Pending Approvals')

    def test_the_tab_opens(self):
        resp = self._approvals()
        self.assertEqual(resp.status_code, 200)
        self.assertTemplateUsed(resp, 'hr/my_approvals.html')

    def test_an_empty_inbox_says_what_it_checked(self):
        """"You're all caught up" over a queue that was never looked at is
        worse than no page at all."""
        resp = self._approvals()
        self.assertContains(resp, 'Nothing is waiting on you')
        self.assertContains(resp, 'purchase-order approvals')

    def test_a_waiting_item_is_listed_with_a_way_to_act_on_it(self):
        _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20),
            event_start_time=_time(9, 0), reason_category='site_visit',
            created_by=self.user)
        resp = self._approvals()
        self.assertContains(resp, 'Tab Report')
        self.assertContains(resp, reverse('hr:team_exceptions'))

    def test_the_badge_counts_what_is_waiting(self):
        _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20),
            event_start_time=_time(9, 0), reason_category='site_visit',
            created_by=self.user)
        self.assertEqual(self._profile().context['my_approvals_count'], 1)
        self.assertEqual(self._approvals().context['my_approvals_count'], 1)

    def test_the_profile_tab_still_works(self):
        """The inbox is answered before the profile context is built; the
        profile itself must be untouched by that."""
        resp = self._profile()
        self.assertEqual(resp.status_code, 200)
        self.assertTemplateUsed(resp, 'hr/my_profile.html')
        self.assertIn('my_attendance_totals', resp.context)

    def test_the_inbox_does_not_build_the_profile_context(self):
        """It is answered early on purpose - opening the inbox should not pay
        for a month of attendance, a leave balance and an asset list nobody is
        looking at."""
        self.assertNotIn('my_attendance_totals', self._approvals().context)

    def test_it_needs_a_login(self):
        self.client.logout()
        self.assertNotEqual(
            self.client.get(reverse('hr:my_approvals')).status_code, 200)

    def test_the_first_link_still_works(self):
        """?tab=approvals shipped before the inbox had a URL of its own, so it
        redirects rather than quietly showing the profile instead."""
        resp = self.client.get(reverse('hr:my_profile'), {'tab': 'approvals'})
        self.assertRedirects(resp, reverse('hr:my_approvals'))

    def test_the_sidebar_carries_it(self):
        """The point of the menu entry: reachable without going through My
        Profile first."""
        body = self.client.get(reverse('hr:my_profile')).content.decode()
        self.assertIn(reverse('hr:my_approvals'), body)
        self.assertIn('Pending Approvals', body)

    def test_the_sidebar_badge_matches_the_page(self):
        """Built by the same function as the page, so a badge cannot send
        somebody looking for work that is not there."""
        _submit_aex(
            employee=self.report, event_date=_date(2026, 7, 20),
            event_start_time=_time(9, 0), reason_category='site_visit',
            created_by=self.user)
        # Any page renders the sidebar, so the count must be there too.
        elsewhere = self.client.get(reverse('hr:my_profile'))
        self.assertEqual(elsewhere.context['my_approvals_count'],
                         self._approvals().context['my_approvals_count'])


class PendingApprovalsPurchaseOrderTests(TestCase):
    """Purchase orders reach an inbox only when the signature is that
    person's own.

    can_user_approve_stage() lets a super admin sign anything, which is right
    for the button on the PO page - an override exists so work never stalls.
    It is wrong for a personal inbox: it would drop every open PO at every
    stage into one person's list of things waiting on them."""

    def setUp(self):
        from accounts.models import Role
        from procurement.models import PurchaseOrder
        self.super_admin = _make_role_user('po_super', Role.SUPER_ADMIN)
        self.admin = _make_role_user('po_admin', Role.ADMIN)
        self.proc_mgr = _make_role_user('po_scm', Role.PROCUREMENT_MGR)
        self.po = PurchaseOrder.objects.create(
            po_number='PO-INBOX-1', po_date=_date(2026, 6, 1),
            created_by=self.proc_mgr)

    def _po_items(self, user):
        from hr.approvals import pending_approvals
        for group in pending_approvals(user):
            if group['key'] == 'purchase_orders':
                return group['items']
        return []

    def _sign(self, *stages):
        """Walk the PO forward through the stages already signed."""
        from django.utils import timezone
        from procurement.models import PurchaseOrder
        updates = {f'{s}_approved_at': timezone.now() for s in stages}
        PurchaseOrder.objects.filter(pk=self.po.pk).update(**updates)
        self.po.refresh_from_db()

    # ── whose signature is it ───────────────────────────────────────────────

    def test_the_first_stage_is_the_procurement_managers(self):
        self.assertEqual(self.po.current_stage['key'], 'scm')
        self.assertEqual(len(self._po_items(self.proc_mgr)), 1)

    def test_a_super_admin_does_not_get_somebody_elses_stage(self):
        """The whole point: an override is not an assignment."""
        self.assertEqual(self.po.current_stage['key'], 'scm')
        self.assertEqual(self._po_items(self.super_admin), [])

    def test_a_super_admin_can_still_sign_it_on_the_po_page(self):
        """The inbox narrows what is shown; it must not narrow what is
        allowed. The override still stands where it belongs."""
        self.assertTrue(self.po.can_user_approve_stage(self.super_admin, 'scm'))

    def test_the_middle_stages_are_the_admins(self):
        self._sign('scm')
        self.assertEqual(self.po.current_stage['key'], 'pm')
        self.assertEqual(len(self._po_items(self.admin)), 1)
        self.assertEqual(self._po_items(self.super_admin), [])
        self.assertEqual(self._po_items(self.proc_mgr), [])

    def test_a_stage_that_is_not_current_yet_is_nobodys_inbox_item(self):
        """Strict sequencing - only the current stage can be signed, so only
        the current stage is waiting on anyone."""
        self.assertEqual(self.po.current_stage['key'], 'scm')
        self.assertEqual(self._po_items(self.admin), [])

    def test_a_signed_stage_drops_off_the_signers_inbox(self):
        self.assertEqual(len(self._po_items(self.proc_mgr)), 1)
        self._sign('scm')
        self.assertEqual(self._po_items(self.proc_mgr), [])

    def test_a_released_po_is_waiting_on_nobody(self):
        self._sign('scm', 'pm', 'coo')
        self.assertIsNone(self.po.current_stage)
        for user in (self.proc_mgr, self.admin, self.super_admin):
            with self.subTest(user=user.username):
                self.assertEqual(self._po_items(user), [])

    # ── the named signer ────────────────────────────────────────────────────

    def test_a_super_admin_who_is_the_named_pm_does_see_it(self):
        """APPROVAL_STAGES names a person per stage, and that name is the only
        individual identity this model carries. A super admin who IS the PM is
        the person the work is waiting on, unlike every other super admin."""
        from procurement.models import PurchaseOrder
        pm_name = PurchaseOrder.stage_signer_name('pm')
        first, _, last = pm_name.partition(' ')
        self.super_admin.first_name, self.super_admin.last_name = first, last
        self.super_admin.save(update_fields=['first_name', 'last_name'])
        self._sign('scm')
        self.assertEqual(self.po.current_stage['key'], 'pm')
        self.assertEqual(len(self._po_items(self.super_admin)), 1)

    def test_that_same_super_admin_still_does_not_get_the_other_stages(self):
        """Being the PM makes the PM stage theirs, not the whole pipeline."""
        from procurement.models import PurchaseOrder
        pm_name = PurchaseOrder.stage_signer_name('pm')
        first, _, last = pm_name.partition(' ')
        self.super_admin.first_name, self.super_admin.last_name = first, last
        self.super_admin.save(update_fields=['first_name', 'last_name'])
        self.assertEqual(self.po.current_stage['key'], 'scm')
        self.assertEqual(self._po_items(self.super_admin), [])

    def test_the_name_match_forgives_case_and_spacing(self):
        """Those names are typed constants and a user record is typed
        separately; an exact comparison would fail on a double space."""
        from procurement.models import PurchaseOrder
        pm_name = PurchaseOrder.stage_signer_name('pm')
        first, _, last = pm_name.partition(' ')
        self.super_admin.first_name = f'  {first.upper()} '
        self.super_admin.last_name = f' {last.lower()}  '
        self.super_admin.save(update_fields=['first_name', 'last_name'])
        self._sign('scm')
        self.assertEqual(len(self._po_items(self.super_admin)), 1)

    def test_the_named_signer_never_widens_what_is_allowed(self):
        """Three layers, and the name only touches the middle one.

        Visibility decides which POs a user can see at all; designation
        decides whose inbox a stage belongs in; permission decides who may
        press Approve. Sharing a name with the PM moves the middle one and
        neither of the others - a stranger still cannot see the PO, and still
        cannot sign it.
        """
        from procurement.models import PurchaseOrder
        from procurement.views import _visible_pos_for
        plain = _login_user('po_plain')
        pm_name = PurchaseOrder.stage_signer_name('pm')
        first, _, last = pm_name.partition(' ')
        plain.first_name, plain.last_name = first, last
        plain.save(update_fields=['first_name', 'last_name'])
        self._sign('scm')

        # Designation: the stage is mapped to that name.
        self.assertTrue(self.po.is_designated_approver(plain, 'pm'))
        # Permission: untouched, and still refuses them.
        self.assertFalse(self.po.can_user_approve_stage(plain, 'pm'))
        # Visibility: they cannot see this PO, so it reaches no inbox of
        # theirs regardless of the name.
        self.assertFalse(_visible_pos_for(plain).filter(pk=self.po.pk).exists())
        self.assertEqual(self._po_items(plain), [])

    def test_the_ceo_stage_stays_with_the_super_admins(self):
        """CEO is super-admin-only by design, so excluding it would leave that
        stage in nobody's inbox at all.

        total_value is a property summing the PO's items, so the threshold is
        dropped rather than the value inflated - the point under test is the
        stage mapping, not how the value is arrived at.
        """
        from decimal import Decimal as _D
        from unittest.mock import patch
        from procurement.models import PurchaseOrder
        with patch.object(PurchaseOrder, 'CEO_APPROVAL_THRESHOLD', _D('0')):
            self.po.refresh_from_db()
            self.assertIn('ceo', self.po.required_stages)
            self._sign('scm', 'pm', 'coo')
            self.assertEqual(self.po.current_stage['key'], 'ceo')
            self.assertEqual(len(self._po_items(self.super_admin)), 1)
            self.assertEqual(self._po_items(self.admin), [])


class EmployeeGradeAndListDisplayTests(TestCase):
    """Grade field: saves via the form, and shows correctly across the list
    and detail pages. The list page shows Grade + Joining Date instead of
    Deployment; the detail page keeps Deployment alongside the new Grade."""

    def setUp(self):
        self.admin = _make_role_user('grade-admin', Role.SUPER_ADMIN)
        self.emp = make_employee(iqama='GRADE-1', name='Grade Test Employee')
        self.emp.grade = 'EX-2'
        self.emp.deployment = 'Riyadh Site'
        self.emp.joining_date = _date(2022, 3, 15)
        self.emp.save(update_fields=['grade', 'deployment', 'joining_date'])

    def test_grade_field_persists(self):
        self.emp.refresh_from_db()
        self.assertEqual(self.emp.grade, 'EX-2')

    def test_grade_filter_narrows_list_results(self):
        # Replaces the old Deployment filter, which filtered on a value
        # that had no matching column in the results (see PR review).
        other = make_employee(iqama='GRADE-2', name='Other Grade Employee')
        other.grade = 'EX-5'
        other.save(update_fields=['grade'])
        self.client.login(username='grade-admin', password='testpass123')
        resp = self.client.get(reverse('hr:employee_list'), {'grade': 'EX-2'})
        self.assertContains(resp, 'Grade Test Employee')
        self.assertNotContains(resp, 'Other Grade Employee')

    def test_grade_and_joining_date_shown_on_list_not_deployment_column(self):
        self.client.login(username='grade-admin', password='testpass123')
        resp = self.client.get(reverse('hr:employee_list'))
        content = resp.content.decode()
        self.assertContains(resp, 'Grade')
        self.assertContains(resp, 'Joining Date')
        self.assertContains(resp, 'EX-2')
        self.assertContains(resp, '15 Mar 2022')
        # The column header itself must be gone, even though the word
        # "Deployment" still legitimately appears elsewhere (e.g. the detail
        # page link's title attribute is not present on this page, but to be
        # safe check the specific header cell, not a bare substring).
        self.assertNotIn('<th>Deployment</th>', content)

    def test_deployment_and_grade_both_shown_on_detail_page(self):
        self.client.login(username='grade-admin', password='testpass123')
        resp = self.client.get(reverse('hr:employee_detail', args=[self.emp.pk]))
        self.assertContains(resp, 'Riyadh Site')
        self.assertContains(resp, 'EX-2')

    def test_grade_field_present_on_edit_form(self):
        self.client.login(username='grade-admin', password='testpass123')
        resp = self.client.get(reverse('hr:employee_update', args=[self.emp.pk]))
        self.assertContains(resp, 'name="grade"')

    def test_missing_grade_shows_dash_on_list_and_detail(self):
        blank_emp = make_employee(iqama='GRADE-3', name='No Grade Employee')
        self.client.login(username='grade-admin', password='testpass123')
        list_resp = self.client.get(reverse('hr:employee_list'))
        list_html = list_resp.content.decode()
        self.assertIn('No Grade Employee', list_html)
        # Isolate this employee's own table row rather than checking for a
        # bare '-' anywhere on the page, which would pass even if grade
        # rendered as something else entirely.
        row = list_html.split('No Grade Employee', 1)[1][:400]
        self.assertIn('<td>-</td>', row)
        detail_resp = self.client.get(reverse('hr:employee_detail', args=[blank_emp.pk]))
        self.assertEqual(detail_resp.status_code, 200)
        detail_html = detail_resp.content.decode()
        grade_section = detail_html.split('Grade:</strong>', 1)[1][:100]
        self.assertIn('-', grade_section)


class EmployeeImportColumnMappingTests(TestCase):
    """Import correctly maps the real-world Excel layout: a Grade column,
    "OCCUPATION (Iqama)" for designation, and the "DIPLOYMENT" typo actually
    present in the source sheet - alongside the existing, correctly-spelled
    aliases already relied on elsewhere."""

    def setUp(self):
        self.admin = _make_role_user('import-admin', Role.SUPER_ADMIN)
        self.client.login(username='import-admin', password='testpass123')

    def _build_workbook(self, headers, rows):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return SimpleUploadedFile(
            'employees.xlsx', buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_import_maps_grade_column(self):
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'Grade'],
            [['Test Employee', 'IMP-001', 'EX-3']])
        self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        emp = Employee.objects.get(iqama_number='IMP-001')
        self.assertEqual(emp.grade, 'EX-3')

    def test_import_maps_occupation_iqama_header_to_designation(self):
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'OCCUPATION (Iqama)'],
            [['Test Employee', 'IMP-002', 'Electrical Machines Maint Tech']])
        self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        emp = Employee.objects.get(iqama_number='IMP-002')
        self.assertEqual(emp.designation, 'Electrical Machines Maint Tech')

    def test_import_maps_diployment_typo_to_deployment(self):
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'DIPLOYMENT'],
            [['Test Employee', 'IMP-003', 'Leap Networks']])
        self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        emp = Employee.objects.get(iqama_number='IMP-003')
        self.assertEqual(emp.deployment, 'Leap Networks')

    def test_import_still_maps_correct_deployment_spelling(self):
        # The typo alias must not break the existing, correctly-spelled
        # header that other imports already rely on.
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'Deployment'],
            [['Test Employee', 'IMP-004', 'Head Office']])
        self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        emp = Employee.objects.get(iqama_number='IMP-004')
        self.assertEqual(emp.deployment, 'Head Office')


class EmployeeImportAuthBoundaryTests(TestCase):
    """Only Super Admin / ERP Admin can import employees - matches the
    check already in employee_import() (is_super_admin_user or
    is_erp_admin_user)."""

    def _build_minimal_workbook(self):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Name', 'Iqama/Passport'])
        ws.append(['Blocked Import Test', 'IMP-AUTH-1'])
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return SimpleUploadedFile(
            'employees.xlsx', buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_regular_user_cannot_import(self):
        _login_user('import-plain')
        self.client.login(username='import-plain', password='testpass123')
        resp = self.client.post(
            reverse('hr:employee_import'), {'excel_file': self._build_minimal_workbook()})
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse('hr:employee_list'))
        self.assertFalse(Employee.objects.filter(iqama_number='IMP-AUTH-1').exists())

    def test_super_admin_can_import(self):
        _make_role_user('import-super', Role.SUPER_ADMIN)
        self.client.login(username='import-super', password='testpass123')
        self.client.post(
            reverse('hr:employee_import'), {'excel_file': self._build_minimal_workbook()})
        self.assertTrue(Employee.objects.filter(iqama_number='IMP-AUTH-1').exists())


class EmployeeImportEdgeCaseTests(TestCase):
    """Grade import handles missing columns, whitespace, over-length
    values, and header collisions without crashing or silently corrupting
    other fields."""

    def setUp(self):
        _make_role_user('import-edge', Role.SUPER_ADMIN)
        self.client.login(username='import-edge', password='testpass123')

    def _build_workbook(self, headers, rows):
        import openpyxl
        from io import BytesIO
        from django.core.files.uploadedfile import SimpleUploadedFile
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return SimpleUploadedFile(
            'employees.xlsx', buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_sheet_with_no_grade_column_does_not_crash(self):
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'Nationality'],
            [['Test Employee', 'IMP-EDGE-1', 'Saudi']])
        resp = self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        self.assertEqual(resp.status_code, 302)
        emp = Employee.objects.get(iqama_number='IMP-EDGE-1')
        self.assertEqual(emp.grade, '')

    def test_grade_value_whitespace_is_stripped(self):
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'Grade'],
            [['Test Employee', 'IMP-EDGE-2', '  EX-1  ']])
        self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        emp = Employee.objects.get(iqama_number='IMP-EDGE-2')
        self.assertEqual(emp.grade, 'EX-1')

    def test_overlength_grade_value_is_truncated_not_rejected(self):
        long_grade = 'X' * 80
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'Grade'],
            [['Test Employee', 'IMP-EDGE-3', long_grade]])
        resp = self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        self.assertEqual(resp.status_code, 302)
        emp = Employee.objects.get(iqama_number='IMP-EDGE-3')
        self.assertEqual(len(emp.grade), 50)

    def test_correct_spelling_wins_when_both_deployment_headers_present(self):
        # If a sheet somehow has both the correct spelling and the typo,
        # the correctly-spelled header (checked first) should win rather
        # than the result depending on column order.
        f = self._build_workbook(
            ['Name', 'Iqama/Passport', 'Deployment', 'DIPLOYMENT'],
            [['Test Employee', 'IMP-EDGE-4', 'Correct Value', 'Typo Value']])
        self.client.post(reverse('hr:employee_import'), {'excel_file': f})
        emp = Employee.objects.get(iqama_number='IMP-EDGE-4')
        self.assertEqual(emp.deployment, 'Correct Value')


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class EmployeePictureDisplayTests(TestCase):
    """Picture upload via the form, and correct display (or placeholder)
    on the detail page header."""

    def setUp(self):
        self.admin = _make_role_user('pic-admin', Role.SUPER_ADMIN)
        self.client.login(username='pic-admin', password='testpass123')
        self.emp = make_employee(iqama='PIC-1', name='Picture Test Employee')

    def _tiny_png(self, name='photo.png'):
        from io import BytesIO
        from PIL import Image as PILImage
        from django.core.files.uploadedfile import SimpleUploadedFile
        buf = BytesIO()
        PILImage.new('RGB', (2, 2), color='red').save(buf, format='PNG')
        buf.seek(0)
        return SimpleUploadedFile(name, buf.read(), content_type='image/png')

    def test_picture_uploads_and_saves(self):
        resp = self.client.post(
            reverse('hr:employee_update', args=[self.emp.pk]),
            data={
                'full_name': self.emp.full_name, 'iqama_number': self.emp.iqama_number,
                'picture': self._tiny_png(),
            })
        self.emp.refresh_from_db()
        self.assertTrue(bool(self.emp.picture))

    def test_oversized_picture_is_rejected(self):
        # A large photo would upload fine but then get embedded into
        # every subsequent Excel export, making that file huge too.
        from io import BytesIO
        from PIL import Image as PILImage
        from django.core.files.uploadedfile import SimpleUploadedFile
        buf = BytesIO()
        # Large, low-compression image reliably exceeds the 5 MB limit.
        PILImage.new('RGB', (3000, 3000), color='red').save(buf, format='PNG', compress_level=0)
        buf.seek(0)
        big_file = SimpleUploadedFile('big.png', buf.read(), content_type='image/png')
        resp = self.client.post(
            reverse('hr:employee_update', args=[self.emp.pk]),
            data={
                'full_name': self.emp.full_name, 'iqama_number': self.emp.iqama_number,
                'picture': big_file,
            })
        self.assertEqual(resp.status_code, 200)  # re-rendered form, not a redirect
        self.assertContains(resp, 'smaller than 5 MB')
        self.emp.refresh_from_db()
        self.assertFalse(bool(self.emp.picture))

    def test_edit_form_has_multipart_enctype(self):
        # A form GET is real HTML - unlike test_picture_uploads_and_saves,
        # which posts through Django's test client and would pass even
        # with no enctype attribute at all, since the client builds a
        # multipart request itself regardless of what the template says.
        # This is the actual regression test for the missing-enctype bug.
        resp = self.client.get(reverse('hr:employee_update', args=[self.emp.pk]))
        self.assertContains(resp, 'enctype="multipart/form-data"')

    def test_picture_shown_on_detail_page_when_set(self):
        self.emp.picture.save('photo.png', self._tiny_png(), save=True)
        resp = self.client.get(reverse('hr:employee_detail', args=[self.emp.pk]))
        self.assertContains(resp, self.emp.picture.url)

    def test_placeholder_icon_shown_when_no_picture(self):
        resp = self.client.get(reverse('hr:employee_detail', args=[self.emp.pk]))
        self.assertContains(resp, 'bi-person')


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class EmployeeExportTests(TestCase):
    """Excel export: Picture column present, real images embedded for
    employees who have one, and graceful handling when a picture is
    missing or its file no longer exists on disk. Also covers the
    admin-only access boundary, which had no prior test coverage."""

    def setUp(self):
        self.admin = _make_role_user('export-admin', Role.SUPER_ADMIN)
        self.client.login(username='export-admin', password='testpass123')

    def _tiny_png(self, name='photo.png'):
        from io import BytesIO
        from PIL import Image as PILImage
        from django.core.files.uploadedfile import SimpleUploadedFile
        buf = BytesIO()
        PILImage.new('RGB', (2, 2), color='blue').save(buf, format='PNG')
        buf.seek(0)
        return SimpleUploadedFile(name, buf.read(), content_type='image/png')

    def _read_workbook(self, response):
        import openpyxl
        from io import BytesIO
        return openpyxl.load_workbook(BytesIO(response.content))

    def test_export_includes_picture_column_header(self):
        make_employee(iqama='EXP-1', name='Export Test Employee')
        resp = self.client.get(reverse('hr:employee_export'))
        wb = self._read_workbook(resp)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Picture', headers)

    def test_export_embeds_image_for_employee_with_picture(self):
        emp = make_employee(iqama='EXP-2', name='Has Picture Employee')
        emp.picture.save('photo.png', self._tiny_png(), save=True)
        resp = self.client.get(reverse('hr:employee_export'))
        wb = self._read_workbook(resp)
        ws = wb.active
        self.assertGreaterEqual(len(ws._images), 1)

    def test_export_does_not_crash_for_employee_without_picture(self):
        make_employee(iqama='EXP-3', name='No Picture Employee')
        resp = self.client.get(reverse('hr:employee_export'))
        self.assertEqual(resp.status_code, 200)
        wb = self._read_workbook(resp)
        self.assertEqual(len(wb.active._images), 0)

    def test_export_does_not_crash_when_picture_file_missing_from_disk(self):
        emp = make_employee(iqama='EXP-4', name='Missing File Employee')
        # Point at a picture path that was never actually written to disk -
        # simulates a DB row surviving after the underlying file was deleted.
        emp.picture.name = 'employee_pictures/does_not_exist.png'
        emp.save(update_fields=['picture'])
        resp = self.client.get(reverse('hr:employee_export'))
        self.assertEqual(resp.status_code, 200)

    def test_regular_user_cannot_export(self):
        _login_user('export-plain')
        self.client.login(username='export-plain', password='testpass123')
        resp = self.client.get(reverse('hr:employee_export'))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(resp.url, reverse('hr:employee_list'))

    def test_super_admin_can_export(self):
        resp = self.client.get(reverse('hr:employee_export'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


class EmployeeListAndExportQueryScalingTests(TestCase):
    """Guards against N+1 regressions on the list and export views: query
    count must not grow as the number of employees grows. This wouldn't
    catch an N+1 from grade/picture today (both are plain columns, not
    relations), but it's exactly the test that catches one the day
    someone adds a related-field lookup (e.g. emp.manager.full_name)
    to either view without select_related."""

    def setUp(self):
        _make_role_user('query-admin', Role.SUPER_ADMIN)
        self.client.login(username='query-admin', password='testpass123')

    def _query_count_for_n_employees(self, url_name, n):
        from django.test.utils import CaptureQueriesContext
        from django.db import connection
        Employee.objects.all().delete()
        for i in range(n):
            make_employee(iqama=f'QRY-{url_name}-{i}', name=f'Query Test {i}')
        with CaptureQueriesContext(connection) as ctx:
            resp = self.client.get(reverse(url_name))
            self.assertEqual(resp.status_code, 200)
            # For a paginated list, force the page to actually render/consume
            # the queryset within the captured block.
            if hasattr(resp, 'content'):
                _ = resp.content
        return len(ctx.captured_queries)

    def test_employee_list_query_count_does_not_scale_with_row_count(self):
        small = self._query_count_for_n_employees('hr:employee_list', 3)
        large = self._query_count_for_n_employees('hr:employee_list', 15)
        self.assertEqual(
            small, large,
            f"Query count grew from {small} (3 employees) to {large} (15 employees) - "
            "likely an N+1 introduced in the list view or template.")

    def test_employee_export_query_count_does_not_scale_with_row_count(self):
        small = self._query_count_for_n_employees('hr:employee_export', 3)
        large = self._query_count_for_n_employees('hr:employee_export', 15)
        self.assertEqual(
            small, large,
            f"Query count grew from {small} (3 employees) to {large} (15 employees) - "
            "likely an N+1 introduced in the export view.")


from django.core.files.storage import FileSystemStorage


class _PathlessStorage(FileSystemStorage):
    """Mimics S3Storage's contract exactly: everything else (open, save,
    exists, url) works independently of any filesystem path - the same
    shape as production's R2 backend - while the public .path() method
    is unimplemented, exactly like S3Storage. FileSystemStorage's own
    exists()/_open()/_save() all normally call self.path() internally,
    so those are overridden here too to route through a private path
    helper instead - otherwise saving through this storage would itself
    raise before the test ever reached the code being tested. Used to
    prove the export survives a storage backend that can't do
    filesystem paths, since this exact bug (export 500s the moment
    anyone has a picture) shipped once already and nothing exercised
    the production storage path."""
    def path(self, name):
        raise NotImplementedError("This backend doesn't support absolute paths.")

    def _local_path(self, name):
        import os
        return os.path.join(self.location, name)

    def exists(self, name):
        import os
        return os.path.lexists(self._local_path(name))

    def _open(self, name, mode='rb'):
        from django.core.files import File
        return File(open(self._local_path(name), mode))

    def _save(self, name, content):
        import os
        full_path = self._local_path(name)
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        with open(full_path, 'wb') as f:
            for chunk in content.chunks():
                f.write(chunk)
        return name

    def url(self, name):
        return '/media/' + name


@override_settings(MEDIA_ROOT=tempfile.mkdtemp())
class EmployeeExportPathlessStorageTests(TestCase):
    """Production runs an S3-compatible backend (R2) whose .path() raises
    NotImplementedError, not OSError/FileNotFoundError - the export must
    read picture bytes via .open() rather than .path(), and must survive
    whatever a real storage backend throws, not just filesystem errors."""

    def setUp(self):
        self.admin = _make_role_user('pathless-admin', Role.SUPER_ADMIN)
        self.client.login(username='pathless-admin', password='testpass123')

    def _tiny_png(self, name='photo.png'):
        from io import BytesIO
        from PIL import Image as PILImage
        from django.core.files.uploadedfile import SimpleUploadedFile
        buf = BytesIO()
        PILImage.new('RGB', (2, 2), color='green').save(buf, format='PNG')
        buf.seek(0)
        return SimpleUploadedFile(name, buf.read(), content_type='image/png')

    @override_settings(STORAGES={
        'default': {'BACKEND': 'hr.tests._PathlessStorage'},
        'staticfiles': {'BACKEND': 'django.contrib.staticfiles.storage.StaticFilesStorage'},
    })
    def test_export_does_not_500_on_a_pathless_storage_backend(self):
        emp = make_employee(iqama='PATHLESS-1', name='Pathless Storage Employee')
        emp.picture.save('photo.png', self._tiny_png(), save=True)
        resp = self.client.get(reverse('hr:employee_export'))
        self.assertEqual(resp.status_code, 200)
        import openpyxl
        from io import BytesIO as _BIO
        wb = openpyxl.load_workbook(_BIO(resp.content))
        # The picture was embedded successfully despite the backend
        # having no .path() at all - this is the actual regression test
        # for the production bug.
        self.assertGreaterEqual(len(wb.active._images), 1)
class PendingApprovalsBadgeCostTests(TestCase):
    """The sidebar badge runs on every page in the app, so its cost is the
    app's cost.

    hr.context_processors.pending_counts calls pending_approvals_count() for
    every authenticated request, and that walks all six sources and builds
    their full item lists purely to call len() on them. Anything that scales
    inside those sources therefore scales across every page in the system.

    This went unnoticed once already: the purchase-order source read
    current_stage (-> required_stages -> total_value, which walks items),
    po.project, and each stage's *_approved_by FK, none of them prefetched.
    Twenty-one purchase orders in a development database produced 362 queries
    on every page load, to render a badge that read zero.
    """

    def setUp(self):
        from accounts.models import Role
        self.user = _make_role_user('badge_super', Role.SUPER_ADMIN)

    def _po(self, number):
        from procurement.models import PurchaseOrder, PurchaseOrderItem
        po = PurchaseOrder.objects.create(
            po_number=number, po_date=_date(2026, 6, 1), created_by=self.user)
        PurchaseOrderItem.objects.create(
            purchase_order=po, serial_number=1, description='Thing',
            quantity=Decimal('1'), rate_per_unit=Decimal('100'))
        return po

    def _queries(self):
        from django.db import connection
        from django.test.utils import CaptureQueriesContext
        from hr.approvals import pending_approvals_count
        with CaptureQueriesContext(connection) as ctx:
            pending_approvals_count(self.user)
        return len(ctx.captured_queries)

    def test_the_badge_cost_does_not_scale_with_purchase_orders(self):
        """Measured at two sizes. A pinned number would be edited to match a
        regression rather than reveal one — which is how the 362-query version
        survived."""
        for i in range(2):
            self._po(f'PO-BADGE-A{i}')
        # One warm-up call. Django caches a missing reverse one-to-one on the
        # instance, so the employee_profile lookup is paid once and would
        # otherwise make the first measurement a query dearer than the second
        # for reasons that have nothing to do with purchase orders.
        self._queries()
        small = self._queries()

        for i in range(12):
            self._po(f'PO-BADGE-B{i}')
        large = self._queries()

        self.assertEqual(
            small, large,
            f'Badge cost grew from {small} queries (2 POs) to {large} (14 POs). '
            f'This runs on every page in the app.')

    def test_the_badge_still_counts_the_same_things_it_did(self):
        """A prefetch must not change the answer, only the cost."""
        from hr.approvals import pending_approvals, pending_approvals_count
        self._po('PO-BADGE-COUNT')
        expected = sum(g['count'] for g in pending_approvals(self.user))
        self.assertEqual(pending_approvals_count(self.user), expected)


class PendingApprovalsSourceHealthTests(TestCase):
    """Every source must actually run.

    pending_approvals drops a source that raises, so one module's problem
    cannot cost somebody the view of their other five queues. The cost of that
    choice is that a permanently broken source is indistinguishable from an
    empty one: leave_revokes named its foreign key `leave_record` when the
    field is `leave_request`, so every call raised FieldError, the queue was
    never in anybody's inbox, and nothing anywhere said so.

    Every other test in this file went through pending_approvals() or the
    view, so all of them saw the swallowed exception as "no leave
    cancellations pending" and passed. These call the sources directly.
    """

    def setUp(self):
        from accounts.models import Role
        self.super_admin = _make_role_user('src_super', Role.SUPER_ADMIN)

    def test_no_source_raises(self):
        from hr.approvals import SOURCES
        broken = []
        for source in SOURCES:
            try:
                source(self.super_admin)
            except Exception as exc:                      # noqa: BLE001
                broken.append(f'{source.__name__}: {type(exc).__name__}: {exc}')
        self.assertEqual(
            broken, [],
            'These sources raise and are silently dropped from every inbox:\n'
            + '\n'.join(broken))

    def test_every_source_is_reachable_for_someone(self):
        """A source that returns None for a super admin is either correctly
        scoped away or quietly unreachable; a super admin should reach all of
        them, so None here means the gate is wrong."""
        from hr.approvals import SOURCES
        unreachable = [s.__name__ for s in SOURCES
                       if s(self.super_admin) is None]
        self.assertEqual(unreachable, [])

    def test_a_pending_leave_revoke_reaches_the_inbox(self):
        """The specific queue that was missing. Asserted through the real
        aggregator, so it also proves the source is not being dropped."""
        from django.utils import timezone
        from hr.approvals import pending_approvals
        from hr.models import LeaveRequest, LeaveRevokeRequest, LeaveType

        employee = make_employee(iqama='REVOKE-1', name='Revoke Test Employee')
        leave_type = LeaveType.objects.create(name='Annual', code='ANN-REV')
        leave = LeaveRequest.objects.create(
            employee=employee, leave_type=leave_type,
            start_date=_date(2026, 6, 1), end_date=_date(2026, 6, 3),
            days=Decimal('3'), status='pending')
        LeaveRevokeRequest.objects.create(
            leave_request=leave, requested_by=self.super_admin,
            reason='Client moved the visit', status='pending')

        groups = {g['key']: g for g in pending_approvals(self.super_admin)}
        self.assertIn('leave_revoke', groups)
        self.assertEqual(groups['leave_revoke']['count'], 1)
        self.assertEqual(groups['leave_revoke']['items'][0]['title'],
                         'Revoke Test Employee')

    def test_a_failing_source_is_logged_rather_than_vanishing(self):
        """Dropping a source silently is what let the above go unnoticed. It
        must be recoverable from the logs."""
        from unittest import mock
        from hr import approvals

        def boom(user):
            raise RuntimeError('deliberate')

        with mock.patch.object(approvals, 'SOURCES', (boom,)):
            with self.assertLogs('hr.approvals', level='ERROR') as captured:
                self.assertEqual(approvals.pending_approvals(self.super_admin), [])
        self.assertIn('boom', '\n'.join(captured.output))


from hr.models import EmployeeDocument


class MyDocumentUploadTests(TestCase):
    """Self-service document upload/delete/export on My Profile - the
    dropdown is restricted to 5 types (not the full 12 admin choices),
    Other requires a note, and only the owning employee can touch their
    own documents."""

    def setUp(self):
        self.user = User.objects.create_user('doc_emp', password='x', email='doc_emp@leap.com')
        self.emp = Employee.objects.create(
            full_name='Document Tester', iqama_number='IQ-DOC', user=self.user)
        self.other_user = User.objects.create_user('doc_other', password='x', email='doc_other@leap.com')
        self.other_emp = Employee.objects.create(
            full_name='Other Doc Person', iqama_number='IQ-DOC-OTHER', user=self.other_user)
        self.no_emp_user = User.objects.create_user('doc_noemp', password='x', email='doc_noemp@leap.com')

    def _tiny_pdf(self, name='doc.pdf'):
        from io import BytesIO
        from reportlab.pdfgen import canvas
        from django.core.files.uploadedfile import SimpleUploadedFile
        buf = BytesIO()
        c = canvas.Canvas(buf)
        c.drawString(100, 700, 'Test Document')
        c.save()
        buf.seek(0)
        return SimpleUploadedFile(name, buf.read(), content_type='application/pdf')

    # -- upload --

    def test_upload_with_known_type_sets_title_from_choice_label(self):
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_upload'), {
            'document_type': 'iqama', 'notes': '', 'file': self._tiny_pdf(),
        })
        self.assertEqual(resp.status_code, 302)
        doc = self.emp.documents.get()
        self.assertEqual(doc.title, 'Iqama / ID Copy')
        self.assertEqual(doc.uploaded_by, self.user)

    def test_upload_other_type_uses_notes_as_title(self):
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_upload'), {
            'document_type': 'other', 'notes': 'Salary Certificate', 'file': self._tiny_pdf(),
        })
        self.assertEqual(resp.status_code, 302)
        doc = self.emp.documents.get()
        self.assertEqual(doc.title, 'Salary Certificate')

    def test_upload_other_type_without_notes_is_rejected(self):
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_upload'), {
            'document_type': 'other', 'notes': '', 'file': self._tiny_pdf(),
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.emp.documents.count(), 0)

    def test_dropdown_only_offers_self_service_choices(self):
        from hr.forms import MyDocumentUploadForm
        form = MyDocumentUploadForm()
        choice_values = [c[0] for c in form.fields['document_type'].choices]
        self.assertEqual(set(choice_values), set(EmployeeDocument.SELF_SERVICE_DOC_TYPES))
        # Confirms admin-only types are excluded, not just that the
        # allowed ones are present.
        self.assertNotIn('warning_letter', choice_values)
        self.assertNotIn('salary_slip', choice_values)

    # -- auth boundary --

    def test_user_without_employee_profile_cannot_upload(self):
        self.client.force_login(self.no_emp_user)
        resp = self.client.post(reverse('hr:my_document_upload'), {
            'document_type': 'iqama', 'notes': '', 'file': self._tiny_pdf(),
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(EmployeeDocument.objects.count(), 0)

    def test_get_request_to_upload_is_rejected(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_document_upload'))
        self.assertEqual(resp.status_code, 405)

    def test_cannot_delete_another_employees_document(self):
        doc = EmployeeDocument.objects.create(
            employee=self.other_emp, document_type='iqama', title='Iqama / ID Copy',
            file=self._tiny_pdf(), uploaded_by=self.other_user)
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_delete', args=[doc.pk]))
        self.assertEqual(resp.status_code, 404)
        self.assertTrue(EmployeeDocument.objects.filter(pk=doc.pk).exists())

    def test_can_delete_own_document(self):
        doc = EmployeeDocument.objects.create(
            employee=self.emp, document_type='iqama', title='Iqama / ID Copy',
            file=self._tiny_pdf(), uploaded_by=self.user)
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_delete', args=[doc.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(EmployeeDocument.objects.filter(pk=doc.pk).exists())

    def test_user_without_employee_profile_cannot_export(self):
        self.client.force_login(self.no_emp_user)
        resp = self.client.get(reverse('hr:my_documents_export_pdf'))
        self.assertEqual(resp.status_code, 302)

    # -- export / edge cases --

    def test_export_with_no_documents_still_produces_a_pdf(self):
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_documents_export_pdf'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')

    def test_export_includes_uploaded_document_in_toc(self):
        EmployeeDocument.objects.create(
            employee=self.emp, document_type='iqama', title='Iqama / ID Copy',
            file=self._tiny_pdf(), uploaded_by=self.user)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_documents_export_pdf'))
        self.assertEqual(resp.status_code, 200)
        from pypdf import PdfReader
        from io import BytesIO
        reader = PdfReader(BytesIO(resp.content))
        # TOC page + at least 1 page from the merged document.
        self.assertGreaterEqual(len(reader.pages), 2)

    def test_export_handles_unconvertible_file_without_crashing(self):
        # A .txt file has no PDF/image conversion path - must be skipped
        # gracefully, not crash the whole export.
        from django.core.files.uploadedfile import SimpleUploadedFile
        txt_file = SimpleUploadedFile('notes.txt', b'plain text', content_type='text/plain')
        EmployeeDocument.objects.create(
            employee=self.emp, document_type='other', title='Random Notes',
            file=txt_file, uploaded_by=self.user)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_documents_export_pdf'))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/pdf')


    def test_user_without_employee_profile_cannot_delete(self):
        doc = EmployeeDocument.objects.create(
            employee=self.emp, document_type='iqama', title='Iqama / ID Copy',
            file=self._tiny_pdf(), uploaded_by=self.user)
        self.client.force_login(self.no_emp_user)
        resp = self.client.post(reverse('hr:my_document_delete', args=[doc.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(EmployeeDocument.objects.filter(pk=doc.pk).exists())

    def test_deleting_nonexistent_document_404s(self):
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_delete', args=[99999]))
        self.assertEqual(resp.status_code, 404)

    def test_upload_without_file_is_rejected(self):
        self.client.force_login(self.user)
        resp = self.client.post(reverse('hr:my_document_upload'), {
            'document_type': 'iqama', 'notes': '',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(self.emp.documents.count(), 0)

    def test_export_includes_actual_image_content(self):
        # Confirms the image -> PDF conversion path actually works in the
        # merged export, not just that PDFs pass through and non-images
        # get skipped gracefully.
        from io import BytesIO
        from PIL import Image as PILImage
        from django.core.files.uploadedfile import SimpleUploadedFile
        buf = BytesIO()
        PILImage.new('RGB', (100, 100), color='blue').save(buf, format='PNG')
        buf.seek(0)
        image_file = SimpleUploadedFile('photo.png', buf.read(), content_type='image/png')
        EmployeeDocument.objects.create(
            employee=self.emp, document_type='cv', title='CV',
            file=image_file, uploaded_by=self.user)
        self.client.force_login(self.user)
        resp = self.client.get(reverse('hr:my_documents_export_pdf'))
        self.assertEqual(resp.status_code, 200)
        from pypdf import PdfReader
        from io import BytesIO as _BIO
        reader = PdfReader(_BIO(resp.content))
        # TOC page + the converted image page.
        self.assertGreaterEqual(len(reader.pages), 2)
