# engineer_calendar/views.py
import calendar
import json
import zipfile
from datetime import date, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from accounts.permissions import require_capability
from hr.models import Employee

from .models import CalendarCell
from .services import generate_draft

SOURCE_CSS = {
    'leave': 'src-leave',
    'holiday': 'src-holiday',
    'weekend': 'src-weekend',
    'wfh': 'src-wfh',
    'timesheet': 'src-timesheet',
    'manual': 'src-manual',
    'blank': 'src-blank',
}


def _resolve_year_month(request):
    """Read year/month from the query string (used for HR paging through
    past months), falling back to today when absent or invalid — same
    'never 500 on bad input' rule timesheet_export already follows."""
    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError('Month out of range')
        calendar.monthrange(year, month)
    except (ValueError, TypeError, OverflowError):
        year, month = today.year, today.month
    return year, month


def _adjacent_month(year, month, delta):
    """delta=-1 for the previous month, +1 for the next, wrapping the year."""
    month += delta
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1
    return year, month


def _clear_merge_span(employee, start_date, end_date):
    """Clear merge markers on every cell in [start_date, end_date]."""
    CalendarCell.objects.filter(
        employee=employee,
        date__gte=start_date,
        date__lte=end_date,
    ).update(merge_start_date=None, merge_end_date=None)


def _build_day_cells(emp_cells, year, month, days_in_month):
    """Build day cell dicts for one employee row, collapsing merge spans
    into a single colspan entry."""
    day_cells = []
    day = 1
    while day <= days_in_month:
        cell_date = date(year, month, day)
        cell = emp_cells.get(day)

        if (
            cell
            and cell.merge_start_date
            and cell.merge_end_date
            and cell.merge_start_date <= cell_date <= cell.merge_end_date
        ):
            # Only emit a cell on the first visible day of the merge in this month
            span_start = max(cell.merge_start_date, date(year, month, 1))
            span_end = min(cell.merge_end_date, date(year, month, days_in_month))
            if cell_date != span_start:
                day += 1
                continue
            colspan = (span_end - span_start).days + 1
            day_cells.append({
                'text': cell.display_text,
                'css_class': SOURCE_CSS.get(cell.source, 'src-blank'),
                'needs_review': cell.needs_review,
                'date': span_start.isoformat(),
                'colspan': colspan,
                'merge_start': cell.merge_start_date.isoformat(),
                'merge_end': cell.merge_end_date.isoformat(),
            })
            day += colspan
            continue

        day_cells.append({
            'text': cell.display_text if cell else '',
            'css_class': SOURCE_CSS.get(cell.source, 'src-blank') if cell else 'src-blank',
            'needs_review': cell.needs_review if cell else False,
            'date': cell_date.isoformat(),
            'colspan': 1,
            'merge_start': '',
            'merge_end': '',
        })
        day += 1
    return day_cells


def _build_calendar_workbook(year, month):
    """Build one month's Engineer Calendar workbook. Pulled out of
    export_excel so the zip download can reuse it for many months without
    duplicating any of the styling/merge logic."""
    days_in_month = calendar.monthrange(year, month)[1]

    cells = CalendarCell.objects.filter(date__year=year, date__month=month).select_related('employee')
    # Active roster, PLUS anyone (even since-deactivated) who actually has
    # attendance data for this specific month -- otherwise a departed
    # employee's history silently disappears from every past export the
    # moment they're deactivated, even though the CalendarCell rows are
    # still sitting in the database untouched.
    employees = Employee.objects.filter(
        Q(is_active=True) | Q(pk__in=cells.values_list('employee_id', flat=True))
    ).distinct()
    cell_map = {}
    for cell in cells:
        cell_map.setdefault(cell.employee_id, {})[cell.date.day] = cell

    wb = Workbook()
    ws = wb.active
    ws.title = calendar.month_abbr[month].upper()

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Cambria', size=10, bold=True)
    normal_font = Font(name='Cambria', size=10)

    FILL_WEEKEND = PatternFill('solid', fgColor='C00000')
    FILL_HOLIDAY = PatternFill('solid', fgColor='C00000')
    FILL_LEAVE = PatternFill('solid', fgColor='FBE5D6')
    FILL_WORK = PatternFill('solid', fgColor='E2EFDA')
    FILL_NAME_ROW = PatternFill('solid', fgColor='DEEBF7')
    FILL_SPACER = PatternFill('solid', fgColor='F2F2F2')

    SOURCE_FILL = {
        'weekend': FILL_WEEKEND,
        'holiday': FILL_HOLIDAY,
        'leave': FILL_LEAVE,
        'wfh': FILL_WORK,
        'timesheet': FILL_WORK,
        'manual': FILL_WORK,
    }

    ws['A1'] = 'Leap Networks Arabia'
    ws['A1'].font = Font(name='Cambria', size=11, bold=True)
    ws['A2'] = 'Al-Khobar, Saudi Arabia'
    ws['A2'].font = header_font
    ws['A4'] = 'Resource Calendar (Detailed)'
    ws['A4'].font = header_font
    ws['A5'] = 'Month'
    ws['A5'].font = header_font
    ws['B5'] = date(year, month, 1)
    ws['B5'].font = Font(name='Cambria', size=10, bold=True, color='FF0000')

    ws.column_dimensions['A'].width = 7.1
    ws.column_dimensions['B'].width = 19.4
    ws.column_dimensions['C'].width = 16.7
    for day in range(1, days_in_month + 1):
        col_letter = ws.cell(row=1, column=3 + day).column_letter
        ws.column_dimensions[col_letter].width = 9.1
    remarks_col = ws.cell(row=1, column=4 + days_in_month).column_letter
    ws.column_dimensions[remarks_col].width = 12.7

    headers = ['S. No.', 'Employee Name', 'Department'] + list(range(1, days_in_month + 1)) + ['Remarks']
    for col, val in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col, value=val)
        c.font = header_font
        c.alignment = center
        c.border = border

    for day in range(1, days_in_month + 1):
        weekday_letter = calendar.day_abbr[date(year, month, day).weekday()][0]
        c = ws.cell(row=9, column=3 + day, value=weekday_letter)
        c.font = header_font
        c.alignment = center

    row = 10
    for i, emp in enumerate(employees, start=1):
        ws.row_dimensions[row].height = 40.2

        for col, val in [(1, i), (2, emp.full_name), (3, emp.designation)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = normal_font
            c.alignment = center
            c.border = border
            c.fill = FILL_NAME_ROW

        emp_cells = cell_map.get(emp.id, {})
        day = 1
        while day <= days_in_month:
            cell = emp_cells.get(day)
            cell_date = date(year, month, day)

            if (cell and cell.merge_start_date and cell.merge_end_date
                    and cell.merge_start_date <= cell_date <= cell.merge_end_date):
                # Same idea as _build_day_cells for the on-screen grid: only
                # emit one cell for the whole merged span, clipped to what's
                # actually visible in this month.
                span_start = max(cell.merge_start_date, date(year, month, 1))
                span_end = min(cell.merge_end_date, date(year, month, days_in_month))
                start_col = 3 + span_start.day
                end_col = 3 + span_end.day

                if end_col > start_col:
                    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

                fill = SOURCE_FILL.get(cell.source)
                for col in range(start_col, end_col + 1):
                    c = ws.cell(row=row, column=col)
                    c.border = border
                    if fill:
                        c.fill = fill

                xcell = ws.cell(row=row, column=start_col, value=cell.display_text)
                xcell.font = normal_font
                xcell.alignment = center

                day = span_end.day + 1
                continue

            xcell = ws.cell(row=row, column=3 + day, value=cell.display_text if cell else '')
            xcell.font = normal_font
            xcell.alignment = center
            xcell.border = border
            if cell and cell.source in SOURCE_FILL:
                xcell.fill = SOURCE_FILL[cell.source]
            day += 1

        remarks_cell = ws.cell(row=row, column=4 + days_in_month, value='')
        remarks_cell.border = border
        remarks_cell.alignment = center

        row += 1
        ws.row_dimensions[row].height = 15
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        if emp.work_location:
            emp_type_cell = ws.cell(row=row, column=1, value=f'LNA {emp.get_work_location_display()} Staff')
        else:
            emp_type_cell = ws.cell(row=row, column=1, value='')
        emp_type_cell.font = normal_font
        if emp.joining_date:
            doj_cell = ws.cell(row=row, column=3, value=f"DOJ: {emp.joining_date.strftime('%d-%B-%Y')}")
            doj_cell.font = normal_font
        row += 1
        ws.row_dimensions[row].height = 8
        ws.cell(row=row, column=1).fill = FILL_SPACER
        row += 1

    filename = f'Engineer_Calendar_{calendar.month_abbr[month]}_{year}.xlsx'
    return wb, filename


@login_required
@require_capability('engineer_calendar.access')
def export_excel(request):
    year, month = _resolve_year_month(request)
    wb, filename = _build_calendar_workbook(year, month)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_capability('engineer_calendar.access')
def download_all_months_zip(request):
    """One zip with a full Engineer Calendar workbook for every month that
    has any CalendarCell data at all — one file per month, same format as
    the single-month Export to Excel button."""
    months = list(
        CalendarCell.objects.values_list('date__year', 'date__month')
        .distinct().order_by('-date__year', '-date__month')
    )
    if not months:
        messages.error(request, 'There is no calendar data to download yet.')
        return redirect('engineer_calendar:grid')

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for yr, mo in months:
            month_wb, month_filename = _build_calendar_workbook(yr, mo)
            month_buf = BytesIO()
            month_wb.save(month_buf)
            zf.writestr(month_filename, month_buf.getvalue())

    response = HttpResponse(buf.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename="Engineer_Calendar_All_Months.zip"'
    return response


@login_required
@require_capability('engineer_calendar.access')
@require_POST
def save_cell(request):
    try:
        payload = json.loads(request.body)
        employee_id = payload['employee_id']
        cell_date = date.fromisoformat(payload['date'])
        text = payload.get('text', '').strip()
    except (KeyError, ValueError, json.JSONDecodeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    employee = get_object_or_404(Employee, pk=employee_id)

    cell, _created = CalendarCell.objects.get_or_create(
        employee=employee,
        date=cell_date,
        defaults={'source': 'manual'},
    )

    # Editing a merged block (clicked start cell): update the whole span text.
    if cell.merge_start_date and cell.merge_end_date:
        merge_start, merge_end = cell.merge_start_date, cell.merge_end_date
        if not text:
            # Clearing the block also breaks the merge so HR can rebuild.
            current = merge_start
            while current <= merge_end:
                span_cell, _ = CalendarCell.objects.get_or_create(
                    employee=employee, date=current, defaults={'source': 'manual'})
                span_cell.display_text = ''
                span_cell.source = CalendarCell.SOURCE_MANUAL
                span_cell.merge_start_date = None
                span_cell.merge_end_date = None
                span_cell.needs_review = False
                span_cell.updated_by = request.user
                span_cell.save()
                current += timedelta(days=1)
            return JsonResponse({
                'text': '',
                'css_class': 'src-manual',
                'reload': True,
            })

        current = merge_start
        while current <= merge_end:
            span_cell, _ = CalendarCell.objects.get_or_create(
                employee=employee, date=current, defaults={'source': 'manual'})
            span_cell.display_text = text
            span_cell.source = CalendarCell.SOURCE_MANUAL
            span_cell.merge_start_date = merge_start
            span_cell.merge_end_date = merge_end
            span_cell.needs_review = False
            span_cell.updated_by = request.user
            span_cell.save()
            current += timedelta(days=1)
        return JsonResponse({
            'text': text,
            'css_class': 'src-manual',
            'reload': False,
        })

    cell.display_text = text
    cell.source = CalendarCell.SOURCE_MANUAL
    cell.merge_start_date = None
    cell.merge_end_date = None
    cell.needs_review = False
    cell.updated_by = request.user
    cell.save()

    return JsonResponse({'text': cell.display_text, 'css_class': 'src-manual', 'reload': False})


@login_required
@require_capability('engineer_calendar.access')
@require_POST
def fill_range(request):
    try:
        data = json.loads(request.body)
        employee_id = data['employee_id']
        start_date = date.fromisoformat(data['start_date'])
        end_date = date.fromisoformat(data['end_date'])
        text = (data.get('text') or '').strip()
    except (KeyError, ValueError, TypeError, json.JSONDecodeError):
        return JsonResponse({'error': 'Invalid request.'}, status=400)

    if end_date < start_date:
        return JsonResponse({'error': 'End date must be on or after start date.'}, status=400)

    employee = get_object_or_404(Employee, pk=employee_id)

    # Clear any merges that overlap this range so stamps stay consistent.
    overlapping = CalendarCell.objects.filter(
        employee=employee,
        date__gte=start_date,
        date__lte=end_date,
        merge_start_date__isnull=False,
    )
    for cell in overlapping:
        if cell.merge_start_date and cell.merge_end_date:
            _clear_merge_span(employee, cell.merge_start_date, cell.merge_end_date)

    CalendarCell.objects.filter(
        employee=employee,
        date__gte=start_date,
        date__lte=end_date,
    ).update(merge_start_date=None, merge_end_date=None)

    updated_days = []
    current_date = start_date
    while current_date <= end_date:
        cell, _ = CalendarCell.objects.get_or_create(
            employee=employee,
            date=current_date,
            defaults={'source': CalendarCell.SOURCE_MANUAL},
        )
        cell.display_text = text
        cell.source = CalendarCell.SOURCE_MANUAL
        cell.merge_start_date = start_date
        cell.merge_end_date = end_date
        cell.needs_review = False
        cell.updated_by = request.user
        cell.save()
        updated_days.append(current_date.day)
        current_date += timedelta(days=1)

    return JsonResponse({
        'updated_days': updated_days,
        'text': text,
        'css_class': 'src-manual',
        'reload': True,
    })


@login_required
@require_capability('engineer_calendar.access')
def calendar_grid(request):
    year, month = _resolve_year_month(request)
    days_in_month = calendar.monthrange(year, month)[1]

    employees = Employee.objects.filter(is_active=True)

    cells = CalendarCell.objects.filter(
        date__year=year, date__month=month
    ).select_related('employee')

    cell_map = {}
    for cell in cells:
        cell_map.setdefault(cell.employee_id, {})[cell.date.day] = cell

    rows = []
    for emp in employees:
        emp_cells = cell_map.get(emp.id, {})
        rows.append({
            'employee': emp,
            'day_cells': _build_day_cells(emp_cells, year, month, days_in_month),
        })

    weekday_letters = [
        calendar.day_abbr[date(year, month, day).weekday()][0]
        for day in range(1, days_in_month + 1)
    ]

    prev_year, prev_month = _adjacent_month(year, month, -1)
    next_year, next_month = _adjacent_month(year, month, 1)

    context = {
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'day_numbers': range(1, days_in_month + 1),
        'weekday_letters': weekday_letters,
        'rows': rows,
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
    }
    return render(request, 'engineer_calendar/calendar_grid.html', context)


@login_required
@require_capability('engineer_calendar.access')
def generate_draft_view(request):
    if request.method == 'POST':
        today = date.today()
        generate_draft(Employee.objects.filter(is_active=True), year=today.year, month=today.month)
        messages.success(request, 'Calendar draft regenerated.')
    return redirect('engineer_calendar:grid')
