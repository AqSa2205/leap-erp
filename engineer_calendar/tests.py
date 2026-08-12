"""Engineer Calendar tests — capability boundaries, CSRF, malformed input,
fill-range / merge behaviour, generate-draft protections, and export.

Style mirrors timesheets/tests.py: every endpoint and branching path gets an
explicit assertion so a future regression (IDOR-ish capability leak, silent
display_text bug, CSRF bypass, merge split) fails loudly.
"""
import json
import zipfile
from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO

from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from accounts.models import Role, User
from accounts.permissions import seed_default_permissions
from hr.models import Employee
from timesheets.models import ActivityCode, TimesheetEntry

from .models import CalendarCell
from .services import generate_draft


def _make_user_with_employee(username, role_name=Role.SALES_REP, full_name=None):
    role, _ = Role.objects.get_or_create(name=role_name)
    user = User.objects.create_user(username=username, password='testpass123', role=role)
    emp = Employee.objects.create(
        full_name=full_name or username,
        iqama_number=f'IQ-{username}',
        user=user,
        is_active=True,
    )
    return user, emp


def _json_post(client, url, payload):
    return client.post(
        url,
        data=json.dumps(payload),
        content_type='application/json',
    )


def _today_month_dates(start_day, end_day=None):
    """Build dates inside the current calendar month (what the grid uses)."""
    today = date.today()
    start = date(today.year, today.month, start_day)
    if end_day is None:
        return start
    return start, date(today.year, today.month, end_day)


class CalendarCapabilityBoundaryTests(TestCase):
    """engineer_calendar.access is admin/super_admin only — plain employees
    must get 403 on every endpoint, never a silent success."""

    def setUp(self):
        self.client = Client()
        self.admin, self.admin_emp = _make_user_with_employee(
            'caladmin', role_name=Role.ADMIN, full_name='Cal Admin')
        self.plain, self.plain_emp = _make_user_with_employee(
            'calplain', role_name=Role.SALES_REP, full_name='Cal Plain')
        seed_default_permissions()
        self.start, self.end = _today_month_dates(1, 3)

    def test_plain_employee_cannot_open_grid(self):
        self.client.login(username='calplain', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.status_code, 403)

    def test_plain_employee_cannot_export(self):
        self.client.login(username='calplain', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:export_excel'))
        self.assertEqual(resp.status_code, 403)

    def test_plain_employee_cannot_save_cell(self):
        self.client.login(username='calplain', password='testpass123')
        resp = _json_post(self.client, reverse('engineer_calendar:save_cell'), {
            'employee_id': self.plain_emp.pk,
            'date': self.start.isoformat(),
            'text': 'HACK',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(CalendarCell.objects.filter(
            employee=self.plain_emp, date=self.start).exists())

    def test_plain_employee_cannot_fill_range(self):
        self.client.login(username='calplain', password='testpass123')
        resp = _json_post(self.client, reverse('engineer_calendar:fill_range'), {
            'employee_id': self.plain_emp.pk,
            'start_date': self.start.isoformat(),
            'end_date': self.end.isoformat(),
            'text': 'HACK RANGE',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(CalendarCell.objects.filter(employee=self.plain_emp).exists())

    def test_plain_employee_cannot_generate_draft(self):
        self.client.login(username='calplain', password='testpass123')
        resp = self.client.post(reverse('engineer_calendar:generate_draft'))
        self.assertEqual(resp.status_code, 403)

    def test_admin_can_open_grid(self):
        self.client.login(username='caladmin', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Engineer Calendar')

    def test_admin_can_export(self):
        self.client.login(username='caladmin', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:export_excel'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resp['Content-Type'],
        )

    def test_anonymous_redirected_from_grid(self):
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)

    def test_anonymous_redirected_from_save_cell(self):
        resp = _json_post(self.client, reverse('engineer_calendar:save_cell'), {
            'employee_id': 1, 'date': self.start.isoformat(), 'text': 'x',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)


class CalendarCSRFEnforcementTests(TestCase):
    """Destructive calendar POSTs must reject missing CSRF tokens."""

    def setUp(self):
        self.client = Client(enforce_csrf_checks=True)
        self.admin, self.emp = _make_user_with_employee(
            'calcsrf', role_name=Role.ADMIN)
        seed_default_permissions()
        self.day = _today_month_dates(5)

    def test_save_cell_without_csrf_rejected(self):
        self.client.login(username='calcsrf', password='testpass123')
        resp = _json_post(self.client, reverse('engineer_calendar:save_cell'), {
            'employee_id': self.emp.pk,
            'date': self.day.isoformat(),
            'text': 'no csrf',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(CalendarCell.objects.filter(
            employee=self.emp, date=self.day).exists())

    def test_fill_range_without_csrf_rejected(self):
        self.client.login(username='calcsrf', password='testpass123')
        start, end = _today_month_dates(5, 7)
        resp = _json_post(self.client, reverse('engineer_calendar:fill_range'), {
            'employee_id': self.emp.pk,
            'start_date': start.isoformat(),
            'end_date': end.isoformat(),
            'text': 'no csrf',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertFalse(CalendarCell.objects.filter(employee=self.emp).exists())

    def test_generate_draft_without_csrf_rejected(self):
        self.client.login(username='calcsrf', password='testpass123')
        resp = self.client.post(reverse('engineer_calendar:generate_draft'))
        self.assertEqual(resp.status_code, 403)


class SaveCellTests(TestCase):
    """save_cell happy path + every invalid-payload / missing-employee branch."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calsave', role_name=Role.ADMIN, full_name='Save Emp')
        seed_default_permissions()
        self.client.login(username='calsave', password='testpass123')
        self.day = _today_month_dates(10)
        self.url = reverse('engineer_calendar:save_cell')

    def test_save_creates_manual_cell(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'date': self.day.isoformat(),
            'text': 'Site visit',
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data['text'], 'Site visit')
        self.assertEqual(data['css_class'], 'src-manual')
        cell = CalendarCell.objects.get(employee=self.emp, date=self.day)
        self.assertEqual(cell.display_text, 'Site visit')
        self.assertEqual(cell.source, CalendarCell.SOURCE_MANUAL)
        self.assertEqual(cell.updated_by, self.admin)
        self.assertFalse(cell.needs_review)
        self.assertIsNone(cell.merge_start_date)

    def test_save_updates_existing_cell(self):
        CalendarCell.objects.create(
            employee=self.emp, date=self.day,
            display_text='Old', source=CalendarCell.SOURCE_TIMESHEET)
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'date': self.day.isoformat(),
            'text': 'Corrected by HR',
        })
        self.assertEqual(resp.status_code, 200)
        cell = CalendarCell.objects.get(employee=self.emp, date=self.day)
        self.assertEqual(cell.display_text, 'Corrected by HR')
        self.assertEqual(cell.source, CalendarCell.SOURCE_MANUAL)

    def test_save_strips_whitespace(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'date': self.day.isoformat(),
            'text': '  padded  ',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.json()['text'], 'padded')

    def test_invalid_json_returns_400(self):
        resp = self.client.post(
            self.url, data='{not-json', content_type='application/json')
        self.assertEqual(resp.status_code, 400)

    def test_missing_employee_id_returns_400(self):
        resp = _json_post(self.client, self.url, {
            'date': self.day.isoformat(), 'text': 'x',
        })
        self.assertEqual(resp.status_code, 400)

    def test_missing_date_returns_400(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk, 'text': 'x',
        })
        self.assertEqual(resp.status_code, 400)

    def test_malformed_date_returns_400(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'date': 'not-a-date',
            'text': 'x',
        })
        self.assertEqual(resp.status_code, 400)

    def test_nonexistent_employee_returns_404(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': 999999,
            'date': self.day.isoformat(),
            'text': 'x',
        })
        self.assertEqual(resp.status_code, 404)

    def test_get_not_allowed(self):
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 405)


class FillRangeTests(TestCase):
    """fill_range must write display_text (not a phantom .text attr), stamp
    merge dates, and reject bad ranges/payloads."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calfill', role_name=Role.ADMIN, full_name='Fill Emp')
        seed_default_permissions()
        self.client.login(username='calfill', password='testpass123')
        self.url = reverse('engineer_calendar:fill_range')
        self.start, self.end = _today_month_dates(2, 5)

    def test_fill_range_sets_display_text_and_merge_on_every_day(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': self.start.isoformat(),
            'end_date': self.end.isoformat(),
            'text': 'ERP AND LEAP AI ENGINE',
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertTrue(data.get('reload'))
        self.assertEqual(data['text'], 'ERP AND LEAP AI ENGINE')
        self.assertEqual(
            sorted(data['updated_days']),
            list(range(self.start.day, self.end.day + 1)),
        )

        cells = list(CalendarCell.objects.filter(
            employee=self.emp,
            date__gte=self.start,
            date__lte=self.end,
        ).order_by('date'))
        self.assertEqual(len(cells), (self.end - self.start).days + 1)
        for cell in cells:
            self.assertEqual(cell.display_text, 'ERP AND LEAP AI ENGINE')
            self.assertEqual(cell.source, CalendarCell.SOURCE_MANUAL)
            self.assertEqual(cell.merge_start_date, self.start)
            self.assertEqual(cell.merge_end_date, self.end)
            self.assertEqual(cell.updated_by, self.admin)
            self.assertFalse(cell.needs_review)

    def test_fill_range_single_day(self):
        day = self.start
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': day.isoformat(),
            'end_date': day.isoformat(),
            'text': 'ONE DAY',
        })
        self.assertEqual(resp.status_code, 200)
        cell = CalendarCell.objects.get(employee=self.emp, date=day)
        self.assertEqual(cell.display_text, 'ONE DAY')
        self.assertEqual(cell.merge_start_date, day)
        self.assertEqual(cell.merge_end_date, day)

    def test_end_before_start_returns_400(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': self.end.isoformat(),
            'end_date': self.start.isoformat(),
            'text': 'bad',
        })
        self.assertEqual(resp.status_code, 400)
        self.assertIn('End date', resp.json()['error'])
        self.assertFalse(CalendarCell.objects.filter(employee=self.emp).exists())

    def test_invalid_json_returns_400(self):
        resp = self.client.post(
            self.url, data='{bad', content_type='application/json')
        self.assertEqual(resp.status_code, 400)

    def test_malformed_dates_return_400(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': '2026-99-99',
            'end_date': self.end.isoformat(),
            'text': 'x',
        })
        self.assertEqual(resp.status_code, 400)

    def test_missing_fields_return_400(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': self.start.isoformat(),
            # no end_date
            'text': 'x',
        })
        self.assertEqual(resp.status_code, 400)

    def test_nonexistent_employee_returns_404(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': 999999,
            'start_date': self.start.isoformat(),
            'end_date': self.end.isoformat(),
            'text': 'x',
        })
        self.assertEqual(resp.status_code, 404)

    def test_get_not_allowed(self):
        resp = self.client.get(self.url)
        self.assertEqual(resp.status_code, 405)

    def test_refill_replaces_previous_merge_cleanly(self):
        _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': self.start.isoformat(),
            'end_date': self.end.isoformat(),
            'text': 'FIRST MERGE',
        })
        # Smaller overlapping refill
        mid_start = self.start + timedelta(days=1)
        mid_end = self.start + timedelta(days=2)
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'start_date': mid_start.isoformat(),
            'end_date': mid_end.isoformat(),
            'text': 'SECOND MERGE',
        })
        self.assertEqual(resp.status_code, 200)
        for d in (mid_start, mid_end):
            cell = CalendarCell.objects.get(employee=self.emp, date=d)
            self.assertEqual(cell.display_text, 'SECOND MERGE')
            self.assertEqual(cell.merge_start_date, mid_start)
            self.assertEqual(cell.merge_end_date, mid_end)


class MergeEditTests(TestCase):
    """Editing a merged block updates the whole span; clearing breaks merge."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calmerge', role_name=Role.ADMIN)
        seed_default_permissions()
        self.client.login(username='calmerge', password='testpass123')
        self.start, self.end = _today_month_dates(8, 11)
        # Seed a merge the same way fill_range would
        current = self.start
        while current <= self.end:
            CalendarCell.objects.create(
                employee=self.emp,
                date=current,
                display_text='ORIGINAL BLOCK',
                source=CalendarCell.SOURCE_MANUAL,
                merge_start_date=self.start,
                merge_end_date=self.end,
            )
            current += timedelta(days=1)
        self.url = reverse('engineer_calendar:save_cell')

    def test_edit_merge_start_updates_all_days_in_span(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'date': self.start.isoformat(),
            'text': 'UPDATED BLOCK',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertFalse(resp.json().get('reload'))
        for cell in CalendarCell.objects.filter(
                employee=self.emp, date__gte=self.start, date__lte=self.end):
            self.assertEqual(cell.display_text, 'UPDATED BLOCK')
            self.assertEqual(cell.merge_start_date, self.start)
            self.assertEqual(cell.merge_end_date, self.end)

    def test_clearing_merge_text_breaks_merge(self):
        resp = _json_post(self.client, self.url, {
            'employee_id': self.emp.pk,
            'date': self.start.isoformat(),
            'text': '',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get('reload'))
        for cell in CalendarCell.objects.filter(
                employee=self.emp, date__gte=self.start, date__lte=self.end):
            # Clearing now regenerates from live attendance data (weekend/
            # holiday/leave/timesheet) instead of leaving the cell blank --
            # see ClearCellFallbackTests for per-day content assertions.
            # This test just confirms the stale manual text is gone and
            # the merge is broken.
            self.assertNotEqual(cell.display_text, 'ORIGINAL BLOCK')
            self.assertIsNone(cell.merge_start_date)
            self.assertIsNone(cell.merge_end_date)


class CalendarGridRenderTests(TestCase):
    """Grid must show active employees and emit colspan for merges."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calgrid', role_name=Role.ADMIN, full_name='Grid Emp Active')
        self.inactive = Employee.objects.create(
            full_name='Inactive Emp', iqama_number='IQ-inactive-cal', is_active=False)
        seed_default_permissions()
        self.client.login(username='calgrid', password='testpass123')
        self.start, self.end = _today_month_dates(12, 14)
        current = self.start
        while current <= self.end:
            CalendarCell.objects.create(
                employee=self.emp,
                date=current,
                display_text='MERGED TEXT',
                source=CalendarCell.SOURCE_MANUAL,
                merge_start_date=self.start,
                merge_end_date=self.end,
            )
            current += timedelta(days=1)

    def test_grid_shows_active_employee_not_inactive(self):
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Grid Emp Active')
        self.assertNotContains(resp, 'Inactive Emp')

    def test_grid_renders_colspan_for_merge(self):
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.status_code, 200)
        colspan = (self.end - self.start).days + 1
        self.assertContains(resp, f'colspan="{colspan}"')
        self.assertContains(resp, 'MERGED TEXT')
        # Only one visible cell for the span — text should appear once
        self.assertEqual(resp.content.decode().count('MERGED TEXT'), 1)

    def test_grid_includes_fill_range_modal(self):
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertContains(resp, 'Fill Date Range')
        self.assertContains(resp, 'fillRangeModal')

    def test_grid_still_shows_departed_employee_for_a_month_they_have_data_in(self):
        """Same roster rule as _build_calendar_workbook: a deactivated
        employee with real CalendarCell data for the month being viewed
        must still appear on screen, so the on-screen grid and the Excel
        export for that same past month never disagree on who's listed."""
        self.emp.is_active = False
        self.emp.save()
        resp = self.client.get(
            reverse('engineer_calendar:grid') + f'?year={self.start.year}&month={self.start.month}')
        self.assertContains(resp, 'Grid Emp Active')

    def test_grid_hides_departed_employee_for_a_month_they_have_no_data_in(self):
        """The fix must not turn into 'show every employee ever' -- a
        departed employee should still disappear from months where they
        genuinely have no calendar data."""
        other_month = self.start.month - 1 or 12
        other_year = self.start.year if self.start.month > 1 else self.start.year - 1
        self.emp.is_active = False
        self.emp.save()
        resp = self.client.get(
            reverse('engineer_calendar:grid') + f'?year={other_year}&month={other_month}')
        self.assertNotContains(resp, 'Grid Emp Active')


class GenerateDraftViewTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calgenview', role_name=Role.ADMIN)
        seed_default_permissions()
        self.client.login(username='calgenview', password='testpass123')
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})

    def test_post_generate_draft_redirects_and_creates_cells(self):
        today = date.today()
        # Use a weekday if possible so we get timesheet text not weekend
        workday = date(today.year, today.month, min(15, 28))
        while workday.weekday() in (4, 5):  # Fri/Sat in this codebase
            workday -= timedelta(days=1)
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='Calendar gen work', activity_code=self.code,
            hours=Decimal('10'))
        resp = self.client.post(reverse('engineer_calendar:generate_draft'))
        self.assertEqual(resp.status_code, 302)
        self.assertRedirects(resp, reverse('engineer_calendar:grid'))
        self.assertTrue(
            CalendarCell.objects.filter(employee=self.emp, date=workday).exists()
            or CalendarCell.objects.filter(employee=self.emp).exists()
        )

    def test_get_generate_draft_does_not_create_cells(self):
        before = CalendarCell.objects.count()
        resp = self.client.get(reverse('engineer_calendar:generate_draft'))
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(CalendarCell.objects.count(), before)


class GenerateDraftServiceTests(TestCase):
    """Service-layer behaviour: timesheet text, manual protection, review flags."""

    def setUp(self):
        self.user, self.emp = _make_user_with_employee(
            'calgensvc', role_name=Role.ADMIN, full_name='Svc Emp')
        seed_default_permissions()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.year, self.month = 2026, 3

    def _weekday(self, day):
        d = date(self.year, self.month, day)
        # Skip Fri=4 / Sat=5 so status falls through to timesheet text
        while d.weekday() in (4, 5):
            d += timedelta(days=1)
        return d

    def test_creates_timesheet_sourced_cell(self):
        workday = self._weekday(2)
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='Build feature', activity_code=self.code,
            hours=Decimal('10'))
        changed = generate_draft([self.emp], year=self.year, month=self.month)
        self.assertGreater(changed, 0)
        cell = CalendarCell.objects.get(employee=self.emp, date=workday)
        self.assertEqual(cell.source, CalendarCell.SOURCE_TIMESHEET)
        self.assertIn('COS_0009', cell.display_text)
        self.assertIn('Build feature', cell.display_text)

    def test_never_overwrites_manual_display_text(self):
        workday = self._weekday(3)
        CalendarCell.objects.create(
            employee=self.emp, date=workday,
            display_text='HR CORRECTION',
            source=CalendarCell.SOURCE_MANUAL,
            auto_generated_text='old auto',
        )
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='New timesheet text', activity_code=self.code,
            hours=Decimal('10'))
        generate_draft([self.emp], year=self.year, month=self.month)
        cell = CalendarCell.objects.get(employee=self.emp, date=workday)
        self.assertEqual(cell.display_text, 'HR CORRECTION')
        self.assertEqual(cell.source, CalendarCell.SOURCE_MANUAL)
        # Drift should flag review and refresh snapshot
        self.assertTrue(cell.needs_review)
        self.assertIn('COS_0009', cell.auto_generated_text)

    def test_same_description_three_days_flags_review(self):
        days = []
        d = self._weekday(2)
        while len(days) < 3:
            if d.weekday() not in (4, 5):
                days.append(d)
            d += timedelta(days=1)
        for day in days:
            TimesheetEntry.objects.create(
                employee=self.emp, date=day,
                task_description='Copy paste', activity_code=self.code,
                hours=Decimal('10'))
        generate_draft([self.emp], year=self.year, month=self.month)
        third = CalendarCell.objects.get(employee=self.emp, date=days[2])
        self.assertTrue(third.needs_review)
        self.assertIn('Same description', third.review_reason)

    def test_multiple_entries_same_day_joined(self):
        workday = self._weekday(4)
        code2, _ = ActivityCode.objects.get_or_create(
            code='COS_0005', defaults={'label': 'Onsite', 'is_active': True})
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='Office', activity_code=self.code, hours=Decimal('5'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='Site', activity_code=code2, hours=Decimal('5'))
        generate_draft([self.emp], year=self.year, month=self.month)
        cell = CalendarCell.objects.get(employee=self.emp, date=workday)
        self.assertIn(' + ', cell.display_text)
        self.assertIn('Office', cell.display_text)
        self.assertIn('Site', cell.display_text)


class ExportExcelTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calexport', role_name=Role.ADMIN, full_name='Export Emp')
        self.plain, _ = _make_user_with_employee(
            'calexportplain', role_name=Role.SALES_REP)
        seed_default_permissions()
        today = date.today()
        CalendarCell.objects.create(
            employee=self.emp,
            date=date(today.year, today.month, 1),
            display_text='Export Me',
            source=CalendarCell.SOURCE_MANUAL,
        )

    def test_export_is_valid_xlsx_with_employee_name(self):
        self.client.login(username='calexport', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:export_excel'))
        self.assertEqual(resp.status_code, 200)
        self.assertIn(
            'attachment',
            resp['Content-Disposition'],
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Flatten all cell values and look for the employee name
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend([str(v) for v in row if v is not None])
        self.assertTrue(any('Export Emp' in v for v in values))

    def test_plain_employee_cannot_export(self):
        self.client.login(username='calexportplain', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:export_excel'))
        self.assertEqual(resp.status_code, 403)


class CalendarCellModelTests(TestCase):
    def setUp(self):
        _, self.emp = _make_user_with_employee('calmodel', role_name=Role.ADMIN)

    def test_unique_together_employee_date(self):
        day = date(2026, 4, 1)
        CalendarCell.objects.create(
            employee=self.emp, date=day, display_text='one',
            source=CalendarCell.SOURCE_BLANK)
        with self.assertRaises(Exception):
            CalendarCell.objects.create(
                employee=self.emp, date=day, display_text='two',
                source=CalendarCell.SOURCE_BLANK)

    def test_is_locked_for_holiday_and_weekend_only(self):
        day = date(2026, 4, 2)
        holiday = CalendarCell.objects.create(
            employee=self.emp, date=day,
            source=CalendarCell.SOURCE_HOLIDAY, display_text='Holiday')
        self.assertTrue(holiday.is_locked)
        weekend = CalendarCell(
            employee=self.emp, date=day + timedelta(days=1),
            source=CalendarCell.SOURCE_WEEKEND)
        self.assertTrue(weekend.is_locked)
        manual = CalendarCell(
            employee=self.emp, date=day + timedelta(days=2),
            source=CalendarCell.SOURCE_MANUAL)
        self.assertFalse(manual.is_locked)
        timesheet = CalendarCell(
            employee=self.emp, date=day + timedelta(days=3),
            source=CalendarCell.SOURCE_TIMESHEET)
        self.assertFalse(timesheet.is_locked)


class SuperAdminAccessTests(TestCase):
    """super_admin is also granted engineer_calendar in DEFAULT_MODULE_ACCESS."""

    def setUp(self):
        self.client = Client()
        _make_user_with_employee('calsuper', role_name=Role.SUPER_ADMIN)
        seed_default_permissions()

    def test_super_admin_can_open_grid(self):
        self.client.login(username='calsuper', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.status_code, 200)


class EngineerCalendarMonthNavigationTests(TestCase):
    """_resolve_year_month / _adjacent_month on the HR grid and export --
    same 'never 500, always fall back to today' contract as the timesheets
    side, plus year-boundary wrapping for Prev/Next."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calnav', role_name=Role.ADMIN, full_name='Cal Nav Emp')
        seed_default_permissions()
        self.client.login(username='calnav', password='testpass123')
        CalendarCell.objects.create(
            employee=self.emp, date=date(2026, 7, 10),
            display_text='JULY CELL', source=CalendarCell.SOURCE_MANUAL)
        CalendarCell.objects.create(
            employee=self.emp, date=date(2026, 10, 10),
            display_text='OCTOBER CELL', source=CalendarCell.SOURCE_MANUAL)

    def test_explicit_month_shows_only_that_months_cells(self):
        resp = self.client.get(reverse('engineer_calendar:grid') + '?year=2026&month=7')
        self.assertContains(resp, 'JULY CELL')
        self.assertNotContains(resp, 'OCTOBER CELL')

    def test_default_no_params_uses_current_month(self):
        today = date.today()
        resp = self.client.get(reverse('engineer_calendar:grid'))
        self.assertEqual(resp.context['year'], today.year)
        self.assertEqual(resp.context['month'], today.month)

    def test_invalid_month_falls_back_to_today_not_500(self):
        today = date.today()
        resp = self.client.get(reverse('engineer_calendar:grid') + '?year=2026&month=0')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['year'], today.year)
        self.assertEqual(resp.context['month'], today.month)

    def test_garbage_year_falls_back_to_today_not_500(self):
        today = date.today()
        resp = self.client.get(reverse('engineer_calendar:grid') + '?year=xx&month=7')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['year'], today.year)
        self.assertEqual(resp.context['month'], today.month)

    def test_prev_next_wrap_december_to_january(self):
        resp = self.client.get(reverse('engineer_calendar:grid') + '?year=2026&month=12')
        self.assertEqual(resp.context['next_year'], 2027)
        self.assertEqual(resp.context['next_month'], 1)
        self.assertEqual(resp.context['prev_year'], 2026)
        self.assertEqual(resp.context['prev_month'], 11)

    def test_prev_next_wrap_january_to_december(self):
        resp = self.client.get(reverse('engineer_calendar:grid') + '?year=2026&month=1')
        self.assertEqual(resp.context['prev_year'], 2025)
        self.assertEqual(resp.context['prev_month'], 12)

    def test_export_link_carries_the_month_being_viewed(self):
        resp = self.client.get(reverse('engineer_calendar:grid') + '?year=2026&month=7')
        self.assertContains(resp, 'year=2026')
        self.assertContains(resp, 'month=7')

    def test_export_excel_respects_query_params_not_just_today(self):
        resp = self.client.get(reverse('engineer_calendar:export_excel') + '?year=2026&month=7')
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        values = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
        self.assertTrue(any('JULY CELL' in v for v in values))
        self.assertFalse(any('OCTOBER CELL' in v for v in values))

    def test_generate_draft_ignores_query_params_always_uses_today(self):
        """Deliberate design decision: Generate Draft never regenerates a
        month HR is just browsing -- it always runs against today."""
        before = CalendarCell.objects.filter(date__year=2020).count()
        self.client.post(reverse('engineer_calendar:generate_draft') + '?year=2020&month=1')
        after = CalendarCell.objects.filter(date__year=2020).count()
        self.assertEqual(before, after)


class EngineerCalendarZipDownloadTests(TestCase):
    """download_all_months_zip: role-gated, one workbook per month with
    data, and no cross-month leakage inside a single month's workbook."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calzip', role_name=Role.ADMIN, full_name='Cal Zip Emp')
        self.plain, _ = _make_user_with_employee('calzipplain', role_name=Role.SALES_REP)
        seed_default_permissions()
        CalendarCell.objects.create(
            employee=self.emp, date=date(2026, 7, 10),
            display_text='JULY CELL', source=CalendarCell.SOURCE_MANUAL)
        CalendarCell.objects.create(
            employee=self.emp, date=date(2026, 8, 10),
            display_text='AUGUST CELL', source=CalendarCell.SOURCE_MANUAL)

    def test_plain_employee_cannot_download_all_zip(self):
        self.client.login(username='calzipplain', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:download_all_zip'))
        self.assertEqual(resp.status_code, 403)

    def test_anonymous_redirected(self):
        resp = self.client.get(reverse('engineer_calendar:download_all_zip'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)

    def test_all_zip_contains_one_workbook_per_month_with_data(self):
        self.client.login(username='calzip', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:download_all_zip'))
        self.assertEqual(resp.status_code, 200)
        zf = zipfile.ZipFile(BytesIO(resp.content))
        names = zf.namelist()
        self.assertIn('Engineer_Calendar_Jul_2026.xlsx', names)
        self.assertIn('Engineer_Calendar_Aug_2026.xlsx', names)

    def test_all_zip_isolates_data_per_month(self):
        self.client.login(username='calzip', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:download_all_zip'))
        zf = zipfile.ZipFile(BytesIO(resp.content))
        july_wb = load_workbook(BytesIO(zf.read('Engineer_Calendar_Jul_2026.xlsx')))
        ws = july_wb.active
        values = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
        self.assertTrue(any('JULY CELL' in v for v in values))
        self.assertFalse(any('AUGUST CELL' in v for v in values))

    def test_no_data_redirects_with_message(self):
        CalendarCell.objects.all().delete()
        self.client.login(username='calzip', password='testpass123')
        resp = self.client.get(reverse('engineer_calendar:download_all_zip'))
        self.assertEqual(resp.status_code, 302)
        self.assertRedirects(resp, reverse('engineer_calendar:grid'))

    def test_departed_employees_data_present_in_historical_export(self):
        """Regression guard, not an open bug: _build_calendar_workbook
        derives its employee set per-month from CalendarCell data (not
        just the live is_active roster), so a deactivated employee's
        past cells still appear in an export for a month they were
        actively employed. (calendar_grid, the on-screen page, has the
        matching fix covered separately, in CalendarGridRenderTests.)"""
        self.emp.is_active = False
        self.emp.save()
        self.client.login(username='calzip', password='testpass123')
        resp = self.client.get(
            reverse('engineer_calendar:export_excel') + '?year=2026&month=7')
        wb = load_workbook(BytesIO(resp.content))
        ws = wb.active
        values = [str(v) for row in ws.iter_rows(values_only=True) for v in row if v is not None]
        self.assertTrue(any('JULY CELL' in v for v in values))


class ClearCellFallbackTests(TestCase):
    """save_cell clearing (single cell and merged block) must fall back to
    whatever generate_draft would normally produce for that day, instead
    of leaving it permanently blank until someone happens to click
    Generate Draft separately. Regression guard for the bug where clearing
    a manually-typed description wiped out the real timesheet/weekend data
    underneath it instead of revealing it."""

    def setUp(self):
        self.client = Client()
        self.admin, self.emp = _make_user_with_employee(
            'calclear', role_name=Role.ADMIN, full_name='Clear Fallback Emp')
        seed_default_permissions()
        self.client.login(username='calclear', password='testpass123')
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.save_url = reverse('engineer_calendar:save_cell')
        self.fill_url = reverse('engineer_calendar:fill_range')

    def _workday(self, year, month, start_day):
        """First non-Fri/Sat day on/after start_day (same weekend
        convention as GenerateDraftServiceTests._weekday elsewhere in this
        file: weekday() 4/5 == Fri/Sat in this codebase)."""
        d = date(year, month, start_day)
        while d.weekday() in (4, 5):
            d += timedelta(days=1)
        return d

    def test_clearing_single_manual_cell_reveals_timesheet_text(self):
        workday = self._workday(2026, 3, 2)
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='Underlying timesheet work',
            activity_code=self.code, hours=Decimal('8'))

        # HR types over the auto-generated text.
        _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': workday.isoformat(),
            'text': 'HR note',
        })
        self.assertEqual(
            CalendarCell.objects.get(employee=self.emp, date=workday).source,
            CalendarCell.SOURCE_MANUAL)

        # HR clears it.
        resp = _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': workday.isoformat(),
            'text': '',
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertIn('Underlying timesheet work', data['text'])
        self.assertEqual(data['css_class'], 'src-timesheet')

        cell = CalendarCell.objects.get(employee=self.emp, date=workday)
        self.assertEqual(cell.source, CalendarCell.SOURCE_TIMESHEET)
        self.assertIn('Underlying timesheet work', cell.display_text)

    def test_clearing_single_manual_cell_reveals_weekend(self):
        workday = self._workday(2026, 3, 1)
        weekend_day = workday
        while weekend_day.weekday() not in (4, 5):
            weekend_day += timedelta(days=1)

        _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': weekend_day.isoformat(),
            'text': 'HR note on a weekend',
        })
        resp = _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': weekend_day.isoformat(),
            'text': '',
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data['text'], 'Weekend')
        cell = CalendarCell.objects.get(employee=self.emp, date=weekend_day)
        self.assertEqual(cell.source, CalendarCell.SOURCE_WEEKEND)

    def test_clearing_single_cell_with_no_underlying_data_stays_blank(self):
        workday = self._workday(2026, 3, 3)
        _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': workday.isoformat(),
            'text': 'HR note, nothing underneath',
        })
        resp = _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': workday.isoformat(),
            'text': '',
        })
        self.assertEqual(resp.status_code, 200)
        data = resp.json()
        self.assertEqual(data['text'], '')
        self.assertEqual(data['css_class'], 'src-blank')
        cell = CalendarCell.objects.get(employee=self.emp, date=workday)
        self.assertEqual(cell.source, CalendarCell.SOURCE_BLANK)

    def test_clearing_merged_range_falls_back_per_day_not_uniformly_blank(self):
        """The regression case: a range covering both a real workday and a
        weekend day must reveal DIFFERENT text on each day after clearing.
        A 'blank the whole range' bug would make both days identical
        (empty) instead of each falling back to its own real value."""
        workday = self._workday(2026, 3, 1)
        weekend_day = workday
        while weekend_day.weekday() not in (4, 5):
            weekend_day += timedelta(days=1)
        start, end = workday, weekend_day

        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='Range underlying work',
            activity_code=self.code, hours=Decimal('8'))

        _json_post(self.client, self.fill_url, {
            'employee_id': self.emp.pk,
            'start_date': start.isoformat(),
            'end_date': end.isoformat(),
            'text': 'MERGED HR NOTE',
        })
        self.assertEqual(
            CalendarCell.objects.get(employee=self.emp, date=workday).display_text,
            'MERGED HR NOTE')

        resp = _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': start.isoformat(),
            'text': '',
        })
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.json().get('reload'))

        workday_cell = CalendarCell.objects.get(employee=self.emp, date=workday)
        weekend_cell = CalendarCell.objects.get(employee=self.emp, date=weekend_day)

        self.assertIn('Range underlying work', workday_cell.display_text)
        self.assertEqual(workday_cell.source, CalendarCell.SOURCE_TIMESHEET)
        self.assertEqual(weekend_cell.display_text, 'Weekend')
        self.assertEqual(weekend_cell.source, CalendarCell.SOURCE_WEEKEND)

        # The merge itself must still be broken on every day in the old span.
        for d in (start, end):
            c = CalendarCell.objects.get(employee=self.emp, date=d)
            self.assertIsNone(c.merge_start_date)
            self.assertIsNone(c.merge_end_date)

    def test_clearing_merge_spanning_two_months_refreshes_both_months(self):
        """The fix groups affected days by (year, month) and calls
        generate_draft once per month touched -- a range crossing a month
        boundary must correctly refresh both months, not just the first."""
        march_workday = self._workday(2026, 3, 28)
        april_workday = self._workday(2026, 4, 1)
        start, end = march_workday, april_workday

        TimesheetEntry.objects.create(
            employee=self.emp, date=march_workday,
            task_description='March side of the range',
            activity_code=self.code, hours=Decimal('8'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=april_workday,
            task_description='April side of the range',
            activity_code=self.code, hours=Decimal('8'))

        _json_post(self.client, self.fill_url, {
            'employee_id': self.emp.pk,
            'start_date': start.isoformat(),
            'end_date': end.isoformat(),
            'text': 'CROSS MONTH NOTE',
        })
        resp = _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk,
            'date': start.isoformat(),
            'text': '',
        })
        self.assertEqual(resp.status_code, 200)

        march_cell = CalendarCell.objects.get(employee=self.emp, date=march_workday)
        april_cell = CalendarCell.objects.get(employee=self.emp, date=april_workday)
        self.assertIn('March side of the range', march_cell.display_text)
        self.assertIn('April side of the range', april_cell.display_text)

    def test_clearing_single_cell_does_not_force_full_page_reload(self):
        """Single-cell clear must return reload: false -- the JS swaps the
        cell's text/class in place rather than reloading the whole grid."""
        workday = self._workday(2026, 3, 9)
        TimesheetEntry.objects.create(
            employee=self.emp, date=workday,
            task_description='In-place swap check',
            activity_code=self.code, hours=Decimal('8'))
        _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk, 'date': workday.isoformat(), 'text': 'note',
        })
        resp = _json_post(self.client, self.save_url, {
            'employee_id': self.emp.pk, 'date': workday.isoformat(), 'text': '',
        })
        self.assertFalse(resp.json().get('reload', False))