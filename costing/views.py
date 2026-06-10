from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Q
from django.template.loader import render_to_string
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.text import slugify
from django.utils import timezone
from decimal import Decimal, InvalidOperation
import logging

logger = logging.getLogger(__name__)


def _safe_filename(name, suffix='', extension=''):
    """Build a safe filename for Content-Disposition headers.

    Strips quotes, newlines, slashes, and any other characters that could
    be used to inject header values or path traversal. Falls back to
    'export' if the result would be empty.
    """
    safe = slugify(str(name or ''))[:80] or 'export'
    if suffix:
        safe = f'{safe}_{suffix}'
    if extension and not extension.startswith('.'):
        extension = f'.{extension}'
    return f'{safe}{extension}'

from .models import ExchangeRate, CostingSheet, CostingSection, CostingLineItem, TermsTemplate, ScopeOfWorkItem, ClientRemarkTemplate, ClientRemarkPair
from notifications.services import notify_users


def _conversion_rate(output_currency, rates_dict):
    """Return factor to multiply a SAR amount by to get `output_currency`.

    rates_dict holds rate_to_usd values, i.e. 1 USD = rates_dict[code] code.
    So 1 SAR = (1/sar_rate) USD = (target_rate/sar_rate) target_currency.
    """
    if output_currency == 'SAR':
        return Decimal('1')
    sar_rate = rates_dict.get('SAR')
    target_rate = rates_dict.get(output_currency)
    if sar_rate and target_rate:
        return (target_rate / sar_rate).quantize(Decimal('0.000001'))
    return Decimal('1')
from .forms import (
    CostingSheetForm, CostingSectionForm, CostingLineItemForm,
    ExchangeRateForm, CostingFilterForm, TermsTemplateForm,
    ClientRemarkTemplateForm, ClientRemarkPairFormSet,
)


def _user_can_view_sheet(user, sheet):
    """Check whether user has view access to a costing sheet."""
    if not user.is_authenticated:
        return False
    if user.is_super_admin_user:
        return True
    if sheet.created_by_id == user.id:
        return True
    if user.is_admin_user or user.is_manager_user:
        return bool(sheet.project and sheet.project.region_id == user.region_id)
    # Proposal team sees every sheet as a BOM (no pricing) regardless of region
    if getattr(user, 'is_proposal_team_user', False):
        return True
    # Finance team can view sheets in their region — needed for budgeting
    # review and ongoing finance reference.
    if getattr(user, 'is_finance_team_user', False):
        return bool(sheet.project and sheet.project.region_id == user.region_id)
    # Procurement only sees BOMs whose project is Won and in their region.
    if getattr(user, 'is_procurement_user', False):
        return bool(
            sheet.project
            and sheet.project.region_id == user.region_id
            and sheet.project.status
            and sheet.project.status.category == 'won'
        )
    return False


def _user_team(user):
    """Which team a user belongs to for change-log / notification purposes."""
    if not getattr(user, 'is_authenticated', False):
        return 'other'
    if getattr(user, 'is_proposal_team_user', False):
        return 'proposal'
    if getattr(user, 'is_finance_team_user', False):
        return 'finance'
    if getattr(user, 'is_sales_team_user', False):
        return 'sales'
    return 'other'


def _record_costing_change(sheet, actor, *, scope, scope_label, field, before, after, is_pricing=False):
    """Persist an entry to CostingSheetChangeLog and notify the opposite team.

    `actor` is the user who made the change. The opposite team (proposal
    if the actor is sales, and vice-versa) is notified so both sides stay
    aware of concurrent edits.
    """
    from .models import CostingSheetChangeLog
    from accounts.models import User

    team = _user_team(actor)
    try:
        CostingSheetChangeLog.objects.create(
            sheet=sheet,
            actor=actor if getattr(actor, 'is_authenticated', False) else None,
            actor_team=team,
            scope=scope,
            scope_label=scope_label[:255] if scope_label else '',
            field=field[:64] if field else '',
            before=str(before or '')[:2000],
            after=str(after or '')[:2000],
            is_pricing_change=bool(is_pricing),
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning('Failed to record costing change: %s', e)
        return

    # Notify the opposite team. Resolve recipients via role flags.
    try:
        if team == 'proposal':
            recipients = User.objects.filter(
                role__name__in=['sales_rep', 'manager', 'admin', 'super_admin'],
            ).filter(
                Q(pk=sheet.created_by_id) |
                Q(region_id=sheet.project.region_id if sheet.project_id else None) |
                Q(role__name='super_admin')
            ).distinct()
            opposite = 'Sales'
        elif team == 'sales':
            recipients = User.objects.filter(role__name__in=['proposal_head', 'proposal_rep']).distinct()
            opposite = 'Proposal'
        else:
            return

        actor_name = (actor.get_full_name() or actor.username) if actor else 'Someone'
        verb = f'edited {scope_label or scope} in "{sheet.title}"'
        description = (
            f'Field: {field} · {before or "—"} → {after or "—"}'
        )
        notify_users(
            recipients=list(recipients),
            verb=verb,
            actor=actor,
            target_url=f'/costing/{sheet.pk}/',
            description=description,
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning('Failed to notify cross-team on costing change: %s', e)


#: Fields proposal team members CANNOT edit — pricing-related only.
PRICING_FIELDS = frozenset({
    'base_unit_cost', 'discount_pct', 'shipping_pct', 'customs_pct',
    'finances_pct', 'installation_pct', 'margin',
    'discount_rate', 'shipping_rate', 'customs_rate',
    'finances_rate', 'installation_rate',
    'supplier_currency', 'output_currency',
    'include_optional_in_total',
})


def _field_is_pricing(field):
    """True if the named field touches pricing/rates/currency."""
    return field in PRICING_FIELDS


FINANCE_LOCKED_STAGES = ('finance_review', 'finance_approved')


def _user_can_edit_sheet(user, sheet):
    """Check whether user has edit access to a costing sheet.

    Stage-aware lock:
      * Once the sheet enters finance_review, sales/proposal/admin/manager
        are read-only — only the finance team (and super admin) can touch
        it during budgeting.
      * Once finance_approved, the sheet is fully locked even to finance —
        super admin only at that point.

    Edit rules outside the finance window: super_admin can edit anything,
    admin/manager can edit within region, creator can edit their own.
    Proposal team can edit non-pricing fields on every sheet (pricing is
    gated elsewhere).
    """
    if not user.is_authenticated:
        return False
    if user.is_super_admin_user:
        return True

    stage = sheet.workflow_stage
    if stage == 'finance_review':
        # Finance owns the sheet during budgeting.
        return bool(
            getattr(user, 'is_finance_team_user', False)
            and sheet.project
            and sheet.project.region_id == user.region_id
        )
    if stage == 'finance_approved':
        # Sheet locked once finance signs off — only super admin can edit.
        return False

    # Pre-finance stages: original ruleset.
    # Creator override is region-scoped — without this, a former employee
    # whose role was downgraded or who moved regions would retain a back-
    # door onto every sheet they ever created.
    if (
        sheet.created_by_id == user.id
        and sheet.project
        and sheet.project.region_id == user.region_id
    ):
        return True
    if user.is_admin_user and sheet.project and sheet.project.region_id == user.region_id:
        return True
    if user.is_manager_user and sheet.project and sheet.project.region_id == user.region_id:
        return True
    if getattr(user, 'is_proposal_team_user', False):
        return True
    return False


class CostingPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    def get_queryset(self):
        queryset = CostingSheet.objects.select_related('project', 'created_by').all()
        user = self.request.user
        if (
            user.is_super_admin_user
            or getattr(user, 'is_proposal_team_user', False)
        ):
            # Super admins and proposal team see every sheet (proposal team
            # is the source of BOMs, so they need full visibility).
            return queryset
        elif user.is_admin_user or user.is_manager_user:
            return queryset.filter(
                Q(created_by=user) |
                Q(project__region=user.region)
            )
        elif getattr(user, 'is_finance_team_user', False):
            # Finance team sees every sheet in their region — they need
            # the full pricing breakdown to budget.
            return queryset.filter(project__region=user.region)
        elif getattr(user, 'is_procurement_user', False):
            # Procurement sees BOMs (no pricing) for Won projects in their
            # own region only — that's when their work begins.
            return queryset.filter(
                project__region=user.region,
                project__status__category='won',
            )
        else:
            return queryset.filter(created_by=user)


class _SheetChildPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Ownership check for views that operate on a child of a CostingSheet
    (Section, LineItem). The subclass must implement get_sheet()."""

    def get_sheet(self):
        raise NotImplementedError

    def test_func(self):
        sheet = self.get_sheet()
        return _user_can_edit_sheet(self.request.user, sheet)


# ─── Costing Sheet CRUD ───────────────────────────────────────

class CostingListView(CostingPermissionMixin, ListView):
    model = CostingSheet
    template_name = 'costing/costing_list.html'
    context_object_name = 'sheets'
    paginate_by = 25

    def test_func(self):
        return True

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        stage = self.request.GET.get('stage')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(customer_reference__icontains=search) |
                Q(project__project_name__icontains=search)
            )
        if status:
            queryset = queryset.filter(status=status)
        if stage:
            queryset = queryset.filter(workflow_stage=stage)
        return queryset.order_by('-updated_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = CostingFilterForm(self.request.GET)
        context['total_count'] = self.get_queryset().count()
        context['workflow_stage_choices'] = CostingSheet.WORKFLOW_STAGE_CHOICES
        context['selected_stage'] = self.request.GET.get('stage', '')
        return context


class CostingCreateView(LoginRequiredMixin, CreateView):
    model = CostingSheet
    form_class = CostingSheetForm
    template_name = 'costing/costing_form.html'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Build a {project_pk: {customer, end_user, contact_with, ref}} map
        # so the create form can auto-fill the PDF header fields as soon as
        # the user picks a project from the dropdown. Scope mirrors the
        # form's project queryset.
        import json
        from projects.models import Project
        user = self.request.user
        proj_qs = Project.objects.select_related('region').all()
        if not user.is_super_admin_user:
            if user.is_admin_user or user.is_manager_user:
                proj_qs = proj_qs.filter(region=user.region)
            else:
                proj_qs = proj_qs.filter(owner=user)
        context['project_data_json'] = json.dumps({
            str(p.pk): {
                'customer':       p.customer or '',
                'end_user':       p.end_user or '',
                'contact_person': p.contact_with or '',
                'reference':      p.proposal_reference or '',
            }
            for p in proj_qs
        })
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Costing sheet created successfully.')
        response = super().form_valid(form)

        # Notify managers about new costing sheet
        from accounts.models import User, Role
        managers = User.objects.filter(role__name=Role.MANAGER, is_active=True)
        notify_users(
            recipients=managers,
            verb='created a new costing sheet',
            actor=self.request.user,
            target=self.object,
            target_url=reverse('costing:detail', kwargs={'pk': self.object.pk}),
            description=f'New costing sheet: {self.object.title}',
            level='info',
        )
        return response

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.object.pk})


class CostingDetailView(CostingPermissionMixin, DetailView):
    model = CostingSheet
    template_name = 'costing/costing_detail.html'
    context_object_name = 'sheet'

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        sheet = self.object

        # Pre-load exchange rates and build rates dict (single query)
        exchange_rates = list(ExchangeRate.objects.all())
        rates_dict = {r.currency_code: r.rate_to_usd for r in exchange_rates}

        # Load sections with item counts + precompute section subtotals and
        # sheet totals in a single pass (items are NOT sent to the template —
        # they are lazy-loaded via AJAX when a section is expanded).
        sections = list(sheet.sections.prefetch_related('line_items').all())

        sheet_totals = {
            'grand_total': Decimal('0'),
            'total_cost': Decimal('0'),
            'total_base_cost': Decimal('0'),
            'total_discount': Decimal('0'),
            'total_margin_amount': Decimal('0'),
            'total_shipping_amount': Decimal('0'),
            'total_customs_amount': Decimal('0'),
            'total_finances_amount': Decimal('0'),
            'total_installation_amount': Decimal('0'),
            'optional_subtotal': Decimal('0'),
        }
        include_opt = sheet.include_optional_in_total
        for section in sections:
            section_sub = {
                'subtotal': Decimal('0'),
                'total_cost': Decimal('0'),
                'base_unit_cost': Decimal('0'),
                'discount': Decimal('0'),
                'unit_cost': Decimal('0'),
                'base_unit_price': Decimal('0'),
                'base_total_price': Decimal('0'),
            }
            item_count = 0
            section_is_optional = section.is_optional
            for item in section.line_items.all():
                item.set_exchange_rates_cache(rates_dict)
                item.set_sheet_cache(sheet)
                item_count += 1
                qty = item.quantity
                # Convert per-item supplier-currency values to SAR so the
                # aggregates are unit-consistent regardless of how many
                # different supplier currencies are in play. Everything below
                # is in SAR; the template multiplies by conversion_rate once
                # at display time to show in output_currency.
                ex2sar = item.exchange_rate_to_sar
                total_cost_sar   = item.unit_cost_sar * qty
                base_cost_sar    = (item.base_unit_cost * ex2sar) * qty
                discount_sar     = (item.discount_amount * ex2sar) * qty

                section_sub['subtotal'] += item.final_total_price
                section_sub['total_cost'] += total_cost_sar
                section_sub['base_unit_cost'] += base_cost_sar
                section_sub['discount'] += discount_sar
                section_sub['unit_cost'] += total_cost_sar
                section_sub['base_unit_price'] += item.base_unit_price * qty
                section_sub['base_total_price'] += item.base_total_price

                # Optional sections are tracked separately and excluded from the
                # main sheet totals unless the include flag is on.
                if section_is_optional:
                    sheet_totals['optional_subtotal'] += item.final_total_price
                    if not include_opt:
                        continue

                ucs = item.unit_cost_sar
                sheet_totals['grand_total'] += item.final_total_price
                sheet_totals['total_cost'] += total_cost_sar
                sheet_totals['total_base_cost'] += base_cost_sar
                sheet_totals['total_discount'] += discount_sar
                sheet_totals['total_margin_amount'] += item.base_total_price - total_cost_sar
                sheet_totals['total_shipping_amount'] += ucs * item.effective_shipping_pct * qty
                sheet_totals['total_customs_amount'] += ucs * item.effective_customs_pct * qty
                sheet_totals['total_finances_amount'] += ucs * item.effective_finances_pct * qty
                sheet_totals['total_installation_amount'] += ucs * item.effective_installation_pct * qty

            section._subtotals = {k: v.quantize(Decimal('0.01')) for k, v in section_sub.items()}
            section.item_count_cached = item_count

        sheet._totals = {k: v.quantize(Decimal('0.01')) for k, v in sheet_totals.items()}

        # Compute SAR → output_currency conversion rate
        conversion_rate = _conversion_rate(sheet.output_currency, rates_dict)

        context['sections'] = sections
        context['section_form'] = CostingSectionForm()
        context['lineitem_form'] = CostingLineItemForm()
        context['exchange_rates'] = exchange_rates
        context['conversion_rate'] = conversion_rate

        # Terms templates for PDF selection
        all_terms = TermsTemplate.objects.all()
        selected_ids = set(sheet.selected_terms.values_list('pk', flat=True))
        terms_by_category = {}
        for cat_value, cat_label in TermsTemplate.CATEGORY_CHOICES:
            terms_by_category[cat_label] = [
                {'template': t, 'selected': t.pk in selected_ids}
                for t in all_terms if t.category == cat_value
            ]
        context['terms_by_category'] = terms_by_category

        # Client remark Q&A bundles for PDF selection
        selected_remark_ids = set(sheet.selected_client_remarks.values_list('pk', flat=True))
        context['client_remark_templates'] = [
            {
                'template': t,
                'selected': t.pk in selected_remark_ids,
                'pair_count': t.pairs.count(),
            }
            for t in ClientRemarkTemplate.objects.all().prefetch_related('pairs')
        ]

        # Additional contacts: only Sales-team users (the only people whose
        # contact info is acceptable to show on a client-facing PDF). Region
        # scoping mirrors the rest of the app — super admins see every
        # region; everyone else only sees their own region's sales team.
        from accounts.models import User, Role
        selected_contact_ids = set(sheet.additional_contacts.values_list('pk', flat=True))
        sales_role_names = [
            Role.SALES_REP, Role.MANAGER, Role.ADMIN, Role.SUPER_ADMIN,
        ]
        sales_users = (
            User.objects
            .filter(is_active=True, role__name__in=sales_role_names)
            .exclude(pk=sheet.created_by_id or 0)
            .select_related('region', 'role')
            .order_by('region__name', 'first_name', 'last_name', 'email')
        )
        if not self.request.user.is_super_admin_user:
            if self.request.user.region_id:
                sales_users = sales_users.filter(region=self.request.user.region)
            else:
                sales_users = sales_users.none()
        contacts_by_region = {}
        for u in sales_users:
            region_name = u.region.name if u.region else 'Unassigned region'
            contacts_by_region.setdefault(region_name, []).append({
                'user': u,
                'selected': u.pk in selected_contact_ids,
            })
        context['additional_contacts_by_region'] = contacts_by_region

        # Scope of Work items (A.2 section). Sum line totals if any items
        # exist; fall back to the sheet's flat scope_of_work_total field.
        sow_items = list(sheet.scope_of_work_items.all())
        context['sow_items'] = sow_items
        sow_total_oc = (
            sum((i.total_price for i in sow_items), Decimal('0'))
            if sow_items else sheet.scope_of_work_total
        )
        context['sow_total'] = sow_total_oc

        # Contract total = (line-item grand total in output currency) + SOW.
        # The detail page Grand Total column already reflects the user's
        # "include optional in total" toggle via sheet.grand_total, so we just
        # add SOW on top to mirror the PDF's MAIN — TOTAL CONTRACT PRICE row.
        grand_total_oc = (sheet.grand_total * conversion_rate).quantize(Decimal('0.01'))
        context['contract_total'] = grand_total_oc + Decimal(str(sow_total_oc))

        # PDF revisions (saved snapshots from previous Export PDF clicks)
        context['pdf_revisions'] = sheet.pdf_revisions.select_related('created_by').all()

        # Recent change log (for cross-team visibility).
        # The UI lets the user dial how many entries are shown via ?change_log_limit=N.
        # Default 20, clamp 1..500 to avoid rendering thousands of rows.
        try:
            _cl_limit = int(self.request.GET.get('change_log_limit') or 20)
        except (TypeError, ValueError):
            _cl_limit = 20
        _cl_limit = max(1, min(_cl_limit, 500))
        _cl_qs = sheet.change_log.select_related('actor')
        context['change_log_total'] = _cl_qs.count()
        context['change_log_limit'] = _cl_limit
        context['change_log'] = _cl_qs.all()[:_cl_limit]

        # Vendor quotes attached to this sheet
        context['vendor_quotes'] = sheet.vendor_quotes.select_related('uploaded_by').all()

        return context


class CostingUpdateView(CostingPermissionMixin, UpdateView):
    model = CostingSheet
    form_class = CostingSheetForm
    template_name = 'costing/costing_form.html'

    def test_func(self):
        return _user_can_edit_sheet(self.request.user, self.get_object())

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        old_status = CostingSheet.objects.get(pk=self.object.pk).status
        messages.success(self.request, 'Costing sheet updated successfully.')
        response = super().form_valid(form)

        # Notify on draft → final
        if old_status == 'draft' and self.object.status == 'final':
            from accounts.models import User, Role
            recipients = set(User.objects.filter(role__name__in=[Role.SUPER_ADMIN, Role.ADMIN], is_active=True))
            if self.object.project and self.object.project.owner:
                recipients.add(self.object.project.owner)
            notify_users(
                recipients=recipients,
                verb=f'finalized costing sheet "{self.object.title}"',
                actor=self.request.user,
                target=self.object,
                target_url=reverse('costing:detail', kwargs={'pk': self.object.pk}),
                description=f'Costing sheet "{self.object.title}" has been finalized.',
                level='success',
                send_email=True,
            )
        return response

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.object.pk})


class CostingDeleteView(CostingPermissionMixin, DeleteView):
    model = CostingSheet
    template_name = 'costing/costing_confirm_delete.html'
    success_url = reverse_lazy('costing:list')

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user:
            return True
        sheet = self.get_object()
        if user.is_admin_user and sheet.project and sheet.project.region == user.region:
            return True
        return sheet.created_by == user

    def form_valid(self, form):
        messages.success(self.request, 'Costing sheet deleted successfully.')
        return super().form_valid(form)


# ─── Section CRUD ─────────────────────────────────────────────

class SectionCreateView(_SheetChildPermissionMixin, CreateView):
    model = CostingSection
    form_class = CostingSectionForm
    template_name = 'costing/section_form.html'

    def get_sheet(self):
        return get_object_or_404(CostingSheet, pk=self.kwargs['sheet_pk'])

    def form_valid(self, form):
        sheet = self.get_sheet()
        form.instance.costing_sheet = sheet
        messages.success(self.request, 'Section added successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.kwargs['sheet_pk']})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sheet'] = self.get_sheet()
        return context


class SectionUpdateView(_SheetChildPermissionMixin, UpdateView):
    model = CostingSection
    form_class = CostingSectionForm
    template_name = 'costing/section_form.html'

    def get_sheet(self):
        return self.get_object().costing_sheet

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.object.costing_sheet.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Section updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sheet'] = self.object.costing_sheet
        return context


class SectionDeleteView(_SheetChildPermissionMixin, DeleteView):
    model = CostingSection
    template_name = 'costing/costing_confirm_delete.html'

    def get_sheet(self):
        return self.get_object().costing_sheet

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.object.costing_sheet.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Section deleted successfully.')
        return super().form_valid(form)


# ─── Line Item CRUD ───────────────────────────────────────────

class LineItemCreateView(_SheetChildPermissionMixin, CreateView):
    model = CostingLineItem
    form_class = CostingLineItemForm
    template_name = 'costing/lineitem_form.html'

    def _get_section(self):
        return get_object_or_404(CostingSection, pk=self.kwargs['section_pk'])

    def get_sheet(self):
        return self._get_section().costing_sheet

    def form_valid(self, form):
        form.instance.section = self._get_section()
        messages.success(self.request, 'Line item added successfully.')
        return super().form_valid(form)

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self._get_section().costing_sheet.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['section'] = self._get_section()
        return context


class LineItemUpdateView(_SheetChildPermissionMixin, UpdateView):
    model = CostingLineItem
    form_class = CostingLineItemForm
    template_name = 'costing/lineitem_form.html'

    def get_sheet(self):
        return self.get_object().section.costing_sheet

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.object.section.costing_sheet.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Line item updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['section'] = self.object.section
        return context


class LineItemDeleteView(_SheetChildPermissionMixin, DeleteView):
    model = CostingLineItem
    template_name = 'costing/costing_confirm_delete.html'

    def get_sheet(self):
        return self.get_object().section.costing_sheet

    def get_success_url(self):
        return reverse('costing:detail', kwargs={'pk': self.object.section.costing_sheet.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Line item deleted successfully.')
        return super().form_valid(form)


# ─── Exchange Rates ───────────────────────────────────────────

class ExchangeRateListView(LoginRequiredMixin, ListView):
    model = ExchangeRate
    template_name = 'costing/exchange_rates.html'
    context_object_name = 'rates'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = ExchangeRateForm()
        return context


class ExchangeRateCreateView(LoginRequiredMixin, CreateView):
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'costing/exchange_rates.html'
    success_url = reverse_lazy('costing:exchange_rates')

    def form_valid(self, form):
        messages.success(self.request, 'Exchange rate added.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Error adding exchange rate.')
        return redirect('costing:exchange_rates')


class ExchangeRateUpdateView(LoginRequiredMixin, UpdateView):
    model = ExchangeRate
    form_class = ExchangeRateForm
    template_name = 'costing/exchange_rate_form.html'
    success_url = reverse_lazy('costing:exchange_rates')

    def form_valid(self, form):
        messages.success(self.request, 'Exchange rate updated.')
        return super().form_valid(form)


class ExchangeRateDeleteView(LoginRequiredMixin, DeleteView):
    model = ExchangeRate
    template_name = 'costing/costing_confirm_delete.html'
    success_url = reverse_lazy('costing:exchange_rates')

    def form_valid(self, form):
        messages.success(self.request, 'Exchange rate deleted.')
        return super().form_valid(form)


# ─── Terms Template CRUD ─────────────────────────────────────

def _safe_next_url(request, candidate):
    """Only return `candidate` if it's a same-origin path. Falls back to None.

    Prevents the next-redirect from being abused as an open redirect to an
    external site.
    """
    if not candidate:
        return None
    return candidate if url_has_allowed_host_and_scheme(
        url=candidate,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ) else None


class TermsTemplateListView(LoginRequiredMixin, ListView):
    model = TermsTemplate
    template_name = 'costing/terms_templates.html'
    context_object_name = 'templates'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = TermsTemplateForm()
        # Pass through ?next=… so the form (and Back button) can return the
        # user to wherever they came from (e.g. a PO edit page).
        context['next_url'] = _safe_next_url(self.request, self.request.GET.get('next'))
        return context


class TermsTemplateCreateView(LoginRequiredMixin, CreateView):
    model = TermsTemplate
    form_class = TermsTemplateForm
    template_name = 'costing/terms_templates.html'
    success_url = reverse_lazy('costing:terms_templates')

    def get_success_url(self):
        nxt = _safe_next_url(self.request, self.request.POST.get('next'))
        return nxt or str(self.success_url)

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Terms template added.')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Error adding terms template.')
        nxt = _safe_next_url(self.request, self.request.POST.get('next'))
        return redirect(nxt or 'costing:terms_templates')


@login_required
@require_POST
def ajax_create_terms_template(request):
    """Create a TermsTemplate inline (no page navigation).

    Used by the PO form / detail page so the "create the template I
    realised I need right now" flow doesn't blow away unsaved PO data via a
    full-page navigation. Role-gated — only sales / procurement / admin
    can add to the global library so junior accounts can't spam or
    deface terms that a real PO might pull in.
    """
    user = request.user
    if not (
        user.is_super_admin_user
        or user.is_admin_user
        or user.is_manager_user
        or getattr(user, 'is_sales_rep_user', False)
        or getattr(user, 'is_procurement_user', False)
    ):
        return JsonResponse(
            {'error': 'Your role cannot add to the terms library.'},
            status=403,
        )
    name = (request.POST.get('name') or '').strip()
    category = (request.POST.get('category') or '').strip()
    content = (request.POST.get('content') or '').strip()
    if not name or not category or not content:
        return JsonResponse({'error': 'name, category, and content are all required.'}, status=400)
    if category not in dict(TermsTemplate.CATEGORY_CHOICES):
        return JsonResponse({'error': 'Invalid category.'}, status=400)
    # Cap length to prevent runaway text in the global library.
    if len(name) > 255 or len(content) > 10_000:
        return JsonResponse({'error': 'Name or content too long.'}, status=400)
    template = TermsTemplate.objects.create(
        name=name, category=category, content=content,
        created_by=request.user,
    )
    preview = ' '.join(template.content.split()[:15])
    if len(template.content.split()) > 15:
        preview += '…'
    return JsonResponse({
        'pk': template.pk,
        'name': template.name,
        'category': template.category,
        'category_label': template.get_category_display(),
        'content': template.content,
        'content_preview': preview,
    })


class TermsTemplateUpdateView(LoginRequiredMixin, UpdateView):
    model = TermsTemplate
    form_class = TermsTemplateForm
    template_name = 'costing/terms_template_form.html'
    success_url = reverse_lazy('costing:terms_templates')

    def form_valid(self, form):
        messages.success(self.request, 'Terms template updated.')
        return super().form_valid(form)


class TermsTemplateDeleteView(LoginRequiredMixin, DeleteView):
    model = TermsTemplate
    template_name = 'costing/costing_confirm_delete.html'
    success_url = reverse_lazy('costing:terms_templates')

    def form_valid(self, form):
        messages.success(self.request, 'Terms template deleted.')
        return super().form_valid(form)


@login_required
@require_POST
def ajax_toggle_term(request, pk):
    """Toggle a TermsTemplate on/off for a costing sheet."""
    # Permission check uses an unlocked read
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    term_id = request.POST.get('term_id')
    if not term_id:
        return JsonResponse({'error': 'term_id required'}, status=400)
    term = get_object_or_404(TermsTemplate, pk=term_id)

    # Lock the sheet row inside a transaction so concurrent toggles
    # against the same M2M can't race.
    with transaction.atomic():
        locked_sheet = CostingSheet.objects.select_for_update().get(pk=sheet.pk)
        if locked_sheet.selected_terms.filter(pk=term.pk).exists():
            locked_sheet.selected_terms.remove(term)
            selected = False
        else:
            locked_sheet.selected_terms.add(term)
            selected = True
    return JsonResponse({'ok': True, 'selected': selected})


class ClientRemarkTemplateListView(LoginRequiredMixin, ListView):
    model = ClientRemarkTemplate
    template_name = 'costing/client_remark_templates.html'
    context_object_name = 'templates'

    def get_queryset(self):
        return ClientRemarkTemplate.objects.all().prefetch_related('pairs')


@login_required
def client_remark_template_edit(request, pk=None):
    """Create or edit a ClientRemarkTemplate (bundle of Q&A pairs).
    Single function-based view so we can save the parent + inline formset
    in one transaction."""
    if pk is None:
        template = ClientRemarkTemplate(created_by=request.user)
        is_new = True
    else:
        template = get_object_or_404(ClientRemarkTemplate, pk=pk)
        is_new = False

    if request.method == 'POST':
        form = ClientRemarkTemplateForm(request.POST, instance=template)
        # Bind the formset against an in-memory parent for new templates;
        # we save the parent first below, then re-bind if necessary.
        if form.is_valid():
            with transaction.atomic():
                template = form.save(commit=False)
                if is_new and not template.created_by:
                    template.created_by = request.user
                template.save()
                formset = ClientRemarkPairFormSet(request.POST, instance=template)
                if formset.is_valid():
                    formset.save()
                    messages.success(
                        request,
                        'Client remarks bundle created.' if is_new else 'Client remarks bundle updated.',
                    )
                    return redirect('costing:client_remark_templates')
                # Rollback parent save if the formset rejects.
                transaction.set_rollback(True)
        else:
            formset = ClientRemarkPairFormSet(request.POST, instance=template)
        for field, errs in form.errors.items():
            messages.error(request, f'{field}: {errs[0]}')
    else:
        form = ClientRemarkTemplateForm(instance=template)
        formset = ClientRemarkPairFormSet(instance=template)

    return render(request, 'costing/client_remark_template_form.html', {
        'form': form,
        'formset': formset,
        'template_obj': template if not is_new else None,
        'is_new': is_new,
    })


class ClientRemarkTemplateDeleteView(LoginRequiredMixin, DeleteView):
    model = ClientRemarkTemplate
    template_name = 'costing/costing_confirm_delete.html'
    success_url = reverse_lazy('costing:client_remark_templates')

    def form_valid(self, form):
        messages.success(self.request, 'Client remarks bundle deleted.')
        return super().form_valid(form)


@login_required
@require_POST
def ajax_toggle_client_remark(request, pk):
    """Toggle a ClientRemarkTemplate on/off for a costing sheet."""
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    remark_id = request.POST.get('remark_id')
    if not remark_id:
        return JsonResponse({'error': 'remark_id required'}, status=400)
    remark = get_object_or_404(ClientRemarkTemplate, pk=remark_id)

    with transaction.atomic():
        locked_sheet = CostingSheet.objects.select_for_update().get(pk=sheet.pk)
        if locked_sheet.selected_client_remarks.filter(pk=remark.pk).exists():
            locked_sheet.selected_client_remarks.remove(remark)
            selected = False
        else:
            locked_sheet.selected_client_remarks.add(remark)
            selected = True
    return JsonResponse({'ok': True, 'selected': selected})


@login_required
@require_POST
def ajax_toggle_additional_contact(request, pk):
    """Toggle a User in the costing sheet's additional_contacts M2M
    (people credited in the PDF header alongside the auto-creator).
    Only Sales-team users are eligible — the PDF is client-facing."""
    from accounts.models import User, Role
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    user_id = request.POST.get('user_id')
    if not user_id:
        return JsonResponse({'error': 'user_id required'}, status=400)
    user = get_object_or_404(User, pk=user_id)

    sales_role_names = [Role.SALES_REP, Role.MANAGER, Role.ADMIN, Role.SUPER_ADMIN]
    if not user.role or user.role.name not in sales_role_names:
        return JsonResponse(
            {'error': 'Only Sales-team users can be added as PDF contacts'},
            status=400,
        )

    # Region scoping: non-super-admins can only credit users from their
    # own region. Mirrors the picker filter on the costing detail view.
    if not request.user.is_super_admin_user:
        if not request.user.region_id or user.region_id != request.user.region_id:
            return JsonResponse(
                {'error': 'You can only add contacts from your own region'},
                status=403,
            )

    with transaction.atomic():
        locked_sheet = CostingSheet.objects.select_for_update().get(pk=sheet.pk)
        if locked_sheet.additional_contacts.filter(pk=user.pk).exists():
            locked_sheet.additional_contacts.remove(user)
            selected = False
        else:
            locked_sheet.additional_contacts.add(user)
            selected = True
    return JsonResponse({'ok': True, 'selected': selected})


# ─── AJAX Section Items (lazy-load) ──────────────────────────

@login_required
def ajax_section_items(request, pk):
    """Return HTML fragment with a section's line items for lazy-loading."""
    section = get_object_or_404(CostingSection, pk=pk)
    sheet = section.costing_sheet
    if not _user_can_view_sheet(request.user, sheet):
        return HttpResponse(status=403)
    items = list(section.line_items.all())

    exchange_rates = list(ExchangeRate.objects.all())
    rates_dict = {r.currency_code: r.rate_to_usd for r in exchange_rates}

    for item in items:
        item.set_exchange_rates_cache(rates_dict)
        item.set_sheet_cache(sheet)

    conversion_rate = _conversion_rate(sheet.output_currency, rates_dict)

    html = render_to_string('costing/_section_items.html', {
        'section': section,
        'items': items,
        'exchange_rates': exchange_rates,
        'conversion_rate': conversion_rate,
        'output_currency': sheet.output_currency,
    }, request=request)
    return HttpResponse(html)


# ─── AJAX Inline Editing ──────────────────────────────────────

@login_required
@require_POST
def ajax_update_sheet_params(request, pk):
    # Permission check via an unlocked read first.
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    field = request.POST.get('field')
    value = request.POST.get('value', '').strip()

    allowed = ('margin', 'discount_rate', 'shipping_rate', 'customs_rate', 'finances_rate', 'installation_rate', 'output_currency', 'default_supplier_currency', 'include_optional_in_total')
    if field not in allowed:
        return JsonResponse({'error': 'Invalid field'}, status=400)

    # Proposal team cannot edit pricing/rate fields.
    if getattr(request.user, 'is_proposal_team_user', False) and _field_is_pricing(field):
        return JsonResponse({'error': 'Pricing fields are read-only for proposal team — only the sales team can change these.'}, status=403)

    # Validate the value before opening a write transaction
    if field in ('output_currency', 'default_supplier_currency'):
        new_value = (value or 'SAR').upper()[:10]
    elif field == 'include_optional_in_total':
        new_value = value in ('1', 'true', 'True', 'on')
    else:
        try:
            new_value = Decimal(value)
        except (InvalidOperation, ValueError):
            return JsonResponse({'error': 'Invalid number'}, status=400)
        # Clamp percentage fields to sane bounds. Margin must be in
        # [0, 99] to keep the selling-price formula 1/(1-margin) stable.
        # Other rates are allowed in [0, 100].
        if field == 'margin':
            new_value = max(Decimal('0'), min(new_value, Decimal('99')))
        else:
            new_value = max(Decimal('0'), min(new_value, Decimal('100')))

    previous = getattr(sheet, field, '')
    # Single atomic UPDATE WHERE pk=X SET field=value. This only writes
    # the one field, so concurrent edits to other fields on the same sheet
    # cannot clobber each other (the previous bug came from sheet.save()
    # writing the entire in-memory snapshot).
    CostingSheet.objects.filter(pk=sheet.pk).update(**{field: new_value})
    if str(previous) != str(new_value):
        _record_costing_change(
            sheet, request.user,
            scope='sheet',
            scope_label=sheet.title or f'Sheet {sheet.pk}',
            field=field, before=previous, after=new_value,
            is_pricing=_field_is_pricing(field),
        )
    return JsonResponse({'ok': True})


@login_required
@require_POST
def ajax_update_exchange_rate(request, pk):
    # Exchange rates are global; only super admin or admin may modify them.
    user = request.user
    if not (user.is_super_admin_user or user.is_admin_user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    # Make sure the row exists (404 if not)
    if not ExchangeRate.objects.filter(pk=pk).exists():
        return JsonResponse({'error': 'Not found'}, status=404)
    value = request.POST.get('value', '').strip()
    try:
        new_value = Decimal(value)
    except (InvalidOperation, ValueError):
        return JsonResponse({'error': 'Invalid number'}, status=400)
    # Atomic single-field UPDATE — no read-modify-write race.
    ExchangeRate.objects.filter(pk=pk).update(rate_to_usd=new_value)
    return JsonResponse({'ok': True})


@login_required
@require_POST
def ajax_update_item_margin(request, pk):
    item = get_object_or_404(CostingLineItem, pk=pk)
    if not _user_can_edit_sheet(request.user, item.section.costing_sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if getattr(request.user, 'is_proposal_team_user', False):
        return JsonResponse({'error': 'Pricing fields are read-only for proposal team.'}, status=403)
    value = request.POST.get('margin', '').strip()
    if not value:
        new_margin = None  # Clear to use sheet margin
    else:
        try:
            new_margin = Decimal(value)
        except (InvalidOperation, ValueError):
            return JsonResponse({'error': 'Invalid number'}, status=400)
        # Clamp to [0, 99] to keep the selling-price formula stable.
        new_margin = max(Decimal('0'), min(new_margin, Decimal('99')))
    previous = item.margin
    # Atomic single-field UPDATE — no read-modify-write race against
    # other concurrent edits to different fields on the same line item.
    CostingLineItem.objects.filter(pk=pk).update(margin=new_margin)
    if str(previous) != str(new_margin):
        _record_costing_change(
            item.section.costing_sheet, request.user,
            scope='item',
            scope_label=f'{item.item_number or "item"} — {item.description[:60] if item.description else ""}',
            field='margin', before=previous, after=new_margin, is_pricing=True,
        )
    return JsonResponse({'ok': True})


@login_required
@require_POST
def ajax_update_item_field(request, pk):
    # Permission check on a normal (unlocked) read.
    item = get_object_or_404(CostingLineItem.objects.select_related('section__costing_sheet'), pk=pk)
    sheet = item.section.costing_sheet
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    field = request.POST.get('field', '').strip()
    value = request.POST.get('value', '').strip()

    numeric_fields = ('base_unit_cost', 'discount_pct', 'shipping_pct', 'customs_pct', 'finances_pct', 'installation_pct', 'margin')
    text_fields = ('description', 'make', 'model_number', 'vendor_name', 'unit', 'item_number')
    allowed_fields = numeric_fields + text_fields + ('supplier_currency', 'quantity')
    if field not in allowed_fields:
        return JsonResponse({'error': 'Invalid field'}, status=400)

    # Proposal team cannot edit pricing/rate fields.
    if getattr(request.user, 'is_proposal_team_user', False) and _field_is_pricing(field):
        return JsonResponse({'error': 'Pricing fields are read-only for proposal team — only the sales team can change these.'}, status=403)

    # Validate / coerce the value before opening the transaction.
    if field == 'supplier_currency':
        new_value = value if value else 'SAR'
    elif field in text_fields:
        # Truncate to the model field's max_length so a long paste from
        # Excel doesn't blow up Postgres with "value too long for type
        # character varying(N)".
        try:
            max_len = CostingLineItem._meta.get_field(field).max_length
        except Exception:
            max_len = None
        new_value = value if max_len is None else value[:max_len]
    elif field == 'quantity':
        try:
            new_value = Decimal(value) if value else Decimal('1')
        except (InvalidOperation, ValueError):
            return JsonResponse({'error': 'Invalid quantity'}, status=400)
        if new_value < 0:
            new_value = Decimal('0')
    elif field == 'margin' and not value:
        new_value = None
    else:
        try:
            new_value = Decimal(value) if value else Decimal('0')
        except (InvalidOperation, ValueError):
            return JsonResponse({'error': 'Invalid number'}, status=400)
        # Clamp percentage fields to sane bounds.
        if field == 'margin':
            new_value = max(Decimal('0'), min(new_value, Decimal('99')))
        elif field in ('discount_pct', 'shipping_pct', 'customs_pct', 'finances_pct', 'installation_pct'):
            new_value = max(Decimal('0'), min(new_value, Decimal('100')))

    # Capture previous value for the change log.
    previous_value = getattr(item, field, '')

    # Atomic single-field UPDATE — only writes the one field, so it cannot
    # clobber concurrent edits to other fields on the same line item.
    # Re-fetch after the update so computed properties reflect the fresh
    # committed state.
    CostingLineItem.objects.filter(pk=pk).update(**{field: new_value})
    item = CostingLineItem.objects.select_related('section__costing_sheet').get(pk=pk)

    # Log + cross-team notify when the value actually changed
    if str(previous_value) != str(new_value):
        _record_costing_change(
            sheet, request.user,
            scope='item',
            scope_label=f'{item.item_number or "item"} — {item.description[:60] if item.description else ""}',
            field=field,
            before=previous_value,
            after=new_value,
            is_pricing=_field_is_pricing(field),
        )

    # Recompute values with caches and return them. Cost columns stay in
    # the item's supplier currency; price columns convert SAR → output_currency
    # so the updated cells match what the template originally rendered.
    rates_dict = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}
    item.set_exchange_rates_cache(rates_dict)
    item.set_sheet_cache(item.section.costing_sheet)
    conv = _conversion_rate(sheet.output_currency, rates_dict)
    q = Decimal('0.01')

    return JsonResponse({
        'ok': True,
        'computed': {
            'unit_cost':        str(item.unit_cost),
            'total_cost':       str(item.total_cost),
            'base_unit_price':  str((item.base_unit_price * conv).quantize(q)),
            'base_total_price': str((item.base_total_price * conv).quantize(q)),
            'final_unit_price': str((item.final_unit_price * conv).quantize(q)),
            'final_total_price':str((item.final_total_price * conv).quantize(q)),
        },
    })


# ─── Paste from Excel (AJAX) ─────────────────────────────────

@login_required
@require_POST
def ajax_paste_line_items(request, pk):
    """Bulk-create line items from TSV pasted from Excel.

    Expected columns (in order, tab-separated):
      Description | Make | Model | Qty | Unit | Base Cost | Currency

    Missing trailing columns are allowed — they default to blank/0/EA/SAR.
    """
    section = get_object_or_404(CostingSection, pk=pk)
    sheet = section.costing_sheet
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    raw = request.POST.get('data', '').strip()
    if not raw:
        return JsonResponse({'error': 'No data pasted'}, status=400)

    # Case-insensitive unit map: lowercase input -> canonical code
    valid_units = {code.lower(): code for code, _ in CostingLineItem.UNIT_CHOICES}
    valid_currencies = {r.currency_code for r in ExchangeRate.objects.all()}

    existing_count = section.line_items.count()
    created_items = []
    errors = []

    lines = [ln for ln in raw.splitlines() if ln.strip()]
    # Detect & skip an obvious header row
    if lines and 'description' in lines[0].lower() and '\t' in lines[0]:
        lines = lines[1:]

    with transaction.atomic():
        for idx, line in enumerate(lines):
            cols = line.split('\t')
            # Pad to 7 columns
            cols += [''] * (7 - len(cols))
            description = cols[0].strip()
            if not description:
                errors.append(f'Row {idx + 1}: empty description — skipped')
                continue
            make = cols[1].strip()
            model_number = cols[2].strip()
            qty_raw = cols[3].strip().replace(',', '') or '1'
            unit = cols[4].strip() or 'EA'
            cost_raw = cols[5].strip().replace(',', '') or '0'
            currency = cols[6].strip().upper() or (sheet.default_supplier_currency or 'SAR')

            try:
                qty = Decimal(qty_raw) if qty_raw else Decimal('1')
            except (InvalidOperation, ValueError):
                qty = Decimal('1')
            try:
                cost = Decimal(cost_raw) if cost_raw else Decimal('0')
            except (InvalidOperation, ValueError):
                cost = Decimal('0')

            unit = valid_units.get(unit.lower(), 'EA')
            if currency not in valid_currencies:
                currency = 'SAR'

            next_number = f'{section.section_number}.{existing_count + len(created_items) + 1}'
            next_order = existing_count + len(created_items)

            item = CostingLineItem.objects.create(
                section=section,
                item_number=next_number,
                description=description[:500],
                make=make[:100],
                model_number=model_number[:100],
                quantity=qty,
                unit=unit,
                supplier_currency=currency,
                base_unit_cost=cost,
                order=next_order,
            )
            created_items.append(item)

    # Render all created rows
    exchange_rates = list(ExchangeRate.objects.all())
    rates_dict = {r.currency_code: r.rate_to_usd for r in exchange_rates}
    conversion_rate = _conversion_rate(sheet.output_currency, rates_dict)

    rows_html = []
    for item in created_items:
        item.set_exchange_rates_cache(rates_dict)
        item.set_sheet_cache(sheet)
        rows_html.append(render_to_string('costing/_section_item_row.html', {
            'item': item,
            'section': section,
            'exchange_rates': exchange_rates,
            'output_currency': sheet.output_currency,
            'conversion_rate': conversion_rate,
        }))

    return JsonResponse({
        'ok': True,
        'created': len(created_items),
        'errors': errors,
        'html': ''.join(rows_html),
    })


# ─── Inline Add Row (AJAX) ───────────────────────────────────

def _renumber_section(section, ordered_items=None):
    """Resequence a section's line items to a clean 1..N order.

    Sets each item's ``order`` to its index and ``item_number`` to
    ``"<section_number>.<idx+1>"``. Pass ``ordered_items`` to control the
    sequence (e.g. after splicing a freshly inserted row into place);
    otherwise the section's current ordering is used.

    Returns a list of ``{'pk', 'item_number'}`` dicts for every row whose
    number/order actually changed, so callers can patch the DOM in place
    instead of re-rendering the whole section.
    """
    if ordered_items is None:
        ordered_items = list(section.line_items.all().order_by('order', 'item_number'))
    changes = []
    for idx, item in enumerate(ordered_items):
        new_number = f'{section.section_number}.{idx + 1}'
        if item.item_number != new_number or item.order != idx:
            CostingLineItem.objects.filter(pk=item.pk).update(
                item_number=new_number, order=idx,
            )
            item.item_number = new_number
            item.order = idx
            changes.append({'pk': item.pk, 'item_number': new_number})
    return changes


@login_required
@require_POST
def ajax_add_line_item(request, pk):
    """Create a blank line item in a section and return the rendered row."""
    section = get_object_or_404(CostingSection, pk=pk)
    sheet = section.costing_sheet
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    # Auto-generate the next item number: section_number + "." + next index
    existing_count = section.line_items.count()
    next_number = f'{section.section_number}.{existing_count + 1}'
    next_order = existing_count

    item = CostingLineItem.objects.create(
        section=section,
        item_number=next_number,
        description='',
        quantity=Decimal('1'),
        unit='EA',
        supplier_currency=sheet.default_supplier_currency or 'SAR',
        base_unit_cost=Decimal('0'),
        order=next_order,
    )

    # Render the single row with the cached sheet / rates so computed fields work
    exchange_rates = list(ExchangeRate.objects.all())
    rates_dict = {r.currency_code: r.rate_to_usd for r in exchange_rates}
    item.set_exchange_rates_cache(rates_dict)
    item.set_sheet_cache(sheet)

    conversion_rate = _conversion_rate(sheet.output_currency, rates_dict)
    html = render_to_string('costing/_section_item_row.html', {
        'item': item,
        'section': section,
        'exchange_rates': exchange_rates,
        'output_currency': sheet.output_currency,
        'conversion_rate': conversion_rate,
    })
    return JsonResponse({'ok': True, 'html': html, 'item_pk': item.pk})


@login_required
@require_POST
def ajax_insert_line_item_after(request, pk):
    """Insert a blank line item directly below ``pk`` and renumber the section.

    Unlike :func:`ajax_add_line_item` (append-to-end), this places the new row
    immediately after the anchor item, then resequences every row in the
    section so numbering stays clean (1.2, 1.3, 1.4 ...). Returns the rendered
    new row plus the list of existing rows whose numbers shifted, so the client
    can patch them in place without re-rendering the section.
    """
    anchor = get_object_or_404(CostingLineItem, pk=pk)
    section = anchor.section
    sheet = section.costing_sheet
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    with transaction.atomic():
        items = list(section.line_items.all().order_by('order', 'item_number'))
        anchor_idx = next(i for i, it in enumerate(items) if it.pk == anchor.pk)

        item = CostingLineItem.objects.create(
            section=section,
            item_number='',  # assigned by _renumber_section below
            description='',
            quantity=Decimal('1'),
            unit='EA',
            supplier_currency=sheet.default_supplier_currency or 'SAR',
            base_unit_cost=Decimal('0'),
            order=anchor_idx + 1,
        )
        # Splice the new row into place, then resequence the whole section.
        items.insert(anchor_idx + 1, item)
        changes = _renumber_section(section, items)

    # The new row carries its own (already correct) number; the rest are the
    # existing rows whose numbers shifted.
    renumbered = [c for c in changes if c['pk'] != item.pk]

    exchange_rates = list(ExchangeRate.objects.all())
    rates_dict = {r.currency_code: r.rate_to_usd for r in exchange_rates}
    item.set_exchange_rates_cache(rates_dict)
    item.set_sheet_cache(sheet)

    conversion_rate = _conversion_rate(sheet.output_currency, rates_dict)
    html = render_to_string('costing/_section_item_row.html', {
        'item': item,
        'section': section,
        'exchange_rates': exchange_rates,
        'output_currency': sheet.output_currency,
        'conversion_rate': conversion_rate,
    })
    return JsonResponse({
        'ok': True, 'html': html, 'item_pk': item.pk, 'renumbered': renumbered,
    })


# ─── Bulk Delete Items (dedicated page) ──────────────────────

@login_required
def bulk_delete_page(request, pk):
    """Page with checkboxes to select and delete multiple line items."""
    section = get_object_or_404(CostingSection, pk=pk)
    sheet = section.costing_sheet
    if not _user_can_edit_sheet(request.user, sheet):
        messages.error(request, 'Permission denied.')
        return redirect('costing:detail', pk=sheet.pk)

    items = section.line_items.all()

    if request.method == 'POST':
        selected_ids = request.POST.getlist('item_ids')
        if not selected_ids:
            messages.warning(request, 'No items selected.')
            return redirect('costing:bulk_delete_page', pk=pk)

        with transaction.atomic():
            deleted = CostingLineItem.objects.filter(
                pk__in=selected_ids, section=section
            ).delete()[0]

            # Renumber remaining items
            _renumber_section(section)

        messages.success(request, f'Deleted {deleted} item(s). Remaining items renumbered.')
        return redirect('costing:detail', pk=sheet.pk)

    return render(request, 'costing/bulk_delete.html', {
        'section': section,
        'sheet': sheet,
        'items': items,
    })


# ─── Section Rate Override AJAX ───────────────────────────────

@login_required
@require_POST
def ajax_update_section_rate(request, pk):
    """Update a rate field on a CostingSection (margin, discount, shipping, etc.)."""
    section = get_object_or_404(CostingSection, pk=pk)
    if not _user_can_edit_sheet(request.user, section.costing_sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if getattr(request.user, 'is_proposal_team_user', False):
        return JsonResponse({'error': 'Section rate overrides are read-only for proposal team.'}, status=403)

    field = request.POST.get('field', '').strip()
    value = request.POST.get('value', '').strip()

    allowed = ('margin', 'discount_rate', 'shipping_rate', 'customs_rate', 'finances_rate', 'installation_rate')
    if field not in allowed:
        return JsonResponse({'error': 'Invalid field'}, status=400)

    if not value:
        new_value = None  # Clear to fall back to sheet rate
    else:
        try:
            new_value = Decimal(value)
        except (InvalidOperation, ValueError):
            return JsonResponse({'error': 'Invalid number'}, status=400)
        if field == 'margin':
            new_value = max(Decimal('0'), min(new_value, Decimal('99')))
        else:
            new_value = max(Decimal('0'), min(new_value, Decimal('100')))

    previous = getattr(section, field, None)
    CostingSection.objects.filter(pk=pk).update(**{field: new_value})
    if str(previous) != str(new_value):
        _record_costing_change(
            section.costing_sheet, request.user,
            scope='section',
            scope_label=f'Section {section.section_number} — {section.title}',
            field=field, before=previous, after=new_value, is_pricing=True,
        )
    return JsonResponse({'ok': True})


# ─── Scope of Work AJAX ──────────────────────────────────────

@login_required
@require_POST
def ajax_add_sow_item(request, pk):
    """Add a Scope of Work line item to a costing sheet."""
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    description = request.POST.get('description', '').strip()
    if not description:
        return JsonResponse({'error': 'Description required'}, status=400)

    try:
        qty = Decimal(request.POST.get('quantity', '1'))
        total_price = Decimal(request.POST.get('total_price', '0'))
    except (InvalidOperation, ValueError):
        return JsonResponse({'error': 'Invalid number'}, status=400)

    uom = request.POST.get('uom', 'LOT').strip() or 'LOT'
    price_text = request.POST.get('price_text', '').strip()
    last_order = sheet.scope_of_work_items.order_by('-order').values_list('order', flat=True).first() or 0
    last_sn = sheet.scope_of_work_items.order_by('-serial_number').values_list('serial_number', flat=True).first() or 0

    item = ScopeOfWorkItem.objects.create(
        costing_sheet=sheet,
        serial_number=last_sn + 1,
        description=description,
        quantity=qty,
        uom=uom,
        total_price=total_price,
        price_text=price_text,
        order=last_order + 1,
    )
    return JsonResponse({
        'ok': True,
        'item': {
            'pk': item.pk,
            'serial_number': item.serial_number,
            'description': item.description,
            'quantity': str(item.quantity),
            'uom': item.uom,
            'display_price': item.display_price,
        },
    })


@login_required
@require_POST
def ajax_delete_sow_item(request, pk):
    """Delete a Scope of Work line item."""
    item = get_object_or_404(ScopeOfWorkItem, pk=pk)
    if not _user_can_edit_sheet(request.user, item.costing_sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    item.delete()
    return JsonResponse({'ok': True})


@login_required
@require_POST
def ajax_update_sow_item(request, pk):
    """Update an existing Scope of Work line item."""
    item = get_object_or_404(ScopeOfWorkItem, pk=pk)
    if not _user_can_edit_sheet(request.user, item.costing_sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    description = request.POST.get('description', '').strip()
    if not description:
        return JsonResponse({'error': 'Description required'}, status=400)

    try:
        qty = Decimal(request.POST.get('quantity', '1'))
        total_price = Decimal(request.POST.get('total_price', '0'))
    except (InvalidOperation, ValueError):
        return JsonResponse({'error': 'Invalid number'}, status=400)

    uom = request.POST.get('uom', 'LOT').strip() or 'LOT'
    price_text = request.POST.get('price_text', '').strip()

    item.description = description
    item.quantity = qty
    item.uom = uom
    item.total_price = total_price
    item.price_text = price_text
    item.save()

    return JsonResponse({
        'ok': True,
        'item': {
            'pk': item.pk,
            'serial_number': item.serial_number,
            'description': item.description,
            'quantity': str(item.quantity),
            'uom': item.uom,
            'total_price': str(item.total_price),
            'price_text': item.price_text,
            'display_price': item.display_price,
        },
    })


@login_required
@require_POST
def ajax_upload_vendor_quote(request, pk):
    """Attach a vendor quote file to a costing sheet (sales team only)."""
    from .models import VendorQuote
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if getattr(request.user, 'is_proposal_team_user', False):
        return JsonResponse({'error': 'Vendor quotes are sales-team only.'}, status=403)

    uploaded = request.FILES.get('file')
    if not uploaded:
        return JsonResponse({'error': 'No file received'}, status=400)
    if uploaded.size > 150 * 1024 * 1024:
        return JsonResponse({'error': 'File too large (max 150MB)'}, status=400)

    import os
    ext = os.path.splitext(uploaded.name)[1].lower().lstrip('.')
    allowed_ext = {'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',
                   'png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp'}
    if ext not in allowed_ext:
        return JsonResponse({'error': f'File type ".{ext}" not allowed.'}, status=400)

    vendor_name = (request.POST.get('vendor_name') or '').strip() or 'Unnamed Vendor'
    quote_ref = (request.POST.get('quote_reference') or '').strip()
    amount_raw = (request.POST.get('amount') or '').strip()
    currency = (request.POST.get('currency') or sheet.default_supplier_currency or 'SAR').upper()
    notes = (request.POST.get('notes') or '').strip()
    amount = None
    if amount_raw:
        try:
            amount = Decimal(amount_raw)
        except (InvalidOperation, ValueError):
            pass

    vq = VendorQuote.objects.create(
        sheet=sheet,
        vendor_name=vendor_name[:255],
        quote_reference=quote_ref[:100],
        file=uploaded,
        original_filename=uploaded.name[:255],
        amount=amount,
        currency=currency[:10],
        notes=notes,
        uploaded_by=request.user if request.user.is_authenticated else None,
    )
    return JsonResponse({
        'ok': True,
        'id': vq.pk,
        'vendor_name': vq.vendor_name,
        'quote_reference': vq.quote_reference,
        'amount': str(vq.amount) if vq.amount else None,
        'currency': vq.currency,
        'original_filename': vq.original_filename,
        'url': vq.file.url,
        'uploaded_at': vq.uploaded_at.strftime('%d %b %Y %H:%M'),
    })


@login_required
@require_POST
def ajax_delete_vendor_quote(request, pk):
    """Delete a vendor quote (sales team only)."""
    from .models import VendorQuote
    vq = get_object_or_404(VendorQuote, pk=pk)
    sheet = vq.sheet
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if getattr(request.user, 'is_proposal_team_user', False):
        return JsonResponse({'error': 'Vendor quotes are sales-team only.'}, status=403)
    try:
        vq.file.delete(save=False)
    except Exception:
        pass
    vq.delete()
    return JsonResponse({'ok': True})


@login_required
def vendor_quote_list(request):
    """Standalone list of every vendor quote the user can see."""
    from .models import VendorQuote
    user = request.user
    qs = VendorQuote.objects.select_related('sheet', 'sheet__project', 'uploaded_by').all()
    # Scope to sheets the user can view
    if not user.is_super_admin_user:
        from django.db.models import Q as _Q
        if user.is_admin_user or user.is_manager_user:
            qs = qs.filter(_Q(sheet__created_by=user) | _Q(sheet__project__region=user.region))
        elif getattr(user, 'is_proposal_team_user', False):
            # Proposal team does not see vendor quotes
            qs = qs.none()
        else:
            qs = qs.filter(sheet__created_by=user)

    search = request.GET.get('search', '').strip()
    if search:
        from django.db.models import Q as _Q
        qs = qs.filter(
            _Q(vendor_name__icontains=search) |
            _Q(quote_reference__icontains=search) |
            _Q(sheet__title__icontains=search) |
            _Q(notes__icontains=search)
        )
    return render(request, 'costing/vendor_quote_list.html', {
        'quotes': qs.order_by('-uploaded_at'),
        'search': search,
    })


@login_required
@require_POST
def clear_changelog(request, pk):
    """Wipe every CostingSheetChangeLog row for one sheet.

    Reserved for super-admin and admin users - a destructive action that
    erases cross-team audit history, so keep the gate tight.
    """
    from .models import CostingSheetChangeLog
    sheet = get_object_or_404(CostingSheet, pk=pk)
    user = request.user
    if not (getattr(user, 'is_super_admin_user', False) or getattr(user, 'is_admin_user', False)):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    deleted, _ = CostingSheetChangeLog.objects.filter(sheet=sheet).delete()
    messages.success(request, f'Cleared {deleted} change-log entries.')
    return redirect('costing:detail', pk=sheet.pk)


@login_required
def commercial_proposal_pdf_list(request):
    """List every Commercial Proposal PDF generated from a costing sheet.

    A Commercial Proposal PDF is a CostingSheetRevision with
    export_format='pdf'. Shown under Proposals → Commercial → Commercial
    Proposal so users can browse / download prior exports without
    wading through the costing list itself.
    """
    from .models import CostingSheetRevision
    user = request.user
    qs = CostingSheetRevision.objects.filter(export_format='pdf').select_related(
        'sheet', 'sheet__project', 'created_by',
    )
    if not user.is_super_admin_user:
        from django.db.models import Q as _Q
        if user.is_admin_user or user.is_manager_user:
            qs = qs.filter(_Q(sheet__created_by=user) | _Q(sheet__project__region=user.region))
        elif getattr(user, 'is_proposal_team_user', False):
            # Proposal team sees every sheet (same as CostingPermissionMixin)
            pass
        else:
            qs = qs.filter(sheet__created_by=user)

    search = (request.GET.get('search') or '').strip()
    if search:
        from django.db.models import Q as _Q
        qs = qs.filter(
            _Q(revision_label__icontains=search) |
            _Q(sheet__title__icontains=search) |
            _Q(sheet__project__proposal_reference__icontains=search) |
            _Q(sheet__project__project_name__icontains=search) |
            _Q(change_summary__icontains=search)
        )
    return render(request, 'costing/commercial_proposal_list.html', {
        'revisions': qs.order_by('-created_at'),
        'search': search,
    })


@login_required
@require_POST
def ajax_apply_default_currency(request, pk):
    """Bulk-set every line item's supplier_currency to the sheet's default.

    One-click helper for teams that set up a sheet in (e.g.) SAR and then
    realise they actually enter costs in GBP — switches every existing item
    without manual re-entry. Audit: one change-log row records the bulk op.
    """
    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    if getattr(request.user, 'is_proposal_team_user', False):
        return JsonResponse({'error': 'Currency changes are sales-team only.'}, status=403)

    target_currency = (sheet.default_supplier_currency or 'SAR').upper()
    # Update every line item across every section in this sheet
    updated = CostingLineItem.objects.filter(section__costing_sheet=sheet).exclude(
        supplier_currency=target_currency,
    ).update(supplier_currency=target_currency)

    if updated:
        _record_costing_change(
            sheet, request.user,
            scope='sheet',
            scope_label=sheet.title or f'Sheet {sheet.pk}',
            field='bulk_supplier_currency',
            before='(mixed)',
            after=f'{target_currency} on {updated} item(s)',
            is_pricing=True,
        )
    return JsonResponse({'ok': True, 'updated': updated, 'currency': target_currency})


@login_required
@require_POST
def costing_workflow_transition(request, pk):
    """Advance the workflow stage of a costing sheet.

    Transitions (button-driven, not automatic):
    - bom_in_progress → ready_for_costing  (proposal hands over to sales)
    - ready_for_costing → costing_in_progress  (sales starts pricing)
    - costing_in_progress → finalized  (sales finalises)
    - any → bom_in_progress  (super-admin only, rewind)
    """
    sheet = get_object_or_404(CostingSheet, pk=pk)
    user = request.user
    if not _user_can_view_sheet(user, sheet):
        messages.error(request, 'Permission denied.')
        return redirect('costing:list')

    action = request.POST.get('action', '').strip()
    note = (request.POST.get('note') or '').strip()
    now = timezone.now()

    transitions = {
        'handover': {
            'from': {'bom_in_progress'},
            'to':   'ready_for_costing',
            'allowed_teams': {'proposal', 'sales'},  # anyone on either team
            'sets':  lambda: {
                'workflow_stage': 'ready_for_costing',
                'handed_over_at': now,
                'handed_over_by': user,
                'handover_note':  note or sheet.handover_note,
            },
            'verb': 'handed the BOM over to the sales team',
            'notify_team': 'sales',
        },
        'start_costing': {
            'from': {'ready_for_costing'},
            'to':   'costing_in_progress',
            'allowed_teams': {'sales'},
            'sets':  lambda: {
                'workflow_stage': 'costing_in_progress',
                'costing_started_at': now,
                'costing_started_by': user,
            },
            'verb': 'started costing on the sheet',
            'notify_team': 'proposal',
        },
        'finalize': {
            'from': {'costing_in_progress', 'ready_for_costing'},
            'to':   'finalized',
            'allowed_teams': {'sales'},
            'sets':  lambda: {
                'workflow_stage': 'finalized',
                'finalized_at': now,
                'finalized_by': user,
            },
            'verb': 'finalized the sheet',
            'notify_team': 'proposal',
        },
        'send_to_finance': {
            'from': {'finalized'},
            'to':   'finance_review',
            'allowed_teams': {'sales'},
            'requires_won_project': True,
            'sets':  lambda: {
                'workflow_stage': 'finance_review',
                'finance_review_at': now,
                'finance_review_by': user,
            },
            'verb': 'sent the sheet to finance for budgeting',
            'notify_team': 'finance',
        },
        'finance_approve': {
            'from': {'finance_review'},
            'to':   'finance_approved',
            'allowed_teams': {'finance'},
            'sets':  lambda: {
                'workflow_stage': 'finance_approved',
                'finance_approved_at': now,
                'finance_approved_by': user,
                'finance_note': note or sheet.finance_note,
            },
            'verb': 'approved the sheet — released for procurement',
            'notify_team': 'procurement',
        },
        'reopen': {
            'from': {
                'finalized', 'costing_in_progress', 'ready_for_costing',
                'finance_review', 'finance_approved',
            },
            'to':   'bom_in_progress',
            'allowed_teams': {'sales'},  # super-admin override below
            'sets':  lambda: {'workflow_stage': 'bom_in_progress'},
            'verb': 're-opened the sheet (back to BOM)',
            'notify_team': 'proposal',
        },
    }

    tx = transitions.get(action)
    if not tx:
        messages.error(request, f'Unknown action: {action}')
        return redirect('costing:detail', pk=sheet.pk)

    # Actor-team check
    team = _user_team(user)
    if team not in tx['allowed_teams'] and not user.is_super_admin_user:
        messages.error(request, 'Your role is not allowed to perform this action.')
        return redirect('costing:detail', pk=sheet.pk)

    # Stage pre-condition check
    if sheet.workflow_stage not in tx['from']:
        messages.error(request, f'This action requires the sheet to be in one of: {", ".join(tx["from"])}. Current: {sheet.workflow_stage}.')
        return redirect('costing:detail', pk=sheet.pk)

    # Reopen guard: rewinding a sheet that finance has touched would
    # silently discard the budget approval. Restrict that origin to finance
    # team or super-admin so sales can't roll back finance work unilaterally.
    if action == 'reopen' and sheet.workflow_stage in ('finance_review', 'finance_approved'):
        if not user.is_super_admin_user and team != 'finance':
            messages.error(
                request,
                'Reopening a sheet from a finance stage requires the finance '
                'team or a super administrator.',
            )
            return redirect('costing:detail', pk=sheet.pk)

    # Won-project gate (currently only used by send_to_finance — finance
    # budgeting only kicks in once the project is actually awarded).
    if tx.get('requires_won_project'):
        proj = sheet.project
        if not proj or not proj.status or proj.status.category != 'won':
            messages.error(
                request,
                'This action is only available once the linked project is in a Won status.',
            )
            return redirect('costing:detail', pk=sheet.pk)

    # Apply
    updates = tx['sets']()
    CostingSheet.objects.filter(pk=sheet.pk).update(**updates)
    sheet.refresh_from_db()

    # Log to the change log
    _record_costing_change(
        sheet, user,
        scope='sheet',
        scope_label=sheet.title or f'Sheet {sheet.pk}',
        field='workflow_stage',
        before='',
        after=sheet.get_workflow_stage_display(),
        is_pricing=False,
    )

    # Notify the opposite team
    try:
        from accounts.models import User as _U
        if tx['notify_team'] == 'proposal':
            recipients = _U.objects.filter(role__name__in=['proposal_head', 'proposal_rep'])
        elif tx['notify_team'] == 'finance':
            recipients = _U.objects.filter(role__name__in=['finance_head', 'finance_manager', 'finance_rep']).filter(
                Q(region_id=sheet.project.region_id if sheet.project_id else None) |
                Q(role__name='finance_head')
            ).distinct()
        elif tx['notify_team'] == 'procurement':
            recipients = _U.objects.filter(role__name__in=['procurement_mgr', 'procurement_off']).filter(
                region_id=sheet.project.region_id if sheet.project_id else None
            ).distinct()
        else:
            recipients = _U.objects.filter(role__name__in=['sales_rep', 'manager', 'admin', 'super_admin']).filter(
                Q(pk=sheet.created_by_id) |
                Q(region_id=sheet.project.region_id if sheet.project_id else None) |
                Q(role__name='super_admin')
            ).distinct()

        notify_users(
            recipients=list(recipients),
            verb=f'{tx["verb"]}: "{sheet.title}"',
            actor=user,
            target_url=f'/costing/{sheet.pk}/',
            description=note or f'Stage is now: {sheet.get_workflow_stage_display()}',
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning('Failed to notify workflow transition: %s', e)

    messages.success(request, f'Sheet is now: {sheet.get_workflow_stage_display()}')
    return redirect('costing:detail', pk=sheet.pk)


@require_POST
def delete_costing_revision(request, pk):
    """Delete a saved costing-sheet revision (PDF or Excel)."""
    from .models import CostingSheetRevision
    rev = get_object_or_404(CostingSheetRevision, pk=pk)
    sheet = rev.sheet
    if not _user_can_edit_sheet(request.user, sheet):
        messages.error(request, 'Permission denied.')
        return redirect('costing:detail', pk=sheet.pk)
    label = rev.revision_label
    try:
        rev.file.delete(save=False)
    except Exception:
        pass
    rev.delete()
    messages.success(request, f'Revision {label} deleted.')
    return redirect('costing:detail', pk=sheet.pk)


# ── Revision helpers: snapshot + diff ─────────────────────────

def _build_costing_snapshot(sheet):
    """Capture key state of a costing sheet for change tracking.
    Kept compact: top-level numbers + per-section/item counts +
    rate values. Enough to produce a useful diff summary."""
    from decimal import Decimal
    def _d(v):
        return None if v is None else str(v)

    sections_data = []
    total_items = 0
    for section in sheet.sections.all():
        items = list(section.line_items.all())
        total_items += len(items)
        sections_data.append({
            'section_number': section.section_number,
            'title': section.title,
            'is_optional': section.is_optional,
            'item_count': len(items),
            'subtotal': _d(section.subtotal),
        })

    return {
        'snapshot_version': 1,
        'totals': {
            'grand_total': _d(sheet.grand_total),
            'total_cost': _d(sheet.total_cost),
            'total_margin_amount': _d(sheet.total_margin_amount),
            'total_discount': _d(sheet.total_discount),
            'total_shipping_amount': _d(sheet.total_shipping_amount),
            'total_customs_amount': _d(sheet.total_customs_amount),
            'total_finances_amount': _d(sheet.total_finances_amount),
            'total_installation_amount': _d(sheet.total_installation_amount),
            'optional_subtotal': _d(sheet.optional_subtotal),
        },
        'sheet_rates': {
            'margin': _d(sheet.margin),
            'discount_rate': _d(sheet.discount_rate),
            'shipping_rate': _d(sheet.shipping_rate),
            'customs_rate': _d(sheet.customs_rate),
            'finances_rate': _d(sheet.finances_rate),
            'installation_rate': _d(sheet.installation_rate),
        },
        'output_currency': sheet.output_currency,
        'include_optional_in_total': sheet.include_optional_in_total,
        'status': sheet.status,
        'section_count': len(sections_data),
        'item_count': total_items,
        'sections': sections_data,
    }


def _compute_costing_diff(prev, current):
    """Compare two costing snapshots. Returns (summary_text, details_list).
    `prev` may be None (means first revision)."""
    if not prev:
        return '(initial export)', []

    from decimal import Decimal, InvalidOperation
    details = []

    def _dec(v):
        try: return Decimal(v)
        except (TypeError, ValueError, InvalidOperation): return None

    # Totals
    for key in ('grand_total', 'total_cost', 'total_margin_amount',
                'total_discount', 'optional_subtotal'):
        old_v = (prev.get('totals') or {}).get(key)
        new_v = (current.get('totals') or {}).get(key)
        if old_v != new_v:
            od, nd = _dec(old_v), _dec(new_v)
            if od is not None and nd is not None:
                delta = nd - od
                sign = '+' if delta >= 0 else ''
                details.append({
                    'field': key, 'before': str(od), 'after': str(nd),
                    'delta': f'{sign}{delta}', 'label': key.replace('_', ' ').title(),
                })
            else:
                details.append({'field': key, 'before': old_v, 'after': new_v, 'label': key.replace('_', ' ').title()})

    # Rates
    for key in ('margin', 'discount_rate', 'shipping_rate',
                'customs_rate', 'finances_rate', 'installation_rate'):
        old_v = (prev.get('sheet_rates') or {}).get(key)
        new_v = (current.get('sheet_rates') or {}).get(key)
        if old_v != new_v:
            details.append({
                'field': f'rate.{key}', 'before': old_v, 'after': new_v,
                'label': key.replace('_', ' ').title() + ' %',
            })

    # Currency / flags
    for key in ('output_currency', 'status'):
        if prev.get(key) != current.get(key):
            details.append({'field': key, 'before': prev.get(key), 'after': current.get(key), 'label': key.replace('_', ' ').title()})
    if prev.get('include_optional_in_total') != current.get('include_optional_in_total'):
        details.append({
            'field': 'include_optional_in_total',
            'before': prev.get('include_optional_in_total'),
            'after': current.get('include_optional_in_total'),
            'label': 'Include Optional in Total',
        })

    # Section / item counts
    if prev.get('section_count') != current.get('section_count'):
        delta = (current.get('section_count') or 0) - (prev.get('section_count') or 0)
        sign = '+' if delta >= 0 else ''
        details.append({
            'field': 'section_count', 'before': prev.get('section_count'),
            'after': current.get('section_count'), 'delta': f'{sign}{delta}',
            'label': 'Sections',
        })
    if prev.get('item_count') != current.get('item_count'):
        delta = (current.get('item_count') or 0) - (prev.get('item_count') or 0)
        sign = '+' if delta >= 0 else ''
        details.append({
            'field': 'item_count', 'before': prev.get('item_count'),
            'after': current.get('item_count'), 'delta': f'{sign}{delta}',
            'label': 'Line items',
        })

    # Build a short summary line (top ~3 changes, with deltas where available)
    if not details:
        return 'no changes vs previous', []
    summary_bits = []
    for d in details[:3]:
        if 'delta' in d:
            summary_bits.append(f"{d['label']} {d['delta']}")
        else:
            summary_bits.append(f"{d['label']}: {d.get('before')} → {d.get('after')}")
    summary = ' · '.join(summary_bits)
    if len(details) > 3:
        summary += f' · +{len(details) - 3} more'
    return summary, details


def _save_costing_revision(sheet, file_bytes, filename, export_format, user):
    """Helper: save a new CostingSheetRevision with snapshot + diff."""
    from django.core.files.base import ContentFile
    from .models import CostingSheetRevision
    snapshot = _build_costing_snapshot(sheet)
    prev = CostingSheetRevision.objects.filter(sheet=sheet).order_by('-created_at').first()
    summary, details = _compute_costing_diff(prev.snapshot if prev else None, snapshot)
    rev_label = CostingSheetRevision.next_label_for(sheet)
    rev = CostingSheetRevision(
        sheet=sheet,
        revision_label=rev_label,
        export_format=export_format,
        original_filename=filename,
        created_by=user if getattr(user, 'is_authenticated', False) else None,
        snapshot=snapshot,
        change_summary=summary[:500],
        change_details=details,
    )
    rev.file.save(f'{sheet.pk}_{rev_label}_{filename}', ContentFile(file_bytes), save=True)
    return rev


# ─── Excel Export ─────────────────────────────────────────────

@login_required
def costing_export_excel(request, pk):
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    except ImportError:
        messages.error(request, 'openpyxl is required for Excel export.')
        return redirect('costing:detail', pk=pk)

    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_view_sheet(request.user, sheet):
        messages.error(request, 'You do not have permission to export this costing sheet.')
        return redirect('costing:list')
    if getattr(request.user, 'is_proposal_team_user', False):
        messages.error(request, 'Excel export includes pricing — only the sales team can download it.')
        return redirect('costing:detail', pk=pk)
    sections = sheet.sections.prefetch_related('line_items').all()
    rates = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}

    # Inject caches into all line items to avoid N+1 queries
    for section in sections:
        for item in section.line_items.all():
            item.set_exchange_rates_cache(rates)
            item.set_sheet_cache(sheet)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BOM'[:31]

    # Styles
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    hdr_white = Font(bold=True, color='FFFFFF', size=9)
    section_font = Font(bold=True, size=10)
    num_fmt = '#,##0.00'
    # Grey for cost columns (second half)
    cost_fill = PatternFill(start_color='D0D0D0', end_color='D0D0D0', fill_type='solid')
    # Darker grey header for cost columns
    hdr_cost = PatternFill(start_color='495057', end_color='495057', fill_type='solid')
    # Dark header for info columns (first half)
    hdr_dark = PatternFill(start_color='212529', end_color='212529', fill_type='solid')

    # Row 2: Sheet parameters + Exchange Rates
    ws['L2'] = 'Margin'
    ws['M2'] = float(sheet.margin)
    ws['L3'] = 'Discount'
    ws['M3'] = float(sheet.discount_rate)
    ws['L4'] = 'Shipping'
    ws['M4'] = float(sheet.shipping_rate)
    ws['L5'] = 'Customs'
    ws['M5'] = float(sheet.customs_rate)
    ws['L6'] = 'Finances'
    ws['M6'] = float(sheet.finances_rate)
    ws['L7'] = 'Installation'
    ws['M7'] = float(sheet.installation_rate)

    ws['O2'] = 'EXCHANGE RATES'
    ws['O2'].font = Font(bold=True)
    rate_row = 3
    for code, rate in sorted(rates.items()):
        ws.cell(row=rate_row, column=15, value=code).font = Font(bold=True)  # O
        ws.cell(row=rate_row, column=16, value=float(rate))  # P
        rate_row += 1

    # Row 5: Project info
    ws['A5'] = 'Project:'
    ws['B5'] = sheet.project.project_name if sheet.project else 'N/A'
    ws['C5'] = 'LN Ref:'
    ws['D5'] = sheet.project.proposal_reference if sheet.project else 'N/A'
    ws['F5'] = 'Cust Ref:'
    ws['G5'] = sheet.customer_reference or 'N/A'

    # Row 7: Customer / End User
    ws['A7'] = 'Customer:'
    ws['C7'] = 'End User:'
    ws['F7'] = 'Date:'

    # Row 8: Title
    ws['A8'] = 'BILL OF MATERIAL - COMMERCIAL'
    ws['A8'].font = Font(bold=True, size=12)

    # Column headers row
    r = 10
    headers = [
        'Item No', 'Description', 'Vendor', 'Make', 'Model', 'Qty', 'Unit',
        'Unit Price', 'Total Price',
        'Currency', 'Base Unit Cost', 'Discount', 'Unit Cost', 'Total Cost',
        'Margin', 'Base Unit Price', 'Base Total Price',
        'Shipping %', 'Customs %', 'Finances %', 'Installation %',
        'Unit Price', 'Total Price',
    ]
    col_fills = (
        [hdr_dark]*9 +
        [hdr_cost]*14
    )
    for col, (h, fill) in enumerate(zip(headers, col_fills), 1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = hdr_white
        cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Data rows
    row = 11
    data_fills = (
        [None]*9 +
        [cost_fill]*14
    )

    for section in sections:
        # Section header
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=23)
        cell = ws.cell(row=row, column=1, value=f"{section.section_number}  {section.title}")
        cell.font = section_font
        cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        row += 1

        for item in section.line_items.all():
            values = [
                item.item_number,
                item.description,
                item.vendor_name,
                item.make,
                item.model_number,
                float(item.quantity),
                item.unit,
                # Unit Price and Total Price (calculated final)
                round(float(item.final_unit_price), 2),
                round(float(item.final_total_price), 2),
                # Cost breakdown (grey section)
                item.supplier_currency,
                round(float(item.base_unit_cost), 2),
                # "Discount" column shows the discount amount in SAR so the
                # math reads as Base Unit Cost - Discount ~ Unit Cost.
                round(float(item.discount_amount), 2),
                round(float(item.unit_cost), 2),
                round(float(item.total_cost), 2),
                float(item.effective_margin),
                round(float(item.base_unit_price), 2),
                round(float(item.base_total_price), 2),
                float(item.effective_shipping_pct),
                float(item.effective_customs_pct),
                float(item.effective_finances_pct),
                float(item.effective_installation_pct),
                # Final Unit Price and Total Price at end of cost section
                round(float(item.final_unit_price), 2),
                round(float(item.final_total_price), 2),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin
                if col <= len(data_fills) and data_fills[col - 1]:
                    cell.fill = data_fills[col - 1]
                if isinstance(val, float) and col not in (5,):
                    cell.number_format = num_fmt
            row += 1

        # Subtotal
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"Sub-Total")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='right')
        st_cell = ws.cell(row=row, column=9, value=round(float(section.subtotal), 2))
        st_cell.font = Font(bold=True)
        st_cell.number_format = num_fmt
        row += 1

    # Grand total
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value=f'GRAND TOTAL ({sheet.output_currency})')
    cell.font = Font(bold=True, size=12, color='FFFFFF')
    cell.fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    cell.alignment = Alignment(horizontal='right')
    gt_cell = ws.cell(row=row, column=9, value=round(float(sheet.grand_total), 2))
    gt_cell.font = Font(bold=True, size=12, color='FFFFFF')
    gt_cell.fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    gt_cell.number_format = num_fmt

    # Column widths
    widths = {
        'A': 10, 'B': 40, 'C': 12, 'D': 12, 'E': 6, 'F': 6,
        'G': 12, 'H': 12, 'I': 12,  # Vendor, Unit Price, Total Price
        'J': 8,  # Currency
        'K': 12, 'L': 10, 'M': 10, 'N': 12,  # Base Unit Cost, Discount, Unit Cost, Total Cost
        'O': 8, 'P': 12, 'Q': 12,  # Margin, Base Unit Price, Base Total Price
        'R': 10, 'S': 10, 'T': 10, 'U': 10,  # Percentages
        'V': 12, 'W': 12,  # Final Unit Price, Total Price
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    filename = _safe_filename(sheet.title, suffix='BOM', extension='xlsx')
    # Write to a buffer so we can both save a revision copy and return it
    import io as _io
    xlsx_buffer = _io.BytesIO()
    wb.save(xlsx_buffer)
    xlsx_bytes = xlsx_buffer.getvalue()

    try:
        _save_costing_revision(sheet, xlsx_bytes, filename, export_format='excel', user=request.user)
    except Exception as _e:
        import logging
        logging.getLogger(__name__).warning('Failed to save costing Excel revision: %s', _e)

    response = HttpResponse(
        xlsx_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── PDF Summary Export ────────────────────────────────────────

@login_required
def costing_export_pdf(request, pk):
    """Export a professional PDF summary matching the commercial offer format"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, BaseDocTemplate, PageTemplate, Frame,
            Table, TableStyle, Paragraph, Spacer,
            PageBreak, Image, NextPageTemplate,
        )
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
        from datetime import datetime
        import string
    except ImportError:
        messages.error(request, 'reportlab is required for PDF export. Install with: pip install reportlab')
        return redirect('costing:detail', pk=pk)

    sheet = get_object_or_404(CostingSheet, pk=pk)
    # `?unpriced=1` produces a price-free BOM PDF — same layout, prices and
    # totals removed. Proposal team can grab this version; the priced one
    # stays sales/finance/procurement-only.
    unpriced = request.GET.get('unpriced') == '1'
    # `?services_only=1` prints the page-1 summary table with the A.2 SERVICES
    # block only — row A (contract total), A.1 (scope of supply) and A.3
    # (optional items) are dropped. Everything else — terms, client remarks
    # and the page-2+ detailed BOM — is unchanged.
    services_only = request.GET.get('services_only') == '1'
    if not _user_can_view_sheet(request.user, sheet):
        messages.error(request, 'You do not have permission to export this costing sheet.')
        return redirect('costing:list')
    if not unpriced and getattr(request.user, 'is_proposal_team_user', False):
        messages.error(
            request,
            'The priced PDF is sales-only. Use "Export BOM PDF (unpriced)" instead.',
        )
        return redirect('costing:detail', pk=pk)
    sections = list(sheet.sections.prefetch_related('line_items').all())

    # Inject caches into all line items to avoid N+1 queries
    rates_dict = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}
    for section in sections:
        for item in section.line_items.all():
            item.set_exchange_rates_cache(rates_dict)
            item.set_sheet_cache(sheet)

    # SAR → output_currency factor for everything the PDF prints. Computed
    # aggregates (line items, section.subtotal, sheet.grand_total) are stored
    # in SAR; the user enters Scope-of-Work totals already in output_currency,
    # so SOW values are NOT multiplied by `conv`.
    conv = _conversion_rate(sheet.output_currency, rates_dict)

    # Build the PDF into a BytesIO so we can both save a revision copy and
    # return it to the user.
    import io as _io
    pdf_buffer = _io.BytesIO()
    if unpriced:
        _pdf_suffix = 'BOM_Unpriced'
    elif services_only:
        _pdf_suffix = 'Services'
    else:
        _pdf_suffix = 'Commercial_Offer'
    filename = _safe_filename(sheet.title, suffix=_pdf_suffix, extension='pdf')

    import os
    from django.contrib.staticfiles.finders import find as find_static

    # Logo path (resolved once, used in every page callback)
    _logo_path = find_static('images/leap_logo.jpg')
    if not _logo_path:
        _logo_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'images', 'leap_logo.jpg')
    if _logo_path and not os.path.exists(_logo_path):
        _logo_path = None

    from reportlab.pdfgen.canvas import Canvas

    class NumberedCanvas(Canvas):
        """Custom canvas that draws logo header + page footer with
        'Page X of Y' on every page. Uses a two-pass approach:
        showPage() saves page state without writing; save() replays
        all pages with the header/footer drawn, now knowing the total.

        Each page template's onPage callback stamps _leap_template on
        the canvas so we know which template (summary/bom) owned each
        page when replaying."""

        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_page_states = []
            self._leap_template = 'summary'  # default

        def showPage(self):
            # Save current page drawing commands without writing to PDF.
            state = dict(self.__dict__)
            state['_leap_template'] = getattr(self, '_leap_template', 'summary')
            self._saved_page_states.append(state)
            self._startPage()  # reset canvas for next page

        def save(self):
            total = len(self._saved_page_states)
            for page_num, state in enumerate(self._saved_page_states, 1):
                self.__dict__.update(state)
                template_id = state.get('_leap_template', 'summary')
                self._draw_header_footer(page_num, total, template_id)
                Canvas.showPage(self)
            Canvas.save(self)

        def _draw_header_footer(self, page_num, total_pages, template_id='summary'):
            self.saveState()
            page_w, page_h = A4

            # ── Watermark: centered, semi-transparent logo ──
            if _logo_path:
                self.saveState()
                self.setFillAlpha(0.06)  # very faint
                wm_w, wm_h = 300, 110
                self.drawImage(
                    _logo_path,
                    (page_w - wm_w) / 2,
                    (page_h - wm_h) / 2,
                    width=wm_w, height=wm_h,
                    preserveAspectRatio=True, mask='auto',
                )
                self.restoreState()

            # ── Header: logo top-left ──
            if _logo_path:
                self.drawImage(_logo_path, 15*mm, page_h - 18*mm, width=100, height=35, preserveAspectRatio=True, mask='auto')

            # ── Header: project details table (full width, below logo, BOM pages only) ──
            if template_id == 'bom':
                header_cell_style = ParagraphStyle('HdrCell', fontName='Helvetica', fontSize=7, leading=8.5)
                header_cell_bold = ParagraphStyle('HdrCellBold', fontName='Helvetica-Bold', fontSize=7, leading=8.5)
                hdr_data = [
                    [Paragraph('Project:', header_cell_bold),
                     Paragraph(str(project_name), header_cell_style), '',
                     Paragraph('LN Ref:', header_cell_bold),
                     Paragraph(str(ln_ref), header_cell_style)],
                    [Paragraph('Cust Ref:', header_cell_bold),
                     Paragraph(str(cust_ref), header_cell_style), '',
                     Paragraph('Date:', header_cell_bold),
                     Paragraph(current_date, header_cell_style)],
                    [Paragraph('Customer:', header_cell_bold),
                     Paragraph(str(sheet.customer_name or ''), header_cell_style), '',
                     Paragraph('End User:', header_cell_bold),
                     Paragraph(str(sheet.end_user or ''), header_cell_style)],
                ]
                avail_w = page_w - 30*mm  # minus left+right margins
                hdr_table = Table(hdr_data, colWidths=[55, avail_w*0.40, 10, 55, avail_w*0.32])
                hdr_table.setStyle(TableStyle([
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
                    ('TOPPADDING', (0, 0), (-1, -1), 1),
                ]))
                hdr_w, hdr_h = hdr_table.wrapOn(self, avail_w, 30*mm)
                # Position the table so its bottom sits just above the flowable content area
                # (content area starts at page_h - topMargin = page_h - 42mm).
                # Leave a 2mm gap above content.
                hdr_table.drawOn(self, 15*mm, page_h - 40*mm)

            # ── Footer: confidential left, page number right ──
            footer_y = 10*mm
            self.setFont('Helvetica', 6)
            self.drawString(15*mm, footer_y,
                'Confidential. \u00A9 Leap Networks. All rights reserved.')
            self.drawRightString(page_w - 15*mm, footer_y,
                f'Page {page_num} of {total_pages}')

            # ── Summary page 1: fill the empty "Page X of Y" cell in the
            #    flowable project-details header. The value is unknown at
            #    flowable-build time so it's overlaid by the canvas here.
            #    Coordinates come from the actual layout of header_table
            #    (computed in the enclosing scope after wrap()), so they
            #    track row-wrapping (e.g. LN Ref taking two lines). ──
            if template_id == 'summary' and page_num == 1:
                self.setFont('Helvetica', 8)
                self.setFillColor(colors.black)
                self.drawString(_page_cell_x, _page_row_baseline_y,
                    f'Page {page_num} of {total_pages}')
            self.restoreState()

    # Create PDF document — use BaseDocTemplate with two page templates:
    #  - "summary"  = page 1 (topMargin=25mm, no project details header)
    #  - "bom"      = pages 2+ (topMargin=42mm, project details header drawn by canvas)
    doc = BaseDocTemplate(
        pdf_buffer,
        pagesize=A4,
        rightMargin=15*mm,
        leftMargin=15*mm,
        topMargin=25*mm,
        bottomMargin=22*mm,
    )
    summary_frame = Frame(
        15*mm, 22*mm,
        A4[0] - 30*mm, A4[1] - 25*mm - 22*mm,
        id='summary', leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
    )
    bom_frame = Frame(
        15*mm, 22*mm,
        A4[0] - 30*mm, A4[1] - 42*mm - 22*mm,
        id='bom', leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0,
    )
    def _mark_summary(canvas, _doc): canvas._leap_template = 'summary'
    def _mark_bom(canvas, _doc): canvas._leap_template = 'bom'
    doc.addPageTemplates([
        PageTemplate(id='summary', frames=[summary_frame], onPage=_mark_summary),
        PageTemplate(id='bom', frames=[bom_frame], onPage=_mark_bom),
    ])

    elements = []
    styles = getSampleStyleSheet()

    # Page usable width (A4=595pt minus 15mm margins each side)
    PAGE_WIDTH = A4[0] - 15*mm - 15*mm  # ≈510pt

    # ─── Colors (matched to the reference PDF) ───
    TITLE_BLUE = colors.HexColor('#528DD4')   # "COMMERCIAL OFFER SUMMARY" bg
    HEADER_GREY = colors.HexColor('#D9D9D9')   # Column header rows
    ALT_ROW = colors.HexColor('#F1DCDB')      # Alternating light pink rows
    SECTION_BG_P1 = colors.white              # Summary rows on page 1
    BORDER_COLOR = colors.black

    # ─── Reusable paragraph styles (sizes from reference PDF) ───
    # Page 1 info: 7.6pt equivalent
    cell_style = ParagraphStyle('Cell', fontName='Helvetica', fontSize=8, leading=10)
    cell_bold = ParagraphStyle('CellBold', fontName='Helvetica-Bold', fontSize=8, leading=10)
    cell_right = ParagraphStyle('CellRight', fontName='Helvetica', fontSize=8, leading=10, alignment=TA_RIGHT)
    cell_center = ParagraphStyle('CellCenter', fontName='Helvetica', fontSize=8, leading=10, alignment=TA_CENTER)
    desc_style = ParagraphStyle('Desc', fontName='Helvetica', fontSize=8, leading=10, wordWrap='CJK')
    # Title bar: 10pt white on blue
    title_style = ParagraphStyle('TitleBar', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER, textColor=colors.white)

    # Placeholder for blanked numbers on the unpriced variant — plain
    # empty string so the price cells render visually empty.
    _BLANK = ''

    def fmt_num(val):
        """Format a Decimal/number with thousand separators and 2 decimals.

        Returns the blank placeholder in unpriced mode so the visual layout
        stays identical to the priced PDF.
        """
        if unpriced:
            return _BLANK
        try:
            return f'{float(val):,.2f}'
        except (TypeError, ValueError):
            return ''

    def fmt_money(val):
        """Format a SAR-denominated value into the sheet's output currency.

        Use this for any aggregate computed from line-item costs (which are
        stored in SAR). Use plain fmt_num for values the user already
        entered in the output currency (Scope-of-Work item totals).
        Returns the blank placeholder in unpriced mode.
        """
        if unpriced:
            return _BLANK
        try:
            return f'{float(Decimal(str(val)) * conv):,.2f}'
        except (TypeError, ValueError, InvalidOperation):
            return ''

    # ─── HEADER SECTION (logo drawn by on_page callback, info table here) ───
    project_name = sheet.project.project_name if sheet.project else sheet.title
    region_name = sheet.project.region.name if sheet.project and sheet.project.region else 'N/A'
    ln_ref = sheet.project.proposal_reference if sheet.project else ''
    cust_ref = sheet.customer_reference or ''
    created_by_name = sheet.created_by.get_full_name() if sheet.created_by else ''
    created_by_email = sheet.created_by.email if sheet.created_by else ''
    current_date = datetime.now().strftime('%d/%m/%Y')

    # Page X of Y placeholder — filled in by the on_page callback.
    # We also add it inline in the header for the summary page.
    total_pages_text = ''  # will be set after build

    header_data = [
        [Paragraph('<b>Project:</b>', cell_style),
         Paragraph(project_name, cell_style), '',
         Paragraph('<b>Sales Office:</b>', cell_style),
         Paragraph(f'LN-{region_name}', cell_style), ''],
        [Paragraph('<b>Customer:</b>', cell_style),
         Paragraph(sheet.customer_name, cell_style), '',
         Paragraph('<b>Telephone:</b>', cell_style),
         Paragraph(sheet.telephone, cell_style), ''],
        [Paragraph('<b>End User:</b>', cell_style),
         Paragraph(sheet.end_user, cell_style), '',
         Paragraph('<b>Fax:</b>', cell_style),
         Paragraph(sheet.fax, cell_style), ''],
        [Paragraph('<b>Cust. Ref:</b>', cell_style),
         Paragraph(cust_ref, cell_style), '',
         Paragraph('<b>Contact:</b>', cell_style),
         Paragraph(sheet.contact_person, cell_style), ''],
        [Paragraph('<b>Contact Person:</b>', cell_style),
         Paragraph(created_by_name, cell_style), '', '', '', ''],
        [Paragraph('<b>LN Ref:</b>', cell_style),
         Paragraph(ln_ref, cell_style), '',
         Paragraph('<b>Email:</b>', cell_style),
         Paragraph(created_by_email, cell_style), ''],
    ]

    # Insert one additional Contact/Email row per extra contributor selected
    # on the sheet. Labels are numbered (Contact 2 / Email 2, ...) so the
    # auto creator stays as the unnumbered "Contact Person / Email" above.
    extra_contacts = list(sheet.additional_contacts.all().order_by('first_name', 'last_name', 'email'))
    for idx, u in enumerate(extra_contacts, start=2):
        full_name = u.get_full_name() or u.username
        header_data.append([
            Paragraph(f'<b>Contact {idx}:</b>', cell_style),
            Paragraph(full_name, cell_style), '',
            Paragraph(f'<b>Email {idx}:</b>', cell_style),
            Paragraph(u.email or '', cell_style), '',
        ])

    header_data.append([
        Paragraph('<b>Date:</b>', cell_style),
        Paragraph(current_date, cell_style), '',
        Paragraph('<b>Page</b>', cell_style),
        # Value drawn by NumberedCanvas overlay (needs total pages from
        # second pass). Leaving empty Paragraph keeps cell width stable.
        Paragraph('', cell_style), '',
    ])

    # Split page width into 6 columns for the info block
    header_table = Table(header_data, colWidths=[70, PAGE_WIDTH*0.35, 10, 70, PAGE_WIDTH*0.35, None])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('LINEBELOW', (0, 0), (-1, -2), 0.25, colors.HexColor('#CCCCCC')),
        ('LINEAFTER', (1, 0), (1, -1), 0.25, colors.HexColor('#CCCCCC')),
    ]))

    # Pre-layout the header table so we know its actual row heights (some
    # rows wrap to two lines, e.g. LN Ref). The NumberedCanvas overlay for
    # "Page X of Y" needs the real Y of the last row.
    header_table.wrap(PAGE_WIDTH, A4[1])
    _hdr_row_heights = list(header_table._rowHeights)
    _page_row_top_y = (A4[1] - 25*mm) - sum(_hdr_row_heights[:-1])
    _page_row_baseline_y = _page_row_top_y - _hdr_row_heights[-1] + 4
    # Column 4 (the value cell) starts at: leftMargin + col0(70) + col1(PAGE_WIDTH*0.35)
    # + col2(10) + col3(70) + default LEFTPADDING(6).
    _page_cell_x = 15*mm + 70 + PAGE_WIDTH*0.35 + 10 + 70 + 6

    elements.append(header_table)
    elements.append(Spacer(1, 5 * mm))

    # ─── TITLE BAR + DATA TABLE ───
    # The summary page (contract total, A.1/A.2/A.3 breakdown) is the only
    # part of this PDF that prints prices. When unpriced=1 we still build
    # the summary objects (cheap) but skip appending them to the document.
    # Define column widths first so the title bar matches
    # 5 columns spanning the full page width
    col_widths = [40, PAGE_WIDTH - 40 - 35 - 35 - 120, 35, 35, 120]
    table_total_width = PAGE_WIDTH
    title_data = [[Paragraph('<b>COMMERCIAL OFFER SUMMARY</b>', title_style)]]
    title_table = Table(title_data, colWidths=[table_total_width])
    title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), TITLE_BLUE),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
    ]))
    elements.append(title_table)
    elements.append(Spacer(1, 2 * mm))

    # ─── MAIN DATA TABLE ───
    currency = sheet.output_currency

    # Column headers
    hdr = [
        Paragraph('<b>Item No</b>', cell_center),
        Paragraph('<b>Item Description</b>', cell_bold),
        Paragraph('<b>Qty</b>', cell_center),
        Paragraph('<b>UOM</b>', cell_center),
        Paragraph(f'<b>Total Price ({currency})</b>', cell_center),
    ]
    data = [hdr]

    # Scope of Work items & total — SOW values are user-entered in
    # output_currency already, so no conversion is applied to them.
    sow_items = list(sheet.scope_of_work_items.all())
    sow_total = sum(i.total_price for i in sow_items) if sow_items else sheet.scope_of_work_total
    # SAR aggregates → output_currency once, here, so the rest of the table
    # can mix them with the SOW values consistently.
    optional_total_oc = (sheet.optional_subtotal * conv).quantize(Decimal('0.01'))
    grand_total_oc = (sheet.grand_total * conv).quantize(Decimal('0.01'))
    has_optional = optional_total_oc and optional_total_oc > 0
    # A.1 always shows the non-optional supply subtotal.
    # sheet.grand_total already excludes optional when include_optional_in_total=False,
    # and includes them when True — so we subtract the optional piece back out to get
    # the stable "non-optional supply" figure for A.1.
    if sheet.include_optional_in_total:
        non_optional_supply_oc = grand_total_oc - optional_total_oc
    else:
        non_optional_supply_oc = grand_total_oc
    # Contract total: A.1 + A.2 always, plus A.3 only when the flag is on.
    contract_total_oc = non_optional_supply_oc + Decimal(str(sow_total))
    if sheet.include_optional_in_total and has_optional:
        contract_total_oc += optional_total_oc

    # Row A (contract total) + Row A.1 (scope of supply) — dropped entirely
    # when ?services_only=1, which prints the A.2 SERVICES block only.
    if not services_only:
        # Row A: MAIN - TOTAL CONTRACT PRICE
        contract_label_parts = 'A.1+A.2+A.3' if (has_optional and sheet.include_optional_in_total) else 'A.1+A.2'
        contract_label = f'MAIN - TOTAL CONTRACT PRICE ({contract_label_parts})'
        data.append([
            Paragraph('<b>A</b>', cell_bold),
            Paragraph(f'<b>{contract_label}</b>', cell_bold),
            '', '',
            Paragraph(f'<b>{fmt_num(contract_total_oc)}</b>', cell_right),
        ])

        # Row A.1: SCOPE OF SUPPLY (pink bg) — always non-optional
        data.append([
            Paragraph('<b>A.1</b>', cell_bold),
            Paragraph('<b>SCOPE OF SUPPLY</b>', cell_bold),
            '', '',
            Paragraph(f'<b>{fmt_num(non_optional_supply_oc)}</b>', cell_right),
        ])

    # Consolidate sections with the same title so that duplicate
    # imports (e.g. two "DATA NETWORK SYSTEM" sections) are merged
    # into a single summary row with their subtotals added together.
    # Optional sections are pulled into a separate list so they render
    # under their own A.3 block at the bottom.
    from collections import OrderedDict
    consolidated = OrderedDict()
    optional_groups = OrderedDict()
    for section in sections:
        if not section.section_number:
            continue  # skip divider rows
        target = optional_groups if section.is_optional else consolidated
        title = section.title.strip().upper()
        if title in target:
            target[title]['subtotal'] += section.subtotal
        else:
            target[title] = {
                'title': section.title,
                'subtotal': section.subtotal,
            }

    # Supply line items (one per consolidated group, renumbered) + the blank
    # separator row before A.2 — both omitted in services-only mode.
    if not services_only:
        for idx, (title_key, group) in enumerate(consolidated.items(), 1):
            data.append([
                Paragraph(str(idx), cell_center),
                Paragraph(f' {group["title"]}', cell_style),
                Paragraph('1', cell_center),
                Paragraph('LOT', cell_center),
                Paragraph(fmt_money(group['subtotal']), cell_right),
            ])

        # Blank separator row
        data.append(['', '', '', '', ''])

    # Row A.2: SERVICES (pink bg) — sow_total is already in output_currency
    data.append([
        Paragraph('<b>A.2</b>', cell_bold),
        Paragraph('<b>SERVICES</b>', cell_bold),
        '', '',
        Paragraph(f'<b>{fmt_num(sow_total)}</b>', cell_right),
    ])

    # Scope of work line items
    for item in sow_items:
        data.append([
            Paragraph(str(item.serial_number), cell_center),
            Paragraph(f' {item.description}', desc_style),
            Paragraph(str(int(item.quantity) if item.quantity == int(item.quantity) else item.quantity), cell_center),
            Paragraph(item.uom, cell_center),
            Paragraph(_BLANK if unpriced else item.display_price, cell_right),
        ])

    # Row indices for special styling
    num_consolidated = len(consolidated)
    if services_only:
        # Only the header row precedes A.2; A.1/A.3 are not rendered.
        a1_row = None
        a2_row = 1
    else:
        a1_row = 2
        a2_row = 2 + num_consolidated + 1 + 1  # +1 blank +1 for A.2 itself
    a3_row = None  # set below if optional section exists
    PINK_BG = colors.HexColor('#F1DCDB')

    # ─── A.3 OPTIONAL ITEMS (only when there are optional sections) ───
    if has_optional and not services_only:
        # Blank separator + A.3 header
        data.append(['', '', '', '', ''])
        # Capture A.3 row index right before appending so indexing is bulletproof
        a3_row = len(data)
        data.append([
            Paragraph('<b>A.3</b>', cell_bold),
            Paragraph('<b>OPTIONAL ITEMS</b>', cell_bold),
            '', '',
            Paragraph(f'<b>{fmt_num(optional_total_oc)}</b>', cell_right),
        ])
        # Optional section line items (renumbered starting from 1 within this block)
        for idx, (title_key, group) in enumerate(optional_groups.items(), 1):
            data.append([
                Paragraph(str(idx), cell_center),
                Paragraph(f' {group["title"]}', cell_style),
                Paragraph('1', cell_center),
                Paragraph('LOT', cell_center),
                Paragraph(fmt_money(group['subtotal']), cell_right),
            ])

    main_table = Table(data, colWidths=col_widths, repeatRows=1)

    SUMMARY_GOLD = colors.HexColor('#FFC000')  # Gold for summary header only

    style_cmds = [
        # Header row — gold (summary page keeps gold)
        ('BACKGROUND', (0, 0), (-1, 0), SUMMARY_GOLD),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # All rows
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        # Orange/red thin line below header (above Row A, or above A.2 in
        # services-only mode).
        ('LINEABOVE', (0, 1), (-1, 1), 2, colors.HexColor('#E26B0A')),
        # A.2 SERVICES — pink bg
        ('BACKGROUND', (0, a2_row), (-1, a2_row), PINK_BG),
    ]
    # A.1 SCOPE OF SUPPLY — pink bg (omitted in services-only mode)
    if a1_row is not None:
        style_cmds.append(('BACKGROUND', (0, a1_row), (-1, a1_row), PINK_BG))
    if a3_row is not None:
        style_cmds.append(('BACKGROUND', (0, a3_row), (-1, a3_row), PINK_BG))

    main_table.setStyle(TableStyle(style_cmds))
    elements.append(main_table)
    elements.append(Spacer(1, 6 * mm))

    # ─── TERMS & CONDITIONS (on the COS summary page, page 1) ───
    section_hdr_style = ParagraphStyle('SectionHdr', fontName='Helvetica-Bold', fontSize=10, leading=14, spaceAfter=4)
    sub_hdr_style = ParagraphStyle('SubHdr', fontName='Helvetica-Bold', fontSize=9, leading=12, spaceAfter=2)
    body_style = ParagraphStyle('Body', fontName='Helvetica', fontSize=9, leading=12, spaceAfter=2)

    selected = sheet.selected_terms.all()
    terms_grouped = {}
    for cat_value, cat_label in TermsTemplate.CATEGORY_CHOICES:
        items_list = [t for t in selected if t.category == cat_value]
        if items_list:
            terms_grouped[cat_value] = items_list

    # Terms & Conditions — lines rendered as-is; user supplies their own numbering
    tc_templates = terms_grouped.get('terms_and_conditions', [])
    tc_legacy = sheet.terms_and_conditions.strip()
    if tc_templates or tc_legacy:
        elements.append(Paragraph('Terms &amp; Conditions:', section_hdr_style))
        for tmpl in tc_templates:
            elements.append(Paragraph(f'<b>{tmpl.name}</b>', sub_hdr_style))
            for line in [l.strip() for l in tmpl.content.strip().splitlines() if l.strip()]:
                elements.append(Paragraph(line, body_style))
        if tc_legacy and not tc_templates:
            for line in [l.strip() for l in tc_legacy.splitlines() if l.strip()]:
                elements.append(Paragraph(line, body_style))
        elements.append(Spacer(1, 3 * mm))

    # Exclusions
    excl_templates = terms_grouped.get('exclusions', [])
    excl_legacy = sheet.exclusions.strip()
    if excl_templates or excl_legacy:
        last_num = sections[-1].section_number if sections else '1'
        try:
            excl_num = str(int(last_num) + 1)
        except ValueError:
            excl_num = 'B'
        elements.append(Paragraph(f'{excl_num}. Exclusion', section_hdr_style))
        letters = list(string.ascii_lowercase)
        idx = 0
        for tmpl in excl_templates:
            for line in [l.strip() for l in tmpl.content.strip().splitlines() if l.strip()]:
                letter = letters[idx] if idx < len(letters) else str(idx + 1)
                elements.append(Paragraph(f'{letter}. {line}', body_style))
                idx += 1
        if excl_legacy and not excl_templates:
            for line in [l.strip() for l in excl_legacy.splitlines() if l.strip()]:
                letter = letters[idx] if idx < len(letters) else str(idx + 1)
                elements.append(Paragraph(f'{letter}. {line}', body_style))
                idx += 1
        elements.append(Spacer(1, 3 * mm))

    # Payment Terms
    pt_templates = terms_grouped.get('payment_terms', [])
    pt_legacy = sheet.payment_terms.strip()
    if pt_templates or pt_legacy:
        elements.append(Paragraph('Payment Terms:', section_hdr_style))
        for tmpl in pt_templates:
            for line in [l.strip() for l in tmpl.content.strip().splitlines() if l.strip()]:
                elements.append(Paragraph(line, body_style))
        if pt_legacy and not pt_templates:
            for line in [l.strip() for l in pt_legacy.splitlines() if l.strip()]:
                elements.append(Paragraph(line, body_style))
        elements.append(Spacer(1, 3 * mm))

    # Conclusion
    conc_templates = terms_grouped.get('conclusion', [])
    conc_legacy = sheet.conclusion.strip()
    if conc_templates or conc_legacy:
        elements.append(Paragraph('Conclusion', section_hdr_style))
        for tmpl in conc_templates:
            for line in [l.strip() for l in tmpl.content.strip().splitlines() if l.strip()]:
                elements.append(Paragraph(line, body_style))
        if conc_legacy and not conc_templates:
            for line in [l.strip() for l in conc_legacy.splitlines() if l.strip()]:
                elements.append(Paragraph(line, body_style))
        elements.append(Spacer(1, 3 * mm))

    # ─── CLIENT REMARKS & LEAP'S ANSWERS ───
    # Two-column table; iterates every pair across every selected bundle.
    # Each cell uses its own color (pair.remark_color / pair.answer_color).
    from xml.sax.saxutils import escape as _xml_escape
    client_pairs = list(
        ClientRemarkPair.objects.filter(template__in=sheet.selected_client_remarks.all())
        .select_related('template')
        .order_by('template__name', 'order', 'pk')
    )
    if client_pairs:
        elements.append(Paragraph("Client Remarks &amp; Leap's Answers", section_hdr_style))

        cr_header_style = ParagraphStyle(
            'CRHeader', fontName='Helvetica-Bold', fontSize=9, leading=11,
            alignment=TA_CENTER, textColor=colors.HexColor('#333333'),
        )
        cr_rows = [[
            Paragraph('Client Remark', cr_header_style),
            Paragraph("Leap's Answer", cr_header_style),
        ]]
        cr_styles = [
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5D7DC')),
            ('LINEBELOW', (0, 0), (-1, 0), 1.2, colors.HexColor('#C41E3A')),
            # Heavy outer box
            ('BOX', (0, 0), (-1, -1), 1.0, colors.HexColor('#333333')),
            # Strong horizontal dividers between every row (emphasise "row" structure)
            ('LINEBELOW', (0, 0), (-1, -2), 0.9, colors.HexColor('#555555')),
            # Vertical line between remark and answer columns
            ('LINEAFTER', (0, 0), (0, -1), 0.7, colors.HexColor('#888888')),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]

        def _hex(value, fallback=colors.black):
            try:
                return colors.HexColor(value or '#000000')
            except Exception:
                return fallback

        for pair in client_pairs:
            remark_style = ParagraphStyle(
                f'CRRemark-{pair.pk}', fontName='Helvetica', fontSize=9,
                leading=12, textColor=_hex(pair.remark_color),
            )
            answer_style = ParagraphStyle(
                f'CRAnswer-{pair.pk}', fontName='Helvetica', fontSize=9,
                leading=12, textColor=_hex(pair.answer_color),
            )
            remark_html = _xml_escape(pair.remark or '').replace('\n', '<br/>')
            answer_html = _xml_escape(pair.answer or '').replace('\n', '<br/>')
            cr_rows.append([
                Paragraph(remark_html, remark_style),
                Paragraph(answer_html, answer_style),
            ])

        # splitInRow=1 lets a single tall cell (e.g. a long pasted remark)
        # break across pages instead of raising LayoutError; repeatRows=1
        # re-prints the header on subsequent pages.
        cr_table = Table(
            cr_rows,
            colWidths=[PAGE_WIDTH * 0.5, PAGE_WIDTH * 0.5],
            repeatRows=1,
            splitByRow=1,
            splitInRow=1,
        )
        cr_table.setStyle(TableStyle(cr_styles))
        elements.append(cr_table)
        elements.append(Spacer(1, 3 * mm))

    # ─── PAGE 2+: DETAILED BOM WITH ALL LINE ITEMS ───
    # Switch to the "bom" page template (larger top margin + canvas header)
    elements.append(NextPageTemplate('bom'))
    elements.append(PageBreak())

    # Project details header is now drawn by NumberedCanvas on every page — no inline table here.

    # BOM title bar — same 8-column layout in both modes; only the
    # numeric values are blanked when unpriced.
    bom_col_widths = [40, PAGE_WIDTH - 40 - 55 - 55 - 30 - 30 - 65 - 65, 55, 55, 30, 30, 65, 65]
    bom_title_data = [[Paragraph(
        f'<b>BILL OF MATERIAL - {"UNPRICED" if unpriced else "PRICED"}</b>',
        title_style,
    )]]
    bom_title_table = Table(bom_title_data, colWidths=[PAGE_WIDTH])
    bom_title_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), TITLE_BLUE),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
    ]))
    elements.append(bom_title_table)

    # Subtitle: MAIN - SCOPE OF SUPPLY
    bom_subtitle_data = [[Paragraph(
        f'<b>MAIN - SCOPE OF SUPPLY - {project_name}</b>',
        ParagraphStyle('SubTitle', fontName='Helvetica-Bold', fontSize=9, alignment=TA_CENTER),
    )]]
    bom_subtitle_table = Table(bom_subtitle_data, colWidths=[PAGE_WIDTH])
    bom_subtitle_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('BOX', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
    ]))
    elements.append(bom_subtitle_table)

    # BOM column styles — smaller font to fit 8 columns (reference uses 4.8-5.4pt)
    cell_right_bold = ParagraphStyle('CellRightBold', fontName='Helvetica-Bold', fontSize=7, leading=8.5, alignment=TA_RIGHT)
    cell_sm = ParagraphStyle('CellSm', fontName='Helvetica', fontSize=7, leading=8.5)
    cell_sm_bold = ParagraphStyle('CellSmBold', fontName='Helvetica-Bold', fontSize=7, leading=8.5)
    cell_sm_center = ParagraphStyle('CellSmCenter', fontName='Helvetica', fontSize=7, leading=8.5, alignment=TA_CENTER)
    cell_sm_center_bold = ParagraphStyle('CellSmCenterBold', fontName='Helvetica-Bold', fontSize=7, leading=8.5, alignment=TA_CENTER)
    cell_sm_right = ParagraphStyle('CellSmRight', fontName='Helvetica', fontSize=7, leading=8.5, alignment=TA_RIGHT)
    desc_sm_style = ParagraphStyle('DescSm', fontName='Helvetica', fontSize=7, leading=8.5, wordWrap='CJK')

    bom_hdr = [
        Paragraph('<b>Item No</b>', cell_sm_center),
        Paragraph('<b>Description</b>', cell_sm_bold),
        Paragraph('<b>Make</b>', cell_sm_center),
        Paragraph('<b>Model</b>', cell_sm_center),
        Paragraph('<b>Qty</b>', cell_sm_center),
        Paragraph('<b>Unit</b>', cell_sm_center),
        Paragraph(f'<b>Unit Price ({currency})</b>', cell_sm_center),
        Paragraph(f'<b>Total Price ({currency})</b>', cell_sm_center),
    ]
    bom_data = [bom_hdr]

    # Build all rows: divider banners + section headers + line items + subtotals
    SECTION_BG = colors.HexColor('#D9E2F3')    # Light blue for section headers
    SUBTOTAL_BG = colors.HexColor('#BFBFBF')   # Medium grey for subtotals
    DIVIDER_BG = colors.HexColor('#FFC000')    # Gold/yellow for import dividers
    divider_style = ParagraphStyle('Divider', fontName='Helvetica-Bold', fontSize=8, leading=10, alignment=TA_CENTER, textColor=colors.HexColor('#333333'))
    section_rows = []  # track (row_index, type) for styling
    BATCH_TOTAL_BG = colors.HexColor('#E6B8AF')  # Darker pink for batch totals
    batch_total_style = ParagraphStyle('BatchTotal', fontName='Helvetica-Bold', fontSize=7, leading=8.5, alignment=TA_RIGHT)
    row_idx = 1  # 0 is header
    batch_running_total = Decimal('0')
    # Use the sheet title as the default batch name for sections that
    # appear before any divider. This way the first group always gets a total.
    batch_name = sheet.title
    has_sections_in_batch = False

    def _insert_batch_total():
        """Insert a batch total row for the current batch."""
        nonlocal row_idx, has_sections_in_batch
        if has_sections_in_batch:
            bom_data.append([
                '', '', '', '', '',
                Paragraph('<b>Total</b>', batch_total_style),
                '',
                Paragraph(f'<b>{fmt_money(batch_running_total)}</b>', cell_right_bold),
            ])
            section_rows.append((row_idx, 'batch_total'))
            row_idx += 1
            has_sections_in_batch = False

    for section in sections:
        # Divider row (blank section_number = import batch heading)
        if not section.section_number:
            # Insert total for the previous batch before starting a new one
            _insert_batch_total()

            bom_data.append([
                Paragraph(f'<b>{section.title}</b>', divider_style),
                '', '', '', '', '', '', '',
            ])
            section_rows.append((row_idx, 'divider'))
            row_idx += 1
            batch_name = section.title
            batch_running_total = Decimal('0')
            continue

        # Section header row
        bom_data.append([
            Paragraph(f'<b>{section.section_number}</b>', cell_sm_center_bold),
            Paragraph(f'<b>{section.title}</b>', cell_sm_bold),
            '', '', '', '', '', '',
        ])
        section_rows.append((row_idx, 'section'))
        row_idx += 1

        # Line items — detect sub-headings (e.g. 1.14 that has children 1.14.1, 1.14.2)
        items_list = list(section.line_items.all())
        item_numbers = [i.item_number for i in items_list]
        # Build a set of prefixes that have children
        sub_heading_numbers = set()
        for num in item_numbers:
            # If "1.14.1" exists, then "1.14" is a sub-heading
            parts = num.rsplit('.', 1)
            if len(parts) == 2 and parts[0]:
                sub_heading_numbers.add(parts[0])

        SUBHEADING_BG = colors.HexColor('#E8EEF4')  # Light blue-grey

        for item in items_list:
            is_subheading = item.item_number in sub_heading_numbers

            if is_subheading:
                # Sub-heading row: bold, highlighted bg
                bom_data.append([
                    Paragraph(f'<b>{item.item_number}</b>', cell_sm_center_bold),
                    Paragraph(f'<b>{item.description}</b>', cell_sm_bold),
                    Paragraph(item.make or '', cell_sm),
                    Paragraph(item.model_number or '', cell_sm),
                    Paragraph(str(int(item.quantity) if item.quantity == int(item.quantity) else item.quantity), cell_sm_center),
                    Paragraph(item.unit, cell_sm_center),
                    Paragraph(fmt_money(item.final_unit_price), cell_sm_right),
                    Paragraph(fmt_money(item.final_total_price), cell_sm_right),
                ])
                section_rows.append((row_idx, 'subheading'))
            else:
                bom_data.append([
                    Paragraph(item.item_number, cell_sm_center),
                    Paragraph(item.description, desc_sm_style),
                    Paragraph(item.make or '', cell_sm),
                    Paragraph(item.model_number or '', cell_sm),
                    Paragraph(str(int(item.quantity) if item.quantity == int(item.quantity) else item.quantity), cell_sm_center),
                    Paragraph(item.unit, cell_sm_center),
                    Paragraph(fmt_money(item.final_unit_price), cell_sm_right),
                    Paragraph(fmt_money(item.final_total_price), cell_sm_right),
                ])
            row_idx += 1

        # Subtotal row
        bom_data.append([
            '', '', '', '', '', '',
            Paragraph('<b>Sub-Total</b>', cell_right_bold),
            Paragraph(f'<b>{fmt_money(section.subtotal)}</b>', cell_right_bold),
        ])
        section_rows.append((row_idx, 'subtotal'))
        row_idx += 1
        batch_running_total += section.subtotal
        has_sections_in_batch = True

    # Insert total for the last batch
    _insert_batch_total()

    # Build the BOM table
    bom_table = Table(bom_data, colWidths=bom_col_widths, repeatRows=1)

    bom_style_cmds = [
        # Header row — yellow
        ('BACKGROUND', (0, 0), (-1, 0), HEADER_GREY),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        # All rows
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_COLOR),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]

    # Style section header, subtotal, divider, and batch total rows
    for ridx, rtype in section_rows:
        if rtype == 'divider':
            # Gold banner spanning all columns — import batch heading
            bom_style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), DIVIDER_BG))
            bom_style_cmds.append(('SPAN', (0, ridx), (-1, ridx)))
            bom_style_cmds.append(('ALIGN', (0, ridx), (-1, ridx), 'CENTER'))
        elif rtype == 'section':
            bom_style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), SECTION_BG))
            bom_style_cmds.append(('SPAN', (1, ridx), (-1, ridx)))
        elif rtype == 'subtotal':
            bom_style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), SUBTOTAL_BG))
            bom_style_cmds.append(('SPAN', (0, ridx), (5, ridx)))
        elif rtype == 'subheading':
            # Light blue-grey for sub-heading items (e.g. 1.14 with children 1.14.1)
            bom_style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), SUBHEADING_BG))
        elif rtype == 'batch_total':
            # Darker pink total row for the entire import batch
            bom_style_cmds.append(('BACKGROUND', (0, ridx), (-1, ridx), BATCH_TOTAL_BG))
            bom_style_cmds.append(('SPAN', (0, ridx), (4, ridx)))
            bom_style_cmds.append(('SPAN', (5, ridx), (6, ridx)))
            bom_style_cmds.append(('LINEABOVE', (0, ridx), (-1, ridx), 1.5, BORDER_COLOR))

    bom_table.setStyle(TableStyle(bom_style_cmds))
    elements.append(bom_table)
    elements.append(Spacer(1, 10 * mm))

    # ─── END OF DOCUMENT ───
    end_style = ParagraphStyle(
        'EndDoc', fontName='Helvetica-Bold', fontSize=11,
        alignment=TA_CENTER, textColor=colors.HexColor('#666666'),
    )
    elements.append(Paragraph('— End of Document —', end_style))

    # Build PDF with logo header and Page X of Y footer on every page.
    # NumberedCanvas handles the drawing after all pages are rendered
    # so it knows the total page count for "Page X of Y".
    doc.build(elements, canvasmaker=NumberedCanvas)

    pdf_bytes = pdf_buffer.getvalue()

    # Save this export as a revision (with state snapshot + diff vs previous).
    # Unpriced exports are a derivative of the same sheet state, so we do
    # not snapshot them as separate revisions — the priced PDF remains the
    # canonical record.
    if not unpriced:
        try:
            _save_costing_revision(sheet, pdf_bytes, filename, export_format='pdf', user=request.user)
        except Exception as _e:
            import logging
            logging.getLogger(__name__).warning('Failed to save costing PDF revision: %s', _e)

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


# ─── Excel Import ────────────────────────────────────────────────

@login_required
def costing_import_excel(request, pk):
    """Import line items from Excel BOQ file"""
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from decimal import Decimal, InvalidOperation
    import re

    sheet = get_object_or_404(CostingSheet, pk=pk)
    if not _user_can_edit_sheet(request.user, sheet):
        messages.error(request, 'You do not have permission to import into this costing sheet.')
        return redirect('costing:list')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'Please select an Excel file to import.')
            return redirect('costing:import_excel', pk=pk)
        # Optional heading the user wants stamped on top of this import
        # batch so they can visually separate it from earlier content.
        import_heading = request.POST.get('import_heading', '').strip()

        # Parse the file outside any transaction so a bad file doesn't
        # hold a DB lock and we can return clean error messages.
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active  # Use first sheet
        except InvalidFileException:
            messages.error(request, 'The uploaded file is not a valid Excel workbook.')
            return redirect('costing:import_excel', pk=pk)
        except (KeyError, ValueError, OSError) as e:
            logger.warning('Costing import: failed to open Excel file', exc_info=e)
            messages.error(request, 'Could not read the Excel file. Please check the format and try again.')
            return redirect('costing:import_excel', pk=pk)

        # Find header row (look for row containing 'description' or 'vendor')
        header_row = None
        headers = {}
        for row_num in range(1, 15):
            row = [cell.value for cell in ws[row_num]]
            row_lower = [str(c).lower() if c else '' for c in row]
            if 'vendor' in row_lower or 'description' in row_lower:
                header_row = row_num
                for i, val in enumerate(row):
                    if val:
                        headers[str(val).lower().strip()] = i
                break

        if not header_row:
            messages.error(request, 'Could not find header row in Excel file. Looking for columns: Item No, Description, Make, Model, Qty, Unit, Vendor')
            return redirect('costing:import_excel', pk=pk)

        # Map column indices - matches the Item No / Description / Make /
        # Model / Qty / Unit / Vendor template used by costing_import_new.
        # unit_price / total_price are optional — when present the value
        # lands in base_unit_cost (raw supplier cost). Spares users from
        # typing 400 prices by hand for an existing sheet.
        col_map = {
            'item_no': headers.get('item no', headers.get('item #', headers.get('rfp item #', headers.get('rfp item', headers.get('#', -1))))),
            'description': headers.get('description', headers.get('desc', -1)),
            'make': headers.get('make', headers.get('manufacturer', -1)),
            'model': headers.get('model', headers.get('part#', headers.get('part', headers.get('part no', -1)))),
            'qty': headers.get('qty', headers.get('quantity', -1)),
            'unit': headers.get('unit', headers.get('uom', -1)),
            'vendor': headers.get('vendor', headers.get('vendor name', -1)),
            'unit_price': headers.get('unit price', headers.get('unit cost', headers.get('price', headers.get('cost', headers.get('rate', headers.get('unit rate', -1)))))),
            'total_price': headers.get('total price', headers.get('total cost', headers.get('total', headers.get('amount', headers.get('extended price', -1))))),
        }

        # All DB writes happen inside a single atomic block.
        # If any row raises, the entire import is rolled back so we
        # never leave orphaned sections/items behind.
        try:
            with transaction.atomic():
                # Continue numbering from any existing data on the sheet
                # so re-importing into the same sheet appends instead of
                # colliding with existing rows.
                existing_sections = list(sheet.sections.all().order_by('order', 'pk'))
                existing_max_order = max((s.order for s in existing_sections), default=-1)
                # Highest existing section number that parses as an int
                # (so we can keep auto-numbering 1, 2, 3, ...).
                existing_max_num = 0
                for s in existing_sections:
                    try:
                        existing_max_num = max(existing_max_num, int(s.section_number))
                    except (TypeError, ValueError):
                        pass

                sections_created = 0
                items_created = 0
                current_section = None
                item_order = 0
                next_order = existing_max_order + 1
                next_auto_num = existing_max_num + 1

                # If the admin entered a heading, create a divider section
                # at the top of this import batch so it's visually marked
                # off from earlier content. The divider has NO section
                # number (blank) so the regular numbering continues
                # uninterrupted: existing 1, 2, 3 -> heading -> 4, 5, 6.
                if import_heading and existing_sections:
                    CostingSection.objects.create(
                        costing_sheet=sheet,
                        section_number='',
                        title=import_heading[:255],
                        order=next_order,
                    )
                    next_order += 1
                    sections_created += 1

                # Helper to safely read a cell from a row by column index.
                def cell(row, key):
                    idx = col_map[key]
                    if idx < 0 or idx >= len(row):
                        return ''
                    return str(row[idx] or '').strip()

                def cell_raw(row, key):
                    idx = col_map[key]
                    if idx < 0 or idx >= len(row):
                        return None
                    return row[idx]

                for row_num in range(header_row + 1, ws.max_row + 1):
                    row = [c.value for c in ws[row_num]]
                    if not any(row):
                        continue

                    # Get values
                    item_no = cell(row, 'item_no')
                    description = cell(row, 'description')
                    make = cell(row, 'make')
                    model = cell(row, 'model')
                    qty_val = cell_raw(row, 'qty')
                    unit = cell(row, 'unit') or 'EA'
                    vendor = cell(row, 'vendor')
                    unit_price_raw = cell_raw(row, 'unit_price')
                    total_price_raw = cell_raw(row, 'total_price')

                    # Skip empty rows or sheet-level headings.
                    if not item_no and not description:
                        continue
                    if item_no and not description and not make and not model and qty_val is None:
                        continue

                    # Parse quantity
                    try:
                        qty = Decimal(str(qty_val)) if qty_val else Decimal('1')
                    except (InvalidOperation, ValueError):
                        qty = Decimal('1')

                    # Parse price (unit preferred; fall back to total/qty).
                    # Strip thousands separators carried over from Excel.
                    def _parse_money(raw):
                        if raw is None or raw == '':
                            return None
                        try:
                            return Decimal(str(raw).replace(',', '').strip())
                        except (InvalidOperation, ValueError):
                            return None
                    unit_cost = Decimal('0')
                    up = _parse_money(unit_price_raw)
                    if up is not None:
                        unit_cost = up
                    else:
                        tp = _parse_money(total_price_raw)
                        if tp is not None and qty > 0:
                            unit_cost = (tp / qty).quantize(Decimal('0.0001'))

                    # Detect if this is a section header.
                    # Section headers: whole numbers (1, 2, 3) or roman numerals,
                    # with a title but NO quantity data and NO decimal in item
                    # number (1.1, 2.3 are line items).
                    is_section = False
                    if item_no and description:
                        has_qty = qty_val is not None and qty_val != '' and qty_val != 0
                        is_decimal_item = '.' in item_no
                        if not is_decimal_item and not has_qty:
                            if re.match(r'^[IVX]+$', item_no) or re.match(r'^\d+$', item_no):
                                is_section = True

                    if is_section:
                        # Create section. If the imported section number
                        # collides with an existing one (or with one we just
                        # created in this import), bump it to the next free
                        # auto-numbered slot so existing items stay intact.
                        section_num = item_no
                        existing_nums = set(
                            CostingSection.objects.filter(costing_sheet=sheet)
                            .values_list('section_number', flat=True)
                        )
                        if section_num in existing_nums:
                            section_num = str(next_auto_num)
                            next_auto_num += 1
                        else:
                            try:
                                next_auto_num = max(next_auto_num, int(section_num) + 1)
                            except (TypeError, ValueError):
                                pass

                        current_section = CostingSection.objects.create(
                            costing_sheet=sheet,
                            section_number=section_num,
                            title=description[:255],
                            order=next_order,
                        )
                        next_order += 1
                        sections_created += 1
                        item_order = 0
                    else:
                        # Create line item
                        if not current_section:
                            # No section header in the file yet. Either
                            # continue appending to the last existing
                            # section or create a default "Imported Items"
                            # bucket.
                            if existing_sections:
                                current_section = existing_sections[-1]
                                last_item = current_section.line_items.order_by('-order').first()
                                item_order = (last_item.order + 1) if last_item else 0
                            else:
                                current_section = CostingSection.objects.create(
                                    costing_sheet=sheet,
                                    section_number=str(next_auto_num),
                                    title='Imported Items',
                                    order=next_order,
                                )
                                next_order += 1
                                next_auto_num += 1
                                sections_created += 1
                                item_order = 0

                        CostingLineItem.objects.create(
                            section=current_section,
                            item_number=item_no or f'{current_section.section_number}.{item_order + 1}',
                            description=description[:500] if description else 'No description',
                            make=make[:100] if make else '',
                            model_number=model[:100] if model else '',
                            quantity=qty,
                            unit=unit[:20] if unit else 'EA',
                            vendor_name=vendor[:255] if vendor else '',
                            base_unit_cost=unit_cost,
                            order=item_order
                        )
                        items_created += 1
                        item_order += 1

            messages.success(request, f'Import successful! Added {sections_created} sections and {items_created} line items.')
            return redirect('costing:detail', pk=pk)

        except (KeyError, ValueError, TypeError) as e:
            # Known data-shape problems (missing columns, bad cell types).
            # Roll back is automatic from the atomic block.
            logger.warning('Costing import: bad row data', exc_info=e)
            messages.error(request, 'The Excel file has an unexpected structure. Please check the column headers and row data.')
            return redirect('costing:import_excel', pk=pk)
        except Exception:
            # Anything else is a real bug — log full traceback for the
            # admin, show a generic message to the user.
            logger.exception('Costing import failed unexpectedly for sheet pk=%s', pk)
            messages.error(request, 'An unexpected error occurred while importing. The administrator has been notified.')
            return redirect('costing:import_excel', pk=pk)

    # GET request - show import form
    return render(request, 'costing/import_excel.html', {
        'sheet': sheet,
    })


# ─── Excel Import (New Sheet) ────────────────────────────────────

@login_required
def costing_import_new(request):
    """Import Excel BOQ to create a new costing sheet"""
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    from decimal import Decimal, InvalidOperation
    import re

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        sheet_title = request.POST.get('title', '').strip()

        if not excel_file:
            messages.error(request, 'Please select an Excel file to import.')
            return redirect('costing:import_new')

        if not sheet_title:
            # Use filename as title
            sheet_title = excel_file.name.replace('.xlsx', '').replace('.xls', '').replace('_', ' ')

        # Parse the file outside any DB transaction so a bad upload
        # never holds a write lock and we can return clean errors.
        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active  # Use first sheet
        except InvalidFileException:
            messages.error(request, 'The uploaded file is not a valid Excel workbook.')
            return redirect('costing:import_new')
        except (KeyError, ValueError, OSError) as e:
            logger.warning('Costing import (new): failed to open Excel file', exc_info=e)
            messages.error(request, 'Could not read the Excel file. Please check the format and try again.')
            return redirect('costing:import_new')

        # Find header row (look for row with "Vendor", "Part#", "Description")
        header_row = None
        headers = {}
        for row_num in range(1, 15):
            row = [cell.value for cell in ws[row_num]]
            row_lower = [str(c).lower() if c else '' for c in row]
            if 'vendor' in row_lower or 'description' in row_lower:
                header_row = row_num
                for i, val in enumerate(row):
                    if val:
                        headers[str(val).lower().strip()] = i
                break

        if not header_row:
            messages.error(request, 'Could not find header row in Excel file. Looking for columns: Item No, Description, Vendor, Make, Model, Qty, Unit')
            return redirect('costing:import_new')

        # Map column indices. Unit price / total price are optional —
        # only consumed when the BOQ already has prices (saves manual
        # entry for hundreds of line items). Currency is intentionally
        # NOT read from the file; line items take supplier_currency from
        # the sheet default, which the user can flip on the detail page.
        col_map = {
            'item_no': headers.get('item no', headers.get('item #', headers.get('rfp item #', headers.get('rfp item', headers.get('#', -1))))),
            'description': headers.get('description', headers.get('desc', -1)),
            'make': headers.get('make', headers.get('manufacturer', -1)),
            'model': headers.get('model', headers.get('part#', headers.get('part', headers.get('part no', -1)))),
            'qty': headers.get('qty', headers.get('quantity', -1)),
            'unit': headers.get('unit', headers.get('uom', -1)),
            'vendor': headers.get('vendor', headers.get('vendor name', -1)),
            'unit_price': headers.get('unit price', headers.get('unit cost', headers.get('price', headers.get('cost', headers.get('rate', headers.get('unit rate', -1)))))),
            'total_price': headers.get('total price', headers.get('total cost', headers.get('total', headers.get('amount', headers.get('extended price', -1))))),
        }

        # All DB writes (including the new sheet itself) happen inside
        # a single atomic block so a partial failure rolls everything
        # back instead of leaving an empty/orphaned costing sheet.
        try:
            with transaction.atomic():
                sheet = CostingSheet.objects.create(
                    title=sheet_title,
                    created_by=request.user,
                    status='draft'
                )

                sections_created = 0
                items_created = 0
                current_section = None
                item_order = 0

                for row_num in range(header_row + 1, ws.max_row + 1):
                    row = [cell.value for cell in ws[row_num]]
                    if not any(row):
                        continue

                    # Get values
                    item_no = str(row[col_map['item_no']] or '').strip() if col_map['item_no'] >= 0 and col_map['item_no'] < len(row) else ''
                    description = str(row[col_map['description']] or '').strip() if col_map['description'] >= 0 and col_map['description'] < len(row) else ''
                    make = str(row[col_map['make']] or '').strip() if col_map['make'] >= 0 and col_map['make'] < len(row) else ''
                    model = str(row[col_map['model']] or '').strip() if col_map['model'] >= 0 and col_map['model'] < len(row) else ''
                    qty_val = row[col_map['qty']] if col_map['qty'] >= 0 and col_map['qty'] < len(row) else None
                    unit = str(row[col_map['unit']] or '').strip() if col_map['unit'] >= 0 and col_map['unit'] < len(row) else 'EA'
                    vendor = str(row[col_map['vendor']] or '').strip() if col_map['vendor'] >= 0 and col_map['vendor'] < len(row) else ''
                    unit_price_raw = row[col_map['unit_price']] if col_map['unit_price'] >= 0 and col_map['unit_price'] < len(row) else None
                    total_price_raw = row[col_map['total_price']] if col_map['total_price'] >= 0 and col_map['total_price'] < len(row) else None

                    # Skip empty rows or sheet-level headings
                    if not item_no and not description:
                        continue
                    if item_no and not description and not make and not model and qty_val is None:
                        continue

                    # Parse quantity
                    try:
                        qty = Decimal(str(qty_val)) if qty_val else Decimal('1')
                    except (InvalidOperation, ValueError):
                        qty = Decimal('1')

                    # Parse price. Prefer unit price; fall back to total/qty.
                    # Strip thousands separators that Excel sometimes carries through.
                    def _parse_money(raw):
                        if raw is None or raw == '':
                            return None
                        try:
                            return Decimal(str(raw).replace(',', '').strip())
                        except (InvalidOperation, ValueError):
                            return None
                    unit_cost = Decimal('0')
                    up = _parse_money(unit_price_raw)
                    if up is not None:
                        unit_cost = up
                    else:
                        tp = _parse_money(total_price_raw)
                        if tp is not None and qty > 0:
                            unit_cost = (tp / qty).quantize(Decimal('0.0001'))

                    # Detect if this is a section header
                    is_section = False
                    if item_no and description:
                        has_qty = qty_val is not None and qty_val != '' and qty_val != 0
                        is_decimal_item = '.' in item_no
                        if not is_decimal_item and not has_qty:
                            if re.match(r'^[IVX]+$', item_no) or re.match(r'^\d+$', item_no):
                                is_section = True

                    if is_section:
                        current_section = CostingSection.objects.create(
                            costing_sheet=sheet,
                            section_number=item_no,
                            title=description[:255],
                            order=sections_created
                        )
                        sections_created += 1
                        item_order = 0
                    else:
                        if not current_section:
                            current_section = CostingSection.objects.create(
                                costing_sheet=sheet,
                                section_number='1',
                                title='Imported Items',
                                order=0
                            )
                            sections_created += 1

                        CostingLineItem.objects.create(
                            section=current_section,
                            item_number=item_no or f'{current_section.section_number}.{item_order + 1}',
                            description=description[:500] if description else 'No description',
                            make=make[:100] if make else '',
                            model_number=model[:100] if model else '',
                            quantity=qty,
                            unit=unit[:20] if unit else 'EA',
                            vendor_name=vendor[:255] if vendor else '',
                            base_unit_cost=unit_cost,
                            order=item_order
                        )
                        items_created += 1
                        item_order += 1

            messages.success(request, f'Import successful! Created costing sheet "{sheet.title}" with {sections_created} sections and {items_created} line items.')
            return redirect('costing:detail', pk=sheet.pk)

        except (KeyError, ValueError, TypeError) as e:
            # Known data-shape problems. Atomic block already rolled back.
            logger.warning('Costing import (new): bad row data', exc_info=e)
            messages.error(request, 'The Excel file has an unexpected structure. Please check the column headers and row data.')
            return redirect('costing:import_new')
        except Exception:
            logger.exception('Costing import (new) failed unexpectedly')
            messages.error(request, 'An unexpected error occurred while importing. The administrator has been notified.')
            return redirect('costing:import_new')

    # GET request - show import form
    return render(request, 'costing/import_new.html', {})
