from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Count, Q
from django.core.paginator import Paginator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.urls import reverse_lazy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, date, timedelta
from dateutil.parser import parse as parse_date

from projects.models import Project, Region, ProjectStatus
from accounts.decorators import manager_or_admin_required
from .models import Vendor, EPC, Exhibition, ProcurementPortal, Certification, SalesContact, SalesCallReport, SalesCallResponse
from .forms import SalesCallReportForm, SalesCallReportFilterForm


@login_required
def index(request):
    """Reports index page"""
    user = request.user

    # Get project counts by category
    if user.is_admin_user:
        projects = Project.objects.all()
    elif user.is_manager_user:
        projects = Project.objects.filter(region=user.region)
    else:
        projects = Project.objects.filter(owner=user)

    stats = {
        'total': projects.count(),
        'active': projects.filter(status__category='active').count(),
        'hot_leads': projects.filter(status__category='hot_lead').count(),
        'won': projects.filter(status__category='won').count(),
        'lost': projects.filter(status__category='lost').count(),
    }

    return render(request, 'reports/index.html', {'stats': stats})


@login_required
def export_excel(request):
    """Export projects to Excel"""
    user = request.user

    # Filter based on role
    if user.is_admin_user:
        projects = Project.objects.all()
    elif user.is_manager_user:
        projects = Project.objects.filter(region=user.region)
    else:
        projects = Project.objects.filter(owner=user)

    # Apply filters from request
    region_id = request.GET.get('region')
    status_id = request.GET.get('status')
    category = request.GET.get('category')

    if region_id:
        projects = projects.filter(region_id=region_id)
    if status_id:
        projects = projects.filter(status_id=status_id)
    if category:
        projects = projects.filter(status__category=category)

    projects = projects.select_related('status', 'region', 'owner')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1c2c", end_color="1a1c2c", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        'S/N', 'Proposal Reference', 'Project Name', 'Client RFQ Ref',
        'Region', 'Status', 'Owner', 'EPC', 'Est. Value',
        'PO Quarter', 'Success Quotient', 'Submission Date',
        'Est. PO Date', 'Remarks', 'Notes'
    ]

    # Write title
    ws.merge_cells('A1:O1')
    ws['A1'] = f"LEAP NETWORKS - SALES PIPELINE REPORT - {datetime.now().strftime('%d/%m/%Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row, project in enumerate(projects, 4):
        data = [
            row - 3,
            project.proposal_reference,
            project.project_name,
            project.client_rfq_reference,
            project.region.code if project.region else '',
            project.status.name if project.status else '',
            str(project.owner) if project.owner else '',
            project.epc,
            float(project.estimated_value),
            project.po_award_quarter,
            float(project.success_quotient),
            project.submission_deadline.strftime('%Y-%m-%d') if project.submission_deadline else '',
            project.estimated_po_date.strftime('%Y-%m-%d') if project.estimated_po_date else '',
            project.remarks,
            project.notes
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border

    # Adjust column widths
    column_widths = [5, 20, 40, 20, 8, 12, 15, 15, 15, 10, 12, 15, 15, 30, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)].width = width

    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"leap_pipeline_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


@manager_or_admin_required
def import_excel(request):
    """Import projects from Excel"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            # Read Excel file with openpyxl
            wb = load_workbook(excel_file, data_only=True)

            imported = 0
            errors = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Skip empty sheets
                if ws.max_row < 2:
                    continue

                # Find header row
                header_row = None
                headers = {}
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
                    row_str = ' '.join([str(v) for v in row if v])
                    if 'Project Name' in row_str or 'Proposal Reference' in row_str:
                        header_row = row_idx
                        for col_idx, cell in enumerate(row):
                            if cell:
                                headers[str(cell).strip()] = col_idx
                        break

                if not header_row:
                    continue

                # Column mapping
                col_map = {
                    'project_name': headers.get('Project Name'),
                    'proposal_reference': headers.get('Leap Proposal Reference'),
                    'client_rfq_reference': headers.get('Client RFQ Ref Number'),
                    'epc': headers.get('EPC'),
                    'estimated_value': headers.get('Est. Value (GBP)') or headers.get('Est. Value (SAR)'),
                    'po_award_quarter': headers.get('PO Award - Q'),
                    'remarks': headers.get('Remarks'),
                    'notes': headers.get('Notes'),
                }

                # Determine region from sheet name
                region_code = 'UK'
                if 'LNA' in sheet_name.upper():
                    region_code = 'LNA'
                elif 'PA' in sheet_name.upper():
                    region_code = 'PA'
                elif 'GLOBAL' in sheet_name.upper():
                    region_code = 'GLB'

                region = Region.objects.filter(code=region_code).first() or Region.objects.first()

                # Determine status category
                status_category = 'active'
                if 'HOT' in sheet_name.upper():
                    status_category = 'hot_lead'
                elif 'WON' in sheet_name.upper():
                    status_category = 'won'
                elif 'LOST' in sheet_name.upper():
                    status_category = 'lost'

                default_status = ProjectStatus.objects.filter(category=status_category).first() or ProjectStatus.objects.first()

                # Process data rows
                for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
                    try:
                        def get_val(key):
                            idx = col_map.get(key)
                            return row[idx] if idx is not None and idx < len(row) else None

                        proposal_ref = get_val('proposal_reference')
                        project_name = get_val('project_name')

                        if not proposal_ref or not project_name:
                            continue

                        # Parse estimated value
                        est_value = get_val('estimated_value') or 0
                        try:
                            est_value = float(str(est_value).replace(',', '').replace('£', '').replace('$', ''))
                        except (ValueError, TypeError):
                            est_value = 0

                        Project.objects.update_or_create(
                            proposal_reference=str(proposal_ref).strip(),
                            defaults={
                                'project_name': str(project_name).strip()[:500],
                                'client_rfq_reference': str(get_val('client_rfq_reference') or '')[:255],
                                'region': region,
                                'status': default_status,
                                'epc': str(get_val('epc') or '')[:200],
                                'estimated_value': est_value,
                                'po_award_quarter': str(get_val('po_award_quarter') or '')[:5],
                                'remarks': str(get_val('remarks') or '')[:1000],
                                'notes': str(get_val('notes') or '')[:1000],
                                'created_by': request.user,
                            }
                        )
                        imported += 1

                    except Exception as e:
                        errors.append(f"Row {row_idx} in {sheet_name}: {str(e)}")

            if imported > 0:
                messages.success(request, f'Successfully imported {imported} projects.')
            if errors:
                messages.warning(request, f'Some rows had errors: {len(errors)} errors.')

            return redirect('reports:index')

        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')

    regions = Region.objects.filter(is_active=True)
    return render(request, 'reports/import.html', {'regions': regions})


@login_required
def summary_report(request):
    """Summary report view"""
    user = request.user

    if user.is_admin_user:
        projects = Project.objects.all()
    elif user.is_manager_user:
        projects = Project.objects.filter(region=user.region)
    else:
        projects = Project.objects.filter(owner=user)

    # Summary by region
    region_summary = Region.objects.filter(is_active=True).annotate(
        total_projects=Count('projects', filter=projects.filter(id__isnull=False).values('id')),
        active_count=Count('projects', filter=projects.filter(status__category='active').values('id')),
        hot_leads_count=Count('projects', filter=projects.filter(status__category='hot_lead').values('id')),
        won_count=Count('projects', filter=projects.filter(status__category='won').values('id')),
        lost_count=Count('projects', filter=projects.filter(status__category='lost').values('id')),
        total_value=Sum('projects__estimated_value'),
    )

    # Summary by status
    status_summary = projects.values('status__name', 'status__category', 'status__color').annotate(
        count=Count('id'),
        value=Sum('estimated_value')
    ).order_by('status__category')

    # Summary by quarter
    quarter_summary = projects.exclude(po_award_quarter='').values('po_award_quarter').annotate(
        count=Count('id'),
        value=Sum('estimated_value')
    ).order_by('po_award_quarter')

    context = {
        'region_summary': region_summary,
        'status_summary': status_summary,
        'quarter_summary': quarter_summary,
        'total_projects': projects.count(),
        'total_value': projects.aggregate(Sum('estimated_value'))['estimated_value__sum'] or 0,
    }

    return render(request, 'reports/summary.html', context)


@login_required
def annual_report(request):
    """Annual Report view - shows vendors, EPCs, exhibitions, certifications, etc."""
    user = request.user

    # Get project pipeline stats
    if user.is_admin_user:
        projects = Project.objects.all()
    elif user.is_manager_user:
        projects = Project.objects.filter(region=user.region)
    else:
        projects = Project.objects.filter(owner=user)

    # Pipeline stats by region
    def get_region_pipeline(region_codes):
        region_projects = projects.filter(region__code__in=region_codes)
        return {
            'active': region_projects.filter(status__category='active').count(),
            'hot_leads': region_projects.filter(status__category='hot_lead').count(),
            'won': region_projects.filter(status__category='won').count(),
            'lost': region_projects.filter(status__category='lost').count(),
            'active_value': region_projects.filter(status__category='active').aggregate(Sum('estimated_value'))['estimated_value__sum'] or 0,
            'won_value': region_projects.filter(status__category='won').aggregate(Sum('estimated_value'))['estimated_value__sum'] or 0,
        }

    pipeline_stats = {
        'lnuk': get_region_pipeline(['UK', 'GLB']),
        'lna': get_region_pipeline(['LNA']),
        'pa': get_region_pipeline(['PA']),
    }

    # Get annual report data
    vendors = Vendor.objects.filter(is_active=True)
    epcs = EPC.objects.filter(is_active=True)
    exhibitions = Exhibition.objects.all().order_by('-year', 'name')
    portals = ProcurementPortal.objects.filter(is_active=True)
    certifications = Certification.objects.all()
    contacts = SalesContact.objects.all()

    # Stats
    vendor_stats = {
        'total': vendors.count(),
        'vendors': vendors.filter(vendor_type='vendor').count(),
        'distributors': vendors.filter(vendor_type='distributor').count(),
        'partners': vendors.filter(vendor_type='partner').count(),
        'oems': vendors.filter(vendor_type='oem').count(),
    }

    portal_stats = {
        'total': portals.count(),
        'free': portals.filter(registration_type='free').count(),
        'freemium': portals.filter(registration_type='freemium').count(),
        'paid': portals.filter(registration_type='paid').count(),
    }

    cert_stats = {
        'total': certifications.count(),
        'obtained': certifications.filter(status='obtained').count(),
        'in_progress': certifications.filter(status='in_progress').count(),
        'pending': certifications.filter(status='pending').count(),
    }

    contact_stats = {
        'total': contacts.count(),
        'contacted': contacts.filter(is_contacted=True).count(),
        'pending': contacts.filter(is_contacted=False).count(),
    }

    context = {
        'pipeline_stats': pipeline_stats,
        'vendors': vendors,
        'vendor_stats': vendor_stats,
        'epcs': epcs,
        'exhibitions': exhibitions,
        'portals': portals,
        'portal_stats': portal_stats,
        'certifications': certifications,
        'cert_stats': cert_stats,
        'contacts': contacts,
        'contact_stats': contact_stats,
    }

    return render(request, 'reports/annual_report.html', context)


# ============================================
# Sales Call Report Views
# ============================================

class SalesCallReportListView(LoginRequiredMixin, ListView):
    """List all sales call reports with filtering and role-based access"""
    model = SalesCallReport
    template_name = 'reports/sales_call_list.html'
    context_object_name = 'reports'
    paginate_by = 25

    def get_queryset(self):
        user = self.request.user

        # Role-based filtering
        if user.is_admin_user:
            queryset = SalesCallReport.objects.all()
        elif user.is_manager_user:
            queryset = SalesCallReport.objects.all()
        else:
            queryset = SalesCallReport.objects.filter(sales_rep=user)

        queryset = queryset.select_related('sales_rep', 'sales_rep__region')

        # Region filter
        region_code = self.request.GET.get('region')
        if region_code and (user.is_admin_user or user.is_manager_user):
            queryset = queryset.filter(sales_rep__region__code=region_code)

        # Apply filters
        search = self.request.GET.get('search')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        action_type = self.request.GET.get('action_type')
        contact_type = self.request.GET.get('contact_type')
        system_category = self.request.GET.get('system_category')
        goal = self.request.GET.get('goal')
        sales_rep_id = self.request.GET.get('sales_rep')

        if search:
            queryset = queryset.filter(
                Q(company_name__icontains=search) |
                Q(contact_name__icontains=search) |
                Q(email__icontains=search) |
                Q(comments__icontains=search)
            )

        if date_from:
            queryset = queryset.filter(call_date__gte=date_from)

        if date_to:
            queryset = queryset.filter(call_date__lte=date_to)

        if action_type:
            queryset = queryset.filter(action_type=action_type)

        if contact_type:
            queryset = queryset.filter(contact_type=contact_type)

        if system_category:
            queryset = queryset.filter(system_categories__icontains=system_category)

        if goal:
            queryset = queryset.filter(goal=goal)

        if sales_rep_id and (user.is_admin_user or user.is_manager_user):
            queryset = queryset.filter(sales_rep_id=sales_rep_id)

        return queryset.order_by('-call_date', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        user = self.request.user
        active_region = self.request.GET.get('region', '')

        # Regions for tabs (admin/manager only)
        if user.is_admin_user or user.is_manager_user:
            from projects.models import Region
            context['regions'] = Region.objects.filter(is_active=True).order_by('code')

        context['active_region'] = active_region

        context['filter_form'] = SalesCallReportFilterForm(
            self.request.GET,
            user=user,
            region_code=active_region or None,
        )

        queryset = self.get_queryset()
        context['total_count'] = queryset.count()
        context['today_count'] = queryset.filter(call_date=date.today()).count()
        context['this_week_count'] = queryset.filter(
            call_date__gte=date.today() - timedelta(days=7)
        ).count()

        return context


class SalesCallReportCreateView(LoginRequiredMixin, CreateView):
    """Create a new sales call report"""
    model = SalesCallReport
    form_class = SalesCallReportForm
    template_name = 'reports/sales_call_form.html'
    success_url = reverse_lazy('reports:sales_call_list')

    def get_initial(self):
        initial = super().get_initial()
        initial['call_date'] = date.today()
        return initial

    def form_valid(self, form):
        form.instance.sales_rep = self.request.user
        messages.success(self.request, 'Sales call report created successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add Sales Call Report'
        context['button_text'] = 'Save Report'
        return context


class SalesCallReportUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    """Update a sales call report"""
    model = SalesCallReport
    form_class = SalesCallReportForm
    template_name = 'reports/sales_call_form.html'
    success_url = reverse_lazy('reports:sales_call_list')

    def test_func(self):
        report = self.get_object()
        user = self.request.user
        # Only allow editing own reports or admin/manager
        return user.is_admin_user or user.is_manager_user or report.sales_rep == user

    def form_valid(self, form):
        messages.success(self.request, 'Sales call report updated successfully.')
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Edit Sales Call Report'
        context['button_text'] = 'Update Report'
        return context


class SalesCallReportDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """View sales call report details"""
    model = SalesCallReport
    template_name = 'reports/sales_call_detail.html'
    context_object_name = 'report'

    def test_func(self):
        report = self.get_object()
        user = self.request.user
        return user.is_admin_user or user.is_manager_user or report.sales_rep == user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['responses'] = self.object.responses.select_related('responder').all()
        context['can_respond'] = self.request.user.is_admin_user or self.request.user.is_manager_user
        return context

    def post(self, request, *args, **kwargs):
        """Handle adding a response"""
        self.object = self.get_object()
        user = request.user

        # Only admin and manager can add responses
        if not (user.is_admin_user or user.is_manager_user):
            messages.error(request, 'You do not have permission to add responses.')
            return redirect('reports:sales_call_detail', pk=self.object.pk)

        message = request.POST.get('response_message', '').strip()
        if message:
            SalesCallResponse.objects.create(
                sales_call=self.object,
                responder=user,
                message=message
            )
            messages.success(request, 'Response added successfully.')
        else:
            messages.warning(request, 'Please enter a response message.')

        return redirect('reports:sales_call_detail', pk=self.object.pk)


class SalesCallReportDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Delete a sales call report"""
    model = SalesCallReport
    template_name = 'reports/sales_call_confirm_delete.html'
    success_url = reverse_lazy('reports:sales_call_list')

    def test_func(self):
        report = self.get_object()
        user = self.request.user
        # Admin and Manager can delete any, Sales Rep can only delete their own
        return user.is_admin_user or user.is_manager_user or report.sales_rep == user

    def form_valid(self, form):
        messages.success(self.request, 'Sales call report deleted successfully.')
        return super().form_valid(form)


@login_required
def export_sales_call_reports(request):
    """Export sales call reports to Excel"""
    user = request.user

    # Get queryset based on role
    if user.is_admin_user:
        queryset = SalesCallReport.objects.all()
    elif user.is_manager_user:
        queryset = SalesCallReport.objects.all()
    else:
        queryset = SalesCallReport.objects.filter(sales_rep=user)

    # Apply filters from request
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if date_from:
        queryset = queryset.filter(call_date__gte=date_from)
    if date_to:
        queryset = queryset.filter(call_date__lte=date_to)

    queryset = queryset.select_related('sales_rep').order_by('-call_date')

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Call Reports"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a1c2c", end_color="1a1c2c", fill_type="solid")

    # Headers
    headers = [
        'Date', 'Sales Rep', 'Action Type', 'Contact Type', 'System',
        'Company', 'Contact Name', 'Title', 'Role', 'Phone', 'Email',
        'Goal', 'Comments', 'Next Action Date', 'Next Action Type'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Write data
    for row, report in enumerate(queryset, 2):
        data = [
            report.call_date.strftime('%Y-%m-%d'),
            str(report.sales_rep),
            report.get_action_type_display(),
            report.get_contact_type_display(),
            report.get_system_category_display(),
            report.company_name,
            report.contact_name,
            report.get_title_display() if report.title else '',
            report.role,
            report.phone,
            report.email,
            report.get_goal_display(),
            report.comments,
            report.next_action_date.strftime('%Y-%m-%d') if report.next_action_date else '',
            report.get_next_action_type_display() if report.next_action_type else ''
        ]
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"sales_call_reports_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response
