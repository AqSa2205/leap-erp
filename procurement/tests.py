import base64
import io
from datetime import date
from decimal import Decimal
from unittest import mock

from django.test import TestCase
from django.urls import reverse
from PIL import Image

from accounts.models import User, Role
from procurement.models import PurchaseOrder


def _png_data_url():
    buf = io.BytesIO()
    Image.new('RGBA', (8, 8), (0, 0, 0, 0)).save(buf, 'PNG')
    return 'data:image/png;base64,' + base64.b64encode(buf.getvalue()).decode()


class POApproveStageTests(TestCase):
    """The Approve-&-Sign endpoint must always answer with JSON so the modal can
    show a real message — a storage failure must not leak an HTML 500 that the
    browser reports as an opaque 'Network error'."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user(
            'boss', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-TEST-1',
            vendor_name='ACME', po_issued_by='Tester',
            created_by=self.user)

    def _url(self, stage='scm'):
        return reverse('procurement:po_approve_stage',
                       kwargs={'pk': self.po.pk, 'stage': stage})

    def test_happy_path_signs_first_stage(self):
        r = self.client.post(self._url('scm'), {'signature_data': _png_data_url()})
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()['ok'])
        self.po.refresh_from_db()
        self.assertIsNotNone(self.po.scm_approved_at)

    def test_missing_signature_is_json_400(self):
        r = self.client.post(self._url('scm'), {})
        self.assertEqual(r.status_code, 400)
        self.assertEqual(r['Content-Type'], 'application/json')
        self.assertIn('signature', r.json()['error'].lower())

    def test_out_of_sequence_stage_is_json_400(self):
        # 'pm' before 'scm' is signed — rejected as out of sequence.
        r = self.client.post(self._url('pm'), {'signature_data': _png_data_url()})
        self.assertEqual(r.status_code, 400)
        self.assertIn('sequence', r.json()['error'].lower())

    def test_unpriced_pdf_omits_pricing(self):
        from procurement.models import PurchaseOrderItem
        PurchaseOrderItem.objects.create(
            purchase_order=self.po, serial_number=1,
            description='Widget', quantity=2, rate_per_unit=100)
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader

        def pdf_text(url_name):
            r = self.client.get(
                reverse(url_name, kwargs={'pk': self.po.pk}))
            self.assertEqual(r.status_code, 200)
            self.assertEqual(r['Content-Type'], 'application/pdf')
            reader = PdfReader(io.BytesIO(r.content))
            return '\n'.join((p.extract_text() or '') for p in reader.pages)

        priced = pdf_text('procurement:po_export_pdf')
        self.assertIn('Rate/Unit', priced)
        self.assertIn('Total (SAR)', priced)

        unpriced = pdf_text('procurement:po_export_pdf_unpriced')
        self.assertNotIn('Rate/Unit', unpriced)
        self.assertNotIn('Total (SAR)', unpriced)
        self.assertNotIn('Amount in words', unpriced)
        self.assertIn('UNPRICED', unpriced)

    def test_out_of_scope_user_gets_404(self):
        # A plain user who didn't create the PO can't see (or sign) it.
        outsider = User.objects.create_user('outsider', password='pw')
        self.client.force_login(outsider)
        r = self.client.post(self._url('scm'), {'signature_data': _png_data_url()})
        self.assertEqual(r.status_code, 404)
        self.po.refresh_from_db()
        self.assertIsNone(self.po.scm_approved_at)

    def test_storage_failure_returns_json_500_not_html(self):
        # Simulate object storage / DB write blowing up during the save. The
        # endpoint must answer with JSON 500 carrying the exception detail, not
        # an HTML 500 page (which the browser reports as "Network error").
        with mock.patch.object(PurchaseOrder, 'save', side_effect=Exception('R2 down')):
            r = self.client.post(self._url('scm'),
                                 {'signature_data': _png_data_url()})
        self.assertEqual(r.status_code, 500)
        self.assertEqual(r['Content-Type'], 'application/json')
        err = r.json()['error']
        self.assertIn('Approval failed', err)
        self.assertIn('R2 down', err)
        # Nothing should have been committed.
        self.po.refresh_from_db()
        self.assertIsNone(self.po.scm_approved_at)


class POPdfFooterTests(TestCase):
    """The PO PDF footer shows the company (left), PO number (centre) and page
    number (right) on every page — on both the priced and unpriced exports."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('boss', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-FOOT-1', vendor_name='ACME',
            po_issued_by='Tester', cost_center='projects', created_by=self.user)

    def _pdf_text(self, url_name='procurement:po_export_pdf'):
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader
        r = self.client.get(reverse(url_name, kwargs={'pk': self.po.pk}))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')
        return '\n'.join((p.extract_text() or '') for p in PdfReader(io.BytesIO(r.content)).pages)

    def test_footer_has_company_mr_po_number_and_page(self):
        text = ' '.join(self._pdf_text().split())   # normalise wrapping
        self.assertIn('Leap Networks Arabia', text)          # company (left, line 1)
        self.assertIn('Material Requisition R00', text)      # MR + default revision (left, line 2)
        self.assertIn('Page 1 of', text)                     # page label (right)
        # PO number appears in the header AND the footer centre -> at least twice.
        self.assertGreaterEqual(text.count('PO-FOOT-1'), 2)

    def test_footer_reflects_edited_revision(self):
        self.po.mr_revision = 'R02'
        self.po.save(update_fields=['mr_revision'])
        text = ' '.join(self._pdf_text().split())
        self.assertIn('Material Requisition R02', text)
        self.assertNotIn('Material Requisition R00', text)

    def test_revision_is_editable_on_the_form(self):
        from procurement.forms import PurchaseOrderForm
        self.assertIn('mr_revision', PurchaseOrderForm().fields)

    def test_revision_defaults_to_r00(self):
        fresh = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 2), po_number='PO-FOOT-2', vendor_name='X',
            po_issued_by='T', cost_center='projects', created_by=self.user)
        self.assertEqual(fresh.mr_revision, 'R00')

    def test_unpriced_pdf_also_has_footer(self):
        text = ' '.join(self._pdf_text('procurement:po_export_pdf_unpriced').split())
        self.assertIn('Leap Networks Arabia', text)
        self.assertIn('Material Requisition R00', text)
        self.assertIn('Page 1 of', text)


class POPdfHeaderTests(TestCase):
    """The PO PDF header shows the company block (left), 'Purchase Order'
    (centre) and the logo (right). The Arabic company name renders when the
    Arabic font is present; when it isn't, the export still succeeds."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('boss', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-HDR-1', vendor_name='ACME',
            po_issued_by='Tester', cost_center='projects', created_by=self.user)

    def _text(self):
        try:
            from pypdf import PdfReader
        except ImportError:
            from PyPDF2 import PdfReader
        r = self.client.get(reverse('procurement:po_export_pdf', kwargs={'pk': self.po.pk}))
        self.assertEqual(r.status_code, 200)
        return '\n'.join((p.extract_text() or '') for p in PdfReader(io.BytesIO(r.content)).pages)

    def test_header_has_company_block_and_centre_title(self):
        text = self._text()
        self.assertIn('Leap Networks Arabia', text)
        self.assertIn('Al-Khobar, Saudi Arabia', text)
        self.assertIn('www.leap-arabia.com', text)
        self.assertIn('Purchase Order', text)

    def test_arabic_font_is_available_and_export_succeeds(self):
        # The Amiri font ships in static/fonts/, so it registers and the Arabic
        # company name renders; the export returns a valid PDF.
        from procurement.views import _arabic_font
        self.assertEqual(_arabic_font(), 'ArabicHeader')
        r = self.client.get(reverse('procurement:po_export_pdf', kwargs={'pk': self.po.pk}))
        self.assertEqual(r.status_code, 200)
        self.assertEqual(r['Content-Type'], 'application/pdf')

    def test_shape_arabic_transforms_text(self):
        from procurement.views import _shape_arabic
        raw = 'شركة لييب نتوركس أرابيا'
        shaped = _shape_arabic(raw)
        self.assertEqual(len(shaped), len(raw))   # same characters, reordered/reshaped
        self.assertNotEqual(shaped, raw)          # bidi + reshape actually changed it


class POEditFormSetTests(TestCase):
    """Editing a PO must persist line-item changes even when an existing row
    has a blank description — previously one such row silently blocked the
    entire save (the formset was invalid and no error was surfaced)."""

    def setUp(self):
        from procurement.models import PurchaseOrderItem
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('boss', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-EDIT-1', vendor_name='V',
            po_issued_by='T', cost_center='projects', created_by=self.user)
        self.good = PurchaseOrderItem.objects.create(
            purchase_order=self.po, serial_number=1, description='Good',
            quantity=1, uom='Nos', rate_per_unit=10, order=0)
        self.blank = PurchaseOrderItem.objects.create(
            purchase_order=self.po, serial_number=2, description='',
            quantity=1, uom='Nos', rate_per_unit=5, order=1)

    def _row(self, i, item, desc, qty):
        return {
            f'items-{i}-id': str(item.pk) if item else '',
            f'items-{i}-serial_number': str(i + 1),
            f'items-{i}-description': desc,
            f'items-{i}-quantity': qty, f'items-{i}-uom': 'Nos',
            f'items-{i}-rate_per_unit': '10', f'items-{i}-order': str(i),
            f'items-{i}-system': '', f'items-{i}-make_model': '',
            f'items-{i}-remarks': '', f'items-{i}-po_value_usd': '',
            f'items-{i}-advance_payment_sar': '', f'items-{i}-delivery_status': '',
            f'items-{i}-scm': '',
        }

    def test_edit_persists_despite_blank_description_row(self):
        data = {
            'po_date': '2026-01-01', 'po_number': 'PO-EDIT-1',
            'cost_center': 'projects', 'status': self.po.status,
            'vendor_name': 'V', 'po_issued_by': 'T', 'vendor_contact_email': '',
            'issuer_email': '', 'project_name': '', 'currency': 'SAR',
            'discount_rate': '0',
            'vat_rate': '15', 'lead_time': '', 'payment_terms_text': '',
            'warranty': '', 'terms_and_conditions': '', 'delivery_incoterms': '',
            'delivery_location': '',
            'items-TOTAL_FORMS': '3', 'items-INITIAL_FORMS': '2',
            'items-MIN_NUM_FORMS': '0', 'items-MAX_NUM_FORMS': '1000',
        }
        data.update(self._row(0, self.good, 'Good EDITED', '9'))
        data.update(self._row(1, self.blank, '', '7'))            # blank desc kept
        data.update(self._row(2, None, 'BRAND NEW', '4'))         # new row

        r = self.client.post(
            reverse('procurement:po_update', kwargs={'pk': self.po.pk}), data)
        self.assertEqual(r.status_code, 302)  # saved + redirected

        self.good.refresh_from_db()
        self.blank.refresh_from_db()
        self.assertEqual(self.good.quantity, 9)        # edit persisted
        self.assertEqual(self.blank.quantity, 7)       # blank-desc row updated
        self.assertTrue(
            self.po.items.filter(description='BRAND NEW').exists())  # new row added


class POImportExcelTests(TestCase):
    """Excel import must bring in ALL line items, not just the first ~18 — the
    old parser hard-capped at row 30 and silently dropped the rest, and read
    discount/VAT from fixed rows that large POs push out of range."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('imp', password='pw', role=self.sa_role)
        self.client.force_login(self.user)

    def _build_workbook(self, n_items, po_number='IMP-TEST'):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'PURCHASE ORDER'
        ws['B1'] = '01-Jan-2026'; ws['B2'] = po_number; ws['B3'] = 'Projects'
        ws['B4'] = 'Vendor X'
        ws['F1'] = 'Issuer'
        ws.cell(row=11, column=1, value='S No.')
        r = 12
        for i in range(1, n_items + 1):
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=f'Make-{i}')
            ws.cell(row=r, column=3, value=f'Item {i}')
            ws.cell(row=r, column=6, value=i)
            ws.cell(row=r, column=7, value='Nos')
            ws.cell(row=r, column=8, value=10.0)
            ws.cell(row=r, column=10, value=f'rem{i}')
            r += 1
        r += 1  # blank spacer
        for label, c6, c9 in [('Base Amount', None, 1000), ('Discount', 0.05, 50),
                              ('Gross Value', None, 950), ('VAT', 0.15, 142.5),
                              ('Total Value in SAR', None, 1092.5)]:
            ws.cell(row=r, column=3, value=label)
            if c6 is not None:
                ws.cell(row=r, column=6, value=c6)
            ws.cell(row=r, column=9, value=c9)
            r += 1
        r += 2
        ws.cell(row=r, column=1, value='TERMS AND CONDITIONS')
        ws.cell(row=r + 1, column=2, value='Payment 30 days')
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.read()

    def _import(self, content):
        from django.core.files.uploadedfile import SimpleUploadedFile
        up = SimpleUploadedFile(
            'po.xlsx', content,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        return self.client.post(reverse('procurement:po_import'), {'excel_file': up})

    def test_imports_all_items_beyond_old_row_cap(self):
        r = self._import(self._build_workbook(25, 'IMP-25'))
        self.assertEqual(r.status_code, 302)
        po = PurchaseOrder.objects.get(po_number='IMP-25')
        self.assertEqual(po.items.count(), 25)
        # discount/VAT parsed from the totals block (now well below row 30)
        self.assertEqual(po.discount_rate, 5)
        self.assertEqual(po.vat_rate, 15)
        self.assertEqual(po.items.order_by('order').last().description, 'Item 25')

    def test_small_po_still_imports(self):
        r = self._import(self._build_workbook(3, 'IMP-3'))
        self.assertEqual(r.status_code, 302)
        po = PurchaseOrder.objects.get(po_number='IMP-3')
        self.assertEqual(po.items.count(), 3)
        self.assertIn('Payment 30 days', po.terms_and_conditions)


class POCurrencyTests(TestCase):
    """A PO carries a selectable currency (SAR/USD/EUR/AED) that drives the
    rate/total labels, the amount-in-words fraction unit, and the exports."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('cur', password='pw', role=self.sa_role)
        self.client.force_login(self.user)

    def test_amount_in_words_fraction_unit_per_currency(self):
        from procurement.views import _amount_in_words
        from decimal import Decimal
        self.assertIn('Halalas', _amount_in_words(Decimal('10.50'), 'SAR'))
        self.assertIn('Cents', _amount_in_words(Decimal('10.50'), 'USD'))
        self.assertIn('Cents', _amount_in_words(Decimal('10.50'), 'EUR'))
        self.assertIn('Fils', _amount_in_words(Decimal('10.50'), 'AED'))
        self.assertIn('Pence', _amount_in_words(Decimal('10.50'), 'GBP'))
        self.assertIn('Penny', _amount_in_words(Decimal('10.01'), 'GBP'))  # singular
        self.assertIn('GBP', _amount_in_words(Decimal('10.00'), 'GBP'))

    def test_gbp_is_a_currency_choice(self):
        from procurement.models import PurchaseOrder
        codes = dict(PurchaseOrder.CURRENCY_CHOICES)
        self.assertIn('GBP', codes)
        self.assertEqual(codes['GBP'], 'GBP — British Pound')

    def test_excel_headers_show_currency(self):
        from procurement.models import PurchaseOrderItem
        from django.utils import timezone
        import io, openpyxl
        po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='CUR-1', vendor_name='V',
            po_issued_by='T', cost_center='projects', currency='AED',
            created_by=self.user)
        PurchaseOrderItem.objects.create(
            purchase_order=po, serial_number=1, description='X', quantity=1,
            uom='Nos', rate_per_unit=10, order=0)
        for s in po.required_stages:        # release so export is unlocked
            setattr(po, f'{s}_approved_at', timezone.now())
            setattr(po, f'{s}_approved_by', self.user)
        po.save()
        r = self.client.get(reverse('procurement:po_export', kwargs={'pk': po.pk}))
        self.assertEqual(r.status_code, 200)
        ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
        # Match on the header text rather than fixed coordinates so the
        # assertion survives layout changes to the sheet.
        header_texts = {c.value for row in ws.iter_rows() for c in row if c.value}
        self.assertIn('Rate/unit (AED)', header_texts)
        self.assertIn('Total Value (AED)', header_texts)


class QuotationImportTests(TestCase):
    """Supplier quotation PDFs are AI-extracted into a normalized structure,
    reviewed in the PO editor, then turned into a PO. The Anthropic call is
    mocked — these tests cover normalization + the upload/review/create flow."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('proc', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.fake = {
            'vendor_name': 'Al-Oufy', 'currency': 'USD', 'vat_rate': 15,
            'project_name': 'Nariyah', 'quotation_reference': 'Q-9',
            'line_items': [
                {'description': 'RGS Conduit 1in', 'make_model': 'ITCC',
                 'quantity': 150, 'uom': 'PCS', 'unit_price': 49.5},
                {'description': 'RGS Lock Nut 1in', 'make_model': '',
                 'quantity': 100, 'uom': 'PCS', 'unit_price': 1.0},
            ],
        }

    def test_normalize_maps_currency_and_unit_price(self):
        from procurement.quotation_extract import _normalize
        out = _normalize({
            'currency': 'SR',
            'line_items': [
                {'description': 'A', 'quantity': '5', 'unit_price': '2.50', 'uom': ''},
                {'description': '', 'quantity': 1, 'unit_price': 1},  # dropped (no desc)
            ],
        })
        self.assertEqual(out['currency'], 'SAR')
        self.assertEqual(len(out['line_items']), 1)
        self.assertEqual(out['line_items'][0]['uom'], 'Nos')   # default
        self.assertEqual(out['line_items'][0]['unit_price'], 2.5)

    def test_normalize_maps_gbp_currency(self):
        from procurement.quotation_extract import _normalize
        for raw in ('GBP', '£', 'POUND', 'Sterling'):
            out = _normalize({'currency': raw, 'line_items': [
                {'description': 'A', 'quantity': '1', 'unit_price': '1', 'uom': ''}]})
            self.assertEqual(out['currency'], 'GBP', raw)

    def _upload(self):
        from unittest import mock
        from django.core.files.uploadedfile import SimpleUploadedFile
        with mock.patch('procurement.quotation_extract.extract_text_from_pdf', return_value='text'), \
             mock.patch('procurement.quotation_extract.extract_quotation',
                        return_value=(self.fake, 'claude-sonnet-4-6')):
            up = SimpleUploadedFile('al.pdf', b'%PDF-1.4', content_type='application/pdf')
            return self.client.post(reverse('procurement:quotation_import'),
                                    {'quotation_file': up})

    def test_upload_extracts_and_redirects_to_review(self):
        from procurement.models import QuotationImport
        r = self._upload()
        qi = QuotationImport.objects.latest('id')
        self.assertEqual(qi.status, 'extracted')
        self.assertEqual(len(qi.line_items), 2)
        self.assertRedirects(
            r, reverse('procurement:quotation_review', kwargs={'pk': qi.pk}))

    def test_review_page_prefills_extraction(self):
        from procurement.models import QuotationImport
        self._upload()
        qi = QuotationImport.objects.latest('id')
        r = self.client.get(reverse('procurement:quotation_review', kwargs={'pk': qi.pk}))
        self.assertEqual(r.status_code, 200)
        body = r.content.decode()
        self.assertIn('Al-Oufy', body)
        self.assertIn('RGS Conduit 1in', body)
        self.assertIn('"USD" selected', body)

    def test_extraction_failure_marks_failed(self):
        from unittest import mock
        from django.core.files.uploadedfile import SimpleUploadedFile
        from procurement.models import QuotationImport
        with mock.patch('procurement.quotation_extract.extract_text_from_pdf', return_value='text'), \
             mock.patch('procurement.quotation_extract.extract_quotation',
                        side_effect=RuntimeError('no API key')):
            up = SimpleUploadedFile('x.pdf', b'%PDF-1.4', content_type='application/pdf')
            self.client.post(reverse('procurement:quotation_import'), {'quotation_file': up})
        qi = QuotationImport.objects.latest('id')
        self.assertEqual(qi.status, 'failed')
        self.assertIn('no API key', qi.error)

    def test_review_creates_po_with_items(self):
        from procurement.models import QuotationImport, PurchaseOrder
        self._upload()
        qi = QuotationImport.objects.latest('id')
        data = {
            'po_date': '2026-01-01', 'po_number': 'PO-Q-1', 'cost_center': 'projects',
            'status': 'draft', 'vendor_name': 'Al-Oufy', 'po_issued_by': 'Q',
            'vendor_contact_email': '', 'issuer_email': '', 'project_name': 'Nariyah',
            'currency': 'USD', 'discount_rate': '0', 'vat_rate': '15', 'lead_time': '',
            'payment_terms_text': '', 'warranty': '', 'terms_and_conditions': '',
            'delivery_incoterms': '', 'delivery_location': '',
            'items-TOTAL_FORMS': '2', 'items-INITIAL_FORMS': '0',
            'items-MIN_NUM_FORMS': '0', 'items-MAX_NUM_FORMS': '1000',
        }
        def row(i, desc, qty, price):
            return {f'items-{i}-id': '', f'items-{i}-serial_number': str(i + 1),
                    f'items-{i}-description': desc, f'items-{i}-make_model': '',
                    f'items-{i}-quantity': str(qty), f'items-{i}-uom': 'PCS',
                    f'items-{i}-rate_per_unit': str(price), f'items-{i}-order': str(i),
                    f'items-{i}-system': '', f'items-{i}-remarks': '',
                    f'items-{i}-po_value_usd': '', f'items-{i}-advance_payment_sar': '',
                    f'items-{i}-delivery_status': '', f'items-{i}-scm': ''}
        data.update(row(0, 'RGS Conduit 1in', 150, 49.5))
        data.update(row(1, 'RGS Lock Nut 1in', 100, 1.0))
        r = self.client.post(
            reverse('procurement:quotation_review', kwargs={'pk': qi.pk}), data)
        po = PurchaseOrder.objects.get(po_number='PO-Q-1')
        self.assertRedirects(r, reverse('procurement:po_detail', kwargs={'pk': po.pk}))
        self.assertEqual(po.items.count(), 2)
        self.assertEqual(po.currency, 'USD')
        qi.refresh_from_db()
        self.assertEqual(qi.status, 'converted')
        self.assertEqual(qi.purchase_order_id, po.pk)


class QuotationRetryTests(TestCase):
    """A pending/failed quotation can be re-extracted without re-uploading."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('proc2', password='pw', role=self.sa_role)
        self.client.force_login(self.user)

    def _make(self, status='failed'):
        from django.core.files.base import ContentFile
        from procurement.models import QuotationImport
        qi = QuotationImport.objects.create(
            original_filename='q.pdf', status=status, created_by=self.user,
            error='AI extraction failed: timeout')
        qi.file.save('q.pdf', ContentFile(b'%PDF-1.4'), save=True)
        return qi

    def test_retry_reextracts_failed(self):
        from unittest import mock
        from procurement.models import QuotationImport
        qi = self._make('failed')
        fake = {'vendor_name': 'V', 'currency': 'SAR', 'vat_rate': 15,
                'line_items': [{'description': 'A', 'quantity': 1, 'uom': 'Nos', 'unit_price': 2}]}
        with mock.patch('procurement.quotation_extract.extract_text_from_pdf', return_value='t'), \
             mock.patch('procurement.quotation_extract.extract_quotation', return_value=(fake, 'm')):
            r = self.client.post(reverse('procurement:quotation_retry', kwargs={'pk': qi.pk}))
        qi.refresh_from_db()
        self.assertEqual(qi.status, 'extracted')
        self.assertEqual(qi.error, '')
        self.assertRedirects(r, reverse('procurement:quotation_review', kwargs={'pk': qi.pk}))

    def test_retry_records_failure_again(self):
        from unittest import mock
        qi = self._make('pending')
        with mock.patch('procurement.quotation_extract.extract_text_from_pdf', return_value='t'), \
             mock.patch('procurement.quotation_extract.extract_quotation',
                        side_effect=RuntimeError('still down')):
            self.client.post(reverse('procurement:quotation_retry', kwargs={'pk': qi.pk}))
        qi.refresh_from_db()
        self.assertEqual(qi.status, 'failed')          # never stuck at pending
        self.assertIn('still down', qi.error)


class TermsUsageSeparationTests(TestCase):
    """Sales and procurement terms are kept in separate pickers."""

    def setUp(self):
        from costing.models import TermsTemplate
        self.sales = TermsTemplate.objects.create(
            name='Sales only', category='payment_terms', usage='sales', content='x')
        self.proc = TermsTemplate.objects.create(
            name='Proc only', category='payment_terms', usage='procurement', content='x')
        self.both = TermsTemplate.objects.create(
            name='Shared', category='payment_terms', usage='both', content='x')

    def _names(self, terms_by_category):
        out = []
        for rows in terms_by_category.values():
            out += [r['template'].name for r in rows]
        return set(out)

    def test_po_picker_excludes_sales_terms(self):
        from procurement.views import _po_terms_by_category
        names = self._names(_po_terms_by_category([]))
        self.assertIn('Proc only', names)
        self.assertIn('Shared', names)
        self.assertNotIn('Sales only', names)

    def test_po_quick_add_creates_procurement_term(self):
        from costing.models import TermsTemplate
        sa, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        u = User.objects.create_user('mgr', password='pw', role=sa)
        self.client.force_login(u)
        r = self.client.post(reverse('costing:terms_template_ajax_create'), {
            'name': 'From PO', 'category': 'exclusions',
            'content': 'no warranty', 'usage': 'procurement'})
        self.assertEqual(r.status_code, 200)
        t = TermsTemplate.objects.get(name='From PO')
        self.assertEqual(t.usage, 'procurement')


class BudgetProcurementFlowTests(TestCase):
    """Procurement works from the finance-approved budget, not the costing
    sheet: budgeted price flows onto the PO and the costing sheet is off-limits."""

    def setUp(self):
        from projects.models import Region, ProjectStatus, Project
        from costing.models import CostingSheet, CostingSection, CostingLineItem
        self.sa, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.proc_role, _ = Role.objects.get_or_create(name=Role.PROCUREMENT_MGR)
        self.region = Region.objects.create(name='Arabia')
        self.won = ProjectStatus.objects.create(name='Won', category='won')
        self.project = Project.objects.create(
            project_name='P', status=self.won, region=self.region)
        self.sheet = CostingSheet.objects.create(
            title='S', project=self.project, margin=Decimal('30'),
            discount_rate=Decimal('0'), workflow_stage='finance_approved')
        self.sec = CostingSection.objects.create(
            costing_sheet=self.sheet, section_number='A.1', title='CCTV', order=0)
        self.item = CostingLineItem.objects.create(
            section=self.sec, item_number='1', description='Cam', quantity=Decimal('2'),
            unit='EA', make='Bosch', model_number='X1', vendor_name='Acme',
            base_unit_cost=Decimal('100'), supplier_currency='SAR')
        self.proc = User.objects.create_user(
            'proc', password='pw', role=self.proc_role, region=self.region)

    def test_budget_flows_to_po_at_budgeted_price(self):
        from procurement.models import PurchaseOrder, PurchaseOrderItem
        self.client.force_login(self.proc)
        # Approved-budget list shows the sheet; tracker opens.
        self.assertEqual(self.client.get(reverse('procurement:approved_budgets')).status_code, 200)
        turl = reverse('procurement:bom_procurement_tracker', kwargs={'sheet_pk': self.sheet.pk})
        self.assertEqual(self.client.get(turl).status_code, 200)
        # Create a PO from the one supply line.
        r = self.client.post(turl, {'item_ids': [str(self.item.pk)]})
        self.assertEqual(r.status_code, 302)
        po = PurchaseOrder.objects.filter(project=self.project).latest('id')
        pi = po.items.get(source_bom_item=self.item)
        # rate = budgeted unit price; vendor carried over; PO vendor pre-filled.
        self.assertEqual(pi.rate_per_unit, self.item.budget_unit_price())
        self.assertEqual(pi.vendor_name, 'Acme')
        self.assertEqual(po.vendor_name, 'Acme')
        self.assertEqual(pi.make_model, 'Bosch X1')
        # Budgeted price at default equals the costing line price (margin 30%).
        self.item.refresh_from_db()
        self.assertEqual(self.item.budget_line_price(), self.item.base_total_price)

    def test_procurement_cannot_open_costing_sheet(self):
        self.client.force_login(self.proc)
        r = self.client.get(reverse('costing:detail', kwargs={'pk': self.sheet.pk}))
        self.assertNotEqual(r.status_code, 200)  # locked out (404/403/redirect)


class POTermOverrideTests(TestCase):
    """Terms are picked from the shared TermsTemplate library, but a PO can
    reword one for itself. The override is scoped to that PO — the template
    and every other PO using it must stay exactly as they were."""

    def setUp(self):
        from costing.models import TermsTemplate
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('ovr', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.original = 'ORIGINAL LINE ONE\nORIGINAL LINE TWO'
        self.tmpl = TermsTemplate.objects.create(
            name='Payment', category='payment_terms', usage='procurement',
            content=self.original)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='OVR-1', vendor_name='V',
            po_issued_by='T', cost_center='projects', created_by=self.user)
        self.po.selected_terms.add(self.tmpl)

    def _save(self, text, po=None):
        from django.http import QueryDict
        from procurement.views import _save_po_term_overrides
        qd = QueryDict(mutable=True)
        qd[f'term_content_{self.tmpl.pk}'] = text
        _save_po_term_overrides(po or self.po, qd, self.user)

    def test_no_override_falls_back_to_template(self):
        entry = self.po.resolved_terms()[0]
        self.assertEqual(entry['content'], self.original)
        self.assertFalse(entry['is_overridden'])

    def test_edit_is_scoped_to_this_po(self):
        self._save('CUSTOM WORDING')
        entry = self.po.resolved_terms()[0]
        self.assertEqual(entry['content'], 'CUSTOM WORDING')
        self.assertTrue(entry['is_overridden'])

        # The shared template must not have been written to.
        self.tmpl.refresh_from_db()
        self.assertEqual(self.tmpl.content, self.original)

        # …and a different PO on the same template still sees the original.
        other = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='OVR-2', vendor_name='V2',
            po_issued_by='T', cost_center='projects', created_by=self.user)
        other.selected_terms.add(self.tmpl)
        self.assertEqual(other.resolved_terms()[0]['content'], self.original)

    def test_editing_back_to_template_text_drops_the_override(self):
        from procurement.models import POTermOverride
        self._save('CUSTOM WORDING')
        self.assertEqual(POTermOverride.objects.filter(purchase_order=self.po).count(), 1)
        # Trailing whitespace / CRLF differences must not count as an edit.
        self._save(self.original.replace('\n', '\r\n') + '   ')
        self.assertEqual(POTermOverride.objects.filter(purchase_order=self.po).count(), 0)

    def test_deselecting_a_term_drops_its_override(self):
        from procurement.models import POTermOverride
        self._save('CUSTOM WORDING')
        self.po.selected_terms.clear()
        self._save('CUSTOM WORDING')
        self.assertEqual(POTermOverride.objects.filter(purchase_order=self.po).count(), 0)

    def test_exports_use_the_edited_text(self):
        import io
        from django.utils import timezone
        from procurement.models import PurchaseOrderItem
        PurchaseOrderItem.objects.create(
            purchase_order=self.po, serial_number=1, description='D',
            quantity=1, uom='Nos', rate_per_unit=10, order=0)
        self._save('CUSTOM WORDING FOR EXPORT')
        for s in self.po.required_stages:      # release so Excel is unlocked
            setattr(self.po, f'{s}_approved_at', timezone.now())
            setattr(self.po, f'{s}_approved_by', self.user)
        self.po.save()

        import openpyxl
        r = self.client.get(reverse('procurement:po_export', kwargs={'pk': self.po.pk}))
        ws = openpyxl.load_workbook(io.BytesIO(r.content)).active
        cells = {str(c.value) for row in ws.iter_rows() for c in row if c.value}
        self.assertIn('CUSTOM WORDING FOR EXPORT', cells)
        self.assertNotIn('ORIGINAL LINE ONE', cells)

        from pypdf import PdfReader
        r = self.client.get(reverse('procurement:po_export_pdf', kwargs={'pk': self.po.pk}))
        text = ''.join(p.extract_text() for p in PdfReader(io.BytesIO(r.content)).pages)
        self.assertIn('CUSTOM WORDING FOR EXPORT', text)
        self.assertNotIn('ORIGINAL LINE ONE', text)


class POTermsPdfNumberingTests(TestCase):
    """Procurement PO PDF must auto-number the terms (users were typing serial
    numbers by hand because nothing sequenced them) and start the Terms &
    Conditions on their own fresh page."""

    def setUp(self):
        from costing.models import TermsTemplate
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('tcnum', password='pw', role=self.sa_role)
        self.client.force_login(self.user)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='TCNUM-1', vendor_name='V',
            po_issued_by='T', cost_center='projects', created_by=self.user)
        from procurement.models import PurchaseOrderItem
        PurchaseOrderItem.objects.create(
            purchase_order=self.po, serial_number=1, description='Widget',
            quantity=1, uom='Nos', rate_per_unit=10, order=0)
        # Two templates whose content includes the user's own numbering — the
        # PDF must reproduce it verbatim and add none of its own.
        self.t1 = TermsTemplate.objects.create(
            name='Payment', category='terms_and_conditions', usage='procurement',
            content='1. Payment 30 days\n2. Prices in SAR')
        self.t2 = TermsTemplate.objects.create(
            name='Warranty', category='terms_and_conditions', usage='procurement',
            content='(a) Warranty 24 months\nUnnumbered clause here')
        self.po.selected_terms.add(self.t1, self.t2)
        from django.utils import timezone
        for s in self.po.required_stages:
            setattr(self.po, f'{s}_approved_at', timezone.now())
            setattr(self.po, f'{s}_approved_by', self.user)
        self.po.save()

    def _pdf_pages(self):
        from pypdf import PdfReader
        r = self.client.get(reverse('procurement:po_export_pdf', kwargs={'pk': self.po.pk}))
        self.assertEqual(r.status_code, 200)
        reader = PdfReader(io.BytesIO(r.content))
        return [(p.extract_text() or '') for p in reader.pages]

    def test_terms_print_content_verbatim_with_no_auto_numbering(self):
        pages = self._pdf_pages()
        text = '\n'.join(pages)
        # The user's own text/numbering is preserved exactly...
        self.assertIn('1. Payment 30 days', text)
        self.assertIn('2. Prices in SAR', text)
        self.assertIn('(a) Warranty 24 months', text)
        self.assertIn('Unnumbered clause here', text)
        # ...and the app adds no numbering of its own: the unnumbered clause
        # is not given a "3." (or any) prefix.
        self.assertNotIn('3. Unnumbered clause here', text)
        self.assertNotIn('4. Unnumbered clause here', text)

    def test_terms_start_on_their_own_page(self):
        pages = self._pdf_pages()
        self.assertGreaterEqual(len(pages), 2)
        # Items are on page 1; the T&C heading must not share it.
        self.assertNotIn('TERMS AND CONDITIONS', pages[0])
        self.assertTrue(any('TERMS AND CONDITIONS' in p for p in pages[1:]))


class POTermsOrderingTests(TestCase):
    """Terms must print in the order the user selected them, not alphabetically."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('tcord', password='pw', role=self.sa_role)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='TCORD-1', vendor_name='V',
            po_issued_by='T', cost_center='projects', created_by=self.user)

    def test_resolved_terms_follow_selection_order_not_alphabetical(self):
        from costing.models import TermsTemplate
        alpha = TermsTemplate.objects.create(
            name='Alpha', category='terms_and_conditions', usage='procurement', content='a')
        zebra = TermsTemplate.objects.create(
            name='Zebra', category='terms_and_conditions', usage='procurement', content='z')
        self.po.selected_terms.add(alpha, zebra)
        # terms_order says Zebra-then-Alpha, which is neither alphabetical nor
        # pk order — resolved_terms must honour it.
        self.po.terms_order = f'{zebra.pk},{alpha.pk}'
        self.po.save(update_fields=['terms_order'])
        names = [e['template'].name for e in self.po.resolved_terms()]
        self.assertEqual(names, ['Zebra', 'Alpha'])

    def test_selected_term_missing_from_order_is_not_dropped(self):
        from costing.models import TermsTemplate
        a = TermsTemplate.objects.create(
            name='A', category='terms_and_conditions', usage='procurement', content='a')
        b = TermsTemplate.objects.create(
            name='B', category='terms_and_conditions', usage='procurement', content='b')
        self.po.selected_terms.add(a, b)
        self.po.terms_order = str(b.pk)      # only B listed; A missing
        self.po.save(update_fields=['terms_order'])
        names = [e['template'].name for e in self.po.resolved_terms()]
        self.assertEqual(names, ['B', 'A'])  # listed first, the rest appended

    def test_ordered_selected_term_ids_respects_hint_and_ignores_unchecked(self):
        from django.http import QueryDict
        from procurement.views import _ordered_selected_term_ids
        post = QueryDict(mutable=True)
        post.setlist('selected_terms', ['5', '3', '9'])   # checked boxes (DOM order)
        post['selected_terms_ordered'] = '9,5,3'          # user's click order
        self.assertEqual(_ordered_selected_term_ids(post), [9, 5, 3])

        # A stale id in the hint that is no longer checked is dropped.
        post['selected_terms_ordered'] = '9,99,5,3'
        self.assertEqual(_ordered_selected_term_ids(post), [9, 5, 3])

        # A checked box missing from the hint (JS off) still appears, at the end.
        post['selected_terms_ordered'] = '9'
        self.assertEqual(_ordered_selected_term_ids(post), [9, 5, 3])


def _png_bytes():
    buf = io.BytesIO()
    Image.new('RGBA', (8, 8), (0, 0, 0, 0)).save(buf, 'PNG')
    return buf.getvalue()


class POEditSignatureTests(TestCase):
    """Editing an already-signed stage's signature: allowed for the original
    signer or a super admin; the recorded approver + timestamp never change."""

    def setUp(self):
        from django.utils import timezone
        from django.core.files.base import ContentFile
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.pm_role, _ = Role.objects.get_or_create(name=Role.PROCUREMENT_MGR)
        self.super_admin = User.objects.create_user('sa', password='pw', role=self.sa_role)
        self.signer = User.objects.create_user('shaker', password='pw', role=self.pm_role)
        self.other_pm = User.objects.create_user('other_pm', password='pw', role=self.pm_role)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-SIG-1',
            vendor_name='ACME', po_issued_by='Tester', created_by=self.super_admin)
        # Sign SCM directly (bypassing the endpoint) so we start from a signed stage.
        self.signed_at = timezone.now()
        self.po.scm_approved_at = self.signed_at
        self.po.scm_approved_by = self.signer
        self.po.scm_signature.save('orig.png', ContentFile(_png_bytes()), save=False)
        self.po.save()

    def _url(self, stage='scm'):
        return reverse('procurement:po_edit_signature',
                       kwargs={'pk': self.po.pk, 'stage': stage})

    def _post_new_sig(self):
        return self.client.post(self._url('scm'), {'signature_data': _png_data_url()})

    def test_original_signer_can_edit_metadata_unchanged(self):
        self.client.force_login(self.signer)
        r = self._post_new_sig()
        self.assertEqual(r.status_code, 200)
        self.assertTrue(r.json()['ok'])
        self.po.refresh_from_db()
        # Approver + timestamp must be untouched — only the image was swapped.
        self.assertEqual(self.po.scm_approved_by_id, self.signer.id)
        self.assertEqual(self.po.scm_approved_at, self.signed_at)
        self.assertTrue(self.po.scm_signature)

    def test_super_admin_can_edit(self):
        self.client.force_login(self.super_admin)
        r = self._post_new_sig()
        self.assertEqual(r.status_code, 200)
        self.po.refresh_from_db()
        self.assertEqual(self.po.scm_approved_by_id, self.signer.id)  # still the signer

    def test_non_signer_non_superadmin_forbidden(self):
        # A different procurement manager can SEE the PO but isn't the signer.
        self.client.force_login(self.other_pm)
        r = self._post_new_sig()
        self.assertEqual(r.status_code, 403)
        self.po.refresh_from_db()
        self.assertEqual(self.po.scm_approved_by_id, self.signer.id)

    def test_edit_unsigned_stage_is_400(self):
        self.client.force_login(self.super_admin)
        r = self.client.post(self._url('pm'), {'signature_data': _png_data_url()})
        self.assertEqual(r.status_code, 400)
        self.assertIn('not been signed', r.json()['error'].lower())

    def test_out_of_scope_user_gets_404(self):
        outsider = User.objects.create_user('outsider', password='pw')
        self.client.force_login(outsider)
        r = self._post_new_sig()
        self.assertEqual(r.status_code, 404)

    def test_missing_signature_is_400(self):
        self.client.force_login(self.signer)
        r = self.client.post(self._url('scm'), {})
        self.assertEqual(r.status_code, 400)
        self.assertIn('signature', r.json()['error'].lower())

    def test_detail_page_edit_button_gated_to_signer(self):
        detail = reverse('procurement:po_detail', kwargs={'pk': self.po.pk})
        # Assert on the button's edit URL, which only renders when the button
        # does (the class name alone also appears in the static JS selector).
        edit_url = self._url('scm')
        # The original signer sees the "Edit signature" control...
        self.client.force_login(self.signer)
        r = self.client.get(detail)
        self.assertEqual(r.status_code, 200)
        self.assertIn(edit_url, r.content.decode())
        # ...a different procurement manager (can view the PO) does not.
        self.client.force_login(self.other_pm)
        self.assertNotIn(edit_url, self.client.get(detail).content.decode())


class TermsSanitizeTests(TestCase):
    """Rich-text Terms & Conditions guards: user HTML is sanitised (no stored
    XSS), allowed formatting survives, and legacy plain text keeps its line
    breaks in both the web render and the PDF converter."""

    def test_render_terms_strips_scripts_and_handlers(self):
        from procurement.templatetags.terms_extras import render_terms
        out = str(render_terms('<p>ok</p><script>alert(1)</script><img src=x onerror=alert(2)>'))
        self.assertNotIn('<script', out.lower())
        self.assertNotIn('onerror', out.lower())
        self.assertIn('ok', out)

    def test_render_terms_keeps_allowed_formatting(self):
        from procurement.templatetags.terms_extras import render_terms
        out = str(render_terms('<span style="color:#ff0000;font-size:14pt">Red</span>'))
        self.assertIn('color', out)
        self.assertIn('14pt', out)

    def test_render_terms_linebreaks_legacy_plain_text(self):
        from procurement.templatetags.terms_extras import render_terms
        out = str(render_terms('Line one\nLine two'))
        self.assertIn('<br', out)

    def test_terms_preview_strips_all_tags(self):
        from procurement.templatetags.terms_extras import terms_preview
        out = str(terms_preview('<b>Bold</b> <script>x</script> text', 15))
        self.assertNotIn('<', out)
        self.assertIn('Bold', out)

    def test_pdf_converter_splits_legacy_plain_text_lines(self):
        from procurement.views import _tinymce_html_to_reportlab_lines
        lines = _tinymce_html_to_reportlab_lines('1. Payment 30 days\n2. Delivery FOB')
        self.assertEqual(len(lines), 2)  # not one run-on line

    def test_pdf_converter_escapes_amp_and_drops_scripts(self):
        from procurement.views import _tinymce_html_to_reportlab_lines
        lines = _tinymce_html_to_reportlab_lines('<p>Safe &amp; sound <script>bad</script></p>')
        joined = ' '.join(m for m, _ in lines)
        self.assertNotIn('<script', joined.lower())
        self.assertIn('&amp;', joined)


class POClientAcknowledgedLockTests(TestCase):
    """Once the client acknowledges a PO, nothing about it may change.

    The lock has many doors — the PO form, line items, terms, signatures, the
    summary editor, both imports and delete are all separate endpoints. Each
    test below asserts the DATA is unchanged, not merely that the response was
    a refusal: a view that writes and then returns 423 passes a status-code
    test and still loses the guarantee."""

    def setUp(self):
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.super_admin = User.objects.create_user('lock_super', password='x')
        self.super_admin.role = role
        self.super_admin.save()

        proc, _ = Role.objects.get_or_create(name=Role.PROCUREMENT_MGR)
        self.procurement = User.objects.create_user('lock_proc', password='x')
        self.procurement.role = proc
        self.procurement.save()

        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-LOCK-1',
            vendor_name='ACME', po_issued_by='Tester',
            created_by=self.procurement, status='issued')
        self.item = self.po.items.create(
            description='Widget', quantity=Decimal('2'),
            rate_per_unit=Decimal('100'))

    def _lock(self):
        self.po.record_status_change(
            to_status='client_acknowledged', changed_by=self.procurement)
        self.po.refresh_from_db()

    # ── the status itself ───────────────────────────────────────────────────

    def test_only_the_new_status_locks(self):
        for status in ('draft', 'issued', 'completed', 'cancelled'):
            self.po.status = status
            with self.subTest(status=status):
                self.assertFalse(self.po.is_locked)
        self.po.status = 'client_acknowledged'
        self.assertTrue(self.po.is_locked)

    def test_acknowledging_stamps_who_and_when(self):
        self._lock()
        self.assertIsNotNone(self.po.client_acknowledged_at)
        self.assertEqual(self.po.client_acknowledged_by, self.procurement)

    def test_every_status_change_is_logged(self):
        self._lock()
        change = self.po.status_changes.first()
        self.assertEqual((change.from_status, change.to_status),
                         ('issued', 'client_acknowledged'))
        self.assertEqual(change.changed_by, self.procurement)

    def test_a_no_op_transition_writes_nothing(self):
        """A log full of rows saying nothing changed is a log nobody reads."""
        self.po.record_status_change(to_status='issued',
                                     changed_by=self.procurement)
        self.assertEqual(self.po.status_changes.count(), 0)

    # ── the doors ───────────────────────────────────────────────────────────

    def test_the_edit_form_refuses(self):
        self._lock()
        self.client.force_login(self.procurement)
        resp = self.client.post(
            reverse('procurement:po_update', args=[self.po.pk]),
            {'po_date': '2026-02-02', 'po_number': 'PO-LOCK-CHANGED',
             'vendor_name': 'OTHER', 'po_issued_by': 'Tester',
             'cost_center': 'projects', 'status': 'issued',
             'items-TOTAL_FORMS': '0', 'items-INITIAL_FORMS': '0'})
        self.po.refresh_from_db()
        self.assertEqual(self.po.po_number, 'PO-LOCK-1')
        self.assertEqual(self.po.vendor_name, 'ACME')
        self.assertEqual(resp.status_code, 302)

    def test_deleting_refuses(self):
        self._lock()
        self.client.force_login(self.super_admin)
        self.client.post(reverse('procurement:po_delete', args=[self.po.pk]))
        self.assertTrue(PurchaseOrder.objects.filter(pk=self.po.pk).exists())

    def test_editing_a_line_item_refuses(self):
        self._lock()
        self.client.force_login(self.procurement)
        resp = self.client.post(
            reverse('procurement:po_item_update_field', args=[self.item.pk]),
            {'field': 'delivery_status', 'value': 'Tampered'})
        self.item.refresh_from_db()
        self.assertEqual(self.item.delivery_status, '')
        self.assertEqual(resp.status_code, 423)

    def test_editing_a_po_header_field_inline_refuses(self):
        """The same endpoint writes PO-level fields as well as item fields, so
        the guard has to cover both branches."""
        self._lock()
        self.client.force_login(self.procurement)
        resp = self.client.post(
            reverse('procurement:po_item_update_field', args=[self.item.pk]),
            {'field': 'lead_time', 'value': '99 weeks'})
        self.po.refresh_from_db()
        self.assertNotEqual(self.po.lead_time, '99 weeks')
        self.assertEqual(resp.status_code, 423)

    def test_toggling_a_term_refuses(self):
        self._lock()
        self.client.force_login(self.procurement)
        before = list(self.po.selected_terms.values_list('pk', flat=True))
        resp = self.client.post(
            reverse('procurement:po_toggle_term', args=[self.po.pk]),
            {'term_id': '1', 'enabled': '1'})
        self.assertEqual(
            list(self.po.selected_terms.values_list('pk', flat=True)), before)
        self.assertEqual(resp.status_code, 423)

    def test_signing_an_approval_stage_refuses(self):
        """A signature added after the client accepted the document would be a
        signature on something that was already agreed.

        Sends a VALID signature on purpose: an empty one is rejected by the
        signature validator whether or not the lock exists, so the test would
        pass against no lock at all.
        """
        self._lock()
        self.client.force_login(self.super_admin)
        resp = self.client.post(
            reverse('procurement:po_approve_stage', args=[self.po.pk, 'scm']),
            {'signature_data': _png_data_url()})
        self.po.refresh_from_db()
        self.assertIsNone(self.po.scm_approved_at)
        self.assertEqual(resp.status_code, 423)

    def test_replacing_a_signature_image_refuses(self):
        """po_edit_signature is a separate endpoint from po_approve_stage and
        needs its own guard."""
        from django.utils import timezone
        PurchaseOrder.objects.filter(pk=self.po.pk).update(
            scm_approved_at=timezone.now(), scm_approved_by=self.super_admin)
        self._lock()
        self.client.force_login(self.super_admin)
        resp = self.client.post(
            reverse('procurement:po_edit_signature', args=[self.po.pk, 'scm']),
            {'signature_data': _png_data_url()})
        self.assertEqual(resp.status_code, 423)

    def test_an_unlocked_po_still_accepts_edits(self):
        """The guard must refuse the locked case only — a lock that blocks
        everything is indistinguishable from a broken page."""
        self.client.force_login(self.procurement)
        resp = self.client.post(
            reverse('procurement:po_item_update_field', args=[self.item.pk]),
            {'field': 'delivery_status', 'value': 'Shipped'})
        self.assertEqual(resp.status_code, 200)
        self.item.refresh_from_db()
        self.assertEqual(self.item.delivery_status, 'Shipped')

    # ── the way out ─────────────────────────────────────────────────────────

    def test_a_super_admin_can_release_it_with_a_reason(self):
        self._lock()
        self.client.force_login(self.super_admin)
        self.client.post(
            reverse('procurement:po_release_lock', args=[self.po.pk]),
            {'reason': 'Wrong delivery address, client asked us to reissue'})
        self.po.refresh_from_db()
        self.assertFalse(self.po.is_locked)
        self.assertEqual(self.po.status, 'issued')

    def test_releasing_records_who_and_why(self):
        self._lock()
        self.client.force_login(self.super_admin)
        self.client.post(
            reverse('procurement:po_release_lock', args=[self.po.pk]),
            {'reason': 'Wrong delivery address'})
        change = self.po.status_changes.first()
        self.assertEqual(change.to_status, 'issued')
        self.assertEqual(change.changed_by, self.super_admin)
        self.assertIn('delivery address', change.reason)

    def test_releasing_without_a_reason_is_refused(self):
        """This is the only door out of the lock; an unexplained release is
        worth less than no record at all."""
        self._lock()
        self.client.force_login(self.super_admin)
        self.client.post(
            reverse('procurement:po_release_lock', args=[self.po.pk]),
            {'reason': '   '})
        self.po.refresh_from_db()
        self.assertTrue(self.po.is_locked)

    def test_procurement_cannot_release_it(self):
        self._lock()
        self.client.force_login(self.procurement)
        self.client.post(
            reverse('procurement:po_release_lock', args=[self.po.pk]),
            {'reason': 'Let me back in'})
        self.po.refresh_from_db()
        self.assertTrue(self.po.is_locked)

    def test_releasing_clears_the_acknowledgement_stamps(self):
        """They describe the CURRENT acknowledgement. Leaving them behind would
        show a PO as accepted on a date it is no longer accepted."""
        self._lock()
        self.client.force_login(self.super_admin)
        self.client.post(
            reverse('procurement:po_release_lock', args=[self.po.pk]),
            {'reason': 'Reissuing'})
        self.po.refresh_from_db()
        self.assertIsNone(self.po.client_acknowledged_at)
        self.assertIsNone(self.po.client_acknowledged_by)
        # The history still knows it happened.
        self.assertEqual(self.po.status_changes.count(), 2)


class POWorkflowStatusTests(TestCase):
    """What a PO is waiting on, derived rather than stored."""

    def setUp(self):
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('wf_user', password='x')
        self.user.role = role
        self.user.save()
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-WF-1', vendor_name='ACME',
            po_issued_by='Tester', created_by=self.user)

    def _sign(self, *stages):
        from django.utils import timezone
        PurchaseOrder.objects.filter(pk=self.po.pk).update(
            **{f'{s}_approved_at': timezone.now() for s in stages})
        self.po.refresh_from_db()

    def test_an_untouched_po_names_its_first_stage(self):
        w = self.po.workflow_status
        self.assertIn('SCM', w['label'])
        self.assertEqual(w['stage_key'], 'scm')

    def test_signing_moves_the_label_to_the_next_stage(self):
        self._sign('scm')
        w = self.po.workflow_status
        self.assertEqual(w['label'], 'Awaiting PM approval')
        self.assertEqual(w['tone'], 'warning')

    def test_a_fully_signed_po_reads_as_released(self):
        self._sign('scm', 'pm', 'coo')
        w = self.po.workflow_status
        self.assertEqual(w['tone'], 'success')
        self.assertIsNone(w['stage_key'])

    def test_a_locked_po_says_locked_rather_than_a_stage(self):
        self.po.record_status_change(to_status='client_acknowledged',
                                     changed_by=self.user)
        self.assertEqual(self.po.workflow_status['tone'], 'locked')

    def test_a_cancelled_po_is_waiting_on_nobody(self):
        self.po.record_status_change(to_status='cancelled', changed_by=self.user)
        w = self.po.workflow_status
        self.assertEqual(w['label'], 'Cancelled')
        self.assertIsNone(w['stage_key'])

    def test_the_label_matches_the_stage_the_button_accepts(self):
        """Derived from current_stage precisely so the list and the Approve
        endpoint cannot disagree about the same PO."""
        self._sign('scm')
        self.assertEqual(self.po.workflow_status['stage_key'],
                         self.po.current_stage['key'])

    def test_the_list_shows_the_approval_column(self):
        self.client.force_login(self.user)
        body = self.client.get(reverse('procurement:po_list')).content.decode()
        self.assertIn('Approval', body)
        self.assertIn('SCM approval', body)


class ApprovedBudgetsReferenceTests(TestCase):
    """Procurement can see the LNA reference without opening the budget."""

    def setUp(self):
        from projects.models import Region, ProjectStatus, Project
        from costing.models import CostingSheet
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('bud_user', password='x')
        self.user.role = role
        self.user.save()
        region = Region.objects.create(name='KSA', code='LNA', currency='SAR')
        status = ProjectStatus.objects.create(name='Open', category='active')
        self.project = Project.objects.create(
            project_name='Ghazlan Substation', proposal_reference='LNA-2026-0417',
            status=status, region=region, estimated_value=Decimal('1'),
            actual_sales=Decimal('0'), year='2026', po_award_quarter='Q2')
        self.sheet = CostingSheet.objects.create(
            title='Budget', project=self.project,
            workflow_stage='finance_approved')

    def test_the_reference_is_on_the_page(self):
        self.client.force_login(self.user)
        body = self.client.get(
            reverse('procurement:approved_budgets')).content.decode()
        self.assertIn('LNA Reference', body)
        self.assertIn('LNA-2026-0417', body)


def _items_workbook(rows, headers=None, preamble=0):
    """A spreadsheet shaped like something a vendor would actually send."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(preamble):
        ws.append(['Quotation', None, None])      # title block above the table
    ws.append(headers or ['S.No', 'Description', 'Qty', 'UOM', 'Rate/Unit (SAR)', 'Remarks'])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class ExcelItemParserTests(TestCase):
    """Reading line items out of somebody else's spreadsheet.

    Column positions cannot be relied on in a file we did not generate, so the
    parser matches headers by name. These tests are mostly about the shapes
    real quotations arrive in."""

    def _parse(self, *args, **kwargs):
        from procurement.excel_import import parse_items
        return parse_items(_items_workbook(*args, **kwargs))

    def test_it_reads_a_plain_table(self):
        rows, skipped = self._parse([[1, 'Camera', 4, 'Nos', 250, 'Outdoor']])
        self.assertEqual(skipped, [])
        self.assertEqual(rows[0]['description'], 'Camera')
        self.assertEqual(rows[0]['quantity'], Decimal('4'))
        self.assertEqual(rows[0]['rate_per_unit'], Decimal('250'))
        self.assertEqual(rows[0]['uom'], 'Nos')

    def test_column_order_does_not_matter(self):
        """The point of matching on names: nobody sends columns in our order."""
        rows, _ = self._parse(
            [['Nos', 250, 'Camera', 4]],
            headers=['UOM', 'Unit Price', 'Item Description', 'Quantity'])
        self.assertEqual(rows[0]['description'], 'Camera')
        self.assertEqual(rows[0]['quantity'], Decimal('4'))
        self.assertEqual(rows[0]['rate_per_unit'], Decimal('250'))

    def test_a_title_block_above_the_table_is_skipped(self):
        """Quotations open with a logo and an address, not a header row."""
        rows, _ = self._parse([[1, 'Camera', 4, 'Nos', 250, '']], preamble=5)
        self.assertEqual(len(rows), 1)

    def test_units_in_the_header_are_ignored(self):
        rows, _ = self._parse(
            [['Camera', 2, 300]],
            headers=['Description', 'Qty', 'Rate / Unit (USD)'])
        self.assertEqual(rows[0]['rate_per_unit'], Decimal('300'))

    def test_extra_columns_are_left_alone(self):
        rows, _ = self._parse(
            [['Camera', 1, 100, 'irrelevant']],
            headers=['Description', 'Qty', 'Rate', 'HS Code'])
        self.assertEqual(len(rows), 1)

    def test_totals_rows_are_skipped_and_reported(self):
        """A totals line imported as an item silently doubles the PO value."""
        rows, skipped = self._parse([
            [1, 'Camera', 4, 'Nos', 250, ''],
            [None, 'Total', None, None, 1000, ''],
        ])
        self.assertEqual(len(rows), 1)
        self.assertEqual(len(skipped), 1)
        self.assertIn('totals', skipped[0][1])

    def test_a_row_with_no_quantity_is_reported_not_guessed(self):
        rows, skipped = self._parse([[1, 'Camera', 0, 'Nos', 250, '']])
        self.assertEqual(rows, [])
        self.assertEqual(len(skipped), 1)

    def test_blank_rows_are_not_reported_as_problems(self):
        """Every quotation has spacer rows. Reporting them as skips would bury
        the real problems."""
        rows, skipped = self._parse([
            [1, 'Camera', 4, 'Nos', 250, ''],
            [None, None, None, None, None, None],
        ])
        self.assertEqual(len(rows), 1)
        self.assertEqual(skipped, [])

    def test_numbers_written_as_text_still_read(self):
        rows, _ = self._parse([[1, 'Camera', '4 nos', 'Nos', 'SAR 1,250.50', '']])
        self.assertEqual(rows[0]['quantity'], Decimal('4'))
        self.assertEqual(rows[0]['rate_per_unit'], Decimal('1250.50'))

    def test_a_sheet_with_no_description_column_says_so(self):
        from procurement.excel_import import ExcelImportError
        with self.assertRaises(ExcelImportError) as caught:
            self._parse([[1, 2]], headers=['Foo', 'Bar'])
        self.assertIn('Description', str(caught.exception))

    def test_a_file_that_is_not_a_spreadsheet_fails_readably(self):
        """A traceback tells the user nothing they can act on."""
        from procurement.excel_import import ExcelImportError, parse_items
        with self.assertRaises(ExcelImportError):
            parse_items(io.BytesIO(b'this is not a workbook'))


class POImportItemsTests(TestCase):
    """Adding items to a PO that already exists."""

    def setUp(self):
        role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.user = User.objects.create_user('imp_user', password='x')
        self.user.role = role
        self.user.save()
        self.client.force_login(self.user)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-IMP-1', vendor_name='ACME',
            po_issued_by='Tester', created_by=self.user, status='draft')
        self.existing = self.po.items.create(
            description='Existing line', quantity=Decimal('1'),
            rate_per_unit=Decimal('50'), serial_number=1)

    def _post(self, *workbooks, follow=False):
        return self.client.post(
            reverse('procurement:po_import_items', args=[self.po.pk]),
            {'item_files': list(workbooks)}, follow=follow)

    def test_items_are_added_not_replaced(self):
        """There is no undo. An import that wiped hand-entered lines would be
        unrecoverable; a duplicated line is obvious and deletable."""
        self._post(_items_workbook([[1, 'Camera', 2, 'Nos', 100, '']]))
        descriptions = list(self.po.items.values_list('description', flat=True))
        self.assertIn('Existing line', descriptions)
        self.assertIn('Camera', descriptions)
        self.assertEqual(self.po.items.count(), 2)

    def test_serial_numbers_continue_from_what_is_there(self):
        self._post(_items_workbook([[1, 'Camera', 2, 'Nos', 100, '']]))
        new = self.po.items.get(description='Camera')
        self.assertEqual(new.serial_number, 2)

    def test_several_files_all_import(self):
        self._post(_items_workbook([[1, 'Camera', 1, 'Nos', 100, '']]),
                   _items_workbook([[1, 'Cable', 5, 'M', 10, '']]))
        self.assertEqual(self.po.items.count(), 3)

    def test_one_bad_file_does_not_lose_the_good_one(self):
        self._post(_items_workbook([[1, 'Camera', 1, 'Nos', 100, '']]),
                   io.BytesIO(b'not a workbook'))
        self.assertTrue(self.po.items.filter(description='Camera').exists())

    def test_the_totals_reflect_the_imported_lines(self):
        before = self.po.total_value
        self._post(_items_workbook([[1, 'Camera', 2, 'Nos', 100, '']]))
        self.po.refresh_from_db()
        self.assertGreater(self.po.total_value, before)

    def test_it_reports_what_was_skipped(self):
        """An import that reads 8 of 12 rows without saying so is worse than
        one that fails outright."""
        resp = self._post(_items_workbook([
            [1, 'Camera', 2, 'Nos', 100, ''],
            [None, 'Total', None, None, 200, ''],
        ]), follow=True)
        text = ' '.join(str(m) for m in resp.context['messages'])
        self.assertIn('skipped', text)

    def test_a_locked_po_refuses_the_import(self):
        self.po.record_status_change(to_status='client_acknowledged',
                                     changed_by=self.user)
        self._post(_items_workbook([[1, 'Camera', 2, 'Nos', 100, '']]))
        self.assertEqual(self.po.items.count(), 1)

    def test_the_panel_is_hidden_on_a_locked_po(self):
        self.po.record_status_change(to_status='client_acknowledged',
                                     changed_by=self.user)
        body = self.client.get(
            reverse('procurement:po_detail', args=[self.po.pk])).content.decode()
        self.assertNotIn('Import items', body)

    def test_the_edit_page_offers_the_import_too(self):
        """Same action, reachable from where people are already editing the
        items rather than only from the detail page."""
        body = self.client.get(
            reverse('procurement:po_update', args=[self.po.pk])).content.decode()
        self.assertIn('Import from Excel', body)
        self.assertIn(
            reverse('procurement:po_import_items', args=[self.po.pk]), body)

    def test_the_edit_page_import_form_is_not_nested(self):
        """A <form> inside a <form> is dropped by browsers, which would make
        the Import button post the PO form instead.

        Anchored on the PO form specifically. An earlier version of this test
        compared against the FIRST </form> on the page, which belongs to the
        navigation bar — so it passed with the modal nested and proved nothing.
        """
        body = self.client.get(
            reverse('procurement:po_update', args=[self.po.pk])).content.decode()
        po_form_opens = body.index('<form method="post" novalidate>')
        po_form_closes = body.index('</form>', po_form_opens)
        import_at = body.index(
            reverse('procurement:po_import_items', args=[self.po.pk]))
        self.assertFalse(
            po_form_opens < import_at < po_form_closes,
            'the import form is nested inside the PO form')

    def test_a_locked_po_offers_no_import_on_the_edit_page(self):
        self.po.record_status_change(to_status='client_acknowledged',
                                     changed_by=self.user)
        resp = self.client.get(
            reverse('procurement:po_update', args=[self.po.pk]), follow=True)
        self.assertNotIn('Import from Excel', resp.content.decode())

    def test_the_create_page_offers_no_import(self):
        """There is nothing to import into until the PO has been saved once."""
        body = self.client.get(
            reverse('procurement:po_create')).content.decode()
        self.assertNotIn('Import from Excel', body)

    def test_submitting_no_file_says_so(self):
        resp = self.client.post(
            reverse('procurement:po_import_items', args=[self.po.pk]),
            {}, follow=True)
        text = ' '.join(str(m) for m in resp.context['messages'])
        self.assertIn('at least one', text)


class POStageRecipientTests(TestCase):
    """Who gets told a PO is waiting on them.

    Separate from who may sign it: can_user_approve_stage() is the permission
    gate and is untouched. This decides the recipient list, which needs a real
    account because the stages name their signers as plain text."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.admin_role, _ = Role.objects.get_or_create(name=Role.ADMIN)
        self.proc_role, _ = Role.objects.get_or_create(name=Role.PROCUREMENT_MGR)
        self.super_admin = User.objects.create_user(
            'nr_super', password='x', email='super@example.com',
            role=self.sa_role)
        self.admin = User.objects.create_user(
            'nr_admin', password='x', email='admin@example.com',
            role=self.admin_role)
        self.proc = User.objects.create_user(
            'nr_proc', password='x', email='proc@example.com',
            role=self.proc_role)

    def _recipients(self, stage):
        from procurement.notifications import stage_recipients
        return stage_recipients(stage)

    def test_a_configured_approver_wins(self):
        from procurement.models import POStageApprover
        POStageApprover.objects.create(stage='scm', user=self.admin)
        self.assertEqual(self._recipients('scm'), [self.admin])

    def test_without_a_mapping_it_falls_back_to_the_role(self):
        """An unconfigured system should still tell somebody. A missed
        approval costs more than a redundant email."""
        self.assertIn(self.proc, self._recipients('scm'))
        self.assertIn(self.admin, self._recipients('pm'))
        self.assertIn(self.super_admin, self._recipients('ceo'))

    def test_the_mapping_replaces_the_role_holders_rather_than_adding_to_them(self):
        """The point of configuring one is to stop emailing everybody."""
        from procurement.models import POStageApprover
        POStageApprover.objects.create(stage='pm', user=self.admin)
        second_admin = User.objects.create_user(
            'nr_admin2', password='x', role=self.admin_role)
        self.assertNotIn(second_admin, self._recipients('pm'))

    def test_an_inactive_mapped_user_is_not_emailed(self):
        from procurement.models import POStageApprover
        leaver = User.objects.create_user('nr_gone', password='x',
                                          role=self.admin_role)
        POStageApprover.objects.create(stage='coo', user=leaver)
        leaver.is_active = False
        leaver.save()
        self.assertNotIn(leaver, self._recipients('coo'))

    def test_a_stage_nobody_can_fill_resolves_to_nobody_not_an_error(self):
        """A PO must still be creatable on a system where the roles have not
        been set up."""
        User.objects.all().update(is_active=False)
        self.assertEqual(self._recipients('scm'), [])


class POApprovalNotificationTests(TestCase):
    """Telling the approver, and not telling anyone who is not waiting."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.admin_role, _ = Role.objects.get_or_create(name=Role.ADMIN)
        self.proc_role, _ = Role.objects.get_or_create(name=Role.PROCUREMENT_MGR)
        self.raiser = User.objects.create_user(
            'an_raiser', password='x', email='raiser@example.com',
            role=self.sa_role)
        self.scm = User.objects.create_user(
            'an_scm', password='x', email='scm@example.com', role=self.proc_role)
        self.pm = User.objects.create_user(
            'an_pm', password='x', email='pm@example.com', role=self.admin_role)
        from procurement.models import POStageApprover
        POStageApprover.objects.create(stage='scm', user=self.scm)
        POStageApprover.objects.create(stage='pm', user=self.pm)
        self.po = PurchaseOrder.objects.create(
            po_date=date(2026, 1, 1), po_number='PO-NOTIFY-1',
            vendor_name='ACME', po_issued_by='Tester', created_by=self.raiser)

    def _notify(self, **kwargs):
        from procurement.notifications import notify_stage_approver
        return notify_stage_approver(self.po, **kwargs)

    def _sign(self, *stages):
        from django.utils import timezone
        PurchaseOrder.objects.filter(pk=self.po.pk).update(
            **{f'{s}_approved_at': timezone.now() for s in stages})
        self.po.refresh_from_db()

    def test_the_current_stages_approver_is_told(self):
        notes = self._notify()
        self.assertEqual([n.recipient for n in notes], [self.scm])

    def test_signing_moves_the_notification_to_the_next_approver(self):
        """Not to the person who just signed — they know."""
        self._sign('scm')
        notes = self._notify()
        self.assertEqual([n.recipient for n in notes], [self.pm])

    def test_a_released_po_tells_nobody(self):
        self._sign('scm', 'pm', 'coo')
        self.assertEqual(self._notify(), [])

    def test_a_cancelled_po_tells_nobody(self):
        self.po.record_status_change(to_status='cancelled', changed_by=self.raiser)
        self.assertEqual(self._notify(), [])

    def test_a_locked_po_tells_nobody(self):
        """Nothing can be signed on it, so an email asking for a signature
        would send somebody to a page that refuses them."""
        self.po.record_status_change(to_status='client_acknowledged',
                                     changed_by=self.raiser)
        self.assertEqual(self._notify(), [])

    def test_the_signer_is_not_told_about_their_own_signature(self):
        notes = self._notify(actor=self.scm)
        self.assertEqual(notes, [])

    def test_the_notification_links_to_the_po(self):
        note = self._notify()[0]
        self.assertIn(
            reverse('procurement:po_detail', args=[self.po.pk]), note.target_url)

    def test_an_absolute_link_is_used_when_the_host_is_known(self):
        """A relative link in an email goes nowhere."""
        note = self._notify(base_url='https://erp.example.com/')[0]
        self.assertTrue(note.target_url.startswith('https://erp.example.com/'))

    def test_the_message_says_what_and_who_it_is_waiting_on(self):
        note = self._notify()[0]
        self.assertIn('PO-NOTIFY-1', note.verb)
        self.assertIn('ACME', note.description)
        self.assertIn('SCM', note.description)

    # ── wiring ──────────────────────────────────────────────────────────────

    def test_approving_a_stage_notifies_the_next_approver(self):
        """captureOnCommitCallbacks because the notification is queued with
        transaction.on_commit — which is the point. TestCase wraps each test in
        a transaction it rolls back, so those callbacks never fire on their own
        and asserting without this would test nothing."""
        from notifications.models import Notification
        self.client.force_login(self.raiser)
        with self.captureOnCommitCallbacks(execute=True):
            self.client.post(
                reverse('procurement:po_approve_stage', args=[self.po.pk, 'scm']),
                {'signature_data': _png_data_url()})
        self.assertTrue(
            Notification.objects.filter(recipient=self.pm,
                                        verb__contains='PO-NOTIFY-1').exists())
        # And not to the person who just signed it.
        self.assertFalse(
            Notification.objects.filter(recipient=self.raiser,
                                        verb__contains='PO-NOTIFY-1').exists())

    def test_a_failed_approval_sends_nothing(self):
        """Queued on commit, so an email never describes a signature that
        rolled back."""
        from notifications.models import Notification
        self.client.force_login(self.raiser)
        # Out of sequence: 'pm' before 'scm' is refused.
        with self.captureOnCommitCallbacks(execute=True):
            self.client.post(
                reverse('procurement:po_approve_stage', args=[self.po.pk, 'pm']),
                {'signature_data': _png_data_url()})
        self.assertFalse(
            Notification.objects.filter(verb__contains='PO-NOTIFY-1').exists())


class POStageApproverScreenTests(TestCase):
    """The routing screen."""

    def setUp(self):
        self.sa_role, _ = Role.objects.get_or_create(name=Role.SUPER_ADMIN)
        self.proc_role, _ = Role.objects.get_or_create(name=Role.PROCUREMENT_MGR)
        self.admin = User.objects.create_user('sa_admin', password='x',
                                              role=self.sa_role)
        self.proc = User.objects.create_user('sa_proc', password='x',
                                             role=self.proc_role)

    def _url(self):
        return reverse('procurement:po_stage_approvers')

    def test_every_stage_is_listed_even_when_unset(self):
        """An unset stage has to be visible as a gap, not simply absent."""
        self.client.force_login(self.admin)
        body = self.client.get(self._url()).content.decode()
        for _key, label, _signer in PurchaseOrder.APPROVAL_STAGES:
            self.assertIn(label, body)
        self.assertIn('Everyone with the role', body)

    def test_an_admin_can_set_an_approver(self):
        from procurement.models import POStageApprover
        self.client.force_login(self.admin)
        self.client.post(self._url(), {'stage': 'scm', 'user': self.proc.pk})
        row = POStageApprover.objects.get(stage='scm')
        self.assertEqual(row.user, self.proc)
        self.assertEqual(row.updated_by, self.admin)

    def test_setting_a_stage_twice_replaces_rather_than_duplicates(self):
        from procurement.models import POStageApprover
        other = User.objects.create_user('sa_other', password='x',
                                         role=self.proc_role)
        self.client.force_login(self.admin)
        self.client.post(self._url(), {'stage': 'scm', 'user': self.proc.pk})
        self.client.post(self._url(), {'stage': 'scm', 'user': other.pk})
        self.assertEqual(POStageApprover.objects.filter(stage='scm').count(), 1)
        self.assertEqual(POStageApprover.objects.get(stage='scm').user, other)

    def test_procurement_cannot_change_the_routing(self):
        """Naming an approver decides who is asked to sign company purchase
        orders — not a self-service setting."""
        from procurement.models import POStageApprover
        self.client.force_login(self.proc)
        self.client.post(self._url(), {'stage': 'scm', 'user': self.proc.pk})
        self.assertFalse(POStageApprover.objects.exists())
