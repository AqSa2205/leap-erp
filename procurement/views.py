from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q
from decimal import Decimal, InvalidOperation

from .models import (
    PurchaseOrder, PurchaseOrderItem,
    POSummaryEntry, QuotationImport,
    DeliveryNote, DeliveryNoteItem,
    InventoryReport, InventoryItem,
    FRCReport, FRCEntry, FRCInventory,
)
from .forms import (
    PurchaseOrderForm, PurchaseOrderItemForm, POItemFormSet, POFilterForm,
    DeliveryNoteForm, DNItemFormSet,
    InventoryReportForm, InventoryItemFormSet,
    FRCReportForm, FRCEntryFormSet, FRCInventoryForm,
)
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, F
from django.utils import timezone
from django.utils.text import slugify
import openpyxl
from datetime import datetime, timedelta
from accounts.permissions import require_capability, CapabilityRequiredMixin


_NUM_UNITS = ['Zero', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine']
_NUM_TEENS = ['Ten', 'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen', 'Seventeen', 'Eighteen', 'Nineteen']
_NUM_TENS = ['', '', 'Twenty', 'Thirty', 'Forty', 'Fifty', 'Sixty', 'Seventy', 'Eighty', 'Ninety']
_NUM_SCALES = ['', 'Thousand', 'Million', 'Billion', 'Trillion']


def _amount_in_words(value, currency='SAR'):
    """Spell out a numeric amount for invoice/PO use.

    Example: Decimal('25420.50') with currency='SAR' →
    'Twenty Five Thousand Four Hundred Twenty SAR and Fifty Halalas Only'.

    Splits into integer SAR and 2-digit fractional Halalas, words each part
    in English, and appends ' Only' as is conventional on a payment doc.
    Returns an empty string for None / unparseable input.
    """
    if value is None:
        return ''
    try:
        d = Decimal(str(value)).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError, TypeError):
        return ''

    # Operate on the absolute value so the sub-helpers never see a negative.
    sign = ''
    if d < 0:
        sign = 'Negative '
        d = -d

    integer_part = int(d)
    halala_part = int(((d - integer_part) * 100).to_integral_value())

    def _under_hundred(n):
        if n < 10:
            return _NUM_UNITS[n]
        if n < 20:
            return _NUM_TEENS[n - 10]
        t, u = divmod(n, 10)
        return _NUM_TENS[t] if u == 0 else f'{_NUM_TENS[t]} {_NUM_UNITS[u]}'

    def _under_thousand(n):
        if n < 100:
            return _under_hundred(n)
        h, r = divmod(n, 100)
        return f'{_NUM_UNITS[h]} Hundred' if r == 0 else f'{_NUM_UNITS[h]} Hundred {_under_hundred(r)}'

    def _spell(n):
        if n == 0:
            return 'Zero'
        parts, i = [], 0
        while n > 0:
            chunk = n % 1000
            if chunk:
                scale = _NUM_SCALES[i] if i < len(_NUM_SCALES) else ''
                parts.append(f'{_under_thousand(chunk)} {scale}'.strip())
            n //= 1000
            i += 1
        return ' '.join(reversed(parts))

    main_words = _spell(integer_part)
    base = f'{sign}{main_words} {currency}'
    if halala_part:
        # Fractional unit name depends on the currency (Halalas / Cents / Fils).
        singular, plural = PurchaseOrder.CURRENCY_FRACTIONS.get(
            currency, ('Halala', 'Halalas'))
        suffix = singular if halala_part == 1 else plural
        return f'{base} and {_under_hundred(halala_part)} {suffix} Only'
    return f'{base} Only'


def _safe_filename(name, prefix='', suffix='', extension=''):
    """Build a safe filename for Content-Disposition headers.

    Strips quotes, newlines, slashes, and any other characters that could
    inject header values or cause path traversal. Falls back to 'export'
    if the result would be empty.
    """
    safe = slugify(str(name or ''))[:80] or 'export'
    parts = [p for p in (prefix, safe, suffix) if p]
    base = '_'.join(parts)
    if extension and not extension.startswith('.'):
        extension = f'.{extension}'
    return f'{base}{extension}'


def _make_numbered_canvas(draft=False):
    """Return a two-pass Canvas subclass that draws "Page X of Y" at the
    bottom of every page. Used by procurement PDF exports for consistent
    pagination across PO / DN / Inventory / Summary outputs.

    When ``draft=True`` a large diagonal "DRAFT" watermark is layered on
    every page so an unapproved export cannot be mistaken for a final,
    signed document.
    """
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.pdfgen.canvas import Canvas

    class NumberedCanvas(Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved_pages = []

        def showPage(self):
            self._saved_pages.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            page_count = len(self._saved_pages)
            for state in self._saved_pages:
                self.__dict__.update(state)
                try:
                    page_w, page_h = self._pagesize
                    if draft:
                        self.saveState()
                        self.setFillColor(colors.HexColor('#C41E3A'))
                        try:
                            self.setFillAlpha(0.10)
                        except Exception:
                            pass
                        self.setFont('Helvetica-Bold', 120)
                        self.translate(page_w / 2, page_h / 2)
                        self.rotate(45)
                        self.drawCentredString(0, 0, 'DRAFT')
                        self.restoreState()
                    self.setFont('Helvetica', 8)
                    self.setFillColor(colors.HexColor('#6c757d'))
                    self.drawCentredString(
                        page_w / 2, 8 * mm,
                        f'Page {self._pageNumber} of {page_count}',
                    )
                except Exception:
                    pass
                super().showPage()
            super().save()

    return NumberedCanvas


# ═══════════════════════════════════════════════════════════════
# PROCUREMENT DASHBOARD
# ═══════════════════════════════════════════════════════════════

@login_required
@require_capability('procurement.access')
def procurement_dashboard(request):
    """Procurement dashboard with KPIs, recent activity, and stats."""
    user = request.user

    # Procurement roles get full access
    has_full_access = user.is_super_admin_user or user.is_procurement_user

    # PO stats
    po_qs = PurchaseOrder.objects.all()
    if not has_full_access:
        if user.is_admin_user or user.is_manager_user:
            po_qs = po_qs.filter(Q(created_by=user) | Q(project__region=user.region))
        else:
            po_qs = po_qs.filter(created_by=user)

    po_total = po_qs.count()
    po_by_status = {}
    for s in ['draft', 'issued', 'acknowledged', 'completed', 'cancelled']:
        po_by_status[s] = po_qs.filter(status=s).count()

    # DN stats
    dn_qs = DeliveryNote.objects.all()
    if not has_full_access:
        if user.is_admin_user or user.is_manager_user:
            dn_qs = dn_qs.filter(Q(created_by=user) | Q(project__region=user.region))
        else:
            dn_qs = dn_qs.filter(created_by=user)
    dn_total = dn_qs.count()

    # Summary stats — counted as the number of POs visible to the user
    # (the summary itself is a derived view of POs, not a separate record).
    summary_total = po_qs.count()

    # Inventory stats
    inv_qs = InventoryReport.objects.all()
    if not has_full_access:
        if user.is_admin_user or user.is_manager_user:
            inv_qs = inv_qs.filter(Q(created_by=user) | Q(project__region=user.region))
        else:
            inv_qs = inv_qs.filter(created_by=user)
    inv_total = inv_qs.count()
    total_inventory_items = InventoryItem.objects.filter(report__in=inv_qs).count()

    # Low stock items - pushed entirely into SQL via the custom queryset.
    # Previously this loaded every InventoryItem and ran the is_low_stock
    # Python property per row (N+1 over potentially thousands of items).
    low_stock_qs = (
        InventoryItem.objects
        .filter(report__in=inv_qs)
        .low_stock()
        .select_related('report')
    )
    low_stock_count = low_stock_qs.count()
    low_stock_items = list(low_stock_qs[:10])  # template only uses first 10

    # Recent activity
    recent_pos = po_qs.order_by('-created_at')[:5]
    recent_dns = dn_qs.order_by('-created_at')[:5]

    context = {
        'po_total': po_total,
        'po_by_status': po_by_status,
        'dn_total': dn_total,
        'summary_total': summary_total,
        'inv_total': inv_total,
        'total_inventory_items': total_inventory_items,
        'low_stock_count': low_stock_count,
        'low_stock_items': low_stock_items,
        'recent_pos': recent_pos,
        'recent_dns': recent_dns,
    }
    return render(request, 'procurement/dashboard.html', context)


def _procurement_team_users():
    """Active users who realistically issue POs.

    Used to populate the optional "Pick teammate" dropdown next to the
    PO Issuer text field — the form still saves whatever the text input
    holds, so the dropdown is just a UX shortcut.
    """
    from accounts.models import User, Role
    return User.objects.filter(
        is_active=True,
        role__name__in=[
            Role.SUPER_ADMIN, Role.ADMIN, Role.MANAGER,
            Role.PROCUREMENT_MGR, Role.PROCUREMENT_OFF,
        ],
    ).select_related('role').order_by('first_name', 'last_name', 'username')


def _visible_pos_for(user):
    """Function-based equivalent of ProcurementPermissionMixin.get_queryset.

    Used by export and approve endpoints (function-based views) so a user
    cannot read or sign a PO outside their region by guessing the PK.
    """
    qs = PurchaseOrder.objects.select_related('project', 'created_by').all()
    if not getattr(user, 'is_authenticated', False):
        return qs.none()
    if user.is_super_admin_user or user.is_procurement_user:
        return qs
    if user.is_admin_user or user.is_manager_user:
        return qs.filter(
            Q(created_by=user) |
            Q(project__region=user.region)
        )
    return qs.filter(created_by=user)


def _validate_signature_image(raw_bytes, *, max_bytes=2 * 1024 * 1024):
    """Verify ``raw_bytes`` is a real image and re-encode as a clean PNG.

    Defends against Pillow CVE chains and stored malicious payloads — the
    original bytes are discarded and only Pillow's safe re-encoded output
    is kept. Raises ValueError with a user-friendly message on failure.
    """
    from io import BytesIO
    from PIL import Image, UnidentifiedImageError
    if not raw_bytes:
        raise ValueError('Signature is empty.')
    if len(raw_bytes) > max_bytes:
        raise ValueError(f'Signature too large (max {max_bytes // (1024 * 1024)} MB).')
    # First pass — verify() catches structural problems before .load() runs
    # the heavier decoders.
    try:
        Image.open(BytesIO(raw_bytes)).verify()
    except (UnidentifiedImageError, Exception) as e:
        raise ValueError(f'Not a valid image: {e}') from e
    # Re-decode through a fresh handle (verify() consumes the stream) and
    # re-encode as PNG to strip any malicious metadata or polyglot tricks.
    img = Image.open(BytesIO(raw_bytes))
    img.load()
    if img.mode in ('RGBA', 'LA'):
        out_img = img
    else:
        out_img = img.convert('RGB')
    out = BytesIO()
    out_img.save(out, format='PNG', optimize=True)
    return out.getvalue()


class ProcurementPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    def get_queryset(self):
        return _visible_pos_for(self.request.user)


# ─── PO CRUD ─────────────────────────────────────────────────

class POListView(CapabilityRequiredMixin, ProcurementPermissionMixin, ListView):
    capability = 'po.access'
    model = PurchaseOrder
    template_name = 'procurement/po_list.html'
    context_object_name = 'purchase_orders'
    paginate_by = 25

    def test_func(self):
        return True

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        status = self.request.GET.get('status')
        if search:
            queryset = queryset.filter(
                Q(po_number__icontains=search) |
                Q(vendor_name__icontains=search) |
                Q(project_name__icontains=search) |
                Q(end_user__icontains=search)
            )
        if status:
            queryset = queryset.filter(status=status)
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['filter_form'] = POFilterForm(self.request.GET)
        context['total_count'] = self.get_queryset().count()
        return context


def _po_terms_by_category(selected_ids):
    """Build the terms_by_category structure used in PO create/edit/detail."""
    from costing.models import TermsTemplate
    selected_ids = set(selected_ids or [])
    # Procurement sees procurement-tagged terms (and shared 'both'), not sales.
    all_terms = TermsTemplate.objects.filter(usage__in=['procurement', 'both'])
    terms_by_category = {}
    for cat_value, cat_label in TermsTemplate.CATEGORY_CHOICES:
        terms_by_category[cat_label] = [
            {'template': t, 'selected': t.pk in selected_ids}
            for t in all_terms if t.category == cat_value
        ]
    return terms_by_category


class POCreateView(ProcurementPermissionMixin, CreateView):
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'procurement/po_form.html'

    def test_func(self):
        return True

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = POItemFormSet(self.request.POST, prefix='items')
            posted_term_ids = [int(x) for x in self.request.POST.getlist('selected_terms') if x.isdigit()]
            context['terms_by_category'] = _po_terms_by_category(posted_term_ids)
        else:
            context['item_formset'] = POItemFormSet(prefix='items')
            context['terms_by_category'] = _po_terms_by_category([])
        context['title'] = 'Create Purchase Order'
        context['system_suggestions'] = PurchaseOrder.SYSTEM_SUGGESTIONS
        context['procurement_team'] = _procurement_team_users()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if item_formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            item_formset.instance = self.object
            item_formset.save()
            term_ids = [int(x) for x in self.request.POST.getlist('selected_terms') if x.isdigit()]
            self.object.selected_terms.set(term_ids)
            messages.success(self.request, f'Purchase Order {self.object.po_number} created successfully.')
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:po_detail', kwargs={'pk': self.object.pk})


# ─── Quotation import → PO (AI-assisted) ──────────────────────

def _can_procure(user):
    return bool(user.is_super_admin_user or user.is_admin_user
                or getattr(user, 'is_procurement_user', False))


def _prefilled_item_formset(items):
    """A PO item formset pre-populated (unbound) from extracted quotation rows."""
    from django.forms import inlineformset_factory
    FS = inlineformset_factory(
        PurchaseOrder, PurchaseOrderItem, form=PurchaseOrderItemForm,
        extra=max(len(items), 1), can_delete=True)
    initial = []
    for i, it in enumerate(items):
        initial.append({
            'serial_number': i + 1,
            'description': it.get('description', ''),
            'make_model': it.get('make_model', ''),
            'quantity': it.get('quantity') or 0,
            'uom': it.get('uom') or 'Nos',
            'rate_per_unit': it.get('unit_price') or 0,
            'order': i,
        })
    return FS(prefix='items', initial=initial, instance=PurchaseOrder())


@login_required
def quotation_import(request):
    """Upload a supplier quotation PDF; extract it into a structured draft the
    user reviews before a PO is created."""
    if not _can_procure(request.user):
        messages.error(request, 'Only procurement team members can import quotations.')
        return redirect('procurement:po_list')

    if request.method == 'POST':
        f = request.FILES.get('quotation_file')
        if not f:
            messages.error(request, 'Please choose a PDF quotation to upload.')
            return redirect('procurement:quotation_import')

        qi = QuotationImport.objects.create(
            file=f, original_filename=f.name, created_by=request.user,
            status='pending')
        ok, err = _extract_quotation_into(qi)
        if not ok:
            messages.error(request, f'Could not extract this quotation: {err}')
            return redirect('procurement:quotation_import')
        n = len(qi.line_items)
        messages.success(request, f'Extracted {n} line item{"" if n == 1 else "s"} — review and create the PO.')
        return redirect('procurement:quotation_review', pk=qi.pk)

    imports = QuotationImport.objects.filter(created_by=request.user)[:25] \
        if not (request.user.is_super_admin_user or request.user.is_admin_user) \
        else QuotationImport.objects.all()[:25]
    return render(request, 'procurement/quotation_import.html', {'imports': imports})


def _extract_quotation_into(qi):
    """Run text + AI extraction into a QuotationImport. Always records an
    outcome (extracted/failed) so the row never stays 'pending' as long as the
    request completes. Returns (ok, error_message)."""
    from .quotation_extract import extract_text_from_pdf, extract_quotation
    try:
        qi.file.open('rb')
        text = extract_text_from_pdf(qi.file)
        qi.file.close()
        data, model = extract_quotation(text)
        qi.extracted_data = data
        qi.model_used = model
        qi.status = 'extracted'
        qi.error = ''
        qi.save(update_fields=['extracted_data', 'model_used', 'status', 'error', 'updated_at'])
        return True, ''
    except Exception as e:
        qi.status = 'failed'
        qi.error = str(e)
        qi.save(update_fields=['status', 'error', 'updated_at'])
        return False, str(e)


@login_required
@require_POST
def quotation_retry(request, pk):
    """Re-run extraction on a pending/failed quotation (e.g. after the worker
    was killed mid-call, or a transient API error)."""
    qi = get_object_or_404(QuotationImport, pk=pk)
    if not _can_procure(request.user):
        messages.error(request, 'Permission denied.')
        return redirect('procurement:po_list')
    if qi.status == 'converted':
        messages.info(request, 'This quotation has already been converted to a PO.')
        return redirect('procurement:quotation_import')
    qi.status = 'pending'
    qi.save(update_fields=['status', 'updated_at'])
    ok, err = _extract_quotation_into(qi)
    if not ok:
        messages.error(request, f'Extraction failed again: {err}')
        return redirect('procurement:quotation_import')
    messages.success(request, 'Extracted — review and create the PO.')
    return redirect('procurement:quotation_review', pk=qi.pk)


@login_required
def quotation_review(request, pk):
    """Review the extracted quotation in the PO editor and create the PO."""
    qi = get_object_or_404(QuotationImport, pk=pk)
    if not _can_procure(request.user):
        messages.error(request, 'Permission denied.')
        return redirect('procurement:po_list')

    data = qi.extracted_data or {}

    def _render(form, item_formset):
        return render(request, 'procurement/po_form.html', {
            'form': form,
            'item_formset': item_formset,
            'title': f'Review Quotation → Create PO',
            'system_suggestions': PurchaseOrder.SYSTEM_SUGGESTIONS,
            'procurement_team': _procurement_team_users(),
            'terms_by_category': _po_terms_by_category([]),
            'quotation': qi,
            'is_quotation_review': True,
        })

    if request.method == 'POST':
        form = PurchaseOrderForm(request.POST, user=request.user)
        item_formset = POItemFormSet(request.POST, prefix='items')
        if form.is_valid() and item_formset.is_valid():
            po = form.save(commit=False)
            po.created_by = request.user
            po.save()
            item_formset.instance = po
            item_formset.save()
            term_ids = [int(x) for x in request.POST.getlist('selected_terms') if x.isdigit()]
            po.selected_terms.set(term_ids)
            qi.purchase_order = po
            qi.status = 'converted'
            qi.save(update_fields=['purchase_order', 'status', 'updated_at'])
            messages.success(request, f'Purchase Order {po.po_number} created from quotation.')
            return redirect('procurement:po_detail', pk=po.pk)
        return _render(form, item_formset)

    # GET — pre-fill the PO editor from the extraction.
    today = datetime.now().date()
    initial = {
        'po_date': today,
        'vendor_name': data.get('vendor_name', ''),
        'vendor_contact_person': data.get('vendor_contact_person', ''),
        'vendor_contact_email': data.get('vendor_contact_email', ''),
        'vendor_contact_tel': data.get('vendor_contact_tel', ''),
        'project_name': data.get('project_name', ''),
        'currency': data.get('currency', 'SAR'),
        'vat_rate': data.get('vat_rate', 15),
        'mr_item_number': data.get('quotation_reference', ''),
        'po_issued_by': request.user.get_full_name() or request.user.username,
        'issuer_email': request.user.email or '',
    }
    form = PurchaseOrderForm(initial=initial, user=request.user)
    item_formset = _prefilled_item_formset(data.get('line_items', []))
    return _render(form, item_formset)


class PODetailView(ProcurementPermissionMixin, DetailView):
    model = PurchaseOrder
    template_name = 'procurement/po_detail.html'
    context_object_name = 'po'

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        po = self.object
        context['items'] = po.items.all()
        context['terms_by_category'] = _po_terms_by_category(
            po.selected_terms.values_list('pk', flat=True)
        )
        # Per-stage approval permission for the current viewer — keys
        # ('scm','pm','coo','ceo') the user is authorised to sign.
        context['approvable_stages'] = {
            key for key, _, _ in po.APPROVAL_STAGES
            if po.can_user_approve_stage(self.request.user, key)
        }
        return context


class POUpdateView(ProcurementPermissionMixin, UpdateView):
    model = PurchaseOrder
    form_class = PurchaseOrderForm
    template_name = 'procurement/po_form.html'

    def test_func(self):
        obj = self.get_object()
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return obj.created_by == user

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = POItemFormSet(self.request.POST, instance=self.object, prefix='items')
            posted_term_ids = [int(x) for x in self.request.POST.getlist('selected_terms') if x.isdigit()]
            context['terms_by_category'] = _po_terms_by_category(posted_term_ids)
        else:
            context['item_formset'] = POItemFormSet(instance=self.object, prefix='items')
            current_ids = list(self.object.selected_terms.values_list('pk', flat=True))
            context['terms_by_category'] = _po_terms_by_category(current_ids)
        context['title'] = f'Edit PO: {self.object.po_number}'
        context['system_suggestions'] = PurchaseOrder.SYSTEM_SUGGESTIONS
        context['procurement_team'] = _procurement_team_users()
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if item_formset.is_valid():
            self.object = form.save()
            item_formset.instance = self.object
            item_formset.save()
            term_ids = [int(x) for x in self.request.POST.getlist('selected_terms') if x.isdigit()]
            self.object.selected_terms.set(term_ids)
            messages.success(self.request, f'Purchase Order {self.object.po_number} updated successfully.')
            return redirect(self.get_success_url())
        else:
            return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:po_detail', kwargs={'pk': self.object.pk})


def _uniform_vendor(items):
    """The single non-empty vendor shared by all items, else '' (for pre-filling
    a PO's header vendor when every picked line is from the same vendor)."""
    vendors = {(li.vendor_name or '').strip() for li in items}
    vendors = {v for v in vendors if v}
    return vendors.pop() if len(vendors) == 1 else ''


@login_required
def po_create_from_bom(request, sheet_pk):
    """Seed a draft PO with every supply line from a finance-approved budget,
    pre-priced at the budgeted unit price.

    Procurement works from the finance budget (not the costing sheet). Items are
    pre-filled (description / make-model / vendor / qty / uom / rate) so the user
    only confirms the PO number and vendor. A.1 supply lines only.
    """
    user = request.user
    if not (user.is_super_admin_user or user.is_admin_user or user.is_procurement_user):
        messages.error(request, 'Only procurement team members can create POs from a budget.')
        return redirect('procurement:approved_budgets')

    from costing.models import CostingSheet, ExchangeRate
    sheet = get_object_or_404(CostingSheet, pk=sheet_pk)
    rates = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}

    picks = []
    for section in sheet.sections.filter(is_optional=False).order_by('order', 'section_number'):
        for item in section.line_items.all().order_by('order', 'item_number'):
            item.set_exchange_rates_cache(rates)
            item.set_sheet_cache(sheet)
            picks.append((section, item))

    placeholder_po_number = f'DRAFT-S{sheet.pk}-{int(datetime.now().timestamp())}'
    po = PurchaseOrder.objects.create(
        po_date=datetime.now().date(),
        po_number=placeholder_po_number,
        vendor_name=_uniform_vendor([li for _, li in picks]),
        po_issued_by=user.get_full_name() or user.username,
        issuer_email=user.email or '',
        project=sheet.project,
        project_name=sheet.project.project_name if sheet.project else (sheet.title or ''),
        created_by=user,
    )

    serial = 1
    for section, item in picks:
        make_model = ' '.join(filter(None, [item.make, item.model_number])).strip()
        PurchaseOrderItem.objects.create(
            purchase_order=po,
            serial_number=serial,
            system=section.title or '',
            make_model=make_model,
            vendor_name=item.vendor_name or '',
            description=item.description,
            quantity=item.quantity,
            uom=item.unit or 'Nos',
            rate_per_unit=item.budget_unit_price(),
            order=serial,
        )
        serial += 1

    copied = serial - 1
    messages.success(
        request,
        f'Draft PO seeded with {copied} budgeted item{"" if copied == 1 else "s"}. '
        f'Replace the placeholder PO number, confirm the vendor, then save.',
    )
    return redirect('procurement:po_update', pk=po.pk)


@login_required
def approved_budgets(request):
    """Procurement landing — finance-approved budgets ready to procure, in the
    user's region. Procurement works from these budgets, not the costing sheet."""
    user = request.user
    if not (user.is_super_admin_user or user.is_admin_user or user.is_procurement_user):
        messages.error(request, 'Only procurement team members can view approved budgets.')
        return redirect('procurement:dashboard')

    from costing.models import CostingSheet
    sheets = (CostingSheet.objects
              .filter(workflow_stage='finance_approved')
              .select_related('project', 'project__region', 'finance_approved_by')
              .order_by('-finance_approved_at', '-updated_at'))
    if not (user.is_super_admin_user or user.is_admin_user):
        sheets = sheets.filter(project__region=user.region)

    return render(request, 'procurement/approved_budgets.html', {'sheets': sheets})


@login_required
def bom_procurement_tracker(request, sheet_pk):
    """Approved-budget procurement page (procurement never sees the costing sheet).

    Lists the finance-approved budget's A.1 supply lines with their budgeted
    unit/line price, qty, vendor, make/model and unit, plus a per-line
    procurement status. The user picks lines for *this* PO and submits; a draft
    PO is created pre-priced at the budgeted unit price, with each item carrying
    its vendor and a source_bom_item back-link so it can't be re-picked. The
    PO's header vendor is pre-filled when every picked line shares one vendor.
    """
    user = request.user
    if not (user.is_super_admin_user or user.is_admin_user or user.is_procurement_user):
        messages.error(request, 'Only procurement team members can procure a budget.')
        return redirect('procurement:approved_budgets')

    from costing.models import CostingSheet, CostingLineItem, ExchangeRate
    sheet = get_object_or_404(CostingSheet, pk=sheet_pk)

    # Region scope (super admin / admin bypass).
    if (not (user.is_super_admin_user or user.is_admin_user)
            and (not sheet.project or sheet.project.region_id != user.region_id)):
        messages.error(request, 'You can only procure budgets for projects in your region.')
        return redirect('procurement:approved_budgets')

    project_status = sheet.project.status if sheet.project else None
    if not project_status or project_status.category != 'won':
        messages.error(request, 'Procurement is only available for Won projects.')
        return redirect('procurement:approved_budgets')
    if sheet.workflow_stage != 'finance_approved':
        messages.error(
            request,
            f'This budget is not finance-approved yet. Current stage: {sheet.get_workflow_stage_display()}.')
        return redirect('procurement:approved_budgets')

    rates = {r.currency_code: r.rate_to_usd for r in ExchangeRate.objects.all()}

    if request.method == 'POST':
        item_ids = [int(x) for x in request.POST.getlist('item_ids') if x.isdigit()]
        if not item_ids:
            messages.error(request, 'Pick at least one item to add to the PO.')
        else:
            picked = list(
                CostingLineItem.objects
                .filter(pk__in=item_ids, section__costing_sheet=sheet, section__is_optional=False)
                .select_related('section'))
            # Skip items claimed by another PO between render and POST.
            picked = [li for li in picked if not li.procured_po_items.exists()]
            if not picked:
                messages.warning(request, 'Every item you picked is already on another PO — nothing to add.')
                return redirect('procurement:bom_procurement_tracker', sheet_pk=sheet.pk)

            placeholder_po_number = f'DRAFT-S{sheet.pk}-{int(datetime.now().timestamp())}'
            po = PurchaseOrder.objects.create(
                po_date=datetime.now().date(),
                po_number=placeholder_po_number,
                vendor_name=_uniform_vendor(picked),
                po_issued_by=user.get_full_name() or user.username,
                issuer_email=user.email or '',
                project=sheet.project,
                project_name=sheet.project.project_name if sheet.project else (sheet.title or ''),
                created_by=user,
            )
            serial = 1
            for li in picked:
                li.set_exchange_rates_cache(rates)
                li.set_sheet_cache(sheet)
                make_model = ' '.join(filter(None, [li.make, li.model_number])).strip()
                PurchaseOrderItem.objects.create(
                    purchase_order=po,
                    serial_number=serial,
                    system=li.section.title or '',
                    make_model=make_model,
                    vendor_name=li.vendor_name or '',
                    description=li.description,
                    quantity=li.quantity,
                    uom=li.unit or 'Nos',
                    rate_per_unit=li.budget_unit_price(),
                    order=serial,
                    source_bom_item=li,
                )
                serial += 1
            messages.success(
                request,
                f'Draft PO seeded with {serial - 1} budgeted item{"" if serial - 1 == 1 else "s"}. '
                f'Replace the placeholder PO number, confirm the vendor, then save.')
            return redirect('procurement:po_update', pk=po.pk)

    # Build per-section rows (A.1 supply, non-optional) with budgeted prices.
    rows = []
    available_count = procured_count = 0
    budget_total = Decimal('0')
    for section in (sheet.sections.filter(is_optional=False)
                    .prefetch_related('line_items__procured_po_items__purchase_order')
                    .order_by('order', 'section_number')):
        section_items = []
        for li in section.line_items.all().order_by('order', 'item_number'):
            li.set_exchange_rates_cache(rates)
            li.set_sheet_cache(sheet)
            procured = list(li.procured_po_items.select_related('purchase_order').all())
            line_price = li.budget_line_price()
            budget_total += line_price
            if procured:
                procured_count += 1
            else:
                available_count += 1
            section_items.append({
                'item': li, 'procured_in': procured, 'is_available': not procured,
                'unit_price': li.budget_unit_price(), 'line_price': line_price,
            })
        if section_items:
            rows.append({'section': section, 'items': section_items})

    return render(request, 'procurement/bom_procurement_tracker.html', {
        'sheet': sheet,
        'rows': rows,
        'total_count': available_count + procured_count,
        'available_count': available_count,
        'procured_count': procured_count,
        'budget_total': budget_total,
    })


@login_required
@require_POST
def po_approve_stage(request, pk, stage):
    """Sign one approval stage on a PO and stamp the timestamp + approver.

    Strict sequencing: only the *current* pending stage can be signed.
    Once signed a stage is locked — no un-approve. If the freshly signed
    stage was the last required one, the PO is now released and exports
    unlock automatically.

    Region- and role-scoped via _visible_pos_for() — guessing PKs across
    regions returns 404. Stage role gate is enforced separately. The
    read-check-write window runs inside an atomic block holding a row-level
    lock so concurrent Approve clicks can't race past each other.
    """
    # First read the signature payload out of the request *before* the
    # transaction opens — Pillow validation on potentially-malicious
    # bytes shouldn't run inside a row-locked block.
    raw = None
    if 'signature_file' in request.FILES:
        raw = request.FILES['signature_file'].read()
    elif request.POST.get('signature_data', '').startswith('data:image/'):
        import base64
        _, _, b64 = request.POST['signature_data'].partition(',')
        try:
            raw = base64.b64decode(b64)
        except Exception:
            return JsonResponse({'error': 'Could not decode the drawn signature.'}, status=400)
    if raw is None:
        return JsonResponse({
            'error': 'Signature required — upload an image or draw on the pad.',
        }, status=400)

    # Re-encode through Pillow so we never persist arbitrary user-supplied
    # bytes — defends against Pillow-decoder CVEs and polyglot uploads.
    try:
        clean = _validate_signature_image(raw)
    except ValueError as e:
        return JsonResponse({'error': str(e)}, status=400)

    sig_field = f'{stage}_signature'
    from django.core.files.base import ContentFile

    # The whole approval — DB lookup, signature upload to object storage
    # (Cloudflare R2 in production), and response building — is guarded so any
    # failure surfaces as a readable JSON error with the exception detail,
    # instead of an opaque HTML 500 the browser reports as a "Network error".
    # The full traceback is logged for diagnosis.
    try:
        with transaction.atomic():
            # Lock just the bare PO row to serialise concurrent approvals.
            # We must NOT join here: on PostgreSQL "SELECT ... FOR UPDATE"
            # cannot lock the nullable side of an outer join, and
            # _visible_pos_for() select_related()s the nullable project /
            # created_by FKs — so locking through it raises NotSupportedError
            # (Postgres only; SQLite silently allows it). Lock the row, then
            # scope-check region/role with a separate, unlocked query.
            try:
                po = PurchaseOrder.objects.select_for_update().get(pk=pk)
            except PurchaseOrder.DoesNotExist:
                return JsonResponse({'error': 'PO not found or not in your scope.'}, status=404)

            if not _visible_pos_for(request.user).filter(pk=pk).exists():
                return JsonResponse({'error': 'PO not found or not in your scope.'}, status=404)

            if stage not in po.required_stages:
                return JsonResponse({'error': 'This stage is not required for this PO.'}, status=400)

            cur = po.current_stage
            if not cur or cur['key'] != stage:
                return JsonResponse({
                    'error': f'Out of sequence — next pending stage is {(cur["label"] if cur else "(none, already released)")}.',
                }, status=400)

            if not po.can_user_approve_stage(request.user, stage):
                return JsonResponse({'error': 'You do not have permission to approve this stage.'}, status=403)

            getattr(po, sig_field).save(
                f'po{po.pk}_{stage}_sig.png',
                ContentFile(clean),
                save=False,
            )
            setattr(po, f'{stage}_approved_at', timezone.now())
            setattr(po, f'{stage}_approved_by', request.user)
            po.save(update_fields=[
                f'{stage}_approved_at', f'{stage}_approved_by', sig_field, 'updated_at',
            ])

        new_cur = po.current_stage
        return JsonResponse({
            'ok': True,
            'released': po.is_released,
            'stage_signed': stage,
            'next_stage': new_cur['key'] if new_cur else None,
            'next_stage_label': new_cur['label'] if new_cur else None,
            'approved_by': (request.user.get_full_name() or request.user.username),
        })
    except Exception as e:
        import logging
        logging.getLogger(__name__).exception(
            'PO %s stage %s approval failed', pk, stage)
        return JsonResponse({
            'error': f'Approval failed: {type(e).__name__}: {e}',
        }, status=500)


class PODeleteView(ProcurementPermissionMixin, DeleteView):
    model = PurchaseOrder
    template_name = 'procurement/po_confirm_delete.html'
    success_url = reverse_lazy('procurement:po_list')

    def test_func(self):
        user = self.request.user
        obj = self.get_object()
        # Audit-trail protection: once *any* approval stage has been signed,
        # only super admin can delete the PO. Without this gate the original
        # creator could erase a fully-signed PO and destroy the audit chain
        # the approval workflow exists to produce.
        any_signed = any(
            getattr(obj, f'{s}_approved_at') is not None
            for s in ('scm', 'pm', 'coo', 'ceo')
        )
        if any_signed and not user.is_super_admin_user:
            return False
        if user.is_super_admin_user or user.is_admin_user:
            return True
        return obj.created_by == user

    def form_valid(self, form):
        messages.success(self.request, 'Purchase Order deleted successfully.')
        return super().form_valid(form)


# ─── Excel Export ─────────────────────────────────────────────

@login_required
def po_export_excel(request, pk):
    """Export a Purchase Order to Excel matching the original format."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    # Region/role scoped — guessing a PK from outside the user's scope
    # returns 404, so this also enforces the same region rules the list
    # and detail views use.
    po = get_object_or_404(_visible_pos_for(request.user), pk=pk)
    if not po.is_released:
        cur = po.current_stage['label'] if po.current_stage else 'approval'
        messages.error(request, f'PO not released yet — pending {cur}. Excel export is locked until all required approvals are signed.')
        return redirect('procurement:po_detail', pk=pk)
    items = po.items.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PURCHASE ORDER'

    # Styles
    bold = Font(bold=True)
    bold_red = Font(bold=True, color='C41E3A', size=11)
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap = Alignment(wrap_text=True, vertical='top')

    # ── Header Section ──
    headers_left = [
        ('PO Date', po.po_date.strftime('%d-%b-%Y') if po.po_date else ''),
        ('PO S. No.', po.po_number),
        ('C. Center', po.get_cost_center_display()),
        ('Vendor', po.vendor_name),
        ('Contact Person', po.vendor_contact_person),
        ('Contact Email', po.vendor_contact_email),
        ('Contact Tel', po.vendor_contact_tel),
    ]
    headers_right = [
        ('PO issued by', po.po_issued_by),
        ('Contact Email', po.issuer_email),
        ('Project Name', po.project_name),
        ('End User', po.end_user),
        ('MR, Item No.', po.mr_item_number),
        ('Delivery Incoterms', po.get_delivery_incoterms_display() if po.delivery_incoterms else ''),
        ('Delivery Location', po.delivery_location),
    ]

    for i, (label, value) in enumerate(headers_left, 1):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=str(value))

    for i, (label, value) in enumerate(headers_right, 1):
        ws.cell(row=i, column=5, value=label).font = bold
        ws.cell(row=i, column=6, value=str(value))

    # ── Line Items Table ──
    table_row = 11
    col_headers = ['S No.', 'Make/Model', 'Item Descriptions / Specification',
                    '', '', 'Quantity', 'UOM', f'Rate/unit ({po.currency})', f'Total Value ({po.currency})', 'Remarks']
    for col, h in enumerate(col_headers, 1):
        cell = ws.cell(row=table_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Merge description header across C-E
    ws.merge_cells(start_row=table_row, start_column=3, end_row=table_row, end_column=5)

    row = table_row + 1
    for item in items:
        ws.cell(row=row, column=1, value=item.serial_number).border = thin_border
        ws.cell(row=row, column=2, value=item.make_model).border = thin_border
        desc_cell = ws.cell(row=row, column=3, value=item.description)
        desc_cell.border = thin_border
        desc_cell.alignment = wrap
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row=row, column=6, value=float(item.quantity)).border = thin_border
        ws.cell(row=row, column=7, value=item.uom).border = thin_border
        ws.cell(row=row, column=8, value=float(item.rate_per_unit)).border = thin_border
        ws.cell(row=row, column=9, value=float(item.total_value)).border = thin_border
        ws.cell(row=row, column=10, value=item.remarks).border = thin_border
        row += 1

    # ── Totals ──
    row += 1
    totals = [
        ('Base Amount', float(po.base_amount)),
        ('Discount', float(po.discount_amount)),
        ('Gross Value', float(po.gross_value)),
        ('VAT', float(po.vat_amount)),
        (f'Total Value in {po.currency}', float(po.total_value)),
    ]
    for label, val in totals:
        ws.cell(row=row, column=3, value=label).font = bold
        if label == 'VAT':
            ws.cell(row=row, column=6, value=float(po.vat_rate / 100))
            ws.cell(row=row, column=6).number_format = '0%'
        if label == 'Discount':
            ws.cell(row=row, column=6, value=float(po.discount_rate / 100))
            ws.cell(row=row, column=6).number_format = '0%'
        val_cell = ws.cell(row=row, column=9, value=val)
        val_cell.font = bold
        val_cell.number_format = '#,##0.00'
        row += 1

    # ── T&C ──
    selected_terms_xl = list(po.selected_terms.all())
    if po.terms_and_conditions or selected_terms_xl:
        row += 2
        ws.cell(row=row, column=1, value='TERMS AND CONDITIONS').font = bold_red
        row += 1
        for tmpl in selected_terms_xl:
            ws.cell(row=row, column=2, value=tmpl.name).font = bold
            row += 1
            for line in tmpl.content.split('\n'):
                if line.strip():
                    ws.cell(row=row, column=2, value=line.strip()).alignment = wrap
                    row += 1
        if po.terms_and_conditions:
            for line in po.terms_and_conditions.split('\n'):
                if line.strip():
                    ws.cell(row=row, column=2, value=line.strip()).alignment = wrap
                    row += 1

    # Column widths
    widths = [8, 18, 15, 15, 15, 10, 8, 14, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = _safe_filename(po.po_number, prefix='PO', extension='xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── PDF Export ───────────────────────────────────────────────

@login_required
def po_export_pdf(request, pk, unpriced=False):
    """Export a Purchase Order to PDF matching the original format.

    When ``unpriced`` is True, all commercial figures are omitted — the
    Rate/Unit and Total columns, the totals block, and the amount-in-words
    line — producing a scope-only copy safe to share without revealing pricing.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.pdfgen.canvas import Canvas
    from io import BytesIO
    from django.contrib.staticfiles.finders import find as find_static

    po = get_object_or_404(_visible_pos_for(request.user), pk=pk)
    items = po.items.all()
    is_draft = not po.is_released

    NumberedCanvas = _make_numbered_canvas(draft=is_draft)

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle('POTitle', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#C41E3A'), spaceAfter=6)
    label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    value_style = ParagraphStyle('Value', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Norm', parent=styles['Normal'], fontSize=8)
    small_style = ParagraphStyle('Small', parent=styles['Normal'], fontSize=7)
    tc_style = ParagraphStyle('TC', parent=styles['Normal'], fontSize=7, leading=9)
    right_style = ParagraphStyle('Right', parent=styles['Normal'], fontSize=8, alignment=TA_RIGHT)
    right_bold = ParagraphStyle('RightBold', parent=styles['Normal'], fontSize=8, alignment=TA_RIGHT, fontName='Helvetica-Bold')

    # ── Logo + Title ──
    if is_draft:
        pending = po.current_stage['label'] if po.current_stage else 'approval'
        title_html = (
            'PURCHASE ORDER '
            f'<font size="9" color="#C41E3A">— DRAFT (awaiting {pending})</font>'
        )
    else:
        title_html = 'PURCHASE ORDER'
    if unpriced:
        title_html += ' <font size="9" color="#6C757D">— UNPRICED</font>'
    logo_path = find_static('images/leap_logo.jpg')
    if logo_path:
        from reportlab.platypus import Image
        logo = Image(logo_path, width=50*mm, height=15*mm)
        title_table = Table([[logo, Paragraph(title_html, title_style)]], colWidths=[55*mm, 120*mm])
        title_table.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(title_table)
    else:
        elements.append(Paragraph(title_html, title_style))
    elements.append(Spacer(1, 4*mm))

    # ── Header Info ──
    def lv(label, value):
        return [Paragraph(label, label_style), Paragraph(str(value or '-'), value_style)]

    header_data = [
        lv('PO Date', po.po_date.strftime('%d %b %Y') if po.po_date else '-') +
        [''] +
        lv('PO Issued By', po.po_issued_by),

        lv('PO Number', po.po_number) +
        [''] +
        lv('Contact Email', po.issuer_email),

        lv('Cost Center', po.get_cost_center_display()) +
        [''] +
        lv('Project Name', po.project_name),

        lv('Vendor', po.vendor_name) +
        [''] +
        lv('End User', po.end_user),

        lv('Contact Person', po.vendor_contact_person) +
        [''] +
        lv('MR / Item No.', po.mr_item_number),

        lv('Contact Email', po.vendor_contact_email) +
        [''] +
        lv('Delivery Incoterms', po.get_delivery_incoterms_display() if po.delivery_incoterms else '-'),

        lv('Contact Tel', po.vendor_contact_tel) +
        [''] +
        lv('Delivery Location', po.delivery_location),
    ]
    header_table = Table(header_data, colWidths=[22*mm, 60*mm, 5*mm, 22*mm, 60*mm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
        # Vertical lines: only between label↔value pairs (left and right groups).
        # Skip the center spacer column so there's no line through the middle.
        ('LINEAFTER', (0, 0), (0, -1), 0.4, colors.HexColor('#E0E0E0')),
        ('LINEAFTER', (3, 0), (3, -1), 0.4, colors.HexColor('#E0E0E0')),
        # Horizontal lines between rows
        ('LINEBELOW', (0, 0), (-1, -2), 0.4, colors.HexColor('#E0E0E0')),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 5*mm))

    # ── Line Items Table ──
    # Unpriced copies drop the Rate/Unit and Total columns; the freed width is
    # redistributed to Description and Remarks so the table still fills the page.
    if unpriced:
        col_widths = [12*mm, 30*mm, 80*mm, 16*mm, 16*mm, 31*mm]
        item_header = [
            Paragraph('<b>S.No.</b>', small_style),
            Paragraph('<b>Make/Model</b>', small_style),
            Paragraph('<b>Item Description</b>', small_style),
            Paragraph('<b>Qty</b>', small_style),
            Paragraph('<b>UOM</b>', small_style),
            Paragraph('<b>Remarks</b>', small_style),
        ]
    else:
        col_widths = [12*mm, 25*mm, 55*mm, 15*mm, 14*mm, 22*mm, 22*mm, 20*mm]
        item_header = [
            Paragraph('<b>S.No.</b>', small_style),
            Paragraph('<b>Make/Model</b>', small_style),
            Paragraph('<b>Item Description</b>', small_style),
            Paragraph('<b>Qty</b>', small_style),
            Paragraph('<b>UOM</b>', small_style),
            Paragraph('<b>Rate/Unit</b>', small_style),
            Paragraph(f'<b>Total ({po.currency})</b>', small_style),
            Paragraph('<b>Remarks</b>', small_style),
        ]
    item_data = [item_header]

    dark_blue = colors.HexColor('#C41E3A')

    for item in items:
        if unpriced:
            item_data.append([
                Paragraph(str(item.serial_number), normal_style),
                Paragraph(item.make_model or '', small_style),
                Paragraph(item.description, small_style),
                Paragraph(f'{item.quantity:,.0f}', right_style),
                Paragraph(item.uom, small_style),
                Paragraph(item.remarks or '', small_style),
            ])
        else:
            item_data.append([
                Paragraph(str(item.serial_number), normal_style),
                Paragraph(item.make_model or '', small_style),
                Paragraph(item.description, small_style),
                Paragraph(f'{item.quantity:,.0f}', right_style),
                Paragraph(item.uom, small_style),
                Paragraph(f'{item.rate_per_unit:,.2f}', right_style),
                Paragraph(f'{item.total_value:,.2f}', right_bold),
                Paragraph(item.remarks or '', small_style),
            ])

    item_table = Table(item_data, colWidths=col_widths, repeatRows=1)
    item_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5D7DC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#495057')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#C41E3A')),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(item_table)
    elements.append(Spacer(1, 4*mm))

    # ── Totals ── (omitted entirely on unpriced copies)
    if not unpriced:
        totals_data = [
            ['', '', 'Base Amount', '', '', '', f'{po.base_amount:,.2f}', ''],
        ]
        if po.discount_rate:
            totals_data.append(['', '', f'Discount ({po.discount_rate:.0f}%)', '', '', '', f'-{po.discount_amount:,.2f}', ''])
        totals_data.append(['', '', 'Gross Value', '', '', '', f'{po.gross_value:,.2f}', ''])
        totals_data.append(['', '', f'VAT ({po.vat_rate:.0f}%)', '', '', '', f'{po.vat_amount:,.2f}', ''])
        totals_data.append(['', '', f'Total Value in {po.currency}', '', '', '', f'{po.total_value:,.2f}', ''])
        total_row_idx = len(totals_data) - 1  # for SPAN/style refs below

        # Amount-in-words row sits inside the same totals table so it aligns
        # to the totals column block (cols 2-6) instead of free-floating.
        amt_words_style = ParagraphStyle(
            'AmtWords', parent=styles['Normal'], fontSize=8, leading=11,
        )
        amt_words_para = Paragraph(
            f'<b>Amount in words:</b> {_amount_in_words(po.total_value, currency=po.currency)}',
            amt_words_style,
        )
        totals_data.append(['', '', amt_words_para, '', '', '', '', ''])
        amt_row_idx = len(totals_data) - 1

        totals_table = Table(totals_data, colWidths=col_widths)
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (2, 0), (2, total_row_idx), 'Helvetica-Bold'),
            ('FONTNAME', (6, 0), (6, total_row_idx), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (6, 0), (6, total_row_idx), 'RIGHT'),
            ('LINEABOVE', (2, total_row_idx), (6, total_row_idx), 1, dark_blue),
            ('LINEBELOW', (2, total_row_idx), (6, total_row_idx), 1.5, dark_blue),
            ('BACKGROUND', (2, total_row_idx), (6, total_row_idx), colors.HexColor('#FBE8EC')),
            # Amount-in-words row — span across the totals column block, left-aligned,
            # vertically padded so it doesn't sit flush against the total row.
            ('SPAN', (2, amt_row_idx), (6, amt_row_idx)),
            ('VALIGN', (2, amt_row_idx), (6, amt_row_idx), 'TOP'),
            ('TOPPADDING', (2, amt_row_idx), (6, amt_row_idx), 4),
            ('BOTTOMPADDING', (2, amt_row_idx), (6, amt_row_idx), 0),
            ('TOPPADDING', (0, 0), (-1, total_row_idx), 2),
            ('BOTTOMPADDING', (0, 0), (-1, total_row_idx), 2),
        ]))
        elements.append(totals_table)

    # ── Approvals — rendered progressively as each stage is signed.
    # Order: SCM → PM → COO → CEO. CEO is omitted entirely for POs under
    # 1M SAR. po.approved_stages excludes unsigned stages, so a draft
    # export renders only the signatures collected so far (or no signature
    # block at all when the PO has not been signed yet).
    from xml.sax.saxutils import escape as _xml_escape

    elements.append(Spacer(1, 8*mm))
    approvals = po.approved_stages
    if approvals:
        from reportlab.platypus import Image as RLImage
        col_w = max(40*mm, 180*mm / max(len(approvals), 1))
        sig_h = 16*mm  # signature image height — width auto-scales

        # Row 1: signature image (or empty cell if none uploaded yet).
        # Read via FieldFile so this works with both local storage and R2.
        from io import BytesIO as _BytesIO
        sig_row = []
        for s in approvals:
            sig = s['signature']
            placed = False
            if sig:
                try:
                    sig.open('rb')
                    data = sig.read()
                    sig.close()
                    img = RLImage(_BytesIO(data),
                                  width=col_w - 6*mm, height=sig_h,
                                  kind='proportional')
                    sig_row.append(img)
                    placed = True
                except Exception:
                    pass
            if not placed:
                sig_row.append('')

        # Row 2: signature line + role label.
        sig_line = '___________________'
        label_row = [
            Paragraph(
                f'<para align="center">{sig_line}<br/><b>{_xml_escape(s["label"])}</b></para>',
                ParagraphStyle('AppRole', parent=styles['Normal'], fontSize=8, leading=10),
            )
            for s in approvals
        ]
        # Row 3: signer name (hardcoded, e.g. Shaker Alkhalifah).
        name_row = [
            Paragraph(
                f'<para align="center">Name: <u>{_xml_escape(s["signer"])}</u></para>',
                ParagraphStyle('AppName', parent=styles['Normal'], fontSize=8, leading=10),
            )
            for s in approvals
        ]
        approval_table = Table(
            [sig_row, label_row, name_row],
            colWidths=[col_w] * len(approvals),
            rowHeights=[sig_h + 2*mm, None, None],
        )
        approval_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, 0), 'BOTTOM'),
            ('VALIGN', (0, 1), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('LINEBELOW', (0, -1), (-1, -1), 1, colors.HexColor('#C41E3A')),
        ]))
        elements.append(approval_table)
        elements.append(Spacer(1, 4*mm))

    # ── Terms & Conditions ──
    from costing.models import TermsTemplate as _TermsTemplate

    selected_terms = list(po.selected_terms.all())
    legacy_tc = (po.terms_and_conditions or '').strip()

    if selected_terms or legacy_tc:
        elements.append(Spacer(1, 6*mm))
        elements.append(Paragraph(
            '<b>TERMS AND CONDITIONS</b>',
            ParagraphStyle('TCHead', parent=styles['Heading2'], fontSize=10, textColor=colors.HexColor('#C41E3A'))
        ))
        elements.append(Spacer(1, 2*mm))

        sub_hdr_style = ParagraphStyle('TCSubHdr', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, spaceAfter=1)

        terms_grouped = {}
        for cat_value, _cat_label in _TermsTemplate.CATEGORY_CHOICES:
            items_in_cat = [t for t in selected_terms if t.category == cat_value]
            if items_in_cat:
                terms_grouped[cat_value] = items_in_cat

        for cat_value, _cat_label in _TermsTemplate.CATEGORY_CHOICES:
            for tmpl in terms_grouped.get(cat_value, []):
                elements.append(Paragraph(f'<b>{_xml_escape(tmpl.name)}</b>', sub_hdr_style))
                for line in [l.strip() for l in tmpl.content.splitlines() if l.strip()]:
                    elements.append(Paragraph(_xml_escape(line), tc_style))
                elements.append(Spacer(1, 1*mm))

        if legacy_tc:
            for line in legacy_tc.split('\n'):
                if line.strip():
                    elements.append(Paragraph(_xml_escape(line.strip()), tc_style))
                    elements.append(Spacer(1, 1*mm))

    try:
        doc.build(elements, canvasmaker=NumberedCanvas)
    except Exception:
        # Fallback build without page numbers if canvas decoration fails
        import logging, traceback
        logging.getLogger(__name__).warning(
            'PO PDF NumberedCanvas build failed:\n%s', traceback.format_exc()
        )
        buf.seek(0)
        buf.truncate()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
        doc.build(elements)
    buf.seek(0)

    response = HttpResponse(buf.read(), content_type='application/pdf')
    prefix = 'PO_DRAFT' if is_draft else 'PO'
    if unpriced:
        prefix += '_UNPRICED'
    filename = _safe_filename(po.po_number, prefix=prefix, extension='pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def po_export_pdf_unpriced(request, pk):
    """Unpriced PO PDF — same layout with all commercial figures removed."""
    return po_export_pdf(request, pk, unpriced=True)


# ─── Import from Excel ────────────────────────────────────────

@login_required
def po_import_excel(request):
    """Import a Purchase Order from an Excel file matching the PO template format."""
    if request.method != 'POST':
        return redirect('procurement:po_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select an Excel file.')
        return redirect('procurement:po_list')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        # Helper to read cell value
        def cell(row, col):
            v = ws.cell(row=row, column=col).value
            return str(v).strip() if v is not None else ''

        def cell_raw(row, col):
            return ws.cell(row=row, column=col).value

        # ── Parse header (rows 1-7) ──
        po_date_raw = cell_raw(1, 2)
        if isinstance(po_date_raw, datetime):
            po_date = po_date_raw.date()
        else:
            po_date = None

        po_number = cell(2, 2)
        if not po_number:
            messages.error(request, 'PO Number not found in cell B2.')
            return redirect('procurement:po_list')

        # Check for duplicate
        if PurchaseOrder.objects.filter(po_number=po_number).exists():
            messages.error(request, f'PO {po_number} already exists.')
            return redirect('procurement:po_list')

        cost_center_raw = cell(3, 2).lower()
        cost_center = 'projects'
        for code, label in PurchaseOrder.COST_CENTER_CHOICES:
            if cost_center_raw in (code, label.lower()):
                cost_center = code
                break

        # Currency is encoded in the rate/total column headers, e.g.
        # "Rate/unit (USD)". Default to SAR if not found.
        currency = 'SAR'
        rate_header = (cell(11, 8) + ' ' + cell(11, 9)).upper()
        for code, _label in PurchaseOrder.CURRENCY_CHOICES:
            if code in rate_header:
                currency = code
                break

        vendor_name = cell(4, 2)
        vendor_contact_person = cell(5, 2)
        vendor_contact_email = cell(6, 2).rstrip("'")
        vendor_contact_tel = cell(7, 2)

        po_issued_by = cell(1, 6)
        issuer_email = cell(2, 6)
        project_name = cell(3, 6)
        end_user = cell(4, 6)
        mr_item_number = cell(5, 6)
        delivery_incoterms_raw = cell(6, 6).upper().split(' ')[0] if cell(6, 6) else ''
        delivery_location = cell(7, 6)

        # Match incoterm
        delivery_incoterms = ''
        for code, label in PurchaseOrder.INCOTERM_CHOICES:
            if delivery_incoterms_raw == code:
                delivery_incoterms = code
                break

        # ── Parse line items first (row 12 until the totals section) ──
        # The number of items is variable, so NOTHING below the table sits at a
        # fixed row. We scan from row 12 to the end of the sheet (no hard cap —
        # the old code stopped at row 30 and silently dropped any item beyond
        # ~18 rows) and stop when we reach the totals / T&C section. The totals
        # and T&C scans then anchor to where the items actually ended.
        TOTALS_LABELS = ('base amount', 'discount', 'gross value', 'vat', 'total value')
        parsed_items = []
        totals_start_row = ws.max_row + 1
        for r in range(12, ws.max_row + 1):
            sn = cell_raw(r, 1)
            has_serial = isinstance(sn, (int, float))
            desc = cell(r, 3)
            make = cell(r, 2)
            qty = cell_raw(r, 6)
            rate = cell_raw(r, 8)

            # The totals / T&C section starts at a label row with no serial
            # number — stop there. Requiring no serial avoids a false match on a
            # genuine item whose description happens to contain a totals word.
            if not has_serial and (
                (desc and any(t in desc.lower() for t in TOTALS_LABELS))
                or 'TERMS AND CONDITIONS' in cell(r, 1).upper()
            ):
                totals_start_row = r
                break

            # Skip rows with no item content (e.g. the blank spacer row the
            # export writes between the items and the totals).
            if not (has_serial or desc or make or isinstance(qty, (int, float))):
                continue

            parsed_items.append({
                'serial_number': int(sn) if has_serial else len(parsed_items) + 1,
                'make_model': make,
                'description': desc,
                'quantity': Decimal(str(qty)) if isinstance(qty, (int, float)) else Decimal('0'),
                'uom': cell(r, 7) or 'Nos',
                'rate_per_unit': Decimal(str(rate)) if isinstance(rate, (int, float)) else Decimal('0'),
                'remarks': cell(r, 10),
            })

        # ── Discount & VAT — scan the totals block just below the items. ──
        discount_rate = Decimal('0')
        vat_rate = Decimal('15')
        for r in range(totals_start_row, min(totals_start_row + 8, ws.max_row + 1)):
            label = cell(r, 3).lower()
            if 'discount' in label:
                val = cell_raw(r, 6)
                if isinstance(val, (int, float)):
                    discount_rate = Decimal(str(val * 100)) if val < 1 else Decimal(str(val))
            if 'vat' in label:
                val = cell_raw(r, 6)
                if isinstance(val, (int, float)):
                    vat_rate = Decimal(str(val * 100)) if val < 1 else Decimal(str(val))

        # ── T&C — everything under the "TERMS AND CONDITIONS" marker. ──
        terms_lines = []
        tc_started = False
        for r in range(totals_start_row, ws.max_row + 1):
            a_val = cell(r, 1)
            b_val = cell(r, 2)
            if 'TERMS AND CONDITIONS' in a_val.upper():
                tc_started = True
                continue
            if tc_started and (a_val or b_val):
                line = f"{a_val} {b_val}".strip() if a_val else b_val
                if line and 'vendor acknowledgment' not in line.lower():
                    terms_lines.append(line)
                if 'vendor acknowledgment' in (a_val + b_val).lower():
                    break

        # ── Create PO ──
        po = PurchaseOrder.objects.create(
            po_date=po_date or datetime.now().date(),
            po_number=po_number,
            cost_center=cost_center,
            vendor_name=vendor_name,
            vendor_contact_person=vendor_contact_person,
            vendor_contact_email=vendor_contact_email,
            vendor_contact_tel=vendor_contact_tel,
            po_issued_by=po_issued_by,
            issuer_email=issuer_email,
            project_name=project_name,
            end_user=end_user,
            mr_item_number=mr_item_number,
            delivery_incoterms=delivery_incoterms,
            delivery_location=delivery_location,
            currency=currency,
            discount_rate=discount_rate,
            vat_rate=vat_rate,
            terms_and_conditions='\n'.join(terms_lines),
            status='draft',
            created_by=request.user,
        )

        # ── Create the parsed line items ──
        for idx, it in enumerate(parsed_items):
            PurchaseOrderItem.objects.create(purchase_order=po, order=idx, **it)
        item_count = len(parsed_items)

        messages.success(request, f'Imported PO {po.po_number} with {item_count} line items.')
        return redirect('procurement:po_detail', pk=po.pk)

    except Exception as e:
        messages.error(request, f'Error importing file: {str(e)}')
        return redirect('procurement:po_list')


# ═══════════════════════════════════════════════════════════════
# PROCUREMENT SUMMARY (Internal / External)
# ═══════════════════════════════════════════════════════════════

SUMMARY_ENTRY_DATE_FIELDS = {
    'po_plan', 'po_forecast', 'po_actual',
    'delivery_plan', 'delivery_readiness', 'delivery_forecast', 'delivery_actual',
}


def _scoped_items_for_summary(request):
    """PO line items the user can see — scoped through the parent PO."""
    user = request.user
    qs = PurchaseOrderItem.objects.select_related(
        'purchase_order', 'purchase_order__project', 'purchase_order__project__region',
        'purchase_order__created_by',
    ).all()
    if user.is_super_admin_user or user.is_procurement_user:
        return qs
    elif user.is_admin_user or user.is_manager_user:
        return qs.filter(
            Q(purchase_order__created_by=user)
            | Q(purchase_order__project__region=user.region)
        )
    return qs.filter(purchase_order__created_by=user)


def _ensure_summary_entries(items, summary_type):
    """Make sure every line item in `items` has a matching POSummaryEntry of the given type."""
    item_ids = [i.id for i in items]
    existing = {
        e.purchase_order_item_id: e
        for e in POSummaryEntry.objects.filter(
            purchase_order_item_id__in=item_ids, summary_type=summary_type
        )
    }
    missing = [i for i in items if i.id not in existing]
    if missing:
        POSummaryEntry.objects.bulk_create([
            POSummaryEntry(purchase_order_item=i, summary_type=summary_type)
            for i in missing
        ])
        existing = {
            e.purchase_order_item_id: e
            for e in POSummaryEntry.objects.filter(
                purchase_order_item_id__in=item_ids, summary_type=summary_type
            )
        }
    return existing


DATE_FIELD_ORDER = (
    'po_plan', 'po_forecast', 'po_actual',
    'delivery_plan', 'delivery_readiness', 'delivery_forecast', 'delivery_actual',
)


def _row_status(entry):
    """Classify a summary row for visual treatment.
    delivered  — Delivery Actual filled
    inprogress — Delivery Plan/Forecast set, no Actual yet
    pending    — nothing on the delivery side, or no entry at all"""
    if not entry:
        return 'pending'
    if entry.delivery_actual:
        return 'delivered'
    if entry.delivery_plan or entry.delivery_forecast or entry.delivery_readiness:
        return 'inprogress'
    return 'pending'


@login_required
def internal_summary(request):
    """Internal procurement summary — every PO line item grouped by its
    `system`, with editable date / remarks columns added per row."""
    items = list(_scoped_items_for_summary(request).order_by(
        'system', 'purchase_order__po_number', 'serial_number',
    ))
    entries_by_item = _ensure_summary_entries(items, 'internal')

    from collections import OrderedDict
    groups = OrderedDict()
    for it in items:
        key = (it.system or '').strip() or '— Unassigned —'
        groups.setdefault(key, []).append({
            'item': it,
            'entry': entries_by_item.get(it.id),
        })

    rows = []
    group_summary = []
    sn = 0
    total_delivered = 0
    total_inprogress = 0
    total_pending = 0
    for group_key, members in groups.items():
        g_delivered = g_inprogress = g_pending = 0
        for idx, m in enumerate(members):
            sn += 1
            entry = m['entry']
            status = _row_status(entry)
            if status == 'delivered':
                g_delivered += 1
                total_delivered += 1
            elif status == 'inprogress':
                g_inprogress += 1
                total_inprogress += 1
            else:
                g_pending += 1
                total_pending += 1
            date_cells = []
            for field in DATE_FIELD_ORDER:
                val = getattr(entry, field, None) if entry else None
                date_cells.append({
                    'field': field,
                    'iso': val.strftime('%Y-%m-%d') if val else '',
                    'display': val.strftime('%d %b %Y') if val else '',
                    'empty': val is None,
                })
            rows.append({
                'group_key': group_key,
                'is_first_in_group': idx == 0,
                'group_size': len(members),
                'sn': sn,
                'item': m['item'],
                'po': m['item'].purchase_order,
                'entry': entry,
                'date_cells': date_cells,
                'status': status,
            })
        group_summary.append({
            'key': group_key,
            'count': len(members),
            'delivered': g_delivered,
            'inprogress': g_inprogress,
            'pending': g_pending,
        })

    return render(request, 'procurement/summary_internal.html', {
        'rows': rows,
        'po_count': len(items),
        'group_count': len(groups),
        'group_summary': group_summary,
        'totals': {
            'delivered': total_delivered,
            'inprogress': total_inprogress,
            'pending': total_pending,
        },
        'summary_type': 'internal',
    })


@login_required
@require_POST
def ajax_summary_entry_update(request, pk):
    """AJAX update of one date or remarks cell on a POSummaryEntry."""
    entry = get_object_or_404(
        POSummaryEntry.objects.select_related('purchase_order_item__purchase_order'),
        pk=pk,
    )
    po = entry.purchase_order_item.purchase_order

    user = request.user
    can_edit = (
        user.is_super_admin_user or user.is_admin_user or user.is_procurement_user
        or po.created_by_id == user.id
    )
    if not can_edit:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    field = request.POST.get('field', '').strip()
    raw = (request.POST.get('value') or '').strip()

    if field == 'remarks':
        entry.remarks = raw
        entry.save(update_fields=['remarks', 'updated_at'])
        return JsonResponse({'ok': True, 'value': entry.remarks})

    if field not in SUMMARY_ENTRY_DATE_FIELDS:
        return JsonResponse({'error': f'Unknown field {field!r}'}, status=400)

    if not raw:
        parsed = None
    else:
        from datetime import datetime
        parsed = None
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                parsed = datetime.strptime(raw, fmt).date()
                break
            except ValueError:
                continue
        if parsed is None:
            return JsonResponse({'error': f'Unparseable date: {raw!r}'}, status=400)

    setattr(entry, field, parsed)
    entry.save(update_fields=[field, 'updated_at'])
    return JsonResponse({
        'ok': True,
        'value': parsed.isoformat() if parsed else '',
        'display': parsed.strftime('%d %b %Y') if parsed else '',
    })


# ─── External Summary ────────────────────────────────────────

# Per-line-item fields editable inline from the External summary table.
# Maps the cell `data-field` -> a tuple of (model attribute, parser).
def _to_decimal_or_none(raw):
    if raw is None or str(raw).strip() == '':
        return None
    return Decimal(str(raw))


PO_ITEM_INLINE_FIELDS = {
    'po_value_usd': ('po_value_usd', _to_decimal_or_none),
    'advance_payment_sar': ('advance_payment_sar', _to_decimal_or_none),
    'delivery_status': ('delivery_status', lambda v: ('' if v is None else str(v))),
    'scm': ('scm', lambda v: ('' if v is None else str(v))[:10]),
}

# PO-level fields editable inline (when the same value applies to every
# line item under that PO — e.g. lead time, payment terms, warranty).
PO_HEADER_INLINE_FIELDS = {
    'lead_time': 'lead_time',
    'payment_terms_text': 'payment_terms_text',
    'warranty': 'warranty',
}


@login_required
@require_POST
def ajax_po_item_field_update(request, pk):
    """AJAX update of a single field on a PurchaseOrderItem.
    Used by the External summary's inline-editable cells."""
    item = get_object_or_404(
        PurchaseOrderItem.objects.select_related('purchase_order'), pk=pk,
    )
    po = item.purchase_order

    user = request.user
    can_edit = (
        user.is_super_admin_user or user.is_admin_user or user.is_procurement_user
        or po.created_by_id == user.id
    )
    if not can_edit:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    field = request.POST.get('field', '').strip()
    raw = request.POST.get('value', '')

    # PO-level fields go on the parent PO record
    if field in PO_HEADER_INLINE_FIELDS:
        attr = PO_HEADER_INLINE_FIELDS[field]
        setattr(po, attr, (raw or '').strip())
        po.save(update_fields=[attr, 'updated_at'])
        return JsonResponse({'ok': True, 'value': getattr(po, attr) or ''})

    if field not in PO_ITEM_INLINE_FIELDS:
        return JsonResponse({'error': f'Unknown field {field!r}'}, status=400)

    attr, parser = PO_ITEM_INLINE_FIELDS[field]
    try:
        parsed = parser(raw)
    except (InvalidOperation, ValueError):
        return JsonResponse({'error': f'Invalid value for {field!r}'}, status=400)

    setattr(item, attr, parsed)
    item.save(update_fields=[attr])

    if isinstance(parsed, Decimal):
        display = f'{parsed:,.2f}'
    else:
        display = parsed or ''
    return JsonResponse({
        'ok': True,
        'value': str(parsed) if parsed not in (None, '') else '',
        'display': display,
    })


def _po_qty_text(item):
    """Format quantity + uom like the sample: '4 Nos'."""
    q = item.quantity
    if q is None:
        return ''
    # Drop trailing zeros for whole numbers
    qstr = f'{q:.0f}' if q == q.to_integral() else f'{q:,.2f}'
    return f'{qstr} {item.uom}'.strip()


def _build_external_rows(items):
    """Pre-compute display data for each row of the External summary."""
    rows = []
    sn = 0
    for it in items:
        sn += 1
        po = it.purchase_order
        item_label_parts = []
        if it.system:
            item_label_parts.append(it.system)
        if it.description:
            item_label_parts.append(it.description)
        item_label = ' — '.join(item_label_parts) if item_label_parts else (it.description or '—')

        rows.append({
            'sn': sn,
            'item': it,
            'po': po,
            'item_label': item_label,
            'po_value_sar': it.total_value,
            'qty_text': _po_qty_text(it),
        })
    return rows


def _external_totals(items):
    """Sum SAR / USD / Advance Payment across the visible items."""
    total_sar = Decimal('0')
    total_usd = Decimal('0')
    total_advance = Decimal('0')
    for it in items:
        total_sar += it.total_value or Decimal('0')
        if it.po_value_usd:
            total_usd += it.po_value_usd
        if it.advance_payment_sar:
            total_advance += it.advance_payment_sar
    return {
        'sar': total_sar,
        'usd': total_usd,
        'advance': total_advance,
    }


def _external_projects_for_user(request):
    """Distinct projects with at least one PO line item the user can see,
    plus a sentinel 'all' option for the Daily-Status-style view."""
    items = _scoped_items_for_summary(request)
    project_ids = (
        items.exclude(purchase_order__project__isnull=True)
        .values_list('purchase_order__project_id', flat=True).distinct()
    )
    from projects.models import Project
    projects = list(Project.objects.filter(pk__in=list(project_ids))
                    .order_by('project_name'))
    return projects


@login_required
def external_summary(request):
    """External procurement summary — line items grouped by Project, with
    procurement-tracking columns. Filterable to a single project (matching
    the sample's per-package tabs); 'all' shows the Daily-Status view."""
    items_qs = _scoped_items_for_summary(request)
    selected = request.GET.get('project', '').strip()  # 'all' or a project pk
    if selected and selected != 'all':
        try:
            items_qs = items_qs.filter(purchase_order__project_id=int(selected))
        except (TypeError, ValueError):
            selected = ''
    if not selected:
        # Default to "all" if no filter passed.
        selected = 'all'

    items = list(items_qs.order_by(
        'purchase_order__project__project_name',
        'purchase_order__po_number',
        'serial_number',
    ))
    rows = _build_external_rows(items)
    totals = _external_totals(items)
    projects = _external_projects_for_user(request)

    selected_project = None
    if selected != 'all':
        for p in projects:
            if str(p.pk) == selected:
                selected_project = p
                break

    return render(request, 'procurement/summary_external.html', {
        'rows': rows,
        'totals': totals,
        'projects': projects,
        'selected': selected,
        'selected_project': selected_project,
        'show_project_column': selected == 'all',
        'po_count': len(items),
    })


@login_required
def external_summary_export(request):
    """Excel export of the external summary, matching the sample layout."""
    items_qs = _scoped_items_for_summary(request)
    selected = request.GET.get('project', '').strip()
    if selected and selected != 'all':
        try:
            items_qs = items_qs.filter(purchase_order__project_id=int(selected))
        except (TypeError, ValueError):
            selected = 'all'

    items = list(items_qs.order_by(
        'purchase_order__project__project_name',
        'purchase_order__po_number',
        'serial_number',
    ))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'External Summary'

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    bold = Font(bold=True)
    bold_red = Font(bold=True, color='C41E3A')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='top', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')
    head_fill = PatternFill('solid', fgColor='F5D7DC')
    total_fill = PatternFill('solid', fgColor='FBE8EC')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    show_project_col = (selected == 'all')

    # Project header row
    project_label = 'All projects (Daily Status)'
    if not show_project_col:
        from projects.models import Project
        try:
            p = Project.objects.get(pk=int(selected))
            project_label = p.project_name
        except (Project.DoesNotExist, ValueError):
            pass
    ws.cell(row=1, column=2, value='Project:').font = bold
    ws.cell(row=1, column=3, value=project_label).font = bold

    # Header row
    headers = [
        (1, 'Sr.'),
        (2, 'System / Item'),
        (3, 'PO Number'),
        (4, 'PO Status'),
        (5, 'PO QTY'),
        (6, 'Supplier'),
        (7, 'Lead Time'),
        (8, 'Incoterm'),
        (9, 'Payment Terms'),
        (10, 'Warranty'),
        (11, 'PO Value (SAR)'),
        (12, 'PO Value (USD/EUR)'),
        (13, 'Advance Payment'),
        (14, 'Delivery Status'),
        (15, 'SCM'),
    ]
    if show_project_col:
        headers.append((16, 'Project'))

    for col, label in headers:
        c = ws.cell(row=3, column=col, value=label)
        c.font = bold
        c.alignment = center
        c.fill = head_fill
        c.border = border

    row = 4
    sn = 0
    for it in items:
        sn += 1
        po = it.purchase_order
        item_label = ' — '.join([p for p in [it.system, it.description] if p]) or (it.description or '')

        cells = [
            (1, sn),
            (2, item_label),
            (3, po.po_number),
            (4, po.get_status_display()),
            (5, _po_qty_text(it)),
            (6, po.vendor_name),
            (7, po.lead_time or ''),
            (8, po.delivery_incoterms or ''),
            (9, po.payment_terms_text or ''),
            (10, po.warranty or ''),
            (11, float(it.total_value) if it.total_value else 0),
            (12, float(it.po_value_usd) if it.po_value_usd else None),
            (13, float(it.advance_payment_sar) if it.advance_payment_sar else None),
            (14, it.delivery_status or ''),
            (15, it.scm or ''),
        ]
        if show_project_col:
            cells.append((16, po.project.project_name if po.project_id else (po.project_name or '')))

        for col, val in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.border = border
            if col in (1, 4, 5, 8, 15, 16):
                c.alignment = center
            elif col in (11, 12, 13):
                c.alignment = right
                c.number_format = '#,##0.00'
            else:
                c.alignment = left
        row += 1

    # Total row
    totals = _external_totals(items)
    total_row = row
    for col in range(1, max(c for c, _ in headers) + 1):
        c = ws.cell(row=total_row, column=col)
        c.fill = total_fill
        c.border = border
    ws.cell(row=total_row, column=9, value='Total').font = bold_red
    ws.cell(row=total_row, column=9).alignment = right
    ws.cell(row=total_row, column=11, value=float(totals['sar'])).font = bold_red
    ws.cell(row=total_row, column=11).alignment = right
    ws.cell(row=total_row, column=11).number_format = '#,##0.00'
    ws.cell(row=total_row, column=12, value=float(totals['usd'])).font = bold_red
    ws.cell(row=total_row, column=12).alignment = right
    ws.cell(row=total_row, column=12).number_format = '#,##0.00'
    ws.cell(row=total_row, column=13, value=float(totals['advance'])).font = bold_red
    ws.cell(row=total_row, column=13).alignment = right
    ws.cell(row=total_row, column=13).number_format = '#,##0.00'

    widths = {1: 5, 2: 32, 3: 22, 4: 14, 5: 14, 6: 22, 7: 22, 8: 12,
              9: 22, 10: 18, 11: 14, 12: 14, 13: 14, 14: 28, 15: 8, 16: 18}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    suffix = 'all' if show_project_col else f'project_{selected}'
    filename = _safe_filename('external_summary', suffix=suffix, extension='xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def internal_summary_export(request):
    """Excel export of the internal summary, matching the sample layout."""
    items = list(_scoped_items_for_summary(request).order_by(
        'system', 'purchase_order__po_number', 'serial_number',
    ))
    entries_by_item = _ensure_summary_entries(items, 'internal')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Internal Summary'

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    head_fill = PatternFill('solid', fgColor='F5D7DC')
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Layout (1-indexed):
    # B=2 SN | C=3 ITEM DESCRIPTION | D=4 SUB ITEM DESCRIPTION | E=5 PO # |
    # F=6 VENDOR | G=7 PO Issuance | H-J = PO Plan/Forecast/Actual |
    # K-N = Delivery Plan/Readiness/Forecast/Actual | O=15 Remarks
    ws.cell(row=1, column=2, value='SN').font = bold
    ws.cell(row=1, column=3, value='ITEM DESCRIPTION').font = bold
    ws.cell(row=1, column=4, value='SUB ITEM DESCRIPTION').font = bold
    ws.cell(row=1, column=5, value='PO #').font = bold
    ws.cell(row=1, column=6, value='VENDOR').font = bold
    ws.cell(row=1, column=7, value='PO Issuance').font = bold

    group_headers = [
        ('PO', 8, 3),
        ('DELIVERY', 11, 4),
    ]
    for label, col, span in group_headers:
        c = ws.cell(row=1, column=col, value=label)
        c.font = bold
        c.alignment = center
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)

    ws.cell(row=1, column=15, value='Remarks').font = bold

    sub_headers = [
        (8, 'PLAN'), (9, 'FORECAST'), (10, 'ACTUAL'),
        (11, 'PLAN'), (12, 'READINESS'), (13, 'FORECAST'), (14, 'ACTUAL'),
    ]
    for col, label in sub_headers:
        c = ws.cell(row=2, column=col, value=label)
        c.font = bold
        c.alignment = center
        c.fill = head_fill

    for col in [2, 3, 4, 5, 6, 7, 15] + [c for c, _ in sub_headers]:
        ws.cell(row=1, column=col).fill = head_fill
        ws.cell(row=1, column=col).alignment = center
    for col in range(2, 16):
        ws.cell(row=1, column=col).border = border
        ws.cell(row=2, column=col).border = border

    row = 3
    sn = 0
    last_system = None
    for item in items:
        po = item.purchase_order
        sn += 1
        entry = entries_by_item.get(item.id)
        sys_label = (item.system or '').strip() or '— Unassigned —'
        ws.cell(row=row, column=2, value=sn).alignment = center
        ws.cell(row=row, column=3, value=sys_label if sys_label != last_system else '').alignment = left
        ws.cell(row=row, column=4, value=item.description or '').alignment = left
        ws.cell(row=row, column=5, value=po.po_number).alignment = left
        ws.cell(row=row, column=6, value=po.vendor_name).alignment = left
        ws.cell(row=row, column=7, value=po.get_status_display()).alignment = left
        last_system = sys_label

        if entry:
            date_cells = [
                (8, entry.po_plan), (9, entry.po_forecast), (10, entry.po_actual),
                (11, entry.delivery_plan), (12, entry.delivery_readiness),
                (13, entry.delivery_forecast), (14, entry.delivery_actual),
            ]
            for col, val in date_cells:
                if val:
                    c = ws.cell(row=row, column=col, value=val)
                    c.number_format = 'dd-mmm-yyyy'
                    c.alignment = center
            ws.cell(row=row, column=15, value=entry.remarks or '').alignment = left

        for col in range(2, 16):
            ws.cell(row=row, column=col).border = border
        row += 1

    widths = {2: 5, 3: 22, 4: 30, 5: 25, 6: 18, 7: 14, 15: 35}
    for c in range(8, 15):
        widths[c] = 12
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = _safe_filename('internal_summary', extension='xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ═══════════════════════════════════════════════════════════════
# DELIVERY NOTE
# ═══════════════════════════════════════════════════════════════

class DNPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    def get_queryset(self):
        queryset = DeliveryNote.objects.select_related('project', 'created_by', 'purchase_order').all()
        user = self.request.user
        if user.is_super_admin_user or user.is_procurement_user:
            return queryset
        elif user.is_admin_user or user.is_manager_user:
            return queryset.filter(Q(created_by=user) | Q(project__region=user.region))
        else:
            return queryset.filter(created_by=user)


class DNListView(CapabilityRequiredMixin, DNPermissionMixin, ListView):
    capability = 'dn.access'
    model = DeliveryNote
    template_name = 'procurement/dn_list.html'
    context_object_name = 'delivery_notes'
    paginate_by = 25

    def test_func(self):
        return True

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(dn_number__icontains=search) |
                Q(sold_to_company__icontains=search) |
                Q(project_title__icontains=search) |
                Q(leap_po_number__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        return context


class DNCreateView(DNPermissionMixin, CreateView):
    model = DeliveryNote
    form_class = DeliveryNoteForm
    template_name = 'procurement/dn_form.html'

    def test_func(self):
        return True

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = DNItemFormSet(self.request.POST, prefix='items')
        else:
            context['item_formset'] = DNItemFormSet(prefix='items')
        context['title'] = 'Create Delivery Note'
        # POs offered in the "build from PO" picker — region-scoped, only
        # POs that actually have line items.
        pos = (
            PurchaseOrder.objects
            .select_related('project')
            .annotate(item_count=Count('items'))
            .filter(item_count__gt=0)
        )
        user = self.request.user
        if not user.is_super_admin_user:
            pos = pos.filter(Q(project__region=user.region) | Q(project__isnull=True))
        context['available_pos'] = pos.order_by('-po_date', '-id')
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if item_formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            item_formset.instance = self.object
            item_formset.save()
            messages.success(self.request, 'Delivery Note created.')
            return redirect(self.get_success_url())
        return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:dn_detail', kwargs={'pk': self.object.pk})


@login_required
def dn_create_from_po(request, po_pk):
    """Selection screen + DN seeder for building a Delivery Note from a PO.

    GET  — lists every PO line item; items already on a DN are read-only
           with back-links, the rest are pickable checkboxes.
    POST — creates a draft DN prefilled from the PO header, copies the
           picked items in as editable DN rows (source_po_item set so they
           can't be re-picked for a later DN), then drops the user on the
           DN edit screen.

    Multi-DN from one PO works naturally: subsequent visits show only the
    still-undelivered items as checkboxes.
    """
    po = get_object_or_404(
        PurchaseOrder.objects.select_related('project', 'project__region'),
        pk=po_pk,
    )
    user = request.user

    # Region scope — non-super-admins only touch POs in their region.
    if (
        not user.is_super_admin_user
        and po.project
        and po.project.region_id != user.region_id
    ):
        messages.error(request, 'You can only create delivery notes for POs in your region.')
        return redirect('procurement:po_detail', pk=po.pk)

    po_items = list(po.items.prefetch_related('delivered_dn_items__delivery_note'))
    if not po_items:
        messages.error(request, 'This PO has no line items to deliver.')
        return redirect('procurement:po_detail', pk=po.pk)

    if request.method == 'POST':
        item_ids = [int(x) for x in request.POST.getlist('item_ids') if x.isdigit()]
        if not item_ids:
            messages.error(request, 'Pick at least one item to add to the delivery note.')
        else:
            dn = DeliveryNote.objects.create(
                sold_to_company=po.end_user or '',
                delivery_address=po.delivery_location or '',
                date=datetime.now().date(),
                project_title=po.project_name or (po.project.project_name if po.project else ''),
                leap_po_number=po.po_number,
                project=po.project,
                purchase_order=po,
                created_by=user,
            )
            serial = 1
            for pi in po.items.filter(pk__in=item_ids):
                # Defensive: skip items claimed by another DN between the
                # page render and this POST.
                if pi.delivered_dn_items.exists():
                    continue
                DeliveryNoteItem.objects.create(
                    delivery_note=dn,
                    serial_number=serial,
                    make_part_number=pi.make_model,
                    description=pi.description,
                    quantity=pi.quantity,
                    uom=pi.uom or 'Each',
                    remarks=pi.remarks,
                    order=serial,
                    source_po_item=pi,
                )
                serial += 1
            copied = serial - 1
            if copied == 0:
                messages.warning(
                    request,
                    'Every item you picked is already on a delivery note — nothing to add.',
                )
                dn.delete()
                return redirect('procurement:dn_create_from_po', po_pk=po.pk)
            messages.success(
                request,
                f'Delivery Note seeded with {copied} item{"" if copied == 1 else "s"} from PO {po.po_number}. '
                f'Fill in the Sold-To details, adjust quantities if partial, then save.',
            )
            return redirect('procurement:dn_update', pk=dn.pk)

    rows = []
    available_count = 0
    delivered_count = 0
    for pi in po_items:
        delivered = list(pi.delivered_dn_items.all())
        if delivered:
            delivered_count += 1
        else:
            available_count += 1
        rows.append({
            'item': pi,
            'delivered_in': delivered,
            'is_available': not delivered,
        })

    return render(request, 'procurement/dn_select_po_items.html', {
        'po': po,
        'rows': rows,
        'total_count': len(po_items),
        'available_count': available_count,
        'delivered_count': delivered_count,
    })


class DNDetailView(DNPermissionMixin, DetailView):
    model = DeliveryNote
    template_name = 'procurement/dn_detail.html'
    context_object_name = 'dn'

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.all()
        return context


class DNUpdateView(DNPermissionMixin, UpdateView):
    model = DeliveryNote
    form_class = DeliveryNoteForm
    template_name = 'procurement/dn_form.html'

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return self.get_object().created_by == user

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = DNItemFormSet(self.request.POST, instance=self.object, prefix='items')
        else:
            context['item_formset'] = DNItemFormSet(instance=self.object, prefix='items')
        context['title'] = f'Edit DN: {self.object.dn_number or self.object.pk}'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if item_formset.is_valid():
            self.object = form.save()
            item_formset.instance = self.object
            item_formset.save()
            messages.success(self.request, 'Delivery Note updated.')
            return redirect(self.get_success_url())
        return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:dn_detail', kwargs={'pk': self.object.pk})


class DNDeleteView(DNPermissionMixin, DeleteView):
    model = DeliveryNote
    template_name = 'procurement/dn_confirm_delete.html'
    success_url = reverse_lazy('procurement:dn_list')

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return self.get_object().created_by == user

    def form_valid(self, form):
        messages.success(self.request, 'Delivery Note deleted.')
        return super().form_valid(form)


# ─── DN Export Excel ──────────────────────────────────────────

@login_required
def dn_export_excel(request, pk):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    dn = get_object_or_404(DeliveryNote, pk=pk)
    items = dn.items.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DELIVERY NOTE'

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap = Alignment(wrap_text=True, vertical='top')

    # Title
    ws.merge_cells('A1:F1')
    ws.cell(row=1, column=1, value='DELIVERY NOTE').font = Font(bold=True, size=14, color='1F4E79')

    # Header info
    info = [
        (3, 'Sold To:', dn.sold_to_company, 'Project Title:', dn.project_title),
        (4, 'Address:', dn.sold_to_address, 'LEAP PO No:', dn.leap_po_number),
        (5, 'Delivery Address:', dn.delivery_address, 'Client PO No:', dn.client_po_number),
        (6, 'Attention:', dn.attention, 'DN Number:', dn.dn_number),
        (7, 'Mobile:', dn.mobile, '', ''),
        (8, 'Email:', dn.email, '', ''),
        (9, 'Date:', dn.date.strftime('%d-%b-%Y') if dn.date else '', '', ''),
    ]
    for r, l1, v1, l2, v2 in info:
        ws.cell(row=r, column=1, value=l1).font = bold
        c = ws.cell(row=r, column=2, value=str(v1))
        c.alignment = wrap
        if l2:
            ws.cell(row=r, column=4, value=l2).font = bold
            ws.cell(row=r, column=5, value=str(v2))

    # Items table
    row = 11
    headers = ['S No.', 'Make/Part Number', 'Description', 'Quantity', 'UOM', 'Remarks / LNA PO# ref']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for item in items:
        row += 1
        data = [item.serial_number, item.make_part_number, item.description,
                float(item.quantity), item.uom, item.remarks]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = wrap

    # Footer
    row += 2
    ws.cell(row=row, column=1, value='ALL CLAIMS FOR DAMAGES AND SHORTAGES OF GOODS MUST BE MADE UPON RECEIPT OF GOODS.').font = Font(bold=True, size=8)
    row += 2
    ws.cell(row=row, column=1, value='GOODS RECEIVED IN GOOD ORDER BY:').font = bold
    ws.cell(row=row, column=4, value='ITEM VERIFIED/AUTHORIZED BY').font = bold
    row += 3
    ws.cell(row=row, column=1, value='____________________________________________')
    ws.cell(row=row, column=4, value='___________________________________')
    row += 1
    ws.cell(row=row, column=1, value='(SIGNATURE OF RECEIVER)')
    ws.cell(row=row, column=4, value='SIGNATURE / DATE')

    widths = [8, 25, 40, 10, 8, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raw = dn.dn_number or f'DN-{dn.pk}'
    filename = _safe_filename(raw, extension='xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── DN Export PDF ────────────────────────────────────────────

@login_required
def dn_export_pdf(request, pk):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from django.contrib.staticfiles.finders import find as find_static

    dn = get_object_or_404(DeliveryNote, pk=pk)
    items = dn.items.all()

    NumberedCanvas = _make_numbered_canvas()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('DNTitle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#C41E3A'), alignment=TA_CENTER, spaceAfter=4)
    label_style = ParagraphStyle('Lbl', parent=styles['Normal'], fontSize=8, textColor=colors.grey)
    value_style = ParagraphStyle('Val', parent=styles['Normal'], fontSize=9, fontName='Helvetica-Bold')
    small = ParagraphStyle('Sm', parent=styles['Normal'], fontSize=8)
    small_bold = ParagraphStyle('SmB', parent=styles['Normal'], fontSize=8, fontName='Helvetica-Bold')

    # Logo
    logo_path = find_static('images/leap_logo.jpg')
    if logo_path:
        from reportlab.platypus import Image
        logo = Image(logo_path, width=50*mm, height=15*mm)
        elements.append(logo)

    elements.append(Paragraph('DELIVERY NOTE', title_style))
    elements.append(Spacer(1, 4*mm))

    # Header info table
    def lv(label, value):
        return [Paragraph(label, label_style), Paragraph(str(value or '-'), value_style)]

    header_data = [
        lv('Sold To:', dn.sold_to_company) + [''] + lv('Project Title:', dn.project_title),
        lv('Address:', dn.sold_to_address) + [''] + lv('LEAP PO No:', dn.leap_po_number),
        lv('Delivery Address:', dn.delivery_address) + [''] + lv('Client PO No:', dn.client_po_number),
        lv('Attention:', dn.attention) + [''] + lv('DN Number:', dn.dn_number),
        lv('Mobile:', dn.mobile) + [''] + ['', ''],
        lv('Email:', dn.email) + [''] + ['', ''],
        lv('Date:', dn.date.strftime('%d %b %Y') if dn.date else '-') + [''] + ['', ''],
    ]
    ht = Table(header_data, colWidths=[22*mm, 60*mm, 5*mm, 22*mm, 60*mm])
    ht.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#D0D0D0')),
        # Vertical lines: only between label↔value pairs (left and right groups).
        ('LINEAFTER', (0, 0), (0, -1), 0.4, colors.HexColor('#E0E0E0')),
        ('LINEAFTER', (3, 0), (3, -1), 0.4, colors.HexColor('#E0E0E0')),
        ('LINEBELOW', (0, 0), (-1, -2), 0.4, colors.HexColor('#E0E0E0')),
    ]))
    elements.append(ht)
    elements.append(Spacer(1, 5*mm))

    # Items table
    dark_blue = colors.HexColor('#C41E3A')
    col_widths = [12*mm, 35*mm, 60*mm, 18*mm, 15*mm, 35*mm]
    item_header = [
        Paragraph('<b>S.No.</b>', small), Paragraph('<b>Make/Part No.</b>', small),
        Paragraph('<b>Description</b>', small), Paragraph('<b>Qty</b>', small),
        Paragraph('<b>UOM</b>', small), Paragraph('<b>Remarks</b>', small),
    ]
    data = [item_header]
    for item in items:
        data.append([
            Paragraph(str(item.serial_number), small),
            Paragraph(item.make_part_number or '', small),
            Paragraph(item.description, small),
            Paragraph(f'{item.quantity:,.0f}', small),
            Paragraph(item.uom, small),
            Paragraph(item.remarks or '', small),
        ])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5D7DC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#495057')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#C41E3A')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6*mm))

    # Footer
    elements.append(Paragraph(
        '<b>ALL CLAIMS FOR DAMAGES AND SHORTAGES OF GOODS MUST BE MADE UPON RECEIPT OF GOODS.</b>',
        ParagraphStyle('Notice', parent=styles['Normal'], fontSize=7, leading=9)
    ))
    elements.append(Spacer(1, 8*mm))

    sig_data = [
        ['GOODS RECEIVED IN GOOD ORDER BY:', '', 'ITEM VERIFIED/AUTHORIZED BY:'],
        ['', '', 'LEAP NETWORKS'],
        ['', '', ''],
        ['____________________________________________', '', '___________________________________'],
        ['(SIGNATURE OF RECEIVER)', '', 'SIGNATURE / DATE'],
        ['DATE:', '', ''],
        ['Name:', '', ''],
    ]
    sig_table = Table(sig_data, colWidths=[75*mm, 20*mm, 75*mm])
    sig_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(sig_table)

    # Leap Networks Supply Chain Management stamp at the bottom-right
    stamp_path = find_static('images/leap_stamp.png')
    if stamp_path:
        from reportlab.platypus import Image
        elements.append(Spacer(1, 4*mm))
        try:
            stamp_img = Image(stamp_path, width=40*mm, height=30*mm)
            # Right-align: empty cell on the left so the stamp sits under the
            # "ITEM VERIFIED/AUTHORIZED BY" column.
            stamp_table = Table([['', '', stamp_img]], colWidths=[75*mm, 20*mm, 75*mm])
            stamp_table.setStyle(TableStyle([
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ALIGN', (2, 0), (2, 0), 'CENTER'),
            ]))
            elements.append(stamp_table)
        except Exception:
            # Stamp drawing should never block the PDF
            pass

    try:
        doc.build(elements, canvasmaker=NumberedCanvas)
    except Exception:
        import logging, traceback
        logging.getLogger(__name__).warning(
            'DN PDF NumberedCanvas build failed:\n%s', traceback.format_exc()
        )
        buf.seek(0)
        buf.truncate()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=15*mm, bottomMargin=15*mm, leftMargin=15*mm, rightMargin=15*mm)
        doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    raw = dn.dn_number or f'DN-{dn.pk}'
    filename = _safe_filename(raw, extension='pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


# ─── DN Import Excel ──────────────────────────────────────────

@login_required
def dn_import_excel(request):
    if request.method != 'POST':
        return redirect('procurement:dn_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select an Excel file.')
        return redirect('procurement:dn_list')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb.active

        def cv(r, c):
            v = ws.cell(row=r, column=c).value
            return str(v).strip() if v is not None else ''

        def cv_raw(r, c):
            return ws.cell(row=r, column=c).value

        # Parse header
        sold_to = cv(7, 2) or cv(7, 3)
        sold_to_address = cv(8, 2)
        delivery_address = cv(9, 2)
        attention = cv(11, 2)
        mobile = cv(12, 2)
        email = cv(13, 2)
        date_raw = cv_raw(14, 2)
        dn_date = date_raw.date() if isinstance(date_raw, datetime) else None

        project_title = cv(11, 5) or cv(9, 5)
        leap_po = cv(13, 5)
        client_po = cv(14, 5)
        dn_number = cv(15, 5)

        dn_obj = DeliveryNote.objects.create(
            sold_to_company=sold_to,
            sold_to_address=sold_to_address,
            delivery_address=delivery_address,
            attention=attention,
            mobile=mobile,
            email=email,
            date=dn_date or datetime.now().date(),
            project_title=project_title,
            leap_po_number=leap_po,
            client_po_number=client_po,
            dn_number=dn_number,
            created_by=request.user,
        )

        # Parse items (find header row first)
        header_row = None
        for r in range(1, ws.max_row + 1):
            val = cv(r, 1).lower()
            if 's no' in val or 'sr' in val:
                header_row = r
                break

        item_count = 0
        if header_row:
            for r in range(header_row + 1, ws.max_row + 1):
                sn = cv_raw(r, 1)
                desc = cv(r, 3)
                if not desc:
                    if cv(r, 1).upper().startswith('ALL CLAIMS'):
                        break
                    continue
                if not isinstance(sn, (int, float)):
                    continue

                qty_raw = cv_raw(r, 4)
                DeliveryNoteItem.objects.create(
                    delivery_note=dn_obj,
                    serial_number=int(sn),
                    make_part_number=cv(r, 2),
                    description=desc,
                    quantity=Decimal(str(qty_raw)) if isinstance(qty_raw, (int, float)) else Decimal('1'),
                    uom=cv(r, 5) or 'Each',
                    remarks=cv(r, 6),
                    order=item_count,
                )
                item_count += 1

        messages.success(request, f'Imported Delivery Note with {item_count} items.')
        return redirect('procurement:dn_detail', pk=dn_obj.pk)

    except Exception as e:
        messages.error(request, f'Error importing: {str(e)}')
        return redirect('procurement:dn_list')


# ═══════════════════════════════════════════════════════════════
# INVENTORY REPORT
# ═══════════════════════════════════════════════════════════════

class InventoryPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    def get_queryset(self):
        queryset = InventoryReport.objects.select_related('project', 'created_by').all()
        user = self.request.user
        if user.is_super_admin_user or user.is_procurement_user:
            return queryset
        elif user.is_admin_user or user.is_manager_user:
            return queryset.filter(Q(created_by=user) | Q(project__region=user.region))
        else:
            return queryset.filter(created_by=user)


class InventoryListView(InventoryPermissionMixin, ListView):
    model = InventoryReport
    template_name = 'procurement/inventory_list.html'
    context_object_name = 'reports'
    paginate_by = 25

    def test_func(self):
        return True

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(warehouse_location__icontains=search)
            )
        # Annotate item count, low-stock count, and total stock value in a
        # single SQL query so the list page does not run several per-row
        # queries (was N+1 across items.count, low_stock_count property,
        # and total_stock_value property).
        # Note: annotation names cannot start with underscore - Django
        # templates reject variables that begin with one.
        from django.db.models import Q as _Q, DecimalField, ExpressionWrapper
        return queryset.annotate(
            annotated_item_count=Count('items', distinct=True),
            annotated_low_stock_count=Count(
                'items',
                filter=_Q(
                    items__min_stock_level__isnull=False,
                    items__min_stock_level__gt=0,
                    items__balance_qty__isnull=False,
                    items__balance_qty__lte=F('items__min_stock_level'),
                ),
                distinct=True,
            ),
            annotated_stock_value=Sum(
                ExpressionWrapper(
                    F('items__unit_cost') * F('items__balance_qty'),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                ),
            ),
        ).order_by('-updated_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        return context


class InventoryCreateView(InventoryPermissionMixin, CreateView):
    model = InventoryReport
    form_class = InventoryReportForm
    template_name = 'procurement/inventory_form.html'

    def test_func(self):
        return True

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = InventoryItemFormSet(self.request.POST, prefix='items')
        else:
            context['item_formset'] = InventoryItemFormSet(prefix='items')
        context['title'] = 'Create Inventory Report'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if item_formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            item_formset.instance = self.object
            item_formset.save()
            messages.success(self.request, 'Inventory Report created.')
            return redirect(self.get_success_url())
        return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:inventory_detail', kwargs={'pk': self.object.pk})


class InventoryDetailView(InventoryPermissionMixin, DetailView):
    model = InventoryReport
    template_name = 'procurement/inventory_detail.html'
    context_object_name = 'report'

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['items'] = self.object.items.all()
        return context


class InventoryUpdateView(InventoryPermissionMixin, UpdateView):
    model = InventoryReport
    form_class = InventoryReportForm
    template_name = 'procurement/inventory_form.html'

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return self.get_object().created_by == user

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['item_formset'] = InventoryItemFormSet(self.request.POST, instance=self.object, prefix='items')
        else:
            context['item_formset'] = InventoryItemFormSet(instance=self.object, prefix='items')
        context['title'] = f'Edit: {self.object.title}'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        item_formset = context['item_formset']
        if item_formset.is_valid():
            self.object = form.save()
            item_formset.instance = self.object
            item_formset.save()
            messages.success(self.request, 'Inventory Report updated.')
            return redirect(self.get_success_url())
        return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:inventory_detail', kwargs={'pk': self.object.pk})


class InventoryDeleteView(InventoryPermissionMixin, DeleteView):
    model = InventoryReport
    template_name = 'procurement/inventory_confirm_delete.html'
    success_url = reverse_lazy('procurement:inventory_list')

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return self.get_object().created_by == user

    def form_valid(self, form):
        messages.success(self.request, 'Inventory Report deleted.')
        return super().form_valid(form)


# ─── Inventory Export Excel ───────────────────────────────────

@login_required
def inventory_export_excel(request, pk):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    report = get_object_or_404(InventoryReport, pk=pk)
    items = report.items.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF', size=9)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    red_fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    wrap = Alignment(wrap_text=True, vertical='top')

    # Title
    ws.cell(row=1, column=1, value=report.title).font = Font(bold=True, size=13, color='1F4E79')
    if report.warehouse_location:
        ws.cell(row=2, column=1, value=f'Warehouse: {report.warehouse_location}').font = bold

    # Headers
    headers = [
        'S.No', 'Item Code', 'Name', 'Model', 'Make', 'Category', 'Status', 'Unit',
        'Qty Received', 'Qty Issued', 'Balance', 'Min Stock', 'Unit Cost', 'Total Value',
        'TAG', 'Rack', 'Received Date', 'Handover To', 'Handover Date', 'Project', 'PO Ref', 'Remarks'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_num, item in enumerate(items, 5):
        data = [
            item.serial_number, item.item_code, item.name, item.model_number, item.make,
            item.get_category_display(), item.get_status_display(), item.unit,
            float(item.quantity_received), float(item.quantity_issued),
            float(item.balance_qty) if item.balance_qty else 0,
            float(item.min_stock_level) if item.min_stock_level else '',
            float(item.unit_cost) if item.unit_cost else '',
            float(item.total_value),
            item.tag, item.rack_location,
            item.received_date.strftime('%d-%b-%Y') if item.received_date else '',
            item.handover_to,
            item.handover_date.strftime('%d-%b-%Y') if item.handover_date else '',
            item.project_ref, item.po_reference, item.remarks,
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            cell.alignment = wrap
        if item.is_low_stock:
            for col in range(1, len(data) + 1):
                ws.cell(row=row_num, column=col).fill = red_fill

    widths = [6, 10, 28, 14, 14, 16, 10, 8, 10, 10, 10, 10, 10, 12, 10, 8, 12, 14, 12, 14, 10, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = _safe_filename(report.title, prefix='Inventory', extension='xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── Inventory Export PDF ─────────────────────────────────────

@login_required
def inventory_export_pdf(request, pk):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_RIGHT
    from io import BytesIO
    from django.contrib.staticfiles.finders import find as find_static

    report = get_object_or_404(InventoryReport, pk=pk)
    items = report.items.all()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=10*mm, bottomMargin=10*mm, leftMargin=8*mm, rightMargin=8*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('InvTitle', parent=styles['Heading1'], fontSize=13, textColor=colors.HexColor('#C41E3A'), spaceAfter=2)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, spaceAfter=4)
    s = ParagraphStyle('S', parent=styles['Normal'], fontSize=5.5, leading=6.5)
    sr = ParagraphStyle('SR', parent=styles['Normal'], fontSize=5.5, leading=6.5, alignment=TA_RIGHT)
    sb = ParagraphStyle('SB', parent=styles['Normal'], fontSize=5.5, leading=6.5, fontName='Helvetica-Bold')

    # Logo + Title
    logo_path = find_static('images/leap_logo.jpg')
    if logo_path:
        from reportlab.platypus import Image
        logo = Image(logo_path, width=40*mm, height=12*mm)
        t = Table([[logo, Paragraph(f'INVENTORY REPORT: {report.title}', title_style)]], colWidths=[45*mm, 230*mm])
        t.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(t)
    else:
        elements.append(Paragraph(f'INVENTORY REPORT: {report.title}', title_style))

    if report.warehouse_location:
        elements.append(Paragraph(f'<b>Warehouse:</b> {report.warehouse_location}', sub_style))
    elements.append(Paragraph(f'<b>Items:</b> {report.total_items} | <b>Low Stock:</b> {report.low_stock_count} | <b>Total Value:</b> SAR {report.total_stock_value:,.2f}', sub_style))
    elements.append(Spacer(1, 2*mm))

    dark_blue = colors.HexColor('#C41E3A')
    col_widths = [7*mm, 12*mm, 30*mm, 12*mm, 12*mm, 14*mm, 11*mm, 8*mm,
                  10*mm, 10*mm, 10*mm, 10*mm, 10*mm, 12*mm,
                  10*mm, 8*mm, 14*mm, 16*mm, 14*mm, 18*mm, 14*mm]

    header = [
        Paragraph('<b>S.No</b>', s), Paragraph('<b>Code</b>', s), Paragraph('<b>Name</b>', s),
        Paragraph('<b>Model</b>', s), Paragraph('<b>Make</b>', s), Paragraph('<b>Category</b>', s),
        Paragraph('<b>Status</b>', s), Paragraph('<b>Unit</b>', s),
        Paragraph('<b>Rcvd</b>', s), Paragraph('<b>Issued</b>', s), Paragraph('<b>Balance</b>', s),
        Paragraph('<b>Min</b>', s), Paragraph('<b>Cost</b>', s), Paragraph('<b>Value</b>', s),
        Paragraph('<b>TAG</b>', s), Paragraph('<b>Rack</b>', s),
        Paragraph('<b>Rcvd Date</b>', s), Paragraph('<b>Handover</b>', s),
        Paragraph('<b>H.Date</b>', s), Paragraph('<b>Project</b>', s), Paragraph('<b>Remarks</b>', s),
    ]
    data = [header]
    low_stock_rows = []

    for idx, item in enumerate(items):
        row_data = [
            Paragraph(str(item.serial_number), s),
            Paragraph(item.item_code or '', s),
            Paragraph(item.name, s),
            Paragraph(item.model_number or '', s),
            Paragraph(item.make or '', s),
            Paragraph(item.get_category_display(), s),
            Paragraph(item.get_status_display(), s),
            Paragraph(item.unit, s),
            Paragraph(f'{item.quantity_received:,.0f}', sr),
            Paragraph(f'{item.quantity_issued:,.0f}', sr),
            Paragraph(f'{item.balance_qty:,.0f}' if item.balance_qty is not None else '-', sr),
            Paragraph(f'{item.min_stock_level:,.0f}' if item.min_stock_level else '-', sr),
            Paragraph(f'{item.unit_cost:,.2f}' if item.unit_cost else '-', sr),
            Paragraph(f'{item.total_value:,.2f}', sr),
            Paragraph(item.tag or '', s),
            Paragraph(item.rack_location or '', s),
            Paragraph(item.received_date.strftime('%d-%m-%y') if item.received_date else '', s),
            Paragraph(item.handover_to or '', s),
            Paragraph(item.handover_date.strftime('%d-%m-%y') if item.handover_date else '', s),
            Paragraph(item.project_ref or '', s),
            Paragraph(item.remarks or '', s),
        ]
        data.append(row_data)
        if item.is_low_stock:
            low_stock_rows.append(idx + 1)

    table = Table(data, colWidths=col_widths, repeatRows=1)
    style_cmds = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5D7DC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#495057')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#C41E3A')),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]
    for r in low_stock_rows:
        style_cmds.append(('BACKGROUND', (0, r), (-1, r), colors.HexColor('#FFE0E0')))
    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    filename = _safe_filename(report.title, prefix='Inventory', extension='pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Inventory Import Excel ───────────────────────────────────

@login_required
def inventory_import_excel(request):
    if request.method != 'POST':
        return redirect('procurement:inventory_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select an Excel file.')
        return redirect('procurement:inventory_list')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        created_count = 0

        for ws in wb.worksheets:
            if ws.max_row < 3:
                continue

            # Create report per sheet
            report = InventoryReport.objects.create(
                title=ws.title,
                created_by=request.user,
            )

            # Find header row
            header_row = None
            for r in range(1, min(5, ws.max_row + 1)):
                val = str(ws.cell(row=r, column=1).value or '').lower().strip()
                if val in ('no', 'sr.no.', 'sr.no', 's.no', 's.no.', 'sr'):
                    header_row = r
                    break
            if not header_row:
                header_row = 1

            item_count = 0
            for r in range(header_row + 1, ws.max_row + 1):
                sn = ws.cell(row=r, column=1).value
                name = ws.cell(row=r, column=2).value
                if not name:
                    continue
                name_str = str(name).strip()
                if not name_str:
                    continue

                def cv(col):
                    v = ws.cell(row=r, column=col).value
                    return str(v).strip() if v is not None else ''

                def nv(col):
                    v = ws.cell(row=r, column=col).value
                    if isinstance(v, (int, float)):
                        return Decimal(str(v))
                    return None

                def dv(col):
                    v = ws.cell(row=r, column=col).value
                    if isinstance(v, datetime):
                        return v.date()
                    return None

                qty_received = nv(6) or Decimal('0')
                qty_issued = nv(11) or Decimal('0')
                balance = nv(12)

                InventoryItem.objects.create(
                    report=report,
                    serial_number=int(sn) if isinstance(sn, (int, float)) else item_count + 1,
                    name=name_str,
                    model_number=cv(3),
                    make=cv(4),
                    unit=cv(5) or 'Each',
                    quantity_received=qty_received,
                    quantity_issued=qty_issued,
                    balance_qty=balance if balance is not None else qty_received - qty_issued,
                    tag=cv(7),
                    rack_location=cv(8),
                    received_date=dv(9),
                    handover_to=cv(10),
                    handover_date=dv(13),
                    project_ref=cv(14),
                    remarks=cv(15),
                    order=item_count,
                )
                item_count += 1

            if item_count == 0:
                report.delete()
            else:
                created_count += 1

        messages.success(request, f'Imported {created_count} inventory sheet(s).')
        return redirect('procurement:inventory_list')

    except Exception as e:
        messages.error(request, f'Error importing: {str(e)}')
        return redirect('procurement:inventory_list')


# ═══════════════════════════════════════════════════════════════
# FRC / PPE UNIFORM TRACKER
# ═══════════════════════════════════════════════════════════════

class FRCPermissionMixin(LoginRequiredMixin, UserPassesTestMixin):
    def get_queryset(self):
        queryset = FRCReport.objects.select_related('project', 'created_by').all()
        user = self.request.user
        if user.is_super_admin_user or user.is_procurement_user:
            return queryset
        elif user.is_admin_user or user.is_manager_user:
            return queryset.filter(Q(created_by=user) | Q(project__region=user.region))
        else:
            return queryset.filter(created_by=user)


class FRCListView(FRCPermissionMixin, ListView):
    model = FRCReport
    template_name = 'procurement/frc_list.html'
    context_object_name = 'reports'
    paginate_by = 25

    def test_func(self):
        return True

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(title__icontains=search))
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['total_count'] = self.get_queryset().count()
        return context


class FRCCreateView(FRCPermissionMixin, CreateView):
    model = FRCReport
    form_class = FRCReportForm
    template_name = 'procurement/frc_form.html'

    def test_func(self):
        return True

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['entry_formset'] = FRCEntryFormSet(self.request.POST, prefix='entries')
        else:
            context['entry_formset'] = FRCEntryFormSet(prefix='entries')
        context['title'] = 'Create FRC / PPE Report'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['entry_formset']
        if formset.is_valid():
            self.object = form.save(commit=False)
            self.object.created_by = self.request.user
            self.object.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, 'FRC Report created.')
            return redirect(self.get_success_url())
        return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:frc_detail', kwargs={'pk': self.object.pk})


class FRCDetailView(FRCPermissionMixin, DetailView):
    model = FRCReport
    template_name = 'procurement/frc_detail.html'
    context_object_name = 'report'

    def test_func(self):
        return True

    def get_context_data(self, **kwargs):
        from collections import OrderedDict
        from datetime import datetime

        context = super().get_context_data(**kwargs)
        entries = list(self.object.entries.all())
        context['entries'] = entries

        # Collect unique projects and sizes for filters
        projects = set()
        shirt_sizes = set()
        pants_sizes = set()
        shoe_sizes = set()
        for e in entries:
            if e.project_name:
                projects.add(e.project_name.strip())
            if e.shirt_size:
                shirt_sizes.add(e.shirt_size.strip())
            if e.pants_size:
                pants_sizes.add(e.pants_size.strip())
            if e.safety_shoes and e.safety_shoes not in ('N/A', 'n/a', '-', ''):
                shoe_sizes.add(e.safety_shoes.strip())

        context['projects'] = sorted(projects)
        context['shirt_sizes'] = sorted(shirt_sizes)
        context['pants_sizes'] = sorted(pants_sizes)
        context['shoe_sizes'] = sorted(shoe_sizes)

        # ── Group entries by receiving_date (the date PPE was handed over).
        # receiving_date is a CharField — try a few common formats so we can
        # sort newest-first; fall back to the raw string if parsing fails.
        DATE_FORMATS = (
            '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d %b %Y', '%d %B %Y',
            '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y',
        )

        def _parse_date(raw):
            if not raw:
                return None
            s = raw.strip()
            for fmt in DATE_FORMATS:
                try:
                    return datetime.strptime(s, fmt).date()
                except ValueError:
                    continue
            return None

        groups = {}
        for e in entries:
            raw = (e.receiving_date or '').strip()
            parsed = _parse_date(raw)
            key = parsed.isoformat() if parsed else (raw or '__nodate__')
            label = parsed.strftime('%d %B %Y') if parsed else (raw or 'Date not recorded')
            day_label = parsed.strftime('%d') if parsed else '—'
            month_label = parsed.strftime('%b %Y').upper() if parsed else ''
            weekday_label = parsed.strftime('%A') if parsed else ''
            if key not in groups:
                groups[key] = {
                    'parsed': parsed,
                    'label': label,
                    'day': day_label,
                    'month': month_label,
                    'weekday': weekday_label,
                    'entries': [],
                }
            groups[key]['entries'].append(e)

        # Sort groups newest-first; entries with unparseable / missing dates
        # bubble to the bottom.
        sorted_groups = OrderedDict()
        for key, g in sorted(
            groups.items(),
            key=lambda kv: (kv[1]['parsed'] is None, -(kv[1]['parsed'].toordinal() if kv[1]['parsed'] else 0)),
        ):
            sorted_groups[key] = g
        context['entries_by_date'] = sorted_groups
        return context


class FRCUpdateView(FRCPermissionMixin, UpdateView):
    model = FRCReport
    form_class = FRCReportForm
    template_name = 'procurement/frc_form.html'

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return self.get_object().created_by == user

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['entry_formset'] = FRCEntryFormSet(self.request.POST, instance=self.object, prefix='entries')
        else:
            context['entry_formset'] = FRCEntryFormSet(instance=self.object, prefix='entries')
        context['title'] = f'Edit: {self.object.title}'
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        formset = context['entry_formset']
        if formset.is_valid():
            self.object = form.save()
            formset.instance = self.object
            formset.save()
            messages.success(self.request, 'FRC Report updated.')
            return redirect(self.get_success_url())
        return self.form_invalid(form)

    def get_success_url(self):
        return reverse('procurement:frc_detail', kwargs={'pk': self.object.pk})


class FRCDeleteView(FRCPermissionMixin, DeleteView):
    model = FRCReport
    template_name = 'procurement/frc_confirm_delete.html'
    success_url = reverse_lazy('procurement:frc_list')

    def test_func(self):
        user = self.request.user
        if user.is_super_admin_user or user.is_admin_user or user.is_procurement_user:
            return True
        return self.get_object().created_by == user

    def form_valid(self, form):
        messages.success(self.request, 'FRC Report deleted.')
        return super().form_valid(form)


# ─── FRC Export Excel ─────────────────────────────────────────

@login_required
def frc_export_excel(request, pk):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    report = get_object_or_404(FRCReport, pk=pk)
    entries = report.entries.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FRC Report'

    bold = Font(bold=True)
    header_font = Font(bold=True, color='FFFFFF', size=9)
    header_fill = PatternFill(start_color='C41E3A', end_color='C41E3A', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    ws.cell(row=1, column=1, value=report.title).font = Font(bold=True, size=13, color='C41E3A')

    headers = ['Sr.', 'Name', 'Designation', 'Project', 'Shirt Size', 'Shirts', 'Pants Size', 'Pants',
               'Safety Glasses', 'Safety Shoes', 'Shoes Qty', 'Safety Helmet', 'Receiving Date', 'Handed To / Remarks']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row_num, e in enumerate(entries, 4):
        data = [e.serial_number, e.employee_name, e.designation, e.project_name,
                e.shirt_size, e.shirt_qty, e.pants_size, e.pants_qty,
                e.safety_glasses, e.safety_shoes, e.shoes_qty, e.safety_helmet,
                e.receiving_date, e.handed_to]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border

    widths = [6, 22, 20, 16, 12, 8, 12, 8, 12, 12, 8, 12, 18, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = _safe_filename(report.title, prefix='FRC', extension='xlsx')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── FRC Export PDF ───────────────────────────────────────────

@login_required
def frc_export_pdf(request, pk):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER
    from io import BytesIO
    from django.contrib.staticfiles.finders import find as find_static

    report = get_object_or_404(FRCReport, pk=pk)
    entries = report.entries.all()

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=10*mm, bottomMargin=10*mm, leftMargin=10*mm, rightMargin=10*mm)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle('FRCTitle', parent=styles['Heading1'], fontSize=14, textColor=colors.HexColor('#C41E3A'), spaceAfter=2)
    s = ParagraphStyle('S', parent=styles['Normal'], fontSize=7, leading=8)

    logo_path = find_static('images/leap_logo.jpg')
    if logo_path:
        from reportlab.platypus import Image
        logo = Image(logo_path, width=40*mm, height=12*mm)
        t = Table([[logo, Paragraph(report.title, title_style)]], colWidths=[45*mm, 230*mm])
        t.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'MIDDLE')]))
        elements.append(t)
    else:
        elements.append(Paragraph(report.title, title_style))
    elements.append(Spacer(1, 4*mm))

    leap_red = colors.HexColor('#C41E3A')
    col_widths = [8*mm, 32*mm, 28*mm, 22*mm, 16*mm, 10*mm, 16*mm, 10*mm,
                  14*mm, 16*mm, 10*mm, 14*mm, 24*mm, 40*mm]
    header = [Paragraph(f'<b>{h}</b>', s) for h in [
        'Sr.', 'Name', 'Designation', 'Project', 'Shirt Size', 'Shirts', 'Pants Size', 'Pants',
        'Glasses', 'Shoes', 'Shoe Qty', 'Helmet', 'Recv Date', 'Handed To / Remarks'
    ]]
    data = [header]
    for e in entries:
        data.append([
            Paragraph(str(e.serial_number), s), Paragraph(e.employee_name, s),
            Paragraph(e.designation, s), Paragraph(e.project_name, s),
            Paragraph(e.shirt_size, s), Paragraph(str(e.shirt_qty), s),
            Paragraph(e.pants_size, s), Paragraph(str(e.pants_qty), s),
            Paragraph(e.safety_glasses, s), Paragraph(e.safety_shoes, s),
            Paragraph(str(e.shoes_qty), s), Paragraph(e.safety_helmet, s),
            Paragraph(e.receiving_date, s), Paragraph(e.handed_to, s),
        ])

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F5D7DC')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#495057')),
        ('LINEBELOW', (0, 0), (-1, 0), 1, leap_red),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FFF5F5')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    response = HttpResponse(buf.read(), content_type='application/pdf')
    filename = _safe_filename(report.title, prefix='FRC', extension='pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── FRC Import Excel ─────────────────────────────────────────

@login_required
def frc_import_excel(request):
    if request.method != 'POST':
        return redirect('procurement:frc_list')

    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.error(request, 'Please select an Excel file.')
        return redirect('procurement:frc_list')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        created_count = 0

        for ws in wb.worksheets:
            if ws.max_row < 3:
                continue

            title = str(ws.cell(row=1, column=1).value or ws.title).strip()
            report = FRCReport.objects.create(title=title or ws.title, created_by=request.user)

            # Find header row
            header_row = None
            for r in range(1, min(5, ws.max_row + 1)):
                val = str(ws.cell(row=r, column=1).value or '').lower().strip()
                if val in ('sr.', 'sr', 'sr.no', 's.no', 'sr. no'):
                    header_row = r
                    break
                val2 = str(ws.cell(row=r, column=2).value or '').lower().strip()
                if val2 == 'name':
                    header_row = r
                    break
            if not header_row:
                header_row = 2

            entry_count = 0
            for r in range(header_row + 1, ws.max_row + 1):
                name = ws.cell(row=r, column=2).value
                if not name:
                    continue
                name_str = str(name).strip()
                if not name_str:
                    continue

                def cv(col):
                    v = ws.cell(row=r, column=col).value
                    return str(v).strip() if v is not None else ''

                def iv(col):
                    """Extract quantity from values like '2', '2+1(XL)', '1(M)', '2(32)'."""
                    import re
                    v = ws.cell(row=r, column=col).value
                    if v is None:
                        return 0
                    if isinstance(v, (int, float)):
                        return int(v)
                    s = str(v).strip()
                    if not s or s in ('N/A', 'n/a', '-'):
                        return 0
                    # Split by + and extract the leading number from each part
                    # "2+1(XL)" → ["2", "1(XL)"] → 2 + 1 = 3
                    # "2(32)" → ["2(32)"] → 2
                    # "1 N+1 H" → ["1 N", "1 H"] → 1 + 1 = 2
                    total = 0
                    for part in s.split('+'):
                        part = part.strip()
                        m = re.match(r'(\d+)', part)
                        if m:
                            total += int(m.group(1))
                    return total

                sn = ws.cell(row=r, column=1).value

                # Format receiving date properly
                recv_raw = ws.cell(row=r, column=10).value
                if isinstance(recv_raw, datetime):
                    recv_date = recv_raw.strftime('%d-%b-%Y')
                else:
                    recv_date = str(recv_raw).strip() if recv_raw else ''

                FRCEntry.objects.create(
                    report=report,
                    serial_number=int(sn) if isinstance(sn, (int, float)) else entry_count + 1,
                    employee_name=name_str,
                    designation=cv(3),
                    project_name=cv(4),
                    shirt_size=cv(5),
                    shirt_qty=iv(5),
                    pants_size=cv(6),
                    pants_qty=iv(6),
                    safety_glasses=cv(7),
                    safety_shoes=cv(8),
                    shoes_qty=iv(8),
                    safety_helmet=cv(9),
                    receiving_date=recv_date,
                    handed_to=cv(11),
                    order=entry_count,
                )
                entry_count += 1

            if entry_count == 0:
                report.delete()
            else:
                created_count += 1

        messages.success(request, f'Imported {created_count} FRC report(s).')
        return redirect('procurement:frc_list')

    except Exception as e:
        messages.error(request, f'Error importing: {str(e)}')
        return redirect('procurement:frc_list')


# ═══════════════════════════════════════════════════════════════
# FRC INVENTORY (PPE STOCK)
# ═══════════════════════════════════════════════════════════════

@login_required
def frc_inventory_list(request):
    """FRC Inventory dashboard — PPE stock by type and size."""
    items = FRCInventory.objects.all()

    # Filter
    item_type = request.GET.get('type', '')
    if item_type:
        items = items.filter(item_type=item_type)

    # Group by type for dashboard cards
    type_summary = {}
    for choice_val, choice_label in FRCInventory.ITEM_TYPE_CHOICES:
        type_items = FRCInventory.objects.filter(item_type=choice_val)
        if type_items.exists():
            total_avail = sum(i.available_stock for i in type_items)
            total_purchased = sum(i.total_purchased for i in type_items)
            total_issued = sum(i.total_issued for i in type_items)
            low_count = sum(1 for i in type_items if i.is_low_stock)
            type_summary[choice_val] = {
                'label': choice_label,
                'count': type_items.count(),
                'available': total_avail,
                'purchased': total_purchased,
                'issued': total_issued,
                'low_stock': low_count,
            }

    total_value = sum(i.stock_value for i in FRCInventory.objects.all())
    total_low = sum(1 for i in FRCInventory.objects.all() if i.is_low_stock)

    context = {
        'items': items,
        'type_summary': type_summary,
        'active_type': item_type,
        'item_type_choices': FRCInventory.ITEM_TYPE_CHOICES,
        'total_value': total_value,
        'total_low': total_low,
        'total_items': FRCInventory.objects.count(),
    }
    return render(request, 'procurement/frc_inventory.html', context)


class FRCInventoryCreateView(LoginRequiredMixin, CreateView):
    model = FRCInventory
    form_class = FRCInventoryForm
    template_name = 'procurement/frc_inventory_form.html'
    success_url = reverse_lazy('procurement:frc_inventory')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Add PPE Stock Item'
        return context

    def form_valid(self, form):
        form.instance.created_by = self.request.user
        messages.success(self.request, 'PPE stock item added.')
        return super().form_valid(form)


class FRCInventoryUpdateView(LoginRequiredMixin, UpdateView):
    model = FRCInventory
    form_class = FRCInventoryForm
    template_name = 'procurement/frc_inventory_form.html'
    success_url = reverse_lazy('procurement:frc_inventory')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Edit: {self.object}'
        return context

    def form_valid(self, form):
        messages.success(self.request, 'PPE stock item updated.')
        return super().form_valid(form)


class FRCInventoryDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = FRCInventory
    template_name = 'procurement/frc_inventory_confirm_delete.html'
    success_url = reverse_lazy('procurement:frc_inventory')

    def test_func(self):
        return self.request.user.is_super_admin_user or self.request.user.is_admin_user

    def form_valid(self, form):
        messages.success(self.request, 'PPE stock item deleted.')
        return super().form_valid(form)


# ─── PO Terms Toggle (AJAX) ──────────────────────────────────

@login_required
@require_POST
def ajax_po_toggle_term(request, pk):
    """Toggle a TermsTemplate on/off for a Purchase Order."""
    from costing.models import TermsTemplate
    po = get_object_or_404(PurchaseOrder, pk=pk)

    user = request.user
    can_edit = (
        user.is_super_admin_user or user.is_admin_user or user.is_procurement_user
        or po.created_by_id == user.id
    )
    if not can_edit:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    term_id = request.POST.get('term_id')
    if not term_id:
        return JsonResponse({'error': 'term_id required'}, status=400)
    term = get_object_or_404(TermsTemplate, pk=term_id)

    with transaction.atomic():
        locked_po = PurchaseOrder.objects.select_for_update().get(pk=po.pk)
        if locked_po.selected_terms.filter(pk=term.pk).exists():
            locked_po.selected_terms.remove(term)
            selected = False
        else:
            locked_po.selected_terms.add(term)
            selected = True
    return JsonResponse({'ok': True, 'selected': selected})
