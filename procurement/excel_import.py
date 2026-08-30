"""Reading purchase-order line items out of a spreadsheet.

Separate from the whole-PO template importer in views.py on purpose. That one
reads OUR export format, whose cell positions are fixed because we write them.
This one reads a spreadsheet somebody else produced — a vendor quotation, a
client's bill of quantities — where the only thing that can be relied on is
that the columns are labelled.

So it finds the header row and maps columns by name, the same approach
projects.views.ProjectImportView already uses for the sales import. Fixed
offsets break the moment somebody inserts a column, which on a file we did not
generate is the normal case rather than the exception.
"""

from decimal import Decimal, InvalidOperation

# Accepted spellings per field. Compared after lowercasing, collapsing
# whitespace and dropping anything in brackets, so "Rate / Unit (SAR)" and
# "rate per unit" both land on the same field.
COLUMN_ALIASES = {
    'description': ('description', 'item description', 'item', 'particulars',
                    'specification', 'item description / specification',
                    'material description', 'scope'),
    'make_model': ('make', 'model', 'make/model', 'make / model', 'brand',
                   'manufacturer', 'part number', 'part no'),
    'quantity': ('quantity', 'qty', 'no', 'nos', 'count'),
    'uom': ('uom', 'unit', 'units', 'unit of measure', 'u/m'),
    'rate_per_unit': ('rate', 'rate/unit', 'rate per unit', 'unit rate',
                      'unit price', 'price', 'rate/unit sar', 'unit cost'),
    'remarks': ('remarks', 'remark', 'notes', 'note', 'comments'),
    'system': ('system', 'category', 'discipline'),
    'vendor_name': ('vendor', 'vendor name', 'supplier', 'supplier name'),
}

REQUIRED_COLUMNS = ('description',)

# How far down to look for the header before giving up. A quotation with more
# than this much preamble is not a table we should be guessing at.
MAX_HEADER_SCAN_ROWS = 30


class ExcelImportError(Exception):
    """The file cannot be read at all. The message is shown to the user."""


def _normalise(value):
    """Header text reduced to something comparable."""
    if value is None:
        return ''
    text = str(value).lower()
    # Drop units and currency in brackets — "rate/unit (SAR)" is still a rate.
    while '(' in text and ')' in text:
        start = text.index('(')
        end = text.index(')', start)
        text = text[:start] + ' ' + text[end + 1:]
    for ch in '.:_/\\-':
        text = text.replace(ch, ' ')
    return ' '.join(text.split())


def _to_decimal(value, default=Decimal('0')):
    if value is None or value == '':
        return default
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    text = str(value).replace(',', '').strip()
    # Vendors write "12 nos" or "SAR 1,200.00" in a numeric column.
    cleaned = ''.join(c for c in text if c.isdigit() or c in '.-')
    try:
        return Decimal(cleaned) if cleaned not in ('', '-', '.') else default
    except InvalidOperation:
        return default


def _alias_lookup():
    """{normalised alias: field}, built once.

    The aliases above are written the way a header reads — "rate/unit",
    "make/model" — but they are compared against normalised text, where the
    slash has already become a space. Normalising both sides is what stops
    "Rate/Unit (SAR)" silently failing to match its own alias.
    """
    lookup = {}
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            lookup.setdefault(_normalise(alias), field)
    return lookup


ALIAS_LOOKUP = _alias_lookup()


def find_header_row(ws):
    """(row index, {field: column index}) for the first row that looks like a
    header, or (None, {}).

    A row qualifies when it names every required column. Matching on content
    rather than position is what lets this read a file with a title block, a
    logo and three blank rows above the table.
    """
    for row_idx in range(1, min(ws.max_row, MAX_HEADER_SCAN_ROWS) + 1):
        mapping = {}
        for col_idx in range(1, ws.max_column + 1):
            header = _normalise(ws.cell(row=row_idx, column=col_idx).value)
            if not header:
                continue
            field = ALIAS_LOOKUP.get(header)
            # First column wins, so a sheet with two "Rate" columns uses the
            # leftmost rather than whichever happens to be scanned last.
            if field and field not in mapping:
                mapping[field] = col_idx
        if all(field in mapping for field in REQUIRED_COLUMNS):
            return row_idx, mapping
    return None, {}


def parse_items(file_obj):
    """(rows, skipped) read from a spreadsheet.

    `rows` are dicts ready for PurchaseOrderItem(**row). `skipped` is a list of
    (row number, reason) — reported to the user rather than dropped, because an
    import that silently reads 8 of 12 lines is worse than one that fails.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
    except Exception as exc:      # noqa: BLE001 - openpyxl raises many types
        raise ExcelImportError(
            f'That file could not be read as a spreadsheet ({exc}). Save it as '
            '.xlsx and try again.') from exc

    ws = wb.active
    header_row, mapping = find_header_row(ws)
    if header_row is None:
        raise ExcelImportError(
            'No item table found. The sheet needs a header row with a column '
            'named Description (Quantity, Rate and UOM are read if present).')

    rows, skipped = [], []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        def value(field):
            col = mapping.get(field)
            return ws.cell(row=row_idx, column=col).value if col else None

        description = value('description')
        description = str(description).strip() if description is not None else ''
        if not description:
            # Blank descriptions are the spacer and totals rows every quotation
            # ends with. Not worth reporting as a problem.
            continue
        if any(word in description.lower()
               for word in ('total', 'subtotal', 'grand total', 'vat')):
            skipped.append((row_idx, f'looks like a totals row: "{description[:40]}"'))
            continue

        quantity = _to_decimal(value('quantity'), Decimal('1'))
        if quantity <= 0:
            skipped.append((row_idx, f'quantity is {quantity or "blank"}'))
            continue

        rows.append({
            'description': description,
            'make_model': str(value('make_model') or '').strip()[:255],
            'quantity': quantity,
            'uom': str(value('uom') or 'Nos').strip()[:50] or 'Nos',
            'rate_per_unit': _to_decimal(value('rate_per_unit')),
            'remarks': str(value('remarks') or '').strip(),
            'system': str(value('system') or '').strip()[:100],
            'vendor_name': str(value('vendor_name') or '').strip()[:255],
        })

    return rows, skipped


def summarise(added, skipped, filename=''):
    """One sentence saying what an import actually did."""
    prefix = f'{filename}: ' if filename else ''
    if not added and not skipped:
        return f'{prefix}no item rows found.'
    parts = [f'{added} item{"" if added == 1 else "s"} added']
    if skipped:
        parts.append(f'{len(skipped)} row{"" if len(skipped) == 1 else "s"} skipped')
    return prefix + ', '.join(parts) + '.'
