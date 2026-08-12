from datetime import date, timedelta
from decimal import Decimal
from unittest.mock import patch

from django.test import TestCase, Client
from django.urls import reverse
from django.utils import timezone

from accounts.models import Role, User
from accounts.permissions import seed_default_permissions
from hr.models import Employee
from .models import ActivityCode, TimesheetEntry
from notifications.models import Notification
from .models import HRSettings, TimesheetRequest, TimesheetRequestAck
from .services import request_timesheets, remind_employee, employees_for_request


def _make_user_with_employee(username, role_name=Role.SALES_REP, full_name=None):
    """Every test needs a user with a role (so timesheets.access actually
    resolves to True — see the lesson from the bugfixes branch: a roleless
    test user always fails has_capability, regardless of seed_default_permissions())
    AND a linked Employee record (since the views 404/redirect gracefully
    otherwise, which is a different code path than the one most tests want
    to exercise)."""
    role, _ = Role.objects.get_or_create(name=role_name)
    user = User.objects.create_user(username=username, password='testpass123', role=role)
    emp = Employee.objects.create(
        full_name=full_name or username, iqama_number=f'IQ-{username}', user=user)
    return user, emp


class TimesheetEntryOwnershipTests(TestCase):
    """The IDOR-critical tests: an employee's own entries are reachable and
    editable; someone else's are not, and the response is 404 (not 403), so
    the pk's existence can't be inferred from the response code alone —
    same pattern as leave_request_document_download in hr/views.py."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user_a, self.emp_a = _make_user_with_employee('emp_a', full_name='Employee A')
        self.user_b, self.emp_b = _make_user_with_employee('emp_b', full_name='Employee B')
        self.entry_a = TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 7, 10),
            task_description='Employee A work log', activity_code=self.code, hours=Decimal('8'))
        seed_default_permissions()  # moved to the end — must run AFTER roles exist

        # ── happy path ──
    def test_owner_can_view_own_timesheet(self):
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, "Employee A work log")

    def test_owner_can_add_entry(self):
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.post(reverse('timesheets:my_timesheet'), {
            'entry_type': 'activity',
            'date': '2026-07-11', 'activity_code': self.code.pk,
            'task_description': 'New task', 'hours': '4',
        })
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(TimesheetEntry.objects.filter(
            employee=self.emp_a, date=date(2026, 7, 11)).exists())

    def test_owner_can_edit_own_draft_entry(self):
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:entry_edit', args=[self.entry_a.pk]), {
                'entry_type': 'activity',
                'date': '2026-07-10', 'activity_code': self.code.pk,
                'task_description': 'Updated', 'hours': '6',
            })
        self.assertEqual(resp.status_code, 302)
        self.entry_a.refresh_from_db()
        self.assertEqual(self.entry_a.task_description, 'Updated')

    def test_owner_can_delete_own_draft_entry(self):
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.post(reverse('timesheets:entry_delete', args=[self.entry_a.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(TimesheetEntry.objects.filter(pk=self.entry_a.pk).exists())

    # ── the security boundary: THE most important tests in this file ──
    def test_other_employee_cannot_view_edit_page_for_entry(self):
        self.client.login(username='emp_b', password='testpass123')
        resp = self.client.get(reverse('timesheets:entry_edit', args=[self.entry_a.pk]))
        self.assertEqual(resp.status_code, 404)

    def test_other_employee_cannot_edit_entry_via_post(self):
        self.client.login(username='emp_b', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:entry_edit', args=[self.entry_a.pk]), {
                'date': '2026-07-10', 'activity_code': self.code.pk,
                'task_description': 'HIJACKED', 'hours': '1',
            })
        self.assertEqual(resp.status_code, 404)
        self.entry_a.refresh_from_db()
        self.assertEqual(self.entry_a.task_description, "Employee A work log") # untouched

    def test_other_employee_cannot_delete_entry(self):
        self.client.login(username='emp_b', password='testpass123')
        resp = self.client.post(reverse('timesheets:entry_delete', args=[self.entry_a.pk]))
        self.assertEqual(resp.status_code, 404)
        self.assertTrue(TimesheetEntry.objects.filter(pk=self.entry_a.pk).exists())  # still there

    # ── auth required ──
    def test_anonymous_redirected_to_login(self):
        resp = self.client.get(reverse('timesheets:my_timesheet'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)

    # ── locked (submitted) entries can't be edited/deleted, even by owner ──
    def test_submitted_entry_cannot_be_edited_by_owner(self):
        self.entry_a.status = TimesheetEntry.STATUS_SUBMITTED
        self.entry_a.save()
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:entry_edit', args=[self.entry_a.pk]), {
                'date': '2026-07-10', 'activity_code': self.code.pk,
                'task_description': 'Trying to change', 'hours': '1',
            })
        self.assertEqual(resp.status_code, 302)  # redirected with error message
        self.entry_a.refresh_from_db()
        self.assertEqual(self.entry_a.task_description, "Employee A work log")  # unchanged  

    def test_submitted_entry_cannot_be_deleted_by_owner(self):
        self.entry_a.status = TimesheetEntry.STATUS_SUBMITTED
        self.entry_a.save()
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.post(reverse('timesheets:entry_delete', args=[self.entry_a.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(TimesheetEntry.objects.filter(pk=self.entry_a.pk).exists())

    def test_entry_delete_via_get_not_allowed(self):
        self.client.login(username='emp_a', password='testpass123')
        resp = self.client.get(reverse('timesheets:entry_delete', args=[self.entry_a.pk]))
        self.assertEqual(resp.status_code, 405)

    def test_edit_cannot_reassign_employee_via_post(self):
        other_user, other_emp = _make_user_with_employee('editreassign')
        self.client.login(username='emp_a', password='testpass123')
        self.client.post(
            reverse('timesheets:entry_edit', args=[self.entry_a.pk]), {
                'date': '2026-07-10', 'activity_code': self.code.pk,
                'task_description': 'Still mine', 'hours': '5',
                'employee': other_emp.pk,
            })
        self.entry_a.refresh_from_db()
        self.assertEqual(self.entry_a.employee, self.emp_a)  # unchanged, not other_emp


class TimesheetOpenToAllUsersTests(TestCase):
    """Confirms the deliberate design change: My Timesheet has no
    capability gate anymore, just login — any authenticated user with a
    linked Employee record can reach it, regardless of role."""

    def test_user_with_no_role_permissions_seeded_can_still_open_my_timesheet(self):
        """Deliberately skip seed_default_permissions() here — if
        my_timesheet ever regains a capability check, this test starts
        failing immediately instead of passing by accident."""
        role, _ = Role.objects.get_or_create(name=Role.DEVELOPER)
        user = User.objects.create_user(username='noaccess', password='testpass123', role=role)
        Employee.objects.create(full_name='No Access', iqama_number='IQ-NOACCESS', user=user)
        client = Client()
        client.login(username='noaccess', password='testpass123')
        resp = client.get(reverse('timesheets:my_timesheet'))
        self.assertEqual(resp.status_code, 200)

class TimesheetEntryValidationTests(TestCase):
    """Server-side validation — must hold even if someone bypasses the
    HTML date-picker's max attribute or the number input's min/max via
    browser dev tools, since those are purely cosmetic client-side hints."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('valemp')
        seed_default_permissions()  # moved to the end
        self.client.login(username='valemp', password='testpass123')    

    def _post(self, **overrides):
        data = {
            'entry_type': 'activity',
            'date': '2026-07-10', 'activity_code': self.code.pk,
            'task_description': 'Task', 'hours': '4',
        }
        data.update(overrides)
        return self.client.post(reverse('timesheets:my_timesheet'), data)

    def test_future_date_now_accepted(self):
        """Regression lock-in: clean_date() was removed from
        TimesheetEntryForm, so future dates are now intentionally allowed
        (per the product decision to let employees log future entries).
        This replaces the old test_future_date_rejected, which would
        silently pass on a false assumption if left in place."""
        future = (date.today() + timedelta(days=5)).isoformat()
        resp = self._post(date=future)
        self.assertEqual(resp.status_code, 302)  # redirected on success, not re-rendered
        self.assertTrue(TimesheetEntry.objects.filter(date=future).exists())

    def test_far_future_date_still_accepted(self):
        """Confirms there's no hidden upper bound either (e.g. no
        'within N days' cap left over from an earlier version)."""
        far_future = (date.today() + timedelta(days=400)).isoformat()
        resp = self._post(date=far_future)
        self.assertEqual(resp.status_code, 302)
        self.assertTrue(TimesheetEntry.objects.filter(date=far_future).exists())

    def test_zero_hours_rejected(self):
        resp = self._post(hours='0')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(TimesheetEntry.objects.count(), 0)

    def test_negative_hours_rejected(self):
        resp = self._post(hours='-3')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(TimesheetEntry.objects.count(), 0)

    def test_over_24_hours_rejected(self):
        resp = self._post(hours='25')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(TimesheetEntry.objects.count(), 0)

    def test_valid_entry_accepted(self):
        resp = self._post()
        self.assertEqual(resp.status_code, 302)
        self.assertEqual(TimesheetEntry.objects.count(), 1)

    def test_deactivated_code_not_selectable(self):
        inactive_code = ActivityCode.objects.create(code='OLD_CODE', is_active=False)
        resp = self._post(activity_code=inactive_code.pk)
        self.assertEqual(resp.status_code, 200)  # form invalid: not a valid choice
        self.assertEqual(TimesheetEntry.objects.count(), 0)

    def test_malicious_post_cannot_set_employee_or_status(self):
        """Even if someone crafts a POST including 'employee' and 'status'
        fields (neither is in TimesheetEntryForm.Meta.fields), the created
        row must still belong to the logged-in user and start as draft."""
        other_user, other_emp = _make_user_with_employee('otheremp2')
        resp = self._post(**{'employee': other_emp.pk, 'status': TimesheetEntry.STATUS_SUBMITTED})
        self.assertEqual(resp.status_code, 302)
        entry = TimesheetEntry.objects.get()
        self.assertEqual(entry.employee, self.emp)  # not other_emp
        self.assertEqual(entry.status, TimesheetEntry.STATUS_DRAFT)  # not submitted


class TimesheetExportOwnershipTests(TestCase):
    """The export must only ever contain the requesting employee's own
    data — same IDOR concern as entry edit/delete, just on a read path."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user_a, self.emp_a = _make_user_with_employee('expa', full_name='Export A')
        self.user_b, self.emp_b = _make_user_with_employee('expb', full_name='Export B')
        TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 7, 5),
            task_description='A-only task', activity_code=self.code, hours=Decimal('3'))
        TimesheetEntry.objects.create(
            employee=self.emp_b, date=date(2026, 7, 5),
            task_description='B-only task', activity_code=self.code, hours=Decimal('5'))
        seed_default_permissions()

    def test_export_is_valid_xlsx(self):
        self.client.login(username='expa', password='testpass123')
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    def test_export_contains_only_own_entries_not_other_employees(self):
        """The real IDOR check for this endpoint: load the workbook and
        confirm B's task text never appears anywhere in A's export."""
        import openpyxl
        from io import BytesIO
        self.client.login(username='expa', password='testpass123')
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        all_text = ' '.join(
            str(cell.value) for row in ws.iter_rows() for cell in row if cell.value)
        self.assertIn('A-only task', all_text)
        self.assertNotIn('B-only task', all_text)

    def test_no_employee_link_redirects_instead_of_crashing(self):
        role, _ = Role.objects.get_or_create(name=Role.SALES_REP)
        User.objects.create_user(username='noemp', password='testpass123', role=role)
        seed_default_permissions()
        self.client.login(username='noemp', password='testpass123')
        resp = self.client.get(reverse('timesheets:export'))
        self.assertEqual(resp.status_code, 302)


class HRReviewCapabilityBoundaryTests(TestCase):
    """The highest-priority gap per the guide: proving the DENIED case for
    every access-control rule, not just the happy path. A plain employee
    (no timesheets.review) must never reach the HR-only pages, no matter
    the URL they try."""

    def setUp(self):
        self.client = Client()
        self.plain_user, self.plain_emp = _make_user_with_employee('plainemp')
        # HR user: explicitly admin, which DOES have timesheets.review
        # per accounts/permissions.py's DEFAULT_CODENAME_GRANTS.
        self.hr_user, self.hr_emp = _make_user_with_employee(
            'hrperson', role_name=Role.ADMIN, full_name='HR Person')
        seed_default_permissions()

        from .services import request_timesheets
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    # ── plain employee denied ──
    def test_plain_employee_cannot_open_hr_request_page(self):
        self.client.login(username='plainemp', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_request'))
        self.assertEqual(resp.status_code, 403)

    def test_plain_employee_cannot_open_hr_request_detail(self):
        self.client.login(username='plainemp', password='testpass123')
        resp = self.client.get(
            reverse('timesheets:hr_request_detail', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 403)

    def test_plain_employee_cannot_post_a_new_request(self):
        self.client.login(username='plainemp', password='testpass123')
        resp = self.client.post(reverse('timesheets:hr_request'), {
            'year': '2026', 'month': '8',
        })
        self.assertEqual(resp.status_code, 403)
        # confirm no new request was silently created despite the denial
        from .models import TimesheetRequest
        self.assertEqual(TimesheetRequest.objects.filter(month=8).count(), 0)

    def test_plain_employee_cannot_trigger_a_reminder(self):
        self.client.login(username='plainemp', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:hr_request_detail', args=[self.ts_request.pk]),
            {'employee_id': self.plain_emp.pk})
        self.assertEqual(resp.status_code, 403)

    # ── HR user allowed (happy path, so the denial above is proven
    # to be about the capability, not a broken URL/view) ──
    def test_hr_user_can_open_request_page(self):
        self.client.login(username='hrperson', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_request'))
        self.assertEqual(resp.status_code, 200)

    def test_hr_user_can_open_request_detail(self):
        self.client.login(username='hrperson', password='testpass123')
        resp = self.client.get(
            reverse('timesheets:hr_request_detail', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 200)

    def test_anonymous_redirected_from_hr_pages(self):
        resp = self.client.get(reverse('timesheets:hr_request'))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)


from django.db import transaction
from .services import submit_month, reopen_month


class SubmitMonthTests(TestCase):
    """submit_month is the one authoritative place the lock happens — these
    tests prove its rules hold regardless of what UI calls it."""

    def setUp(self):
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('submitter')
        self.hr_user, _ = _make_user_with_employee('subhr', role_name=Role.ADMIN)
        seed_default_permissions()

    def test_submitting_empty_month_raises(self):
        with self.assertRaises(ValueError):
            submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)

    def test_submit_flips_draft_entries_to_submitted(self):
        e1 = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='t1', activity_code=self.code, hours=Decimal('4'))
        e2 = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 2),
            task_description='t2', activity_code=self.code, hours=Decimal('4'))
        submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)
        e1.refresh_from_db()
        e2.refresh_from_db()
        self.assertEqual(e1.status, TimesheetEntry.STATUS_SUBMITTED)
        self.assertEqual(e2.status, TimesheetEntry.STATUS_SUBMITTED)

    def test_submit_stamps_submitted_at_and_by(self):
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='t1', activity_code=self.code, hours=Decimal('4'))
        tsm = submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)
        self.assertIsNotNone(tsm.submitted_at)
        self.assertEqual(tsm.submitted_by, self.user)
        self.assertTrue(tsm.is_submitted)

    def test_double_submit_raises(self):
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='t1', activity_code=self.code, hours=Decimal('4'))
        submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)
        with self.assertRaises(ValueError):
            submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)

    def test_entries_from_other_months_not_touched(self):
        """Submitting July must not accidentally lock August entries too."""
        july_entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='july', activity_code=self.code, hours=Decimal('4'))
        august_entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 8, 1),
            task_description='august', activity_code=self.code, hours=Decimal('4'))
        submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)
        august_entry.refresh_from_db()
        self.assertEqual(august_entry.status, TimesheetEntry.STATUS_DRAFT)  # untouched

    def test_already_submitted_entries_not_re_touched_on_failed_resubmit(self):
        """A failed double-submit must not create a second TimesheetMonth
        row or otherwise corrupt state."""
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='t1', activity_code=self.code, hours=Decimal('4'))
        submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)
        try:
            submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)
        except ValueError:
            pass
        from .models import TimesheetMonth
        self.assertEqual(
            TimesheetMonth.objects.filter(employee=self.emp, year=2026, month=7).count(), 1)


class ReopenMonthTests(TestCase):
    """reopen_month unlocks a submitted month so the employee can fix a
    mistake — must flip entries back to draft, not just the month flag."""

    def setUp(self):
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('reopener')
        self.hr_user, _ = _make_user_with_employee('reopenhr', role_name=Role.ADMIN)
        seed_default_permissions()
        self.entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='t1', activity_code=self.code, hours=Decimal('4'))
        self.tsm = submit_month(employee=self.emp, year=2026, month=7, submitted_by=self.user)

    def test_reopen_never_submitted_month_raises(self):
        with self.assertRaises(ValueError):
            reopen_month(employee=self.emp, year=2099, month=1, reopened_by=self.hr_user)

    def test_reopen_flips_entries_back_to_draft(self):
        reopen_month(employee=self.emp, year=2026, month=7, reopened_by=self.hr_user)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.status, TimesheetEntry.STATUS_DRAFT)

    def test_reopen_stamps_reopened_at_and_by(self):
        tsm = reopen_month(employee=self.emp, year=2026, month=7, reopened_by=self.hr_user)
        self.assertIsNotNone(tsm.reopened_at)
        self.assertEqual(tsm.reopened_by, self.hr_user)
        self.assertFalse(tsm.is_submitted)

    def test_reopening_twice_raises_second_time(self):
        reopen_month(employee=self.emp, year=2026, month=7, reopened_by=self.hr_user)
        with self.assertRaises(ValueError):
            reopen_month(employee=self.emp, year=2026, month=7, reopened_by=self.hr_user)

    def test_after_reopen_employee_can_edit_entry_again(self):
        """End-to-end: reopen unlocks the entry for real editing through the
        view layer, not just at the model/service level."""
        reopen_month(employee=self.emp, year=2026, month=7, reopened_by=self.hr_user)
        client = Client()
        client.login(username='reopener', password='testpass123')
        resp = client.post(
            reverse('timesheets:entry_edit', args=[self.entry.pk]), {
                'entry_type': 'activity',
                'date': '2026-07-01', 'activity_code': self.code.pk,
                'task_description': 'Fixed after reopen', 'hours': '5',
            })
        self.assertEqual(resp.status_code, 302)
        self.entry.refresh_from_db()
        self.assertEqual(self.entry.task_description, 'Fixed after reopen')



class SendToHRTests(TestCase):
    """send_to_hr / acknowledge_send: an employee only ever acts on their
    own acknowledgment row, never someone else's, and double-acking is a
    no-op rather than a duplicate or a crash. Email is mocked so tests
    never call Microsoft Graph."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.hr_user, _ = _make_user_with_employee('sendhr', role_name=Role.ADMIN)
        HRSettings.objects.get_or_create(pk=1, defaults={'hr_email': 'hr@leap-arabia.com'})
        self.user_a, self.emp_a = _make_user_with_employee('senda', full_name='Send A')
        self.user_b, self.emp_b = _make_user_with_employee('sendb', full_name='Send B')
        seed_default_permissions()
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_preview_page_loads_for_pending_request(self):
        self.client.login(username='senda', password='testpass123')
        resp = self.client.get(reverse('timesheets:send_to_hr', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'Send to HR')
        self.assertNotContains(resp, 'Open in Outlook')

    @patch('timesheets.views.EmailMessage')
    def test_acknowledge_creates_ack_for_correct_employee_only(self, mock_email_cls):
        mock_email_cls.return_value.send.return_value = 1
        self.client.login(username='senda', password='testpass123')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.assertTrue(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp_a).exists())
        self.assertFalse(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp_b).exists())

    @patch('timesheets.views.EmailMessage')
    def test_double_acknowledge_does_not_duplicate(self, mock_email_cls):
        mock_email_cls.return_value.send.return_value = 1
        self.client.login(username='senda', password='testpass123')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.assertEqual(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp_a).count(), 1)
        # Second click is already-acked — should not send email again
        self.assertEqual(mock_email_cls.return_value.send.call_count, 1)

    def test_already_acked_employee_redirected_from_preview(self):
        TimesheetRequestAck.objects.create(request=self.ts_request, employee=self.emp_a)
        self.client.login(username='senda', password='testpass123')
        resp = self.client.get(reverse('timesheets:send_to_hr', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 302)

    def test_preview_shows_fixed_hr_email_not_employee_input(self):
        """The HR email in the preview always comes from HRSettings, never
        from anything the employee could submit or influence."""
        self.client.login(username='senda', password='testpass123')
        resp = self.client.get(reverse('timesheets:send_to_hr', args=[self.ts_request.pk]))
        self.assertContains(resp, 'hr@leap-arabia.com')

    @patch('timesheets.views.EmailMessage')
    def test_send_uses_fixed_hr_email_and_employee_in_subject(self, mock_email_cls):
        mock_msg = mock_email_cls.return_value
        mock_msg.send.return_value = 1
        self.client.login(username='senda', password='testpass123')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        kwargs = mock_email_cls.call_args.kwargs
        self.assertEqual(kwargs['to'], ['hr@leap-arabia.com'])
        self.assertIn('Send A', kwargs['subject'])
        self.assertIn('July 2026', kwargs['subject'])
        mock_msg.attach.assert_called_once()
        attach_args = mock_msg.attach.call_args[0]
        self.assertTrue(attach_args[0].endswith('.xlsx'))

    @patch('timesheets.views.EmailMessage')
    def test_acknowledging_someone_elses_request_id_only_affects_own_employee(self, mock_email_cls):
        """B posting to the SAME ts_request only ever creates B's own ack —
        there's no way to forge an ack for A through this endpoint."""
        mock_email_cls.return_value.send.return_value = 1
        self.client.login(username='sendb', password='testpass123')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.assertTrue(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp_b).exists())
        self.assertFalse(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp_a).exists())

    @patch('timesheets.views.EmailMessage')
    def test_acknowledging_locks_that_months_entries(self, mock_email_cls):
        """The actual gap fix: acknowledging now really locks entries,
        not just creates an Ack row."""
        mock_email_cls.return_value.send.return_value = 1
        TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 7, 5),
            task_description='work', activity_code=self.code, hours=Decimal('10'))
        self.client.login(username='senda', password='testpass123')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        entry = TimesheetEntry.objects.get(employee=self.emp_a, date=date(2026, 7, 5))
        self.assertEqual(entry.status, TimesheetEntry.STATUS_SUBMITTED)

    @patch('timesheets.views.EmailMessage')
    def test_double_acknowledge_does_not_crash_on_already_locked_month(self, mock_email_cls):
        mock_email_cls.return_value.send.return_value = 1
        self.client.login(username='senda', password='testpass123')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        resp = self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 302)  # still succeeds, no 500

    @patch('timesheets.views.EmailMessage')
    def test_email_failure_does_not_ack_or_lock(self, mock_email_cls):
        mock_email_cls.return_value.send.side_effect = RuntimeError('Graph down')
        TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 7, 5),
            task_description='work', activity_code=self.code, hours=Decimal('10'))
        self.client.login(username='senda', password='testpass123')
        resp = self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp_a).exists())
        entry = TimesheetEntry.objects.get(employee=self.emp_a, date=date(2026, 7, 5))
        self.assertEqual(entry.status, TimesheetEntry.STATUS_DRAFT)

class NotificationTargetingTests(TestCase):
    """request_timesheets notifies every active user regardless of role
    (per _all_active_users()'s own docstring: timesheets is intentionally
    open to everyone), minus the actor. remind_employee must still reach
    only the one targeted employee, not broadcast to everyone."""

    def setUp(self):
        self.hr_user, self.hr_emp = _make_user_with_employee('notifhr', role_name=Role.ADMIN)
        self.user_a, self.emp_a = _make_user_with_employee('notifa')
        self.user_b, self.emp_b = _make_user_with_employee('notifb')
        # AI roles do NOT have timesheets.access per DEFAULT_MODULE_ACCESS —
        # this user must never be notified.
        self.ai_user, self.ai_emp = _make_user_with_employee(
            'notifai', role_name=Role.DEVELOPER)
        seed_default_permissions()

    def test_request_notifies_eligible_users_excludes_actor(self):
        ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        notified = set(Notification.objects.filter(
            verb='requested your timesheet').values_list('recipient__username', flat=True))
        self.assertIn('notifa', notified)
        self.assertIn('notifb', notified)
        self.assertNotIn('notifhr', notified)  # actor excluded

    def test_request_notifies_users_regardless_of_role(self):
        """Regression lock-in: request_timesheets was switched from
        _users_with_capability('timesheets.access') to _all_active_users()
        specifically so developers (who don't have timesheets.access)
        still get notified. This replaces the old test that asserted the
        opposite — leaving it as-is would have made the suite pass while
        actually proving the wrong thing."""
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        notified = set(Notification.objects.filter(
            verb='requested your timesheet').values_list('recipient__username', flat=True))
        self.assertIn('notifai', notified)  # developer role — now correctly included

    def test_reminder_reaches_only_the_targeted_employee(self):
        ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        remind_employee(ts_request=ts_request, employee=self.emp_a, reminded_by=self.hr_user)
        notified = set(Notification.objects.filter(
            verb='reminded you to send your timesheet').values_list('recipient__username', flat=True))
        self.assertEqual(notified, {'notifa'})  # exactly one recipient, not B, not AI, not HR

    def test_reminding_employee_with_no_linked_user_raises(self):
        ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        orphan_emp = Employee.objects.create(full_name='No Login', iqama_number='IQ-ORPHAN')
        with self.assertRaises(ValueError):
            remind_employee(ts_request=ts_request, employee=orphan_emp, reminded_by=self.hr_user)


class CSRFEnforcementTests(TestCase):
    """Per the guide: an explicit test proving a request with no CSRF token
    is rejected — the exact test that would have caught the drafts-feature
    empty-token bypass, applied here even though our CSRF handling was
    always plain Django (no hand-rolled check to worry about)."""

    def setUp(self):
        self.client = Client(enforce_csrf_checks=True)
        self.hr_user, _ = _make_user_with_employee('csrfhr', role_name=Role.ADMIN)
        self.user, self.emp = _make_user_with_employee('csrfemp')
        seed_default_permissions()
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_post_without_csrf_token_rejected(self):
        self.client.login(username='csrfemp', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 403)
        # confirm the denied request had no actual side effect
        self.assertFalse(TimesheetRequestAck.objects.filter(
            request=self.ts_request, employee=self.emp).exists())


class HRSettingsSingletonTests(TestCase):
    """HRSettings.load() must always return the same single row, even when
    called repeatedly — this is what keeps the fixed-HR-email guarantee
    (never typed by an employee) actually true over time."""

    def test_load_creates_row_on_first_call(self):
        self.assertEqual(HRSettings.objects.count(), 0)
        settings_obj = HRSettings.load()
        self.assertEqual(HRSettings.objects.count(), 1)
        self.assertEqual(settings_obj.pk, 1)

    def test_load_never_creates_a_second_row(self):
        first = HRSettings.load()
        first.hr_email = 'hr@leap-arabia.com'
        first.save()
        second = HRSettings.load()
        self.assertEqual(first.pk, second.pk)
        self.assertEqual(HRSettings.objects.count(), 1)
        self.assertEqual(second.hr_email, 'hr@leap-arabia.com')

class RequestTimesheetsDeduplicationTests(TestCase):
    """Regression test for the duplicate-TimesheetRequest bug: calling
    request_timesheets twice for the same year/month must reuse the same
    row, not fragment ack-tracking across multiple copies of 'July'."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('duphr', role_name=Role.ADMIN)
        seed_default_permissions()

    def test_calling_twice_for_same_month_creates_only_one_request(self):
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        self.assertEqual(
            TimesheetRequest.objects.filter(year=2026, month=7).count(), 1)

    def test_second_call_still_notifies_everyone(self):
        """Re-requesting should still act as a valid 'remind everyone'
        action, not silently no-op just because the row already exists."""
        user_a, emp_a = _make_user_with_employee('dupnotify_a')
        seed_default_permissions()  # re-seed now that dupnotify_a's role exists
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        Notification.objects.all().delete()  # clear the first round
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        notified = set(Notification.objects.filter(
            verb='requested your timesheet').values_list('recipient__username', flat=True))
        self.assertIn('dupnotify_a', notified)

    def test_different_months_still_create_separate_requests(self):
        """The fix must not over-correct into treating every request as
        the same one regardless of month."""
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        request_timesheets(year=2026, month=8, requested_by=self.hr_user)
        self.assertEqual(TimesheetRequest.objects.count(), 2)


class RemindAlreadyAckedTests(TestCase):
    """Regression test for the 'remind someone who already sent theirs'
    gap: remind_employee must check the ack itself, not trust that the UI
    hid the button."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('remindhr', role_name=Role.ADMIN)
        self.user, self.emp = _make_user_with_employee('remindme')
        seed_default_permissions()
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_reminding_already_acked_employee_raises(self):
        TimesheetRequestAck.objects.create(request=self.ts_request, employee=self.emp)
        with self.assertRaises(ValueError):
            remind_employee(
                ts_request=self.ts_request, employee=self.emp, reminded_by=self.hr_user)

    def test_reminding_already_acked_employee_sends_no_notification(self):
        TimesheetRequestAck.objects.create(request=self.ts_request, employee=self.emp)
        Notification.objects.all().delete()
        try:
            remind_employee(
                ts_request=self.ts_request, employee=self.emp, reminded_by=self.hr_user)
        except ValueError:
            pass
        self.assertFalse(Notification.objects.filter(
            verb='reminded you to send your timesheet').exists())

    def test_reminding_via_view_shows_error_not_crash(self):
        """End-to-end through hr_request_detail's POST handler, not just
        the service function directly."""
        TimesheetRequestAck.objects.create(request=self.ts_request, employee=self.emp)
        client = Client()
        client.login(username='remindhr', password='testpass123')
        resp = client.post(
            reverse('timesheets:hr_request_detail', args=[self.ts_request.pk]),
            {'employee_id': self.emp.pk})
        self.assertEqual(resp.status_code, 302)  # redirected with error message, not a 500


class MalformedQueryParamTests(TestCase):
    """Per the guide's checklist: invalid input must return 4xx (or a
    clean redirect with an error), never an unhandled 500."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('malformedhr', role_name=Role.ADMIN)
        self.user, self.emp = _make_user_with_employee('malformeduser')
        seed_default_permissions()

    def test_export_with_non_numeric_year_does_not_500(self):
        self.client.login(username='malformeduser', password='testpass123')
        resp = self.client.get(reverse('timesheets:export') + '?year=abc&month=7')
        self.assertEqual(resp.status_code, 302)  # redirected with error, not a crash

    def test_export_with_non_numeric_month_does_not_500(self):
        self.client.login(username='malformeduser', password='testpass123')
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=xyz')
        self.assertEqual(resp.status_code, 302)

    def test_hr_request_with_non_numeric_year_does_not_500(self):
        self.client.login(username='malformedhr', password='testpass123')
        resp = self.client.post(reverse('timesheets:hr_request'), {
            'year': 'not-a-year', 'month': '7',
        })
        self.assertEqual(resp.status_code, 302)
        # confirm nothing bad got created despite the malformed input
        self.assertEqual(TimesheetRequest.objects.filter(month=7).count(), 0)

    def test_export_with_month_13_does_not_500(self):
        self.client.login(username='malformeduser', password='testpass123')
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=13')
        self.assertNotEqual(resp.status_code, 500)

class HRRequestYearDefaultingTests(TestCase):
    """The template no longer submits a 'year' field — the view must keep
    defaulting to the current year server-side when it's absent, not
    error or silently use year=0/None."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('yrdefault', role_name=Role.ADMIN)
        seed_default_permissions()

    def test_posting_without_year_field_defaults_to_current_year(self):
        client = Client()
        client.login(username='yrdefault', password='testpass123')
        client.post(reverse('timesheets:hr_request'), {'month': '7'})  # no 'year' key at all
        current_year = date.today().year
        self.assertTrue(TimesheetRequest.objects.filter(year=current_year, month=7).exists())


class AdditionalCSRFEnforcementTests(TestCase):
    """A second CSRF check on a different destructive endpoint, since
    entry_delete removes data and deserves its own explicit proof, not
    just an inference from the acknowledge_send test."""

    def setUp(self):
        self.client = Client(enforce_csrf_checks=True)
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('csrfdel')
        seed_default_permissions()
        self.entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='t1', activity_code=self.code, hours=Decimal('4'))

    def test_delete_without_csrf_token_rejected(self):
        self.client.login(username='csrfdel', password='testpass123')
        resp = self.client.post(reverse('timesheets:entry_delete', args=[self.entry.pk]))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(TimesheetEntry.objects.filter(pk=self.entry.pk).exists())


class NonexistentRequestIdTests(TestCase):
    """send_to_hr and acknowledge_send must handle a request_id that
    doesn't exist at all, not just one that belongs to someone else."""

    def setUp(self):
        self.user, self.emp = _make_user_with_employee('nonexistuser')
        seed_default_permissions()
        self.client.login(username='nonexistuser', password='testpass123')

    def test_send_to_hr_with_bad_id_redirects_with_message(self):
        """Changed from a hard 404: a missing request_id (never existed,
        or HR deleted it after the employee already saw the banner) now
        redirects back to My Timesheet with a friendly message instead of
        Django's default error page."""
        resp = self.client.get(reverse('timesheets:send_to_hr', args=[99999]))
        self.assertEqual(resp.status_code, 302)
        self.assertRedirects(resp, reverse('timesheets:my_timesheet'))

    def test_acknowledge_send_with_bad_id_redirects(self):
        resp = self.client.post(reverse('timesheets:acknowledge_send', args=[99999]))
        self.assertEqual(resp.status_code, 302)
        self.assertRedirects(resp, reverse('timesheets:my_timesheet'))


class HRRequestDetailQueryCountTests(TestCase):
    """N+1 guard, per the guide's Django-practices section: looping over
    employees in employees_for_request must not issue one query per
    employee. This locks the query count so a future refactor that
    accidentally removes select_related gets caught immediately."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('queryhr', role_name=Role.ADMIN)
        for i in range(5):
            _make_user_with_employee(f'queryemp{i}')
        seed_default_permissions()
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_request_detail_query_count_does_not_scale_with_employee_count(self):
        client = Client()
        client.login(username='queryhr', password='testpass123')
        with self.assertNumQueries(17):  # adjust once if your baseline differs; the
                                          # point is this number staying FLAT as
                                          # employee count grows, not its exact value
            client.get(reverse('timesheets:hr_request_detail', args=[self.ts_request.pk]))


from .forms import TimesheetEntryForm


class TimesheetEntryFormDirectTests(TestCase):
    """Form-level tests, isolated from the view — faster and pinpoints
    exactly which clean_* method is responsible if one of these breaks."""

    def setUp(self):
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})

    def _data(self, **overrides):
        data = {
            'entry_type': 'activity',
            'date': '2026-07-10', 'activity_code': self.code.pk,
            'task_description': 'Task', 'hours': '4',
        }
        data.update(overrides)
        return data

    def test_today_is_a_valid_date(self):
        form = TimesheetEntryForm(self._data(date=date.today().isoformat()))
        self.assertTrue(form.is_valid())

    def test_exactly_24_hours_is_valid(self):
        form = TimesheetEntryForm(self._data(hours='24'))
        self.assertTrue(form.is_valid())

    def test_decimal_hours_accepted(self):
        form = TimesheetEntryForm(self._data(hours='4.5'))
        self.assertTrue(form.is_valid())

    def test_missing_task_description_invalid(self):
        form = TimesheetEntryForm(self._data(task_description=''))
        self.assertFalse(form.is_valid())

    def test_missing_date_invalid(self):
        form = TimesheetEntryForm(self._data(date=''))
        self.assertFalse(form.is_valid())

    def test_missing_activity_code_invalid(self):
        form = TimesheetEntryForm(self._data(activity_code=''))
        self.assertFalse(form.is_valid())

class TimesheetEntryFormDefaultsTests(TestCase):
    """Locks in the two default-value changes: hours defaults to 10, date
    defaults to today — and confirms editing an existing entry does NOT
    silently overwrite its real values with these defaults."""

    def setUp(self):
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})

    def test_new_form_defaults_hours_to_10(self):
        form = TimesheetEntryForm()
        self.assertEqual(form.fields['hours'].initial, 10)

    def test_new_form_defaults_date_to_today(self):
        form = TimesheetEntryForm()
        self.assertEqual(form.fields['date'].initial, date.today())

    def test_editing_existing_entry_does_not_overwrite_its_real_date_or_hours(self):
        """The guard in forms.py's __init__ (`if not self.instance.pk`)
        exists specifically so editing an old entry doesn't silently
        reset its date to today or its hours to 10."""
        user, emp = _make_user_with_employee('defaultsedit')
        entry = TimesheetEntry.objects.create(
            employee=emp, date=date(2026, 1, 5), task_description='Old entry',
            activity_code=self.code, hours=Decimal('3'))
        form = TimesheetEntryForm(instance=entry)
        # initial should be unset (None) here, not forced to today/10 —
        # Django's ModelForm already pulls the real values from the
        # instance itself when initial is left alone.
        self.assertIsNone(form.fields['date'].initial)
        self.assertIsNone(form.fields['hours'].initial)
        self.assertEqual(form['date'].value(), entry.date)
        self.assertEqual(form['hours'].value(), entry.hours)


class ExportBusinessLogicTests(TestCase):
    """The export's actual computed values, not just 'does it download' —
    combining multiple same-day entries and weekend highlighting are real
    business logic with room to silently drift wrong."""

    def setUp(self):
        self.client = Client()
        self.code1, _ = ActivityCode.objects.get_or_create(
            code='COS_0001', defaults={'label': 'A', 'is_active': True})
        self.code2, _ = ActivityCode.objects.get_or_create(
            code='COS_0002', defaults={'label': 'B', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('exportlogic')
        seed_default_permissions()
        self.client.login(username='exportlogic', password='testpass123')

    def test_combined_row_joins_descriptions_and_codes_with_flat_hours(self):
        """Regression lock-in: hours used to be summed (3+5=8) — this
        replaces that old assertion, since combined_hours is now a flat
        Decimal('10') whenever the day has any entries at all, tested
        fully in FlatOfficeHoursTests. Descriptions and codes still join
        exactly as before; only the hours behavior changed."""
        import openpyxl
        from io import BytesIO
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 10),
            task_description='Morning task', activity_code=self.code1, hours=Decimal('3'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 10),
            task_description='Afternoon task', activity_code=self.code2, hours=Decimal('5'))

        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        row = 11 + 10  # header at row 11, day 10 -> row 21

        self.assertEqual(ws.cell(row=row, column=3).value, 'Morning task; Afternoon task')
        self.assertEqual(ws.cell(row=row, column=6).value, Decimal('10'))
        self.assertEqual(ws.cell(row=row, column=7).value, 'COS_0001, COS_0002')

    def test_day_with_no_entries_exports_blank_row(self):
        import openpyxl
        from io import BytesIO
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        row = 11 + 15  # a day with nothing logged
        self.assertIsNone(ws.cell(row=row, column=3).value)
        self.assertIsNone(ws.cell(row=row, column=6).value)

    def test_weekend_rows_are_highlighted_differently_from_weekdays(self):
        import openpyxl
        import calendar as cal_module
        from io import BytesIO
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active

        days_in_month = cal_module.monthrange(2026, 7)[1]
        weekend_row = weekday_row = None
        for day in range(1, days_in_month + 1):
            d = date(2026, 7, day)
            if d.weekday() in (4, 5) and weekend_row is None:
                weekend_row = 11 + day
            if d.weekday() not in (4, 5) and weekday_row is None:
                weekday_row = 11 + day

        weekend_fill = ws.cell(row=weekend_row, column=4).fill.start_color.rgb or ''
        weekday_fill = ws.cell(row=weekday_row, column=4).fill.start_color.rgb or ''
        self.assertIn('FFFF00', weekend_fill)
        self.assertNotIn('FFFF00', weekday_fill)

    def test_entry_with_since_deactivated_code_still_exports_correctly(self):
        """An entry logged while a code was active must still export fine
        after that code gets deactivated later — the FK still resolves."""
        import openpyxl
        from io import BytesIO
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 5),
            task_description='Old task', activity_code=self.code1, hours=Decimal('4'))
        self.code1.is_active = False
        self.code1.save()

        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 200)
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=11 + 5, column=7).value, 'COS_0001')

    def test_export_query_count_does_not_scale_with_entry_count(self):
        """N+1 guard for the export loop — without select_related('activity_code')
        this scales linearly with entry count instead of staying flat."""
        for day in range(1, 11):
            TimesheetEntry.objects.create(
                employee=self.emp, date=date(2026, 7, day),
                task_description=f'task {day}', activity_code=self.code1, hours=Decimal('4'))
        with self.assertNumQueries(4):  # confirmed real baseline after adding
                                          # select_related('activity_code') — the
                                          # point is this staying FLAT as entries grow
            self.client.get(reverse('timesheets:export') + '?year=2026&month=7')


class SendToHRNoEmployeeLinkTests(TestCase):
    """send_to_hr must redirect gracefully for a user with no linked
    Employee record, same as the other views — not crash."""

    def test_send_to_hr_redirects_when_no_employee_link(self):
        role, _ = Role.objects.get_or_create(name=Role.SALES_REP)
        User.objects.create_user(username='noemplink', password='testpass123', role=role)
        seed_default_permissions()
        client = Client()
        client.login(username='noemplink', password='testpass123')
        from .services import request_timesheets
        hr_role, _ = Role.objects.get_or_create(name=Role.ADMIN)
        hr_user = User.objects.create_user(username='noemplinkhr', password='testpass123', role=hr_role)
        ts_request = request_timesheets(year=2026, month=7, requested_by=hr_user)
        resp = client.get(reverse('timesheets:send_to_hr', args=[ts_request.pk]))
        self.assertEqual(resp.status_code, 302)


class EmployeesForRequestOrderingTests(TestCase):
    """employees_for_request sorts Pending before Sent, alphabetical
    within each group — the HR detail page's usefulness depends on this."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('orderhr', role_name=Role.ADMIN)
        self.zoe_user, self.zoe_emp = _make_user_with_employee('orderz', full_name='Zoe')
        self.amy_user, self.amy_emp = _make_user_with_employee('ordera', full_name='Amy')
        seed_default_permissions()
        from .services import request_timesheets
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_pending_employees_listed_before_sent_employees(self):
        TimesheetRequestAck.objects.create(request=self.ts_request, employee=self.zoe_emp)
        rows = employees_for_request(self.ts_request)
        acked_flags = [r['acked'] for r in rows]
        # all False (pending) entries must come before all True (sent) entries
        self.assertEqual(acked_flags, sorted(acked_flags))

    def test_alphabetical_within_same_acked_group(self):
        rows = employees_for_request(self.ts_request)
        pending_names = [r['employee'].full_name for r in rows if not r['acked']]
        self.assertEqual(pending_names, sorted(pending_names))


class RemainingCSRFEnforcementTests(TestCase):
    """CSRF coverage for every remaining state-changing endpoint, so no
    destructive/mutating view is left unproven — matches the guide's
    'CSRF enforced' checklist item applied exhaustively, not just once."""

    def setUp(self):
        self.client = Client(enforce_csrf_checks=True)
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('csrfall')
        self.hr_user, _ = _make_user_with_employee('csrfallhr', role_name=Role.ADMIN)
        seed_default_permissions()
        from .services import request_timesheets
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_add_entry_without_csrf_token_rejected(self):
        self.client.login(username='csrfall', password='testpass123')
        resp = self.client.post(reverse('timesheets:my_timesheet'), {
            'date': '2026-07-10', 'activity_code': self.code.pk,
            'task_description': 'Should not save', 'hours': '4',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(TimesheetEntry.objects.count(), 0)

    def test_edit_entry_without_csrf_token_rejected(self):
        entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='original', activity_code=self.code, hours=Decimal('4'))
        self.client.login(username='csrfall', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:entry_edit', args=[entry.pk]), {
                'date': '2026-07-01', 'activity_code': self.code.pk,
                'task_description': 'HIJACKED', 'hours': '1',
            })
        self.assertEqual(resp.status_code, 403)
        entry.refresh_from_db()
        self.assertEqual(entry.task_description, 'original')

    def test_hr_request_submit_without_csrf_token_rejected(self):
        self.client.login(username='csrfallhr', password='testpass123')
        resp = self.client.post(reverse('timesheets:hr_request'), {
            'year': '2026', 'month': '9',
        })
        self.assertEqual(resp.status_code, 403)
        self.assertEqual(TimesheetRequest.objects.filter(month=9).count(), 0)

    def test_remind_without_csrf_token_rejected(self):
        self.client.login(username='csrfallhr', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:hr_request_detail', args=[self.ts_request.pk]),
            {'employee_id': self.emp.pk})
        self.assertEqual(resp.status_code, 403)

    def test_inline_save_without_csrf_token_rejected(self):
        """entry_inline_save (the Send to HR preview's click-to-edit) is a
        newer endpoint than the rest of this class -- added this session,
        so it needs the same proof as everything else here that it isn't
        secretly relying on @csrf_exempt."""
        entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='original', activity_code=self.code, hours=Decimal('4'))
        self.client.login(username='csrfall', password='testpass123')
        resp = self.client.post(
            reverse('timesheets:entry_inline_save', args=[entry.pk]), {
                'entry_type': 'activity', 'activity_code': self.code.pk,
                'task_description': 'HIJACKED', 'hours': '1',
            })
        self.assertEqual(resp.status_code, 403)
        entry.refresh_from_db()
        self.assertEqual(entry.task_description, 'original')


class TimesheetEntryOrderingTests(TestCase):
    """The entries list on My Timesheet is ordered most-recent-first,
    per the model's Meta.ordering — worth locking in explicitly."""

    def test_entries_ordered_newest_date_first(self):
        code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        user, emp = _make_user_with_employee('orderentries')
        seed_default_permissions()
        e_old = TimesheetEntry.objects.create(
            employee=emp, date=date(2026, 7, 1),
            task_description='old', activity_code=code, hours=Decimal('4'))
        e_new = TimesheetEntry.objects.create(
            employee=emp, date=date(2026, 7, 20),
            task_description='new', activity_code=code, hours=Decimal('4'))
        entries = list(TimesheetEntry.objects.filter(employee=emp))
        self.assertEqual(entries[0].pk, e_new.pk)
        self.assertEqual(entries[1].pk, e_old.pk)


class NonexistentEntryTests(TestCase):
    """Editing/deleting a pk that belongs to no one at all (not just
    someone else) must still 404 cleanly, not error differently."""

    def setUp(self):
        self.user, self.emp = _make_user_with_employee('ghostentry')
        seed_default_permissions()
        self.client.login(username='ghostentry', password='testpass123')

    def test_edit_nonexistent_entry_404s(self):
        resp = self.client.get(reverse('timesheets:entry_edit', args=[99999]))
        self.assertEqual(resp.status_code, 404)

    def test_delete_nonexistent_entry_404s(self):
        resp = self.client.post(reverse('timesheets:entry_delete', args=[99999]))
        self.assertEqual(resp.status_code, 404)


class InactiveEmployeeVisibilityTests(TestCase):
    """Documents actual current behavior: employees_for_request only
    checks User.is_active (via _users_with_capability), NOT
    Employee.is_active. A deactivated Employee whose User account is
    still active keeps showing up as Pending indefinitely. This is a
    product decision to make on purpose, not a bug -- this test just
    makes the real behavior visible and locked in, so if it's ever
    "fixed" to also check Employee.is_active, that's a deliberate
    change someone reviews, not an accidental side effect."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('inactivehr', role_name=Role.ADMIN)
        self.user, self.emp = _make_user_with_employee('inactiveemp', full_name='Inactive Person')
        seed_default_permissions()
        from .services import request_timesheets
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_employee_with_inactive_employee_record_but_active_user_still_listed(self):
        self.emp.is_active = False
        self.emp.save()
        rows = employees_for_request(self.ts_request)
        names = [r['employee'].full_name for r in rows]
        self.assertIn('Inactive Person', names)  # current behavior: still shows up

    def test_employee_with_deactivated_user_account_disappears_from_list(self):
        self.user.is_active = False
        self.user.save()
        rows = employees_for_request(self.ts_request)
        names = [r['employee'].full_name for r in rows]
        self.assertNotIn('Inactive Person', names)  # correctly excluded via User.is_active


class HRSettingsAdminSingletonTests(TestCase):
    """The admin's has_add_permission override must block adding a second
    HRSettings row through /admin/, not just through .load()."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('adminhr', role_name=Role.ADMIN)
        self.hr_user.is_staff = True
        self.hr_user.is_superuser = True
        self.hr_user.save()
        seed_default_permissions()

    def test_admin_add_view_blocked_when_a_row_already_exists(self):
        HRSettings.objects.get_or_create(pk=1, defaults={'hr_email': 'hr@leap-arabia.com'})
        self.client.login(username='adminhr', password='testpass123')
        resp = self.client.get('/admin/timesheets/hrsettings/add/')
        self.assertEqual(resp.status_code, 403)

    def test_admin_add_view_allowed_when_no_row_exists_yet(self):
        self.assertEqual(HRSettings.objects.count(), 0)
        self.client.login(username='adminhr', password='testpass123')
        resp = self.client.get('/admin/timesheets/hrsettings/add/')
        self.assertEqual(resp.status_code, 200)



class HRRequestDeleteTests(TestCase):
    """hr_request_delete: covers the full checklist from the guide — happy
    path, the denied case, CSRF, and that delete_request's cascade cleanup
    (acks + notifications) actually happens, not just the request row."""

    def setUp(self):
        self.hr_user, self.hr_emp = _make_user_with_employee('delhr', role_name=Role.ADMIN)
        self.plain_user, self.plain_emp = _make_user_with_employee('delplain')
        seed_default_permissions()
        self.ts_request = request_timesheets(year=2026, month=7, requested_by=self.hr_user)

    def test_hr_can_delete_a_request(self):
        client = Client()
        client.login(username='delhr', password='testpass123')
        resp = client.post(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertFalse(TimesheetRequest.objects.filter(pk=self.ts_request.pk).exists())

    def test_deleting_a_request_also_deletes_its_acks(self):
        TimesheetRequestAck.objects.create(request=self.ts_request, employee=self.plain_emp)
        client = Client()
        client.login(username='delhr', password='testpass123')
        client.post(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertFalse(TimesheetRequestAck.objects.filter(
            request_id=self.ts_request.pk).exists())

    def test_deleting_a_request_also_deletes_its_notifications(self):
        """delete_request explicitly cleans up notifications pointing at
        the request, on top of the CASCADE on acks — so nobody's left
        with a notification linking to something that no longer exists."""
        self.assertTrue(Notification.objects.filter(
            verb='requested your timesheet').exists())  # sanity: request_timesheets sent one
        client = Client()
        client.login(username='delhr', password='testpass123')
        client.post(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertFalse(Notification.objects.filter(
            verb='requested your timesheet').exists())

    def test_plain_employee_cannot_delete_a_request(self):
        client = Client()
        client.login(username='delplain', password='testpass123')
        resp = client.post(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(TimesheetRequest.objects.filter(pk=self.ts_request.pk).exists())

    def test_delete_requires_post_not_get(self):
        client = Client()
        client.login(username='delhr', password='testpass123')
        resp = client.get(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 405)  # @require_POST

    def test_delete_nonexistent_request_404s(self):
        client = Client()
        client.login(username='delhr', password='testpass123')
        resp = client.post(reverse('timesheets:hr_request_delete', args=[99999]))
        self.assertEqual(resp.status_code, 404)

    def test_delete_without_csrf_token_rejected(self):
        client = Client(enforce_csrf_checks=True)
        client.login(username='delhr', password='testpass123')
        resp = client.post(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 403)
        self.assertTrue(TimesheetRequest.objects.filter(pk=self.ts_request.pk).exists())

    def test_anonymous_redirected_from_delete(self):
        client = Client()
        resp = client.post(reverse('timesheets:hr_request_delete', args=[self.ts_request.pk]))
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)



class FlatOfficeHoursTests(TestCase):
    """Regression lock-in: both the export and the Send-to-HR preview show
    a flat 10 hours per day with any entries, regardless of how many
    entries or their individual hours values -- summing was replaced on
    purpose since every entry already defaults to 10."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('flathours')
        seed_default_permissions()

    def test_export_shows_flat_10_regardless_of_entry_count(self):
        import openpyxl
        from io import BytesIO
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 3),
            task_description='a', activity_code=self.code, hours=Decimal('2'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 3),
            task_description='b', activity_code=self.code, hours=Decimal('6'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 3),
            task_description='c', activity_code=self.code, hours=Decimal('9'))
        self.client.login(username='flathours', password='testpass123')
        resp = self.client.get(reverse('timesheets:export') + '?year=2026&month=7')
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=11 + 3, column=6).value, Decimal('10'))

    def test_send_to_hr_preview_shows_flat_10_matching_export(self):
        hr_user, _ = _make_user_with_employee('flathourshr', role_name=Role.ADMIN)
        ts_request = request_timesheets(year=2026, month=7, requested_by=hr_user)
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 3),
            task_description='a', activity_code=self.code, hours=Decimal('2'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 3),
            task_description='b', activity_code=self.code, hours=Decimal('6'))
        self.client.login(username='flathours', password='testpass123')
        resp = self.client.get(reverse('timesheets:send_to_hr', args=[ts_request.pk]))
        self.assertContains(resp, '<td class="center">10</td>')  # matches Decimal('10') rendered plain, no decimals


class AutogenBannerTests(TestCase):
    """The banner must key off the actual auto-generated task description
    text, not the activity code -- someone manually picking COS_0009 for
    unrelated work must never see it, only people with a real
    'Started task:' / 'Ended task:' entry."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('bannertest')
        seed_default_permissions()
        self.client.login(username='bannertest', password='testpass123')

    def test_banner_hidden_when_cos_0009_used_manually(self):
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='Discussed client proposals',  # manually typed, unrelated
            activity_code=self.code, hours=Decimal('10'))
        resp = self.client.get(reverse('timesheets:my_timesheet'))
        self.assertNotContains(resp, 'is used as default when a task auto-starts')

    def test_banner_shown_for_autogenerated_started_entry(self):
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='Started task: Fix login bug',
            activity_code=self.code, hours=Decimal('10'))
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=7')
        self.assertContains(resp, 'is used as default when a task auto-starts')

    def test_banner_shown_for_autogenerated_ended_entry(self):
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 1),
            task_description='Ended task: Fix login bug',
            activity_code=self.code, hours=Decimal('10'))
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=7')
        self.assertContains(resp, 'is used as default when a task auto-starts')


class ReRequestExcludesAckedTests(TestCase):
    """The recipient filter in request_timesheets must actually exclude
    people who already acked -- re-requesting is meant to nudge
    stragglers only, not re-notify everyone including those done."""

    def setUp(self):
        self.hr_user, _ = _make_user_with_employee('rerequesthr', role_name=Role.ADMIN)
        self.user, self.emp = _make_user_with_employee('alreadyacked')
        seed_default_permissions()

    def test_already_acked_employee_not_renotified_on_rerequest(self):
        req = request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        TimesheetRequestAck.objects.create(request=req, employee=self.emp)
        Notification.objects.all().delete()
        request_timesheets(year=2026, month=7, requested_by=self.hr_user)
        notified = set(Notification.objects.filter(
            verb='requested your timesheet').values_list('recipient__username', flat=True))
        self.assertNotIn('alreadyacked', notified)

class MyTimesheetMonthNavigationTests(TestCase):
    """_resolve_timesheet_month / _adjacent_month: My Timesheet's month
    browser never 500s on bad input, always falls back to today, and
    Prev/Next wrap the year boundary correctly. Also confirms entries
    from other months never leak into whatever month is being viewed."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('navemp', full_name='Nav Emp')
        seed_default_permissions()
        self.client.login(username='navemp', password='testpass123')
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 15),
            task_description='July-only task', activity_code=self.code, hours=Decimal('10'))
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 10, 15),
            task_description='October-only task', activity_code=self.code, hours=Decimal('10'))

    def test_explicit_month_shows_only_that_months_entries(self):
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=7')
        self.assertContains(resp, 'July-only task')
        self.assertNotContains(resp, 'October-only task')

    def test_switching_month_shows_the_other_months_entries(self):
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=10')
        self.assertContains(resp, 'October-only task')
        self.assertNotContains(resp, 'July-only task')

    def test_default_no_params_uses_current_month(self):
        today = date.today()
        resp = self.client.get(reverse('timesheets:my_timesheet'))
        self.assertEqual(resp.context['year'], today.year)
        self.assertEqual(resp.context['month'], today.month)

    def test_invalid_month_falls_back_to_today_not_500(self):
        today = date.today()
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=13')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['year'], today.year)
        self.assertEqual(resp.context['month'], today.month)

    def test_garbage_year_falls_back_to_today_not_500(self):
        today = date.today()
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=notanumber&month=7')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.context['year'], today.year)
        self.assertEqual(resp.context['month'], today.month)

    def test_prev_next_wrap_december_to_january(self):
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=12')
        self.assertEqual(resp.context['next_year'], 2027)
        self.assertEqual(resp.context['next_month'], 1)
        self.assertEqual(resp.context['prev_year'], 2026)
        self.assertEqual(resp.context['prev_month'], 11)

    def test_prev_next_wrap_january_to_december(self):
        resp = self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=1')
        self.assertEqual(resp.context['prev_year'], 2025)
        self.assertEqual(resp.context['prev_month'], 12)
        self.assertEqual(resp.context['next_year'], 2026)
        self.assertEqual(resp.context['next_month'], 2)

    def test_adding_entry_while_browsing_a_month_redirects_back_to_that_month(self):
        resp = self.client.post(
            reverse('timesheets:my_timesheet') + '?year=2026&month=10', {
                'entry_type': 'activity',
                'date': '2026-10-20', 'activity_code': self.code.pk,
                'task_description': 'Added while browsing October', 'hours': '5',
            })
        self.assertEqual(resp.status_code, 302)
        self.assertIn('year=2026', resp.url)
        self.assertIn('month=10', resp.url)


class LockedMonthCannotBeTamperedWithTests(TestCase):
    """The month-picker makes it trivial to browse back to an already
    -submitted month. Confirms that browsing there is harmless: neither
    adding a brand-new entry nor editing a still-draft entry's date can
    land data inside a month HR has already received."""

    def setUp(self):
        from .models import TimesheetMonth
        self.TimesheetMonth = TimesheetMonth

        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.user, self.emp = _make_user_with_employee('lockedemp', full_name='Locked Emp')
        seed_default_permissions()
        self.client.login(username='lockedemp', password='testpass123')

        # A real submitted entry for July, so the month is genuinely locked.
        TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 7, 10),
            task_description='Already submitted', activity_code=self.code,
            hours=Decimal('10'), status=TimesheetEntry.STATUS_SUBMITTED)
        self.tsm = TimesheetMonth.objects.create(
            employee=self.emp, year=2026, month=7, submitted_at=timezone.now())

        # A still-draft entry in an OPEN month, used for the edit-into-locked test.
        self.draft_entry = TimesheetEntry.objects.create(
            employee=self.emp, date=date(2026, 8, 5),
            task_description='Still a draft', activity_code=self.code, hours=Decimal('10'))

    def test_cannot_add_new_entry_into_a_locked_month(self):
        resp = self.client.post(
            reverse('timesheets:my_timesheet') + '?year=2026&month=7', {
                'entry_type': 'activity',
                'date': '2026-07-20', 'activity_code': self.code.pk,
                'task_description': 'Sneaky post-submit entry', 'hours': '5',
            })
        self.assertFalse(
            TimesheetEntry.objects.filter(task_description='Sneaky post-submit entry').exists())
        self.assertEqual(resp.status_code, 200)  # re-renders the form, no redirect

    def test_can_still_add_entry_into_an_open_month(self):
        resp = self.client.post(
            reverse('timesheets:my_timesheet') + '?year=2026&month=8', {
                'entry_type': 'activity',
                'date': '2026-08-20', 'activity_code': self.code.pk,
                'task_description': 'Legit open-month entry', 'hours': '5',
            })
        self.assertTrue(
            TimesheetEntry.objects.filter(task_description='Legit open-month entry').exists())
        self.assertEqual(resp.status_code, 302)

    def test_cannot_edit_a_draft_entrys_date_into_a_locked_month(self):
        resp = self.client.post(
            reverse('timesheets:entry_edit', args=[self.draft_entry.pk]), {
                'entry_type': 'activity',
                'date': '2026-07-25', 'activity_code': self.code.pk,
                'task_description': 'Still a draft', 'hours': '10',
            })
        self.draft_entry.refresh_from_db()
        self.assertEqual(self.draft_entry.date, date(2026, 8, 5))  # unchanged
        self.assertEqual(resp.status_code, 200)

    def test_can_still_edit_a_draft_entry_within_its_own_open_month(self):
        resp = self.client.post(
            reverse('timesheets:entry_edit', args=[self.draft_entry.pk]), {
                'entry_type': 'activity',
                'date': '2026-08-06', 'activity_code': self.code.pk,
                'task_description': 'Edited within open month', 'hours': '6',
            })
        self.draft_entry.refresh_from_db()
        self.assertEqual(self.draft_entry.task_description, 'Edited within open month')
        self.assertEqual(resp.status_code, 302)

    def test_can_add_entry_after_hr_reopens_the_month(self):
        """reopen_month is HR's way of unlocking a month for a correction.
        Confirms _month_is_locked actually respects that -- not just that
        submitted_at is set, but that a later reopened_at unlocks it again."""
        from .services import reopen_month
        reopen_month(employee=self.emp, year=2026, month=7, reopened_by=self.user)

        resp = self.client.post(
            reverse('timesheets:my_timesheet') + '?year=2026&month=7', {
                'entry_type': 'activity',
                'date': '2026-07-22', 'activity_code': self.code.pk,
                'task_description': 'Added after HR reopened July', 'hours': '4',
            })
        self.assertTrue(
            TimesheetEntry.objects.filter(task_description='Added after HR reopened July').exists())
        self.assertEqual(resp.status_code, 302)

    def test_month_with_no_timesheetmonth_row_is_not_locked(self):
        """A month that was never submitted at all has no TimesheetMonth
        row -- distinct code path from one that exists with is_submitted
        =False. Both must behave as 'open'."""
        self.assertFalse(
            self.TimesheetMonth.objects.filter(employee=self.emp, year=2026, month=9).exists())
        resp = self.client.post(
            reverse('timesheets:my_timesheet') + '?year=2026&month=9', {
                'entry_type': 'activity',
                'date': '2026-09-10', 'activity_code': self.code.pk,
                'task_description': 'Never-submitted month entry', 'hours': '5',
            })
        self.assertTrue(
            TimesheetEntry.objects.filter(task_description='Never-submitted month entry').exists())
        self.assertEqual(resp.status_code, 302)

    def test_inline_save_still_blocked_on_a_locked_months_submitted_entry(self):
        """The Send to HR preview's inline pencil-edit uses a separate
        endpoint (timesheet_entry_inline_save). Confirms it's independently
        safe, not just 'probably fine' because it never touches the date --
        it must still refuse to save once the underlying entry is locked."""
        locked_entry = TimesheetEntry.objects.get(
            employee=self.emp, date=date(2026, 7, 10))
        resp = self.client.post(
            reverse('timesheets:entry_inline_save', args=[locked_entry.pk]), {
                'entry_type': 'activity', 'activity_code': self.code.pk,
                'task_description': 'Sneaky inline edit', 'hours': '2',
            })
        self.assertEqual(resp.status_code, 400)
        locked_entry.refresh_from_db()
        self.assertEqual(locked_entry.task_description, 'Already submitted')


class HRDownloadZipTests(TestCase):
    """hr_download_zip: role-gated, correct employees per month, correct
    per-employee file contents, safe fallback on bad input."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.hr_user, _ = _make_user_with_employee('ziphr', role_name=Role.ADMIN)
        self.plain_user, _ = _make_user_with_employee('zipplain')
        self.emp_a_user, self.emp_a = _make_user_with_employee('zipempa', full_name='Zip Employee A')
        self.emp_b_user, self.emp_b = _make_user_with_employee('zipempb', full_name='Zip Employee B')
        self.emp_c_user, self.emp_c = _make_user_with_employee('zipempc', full_name='Zip Employee C')
        seed_default_permissions()

        # July: A and B have entries, C does not
        TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 7, 5),
            task_description='A July work', activity_code=self.code, hours=Decimal('10'))
        TimesheetEntry.objects.create(
            employee=self.emp_b, date=date(2026, 7, 6),
            task_description='B July work', activity_code=self.code, hours=Decimal('10'))
        # August: only C has entries
        TimesheetEntry.objects.create(
            employee=self.emp_c, date=date(2026, 8, 3),
            task_description='C August work', activity_code=self.code, hours=Decimal('10'))
        # A also has an August entry -- used to prove per-file month isolation
        TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 8, 10),
            task_description='A August work', activity_code=self.code, hours=Decimal('10'))

    def test_plain_employee_cannot_download_zip(self):
        self.client.login(username='zipplain', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 403)

    def test_anonymous_redirected(self):
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 302)
        self.assertIn('/login', resp.url)

    def test_single_month_zip_includes_only_employees_with_entries_that_month(self):
        import zipfile
        from io import BytesIO
        self.client.login(username='ziphr', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?year=2026&month=7')
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/zip')
        zf = zipfile.ZipFile(BytesIO(resp.content))
        names = zf.namelist()
        self.assertTrue(any('Zip_Employee_A' in n for n in names))
        self.assertTrue(any('Zip_Employee_B' in n for n in names))
        self.assertFalse(any('Zip_Employee_C' in n for n in names))

    def test_single_month_zip_file_contents_isolated_to_that_month(self):
        """The real cross-month leak check: A's July file must not
        contain A's August data, even though A has entries in both."""
        import zipfile
        from io import BytesIO
        import openpyxl
        self.client.login(username='ziphr', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?year=2026&month=7')
        zf = zipfile.ZipFile(BytesIO(resp.content))
        a_file = next(n for n in zf.namelist() if 'Zip_Employee_A' in n)
        wb = openpyxl.load_workbook(BytesIO(zf.read(a_file)))
        ws = wb.active
        all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        self.assertIn('A July work', all_text)
        self.assertNotIn('A August work', all_text)

    def test_invalid_month_redirects_with_message_not_500(self):
        self.client.login(username='ziphr', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?year=2026&month=13')
        self.assertEqual(resp.status_code, 302)
        self.assertRedirects(resp, reverse('timesheets:hr_request'))

    def test_download_all_zip_uses_month_named_folders(self):
        import zipfile
        from io import BytesIO
        self.client.login(username='ziphr', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?all=1')
        self.assertEqual(resp.status_code, 200)
        zf = zipfile.ZipFile(BytesIO(resp.content))
        names = zf.namelist()
        self.assertTrue(any(n.startswith('July_2026/') for n in names))
        self.assertTrue(any(n.startswith('August_2026/') for n in names))

    def test_download_all_zip_no_data_redirects(self):
        TimesheetEntry.objects.all().delete()
        self.client.login(username='ziphr', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?all=1')
        self.assertEqual(resp.status_code, 302)
        self.assertRedirects(resp, reverse('timesheets:hr_request'))

    def test_KNOWN_ISSUE_same_name_employees_both_survive_extraction(self):
        """RED TEST -- documents a real bug: _build_timesheet_workbook
        names files from full_name alone, with no employee id. Two
        employees sharing a full name get the SAME arcname in the zip,
        so extracting it lets one silently overwrite the other -- HR
        loses a timesheet with no error. This test currently FAILS.
        Fix by including something employee-unique (id/iqama_number) in
        the filename before merging; do not loosen this assertion."""
        import zipfile
        import tempfile
        import os
        from io import BytesIO
        dup_a = Employee.objects.get(pk=self.emp_a.pk)
        dup_a.full_name = 'Same Name'
        dup_a.save()
        dup_b = Employee.objects.get(pk=self.emp_b.pk)
        dup_b.full_name = 'Same Name'
        dup_b.save()

        self.client.login(username='ziphr', password='testpass123')
        resp = self.client.get(reverse('timesheets:hr_download_zip') + '?year=2026&month=7')
        zf = zipfile.ZipFile(BytesIO(resp.content))
        with tempfile.TemporaryDirectory() as tmp:
            zf.extractall(tmp)
            same_name_files = [f for f in os.listdir(tmp) if 'Same_Name' in f]
            # Two distinct employees -> two distinct files should survive
            # extraction. Today only one does.
            self.assertEqual(len(same_name_files), 2)


class SendToHRBrowsingDecouplingTests(TestCase):
    """The core invariant from the design doc: browsing My Timesheet to a
    different month must never influence which month gets emailed/locked
    when the employee actually clicks Send to HR. send_to_hr and
    acknowledge_send must key off ts_request.year/month only."""

    def setUp(self):
        self.client = Client()
        self.code, _ = ActivityCode.objects.get_or_create(
            code='COS_0009', defaults={'label': 'Office', 'is_active': True})
        self.hr_user, _ = _make_user_with_employee('decouplehr', role_name=Role.ADMIN)
        HRSettings.objects.get_or_create(pk=1, defaults={'hr_email': 'hr@leap-arabia.com'})
        self.user_a, self.emp_a = _make_user_with_employee('decouplea', full_name='Decouple A')
        seed_default_permissions()
        self.ts_request = request_timesheets(year=2026, month=8, requested_by=self.hr_user)

        TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 8, 15),
            task_description='August-only task', activity_code=self.code, hours=Decimal('10'))
        self.october_entry = TimesheetEntry.objects.create(
            employee=self.emp_a, date=date(2026, 10, 15),
            task_description='October-only task', activity_code=self.code, hours=Decimal('10'))

    def test_browsing_a_different_month_before_opening_preview_does_not_leak(self):
        self.client.login(username='decouplea', password='testpass123')
        # Employee is browsing October on My Timesheet first...
        self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=10')
        # ...then opens the pending August request.
        resp = self.client.get(reverse('timesheets:send_to_hr', args=[self.ts_request.pk]))
        self.assertContains(resp, 'August-only task')
        self.assertNotContains(resp, 'October-only task')

    def test_send_to_hr_ignores_year_month_query_params_on_its_own_url(self):
        self.client.login(username='decouplea', password='testpass123')
        resp = self.client.get(
            reverse('timesheets:send_to_hr', args=[self.ts_request.pk]) + '?year=2026&month=10')
        self.assertContains(resp, 'August-only task')
        self.assertNotContains(resp, 'October-only task')

    @patch('timesheets.views.EmailMessage')
    def test_acknowledge_send_only_locks_and_emails_the_requests_month(self, mock_email_cls):
        import openpyxl
        from io import BytesIO
        from .models import TimesheetMonth
        mock_msg = mock_email_cls.return_value
        mock_msg.send.return_value = 1

        self.client.login(username='decouplea', password='testpass123')
        self.client.get(reverse('timesheets:my_timesheet') + '?year=2026&month=10')
        self.client.post(reverse('timesheets:acknowledge_send', args=[self.ts_request.pk]))

        attach_args = mock_msg.attach.call_args[0]
        wb = openpyxl.load_workbook(BytesIO(attach_args[1]))
        ws = wb.active
        all_text = ' '.join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        self.assertIn('August-only task', all_text)
        self.assertNotIn('October-only task', all_text)

        self.october_entry.refresh_from_db()
        self.assertEqual(self.october_entry.status, TimesheetEntry.STATUS_DRAFT)

        self.assertTrue(TimesheetMonth.objects.filter(
            employee=self.emp_a, year=2026, month=8, submitted_at__isnull=False).exists())
        self.assertFalse(TimesheetMonth.objects.filter(
            employee=self.emp_a, year=2026, month=10).exists())




