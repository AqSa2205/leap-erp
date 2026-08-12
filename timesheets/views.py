from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_POST
from decimal import Decimal
from collections import defaultdict
from io import BytesIO
import zipfile
from accounts.permissions import require_capability
from .models import TimesheetEntry, TimesheetMonth
from .forms import TimesheetEntryForm
import calendar
import re
from datetime import date as date_cls
from .services import employees_for_request, remind_employee
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.conf import settings
from django.core.mail import EmailMessage
from django.http import HttpResponse, JsonResponse
from .models import TimesheetRequest, TimesheetRequestAck, HRSettings
from .services import request_timesheets, delete_request, submit_month, reopen_month
from engineer_calendar.services import generate_draft



def _entry_label(entry):
    """Short code/reference used to dedupe and display an entry — the
    project's proposal reference for project-based entries, the activity
    code otherwise. Centralised here so the HR preview, the Excel export,
    and the my-timesheet grouping all agree on what to show."""
    return entry.project.proposal_reference if entry.project_id else entry.activity_code.code

def _get_employee_or_none(request):
    """Same lookup my_profile uses in hr/views.py — an authenticated user
    isn't guaranteed to be linked to an Employee record yet."""
    return getattr(request.user, 'employee_profile', None)


def _month_is_locked(employee, on_date):
    """True if `on_date` falls in a month that's already been submitted
    to HR for this employee (and not since reopened). Used to stop a new
    or edited entry from landing in a month HR has already received —
    the month-picker makes it easy to browse back into one, so this has
    to be enforced server-side, not just by hiding the form."""
    tsm = TimesheetMonth.objects.filter(
        employee=employee, year=on_date.year, month=on_date.month).first()
    return bool(tsm and tsm.is_submitted)


def _group_entries_by_date(entries):
    """Group entries (already newest-first per the model's Meta.ordering)
    by calendar date, so several same-day rows — e.g. auto-start/auto-end
    tasks — collapse into a single summary row that expands to show each
    individual entry with its own Edit/Delete."""
    groups = {}
    for e in entries:
        groups.setdefault(e.date, []).append(e)

    result = []
    for entry_date, day_entries in groups.items():
        # Same "one row per unique activity code" rule as the HR preview —
        # a task that auto-starts and auto-ends the same day can leave
        # several COS_0009 rows; only count each code's hours once so the
        # total can never silently exceed a real workday.
        unique_by_code = {}
        order = []
        for e in day_entries:
            code = _entry_label(e)
            if code not in unique_by_code:
                unique_by_code[code] = e
                order.append(code)

        is_multi = len(day_entries) > 1
        if is_multi:
            combined_description = ' '.join(
                f'{i}) {e.task_description}' for i, e in enumerate(day_entries, start=1))
        else:
            combined_description = day_entries[0].task_description  # unchanged, no prefix

        result.append({
            'date': entry_date,
            'entries': day_entries,
            'is_multi': is_multi,
            'combined_description': combined_description,
            'combined_codes': ', '.join(order),
            'combined_hours': sum((unique_by_code[c].hours for c in order), Decimal('0')),
            'locked': any(e.status == TimesheetEntry.STATUS_SUBMITTED for e in day_entries),
        })
    return result


@login_required
def my_timesheet(request):
    """Self-service page: list the employee's own entries + a form to add
    a new one. Mirrors hr.views.my_profile's GET-display / POST-create
    shape for a single-purpose form."""
    emp = _get_employee_or_none(request)
    if emp is None:
        messages.error(request, 'Your account is not linked to an employee record yet. Contact HR.')
        return render(request, 'timesheets/my_timesheet.html', {'employee': None, 'entries': [], 'form': None})

    year, month = _resolve_timesheet_month(request)

    if request.method == 'POST':
        form = TimesheetEntryForm(request.POST)
        if form.is_valid():
            entry_date = form.cleaned_data['date']
            if _month_is_locked(emp, entry_date):
                messages.error(
                    request,
                    f'{calendar.month_name[entry_date.month]} {entry_date.year} has already '
                    'been submitted to HR and is locked. Contact HR if this needs a correction.',
                )
            else:
                entry = form.save(commit=False)
                entry.employee = emp
                entry.save()
                messages.success(request, 'Entry added.')
                return redirect(f"{reverse('timesheets:my_timesheet')}?year={year}&month={month}")
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = TimesheetEntryForm()
    entries = TimesheetEntry.objects.filter(
        employee=emp, date__year=year, date__month=month
    ).select_related('activity_code', 'project')
    grouped_entries = _group_entries_by_date(entries)
    # Only nag about the COS_0009 default when the employee actually has an
    # entry the auto-start/auto-end signal created -- checking the activity
    # code alone was wrong, since a person can manually pick COS_0009 for
    # unrelated work and get shown a banner that has nothing to do with them.
    show_autogen_banner = entries.filter(
        task_description__startswith='Started task: '
    ).exists() or entries.filter(
        task_description__startswith='Ended task: '
    ).exists()
    pending_request = _pending_request_for(emp)
    prev_year, prev_month = _adjacent_month(year, month, -1)
    next_year, next_month = _adjacent_month(year, month, 1)

    return render(request, 'timesheets/my_timesheet.html', {
        'employee': emp,
        'entries': entries,
        'grouped_entries': grouped_entries,
        'show_autogen_banner': show_autogen_banner,
        'form': form,
        'pending_request': pending_request,
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'prev_year': prev_year,
        'prev_month': prev_month,
        'next_year': next_year,
        'next_month': next_month,
    })


@login_required
def timesheet_entry_edit(request, pk):
    """Edit one of the employee's own entries. Ownership is checked before
    anything else — same 'don't distinguish 404 vs 403' spirit as
    leave_request_document_download: we just 404 if it isn't theirs."""
    emp = _get_employee_or_none(request)
    entry = get_object_or_404(TimesheetEntry, pk=pk, employee=emp)

    if entry.status != TimesheetEntry.STATUS_DRAFT:
        messages.error(request, 'This entry has been submitted and can no longer be edited.')
        return redirect('timesheets:my_timesheet')

    if request.method == 'POST':
        form = TimesheetEntryForm(request.POST, instance=entry)
        if form.is_valid():
            entry_date = form.cleaned_data['date']
            if _month_is_locked(emp, entry_date):
                messages.error(
                    request,
                    f'{calendar.month_name[entry_date.month]} {entry_date.year} has already '
                    'been submitted to HR and is locked. Contact HR if this needs a correction.',
                )
            else:
                form.save()
                messages.success(request, 'Entry updated.')
                return redirect('timesheets:my_timesheet')
        else:
            messages.error(request, 'Please fix the errors below.')
    else:
        form = TimesheetEntryForm(instance=entry)

    return render(request, 'timesheets/entry_edit.html', {'form': form, 'entry': entry})


@login_required
@require_POST
def timesheet_entry_delete(request, pk):
    """Delete one of the employee's own draft entries. POST-only, like the
    other delete views in this codebase (e.g. employee_document_delete)."""
    emp = _get_employee_or_none(request)
    entry = get_object_or_404(TimesheetEntry, pk=pk, employee=emp)

    if entry.status != TimesheetEntry.STATUS_DRAFT:
        messages.error(request, 'This entry has been submitted and can no longer be deleted.')
        return redirect('timesheets:my_timesheet')

    entry.delete()
    messages.success(request, 'Entry deleted.')
    return redirect('timesheets:my_timesheet')


@login_required
@require_POST
def timesheet_entry_inline_save(request, pk):
    """AJAX-only save used by the Send to HR preview page's inline edit.
    Same ownership/lock rules as timesheet_entry_edit — the employee just
    never has to leave the preview to make a last-minute fix before sending.
    The date is intentionally not editable here: it's fixed by which row
    on the preview they clicked, so we always keep the entry's real date."""
    emp = _get_employee_or_none(request)
    if emp is None:
        return JsonResponse({'error': 'Your account is not linked to an employee record.'}, status=400)

    entry = get_object_or_404(TimesheetEntry, pk=pk, employee=emp)
    if entry.status != TimesheetEntry.STATUS_DRAFT:
        return JsonResponse({'error': 'This entry has been submitted and can no longer be edited.'}, status=400)

    data = request.POST.copy()
    data['date'] = entry.date.isoformat()
    form = TimesheetEntryForm(data, instance=entry)
    if not form.is_valid():
        first_error = next(iter(form.errors.values()))[0]
        return JsonResponse({'error': first_error}, status=400)

    entry = form.save()
    return JsonResponse({
        'ok': True,
        'task_description': entry.task_description,
        'hours': str(entry.hours),
        'code_label': _entry_label(entry),
    })


def _ordinal(n):
    """1 -> '1st', 2 -> '2nd', 3 -> '3rd', 4 -> '4th', etc."""
    if 11 <= (n % 100) <= 13:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f'{n}{suffix}'


@login_required
def timesheet_export(request):
    """Export the logged-in employee's own month as an .xlsx matching the
    original HR template exactly (styling copied from the real template:
    Century Gothic title, C00000 header fill, thin borders, dd/mmm/yyyy
    dates, yellow-highlighted Fri/Sat rows). One row per calendar day —
    filled from a TimesheetEntry if one exists that day, blank otherwise."""
    emp = _get_employee_or_none(request)
    if emp is None:
        messages.error(request, 'Your account is not linked to an employee record yet. Contact HR.')
        return redirect('timesheets:my_timesheet')

    today = date_cls.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError('Month out of range')
        calendar.monthrange(year, month)[1]
    except (ValueError, TypeError):
        messages.error(request, 'Invalid year or month.')
        return redirect('timesheets:my_timesheet')

    wb, filename = _build_timesheet_workbook(emp, year, month)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def _build_timesheet_workbook(emp, year, month):
    """Build the HR-template timesheet workbook for one employee/month.
    Shared by the download export and the Send-to-HR email attachment so
    both always produce the same file."""
    days_in_month = calendar.monthrange(year, month)[1]
    entries_by_day = defaultdict(list)
    for e in TimesheetEntry.objects.filter(
            employee=emp, date__year=year, date__month=month
        ).select_related('activity_code', 'project').order_by('id'):
        entries_by_day[e.date.day].append(e)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Time_Sheet'

    title_font = Font(name='Century Gothic', bold=True, size=22)
    label_font = Font(name='Trebuchet MS', size=10)
    header_font = Font(name='Trebuchet MS', size=10, color='FFFFFF')
    header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='bottom')
    weekend_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('D2:G2')
    ws['D2'] = 'MONTHLY  TIMESHEET'
    ws['D2'].font = title_font
    ws['D2'].alignment = Alignment(horizontal='center', vertical='center')

    ws['C4'] = 'Employee:'
    ws['C4'].font = label_font
    ws['D4'] = emp.full_name
    ws['D4'].font = label_font
    ws['H4'] = 'Signature:____________________________'
    ws['H4'].font = label_font

    ws['C5'] = 'Designation'
    ws['C5'].font = label_font
    ws['D5'] = emp.designation or ''
    ws['D5'].font = label_font

    ws['C7'] = 'Supervisor Name :'
    ws['C7'].font = label_font
    ws['D7'] = emp.main_manager.full_name if emp.main_manager else ''
    ws['D7'].font = label_font
    ws['H7'] = 'Signature:____________________________'
    ws['H7'].font = label_font

    month_name = calendar.month_name[month]
    ws['C9'] = 'Month of'
    ws['C9'].font = label_font
    ws['D9'] = f'{_ordinal(1)} {month_name} {year} to {_ordinal(days_in_month)} {month_name} {year}'
    ws['D9'].font = label_font

    headers = ['Sr.No', 'Activity_Description', 'Date', 'Type', 'Hours ', 'Activity_code']
    for col, text in enumerate(headers, start=2):
        cell = ws.cell(row=11, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for day in range(1, days_in_month + 1):
        row = 11 + day
        the_date = date_cls(year, month, day)
        day_entries = entries_by_day.get(day, [])
        is_weekend = the_date.weekday() in (4, 5)

        combined_description = '; '.join(e.task_description for e in day_entries) or None
        combined_hours = Decimal('10') if day_entries else None
        seen_codes = []
        for e in day_entries:
            label = _entry_label(e)
            if label not in seen_codes:
                seen_codes.append(label)
        combined_codes = ', '.join(seen_codes) or None

        values = {
            2: day,
            3: combined_description,
            4: the_date,
            5: 'Normal' if day_entries else None,
            6: combined_hours,
            7: combined_codes,
        }
        for col, value in values.items():
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            if col in (3, 7):
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            else:
                cell.alignment = center_alignment
            if col == 4:
                cell.number_format = 'dd/mmm/yyyy'
            if is_weekend:
                cell.fill = weekend_fill

    widths = {'A': 10, 'B': 10, 'C': 60, 'D': 34, 'E': 10, 'G': 16, 'H': 41}
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    safe_name = re.sub(r'[^A-Za-z0-9_-]', '_', emp.full_name.replace(' ', '_'))
    filename = f'{safe_name}_{emp.pk}_Timesheet_{month_name}_{year}.xlsx'
    return wb, filename


@login_required
@require_capability('timesheets.review')
def hr_request_timesheets(request):
    """HR clicks a button to ask everyone for a given month's timesheet.
    Defaults to the current month, but HR can pick an earlier one (e.g. if
    they forgot to request last month)."""
    today = date_cls.today()
    if request.method == 'POST':
        try:
            year = int(request.POST.get('year', today.year))
            month = int(request.POST.get('month', today.month))
        except (ValueError, TypeError):
            messages.error(request, 'Invalid year or month.')
            return redirect('timesheets:hr_request')
        request_timesheets(year=year, month=month, requested_by=request.user)
        messages.success(request, f'Timesheet request sent for {month}/{year}.')
        return redirect('timesheets:hr_request')
    recent_requests = TimesheetRequest.objects.all()[:10]
    month_choices = [(i, calendar.month_name[i]) for i in range(1, 13)]
    return render(request, 'timesheets/hr_request.html', {
        'today': today,
        'month_name': calendar.month_name[today.month],
        'recent_requests': recent_requests,
        'month_choices': month_choices,
    })


@login_required
@require_capability('timesheets.review')
def hr_download_zip(request):
    """HR bulk-downloads timesheets as a zip: either every employee's
    .xlsx for one chosen month, or one zip covering every month that has
    any entries at all, organized into per-month folders inside the zip.
    Reuses _build_timesheet_workbook so the files inside are byte-for-byte
    the same format as the personal export / HR email attachment."""
    from hr.models import Employee

    download_all = request.GET.get('all') == '1'

    if download_all:
        months = list(
            TimesheetEntry.objects.values_list('date__year', 'date__month')
            .distinct().order_by('-date__year', '-date__month')
        )
        if not months:
            messages.error(request, 'There are no timesheet entries to download yet.')
            return redirect('timesheets:hr_request')
        zip_filename = 'Timesheets_All_Months.zip'
    else:
        today = date_cls.today()
        try:
            year = int(request.GET.get('year', today.year))
            month = int(request.GET.get('month', today.month))
            if not (1 <= month <= 12):
                raise ValueError('Month out of range')
            calendar.monthrange(year, month)
        except (ValueError, TypeError):
            messages.error(request, 'Invalid year or month.')
            return redirect('timesheets:hr_request')
        months = [(year, month)]
        zip_filename = f'Timesheets_{calendar.month_name[month]}_{year}.zip'

    buf = BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for yr, mo in months:
            employee_ids = TimesheetEntry.objects.filter(
                date__year=yr, date__month=mo
            ).values_list('employee_id', flat=True).distinct()
            employees = Employee.objects.filter(pk__in=employee_ids)

            if not employees:
                continue

            month_folder = f'{calendar.month_name[mo]}_{yr}'
            for emp in employees:
                wb, filename = _build_timesheet_workbook(emp, yr, mo)
                emp_buf = BytesIO()
                wb.save(emp_buf)
                arcname = f'{month_folder}/{filename}' if download_all else filename
                zf.writestr(arcname, emp_buf.getvalue())

    response = HttpResponse(buf.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    return response


def _build_month_rows(emp, year, month):
    """Shared by both the Excel export and the HR preview page, so the two
    never quietly drift apart into showing different numbers for the same
    month."""
    days_in_month = calendar.monthrange(year, month)[1]
    entries_by_day = defaultdict(list)
    for e in TimesheetEntry.objects.filter(
            employee=emp, date__year=year, date__month=month
        ).select_related('activity_code', 'project').order_by('id'):
        entries_by_day[e.date.day].append(e)

    rows = []
    for day in range(1, days_in_month + 1):
        the_date = date_cls(year, month, day)
        day_entries = entries_by_day.get(day, [])

        unique_by_code = {}
        order = []
        for e in day_entries:
            code = _entry_label(e)
            if code not in unique_by_code:
                unique_by_code[code] = e
                order.append(code)

        rows.append({
            'day': day,
            'date': the_date,
            'description': '; '.join(unique_by_code[c].task_description for c in order),
            'hours': Decimal('10') if order else None,
            'codes': ', '.join(order),
            'is_weekend': the_date.weekday() in (4, 5),
        })
    return rows


def _resolve_timesheet_month(request):
    """Read year/month from the query string for the My Timesheet month
    browser, falling back to today when absent or invalid."""
    today = date_cls.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        if not (1 <= month <= 12):
            raise ValueError('Month out of range')
        calendar.monthrange(year, month)
    except (ValueError, TypeError, OverflowError):
        year, month = today.year, today.month
    return year, month


def _adjacent_month(year, month, delta):
    """delta=-1 for the previous month, +1 for the next, wrapping the year."""
    month += delta
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1
    return year, month


def _pending_request_for(emp):
    """The most recent TimesheetRequest this employee hasn't acknowledged
    yet, or None if they're all caught up."""
    acked_request_ids = TimesheetRequestAck.objects.filter(
        employee=emp).values_list('request_id', flat=True)
    return TimesheetRequest.objects.exclude(pk__in=acked_request_ids).first()


@login_required
def send_to_hr(request, request_id):
    """Preview page before sending. Shows the month's data and a single
    Send to HR button. Ownership-scoped: only usable for a request that
    is still pending for THIS employee."""
    emp = _get_employee_or_none(request)
    if emp is None:
        messages.error(request, 'Your account is not linked to an employee record yet. Contact HR.')
        return redirect('timesheets:my_timesheet')

    try:
        ts_request = TimesheetRequest.objects.get(pk=request_id)
    except TimesheetRequest.DoesNotExist:
        messages.info(request, 'HR deleted this request.')
        return redirect('timesheets:my_timesheet')

    already_acked = TimesheetRequestAck.objects.filter(
        request=ts_request, employee=emp).exists()
    if already_acked:
        messages.info(request, 'You already sent this timesheet to HR.')
        return redirect('timesheets:my_timesheet')

    year, month = ts_request.year, ts_request.month
    rows = _build_month_rows(emp, year, month)
    month_name = calendar.month_name[month]
    days_in_month = calendar.monthrange(year, month)[1]
    date_range_label = f'{_ordinal(1)} {month_name} {year} to {_ordinal(days_in_month)} {month_name} {year}'
    hr_email = HRSettings.load().hr_email

    entries_by_day = defaultdict(list)
    for e in TimesheetEntry.objects.filter(
            employee=emp, date__year=year, date__month=month
        ).select_related('activity_code', 'project').order_by('id'):
        entries_by_day[e.date.day].append(e)
    for row in rows:
        row['entries'] = entries_by_day.get(row['day'], [])

    return render(request, 'timesheets/send_to_hr.html', {
        'employee': emp,
        'ts_request': ts_request,
        'rows': rows,
        'month_name': month_name,
        'date_range_label': date_range_label,
        'hr_email': hr_email,
    })


@login_required
@require_POST
def acknowledge_send(request, request_id):
    """Email the timesheet Excel to HR via Graph, then ack + lock + calendar draft.
    On email failure: no ack, no lock, no draft."""
    emp = _get_employee_or_none(request)
    if emp is None:
        messages.error(request, 'Your account is not linked to an employee record yet. Contact HR.')
        return redirect('timesheets:my_timesheet')

    try:
        ts_request = TimesheetRequest.objects.get(pk=request_id)
    except TimesheetRequest.DoesNotExist:
        messages.info(request, 'HR deleted this request.')
        return redirect('timesheets:my_timesheet')

    if TimesheetRequestAck.objects.filter(request=ts_request, employee=emp).exists():
        messages.info(request, 'You already sent this timesheet to HR.')
        return redirect('timesheets:my_timesheet')

    year, month = ts_request.year, ts_request.month
    month_name = calendar.month_name[month]
    hr_email = (HRSettings.load().hr_email or '').strip()
    if not hr_email:
        messages.error(request, 'HR email is not configured yet. Contact an administrator.')
        return redirect('timesheets:send_to_hr', request_id=ts_request.pk)

    wb, filename = _build_timesheet_workbook(emp, year, month)
    buf = BytesIO()
    wb.save(buf)

    # Stable subject for a future mailbox auto-detect feature.
    subject = f'Timesheet - {emp.full_name} - {month_name} {year}'
    body = (
        f'Hi,\n\n'
        f'Please find attached the timesheet for {emp.full_name} '
        f'({month_name} {year}).\n\n'
        f'Sent automatically from Leap ERP.\n'
    )
    msg = EmailMessage(
        subject=subject,
        body=body,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[hr_email],
    )
    reply_to = (
        (emp.work_email or '').strip()
        or (emp.personal_email or '').strip()
        or (getattr(request.user, 'email', '') or '').strip()
    )
    if reply_to:
        msg.reply_to = [reply_to]
    msg.attach(
        filename,
        buf.getvalue(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    try:
        sent = msg.send(fail_silently=False)
    except Exception:
        messages.error(
            request,
            'Could not send the timesheet email to HR. Please try again in a moment. '
            'Your timesheet was not locked.',
        )
        return redirect('timesheets:send_to_hr', request_id=ts_request.pk)

    if not sent:
        messages.error(
            request,
            'Could not send the timesheet email to HR. Please try again in a moment. '
            'Your timesheet was not locked.',
        )
        return redirect('timesheets:send_to_hr', request_id=ts_request.pk)

    TimesheetRequestAck.objects.get_or_create(request=ts_request, employee=emp)
    generate_draft([emp], year=year, month=month)
    try:
        submit_month(
            employee=emp, year=year, month=month,
            submitted_by=request.user)
    except ValueError:
        pass

    messages.success(request, f'Timesheet sent to HR ({hr_email}).')
    return redirect('timesheets:my_timesheet')


@login_required
@require_capability('timesheets.review')
def hr_request_detail(request, request_id):
    """HR's view of one specific request: who's sent it, who hasn't, with a
    Remind button per pending employee."""
    ts_request = get_object_or_404(TimesheetRequest, pk=request_id)

    if request.method == 'POST':
        emp_pk = request.POST.get('employee_id')
        from hr.models import Employee
        employee = get_object_or_404(Employee, pk=emp_pk)
        try:
            remind_employee(ts_request=ts_request, employee=employee, reminded_by=request.user)
            messages.success(request, f'Reminder sent to {employee.full_name}.')
        except ValueError as e:
            messages.error(request, str(e))
        return redirect('timesheets:hr_request_detail', request_id=ts_request.pk)

    rows = employees_for_request(ts_request)
    sent_count = sum(1 for r in rows if r['acked'])
    return render(request, 'timesheets/hr_request_detail.html', {
        'ts_request': ts_request,
        'rows': rows,
        'sent_count': sent_count,
        'total_count': len(rows),
    })


@login_required
@require_capability('timesheets.review')
@require_POST
def hr_request_delete(request, request_id):
    """Permanently remove a timesheet request — e.g. HR picked the wrong
    month by mistake. Also reopens (un-submits) any employee's entries
    that were locked as a result of acknowledging THIS request, so
    deleting the request doesn't leave entries stuck as 'Submitted' with
    no request left to explain why."""
    ts_request = get_object_or_404(TimesheetRequest, pk=request_id)
    year, month = ts_request.year, ts_request.month

    acked_employees = [
        ack.employee for ack in
        TimesheetRequestAck.objects.filter(request=ts_request).select_related('employee')
    ]
    for emp in acked_employees:
        try:
            reopen_month(employee=emp, year=year, month=month, reopened_by=request.user)
        except ValueError:
            pass

    delete_request(ts_request)
    messages.success(request, f'Timesheet request for {month}/{year} deleted.')
    return redirect('timesheets:hr_request')
